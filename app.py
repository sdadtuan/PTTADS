from __future__ import annotations

import csv
import hashlib
import json
import logging
import os
import re
import secrets
import smtplib
import sqlite3
import threading
import uuid
from calendar import monthrange
from copy import deepcopy
from datetime import date, datetime, timedelta
from email.message import EmailMessage
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any
from urllib.parse import quote

from marketing_chat_service import (
    build_export_html,
    build_export_markdown,
    build_marketing_chat_config,
    build_marketing_strategy_reply,
)
from crm_daily_work_report import daily_work_report_xlsx_response
from crm_daily_work_report_store import (
    daily_work_report_row_to_dict,
    ensure_daily_work_report_schema,
    fetch_daily_work_report_by_id,
    fetch_daily_work_report_by_staff_date,
    fetch_daily_work_reports,
    normalize_tasks,
    update_daily_work_report_by_id,
    upsert_daily_work_report,
    validate_report_date,
)
from crm_lead_care_pipeline import CARE_PIPELINE_STAGES_PUBLIC, complete_lead_care_stage
from crm_lead_store import (
    ACTIVITY_TYPES,
    ACTIVITY_TYPE_LABELS,
    CRM_CARE_CONTACT_TO_ACTIVITY,
    LEAD_LEVELS,
    LEAD_LEVEL_LABELS,
    LEAD_SOURCES,
    LEAD_SOURCE_LABELS,
    LEAD_STATUSES,
    LEAD_STATUS_LABELS,
    activity_row_to_dict,
    assign_lead,
    count_leads,
    create_lead,
    delete_lead,
    ensure_lead_schema,
    apply_lead_score,
    fetch_lead_activities,
    fetch_lead_by_id,
    fetch_lead_stats,
    fetch_lead_stats_extended,
    fetch_lead_assignment_logs,
    fetch_lead_status_logs,
    fetch_leads,
    fetch_max_lead_id,
    fetch_max_lead_id_any,
    fetch_new_assigned_leads,
    fetch_facebook_webhook_repeat_leads,
    find_duplicate_matches,
    lead_row_to_dict,
    log_lead_activity,
    update_lead,
    validate_lead_contacts,
)
from crm_lead_ai import (
    ai_classify_suggestion,
    ai_recommend_lead,
    ai_search_leads,
    ai_suggest_products_for_lead,
    ai_summarize_lead,
    ai_price_list_query,
)
from crm_project_leads import (
    _UNSET,
    add_project_staff,
    assert_staff_portal_project,
    fetch_staff_project_ids,
    list_assignable_staff_for_project,
    list_lead_project_options,
    list_lead_project_options_for_staff,
    list_project_staff,
    parse_re_project_filter,
    parse_re_project_id,
    remove_project_staff,
    staff_can_view_lead,
    update_project_staff,
)
from crm_re_price_lists import (
    PRICE_LIST_STATUS_LABELS,
    apply_price_list,
    compare_price_lists,
    delete_price_list,
    fetch_price_list,
    import_price_list_items_csv,
    list_all_version_codes,
    list_price_list_items,
    list_price_lists,
    products_on_price_version,
    save_price_list,
)
from crm_re_project_accounting import (
    CASH_FLOW_SOURCE_LABELS,
    CASH_FLOW_STATUS_LABELS,
    CASH_FLOW_TYPE_LABELS,
    MARKETING_SUB_CATEGORY_LABELS,
    ai_project_finance_query,
    apply_predicted_risks_to_register,
    build_accounting_export_sheets,
    compute_accounting_dashboard,
    delete_cash_flow_line,
    forecast_financial_outlook,
    import_cash_flow_csv,
    list_cash_flow_lines,
    predict_financial_risks,
    save_cash_flow_line,
    sync_budget_from_plans,
    sync_revenue_from_inventory,
)
from crm_project_deep import (
    hold_product_for_lead,
    import_products_csv,
    inventory_by_price_batch_summary,
    inventory_by_zone_summary,
    list_price_batches,
    list_project_zones,
    normalize_product_line,
    release_product_hold,
    search_available_products,
)
from crm_lead_rules import (
    DUPLICATE_POLICIES,
    allowed_next_statuses,
    fetch_lead_config,
    fetch_lead_duplicates,
    merge_leads,
    save_lead_config,
    transitions_for_ui,
)
from crm_lead_scoring import DEFAULT_SCORING_RULES, SCORING_CONDITIONS, SCORING_FIELD_OPTIONS
from crm_lead_scoring_rubric import DEFAULT_LEAD_SCORING_RUBRIC, EVALUATOR_OPTIONS, default_scoring_rubric
from crm_lead_auto_assign import DEFAULT_TIER_LEVEL_MAP, STRATEGY_DEFS, merge_assign_config
from crm_lead_tiers import DEFAULT_LEVEL_TIERS, UNCLASSIFIED_TIER_ID, fetch_level_tiers, level_labels_map
from crm_lead_report import build_leads_pdf, build_leads_xlsx
from crm_facebook_config import DEFAULT_FACEBOOK_CONFIG
from crm_facebook_leads import (
    facebook_integration_status,
    fetch_leads_live_revision,
    process_facebook_leadgen_id,
    process_facebook_webhook_payload,
    run_facebook_ingest_cycle,
    save_facebook_webhook_receipt,
    sync_facebook_leads,
)
from crm_facebook_autosync import run_facebook_autosync_once, start_facebook_autosync_worker
from crm_lead_webhooks import (
    facebook_verify_token,
    ingest_webhook_leads,
    parse_facebook_webhook,
    parse_facebook_webhook_json,
    parse_zalo_webhook,
    verify_facebook_signature,
    verify_zalo_signature,
    zalo_webhook_secret,
)
from crm_assistant_service import (
    build_crm_assistant_config,
    build_crm_assistant_reply,
    build_crm_assistant_response,
    build_crm_export_html,
    build_crm_export_markdown,
    fetch_crm_context,
    find_payroll_staff_ids_by_query,
)
from marketing_execution import build_kpi_strategy_xlsx, build_multichannel_plan_xlsx, build_weekly_marketing_plan_xlsx
from marketing_campaign_kit import (
    KPI_BRIEF_PROMPT,
    build_campaign_kpi_xlsx,
    build_khmkt_xlsx,
    generate_campaign_kit,
    get_campaign_kit,
)
from service_extras import DEFAULT_SERVICE_LANDING_EXTRAS
from crm_workflow_playbook import (
    get_crm_lead_intake_master_flow,
    get_crm_marketing_ingress_channels,
    get_crm_workflow_playbook,
)
from crm_sales_hub import (
    MARKET_STATUSES,
    MARKET_STATUS_LABELS_VI,
    PARTNER_STATUSES,
    PARTNER_STATUS_LABELS_VI,
    PARTNER_TYPES,
    PARTNER_TYPE_LABELS_VI,
    PLAN_STATUSES,
    PLAN_STATUS_LABELS_VI,
    TARGET_TYPES,
    TARGET_TYPE_LABELS_VI,
    TRAINING_STATUSES,
    TRAINING_STATUS_LABELS_VI,
    TX_STAGES,
    TX_STAGE_LABELS_VI,
    TX_TYPES,
    TX_TYPE_LABELS_VI,
    ensure_sales_hub_schema,
    fetch_sales_summary,
    list_market_research,
    list_partners,
    list_pipeline_cases,
    list_plans,
    list_targets,
    list_trainings,
    list_transactions,
    sales_report_data,
)
from crm_sales_pipeline import (
    LEGACY_STATUS_TO_PIPELINE,
    SALES_PIPELINE_LABELS_VI,
    SALES_PIPELINE_STAGES,
    TERMINAL_STAGES,
    compute_funnel_stats,
    ensure_pipeline_schema,
    enrich_case_row,
    legacy_status_for_stage,
    normalize_pipeline_stage,
    on_case_created,
    on_pipeline_stage_change,
    pipeline_stage_label,
)
from crm_re_projects import (
    BUDGET_CATEGORIES,
    BUDGET_CATEGORY_LABELS,
    KPI_CATEGORIES,
    KPI_CATEGORY_LABELS,
    KPI_METRIC_TEMPLATES,
    KPI_TRACK_STATUSES,
    KPI_TRACK_STATUS_LABELS,
    PRODUCT_LINES,
    PRODUCT_LINE_LABELS,
    PRODUCT_STATUSES,
    PRODUCT_STATUS_LABELS,
    PRODUCT_TYPOLOGIES,
    PRODUCT_TYPOLOGY_LABELS,
    PROJECT_STATUSES,
    PROJECT_STATUS_LABELS,
    PROJECT_TYPES,
    PROJECT_TYPE_LABELS,
    RISK_CATEGORIES,
    RISK_CATEGORY_LABELS,
    RISK_LEVELS,
    RISK_LEVEL_LABELS,
    create_project as re_create_project,
    compute_kpi_board_stats,
    compute_product_inventory_stats,
    compute_project_workflow,
    delete_budget_line,
    delete_kpi,
    delete_product,
    delete_project as re_delete_project,
    delete_project_type,
    delete_risk,
    ensure_re_projects_schema,
    fetch_project,
    fetch_project_export_data,
    fetch_project_summary,
    pull_project_kpis_from_staff,
    refresh_project_re_leads_new_kpi,
    list_crm_kpi_metrics,
    sync_project_kpis_to_staff,
    list_kpis,
    list_products,
    list_projects,
    list_project_types,
    list_risks,
    project_export_budget_rows,
    project_export_kpi_rows,
    project_export_plan_rows,
    project_export_product_rows,
    project_export_risk_rows,
    project_export_summary_rows,
    project_export_workflow_rows,
    project_type_label_map,
    save_budget_line,
    save_kpi,
    save_product,
    save_project_type,
    save_risk,
    seed_re_project_section_permissions,
    update_project as re_update_project,
)
from admin_page_permissions import (
    ADMIN_CRM_PERMISSION_IDS,
    ADMIN_CRM_SECTION_IDS,
    ADMIN_CRM_SECTIONS,
    backfill_position_grants_customized,
    build_position_permission_matrix,
    default_grants_for_position,
    ensure_position_grants_customized_column,
    migrate_kd01_leads_only_permissions,
    migrate_hdsd_position_permissions,
    mark_position_grants_customized,
    parse_position_grants_payload,
    position_can,
    position_leads_only_ui,
    grants_map_to_rows,
    seed_marketing_positions,
    ui_button_can,
)
from ptt_ui_button_permissions import CRM_UI_BUTTON_BY_ID
from crm_sop_seed import LAUNCH_CAMPAIGN_TEMPLATE_CODE, seed_launch_campaign_sop_template
from crm_service_lifecycle import (
    ensure_schema as _ensure_service_lifecycle_schema,
    activate_lifecycle,
    advance_stage as _svc_advance_stage,
    get_by_lead as _svc_get_by_lead,
    get_by_contract as _svc_get_by_contract,
    list_active as _svc_list_active,
    get_stage_context as _svc_get_stage_context,
    VALID_STAGES as SVC_LIFECYCLE_STAGES,
    VALID_SLUGS as SVC_LIFECYCLE_SLUGS,
)
from crm_svc_tasks import ensure_schema as _ensure_svc_tasks_schema
from crm_lead_intake import ensure_schema as _ensure_lead_intake_schema
from crm_svc_risk import ensure_schema as _ensure_svc_risk_schema
from crm_svc_finance import ensure_schema as _ensure_svc_finance_schema
from crm_svc_finance import migrate_contract_billing_type as _migrate_contract_billing
from crm_svc_kpi import ensure_schema as _ensure_svc_kpi_schema
from crm_customer_brief import ensure_schema as _ensure_customer_brief_schema
from crm_aeo import ensure_schema as _ensure_aeo_schema
from crm_proposal import ensure_schema as _ensure_proposal_schema
from cms_permissions import (
    CMS_ACTIONS,
    CMS_ACTION_LABELS_VI,
    CMS_MODULES,
    CMS_MODULE_IDS,
    CMS_ROLES,
    backfill_role_grants_customized,
    build_permission_matrix,
    default_grants_for_role,
    ensure_role_grants_customized_column,
    migrate_cms_role_sidebar_modules,
    parse_grants_payload,
    role_can,
    role_grants_were_customized,
)
from crm_staff_auth import (
    apply_staff_login_from_payload,
    ensure_staff_login_schema,
    hash_password,
    staff_crm_api_allowed,
    staff_crm_api_allowed_extended,
    staff_portal_html_allowed,
    staff_row_for_api,
    verify_password,
)
from crm_staff_competency import (
    METRIC_OPTIONS,
    default_competency_config,
    score_staff_competency,
)
from crm_staff_levels import DEFAULT_STAFF_LEVELS, normalize_sales_level, staff_level_labels_map
from crm_staff_settings import fetch_staff_config, fetch_staff_competency, save_staff_config
from unified_auth import (
    ensure_unified_password_stored,
    migrate_unified_passwords,
    set_unified_password,
    sync_password_hash,
    unified_login,
    verify_unified_password,
)
from crm_care import (
    CRM_CARE_CONTACT_LABELS_VI,
    CRM_CARE_CONTACT_TYPES,
    CRM_CARE_STATUS_LABELS_VI,
    CRM_CARE_STATUS_TYPES,
    care_report_row_to_dict,
    ensure_care_schema,
    fetch_last_care_reports_map,
    format_care_event_body,
    normalize_care_contact,
    normalize_care_status,
)
from crm_customer_360 import (
    CUSTOMER_GENDER_LABELS_VI,
    CUSTOMER_GENDERS,
    CUSTOMER_LEAD_SOURCE_LABELS_VI,
    CUSTOMER_LEAD_SOURCES,
    ISSUE_PRIORITIES,
    ISSUE_PRIORITY_LABELS_VI,
    ISSUE_STATUSES,
    ISSUE_STATUS_LABELS_VI,
    ISSUE_TYPES,
    ISSUE_TYPE_LABELS_VI,
    PURCHASE_STATUSES,
    PURCHASE_STATUS_LABELS_VI,
    RELATION_TYPES,
    RELATION_TYPE_LABELS_VI,
    apply_profile_patch,
    compute_lead_sources,
    enrich_customer_row,
    ensure_customer_360_schema,
    fetch_customer_issues,
    fetch_customer_purchases,
    fetch_customer_relations,
    normalize_gender,
    normalize_issue_priority,
    normalize_issue_status,
    normalize_issue_type,
    normalize_lead_source,
    normalize_purchase_status,
    normalize_relation_type,
    profile_update_sql_values,
    PROFILE_INSERT_COLS,
    _issue_row,
    _purchase_row,
    _relation_row,
)
from crm_attendance_import import parse_timesheet_xlsx
from crm_payroll_engine import (
    compute_staff_payroll,
    dashboard_summary,
    ensure_payroll_policy_schema,
    enrich_attendance_row,
    load_policy,
    load_position_payroll_map,
    normalize_weekday_shifts,
    parse_work_weekdays,
    policy_for_api,
    count_workdays_in_month,
    weekday_shifts_json,
    work_weekdays_from_shifts,
    default_weekday_shifts,
)
from zkteco_iclock import iclock_get_response, iclock_options_response, parse_attlog_body

try:
    from flask import (
        Flask,
        Response,
        abort,
        flash,
        jsonify,
        redirect,
        render_template,
        request,
        send_file,
        session,
        url_for,
    )
except ModuleNotFoundError as exc:  # pragma: no cover - môi trường thiếu deps
    if getattr(exc, "name", None) == "flask":
        import sys

        sys.stderr.write(
            "\n  [PTT] Chưa cài Flask. Trong thư mục PTT chạy:\n"
            "    python3 -m venv .venv\n"
            "    source .venv/bin/activate   # Windows: .venv\\Scripts\\activate\n"
            "    pip install -r requirements.txt\n"
            "  Hoặc: bash run_dev.sh\n\n"
        )
    raise
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
from werkzeug.exceptions import RequestEntityTooLarge
from werkzeug.utils import secure_filename


BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "ptt.db"
UPLOAD_DIR = BASE_DIR / "static" / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
ALLOWED_UPLOAD_EXTENSIONS = {"jpg", "jpeg", "png", "webp", "gif", "svg", "mp4", "webm", "mov"}
MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MB — ảnh CMS
MAX_VIDEO_UPLOAD_BYTES = 80 * 1024 * 1024  # 80 MB — video hero / media
MAX_CV_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB — CV tuyển dụng
EXCEL_PATH = BASE_DIR.parent / "data.xlsx"
CONSULT_SHEET_NAME = "consultations"
CONSULT_EMAIL_TO = "kosoo81@gmail.com"

# Banner quảng cáo trong mega menu (submenu Dịch vụ) — gộp với settings từ DB, DB ghi đè
DEFAULT_SETTINGS_NAV_MEGA_BANNER: dict[str, str] = {
    "nav_mega_banner_image": "https://images.unsplash.com/photo-1552664730-d307ca884978?auto=format&fit=crop&w=1200&q=80",
    "nav_mega_banner_href": "#contact",
    "nav_mega_banner_title": "Ưu đãi tư vấn Creative Martech",
    "nav_mega_banner_subtitle": "Kết nối đội ngũ PTT — đề xuất theo ngành, kênh & mục tiêu tăng trưởng.",
    # Link OA / chat Zalo (https://zalo.me/...) — nút nổi; để trống thì ẩn nút Zalo
    "fab_zalo_url": "https://zalo.me/0938834238",
}

# Nội dung chân trang landing — CMS ghi đè qua bảng settings
DEFAULT_SETTINGS_FOOTER: dict[str, str] = {
    "footer_cta_title": "Khám phá cách chúng tôi hỗ trợ doanh nghiệp phát triển",
    "footer_cta_line": (
        "Tùy biến chiến lược theo mục tiêu, dữ liệu và công nghệ martech — Creative Martech Platform."
    ),
    "footer_contact_label": "Liên hệ tư vấn",
    "footer_cta_button_label": "Nhận 1 đề xuất",
    "footer_tagline_intro": "Creative Martech Platform",
    "footer_tagline_body": (
        "Tích hợp sáng tạo, dữ liệu và công nghệ — giải pháp marketing tăng trưởng bền vững cho doanh nghiệp Việt Nam."
    ),
    "footer_facebook_url": "https://www.facebook.com",
    "footer_linkedin_url": "https://www.linkedin.com",
    "footer_phones_lines": "+84 (24) 7307 7979 (HN)\n+84 (28) 7307 7979 (HCM)",
    "footer_legal_years": "2020 – 2026",
    "footer_legal_rights": ". All rights reserved.",
    "capabilities_intro": (
        "PTT là nền tảng Creative Martech đa ngành — tích hợp tự động hóa, AI và dữ liệu "
        "vào toàn bộ quy trình marketing: từ sáng tạo nội dung, cá nhân hóa trải nghiệm khách hàng "
        "đến tối ưu hóa hiệu suất chiến dịch theo thời gian thực."
    ),
    "capabilities_items_lines": (
        "Marketing Automation\n"
        "AI Content & Personalization\n"
        "Data Analytics & Business Intelligence\n"
        "CRM & Lead Intelligence\n"
        "Paid Media đa kênh & Tối ưu ROI\n"
        "AI Agent & Customer Experience"
    ),
    "capabilities_items_json": "",
    "partner_logos_json": "",
    "cfab_eyebrow": "Miễn phí · Phản hồi trong 2 giờ",
    "cfab_title": "Nhận tư vấn giải pháp",
    "service_aside_eyebrow": "Tư vấn nhanh",
    "service_aside_title": "Nhận lộ trình theo mục tiêu & ngân sách",
    "service_aside_btn_label": "Đăng ký nhận đề xuất",
}

DEFAULT_HERO_SLIDES: list[dict[str, str]] = [
    {
        "image_url": "https://images.unsplash.com/photo-1486406146926-c627a92ad1ab?auto=format&fit=crop&w=1920&q=80",
        "kicker": "Creative Martech Platform — Automation · AI · Data",
        "heading": "Tự động hóa, AI và dữ liệu — nền tảng marketing cho doanh nghiệp hiện đại",
        "lead": "PTT tích hợp công nghệ tự động hóa, AI cá nhân hóa và dữ liệu thực tế để tối ưu hóa chiến dịch marketing — cho mọi ngành, đo được KPI rõ ràng.",
        "cta_primary_label": "Nhận tư vấn miễn phí",
        "cta_primary_href": "#contact",
        "cta_ghost_label": "Xem case study",
        "cta_ghost_href": "#projects",
    },
    {
        "image_url": "https://images.unsplash.com/photo-1551288049-bebda4e38f71?auto=format&fit=crop&w=1920&q=80",
        "kicker": "AI Personalization · Content Automation · Real-time Optimization",
        "heading": "AI tạo nội dung, cá nhân hóa theo ngữ cảnh — tự động hóa từ đầu đến cuối",
        "lead": "Từ brief chiến lược đến nội dung phân phối — PTT tự động hóa toàn bộ luồng sáng tạo và truyền thông, giúp doanh nghiệp tiết kiệm nguồn lực và nâng cao trải nghiệm khách hàng.",
        "cta_primary_label": "Xem demo hệ thống",
        "cta_primary_href": "#contact",
        "cta_ghost_label": "",
        "cta_ghost_href": "",
    },
    {
        "image_url": "https://images.unsplash.com/photo-1560179707-f14e90ef3623?auto=format&fit=crop&w=1920&q=80",
        "kicker": "Data-Driven · ROI Tracking · Business Intelligence",
        "heading": "Dữ liệu thực tế, KPI minh bạch — tối ưu hiệu suất kinh doanh theo thời gian thực",
        "lead": "Chúng tôi không báo cáo lượt tiếp cận — chúng tôi đo hiệu quả chuyển đổi, tối ưu chi phí và tăng trưởng doanh thu cụ thể cho từng doanh nghiệp.",
        "cta_primary_label": "Đánh giá hệ thống marketing — miễn phí",
        "cta_primary_href": "#contact",
        "cta_ghost_label": "",
        "cta_ghost_href": "",
    },
]

# Landing — dữ liệu tham chiếu KHTN / KHQT / CSKH (dùng trên CRM /crm/marketing-plan/segment/…)
MARKETING_LIFECYCLE_SLUGS: tuple[str, ...] = ("khtn", "khqt", "cskh")
MARKETING_LIFECYCLE_PAGES: dict[str, dict[str, Any]] = {
    "khtn": {
        "abbr": "KHTN",
        "title": "Khách hàng tiềm năng",
        "lead": (
            "Tập khách chưa mua hoặc chưa tương tác sâu — cần xác định, thu hút và đủ điểm (scoring) "
            "trước khi chuyển sang giai đoạn quan tâm."
        ),
        "bullets": [
            "Thu thập & làm giàu dữ liệu: form, landing, event, quảng cáo, SEO/AEO, social.",
            "Phân khúc theo nhu cầu, hành vi, ngành — ưu tiên kênh và thông điệp phù hợp.",
            "Đặt mục tiêu đo lường rõ: CPL, tỷ lệ MQL/SQL, chi phí / lead, chất lượng nguồn.",
        ],
    },
    "khqt": {
        "abbr": "KHQT",
        "title": "Khách hàng quan tâm",
        "lead": (
            "Khách đã thể hiện ý định (tải tài liệu, đăng ký demo, nhắn hỏi, vào funnel) "
            "— cần nuôi dưỡng và làm rõ tiêu chí mua."
        ),
        "bullets": [
            "Chuỗi nurture: email, remarketing, nội dung theo vai trò — giảm ma sát trên đường mua.",
            "Đồng bộ sales & marketing: SLA chốt lead, ghi CRM, hội thoại có ngữ cảnh.",
            "Bằng chứng xã hội & case study giúp rút ngắn thời gian cân nhắc.",
        ],
    },
    "cskh": {
        "abbr": "CSKH",
        "title": "Chăm sóc khách hàng",
        "lead": (
            "Sau chuyển đổi: giữ chân, hỗ trợ, mở rộng giá trị — đặt nền cho giới thiệu & tái mua."
        ),
        "bullets": [
            "Onboarding, kênh hỗ trợ đa điểm, thời gian phản hồi và chất lượng giải quyết vấn đề.",
            "Chương trình loyalty, cộng đồng, nội dung giá trị cho khách hiện hữu.",
            "Thu thập NPS/CSAT, vòng phản hồi sản phẩm — giảm churn, tăng LTV.",
        ],
    },
}

# Quy trình 5 bước — chỉ hiển thị trên CRM trang KHTN (/crm/marketing-plan/segment/khtn), popup chi tiết từng bước.
KHTN_PIPELINE_STEPS: list[dict[str, Any]] = [
    {
        "id": 1,
        "label": "Bước 1. Nghiên cứu thị trường",
        "lead": "Làm rõ bối cảnh cạnh tranh, nhu cầu và rủi ro trước khi đổ ngân sách vào thu hút lead.",
        "points": [
            "Khung TAM / SAM / SOM hoặc ước lượng quy mô đủ dùng cho giai đoạn hiện tại (nguồn: báo cáo ngành, sàn, khảo sát).",
            "Đối thủ trực tiếp / gián tiếp: định vị, giá, kênh, thông điệp — bảng so sánh ngắn.",
            "Xu hướng & yếu tố vĩ mô: chính sách, mùa vụ, công nghệ ảnh hưởng tới cầu.",
            "Deliverable gợi ý: 1 trang tóm tắt insight + danh sách giả định cần kiểm chứng.",
        ],
    },
    {
        "id": 2,
        "label": "Bước 2. Phân khúc thị trường",
        "lead": "Chia thị trường thành các nhóm đồng nhất để tập trung nguồn lực.",
        "points": [
            "Tiêu chí phân khúc: địa lý, quy mô, ngành, hành vi mua, mức giá / willingness-to-pay.",
            "Ưu tiên phân khúc theo tiềm năng (size × khả năng thắng × phù hợp sản phẩm).",
            "Tránh phân khúc quá rộng hoặc quá nhiều phân khúc song song khi chưa đủ dữ liệu.",
            "Deliverable: ma trận phân khúc + 1–3 phân khúc ưu tiên kỳ này.",
        ],
    },
    {
        "id": 3,
        "label": "Bước 3. Xác định khách hàng mục tiêu",
        "lead": "Chọn ICP / persona cụ thể để thống nhất brief nội dung và media.",
        "points": [
            "ICP: tiêu chí firmographic + technographic (B2B) hoặc demographic + lifestyle (B2C).",
            "Persona: vai trò quyết định, pain, kênh tin cậy, phản đối thường gặp.",
            "Jobs-to-be-done: khách “thuê” sản phẩm để làm việc gì — làm rõ offer.",
            "Deliverable: 1 trang persona chính (+ phụ nếu cần) dùng chung team sales & marketing.",
        ],
    },
    {
        "id": 4,
        "label": "Bước 4. Lựa chọn chiến lược tiếp cận",
        "lead": "Thiết kế mix kênh, thông điệp và funnel phù hợp phân khúc đã chọn.",
        "points": [
            "Chiến lược tổng thể: inbound / outbound / partner / community — chọn trọng tâm.",
            "Thông điệp & proof: USP, chứng cứ, case, CTAs nhất quán điểm chạm.",
            "Kế hoạch chạm: tần suất, ngân sách theo kênh, KPI từng giai đoạn.",
            "Deliverable: bản phác thảo chiến dịch / media plan 1 trang (có thể chỉnh sau đo lường).",
        ],
    },
    {
        "id": 5,
        "label": "Bước 5. Thử nghiệm quảng cáo trên thị trường mục tiêu",
        "lead": "Chạy thử có kiểm soát, đo kết quả, rút kinh nghiệm trước khi scale.",
        "points": [
            "Thiết kế thử nghiệm: giả thuyết, biến số (creative / audience / landing), ngân sách pilot.",
            "Danh mục thực hành trong form (bổ sung): công cụ vận hành — Phương pháp thử có kiểm soát — Thống kê & chỉ số / ngưỡng ra quyết định.",
            "Phân tích chi tiết kết quả pilot: định lượng & ý nghĩa thống kê, breakdown phân khúc, creative + funnel, định tính (sentiment), so sánh đối chiếu, bài học/rủi ro, khung quyết định.",
            "Theo dõi: CPL, CPA, CTR, CVR, chất lượng lead — so với ngưỡng mục tiêu.",
            "Quy tắc scale: chỉ tăng ngân sách khi đã có tín hiệu ổn định theo cửa sổ thời gian.",
            "Deliverable: báo cáo ngắn sau pilot + quyết định: dừng / tối ưu / nhân rộng.",
        ],
    },
]


def _tel_href_from_display(phone: str) -> str:
    """Chuỗi SĐT hiển thị (VN) → href tel: cho nút gọi nhanh."""
    digits = re.sub(r"\D", "", (phone or "").strip())
    if not digits:
        return ""
    if digits.startswith("84") and len(digits) >= 10:
        return f"tel:+{digits}"
    if digits.startswith("0") and len(digits) >= 9:
        return f"tel:+84{digits[1:]}"
    if len(digits) >= 8:
        return f"tel:+{digits}"
    return f"tel:+{digits}" if digits else ""


def _settings_for_public_pages() -> dict[str, str]:
    return {**DEFAULT_SETTINGS_NAV_MEGA_BANNER, **DEFAULT_SETTINGS_FOOTER, **fetch_settings()}

load_dotenv(BASE_DIR / ".env")
load_dotenv(BASE_DIR.parent / ".env")

from ptt_observability import bind_correlation_id, init_observability

app = Flask(__name__)
init_observability(component="flask", app=app)
app.config["MAX_CONTENT_LENGTH"] = MAX_VIDEO_UPLOAD_BYTES


@app.before_request
def _ptt_bind_correlation_id() -> None:
    cid = (
        request.headers.get("X-Correlation-Id")
        or request.headers.get("X-Request-Id")
        or str(uuid.uuid4())
    )
    bind_correlation_id(cid)


def _flask_secret_key() -> str:
    """Chữ ký session: ưu tiên FLASK_SECRET_KEY, rồi SECRET_KEY (.env thường chỉ có key này). Không dùng chuỗi rỗng."""
    for k in ("FLASK_SECRET_KEY", "SECRET_KEY"):
        raw = os.getenv(k)
        if isinstance(raw, str) and raw.strip():
            return raw.strip()
    return "dev-ptt-career-key-change-in-production"


app.secret_key = _flask_secret_key()
# Tránh trùng cookie "session" với app Flask/Python khác trên cùng localhost (ví dụ admin cổng khác) → hỏng chữ ký, không vào được CRM/CMS.
app.config["SESSION_COOKIE_NAME"] = "ptt_session"
app.config.setdefault("SESSION_COOKIE_HTTPONLY", True)
app.config.setdefault("SESSION_COOKIE_SAMESITE", "Lax")
if os.getenv("SESSION_COOKIE_SECURE", "").strip().lower() in {"1", "true", "yes"}:
    app.config["SESSION_COOKIE_SECURE"] = True
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=14)

ADMIN_SESSION_KEY = "_ptt_admin_ok"
CMS_ROLE_SESSION_KEY = "_ptt_cms_role"
CMS_USER_SESSION_KEY = "_ptt_cms_username"
CMS_POSITION_SESSION_KEY = "_ptt_cms_position_id"
STAFF_SESSION_KEY = "_ptt_staff_ok"
STAFF_ID_SESSION_KEY = "_ptt_staff_id"
STAFF_NAME_SESSION_KEY = "_ptt_staff_name"


def _admin_expected_credentials() -> tuple[str, str]:
    """Đọc user/mật khẩu từ env (ADMIN_USERNAME, ADMIN_PASSWORD); mặc định dev dễ vào được."""
    u = (os.getenv("ADMIN_USERNAME") or "admin").strip()
    pw = os.getenv("ADMIN_PASSWORD")
    pw = pw.strip() if isinstance(pw, str) else ""
    if not pw:
        pw = os.getenv("PTT_ADMIN_PASSWORD", "").strip()
    if not pw:
        pw = "changeme"  # dev — đặt ADMIN_PASSWORD trong .env khi triển khai thật
    return u, pw


def _admin_logged_in() -> bool:
    return bool(session.get(ADMIN_SESSION_KEY))


def _staff_logged_in() -> bool:
    return bool(session.get(STAFF_SESSION_KEY)) and _staff_session_id() is not None


def _staff_session_id() -> int | None:
    raw = session.get(STAFF_ID_SESSION_KEY)
    try:
        sid = int(raw)
        return sid if sid > 0 else None
    except (TypeError, ValueError):
        return None


def _staff_session_name() -> str:
    return str(session.get(STAFF_NAME_SESSION_KEY) or "").strip()


def _internal_logged_in() -> bool:
    return _admin_logged_in() or _staff_logged_in()


def _crm_effective_staff_id() -> int | None:
    """Nhân viên đăng nhập → id của họ; admin CMS → None (xem/tạo toàn bộ)."""
    if _admin_logged_in():
        return None
    if _staff_logged_in():
        return _staff_session_id()
    return None


def _crm_staff_portal_active() -> bool:
    """Chế độ portal nhân viên — không áp dụng khi admin CMS đang đăng nhập."""
    return _staff_logged_in() and not _admin_logged_in()


def _crm_case_assigned_to_staff(conn: sqlite3.Connection, case_id: int, staff_id: int) -> bool:
    row = conn.execute(
        "SELECT assigned_staff_id FROM crm_cases WHERE id = ?",
        (case_id,),
    ).fetchone()
    if row is None:
        return False
    aid = row["assigned_staff_id"]
    try:
        return int(aid) == int(staff_id)
    except (TypeError, ValueError):
        return False


def _crm_forbid_staff_case() -> Any:
    return jsonify({"error": "Bạn chỉ xem và cập nhật khách hàng được gán cho mình."}), 403


def _ensure_crm_session_html() -> Any | None:
    if _internal_logged_in():
        return None
    return redirect(url_for("admin_login", next=request.path or "/crm"))


def _ensure_admin_only_html() -> Any | None:
    if _admin_logged_in():
        return None
    if _staff_logged_in():
        return redirect(url_for("crm_staff_home"))
    return redirect(url_for("admin_login", next=request.path or "/"))


def _cms_session_role() -> str:
    return str(session.get(CMS_ROLE_SESSION_KEY) or "super_admin").strip() or "super_admin"


def _cms_session_username() -> str:
    return str(session.get(CMS_USER_SESSION_KEY) or "").strip()


def _admin_ui_display() -> tuple[str, str]:
    """Tên và vai trò hiển thị trên topbar admin."""
    if _staff_logged_in() and not _admin_logged_in():
        return _staff_session_name() or "Nhân viên", "Nhân viên CSKH"
    if not _admin_logged_in():
        return "", ""
    user = _cms_session_username() or "Admin"
    role_code = _cms_session_role()
    if role_code == "super_admin":
        return user, "Quản trị viên"
    try:
        with get_connection() as conn:
            row = conn.execute(
                "SELECT name FROM cms_roles WHERE code = ?",
                (role_code,),
            ).fetchone()
        role = str(row["name"]) if row else role_code
    except Exception:
        role = role_code
    return user, role


def _start_admin_session(
    username: str,
    role_code: str,
    *,
    position_id: int | None = None,
) -> None:
    session.clear()
    session[ADMIN_SESSION_KEY] = True
    session.permanent = True
    role = str(role_code or "viewer").strip() or "viewer"
    session[CMS_ROLE_SESSION_KEY] = role
    session[CMS_USER_SESSION_KEY] = username.strip()
    # super_admin / cms_admin không gắn chức vụ — tránh giới hạn section CRM
    if role in ("super_admin", "cms_admin"):
        position_id = None
    if position_id is not None and position_id > 0:
        session[CMS_POSITION_SESSION_KEY] = position_id
    else:
        session.pop(CMS_POSITION_SESSION_KEY, None)


def _verify_admin_current_password(
    conn: sqlite3.Connection, username: str, current_password: str
) -> bool:
    eu, epw = _admin_expected_credentials()
    return verify_unified_password(
        conn,
        username,
        current_password,
        env_username=eu,
        env_password=epw,
        const_eq=_const_eq_str,
    )


@app.context_processor
def inject_admin_ui_context() -> dict[str, str]:
    user, role = _admin_ui_display()
    return {"admin_ui_user": user, "admin_ui_role": role}


@app.context_processor
def inject_service_categories() -> dict:
    try:
        return {"service_categories": fetch_service_categories()}
    except Exception:
        return {"service_categories": []}


@app.context_processor
def inject_public_site_settings() -> dict:
    """Settings + tel href cho footer / header trên trang công khai."""
    path = (request.path or "").split("?")[0]
    if path.startswith(("/crm", "/admin", "/cms", "/api")):
        return {}
    try:
        settings = _settings_for_public_pages()
        return {
            "settings": settings,
            "contact_tel_href": _tel_href_from_display(settings.get("contact_phone", "")),
        }
    except Exception:
        return {}


def _admin_full_access() -> bool:
    """super_admin / cms_admin — bỏ qua giới hạn menu CMS và section CRM."""
    if not _admin_logged_in():
        return False
    return _cms_session_role() in ("super_admin", "cms_admin")


def _cms_is_super_admin() -> bool:
    return _cms_session_role() == "super_admin" or _admin_full_access()


def _cms_load_role_grants(conn: sqlite3.Connection, role_code: str) -> dict[str, list[str]]:
    rows = conn.execute(
        """
        SELECT module_id, action FROM cms_role_permissions
        WHERE role_code = ?
        """,
        (role_code,),
    ).fetchall()
    if not rows:
        if role_grants_were_customized(conn, role_code):
            return {mid: [] for mid in CMS_MODULE_IDS}
        return default_grants_for_role(role_code)
    grants: dict[str, set[str]] = {mid: set() for mid in CMS_MODULE_IDS}
    for r in rows:
        mid = str(r["module_id"])
        act = str(r["action"])
        if mid in grants and act in CMS_ACTIONS:
            grants[mid].add(act)
    return {mid: sorted(acts) for mid, acts in grants.items()}


def _cms_current_grants(conn: sqlite3.Connection | None = None) -> dict[str, list[str]]:
    if conn is not None:
        return _cms_load_role_grants(conn, _cms_session_role())
    with get_connection() as c:
        return _cms_load_role_grants(c, _cms_session_role())


def _cms_can(module_id: str, action: str, conn: sqlite3.Connection | None = None) -> bool:
    if _admin_full_access():
        return True
    if _cms_is_super_admin():
        return True
    grants = _cms_current_grants(conn)
    return role_can(grants, module_id, action)


def _cms_forbidden_json(module_id: str, action: str) -> Any:
    return (
        jsonify(
            {
                "error": f"Không có quyền «{action}» trên hạng mục «{module_id}».",
                "module_id": module_id,
                "action": action,
                "role": _cms_session_role(),
            }
        ),
        403,
    )


def _cms_session_position_id() -> int | None:
    raw = session.get(CMS_POSITION_SESSION_KEY)
    try:
        pid = int(raw)
        return pid if pid > 0 else None
    except (TypeError, ValueError):
        return None


def _cms_load_position_grants(conn: sqlite3.Connection, position_id: int) -> dict[str, list[str]]:
    pos_row = conn.execute(
        "SELECT code FROM crm_positions WHERE id = ? AND active = 1",
        (position_id,),
    ).fetchone()
    code = str(pos_row["code"]) if pos_row else ""
    rows = conn.execute(
        """
        SELECT section_id, action FROM crm_position_section_permissions
        WHERE position_id = ?
        """,
        (position_id,),
    ).fetchall()
    customized_row = conn.execute(
        "SELECT grants_customized FROM crm_positions WHERE id = ? AND active = 1",
        (position_id,),
    ).fetchone()
    customized = bool(customized_row and int(customized_row["grants_customized"] or 0))
    if not rows:
        if customized:
            return {sid: [] for sid in ADMIN_CRM_PERMISSION_IDS}
        return default_grants_for_position(code)
    grants: dict[str, set[str]] = {sid: set() for sid in ADMIN_CRM_PERMISSION_IDS}
    for r in rows:
        sid = str(r["section_id"])
        act = str(r["action"])
        if sid in grants and act in CMS_ACTIONS:
            grants[sid].add(act)
    return {sid: sorted(acts) for sid, acts in grants.items()}


def _cms_current_position_grants(conn: sqlite3.Connection | None = None) -> dict[str, list[str]]:
    pid = _cms_session_position_id()
    if pid is None:
        return {}
    if conn is not None:
        return _cms_load_position_grants(conn, pid)
    with get_connection() as c:
        return _cms_load_position_grants(c, pid)


def _patch_position_grants_hdsd(grants: dict[str, list[str]]) -> dict[str, list[str]]:
    """HDSD — ai đăng nhập admin đều thấy menu & đọc tài liệu (export vẫn theo ma trận)."""
    out = {k: list(v) for k, v in grants.items()}
    acts = set(out.get("crm_hdsd") or [])
    acts.add("view")
    out["crm_hdsd"] = sorted(acts)
    return out


def _admin_section_can(section_id: str, action: str, conn: sqlite3.Connection | None = None) -> bool:
    """Quyền section Admin/CRM: admin → full; super_admin → full; chức vụ (staff portal)."""
    if _admin_full_access():
        return True
    if _cms_is_super_admin():
        return True
    sid = str(section_id or "").strip()
    act = str(action or "").strip().lower()
    if sid == "crm_hdsd" and act == "view" and _admin_logged_in():
        return True
    if sid in CMS_MODULE_IDS:
        if not _cms_can(sid, act, conn):
            return False
    if sid in CRM_UI_BUTTON_BY_ID:
        pid = _cms_session_position_id()
        if pid is None:
            return True
        if conn is not None:
            grants = _cms_load_position_grants(conn, pid)
        else:
            grants = _cms_current_position_grants()
        return ui_button_can(grants, sid)
    if sid not in ADMIN_CRM_SECTION_IDS:
        return sid in CMS_MODULE_IDS
    pid = _cms_session_position_id()
    if pid is None:
        return True
    if conn is not None:
        grants = _cms_load_position_grants(conn, pid)
    else:
        grants = _cms_current_position_grants()
    return position_can(grants, sid, act)


def _admin_button_can(button_id: str, conn: sqlite3.Connection | None = None) -> bool:
    if _admin_full_access() or _cms_is_super_admin():
        return True
    bid = str(button_id or "").strip()
    if bid not in CRM_UI_BUTTON_BY_ID:
        return False
    pid = _cms_session_position_id()
    if pid is None:
        return True
    if conn is not None:
        grants = _cms_load_position_grants(conn, pid)
    else:
        grants = _cms_current_position_grants()
    return ui_button_can(grants, bid)


def _admin_section_can_customer_write(action: str, conn: sqlite3.Connection | None = None) -> bool:
    """Tạo/sửa khách hàng: quyền trang Khách hàng hoặc Tạo yêu cầu CSKH."""
    if _admin_full_access():
        return True
    act = str(action or "").strip().lower()
    if act == "create":
        return _admin_section_can("crm_board_customers", "create", conn) or _admin_section_can(
            "crm_board_create", "create", conn
        )
    if act == "edit":
        return _admin_section_can("crm_board_customers", "edit", conn) or _admin_section_can(
            "crm_board_create", "create", conn
        )
    return _admin_section_can("crm_board_customers", act, conn)


def _admin_section_forbidden_json(section_id: str, action: str) -> Any:
    return (
        jsonify(
            {
                "error": f"Chức vụ của bạn không có quyền «{action}» trên «{section_id}».",
                "section_id": section_id,
                "action": action,
                "position_id": _cms_session_position_id(),
            }
        ),
        403,
    )


def _admin_grants_bootstrap_json(conn: sqlite3.Connection | None = None) -> str:
    ui_buttons_payload = {
        bid: {
            "parent_section": meta["parent_section"],
            "requires_action": meta["requires_action"],
        }
        for bid, meta in CRM_UI_BUTTON_BY_ID.items()
    }
    permission_ids = sorted(ADMIN_CRM_PERMISSION_IDS)
    if conn is None:
        with get_connection() as c:
            return _admin_grants_bootstrap_json(c)
    role = _cms_session_role()
    if _admin_full_access():
        payload = {
            "cms_grants": {mid: list(CMS_ACTIONS) for mid in CMS_MODULE_IDS},
            "position_grants": {sid: list(CMS_ACTIONS) for sid in permission_ids},
            "is_super_admin": role == "super_admin",
            "is_full_access": True,
            "is_admin_session": True,
            "position_id": None,
            "crm_section_ids": permission_ids,
            "ui_buttons": ui_buttons_payload,
        }
        return json.dumps(payload, ensure_ascii=False)
    if _admin_logged_in():
        cms_g = _cms_load_role_grants(conn, role)
        pos_g = _patch_position_grants_hdsd(_cms_current_position_grants(conn))
        payload = {
            "cms_grants": cms_g,
            "position_grants": pos_g,
            "is_super_admin": False,
            "is_full_access": False,
            "is_admin_session": True,
            "position_id": _cms_session_position_id(),
            "crm_section_ids": permission_ids,
            "ui_buttons": ui_buttons_payload,
        }
        return json.dumps(payload, ensure_ascii=False)
    cms_g = _cms_load_role_grants(conn, role)
    pos_g = _patch_position_grants_hdsd(_cms_current_position_grants(conn))
    payload = {
        "cms_grants": cms_g,
        "position_grants": pos_g,
        "is_super_admin": role == "super_admin",
        "is_full_access": False,
        "position_id": _cms_session_position_id(),
        "crm_section_ids": permission_ids,
        "ui_buttons": ui_buttons_payload,
    }
    return json.dumps(payload, ensure_ascii=False)


def _admin_grants_template_kwargs(conn: sqlite3.Connection | None = None) -> dict[str, Any]:
    if conn is None:
        with get_connection() as c:
            return _admin_grants_template_kwargs(c)
    return {
        "admin_grants_json": _admin_grants_bootstrap_json(conn),
        "admin_full_access": _admin_full_access(),
        "crm_leads_only_ui": _crm_leads_only_ui(conn),
    }


def _ensure_admin_session_html() -> Any | None:
    """Admin/CMS/CRM quản trị — không dùng cho portal nhân viên."""
    if _admin_logged_in():
        return None
    if _staff_logged_in():
        return redirect(url_for("crm_staff_home"))
    return redirect(url_for("admin_login", next=request.path or "/"))


def _safe_internal_redirect_path(next_path: str | None, *, default_endpoint: str) -> str:
    if not next_path or not isinstance(next_path, str):
        return url_for(default_endpoint)
    p = next_path.strip()
    if not p.startswith("/") or p.startswith("//") or "\n" in p or "\\" in p:
        return url_for(default_endpoint)
    return p


def _const_eq_str(a: str, b: str) -> bool:
    """So sánh chuỗi với thời gian gần như không phụ thuộc độ dài (SHA-256 hex cố định)."""
    ha = hashlib.sha256((a or "").encode("utf-8")).hexdigest()
    hb = hashlib.sha256((b or "").encode("utf-8")).hexdigest()
    return secrets.compare_digest(ha, hb)


_FACEBOOK_WEBHOOK_PATH = "/api/crm/integration/webhooks/facebook"


def _facebook_webhook_path_normalized() -> str:
    return (request.path or "").rstrip("/") or "/"


def _is_facebook_webhook_path() -> bool:
    p = _facebook_webhook_path_normalized()
    return p == _FACEBOOK_WEBHOOK_PATH or p.startswith(_FACEBOOK_WEBHOOK_PATH + "/")


def _facebook_webhook_slug_from_path() -> str | None:
    p = _facebook_webhook_path_normalized()
    prefix = _FACEBOOK_WEBHOOK_PATH + "/"
    if not p.startswith(prefix):
        return None
    slug = p[len(prefix) :].strip("/")
    return slug or None


def _facebook_webhook_raw_body() -> bytes:
    cached = request.environ.get("PTT_FB_WEBHOOK_RAW")
    if isinstance(cached, (bytes, bytearray)):
        return bytes(cached)
    return request.get_data(cache=True)


def _facebook_webhook_signature_header() -> str | None:
    return (
        request.headers.get("X-Hub-Signature-256")
        or request.headers.get("X-Hub-Signature")
        or request.environ.get("HTTP_X_HUB_SIGNATURE_256")
        or request.environ.get("HTTP_X_HUB_SIGNATURE")
    )


@app.before_request
def _cache_facebook_webhook_raw_body():
    if request.method == "POST" and _is_facebook_webhook_path():
        request.environ["PTT_FB_WEBHOOK_RAW"] = request.get_data(cache=True)


@app.before_request
def admin_auth_guard():
    """Bảo vệ /admin (trừ đăng nhập), /cms, /crm và API nội bộ."""
    path = request.path or ""
    if path.startswith("/static/"):
        return None
    if request.endpoint in {"healthz_ptt", "robots_txt", "sitemap_xml", "static"}:
        return None
    if path == "/admin/login":
        return None
    if path == "/admin/logout":
        return None
    if path == "/api/crm/attendance/device":
        return None
    if path.startswith("/api/crm/integration/webhooks/"):
        return None
    if path.startswith("/api/v1/webhooks/") or path == "/api/v1/channels":
        return None
    if path == "/api/crm/agency/sla-sync-cron":
        return None
    if path == "/api/crm/integration/marketing/ingest":
        return None
    if path == "/api/crm/integration/facebook/sync-cron":
        return None
    if path == "/api/crm/finance/kpi-alert-cron":
        return None
    if path == "/api/crm/owner-weekly/alert-cron":
        return None
    if path.startswith("/iclock/"):
        return None

    need_wall = False
    respond_json = False

    if path == "/admin" or path.startswith("/admin/"):
        need_wall = True
    elif path == "/cms" or path.startswith("/cms/"):
        need_wall = True
    elif path == "/crm" or path.startswith("/crm/"):
        need_wall = True
    elif path.startswith("/api/crm"):
        need_wall = True
        respond_json = True
    elif path.startswith("/api/v1/clients") or path.startswith("/api/v1/jobs") or path.startswith(
        "/api/v1/notifications"
    ) or path.startswith("/api/v1/kpi-definitions") or path == "/health/worker":
        need_wall = True
        respond_json = True
    elif path.startswith("/api/projects") or path.startswith("/api/news"):
        need_wall = True
        respond_json = True
    elif path in ("/api/settings", "/api/services"):
        need_wall = True
        respond_json = True
    elif path.startswith("/api/cms/marketing-chat"):
        need_wall = True
        respond_json = True
    elif path.startswith("/api/cms/permissions") or path.startswith("/api/cms/admin-users"):
        need_wall = True
        respond_json = True
    elif path.startswith("/api/cms/live-chat"):
        need_wall = True
        respond_json = True
    elif path.startswith("/api/cms/media"):
        need_wall = True
        respond_json = True
    elif path == "/account/password" or path.startswith("/account/"):
        need_wall = True
    elif path.startswith("/api/account/"):
        need_wall = True
        respond_json = True

    if not need_wall:
        return None

    if _admin_logged_in():
        if _crm_leads_only_ui() and not _crm_leads_only_path_allowed(path):
            if respond_json or path.startswith("/api/"):
                return jsonify(
                    {"error": "Chỉ được truy cập Quản lý Lead.", "redirect": url_for("crm_leads_page")}
                ), 403
            return redirect(url_for("crm_leads_page"))
        return None

    if _staff_logged_in():
        staff_id = _staff_session_id()
        if staff_portal_html_allowed(path):
            return None
        if path.startswith("/api/account/"):
            return None
        if path.startswith("/crm/"):
            if respond_json:
                return jsonify({"error": "Chỉ quản trị viên mới truy cập mục này.", "redirect": url_for("crm_board")}), 403
            return redirect(url_for("crm_board"))
        if path.startswith("/api/crm") and staff_crm_api_allowed_extended(request.method, path):
            if re.match(r"^/api/crm/staff/(\d+)/workspace$", path):
                m = re.match(r"^/api/crm/staff/(\d+)/workspace$", path)
                if m and int(m.group(1)) != int(staff_id):
                    return jsonify({"error": "Không xem được workspace nhân viên khác."}), 403
            return None
        if respond_json:
            return jsonify({"error": "Không có quyền.", "login": url_for("admin_login")}), 403
        return redirect(url_for("crm_board"))

    if respond_json:
        return (
            jsonify(
                {
                    "error": "Chưa đăng nhập.",
                    "login": url_for("admin_login"),
                }
            ),
            401,
        )
    return redirect(url_for("admin_login", next=path))


@app.after_request
def _no_store_internal_pages(resp: Any) -> Any:
    """Tránh trình duyệt / proxy cache trang nội bộ và API CRM (back-forward cache sau đăng xuất)."""
    path = request.path or ""
    ep = request.endpoint or ""
    sensitive = (
        ep in ("admin", "cms", "crm_board", "crm_staff_page", "crm_payroll_page", "crm_kpi_page", "admin_login")
        or path.startswith("/api/crm")
        or path.startswith("/crm")
    )
    if sensitive:
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, private"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
        return resp
    if (
        resp.status_code == 200
        and path.startswith("/static/")
        and not path.endswith("/")  # thư mục
    ):
        raw_sec = os.getenv("PTT_STATIC_CACHE_MAX_AGE", "604800").strip()
        try:
            max_age = max(0, int(raw_sec))
        except ValueError:
            max_age = 604_800  # 7 ngày; URL asset có ?v=mtime để đổi bản không cần bust tay
        if max_age > 0:
            # Ghi đè Cache-Control mặc định của Flask (debug đặt no-cache — setdefault không thay được).
            resp.headers["Cache-Control"] = f"public, max-age={max_age}"
    return resp


@app.template_global()
def static_css_file() -> str:
    """Luôn phục vụ styles.css — deploy sửa CSS rồi restart, không cần build styles.min.css."""
    return "styles.css"


@app.template_global()
def static_v(path_under_static: str) -> str:
    """Tham số ?v= cho file tĩnh — tránh cache trình duyệt khi thay ảnh hoặc asset."""
    p = BASE_DIR / "static" / path_under_static.replace("\\", "/").lstrip("/")
    try:
        return str(int(p.stat().st_mtime))
    except OSError:
        return "0"


@app.template_global()
def webp_static_file(path_under_static: str) -> str | None:
    """Đường dẫn tới file .webp cùng tên trong static/, nếu đã được tạo."""
    norm = path_under_static.replace("\\", "/").lstrip("/")
    base = BASE_DIR / "static" / norm
    if base.suffix.lower() not in {".png", ".jpg", ".jpeg"}:
        return None
    webp_rel = Path(norm).with_suffix(".webp").as_posix()
    if (BASE_DIR / "static" / webp_rel).is_file():
        return webp_rel
    return None


@app.template_global()
def static_js_file(name: str) -> str:
    """Luôn phục vụ file .js nguồn — deploy sửa static/*.js rồi restart, không cần build *.min.js."""
    norm = name.replace("\\", "/").lstrip("/")
    if not norm.endswith(".js") or "/" in norm:
        return norm
    return norm


ALLOWED_CV_EXTENSIONS = frozenset({".pdf", ".doc", ".docx"})
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# Nhóm dịch vụ theo cấu trúc tương tự bngagency.vn (AEO, SEO, thiết kế, quảng cáo, nội dung) — rút gọn so với bản cũ
DEFAULT_SERVICE_CATEGORIES = [
    {
        "title": "Tìm kiếm tự nhiên",
        "items": [
            {
                "slug": "dich-vu-aeo",
                "name": "Dịch vụ AEO",
                "summary": "Tối ưu nội dung theo hướng Answer Engine: giúp thương hiệu xuất hiện trong câu trả lời từ công cụ tìm kiếm và trợ lý AI.",
                "highlights": [
                    "Cấu trúc câu hỏi – câu trả lời, FAQ và schema phù hợp AEO.",
                    "Ưu tiên mục tiêu hiển thị so với chỉ tối ưu từ khóa truyền thống.",
                    "Bám sát hành trình tìm hiểu của người dùng trên Google và dịch vụ tìm bằng AI.",
                    "Kế hoạch nội dung theo từng giai đoạn và theo dõi hiệu quả theo từng khu vực.",
                ],
            },
            {
                "slug": "dich-vu-seo-tong-the",
                "name": "Dịch vụ SEO tổng thể",
                "summary": "Chiến lược SEO từ kỹ thuật, nội dung đến liên kết, giúp tăng thứ hạng và lưu lượng từ tìm kiếm tự nhiên.",
                "highlights": [
                    "Phân tích từ khóa, đối thủ và cơ cấu site.",
                    "Kế hoạch on-page, nội dung và tối ưu tốc độ / Core Web Vitals.",
                    "Chiến lược backlink bền vững, phù hợp chính sách Google.",
                    "Báo cáo theo từ thứ hạng, traffic và mục tiêu chuyển đổi.",
                ],
            },
            {
                "slug": "dich-vu-seo-local",
                "name": "Dịch vụ SEO local",
                "summary": "Tăng mức độ hiển thị cho doanh nghiệp có địa điểm: Google Business Profile, tìm kiếm theo vùng và bản đồ.",
                "highlights": [
                    "Tối ưu hồ sơ doanh nghiệp trên bản đồ và tìm kiếm địa phương.",
                    "Nội dung và từ khóa theo từng khu vực / chi nhánh.",
                    "Quản lý đánh giá, thông tin NAP thống nhất đa nền tảng.",
                    "Báo cáo số lượt xem, cuộc gọi, chỉ đường và form liên hệ từ local.",
                ],
            },
            {
                "slug": "dich-vu-seo-audit",
                "name": "Dịch vụ SEO Audit",
                "summary": "Rà soát toàn diện website, chỉ rõ ưu tiên kỹ thuật và nội dung cần sửa để cải thiện index và ranking.",
                "highlights": [
                    "Quét kỹ thuật: crawl, index, cấu trúc, redirect, sitemap.",
                    "Đánh giá on-page, trùng lặp nội dung, internal link.",
                    "Bảng ưu tiên hành động theo tác động và công sức.",
                    "Trình bày kết quả theo tài liệu và nếu cần hỗ trợ triển khai.",
                ],
            },
            {
                "slug": "dich-vu-quan-tri-website",
                "name": "Dịch vụ quản trị website",
                "summary": "Vận hành, cập nhật và tối ưu an toàn cho website, đảm bảo site ổn định phục vụ marketing và tìm kiếm.",
                "highlights": [
                    "Cập nhật nội dung, cài đặt, plugin/core theo lịch.",
                    "Sao lưu, cập nhật bảo mật, giảm thiểu rủi ro downtime.",
                    "Hỗ trợ tích hợp tag, form, tracking cơ bản.",
                    "Lịch bảo trì rõ ràng, SLA theo từng gói.",
                ],
            },
        ],
    },
    {
        "title": "Thiết kế",
        "items": [
            {
                "slug": "thiet-ke-website",
                "name": "Thiết kế website",
                "summary": "Thiết kế giao diện website chuyên nghiệp, tối ưu trải nghiệm người dùng và chuyển đổi cho thương hiệu.",
                "highlights": [
                    "Wireframe, UI theo brand guideline và từng ngành.",
                    "Thân thiện mobile, tốc độ và chuẩn cơ bản cho SEO kỹ thuật.",
                    "Dễ mở rộng mô-đun, form và kết nối công cụ marketing.",
                    "Bàn giao tài liệu và hướng dẫn cập nhật nội dung.",
                ],
            },
            {
                "slug": "thiet-ke-website-tron-goi",
                "name": "Thiết kế website trọn gói",
                "summary": "Gói trọn từ ý tưởng, thiết kế đến triển khai, phù hợp dự án cần một đầu mối rõ ràng.",
                "highlights": [
                    "Hợp đồng theo mốc: thiết kế, nội dung, tích hợp, go-live.",
                    "Phối hợp tích hợp hosting/domain và email công ty (khi cần).",
                    "Đồng bộ tracking, form liên hệ, pixel cơ bản.",
                    "Hỗ trợ bảo hành / chỉnh sửa theo giai đoạn bàn giao.",
                ],
            },
            {
                "slug": "thiet-ke-landing-page",
                "name": "Thiết kế landing page",
                "summary": "Trang đích tập trung chuyển đổi, phù hợp chiến dịch quảng cáo, ra mắt sản phẩm hoặc thu thập lead.",
                "highlights": [
                    "Bố cục tối ưu theo mục tiêu: form, gọi, chat.",
                    "A/B theo từng giai đoạn nếu đủ traffic.",
                    "Mã theo dõi chuyển đổi và tích hợp CRM / sheet.",
                    "Bàn giao nhanh, trùng với lịch media.",
                ],
            },
        ],
    },
    {
        "title": "Quảng cáo kỹ thuật số",
        "items": [
            {
                "slug": "quang-cao-facebook",
                "name": "Dịch vụ chạy quảng cáo Facebook",
                "summary": "Lập kế hoạch, triển khai và tối ưu quảng cáo trên hệ thống Meta: Facebook, Instagram.",
                "highlights": [
                    "Thiết lập tài khoản, pixel và sự kiện theo mục tiêu (traffic, lead, mua hàng).",
                    "Dựng tệp, creative và tối ưu theo CPA / ROAS.",
                    "Báo cáo theo tuần, đề xuất điều chỉnh budget và tệp.",
                    "Tư vấn kịch bản nội dung theo từng giai đoạn funnel.",
                ],
            },
            {
                "slug": "quang-cao-google",
                "name": "Dịch vụ chạy quảng cáo Google",
                "summary": "Tìm kiếm, hiển thị, Performance Max và mạng lưới lớn: tư vấn cấu trúc tài khoản và tối ưu theo mục tiêu chuyển đổi.",
                "highlights": [
                    "Cấu trúc campaign theo ngành hàng và vùng.",
                    "Từ khóa, bài trừ, remarketing, feed (nếu dùng shopping).",
                    "Bám sát conversion và giá mỗi chuyển đổi.",
                    "Báo cáo dễ đọc, kèm dự báo tối ưu ngân sách.",
                ],
            },
            {
                "slug": "thue-tai-khoan-quang-cao",
                "name": "Dịch vụ cho thuê tài khoản quảng cáo",
                "summary": "Hỗ trợ tài khoản quảng cáo phù hợp với từng mục tiêu, giảm rủi ro ngắt quyền; triển khai kèm quản trị theo hợp đồng.",
                "highlights": [
                    "Làm rõ điều khoản, phí và trách nhiệm từng bên trước chạy.",
                    "Cấu hình tài khoản, thẻ thanh toán theo hướng dẫn nền tảng.",
                    "Báo cáo minh bạch chi phí theo từng tháng / chiến dịch.",
                    "Ưu tiên hành vi kinh doanh sạch, tuân thủ chính sách nền tảng.",
                ],
            },
        ],
    },
    {
        "title": "Tiếp thị nội dung",
        "items": [
            {
                "slug": "tiep-thi-noi-dung",
                "name": "Tiếp thị nội dung",
                "summary": "Kế hoạch và sản xuất nội dung theo từng kênh (web, social, tài liệu) hướng tới tăng lưu lượng và lòng tin thương hiệu.",
                "highlights": [
                    "Lịch nội dung theo persona và hành trình mua hàng.",
                    "Bài dài, bài chuẩn SEO, bài chuyên sâu ngành.",
                    "Hạn chế nội dung chung chung, bám số liệu engagement và lead.",
                    "Có thể kết hợp cùng gói SEO / quảng cáo theo từng dự án.",
                ],
            },
        ],
    },
]

_OPTIONAL_SVC_LANDING: frozenset[str] = frozenset(
    {
        "tagline",
        "overview",
        "stats",
        "pillars",
        "outcomes",
        "deliverable_title",
        "stats_section_title",
        "pillars_section_title",
        "outcomes_section_title",
        "faq",
        "inline_cta",
        "cta_wide_title",
        "cta_wide_lead",
        "meta_description",
        "image_url",
        "hero_eyebrow",
        "hero_cta_primary_label",
        "hero_cta_phone_label",
        "price_label",
        "price_note",
    }
)


def _as_str_list(v: Any) -> list[str]:
    if v is None:
        return []
    if isinstance(v, str):
        t = v.strip()
        return [t] if t else []
    if isinstance(v, list):
        return [str(x).strip() for x in v if str(x).strip()]
    return []


def _as_stats(v: Any) -> list[dict[str, str]]:
    if not isinstance(v, list):
        return []
    out: list[dict[str, str]] = []
    for it in v:
        if not isinstance(it, dict):
            continue
        value = str(it.get("value", "")).strip()
        label = str(it.get("label", "")).strip()
        if value and label:
            out.append({"value": value, "label": label})
    return out[:6]


def _as_pillars(v: Any) -> list[dict[str, str]]:
    if not isinstance(v, list):
        return []
    out: list[dict[str, str]] = []
    for it in v:
        if not isinstance(it, dict):
            continue
        t = str(it.get("title", "")).strip()
        b = str(it.get("body", "")).strip()
        if t and b:
            out.append({"title": t, "body": b})
    return out[:6]


def _as_faq(v: Any) -> list[dict[str, str]]:
    if not isinstance(v, list):
        return []
    out: list[dict[str, str]] = []
    for it in v:
        if not isinstance(it, dict):
            continue
        q = it.get("q") or it.get("question")
        a = it.get("a") or it.get("answer")
        qs = str(q or "").strip()
        an = str(a or "").strip()
        if qs and an:
            out.append({"q": qs, "a": an})
    return out[:20]


def _merge_landing_extras(item: dict[str, Any], slug: str) -> dict[str, Any]:
    ex = deepcopy(DEFAULT_SERVICE_LANDING_EXTRAS.get(slug, {}))
    for k, v in item.items():
        if k in _OPTIONAL_SVC_LANDING and v is not None:
            if isinstance(v, str) and not v.strip():
                continue
            ex[k] = v
    return ex


def _finalize_landing(merged: dict[str, Any], summary: str) -> dict[str, Any]:
    out: dict[str, Any] = dict(merged)
    over = _as_str_list(out.get("overview"))
    s = (summary or "").strip()
    if not over and s:
        over = [s]
    out["overview"] = over
    out["stats"] = _as_stats(out.get("stats"))
    out["pillars"] = _as_pillars(out.get("pillars"))
    out["outcomes"] = _as_str_list(out.get("outcomes"))
    out["faq"] = _as_faq(out.get("faq"))
    out["tagline"] = str(out.get("tagline", "")).strip()
    for key in (
        "deliverable_title",
        "stats_section_title",
        "pillars_section_title",
        "outcomes_section_title",
        "inline_cta",
        "cta_wide_title",
        "cta_wide_lead",
        "meta_description",
        "image_url",
        "hero_eyebrow",
        "hero_cta_primary_label",
        "hero_cta_phone_label",
        "price_label",
        "price_note",
    ):
        v = out.get(key)
        out[key] = str(v).strip() if v is not None and v != "" else ""
    if not out.get("deliverable_title"):
        out["deliverable_title"] = "Nội dung & phạm vi triển khai"
    if not out.get("stats_section_title"):
        out["stats_section_title"] = "Mục tiêu & trọng tâm theo từng giai đoạn"
    if not out.get("pillars_section_title"):
        out["pillars_section_title"] = "Cách PTT triển khai"
    if not out.get("outcomes_section_title"):
        out["outcomes_section_title"] = "Kết quả hướng tới"
    return out


# Logo đối tác (tệp trong static/partner-logos/…)
PARTNER_LOGOS: list[dict[str, str]] = [
    {"name": "BIDV", "site": "https://bidv.com.vn", "logo_file": "partner-logos/admicro-dnhh/BIDV.png"},
    {"name": "VietinBank", "site": "https://vietinbank.vn", "logo_file": "partner-logos/admicro-dnhh/viettinbank1.png"},
    {"name": "Agribank", "site": "https://agribank.com.vn", "logo_file": "partner-logos/admicro-dnhh/agri.png"},
    {"name": "SHB", "site": "https://shb.com.vn", "logo_file": "partner-logos/admicro-dnhh/SHB.png"},
    {"name": "Habeco", "site": "https://habeco.com.vn", "logo_file": "partner-logos/admicro-dnhh/habeco.png"},
    {"name": "BIC", "site": "https://bic.vn", "logo_file": "partner-logos/admicro-dnhh/bic.png"},
    {"name": "ACV", "site": "https://acv.vn", "logo_file": "partner-logos/admicro-dnhh/acv.png"},
    {"name": "Sunshine Group", "site": "https://sunshinegroup.vn", "logo_file": "partner-logos/admicro-dnhh/sunshine.png"},
    {"name": "SYS", "site": "https://sysvietnam.vn", "logo_file": "partner-logos/admicro-dnhh/sys.png"},
]

DEFAULT_CAPABILITIES_ITEMS: list[str] = [
    "Marketing Automation",
    "AI Content & Personalization",
    "Data Analytics & Business Intelligence",
    "CRM & Lead Intelligence",
    "Paid Media đa kênh & Tối ưu ROI",
    "AI Agent & Customer Experience",
]


def _default_capabilities_items() -> list[dict[str, str]]:
    return [
        {"title": title, "icon": str(i % 6), "icon_url": ""}
        for i, title in enumerate(DEFAULT_CAPABILITIES_ITEMS)
    ]


def _normalize_capability_icon(icon_raw: str, index: int) -> str:
    icon = str(icon_raw or "").strip()
    if icon.isdigit() and 0 <= int(icon) <= 5:
        return icon
    return str(index % 6)


def _capabilities_items_for_landing(settings: dict[str, str]) -> list[dict[str, str]]:
    raw_json = (settings.get("capabilities_items_json") or "").strip()
    if raw_json:
        try:
            data = json.loads(raw_json)
            if isinstance(data, list) and data:
                out: list[dict[str, str]] = []
                for i, entry in enumerate(data):
                    if isinstance(entry, dict):
                        title = str(entry.get("title") or entry.get("text") or "").strip()
                        icon_url = str(entry.get("icon_url") or "").strip()
                        if icon_url and not (
                            icon_url.startswith("http") or icon_url.startswith("/")
                        ):
                            icon_url = ""
                        icon = _normalize_capability_icon(
                            str(entry.get("icon") or entry.get("icon_preset") or ""),
                            i,
                        )
                    elif isinstance(entry, str):
                        title = entry.strip()
                        icon = str(i % 6)
                        icon_url = ""
                    else:
                        continue
                    if title:
                        out.append({"title": title, "icon": icon, "icon_url": icon_url})
                if out:
                    return out
        except (json.JSONDecodeError, TypeError):
            pass
    lines_raw = (settings.get("capabilities_items_lines") or "").strip()
    if lines_raw:
        items = [ln.strip() for ln in lines_raw.splitlines() if ln.strip()]
        if items:
            return [
                {"title": title, "icon": str(i % 6), "icon_url": ""}
                for i, title in enumerate(items)
            ]
    return _default_capabilities_items()


def _partner_logos_for_landing(settings: dict[str, str]) -> list[dict[str, str]]:
    raw = (settings.get("partner_logos_json") or "").strip()
    if raw:
        try:
            data = json.loads(raw)
            if isinstance(data, list) and data:
                out: list[dict[str, str]] = []
                for item in data:
                    if not isinstance(item, dict):
                        continue
                    name = str(item.get("name") or "").strip()
                    site = str(item.get("site") or "#").strip() or "#"
                    logo_url = str(item.get("logo_url") or "").strip()
                    logo_file = str(item.get("logo_file") or "").strip()
                    if not name or (not logo_url and not logo_file):
                        continue
                    out.append(
                        {
                            "name": name,
                            "site": site,
                            "logo_url": logo_url,
                            "logo_file": logo_file,
                        }
                    )
                if out:
                    return out
        except (json.JSONDecodeError, TypeError):
            pass
    return PARTNER_LOGOS


def _partner_logos_to_editor_rows(logos: list[dict[str, str]]) -> list[dict[str, str]]:
    """Chuyển logo đang hiển thị (PARTNER_LOGOS hoặc JSON đã lưu) → hàng editor CMS."""
    rows: list[dict[str, str]] = []
    for item in logos:
        name = str(item.get("name") or "").strip()
        site = str(item.get("site") or "#").strip() or "#"
        logo_url = str(item.get("logo_url") or "").strip()
        if not logo_url:
            logo_file = str(item.get("logo_file") or "").strip().lstrip("/")
            if logo_file:
                logo_url = f"/static/{logo_file}"
        if name and logo_url:
            rows.append({"name": name, "site": site, "logo_url": logo_url})
    return rows


def _partner_logos_seed_for_cms(settings: dict[str, str]) -> list[dict[str, str]]:
    """Logo đối tác đang hiển thị trên web — dùng làm dữ liệu khởi tạo editor CMS."""
    merged = {**DEFAULT_SETTINGS_NAV_MEGA_BANNER, **DEFAULT_SETTINGS_FOOTER, **settings}
    return _partner_logos_to_editor_rows(_partner_logos_for_landing(merged))


def _merged_landing_settings() -> dict[str, str]:
    return {**DEFAULT_SETTINGS_NAV_MEGA_BANNER, **DEFAULT_SETTINGS_FOOTER, **fetch_settings()}


def _partner_logos_effective_json(settings: dict[str, str] | None = None) -> str:
    """JSON logo đối tác đang hiển thị trên landing — dùng khởi tạo editor CMS."""
    merged = settings if settings is not None else _merged_landing_settings()
    rows = _partner_logos_to_editor_rows(_partner_logos_for_landing(merged))
    return json.dumps(rows, ensure_ascii=False)


def _capabilities_items_seed_for_cms(settings: dict[str, str]) -> list[dict[str, str]]:
    """Mục capabilities đang hiển thị trên web — dùng khởi tạo editor CMS."""
    merged = {**DEFAULT_SETTINGS_NAV_MEGA_BANNER, **DEFAULT_SETTINGS_FOOTER, **settings}
    return _capabilities_items_for_landing(merged)


def _capabilities_items_effective_json(settings: dict[str, str] | None = None) -> str:
    """JSON capabilities đang hiển thị trên landing — dùng khởi tạo editor CMS."""
    merged = settings if settings is not None else _merged_landing_settings()
    items = _capabilities_items_for_landing(merged)
    return json.dumps(items, ensure_ascii=False)


# Chỉ đọc từ GET /api/settings — không ghi vào DB khi PUT.
READONLY_SETTINGS_KEYS = frozenset({
    "partner_logos_effective_json",
    "capabilities_items_effective_json",
})


RECRUITMENT_EMAIL = "tuyendung@pttadvertising.vn"

_RAW_RECRUITMENT_POSITIONS: list[dict[str, Any]] = [
    {
        "slug": "account-executive-digital-marketing",
        "title": "Account Executive — Digital Marketing",
        "location": "Hà Nội",
        "employment_type": "Toàn thời gian",
        "description": (
            "Tiếp nhận brief khách hàng, xây dựng đề xuất giải pháp truyền thông đa kênh, "
            "phối hợp team creative & media, theo dõi tiến độ và báo cáo hiệu quả chiến dịch."
        ),
        "intro": (
            "Vị trí Account Executive đóng vai trò cầu nối giữa khách hàng và nội bộ, "
            "đảm bảo chiến dịch được triển khai đúng mục tiêu KPI và chất lượng dịch vụ."
        ),
        "responsibilities": [
            "Tiếp nhận brief, làm rõ mục tiêu truyền thông và ngân sách với khách hàng.",
            "Phối hợp Creative, Media và Data để xây dựng proposal & timeline triển khai.",
            "Theo dõi hiệu suất chiến dịch, tổng hợp báo cáo định kỳ và đề xuất tối ưu.",
            "Duy trì mối quan hệ khách hàng, hỗ trợ chăm sóc sau bán và upsell gói dịch vụ phù hợp.",
        ],
        "requirements": [
            "Tối thiểu 2 năm kinh nghiệm account / client service trong agency hoặc nhãn hàng.",
            "Hiểu biết nền tảng về digital marketing, KPI (reach, CTR, CPA, ROAS…).",
            "Kỹ năng thuyết trình, làm việc nhóm và quản lý nhiều dự án song song.",
            "Giao tiếp tiếng Anh đọc hiểu tài liệu chuyên ngành là lợi thế.",
        ],
        "benefits": [
            "Môi trường làm việc chuyên nghiệp, cơ hội học hỏi từ các chiến dịch lớn.",
            "Chế độ BHXH, BHYT đầy đủ; review lương và thăng tiến theo năng lực.",
            "Laptop làm việc; hỗ trợ đào tạo nội bộ và tham gia sự kiện ngành.",
        ],
    },
    {
        "slug": "media-planning-buying",
        "title": "Chuyên viên Media Planning & Buying",
        "location": "TP. Hồ Chí Minh",
        "employment_type": "Toàn thời gian",
        "description": (
            "Lập kế hoạch phân bổ ngân sách, đàm phán với publisher, tối ưu chỉ số hiển thị và CPA/CPC, "
            "phối hợp tracking và báo cáo theo KPI."
        ),
        "intro": (
            "Bạn sẽ tham gia xây dựng chiến lược mua và phân phối media trên các nền tảng display, "
            "video và programmatic, đảm bảo tối ưu ngân sách theo mục tiêu thương hiệu và hiệu suất."
        ),
        "responsibilities": [
            "Xây dựng media plan theo brief, phân bổ ngân sách theo kênh và giai đoạn.",
            "Đàm phán giá, booking và quản lý chất lượng hiển thị với đối tác publisher.",
            "Phối hợp team tracking & analytics để đánh giá và tối ưu realtime.",
            "Báo cáo kết quả theo tuần/tháng; đề xuất điều chỉnh creative và targeting.",
        ],
        "requirements": [
            "Có kinh nghiệm media planning/buying hoặc programmatic (ưu tiên có chứng chỉ GMP/Google Ads).",
            "Thành thạo Excel/Sheet; quen với dashboard báo cáo (Data Studio, Looker…).",
            "Tư duy số, chịu được áp lực deadline và ngân sách chặt.",
            "Tiếng Anh đọc hiểu tài liệu kỹ thuật và email làm việc quốc tế.",
        ],
        "benefits": [
            "Được làm việc với inventory và nền tảng martech hiện đại.",
            "Chính sách phúc lợi cạnh tranh; thưởng theo hiệu quả dự án.",
            "Lịch làm việc linh hoạt một phần theo team (thỏa thuận).",
        ],
    },
    {
        "slug": "frontend-engineer-react",
        "title": "Frontend Engineer (React)",
        "location": "Hà Nội / Hybrid",
        "employment_type": "Toàn thời gian",
        "description": (
            "Phát triển giao diện sản phẩm martech, tối ưu hiệu năng và trải nghiệm người dùng, "
            "làm việc gần với design và backend."
        ),
        "intro": (
            "Tham gia đội phát triển sản phẩm nội bộ và landing hệ thống quảng cáo, "
            "đảm bảo giao diện ổn định, accessible và tối ưu trên đa thiết bị."
        ),
        "responsibilities": [
            "Phát triển UI bằng React/TypeScript theo design system và code review.",
            "Tối ưu hiệu năng (lazy load, bundle size), SEO kỹ thuật cơ bản cho landing.",
            "Phối hợp QA fix bug; viết test (unit/e2e) theo quy trình team.",
            "Tham gia họp sprint, ước lượh và báo cáo tiến độ.",
        ],
        "requirements": [
            "Thành thạo HTML/CSS, React hooks, state management (Redux/Zustand là lợi thế).",
            "Quen với Git flow, REST API; hiểu biết về bảo mật front-end cơ bản.",
            "Tối thiểu 2 năm kinh nghiệm frontend; portfolio hoặc link GitHub.",
            "Ưu tiên ứng viên có kinh nghiệm SSR/Next.js hoặc performance audit.",
        ],
        "benefits": [
            "Hybrid linh hoạt; trang thiết bị làm việc hiện đại.",
            "Đào tạo kỹ thuật định kỳ; tham gia cộng đồng nội bộ tech talk.",
            "Review lương 2 lần/năm theo đóng góp và roadmap cá nhân.",
        ],
    },
]


def _build_recruitment_positions() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for r in _RAW_RECRUITMENT_POSITIONS:
        subj = f"[Ứng tuyển] {r['title']}"
        body = (
            "Kính gửi bộ phận nhân sự,\n\n"
            f"Tôi muốn nộp hồ sơ ứng tuyển vị trí: {r['title']}.\n"
            f"Địa điểm làm việc: {r['location']} · {r['employment_type']}.\n\n"
            "Vui lòng xem CV / portfolio đính kèm email này.\n\n"
            "Trân trọng,"
        )
        href = f"mailto:{RECRUITMENT_EMAIL}?subject={quote(subj)}&body={quote(body)}"
        rows.append({**r, "apply_href": href})
    return rows


RECRUITMENT_POSITIONS: list[dict[str, Any]] = _build_recruitment_positions()


def _job_row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    d = dict(row)
    for k in ("responsibilities", "requirements", "benefits"):
        try:
            d[k] = json.loads(d.get(k) or "[]")
        except (json.JSONDecodeError, TypeError):
            d[k] = []
    subj = f"[Ứng tuyển] {d['title']}"
    body = (
        "Kính gửi bộ phận nhân sự,\n\n"
        f"Tôi muốn nộp hồ sơ ứng tuyển vị trí: {d['title']}.\n"
        f"Địa điểm làm việc: {d['location']} · {d['employment_type']}.\n\n"
        "Vui lòng xem CV / portfolio đính kèm email này.\n\n"
        "Trân trọng,"
    )
    d["apply_href"] = f"mailto:{RECRUITMENT_EMAIL}?subject={quote(subj)}&body={quote(body)}"
    return d


def get_recruitment_jobs(active_only: bool = True) -> list[dict[str, Any]]:
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT * FROM recruitment_jobs"
            + (" WHERE is_active=1" if active_only else "")
            + " ORDER BY sort_order, id",
        ).fetchall()
    return [_job_row_to_dict(r) for r in rows]


def get_recruitment_job(slug: str) -> dict[str, Any] | None:
    with get_connection() as conn:
        row = conn.execute(
            "SELECT * FROM recruitment_jobs WHERE slug=? AND is_active=1", (slug,)
        ).fetchone()
    return _job_row_to_dict(row) if row else None


def get_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db() -> None:
    with get_connection() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS projects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                category TEXT NOT NULL,
                image_url TEXT NOT NULL,
                intro TEXT NOT NULL DEFAULT '',
                description TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS news (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                summary TEXT NOT NULL,
                url TEXT NOT NULL,
                published_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );
            """
        )

        _migrate_projects_schema(conn)
        _migrate_news_schema(conn)

        seed_projects = [
            (
                "TNEX - Trạm thanh xuân 0 phí",
                "Digital Campaign",
                "https://images.unsplash.com/photo-1551434678-e076c223a692?auto=format&fit=crop&w=1200&q=80",
                "Chiến dịch truyền thông tích hợp để gia tăng tệp người dùng trẻ.",
            ),
            (
                "DR.PAPIE - Tay mẹ kỹ càng",
                "FMCG Launch",
                "https://images.unsplash.com/photo-1556745757-8d76bdb6984b?auto=format&fit=crop&w=1200&q=80",
                "Kết hợp content marketing và social amplification cho thương hiệu chăm sóc gia đình.",
            ),
            (
                "VinFast - Tri ân người tiên phong",
                "Brand Marketing",
                "https://images.unsplash.com/photo-1449965408869-eaa3f722e40d?auto=format&fit=crop&w=1200&q=80",
                "Kể câu chuyện thương hiệu theo định hướng hiệu suất và cảm xúc.",
            ),
            (
                "VPBankS - The Investors",
                "Performance Media",
                "https://images.unsplash.com/photo-1559526324-593bc073d938?auto=format&fit=crop&w=1200&q=80",
                "Tập trung tối ưu lead chất lượng cao từ tệp người dùng đầu tư.",
            ),
            (
                "Tiger Beer - Vươn mình bứt phá",
                "Integrated Campaign",
                "https://images.unsplash.com/photo-1532635241-17e820acc59f?auto=format&fit=crop&w=1200&q=80",
                "Chiến dịch tích hợp digital, social và activation để tăng độ phủ thương hiệu.",
            ),
            (
                "Human Act Prize 2025",
                "Social Impact",
                "https://images.unsplash.com/photo-1521737711867-e3b97375f902?auto=format&fit=crop&w=1200&q=80",
                "Lan tỏa các dự án cộng đồng và xây dựng nhận diện giải thưởng.",
            ),
        ]
        existing_project_titles = {
            row["title"] for row in conn.execute("SELECT title FROM projects").fetchall()
        }
        for title, category, image_url, description in seed_projects:
            if title not in existing_project_titles:
                conn.execute(
                    """
                    INSERT INTO projects (title, category, image_url, intro, description, created_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        title,
                        category,
                        image_url,
                        description,
                        description,
                        datetime.now().strftime("%Y-%m-%d"),
                    ),
                )

        seed_news = [
            (
                "Viết tiếp những con số bùng nổ tại WeChoice Awards 2025",
                "Hơn 18 triệu lượt bình chọn và giữ top từ khóa thịnh hành.",
                "https://ptt.com.vn",
            ),
            (
                "PTT Advertising Solutions và Chicilon Media ký kết hợp tác chiến lược",
                "Thúc đẩy trải nghiệm truyền thông tích hợp hiện đại và hiệu quả.",
                "https://ptt.com.vn",
            ),
            (
                "PTT Advertising Solutions và Goldsun Media mở rộng giải pháp Digital - OOH",
                "Gia tăng hiệu quả truyền thông tích hợp cho khối khách hàng doanh nghiệp.",
                "https://ptt.com.vn",
            ),
            (
                "Giải thưởng Human Act Prize 2025: truyền cảm hứng hành động",
                "Tôn vinh các sáng kiến vì cộng đồng có tác động tích cực và bền vững.",
                "https://ptt.com.vn",
            ),
        ]
        existing_news_titles = {
            row["title"] for row in conn.execute("SELECT title FROM news").fetchall()
        }
        for title, summary, url in seed_news:
            if title not in existing_news_titles:
                conn.execute(
                    """
                    INSERT INTO news (title, summary, url, published_at)
                    VALUES (?, ?, ?, ?)
                    """,
                    (title, summary, url, datetime.now().strftime("%Y-%m-%d")),
                )

        defaults = {
            "brand_name": "PTT Advertising Solutions",
            "hero_title": "Nền tảng Creative Martech cho tăng trưởng bền vững",
            "hero_description": "Kết nối thương hiệu và người tiêu dùng bằng hệ sinh thái nội dung, công nghệ và dữ liệu.",
            "contact_email": "contact@pttadvertising.vn",
            "contact_phone": "+84 24 7307 7979",
            "office_hn": "",
            "office_hcm": "73, Tạ Hiện, Phường Cát Lái, TP. Hồ Chí Minh",
            **DEFAULT_SETTINGS_NAV_MEGA_BANNER,
            **DEFAULT_SETTINGS_FOOTER,
        }

        existing = {
            row["key"] for row in conn.execute("SELECT key FROM settings").fetchall()
        }
        for key, value in defaults.items():
            if key not in existing:
                conn.execute(
                    "INSERT INTO settings (key, value) VALUES (?, ?)",
                    (key, value),
                )

        # Thống nhất tên thương hiệu & email từ bản cũ (khi cập nhật từ Admicro/PTT Creative Martech)
        conn.execute(
            "UPDATE settings SET value = ? WHERE key = 'brand_name' AND value IN (?, ?, ?, ?)",
            (
                "PTT Advertising Solutions",
                "PTT Creative Martech",
                "ADMICRO",
                "Admicro",
                "admicro",
            ),
        )
        conn.execute(
            "UPDATE settings SET value = ? WHERE key = 'contact_email' AND value IN (?, ?)",
            ("contact@pttadvertising.vn", "contact@ptt.vn", "contact@admicro.vn"),
        )
        # Gán link Zalo mặc định khi chưa cấu hình (SĐT 0938 834 238)
        conn.execute(
            "UPDATE settings SET value = ? WHERE key = 'fab_zalo_url' AND TRIM(COALESCE(value, '')) = ?",
            ("https://zalo.me/0938834238", ""),
        )
        # Địa chỉ văn phòng HCMC (và gỡ placeholder HN nếu vẫn là bản seed cũ)
        conn.execute(
            """
            UPDATE settings SET value = ?
            WHERE key = 'office_hcm'
              AND TRIM(COALESCE(value, '')) IN (?, ?)
            """,
            (
                "73, Tạ Hiện, Phường Cát Lái, TP. Hồ Chí Minh",
                "Tầng 5, 123 Tower, TP Hồ Chí Minh",
                "Tầng 5 tòa 123 Tower, 123-127 Võ Văn Tần, Phường 6, Quận 3.",
            ),
        )
        conn.execute(
            """
            UPDATE settings SET value = ''
            WHERE key = 'office_hn'
              AND TRIM(COALESCE(value, '')) IN (?, ?)
            """,
            (
                "Tầng 20, Center Building, Hà Nội",
                "Tầng 20, Center Building, Hapulico Complex, số 1 Nguyễn Huy Tưởng, Thanh Xuân.",
            ),
        )

        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS crm_customers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                phone TEXT NOT NULL DEFAULT '',
                email TEXT NOT NULL DEFAULT '',
                address TEXT NOT NULL DEFAULT '',
                company TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS crm_cases (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_id INTEGER NOT NULL REFERENCES crm_customers(id),
                title TEXT NOT NULL,
                description TEXT NOT NULL DEFAULT '',
                channel TEXT NOT NULL DEFAULT 'khac',
                priority TEXT NOT NULL DEFAULT 'binh_thuong',
                status TEXT NOT NULL DEFAULT 'tiep_nhan',
                assigned_to TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS crm_case_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                case_id INTEGER NOT NULL REFERENCES crm_cases(id),
                kind TEXT NOT NULL DEFAULT 'ghi_chu',
                body TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_crm_cases_status ON crm_cases(status);
            CREATE INDEX IF NOT EXISTS idx_crm_cases_updated ON crm_cases(updated_at DESC);
            CREATE INDEX IF NOT EXISTS idx_crm_events_case ON crm_case_events(case_id);
            """
        )

        _migrate_crm_staff_schema(conn)
        _ensure_crm_customers_extended_columns(conn)
        _migrate_live_chat_schema(conn)


def _migrate_news_schema(conn: sqlite3.Connection) -> None:
    cols = {r[1] for r in conn.execute("PRAGMA table_info(news)")}
    if "image_url" not in cols:
        conn.execute("ALTER TABLE news ADD COLUMN image_url TEXT NOT NULL DEFAULT ''")


def _migrate_projects_schema(conn: sqlite3.Connection) -> None:
    cols = {r[1] for r in conn.execute("PRAGMA table_info(projects)")}
    if "intro" not in cols:
        conn.execute("ALTER TABLE projects ADD COLUMN intro TEXT NOT NULL DEFAULT ''")
        conn.execute(
            "UPDATE projects SET intro = description WHERE TRIM(intro) = '' AND TRIM(description) != ''"
        )


def _migrate_live_chat_schema(conn: sqlite3.Connection) -> None:
    tables = {r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}
    if "chat_conversations" not in tables:
        conn.execute(
            """
            CREATE TABLE chat_conversations (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id  TEXT    NOT NULL UNIQUE,
                visitor_name TEXT   NOT NULL DEFAULT '',
                visitor_page TEXT   NOT NULL DEFAULT '',
                status      TEXT    NOT NULL DEFAULT 'open',
                created_at  TEXT    NOT NULL DEFAULT (datetime('now')),
                updated_at  TEXT    NOT NULL DEFAULT (datetime('now'))
            )
            """
        )
    if "chat_messages" not in tables:
        conn.execute(
            """
            CREATE TABLE chat_messages (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                conversation_id INTEGER NOT NULL,
                sender          TEXT    NOT NULL,
                content         TEXT    NOT NULL,
                created_at      TEXT    NOT NULL DEFAULT (datetime('now')),
                FOREIGN KEY (conversation_id) REFERENCES chat_conversations(id) ON DELETE CASCADE
            )
            """
        )


def _chat_openai_reply(messages: list[dict[str, Any]]) -> str | None:
    import urllib.request as _ureq
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        return None
    model = os.environ.get("OPENAI_MODEL") or os.environ.get("AI_CHAT_MODEL") or "gpt-4o-mini"
    payload = json.dumps(
        {"model": model, "messages": messages, "max_tokens": 600, "temperature": 0.6},
        ensure_ascii=False,
    ).encode("utf-8")
    req = _ureq.Request(
        "https://api.openai.com/v1/chat/completions",
        data=payload,
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with _ureq.urlopen(req, timeout=20) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        return str(data["choices"][0]["message"]["content"]).strip()
    except Exception:
        return None


def _ensure_crm_customers_extended_columns(conn: sqlite3.Connection) -> None:
    """Bổ sung cột hồ sơ khách hàng (idempotent)."""
    cols = {r[1] for r in conn.execute("PRAGMA table_info(crm_customers)")}
    if "address" not in cols:
        try:
            conn.execute(
                "ALTER TABLE crm_customers ADD COLUMN address TEXT NOT NULL DEFAULT ''"
            )
        except sqlite3.Error:
            pass
    ensure_customer_360_schema(conn)


def _migrate_crm_staff_schema(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_staff (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT NOT NULL DEFAULT '',
            email TEXT NOT NULL DEFAULT '',
            active INTEGER NOT NULL DEFAULT 1,
            sort_order INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_staff_list ON crm_staff(active, sort_order, name)"
    )
    case_cols = {r[1] for r in conn.execute("PRAGMA table_info(crm_cases)")}
    if "assigned_staff_id" not in case_cols:
        conn.execute(
            "ALTER TABLE crm_cases ADD COLUMN assigned_staff_id INTEGER "
            "REFERENCES crm_staff(id)"
        )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_cases_staff ON crm_cases(assigned_staff_id)"
    )
    try:
        orphans = conn.execute(
            """
            SELECT c.id, TRIM(c.assigned_to) AS legacy
            FROM crm_cases c
            WHERE c.assigned_staff_id IS NULL AND TRIM(COALESCE(c.assigned_to, '')) != ''
            """
        ).fetchall()
        for r in orphans:
            legacy = str(r["legacy"] or "").strip()
            if not legacy:
                continue
            hit = conn.execute(
                """
                SELECT id FROM crm_staff
                WHERE TRIM(name) = ? AND active = 1
                LIMIT 1
                """,
                (legacy,),
            ).fetchone()
            if hit:
                conn.execute(
                    "UPDATE crm_cases SET assigned_staff_id = ? WHERE id = ?",
                    (int(hit["id"]), int(r["id"])),
                )
    except sqlite3.Error:
        pass

    _ensure_crm_staff_extended_columns(conn)


def _ensure_crm_staff_extended_columns(conn: sqlite3.Connection) -> None:
    """Bổ sung cột nghiệp vụ cho crm_staff (idempotent)."""
    cols = {r[1] for r in conn.execute("PRAGMA table_info(crm_staff)")}
    extra: list[tuple[str, str]] = [
        ("job_title", "ALTER TABLE crm_staff ADD COLUMN job_title TEXT NOT NULL DEFAULT ''"),
        ("department", "ALTER TABLE crm_staff ADD COLUMN department TEXT NOT NULL DEFAULT ''"),
        ("internal_code", "ALTER TABLE crm_staff ADD COLUMN internal_code TEXT NOT NULL DEFAULT ''"),
        ("notes", "ALTER TABLE crm_staff ADD COLUMN notes TEXT NOT NULL DEFAULT ''"),
        ("updated_at", "ALTER TABLE crm_staff ADD COLUMN updated_at TEXT NOT NULL DEFAULT ''"),
    ]
    for col_name, ddl in extra:
        if col_name not in cols:
            try:
                conn.execute(ddl)
            except sqlite3.Error:
                pass

    _ensure_enterprise_crm_schema(conn)


def _ensure_enterprise_crm_schema(conn: sqlite3.Connection) -> None:
    """Phòng ban, nhân sự doanh nghiệp, mốc phân công trên hồ sơ."""
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_departments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL DEFAULT '',
            name TEXT NOT NULL,
            description TEXT NOT NULL DEFAULT '',
            sort_order INTEGER NOT NULL DEFAULT 0,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_departments_active ON crm_departments(active, sort_order)"
    )

    sc = {r[1] for r in conn.execute("PRAGMA table_info(crm_staff)")}
    staff_cols: list[tuple[str, str]] = [
        ("department_id", "ALTER TABLE crm_staff ADD COLUMN department_id INTEGER REFERENCES crm_departments(id)"),
        ("reports_to_id", "ALTER TABLE crm_staff ADD COLUMN reports_to_id INTEGER REFERENCES crm_staff(id)"),
        (
            "employment_type",
            "ALTER TABLE crm_staff ADD COLUMN employment_type TEXT NOT NULL DEFAULT 'full_time'",
        ),
        ("started_on", "ALTER TABLE crm_staff ADD COLUMN started_on TEXT NOT NULL DEFAULT ''"),
        ("ended_on", "ALTER TABLE crm_staff ADD COLUMN ended_on TEXT NOT NULL DEFAULT ''"),
    ]
    for col, ddl in staff_cols:
        if col not in sc:
            try:
                conn.execute(ddl)
            except sqlite3.Error:
                pass

    cc = {r[1] for r in conn.execute("PRAGMA table_info(crm_cases)")}
    if "assigned_at" not in cc:
        try:
            conn.execute(
                "ALTER TABLE crm_cases ADD COLUMN assigned_at TEXT NOT NULL DEFAULT ''"
            )
        except sqlite3.Error:
            pass

    n_dept = conn.execute("SELECT COUNT(*) AS n FROM crm_departments").fetchone()
    if n_dept and int(n_dept["n"]) == 0:
        ts = datetime.now().strftime("%Y-%m-%d")
        now = _crm_ts()
        seed = [
            ("CSKH", "Chăm sóc khách hàng", "Tiếp nhận và xử lý yêu cầu khách hàng.", 10),
            ("KD", "Kinh doanh", "Đội ngũ phát triển khách hàng và dự án.", 20),
            ("VH", "Vận hành", "Điều phối triển khai và chất lượng dịch vụ.", 30),
        ]
        for code, name, desc, so in seed:
            conn.execute(
                """
                INSERT INTO crm_departments (code, name, description, sort_order, active, created_at, updated_at)
                VALUES (?, ?, ?, ?, 1, ?, ?)
                """,
                (code, name, desc, so, ts, now),
            )

    try:
        conn.execute(
            """
            UPDATE crm_cases
            SET assigned_at = COALESCE(NULLIF(TRIM(assigned_at), ''), updated_at)
            WHERE assigned_staff_id IS NOT NULL
              AND (assigned_at IS NULL OR TRIM(assigned_at) = '')
            """
        )
    except sqlite3.Error:
        pass

    _ensure_crm_positions_schema(conn)
    _ensure_crm_lead_channels_schema(conn)
    _ensure_crm_attendance_payroll_schema(conn)
    _ensure_crm_kpi_schema(conn)
    _ensure_crm_hub_schema(conn)
    _ensure_crm_sop_schema(conn)
    _ensure_crm_marketing_plan_schema(conn)
    ensure_pipeline_schema(conn)
    ensure_sales_hub_schema(conn)
    ensure_re_projects_schema(conn)
    seed_re_project_section_permissions(conn)
    _ensure_cms_permissions_schema(conn)
    ensure_care_schema(conn)
    ensure_daily_work_report_schema(conn)
    ensure_lead_schema(conn)
    from crm_lead_catalog import ensure_lead_catalog_schema

    ensure_lead_catalog_schema(conn)
    ensure_staff_login_schema(conn)
    from crm_lead_assign_scope import bootstrap_staff_assign_scopes_if_empty

    bootstrap_staff_assign_scopes_if_empty(conn)
    migrate_unified_passwords(conn, updated_at=_crm_ts())
    from crm_cross_module import ensure_cross_module_schema

    ensure_cross_module_schema(conn)
    _ensure_service_lifecycle_schema(conn)
    _ensure_svc_tasks_schema(conn)
    _ensure_lead_intake_schema(conn)
    _ensure_svc_risk_schema(conn)
    _ensure_svc_finance_schema(conn)
    _migrate_contract_billing(conn)
    _ensure_svc_kpi_schema(conn)
    _ensure_customer_brief_schema(conn)
    _ensure_aeo_schema(conn)
    _ensure_proposal_schema(conn)
    from crm_lead_presales import ensure_schema as _ensure_lead_presales_schema

    _ensure_lead_presales_schema(conn)
    from crm_lead_presales_marketing_plan import ensure_r5_schema

    ensure_r5_schema(conn)
    from crm_lead_industry_addon import ensure_r6_schema

    ensure_r6_schema(conn)
    from crm_lead_product_model_p3 import ensure_p3_schema

    ensure_p3_schema(conn)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS recruitment_jobs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            slug TEXT UNIQUE NOT NULL,
            title TEXT NOT NULL,
            location TEXT NOT NULL DEFAULT '',
            employment_type TEXT NOT NULL DEFAULT 'Toàn thời gian',
            description TEXT NOT NULL DEFAULT '',
            intro TEXT NOT NULL DEFAULT '',
            responsibilities TEXT NOT NULL DEFAULT '[]',
            requirements TEXT NOT NULL DEFAULT '[]',
            benefits TEXT NOT NULL DEFAULT '[]',
            is_active INTEGER NOT NULL DEFAULT 1,
            sort_order INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT (strftime('%Y-%m-%dT%H:%M:%SZ','now'))
        )
        """
    )
    # Seed hardcoded positions if table empty
    if conn.execute("SELECT COUNT(*) FROM recruitment_jobs").fetchone()[0] == 0:
        for i, r in enumerate(_RAW_RECRUITMENT_POSITIONS):
            conn.execute(
                """INSERT OR IGNORE INTO recruitment_jobs
                   (slug, title, location, employment_type, description, intro,
                    responsibilities, requirements, benefits, sort_order)
                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (
                    r["slug"], r["title"], r.get("location", ""),
                    r.get("employment_type", "Toàn thời gian"),
                    r.get("description", ""), r.get("intro", ""),
                    json.dumps(r.get("responsibilities", []), ensure_ascii=False),
                    json.dumps(r.get("requirements", []), ensure_ascii=False),
                    json.dumps(r.get("benefits", []), ensure_ascii=False),
                    i,
                ),
            )


def _ensure_cms_permissions_schema(conn: sqlite3.Connection) -> None:
    """Vai trò, ma trận quyền CMS, gán user → role."""
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS cms_roles (
            code TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            description TEXT NOT NULL DEFAULT '',
            is_system INTEGER NOT NULL DEFAULT 1,
            updated_at TEXT NOT NULL DEFAULT ''
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS cms_role_permissions (
            role_code TEXT NOT NULL,
            module_id TEXT NOT NULL,
            action TEXT NOT NULL,
            PRIMARY KEY (role_code, module_id, action),
            FOREIGN KEY (role_code) REFERENCES cms_roles(code) ON DELETE CASCADE
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS cms_admin_users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            display_name TEXT NOT NULL DEFAULT '',
            role_code TEXT NOT NULL,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY (role_code) REFERENCES cms_roles(code)
        )
        """
    )
    ts = _crm_ts()
    ts_date = datetime.now().strftime("%Y-%m-%d")
    for role in CMS_ROLES:
        conn.execute(
            """
            INSERT INTO cms_roles (code, name, description, is_system, updated_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(code) DO UPDATE SET
                name = excluded.name,
                description = excluded.description
            """,
            (
                role["code"],
                role["name"],
                role.get("description", ""),
                1 if role.get("is_system") else 0,
                ts,
            ),
        )
        existing = conn.execute(
            "SELECT COUNT(*) AS n FROM cms_role_permissions WHERE role_code = ?",
            (role["code"],),
        ).fetchone()
        if existing and int(existing["n"]) == 0:
            defaults = default_grants_for_role(role["code"])
            for mid, acts in defaults.items():
                for act in acts:
                    conn.execute(
                        """
                        INSERT OR IGNORE INTO cms_role_permissions (role_code, module_id, action)
                        VALUES (?, ?, ?)
                        """,
                        (role["code"], mid, act),
                    )
    n_users = conn.execute("SELECT COUNT(*) AS n FROM cms_admin_users").fetchone()
    if n_users and int(n_users["n"]) == 0:
        eu, _ = _admin_expected_credentials()
        conn.execute(
            """
            INSERT INTO cms_admin_users (username, display_name, role_code, active, created_at, updated_at)
            VALUES (?, ?, 'super_admin', 1, ?, ?)
            """,
            (eu, "Quản trị mặc định", ts_date, ts),
        )

    au_cols = {r[1] for r in conn.execute("PRAGMA table_info(cms_admin_users)")}
    if "position_id" not in au_cols:
        conn.execute(
            "ALTER TABLE cms_admin_users ADD COLUMN position_id INTEGER REFERENCES crm_positions(id)"
        )
    au_cols = {r[1] for r in conn.execute("PRAGMA table_info(cms_admin_users)")}
    if "password_hash" not in au_cols:
        conn.execute(
            "ALTER TABLE cms_admin_users ADD COLUMN password_hash TEXT NOT NULL DEFAULT ''"
        )

    _ensure_system_admin_access(conn)
    ensure_role_grants_customized_column(conn)
    ensure_position_grants_customized_column(conn)
    backfill_role_grants_customized(conn)
    backfill_position_grants_customized(conn)
    migrate_cms_role_sidebar_modules(conn)

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_position_section_permissions (
            position_id INTEGER NOT NULL REFERENCES crm_positions(id) ON DELETE CASCADE,
            section_id TEXT NOT NULL,
            action TEXT NOT NULL,
            PRIMARY KEY (position_id, section_id, action)
        )
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_crm_pos_section_perm
        ON crm_position_section_permissions(position_id)
        """
    )
    n_pos_perm = conn.execute(
        "SELECT COUNT(*) AS n FROM crm_position_section_permissions"
    ).fetchone()
    if n_pos_perm and int(n_pos_perm["n"]) == 0:
        pos_rows = conn.execute(
            "SELECT id, code FROM crm_positions WHERE active = 1"
        ).fetchall()
        for prow in pos_rows:
            pid = int(prow["id"])
            pcode = str(prow["code"])
            defaults = default_grants_for_position(pcode)
            for sid, acts in defaults.items():
                for act in acts:
                    conn.execute(
                        """
                        INSERT OR IGNORE INTO crm_position_section_permissions
                        (position_id, section_id, action)
                        VALUES (?, ?, ?)
                        """,
                        (pid, sid, act),
                    )
    seed_marketing_positions(conn)
    migrate_kd01_leads_only_permissions(conn)
    migrate_hdsd_position_permissions(conn)


def _ensure_system_admin_access(conn: sqlite3.Connection) -> None:
    """User admin / super_admin: toàn quyền, không gắn chức vụ giới hạn CRM."""
    ts = _crm_ts()
    eu, _ = _admin_expected_credentials()
    admin_names = {str(eu).strip().lower(), "admin"}
    for mid in CMS_MODULE_IDS:
        for act in CMS_ACTIONS:
            conn.execute(
                """
                INSERT OR IGNORE INTO cms_role_permissions (role_code, module_id, action)
                VALUES ('super_admin', ?, ?)
                """,
                (mid, act),
            )
    for uname in admin_names:
        if not uname:
            continue
        conn.execute(
            """
            UPDATE cms_admin_users
            SET role_code = 'super_admin', position_id = NULL, active = 1, updated_at = ?
            WHERE lower(trim(username)) = ?
            """,
            (ts, uname),
        )
    conn.execute(
        """
        UPDATE cms_admin_users
        SET position_id = NULL, updated_at = ?
        WHERE role_code IN ('super_admin', 'cms_admin') AND position_id IS NOT NULL
        """,
        (ts,),
    )


def _ensure_crm_lead_channels_schema(conn: sqlite3.Connection) -> None:
    """Danh mục kênh tiếp nhận lead / yêu cầu CRM (dropdown Yêu cầu mới)."""
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_lead_channels (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL DEFAULT '',
            name TEXT NOT NULL,
            description TEXT NOT NULL DEFAULT '',
            sort_order INTEGER NOT NULL DEFAULT 0,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_crm_lead_channels_code "
        "ON crm_lead_channels(lower(trim(code))) WHERE trim(code) != ''"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_lead_channels_active "
        "ON crm_lead_channels(active, sort_order)"
    )
    n = conn.execute("SELECT COUNT(*) AS n FROM crm_lead_channels").fetchone()
    if n and int(n["n"]) == 0:
        ts_date = datetime.now().strftime("%Y-%m-%d")
        ts = _crm_ts()
        seed = [
            ("dien_thoai", "Điện thoại", "Khách gọi hotline / SĐT sales.", 10),
            ("email", "Email", "Lead qua email hoặc form gửi mail.", 20),
            ("zalo", "Zalo", "Zalo OA / chat cá nhân.", 30),
            ("truc_tiep", "Trực tiếp", "Gặp mặt showroom / sự kiện.", 40),
            ("khac", "Khác", "Kênh khác — ghi rõ trong mô tả.", 50),
        ]
        for code, name, desc, so in seed:
            conn.execute(
                """
                INSERT INTO crm_lead_channels (code, name, description, sort_order, active, created_at, updated_at)
                VALUES (?, ?, ?, ?, 1, ?, ?)
                """,
                (code, name, desc, so, ts_date, ts),
            )


def _ensure_crm_kpi_schema(conn: sqlite3.Connection) -> None:
    """Danh mục chỉ tiêu KPI và bảng KPI theo nhân viên / tháng."""
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_kpi_metrics (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL DEFAULT '',
            name TEXT NOT NULL,
            unit TEXT NOT NULL DEFAULT '',
            description TEXT NOT NULL DEFAULT '',
            sort_order INTEGER NOT NULL DEFAULT 0,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_kpi_metrics_active ON crm_kpi_metrics(active, sort_order)"
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_staff_kpi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            staff_id INTEGER NOT NULL REFERENCES crm_staff(id) ON DELETE CASCADE,
            metric_id INTEGER NOT NULL REFERENCES crm_kpi_metrics(id) ON DELETE CASCADE,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            target_value REAL,
            actual_value REAL,
            status TEXT NOT NULL DEFAULT 'draft',
            note TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE(staff_id, metric_id, year, month),
            CHECK (month >= 1 AND month <= 12 AND year >= 2000 AND year <= 2100)
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_staff_kpi_lookup ON crm_staff_kpi(year, month, staff_id)"
    )

    mi = {r[1] for r in conn.execute("PRAGMA table_info(crm_kpi_metrics)")}
    if "higher_is_better" not in mi:
        try:
            conn.execute(
                "ALTER TABLE crm_kpi_metrics ADD COLUMN higher_is_better INTEGER NOT NULL DEFAULT 1"
            )
        except sqlite3.Error:
            pass
    if "warn_ratio" not in mi:
        try:
            conn.execute("ALTER TABLE crm_kpi_metrics ADD COLUMN warn_ratio REAL")
        except sqlite3.Error:
            pass

    n_met = conn.execute("SELECT COUNT(*) AS n FROM crm_kpi_metrics").fetchone()
    if n_met and int(n_met["n"]) == 0:
        ts_d = datetime.now().strftime("%Y-%m-%d")
        now = _crm_ts()
        seeds = [
            ("SLA_PCT", "SLA phản hồi đúng hạn", "%", "Tỷ lệ yêu cầu được phản hồi trong TAT quy định.", 10),
            ("CSAT", "Điểm hài lòng KH", "điểm", "Khảo sát / đánh giá sau xử lý (thang tuỳ công ty).", 20),
            ("CASES_OK", "Hồ sơ xử lý đúng quy trình", "%", "Đạt checklist / không vi phạm SLA nội bộ.", 30),
            ("TARGET_LEADS", "Lead / cơ hội phụ trách", "lead", "Số lead hoặc pipeline theo chỉ tiêu tháng.", 40),
        ]
        for code, name, unit, desc, so in seeds:
            conn.execute(
                """
                INSERT INTO crm_kpi_metrics (
                    code, name, unit, description, sort_order, active,
                    created_at, updated_at, higher_is_better, warn_ratio
                )
                VALUES (?, ?, ?, ?, ?, 1, ?, ?, 1, 0.9)
                """,
                (code, name, unit, desc, so, ts_d, now),
            )

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_position_kpi_metrics (
            position_id INTEGER NOT NULL REFERENCES crm_positions(id) ON DELETE CASCADE,
            metric_id INTEGER NOT NULL REFERENCES crm_kpi_metrics(id) ON DELETE CASCADE,
            sort_order INTEGER NOT NULL DEFAULT 0,
            PRIMARY KEY (position_id, metric_id)
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_position_kpi_pos ON crm_position_kpi_metrics(position_id)"
    )
    n_map = conn.execute("SELECT COUNT(*) AS n FROM crm_position_kpi_metrics").fetchone()
    if n_map and int(n_map["n"]) == 0:
        pos_rows = conn.execute(
            "SELECT id, code FROM crm_positions WHERE active = 1"
        ).fetchall()
        met_rows = conn.execute(
            "SELECT id, code FROM crm_kpi_metrics WHERE active = 1"
        ).fetchall()
        code_to_mid = {str(r["code"]): int(r["id"]) for r in met_rows}
        pos_code_to_pid = {str(r["code"]): int(r["id"]) for r in pos_rows}
        role_metrics: dict[str, list[str]] = {
            "CSKH-01": ["SLA_PCT", "CSAT", "CASES_OK"],
            "KD-01": ["TARGET_LEADS", "SLA_PCT", "CSAT"],
            "VH-01": ["CASES_OK", "SLA_PCT"],
        }
        so = 0
        for pcode, mcodes in role_metrics.items():
            pid = pos_code_to_pid.get(pcode)
            if not pid:
                continue
            for mcode in mcodes:
                mid = code_to_mid.get(mcode)
                if not mid:
                    continue
                so += 10
                conn.execute(
                    """
                    INSERT OR IGNORE INTO crm_position_kpi_metrics (position_id, metric_id, sort_order)
                    VALUES (?, ?, ?)
                    """,
                    (pid, mid, so),
                )


def _ensure_crm_hub_schema(conn: sqlite3.Connection) -> None:
    """Chiến dịch marketing, hợp đồng, nhắc nhở; gắn campaign_id lên crm_cases."""
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_campaigns (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL DEFAULT '',
            name TEXT NOT NULL,
            channel TEXT NOT NULL DEFAULT 'other',
            external_ref TEXT NOT NULL DEFAULT '',
            utm_campaign TEXT NOT NULL DEFAULT '',
            notes TEXT NOT NULL DEFAULT '',
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_crm_campaigns_code_nn
        ON crm_campaigns(code)
        WHERE TRIM(code) != ''
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_campaigns_active ON crm_campaigns(active, name)"
    )
    _ensure_crm_campaigns_hub_columns(conn)

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_contracts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER NOT NULL REFERENCES crm_customers(id),
            case_id INTEGER REFERENCES crm_cases(id) ON DELETE SET NULL,
            campaign_id INTEGER REFERENCES crm_campaigns(id) ON DELETE SET NULL,
            reference_code TEXT NOT NULL DEFAULT '',
            title TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'draft',
            signed_on TEXT NOT NULL DEFAULT '',
            starts_on TEXT NOT NULL DEFAULT '',
            ends_on TEXT NOT NULL DEFAULT '',
            amount_vnd INTEGER NOT NULL DEFAULT 0,
            renewal_reminder_days INTEGER NOT NULL DEFAULT 30,
            notes TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_contracts_customer ON crm_contracts(customer_id)"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_contracts_dates ON crm_contracts(status, ends_on)"
    )

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_reminders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            scope TEXT NOT NULL,
            ref_id INTEGER NOT NULL DEFAULT 0,
            reminder_kind TEXT NOT NULL DEFAULT 'manual',
            title TEXT NOT NULL,
            body TEXT NOT NULL DEFAULT '',
            remind_at TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'pending',
            staff_id INTEGER REFERENCES crm_staff(id) ON DELETE SET NULL,
            meta_json TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_reminders_due ON crm_reminders(status, remind_at)"
    )

    cc = {r[1] for r in conn.execute("PRAGMA table_info(crm_cases)")}
    if "campaign_id" not in cc:
        try:
            conn.execute(
                "ALTER TABLE crm_cases ADD COLUMN campaign_id INTEGER "
                "REFERENCES crm_campaigns(id)"
            )
        except sqlite3.Error:
            pass


def _ensure_crm_campaigns_hub_columns(conn: sqlite3.Connection) -> None:
    """Phase 2 — agency client + Meta map sync metadata on Hub campaigns."""
    cols = {row[1] for row in conn.execute("PRAGMA table_info(crm_campaigns)").fetchall()}
    migrations = [
        ("agency_client_id", "ALTER TABLE crm_campaigns ADD COLUMN agency_client_id TEXT NOT NULL DEFAULT ''"),
        ("target_cpl_vnd", "ALTER TABLE crm_campaigns ADD COLUMN target_cpl_vnd INTEGER NOT NULL DEFAULT 0"),
        ("hub_map_synced_at", "ALTER TABLE crm_campaigns ADD COLUMN hub_map_synced_at TEXT NOT NULL DEFAULT ''"),
        ("hub_map_last_error", "ALTER TABLE crm_campaigns ADD COLUMN hub_map_last_error TEXT NOT NULL DEFAULT ''"),
    ]
    for name, sql in migrations:
        if name not in cols:
            conn.execute(sql)


def _crm_apply_campaign_hub_fields(merged: dict[str, Any], payload: dict[str, Any]) -> None:
    if "agency_client_id" in payload:
        merged["agency_client_id"] = str(payload.get("agency_client_id") or "").strip()[:64]
    if "target_cpl_vnd" in payload:
        try:
            merged["target_cpl_vnd"] = max(0, int(payload.get("target_cpl_vnd") or 0))
        except (TypeError, ValueError):
            merged["target_cpl_vnd"] = 0


def _crm_sync_hub_campaign_to_pg(campaign_id: int, conn: sqlite3.Connection) -> None:
    try:
        from ptt_agency.hub_campaign_sync import sync_campaign_row

        row = conn.execute("SELECT * FROM crm_campaigns WHERE id = ?", (campaign_id,)).fetchone()
        if row:
            sync_campaign_row(dict(row), sqlite_conn=conn)
            conn.commit()
    except Exception as exc:
        logger.debug("hub_campaign_map sync skipped campaign_id=%s: %s", campaign_id, exc)


def _ensure_crm_sop_schema(conn: sqlite3.Connection) -> None:
    """Quy trình Marketing (SOP): templates, steps, runs, run_tasks."""
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_sop_templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL DEFAULT '',
            name TEXT NOT NULL,
            channel TEXT NOT NULL DEFAULT 'other',
            description TEXT NOT NULL DEFAULT '',
            notes TEXT NOT NULL DEFAULT '',
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_sop_tpl_active ON crm_sop_templates(active, channel)"
    )

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_sop_steps (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            template_id INTEGER NOT NULL REFERENCES crm_sop_templates(id) ON DELETE CASCADE,
            position INTEGER NOT NULL DEFAULT 0,
            title TEXT NOT NULL,
            description TEXT NOT NULL DEFAULT '',
            offset_days INTEGER NOT NULL DEFAULT 0,
            duration_days INTEGER NOT NULL DEFAULT 1,
            role TEXT NOT NULL DEFAULT 'any',
            required INTEGER NOT NULL DEFAULT 1,
            checklist_json TEXT NOT NULL DEFAULT '[]',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_sop_steps_tpl ON crm_sop_steps(template_id, position)"
    )

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_sop_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            campaign_id INTEGER REFERENCES crm_campaigns(id) ON DELETE SET NULL,
            template_id INTEGER REFERENCES crm_sop_templates(id) ON DELETE SET NULL,
            name TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'active',
            start_date TEXT NOT NULL DEFAULT '',
            notes TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_sop_runs_status ON crm_sop_runs(status, start_date)"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_sop_runs_campaign ON crm_sop_runs(campaign_id)"
    )

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_sop_run_tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id INTEGER NOT NULL REFERENCES crm_sop_runs(id) ON DELETE CASCADE,
            step_id INTEGER REFERENCES crm_sop_steps(id) ON DELETE SET NULL,
            position INTEGER NOT NULL DEFAULT 0,
            title TEXT NOT NULL,
            description TEXT NOT NULL DEFAULT '',
            role TEXT NOT NULL DEFAULT 'any',
            due_date TEXT NOT NULL DEFAULT '',
            status TEXT NOT NULL DEFAULT 'todo',
            assigned_staff_id INTEGER REFERENCES crm_staff(id) ON DELETE SET NULL,
            notes TEXT NOT NULL DEFAULT '',
            checklist_json TEXT NOT NULL DEFAULT '[]',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_sop_rtasks_run ON crm_sop_run_tasks(run_id, position)"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_sop_rtasks_due ON crm_sop_run_tasks(status, due_date)"
    )
    seed_launch_campaign_sop_template(conn)


def _ensure_crm_marketing_plan_schema(conn: sqlite3.Connection) -> None:
    """Kế hoạch marketing chiến lược — liên kết chiến dịch, mục tiêu, KPI, cột mốc."""
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_marketing_plans (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL DEFAULT '',
            name TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'draft',
            priority TEXT NOT NULL DEFAULT 'normal',
            fiscal_year INTEGER NOT NULL DEFAULT 2026,
            period_label TEXT NOT NULL DEFAULT '',
            north_star TEXT NOT NULL DEFAULT '',
            objectives TEXT NOT NULL DEFAULT '',
            pillars_json TEXT NOT NULL DEFAULT '[]',
            audiences TEXT NOT NULL DEFAULT '',
            channels_focus_json TEXT NOT NULL DEFAULT '[]',
            budget_planned_vnd INTEGER NOT NULL DEFAULT 0,
            budget_actual_vnd INTEGER NOT NULL DEFAULT 0,
            success_metrics_json TEXT NOT NULL DEFAULT '[]',
            risks_notes TEXT NOT NULL DEFAULT '',
            owner_staff_id INTEGER REFERENCES crm_staff(id) ON DELETE SET NULL,
            start_date TEXT NOT NULL DEFAULT '',
            end_date TEXT NOT NULL DEFAULT '',
            notes TEXT NOT NULL DEFAULT '',
            strategy_framework_json TEXT NOT NULL DEFAULT '{}',
            target_market_prof_json TEXT NOT NULL DEFAULT '{}',
            target_market_steps4_json TEXT NOT NULL DEFAULT '{}',
            khtn_market_research_json TEXT NOT NULL DEFAULT '{}',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    _mp_cols = {r[1] for r in conn.execute("PRAGMA table_info(crm_marketing_plans)")}
    if "strategy_framework_json" not in _mp_cols:
        try:
            conn.execute(
                "ALTER TABLE crm_marketing_plans ADD COLUMN strategy_framework_json TEXT NOT NULL DEFAULT '{}'"
            )
        except sqlite3.Error:
            pass
    if "target_market_prof_json" not in _mp_cols:
        try:
            conn.execute(
                "ALTER TABLE crm_marketing_plans ADD COLUMN target_market_prof_json TEXT NOT NULL DEFAULT '{}'"
            )
        except sqlite3.Error:
            pass
    if "target_market_steps4_json" not in _mp_cols:
        try:
            conn.execute(
                "ALTER TABLE crm_marketing_plans ADD COLUMN target_market_steps4_json TEXT NOT NULL DEFAULT '{}'"
            )
        except sqlite3.Error:
            pass
    if "khtn_market_research_json" not in _mp_cols:
        try:
            conn.execute(
                "ALTER TABLE crm_marketing_plans ADD COLUMN khtn_market_research_json TEXT NOT NULL DEFAULT '{}'"
            )
        except sqlite3.Error:
            pass
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_mplan_year_status ON crm_marketing_plans(fiscal_year, status)"
    )

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_marketing_plan_campaigns (
            plan_id INTEGER NOT NULL REFERENCES crm_marketing_plans(id) ON DELETE CASCADE,
            campaign_id INTEGER NOT NULL REFERENCES crm_campaigns(id) ON DELETE CASCADE,
            PRIMARY KEY (plan_id, campaign_id)
        )
        """
    )

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_marketing_plan_milestones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            plan_id INTEGER NOT NULL REFERENCES crm_marketing_plans(id) ON DELETE CASCADE,
            position INTEGER NOT NULL DEFAULT 0,
            title TEXT NOT NULL,
            due_date TEXT NOT NULL DEFAULT '',
            status TEXT NOT NULL DEFAULT 'pending',
            notes TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_mmile_plan ON crm_marketing_plan_milestones(plan_id, due_date)"
    )


def _ensure_crm_attendance_payroll_schema(conn: sqlite3.Connection) -> None:
    """Chấm công theo ngày và bảng lương tháng (SQLite, idempotent)."""
    sc = {r[1] for r in conn.execute("PRAGMA table_info(crm_staff)")}
    if "base_salary_vnd" not in sc:
        try:
            conn.execute(
                "ALTER TABLE crm_staff ADD COLUMN base_salary_vnd INTEGER NOT NULL DEFAULT 0"
            )
        except sqlite3.Error:
            pass
    if "attendance_pin" not in sc:
        try:
            conn.execute(
                "ALTER TABLE crm_staff ADD COLUMN attendance_pin TEXT NOT NULL DEFAULT ''"
            )
        except sqlite3.Error:
            pass
    if "sales_level" not in sc:
        try:
            conn.execute(
                "ALTER TABLE crm_staff ADD COLUMN sales_level TEXT NOT NULL DEFAULT ''"
            )
        except sqlite3.Error:
            pass

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            staff_id INTEGER NOT NULL REFERENCES crm_staff(id) ON DELETE CASCADE,
            work_date TEXT NOT NULL,
            check_in TEXT NOT NULL DEFAULT '',
            check_out TEXT NOT NULL DEFAULT '',
            break_minutes INTEGER NOT NULL DEFAULT 0,
            note TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE(staff_id, work_date)
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_attendance_range ON crm_attendance(work_date, staff_id)"
    )

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_payroll (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            workdays_standard INTEGER NOT NULL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'draft',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE(year, month)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_payroll_line (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            payroll_id INTEGER NOT NULL REFERENCES crm_payroll(id) ON DELETE CASCADE,
            staff_id INTEGER NOT NULL REFERENCES crm_staff(id),
            days_present INTEGER NOT NULL DEFAULT 0,
            base_salary_vnd INTEGER NOT NULL DEFAULT 0,
            salary_from_attendance_vnd INTEGER NOT NULL DEFAULT 0,
            allowances_vnd INTEGER NOT NULL DEFAULT 0,
            deductions_vnd INTEGER NOT NULL DEFAULT 0,
            net_salary_vnd INTEGER NOT NULL DEFAULT 0,
            note TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE(payroll_id, staff_id)
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_payroll_line_staff ON crm_payroll_line(staff_id)"
    )
    ensure_payroll_policy_schema(conn)


def _crm_weekdays_in_month(year: int, month: int) -> int:
    """Số ngày làm việc trong tháng theo chính sách payroll."""
    with get_connection() as conn:
        policy = load_policy(conn)
    weekdays = parse_work_weekdays(str(policy.get("work_weekdays") or ""))
    return count_workdays_in_month(year, month, weekdays)


def _crm_month_bounds(year: int, month: int) -> tuple[str, str]:
    last = monthrange(year, month)[1]
    return (
        f"{year:04d}-{month:02d}-01",
        f"{year:04d}-{month:02d}-{last:02d}",
    )


def _ensure_crm_positions_schema(conn: sqlite3.Connection) -> None:
    """Danh mục chức vụ (chức danh chuẩn)."""
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_positions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL DEFAULT '',
            name TEXT NOT NULL,
            description TEXT NOT NULL DEFAULT '',
            sort_order INTEGER NOT NULL DEFAULT 0,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_crm_positions_active ON crm_positions(active, sort_order)"
    )
    sc = {r[1] for r in conn.execute("PRAGMA table_info(crm_staff)")}
    if "position_id" not in sc:
        try:
            conn.execute(
                "ALTER TABLE crm_staff ADD COLUMN position_id INTEGER REFERENCES crm_positions(id)"
            )
        except sqlite3.Error:
            pass

    np = conn.execute("SELECT COUNT(*) AS n FROM crm_positions").fetchone()
    if np and int(np["n"]) == 0:
        ts = datetime.now().strftime("%Y-%m-%d")
        now = _crm_ts()
        seed = [
            ("CSKH-01", "Chuyên viên CSKH", "Xử lý tiếp nhận yêu cầu.", 10),
            ("KD-01", "Nhân viên kinh doanh", "Chỉ nhận và chăm sóc lead được phân công.", 20),
            ("VH-01", "Điều phối vận hành", "Phối hợp triển khai dự án.", 30),
        ]
        for code, name, desc, so in seed:
            conn.execute(
                """
                INSERT INTO crm_positions (code, name, description, sort_order, active, created_at, updated_at)
                VALUES (?, ?, ?, ?, 1, ?, ?)
                """,
                (code, name, desc, so, ts, now),
            )


# Quy trình chăm sóc khách hàng (Kanban / ticket)
CRM_STATUSES_ORDER: tuple[str, ...] = (
    "tiep_nhan",
    "dang_xu_ly",
    "cho_khach",
    "da_giai_quyet",
    "dong",
)
CRM_STATUS_LABELS_VI: dict[str, str] = {
    "tiep_nhan": "Tiếp nhận",
    "dang_xu_ly": "Đang xử lý",
    "cho_khach": "Chờ KH phản hồi",
    "da_giai_quyet": "Đã giải quyết",
    "dong": "Đã đóng",
}
CRM_CHANNELS: tuple[str, ...] = (
    "dien_thoai",
    "email",
    "zalo",
    "truc_tiep",
    "khac",
)
CRM_CHANNEL_LABELS_VI: dict[str, str] = {
    "dien_thoai": "Điện thoại",
    "email": "Email",
    "zalo": "Zalo",
    "truc_tiep": "Trực tiếp",
    "khac": "Khác",
}
CRM_PRIORITIES: tuple[str, ...] = ("thap", "binh_thuong", "cao")
CRM_PRIORITY_LABELS_VI: dict[str, str] = {
    "thap": "Thấp",
    "binh_thuong": "Bình thường",
    "cao": "Cao",
}
CRM_CAMPAIGN_CHANNELS: tuple[str, ...] = (
    "meta",
    "email",
    "social",
    "ads",
    "seo",
    "event",
    "partner",
    "other",
)
CRM_CAMPAIGN_CHANNEL_LABELS_VI: dict[str, str] = {
    "meta": "Meta Ads",
    "email": "Email",
    "social": "Mạng xã hội",
    "ads": "Quảng cáo trả phí",
    "seo": "SEO / nội dung",
    "event": "Sự kiện",
    "partner": "Đối tác",
    "other": "Khác",
}
CRM_CONTRACT_STATUSES: tuple[str, ...] = (
    "draft",
    "negotiation",
    "signed",
    "active",
    "expiring",
    "paused",
    "completed",
    "renewed",
    "lost",
    "cancelled",
)
CRM_CONTRACT_STATUS_LABELS_VI: dict[str, str] = {
    "draft": "Nháp",
    "negotiation": "Đang thương lượng",
    "signed": "Đã ký",
    "active": "Đang hiệu lực",
    "expiring": "Sắp hết hạn",
    "paused": "Tạm dừng",
    "completed": "Hoàn tất",
    "renewed": "Đã gia hạn",
    "lost": "Mất cơ hội",
    "cancelled": "Huỷ",
}
CRM_REMINDER_SCOPES: tuple[str, ...] = (
    "case",
    "contract",
    "customer",
    "finance_kpi",
    "owner_weekly",
    "general",
)
CRM_REMINDER_KINDS: tuple[str, ...] = (
    "manual",
    "contract_renewal",
    "status_followup",
    "kpi_alert",
    "owner_weekly_alert",
)
CRM_REMINDER_STATUSES: tuple[str, ...] = ("pending", "done", "dismissed")
CRM_REMINDER_SCOPE_LABELS_VI: dict[str, str] = {
    "case": "Yêu cầu CRM",
    "contract": "Hợp đồng",
    "customer": "Khách hàng",
    "finance_kpi": "KPI tài chính",
    "owner_weekly": "Dashboard tuần (Chủ DN)",
    "general": "Chung",
}
CRM_REMINDER_KIND_LABELS_VI: dict[str, str] = {
    "manual": "Thủ công",
    "contract_renewal": "Gia hạn hợp đồng",
    "status_followup": "Theo dõi trạng thái",
    "kpi_alert": "Cảnh báo KPI",
    "owner_weekly_alert": "Cảnh báo tuần (RAG)",
}
CRM_REMINDER_STATUS_LABELS_VI: dict[str, str] = {
    "pending": "Chờ xử lý",
    "done": "Đã xong",
    "dismissed": "Đã bỏ qua",
}
CRM_SOP_TASK_STATUSES: tuple[str, ...] = ("todo", "in_progress", "done", "skipped")
CRM_SOP_TASK_STATUS_LABELS_VI: dict[str, str] = {
    "todo": "Chờ làm",
    "in_progress": "Đang thực hiện",
    "done": "Hoàn thành",
    "skipped": "Bỏ qua",
}
CRM_SOP_RUN_STATUSES: tuple[str, ...] = ("draft", "active", "paused", "completed", "archived")
CRM_SOP_RUN_STATUS_LABELS_VI: dict[str, str] = {
    "draft": "Nháp",
    "active": "Đang chạy",
    "paused": "Tạm dừng",
    "completed": "Hoàn tất",
    "archived": "Lưu trữ",
}
CRM_SOP_STEP_ROLES: tuple[str, ...] = (
    "any", "strategy", "content", "design", "ads", "seo", "analytics", "approver", "client"
)
CRM_SOP_STEP_ROLE_LABELS_VI: dict[str, str] = {
    "any": "Bất kỳ",
    "strategy": "Chiến lược",
    "content": "Nội dung",
    "design": "Thiết kế",
    "ads": "Quảng cáo",
    "seo": "SEO",
    "analytics": "Phân tích",
    "approver": "Duyệt",
    "client": "Khách hàng",
}
CRM_MARKETING_PLAN_STATUSES: tuple[str, ...] = (
    "draft",
    "review",
    "active",
    "paused",
    "completed",
    "archived",
    "cancelled",
)
CRM_MARKETING_PLAN_STATUS_LABELS_VI: dict[str, str] = {
    "draft": "Nháp",
    "review": "Chờ phê duyệt",
    "active": "Đang triển khai",
    "paused": "Tạm dừng",
    "completed": "Đã đóng chu kỳ",
    "archived": "Lưu trữ",
    "cancelled": "Huỷ bỏ",
}
CRM_MARKETING_PLAN_PRIORITIES: tuple[str, ...] = ("low", "normal", "high", "critical")
CRM_MARKETING_PLAN_PRIORITY_LABELS_VI: dict[str, str] = {
    "low": "Thấp",
    "normal": "Bình thường",
    "high": "Cao",
    "critical": "Rất cao",
}
CRM_MARKETING_PLAN_MS_STATUSES: tuple[str, ...] = ("pending", "in_progress", "done", "cancelled")
CRM_MARKETING_PLAN_MS_STATUS_LABELS_VI: dict[str, str] = {
    "pending": "Chờ",
    "in_progress": "Đang làm",
    "done": "Hoàn tất",
    "cancelled": "Huỷ",
}
# Khối nghiệp vụ marketing chuyên nghiệp — lưu trong strategy_framework_json (object).
CRM_MP_STRATEGY_FRAMEWORK_KEYS: tuple[str, ...] = (
    "target_market",
    "market_message",
    "media_reach",
    "retention_system",
    "nurture_system",
    "conversion_strategy",
    "world_class_experience",
    "lifecycle_extension",
    "referral_engine",
)
CRM_MP_STRATEGY_FRAMEWORK_LABELS_VI: dict[str, str] = {
    "target_market": "TMMT — tóm tắt định vị TMMT ngắn gọn (2–4 dòng elevator); chi tiết bên dưới",
    "market_message": "Thông điệp tới TMMT — USP, định vị, câu chuyện thương hiệu, proof",
    "media_reach": "Phương tiện truyền thông tiếp cận TMMT — paid / owned / earned, mix, tiêu chí đo",
    "retention_system": "Giữ chân người quan tâm — nội dung, onboarding, điểm chạm sau click",
    "nurture_system": "Chăm sóc người quan tâm — CRM flow, scoring, SLA phản hồi",
    "conversion_strategy": "Chuyển đổi khách — funnel, offer, xử lý phản đối, urgency & friction",
    "world_class_experience": "Trải nghiệm đẳng cấp — chuẩn dịch vụ, omnichannel, ‘wow moments’",
    "lifecycle_extension": "Gia tăng vòng đời KH — upsell/cross-sell, loyalty, win-back",
    "referral_engine": "Thúc đẩy lời giới thiệu — chương trình giới thiệu, social proof, NPS",
}

# TMMT chi tiết (chương trình chuyên nghiệp) — cột riêng target_market_prof_json.
CRM_MP_TARGET_MARKET_PROF_KEYS: tuple[str, ...] = (
    "market_context",
    "tam_sam_som",
    "geo_behavior",
    "segmentation_icp",
    "personas_roles",
    "jobs_to_be_done",
    "pains_desired_outcomes",
    "buy_triggers_obstacles",
    "criteria_vs_alternatives",
    "insights_evidence",
    "segment_priorities",
    "success_hypotheses_next",
)
CRM_MP_TARGET_MARKET_PROF_LABELS_VI: dict[str, str] = {
    "market_context": "Bối cảnh thị trường — ngành, chu kỳ, xu hướng, quy định",
    "tam_sam_som": "TAM / SAM / SOM — ước tính, đơn vị, giả định & nguồn chứng cứ",
    "geo_behavior": "Địa lý & hành vi — nơi tìm thông tin / mua, thiết bị, thời điểm",
    "segmentation_icp": "Phân khúc & ICP — tiêu chí chia, cỡ DN/đối tượng",
    "personas_roles": "Persona & vai trò — người dùng / ký duyệt / chặn trong B2B",
    "jobs_to_be_done": "Jobs-to-be-done — tình huống, kỳ vọng “thuê” giải pháp",
    "pains_desired_outcomes": "Pain & gain — vấn đề, chỉ số đo sau khi chuyển đổi",
    "buy_triggers_obstacles": "Kích hoạt mua & rào cản — urgency, tin cậy, switching cost",
    "criteria_vs_alternatives": "Tiêu chí chọn NCC vs thay thế — đối thủ / tự làm / không làm gì",
    "insights_evidence": "Insight & bằng chứng — phỏng vấn, khảo sát, dữ liệu",
    "segment_priorities": "Ưu tiên TMMT kỳ này — phân khúc ưu tiên & lý do",
    "success_hypotheses_next": "Giả thuyết kiểm chứng TMMT & bước nghiên cứu tiếp theo",
}

# TMMT — quy trình 4 bước (chuẩn BĐS, áp dụng được ngành khác) — cột target_market_steps4_json.
CRM_MP_TARGET_MARKET_STEPS4_KEYS: tuple[str, ...] = (
    "step1_market_size",
    "step1_trends",
    "step1_competitors",
    "step1_tools_sources",
    "step2_demographics",
    "step2_geography",
    "step2_behavior",
    "step2_psychographics",
    "step2_data_methods",
    "step3_priority_segments",
    "step3_persona_primary",
    "step3_persona_secondary",
    "step4_criteria_evaluation",
    "step4_final_decision",
)
CRM_MP_TARGET_MARKET_STEPS4_LABELS_VI: dict[str, str] = {
    "step1_market_size": "[Bước 1] Dung lượng thị trường — số KH tiềm năng, doanh thu ước tính, đơn vị (VD BĐS: quy mô KH TP.HCM…)",
    "step1_trends": "[Bước 1] Xu hướng — chuyển dịch hành vi tiêu dùng, tăng trưởng phân khúc (%)",
    "step1_competitors": "[Bước 1] Đối thủ — điểm mạnh/yếu, vị trí, giá, dịch vụ so với bạn",
    "step1_tools_sources": "[Bước 1] Nguồn & công cụ — báo cáo nghiên cứu, sàn giao dịch, tin ngành, nội bộ…",
    "step2_demographics": "[Bước 2] Nhân khẩu — tuổi, giới, thu nhập, học vấn/ nghề nghiệp (áp vào TM của bạn)",
    "step2_geography": "[Bước 2] Địa lý — quận/huyện, tiện ích gần (trường, bệnh viện), không gian sống",
    "step2_behavior": "[Bước 2] Hành vi — mục đích mua (ở/đầu tư/cho thuê), kênh tìm kiếm, tần suất mua",
    "step2_psychographics": "[Bước 2] Tâm lý — sở thích, lối sống, giá trị (chất lượng sống, an toàn tài chính…)",
    "step2_data_methods": "[Bước 2] Thu thập dữ liệu — CRM hiện tại, khảo sát, phỏng vấn, focus group, Analytics…",
    "step3_priority_segments": "[Bước 3] Chọn 1–2 phân khúc ưu tiên — vì sao có tiềm năng mua/chuyển đổi cao?",
    "step3_persona_primary": "[Bước 3] Phân khúc / persona chính — bảng chi tiết (tuổi, nghề, thu nhập, nhu cầu, pain, kênh…)",
    "step3_persona_secondary": "[Bước 3] Phân khúc / persona phụ (nếu có) — hoặc bảng so sánh hai cột Phân khúc 1 / 2",
    "step4_criteria_evaluation": "[Bước 4] Đánh giá theo 4 tiêu chí: phục vụ tốt? doanh thu ổn định? có mở rộng? lợi thế CT so đối thủ?",
    "step4_final_decision": "[Bước 4] Quyết định TMMT cuối — chọn phân khúc nào và lý do (VD Sàn căn hộ→gia đình trẻ)",
}
CRM_MP_TARGET_MARKET_STEPS4_SECTIONS: tuple[tuple[str, tuple[str, ...]], ...] = (
    (
        "Bước 1 — Phân tích tổng quan thị trường · nắm bức tranh (ví dụ thị trường BĐS)",
        (
            "step1_market_size",
            "step1_trends",
            "step1_competitors",
            "step1_tools_sources",
        ),
    ),
    (
        "Bước 2 — Phân khúc thị trường",
        (
            "step2_demographics",
            "step2_geography",
            "step2_behavior",
            "step2_psychographics",
            "step2_data_methods",
        ),
    ),
    (
        "Bước 3 — Xác định khách hàng mục tiêu",
        (
            "step3_priority_segments",
            "step3_persona_primary",
            "step3_persona_secondary",
        ),
    ),
    (
        "Bước 4 — Lựa chọn thị trường mục tiêu cuối cùng",
        ("step4_criteria_evaluation", "step4_final_decision"),
    ),
)

CRM_MP_STRATEGY_FRAMEWORK_KEYS_WITHOUT_TMMT: tuple[str, ...] = tuple(
    k for k in CRM_MP_STRATEGY_FRAMEWORK_KEYS if k != "target_market"
)
CRM_EMPLOYMENT_TYPES: tuple[str, ...] = ("full_time", "contractor", "intern", "collaborator")
CRM_EMPLOYMENT_LABELS_VI: dict[str, str] = {
    "full_time": "Chính thức",
    "contractor": "Cộng tác / Hợp đồng",
    "intern": "Thực tập",
    "collaborator": "Cộng tác viên",
}

CRM_KPI_STATUSES: tuple[str, ...] = ("draft", "achieved", "at_risk", "missed")
CRM_KPI_STATUS_LABELS_VI: dict[str, str] = {
    "draft": "Nháp",
    "achieved": "Đạt",
    "at_risk": "Rủi ro",
    "missed": "Không đạt",
}


def _crm_kpi_access_mode() -> str:
    """Quyền trang/API KPI: full | view | off (biến môi trường CRM_KPI_ACCESS)."""
    if _staff_logged_in():
        return "staff"
    v = (os.getenv("CRM_KPI_ACCESS") or "full").strip().lower()
    return v if v in ("full", "view", "off") else "full"


def _crm_position_metric_ids(conn: sqlite3.Connection, position_id: int | None) -> list[int]:
    if position_id is not None and position_id > 0:
        rows = conn.execute(
            """
            SELECT metric_id FROM crm_position_kpi_metrics
            WHERE position_id = ?
            ORDER BY sort_order ASC, metric_id ASC
            """,
            (position_id,),
        ).fetchall()
        if rows:
            return [int(r["metric_id"]) for r in rows]
    rows = conn.execute(
        "SELECT id FROM crm_kpi_metrics WHERE active = 1 ORDER BY sort_order ASC, id ASC"
    ).fetchall()
    return [int(r["id"]) for r in rows]


def _crm_staff_profile(conn: sqlite3.Connection, staff_id: int) -> dict[str, Any] | None:
    row = conn.execute(
        """
        SELECT s.id, s.name, s.email, s.phone, s.job_title, s.department, s.position_id,
               p.name AS position_name, p.code AS position_code
        FROM crm_staff s
        LEFT JOIN crm_positions p ON p.id = s.position_id
        WHERE s.id = ? AND s.active = 1
        """,
        (staff_id,),
    ).fetchone()
    return dict(row) if row else None


def _crm_kpi_access_json_error() -> Any | None:
    """403 nếu tắt hoàn toàn KPI (admin luôn được truy cập)."""
    if _admin_logged_in():
        return None
    if _staff_logged_in():
        return None
    if _crm_kpi_access_mode() == "off":
        return (
            jsonify(
                {
                    "error": "KPI đã tắt trên server (đặt CRM_KPI_ACCESS=full hoặc view).",
                }
            ),
            403,
        )
    return None


def _crm_kpi_require_edit_json() -> Any | None:
    """403 nếu chỉ được xem (view) hoặc tắt KPI — admin & nhân viên portal được sửa."""
    if _admin_logged_in():
        return None
    if _staff_logged_in():
        return None
    err = _crm_kpi_access_json_error()
    if err is not None:
        return err
    if _crm_kpi_access_mode() != "full":
        return (
            jsonify(
                {
                    "error": "Chế độ chỉ xem KPI — không thể chỉnh sửa (CRM_KPI_ACCESS=view).",
                }
            ),
            403,
        )
    return None


def _crm_kpi_achievement_pct(higher_is_better: int, target_value: Any, actual_value: Any) -> float | None:
    """Phần trăm đạt (0–100) để biểu đồ / export; None nếu không suy ra được."""
    if target_value is None or actual_value is None:
        return None
    try:
        t = float(target_value)
        a = float(actual_value)
    except (TypeError, ValueError):
        return None
    if t == 0:
        return None
    hi = int(higher_is_better or 1) == 1
    if hi:
        return round(100.0 * min(1.0, a / t), 2)
    return round(100.0 * min(1.0, t / max(a, 1e-9)), 2)


def _crm_kpi_derive_alert(
    *,
    status: str,
    higher_is_better: int,
    warn_ratio: Any,
    target_value: Any,
    actual_value: Any,
) -> tuple[str | None, str | None]:
    """(critical|warn|None, reason_code)."""
    st = (status or "draft").strip().lower()
    if st == "missed":
        return "critical", "status_missed"
    if st == "at_risk":
        return "warn", "status_at_risk"
    if warn_ratio is None or target_value is None or actual_value is None:
        return None, None
    try:
        wr = float(warn_ratio)
        t = float(target_value)
        a = float(actual_value)
    except (TypeError, ValueError):
        return None, None
    if wr <= 0 or wr > 10 or t == 0:
        return None, None
    hi = int(higher_is_better or 1) == 1
    if hi:
        if a < t * wr:
            return "warn", "below_threshold"
    else:
        if a > t / wr:
            return "warn", "above_threshold"
    return None, None


def _crm_parse_staff_kpi_query_args() -> tuple[int | None, int | None, int | None, str | None]:
    """year, month, staff_id|None, error message."""
    staff_raw = (request.args.get("staff_id") or "").strip()
    try:
        year = int(request.args.get("year") or 0)
        month = int(request.args.get("month") or 0)
    except (TypeError, ValueError):
        return None, None, None, "year/month không hợp lệ"
    if year < 2000 or year > 2100 or month < 1 or month > 12:
        return None, None, None, "Kỳ không hợp lệ"
    staff_id: int | None = None
    if staff_raw:
        try:
            staff_id = int(staff_raw)
        except ValueError:
            return None, None, None, "staff_id không hợp lệ"
        if staff_id <= 0:
            staff_id = None
    eff = _crm_effective_staff_id()
    if eff is not None:
        if staff_id is not None and staff_id != eff:
            return None, None, None, "Không xem KPI nhân viên khác"
        staff_id = eff
    return year, month, staff_id, None


def _crm_staff_position_id(conn: sqlite3.Connection, staff_id: int) -> int | None:
    row = conn.execute(
        "SELECT position_id FROM crm_staff WHERE id = ? AND active = 1",
        (staff_id,),
    ).fetchone()
    if row is None:
        return None
    try:
        pid = int(row["position_id"] or 0)
        return pid if pid > 0 else None
    except (TypeError, ValueError):
        return None


def _crm_staff_metric_allowed(
    conn: sqlite3.Connection, staff_id: int, metric_id: int
) -> bool:
    position_id = _crm_staff_position_id(conn, staff_id)
    allowed = _crm_position_metric_ids(conn, position_id)
    return metric_id in allowed


def _crm_ensure_staff_kpi_rows(
    conn: sqlite3.Connection,
    staff_id: int,
    position_id: int | None,
    year: int,
    month: int,
) -> None:
    """Tạo dòng KPI nháp cho các chỉ tiêu theo chức vụ (nếu chưa có)."""
    metric_ids = _crm_position_metric_ids(conn, position_id)
    if not metric_ids:
        return
    ts = _crm_ts()
    ts_d = datetime.now().strftime("%Y-%m-%d")
    for mid in metric_ids:
        conn.execute(
            """
            INSERT OR IGNORE INTO crm_staff_kpi (
                staff_id, metric_id, year, month,
                target_value, actual_value, status, note, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, NULL, NULL, 'draft', '', ?, ?)
            """,
            (staff_id, mid, year, month, ts_d, ts),
        )


def _crm_kpi_alert_label_vi(level: str | None, reason: str | None) -> str:
    if level == "critical":
        return "Nghiêm trọng — không đạt (trạng thái)"
    if level == "warn":
        if reason == "status_at_risk":
            return "Cảnh báo — rủi ro (trạng thái)"
        if reason in ("below_threshold", "above_threshold"):
            return "Cảnh báo — vượt ngưỡng rủi ro chỉ tiêu"
    return ""


# JOIN khách hàng + nhân viên phụ trách (LEFT: hồ sơ cũ / chưa gán)
_CRM_CASE_SELECT = """
SELECT c.*,
       cu.name AS customer_name,
       cu.phone AS customer_phone,
       cu.email AS customer_email,
       cu.address AS customer_address,
       cu.company AS customer_company,
       st.name AS staff_display_name,
       camp.code AS campaign_code,
       camp.name AS campaign_name
FROM crm_cases c
JOIN crm_customers cu ON cu.id = c.customer_id
LEFT JOIN crm_staff st ON st.id = c.assigned_staff_id
LEFT JOIN crm_campaigns camp ON camp.id = c.campaign_id
"""


def _crm_ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _crm_slug_lead_channel_code(raw: str) -> str:
    s = re.sub(r"[^\w\s-]", "", str(raw or "").lower())
    s = re.sub(r"[\s_-]+", "_", s).strip("_")
    return (s[:32] or "kenh")


def _crm_lead_channel_labels_map(conn: sqlite3.Connection) -> dict[str, str]:
    labels = dict(CRM_CHANNEL_LABELS_VI)
    rows = conn.execute(
        """
        SELECT code, name FROM crm_lead_channels
        WHERE active = 1 AND trim(code) != ''
        ORDER BY sort_order ASC, name COLLATE NOCASE ASC
        """
    ).fetchall()
    for row in rows:
        labels[str(row["code"])] = str(row["name"])
    return labels


def _crm_resolve_lead_channel(conn: sqlite3.Connection, raw: str | None) -> str:
    code = str(raw or "").strip()
    if not code:
        code = "khac"
    hit = conn.execute(
        """
        SELECT code FROM crm_lead_channels
        WHERE lower(trim(code)) = lower(?) AND active = 1
        LIMIT 1
        """,
        (code,),
    ).fetchone()
    if hit:
        return str(hit["code"])
    fallback = conn.execute(
        """
        SELECT code FROM crm_lead_channels
        WHERE lower(trim(code)) = 'khac' AND active = 1
        LIMIT 1
        """
    ).fetchone()
    if fallback:
        return str(fallback["code"])
    return code if code in CRM_CHANNELS else "khac"


def _crm_customer_staff_can_access(
    conn: sqlite3.Connection, customer_id: int, portal_sid: int
) -> bool:
    """Nhân viên chỉ xem KH có ít nhất một case đang mở được gán cho họ."""
    rows = conn.execute(
        """
        SELECT pipeline_stage, status FROM crm_cases
        WHERE customer_id = ? AND assigned_staff_id = ?
        """,
        (customer_id, portal_sid),
    ).fetchall()
    for r in rows:
        stage = normalize_pipeline_stage(r["pipeline_stage"] or r["status"])
        if stage not in TERMINAL_STAGES:
            return True
    return False


def _crm_customers_overview_rows(
    conn: sqlite3.Connection,
    *,
    portal_sid: int | None = None,
    q_raw: str = "",
    limit: int = 200,
    active_only: bool = False,
) -> list[dict[str, Any]]:
    """Danh sách khách hàng kèm thống kê chăm sóc / pipeline."""
    params: list[Any] = []
    where_parts: list[str] = []
    if q_raw:
        like = f"%{q_raw.lower()}%"
        where_parts.append(
            """(
            lower(coalesce(trim(cu.name), '')) LIKE ?
            OR lower(coalesce(trim(cu.phone), '')) LIKE ?
            OR lower(coalesce(trim(cu.email), '')) LIKE ?
            OR lower(coalesce(trim(cu.company), '')) LIKE ?
            OR lower(coalesce(trim(cu.address), '')) LIKE ?
            )"""
        )
        params.extend([like, like, like, like, like])

    if portal_sid is not None:
        where_parts.append(
            """EXISTS (
                SELECT 1 FROM crm_cases cx
                WHERE cx.customer_id = cu.id
                  AND cx.assigned_staff_id = ?
                  AND COALESCE(cx.pipeline_stage, 'moi') NOT IN ('chot', 'mat')
            )"""
        )
        params.append(portal_sid)

    where_sql = f"WHERE {' AND '.join(where_parts)}" if where_parts else ""
    sql = f"""
        SELECT
            cu.id,
            cu.name,
            cu.phone,
            cu.email,
            cu.address,
            cu.company,
            cu.lead_source,
            cu.created_at,
            (
                SELECT COUNT(*) FROM crm_customer_issues i
                WHERE i.customer_id = cu.id
                  AND i.status NOT IN ('da_xu_ly', 'dong')
            ) AS issues_open,
            (
                SELECT COUNT(*) FROM crm_cases c WHERE c.customer_id = cu.id
            ) AS cases_total,
            (
                SELECT COUNT(*) FROM crm_cases c
                WHERE c.customer_id = cu.id
                  AND COALESCE(c.pipeline_stage, 'moi') NOT IN ('chot', 'mat')
            ) AS cases_open,
            (
                SELECT MAX(c.updated_at) FROM crm_cases c WHERE c.customer_id = cu.id
            ) AS last_case_updated,
            (
                SELECT st.name FROM crm_cases c
                LEFT JOIN crm_staff st ON st.id = c.assigned_staff_id
                WHERE c.customer_id = cu.id
                ORDER BY datetime(c.updated_at) DESC
                LIMIT 1
            ) AS primary_staff_name,
            (
                SELECT c.pipeline_stage FROM crm_cases c
                WHERE c.customer_id = cu.id
                ORDER BY datetime(c.updated_at) DESC
                LIMIT 1
            ) AS primary_pipeline_stage,
            (
                SELECT r.care_status FROM crm_care_reports r
                JOIN crm_cases c ON c.id = r.case_id
                WHERE c.customer_id = cu.id
                ORDER BY r.id DESC
                LIMIT 1
            ) AS last_care_status,
            (
                SELECT r.created_at FROM crm_care_reports r
                JOIN crm_cases c ON c.id = r.case_id
                WHERE c.customer_id = cu.id
                ORDER BY r.id DESC
                LIMIT 1
            ) AS last_care_at,
            (
                SELECT COUNT(*) FROM crm_contracts ct WHERE ct.customer_id = cu.id
            ) AS contracts_total
        FROM crm_customers cu
        {where_sql}
        ORDER BY datetime(COALESCE(
            (SELECT MAX(c.updated_at) FROM crm_cases c WHERE c.customer_id = cu.id),
            cu.created_at
        )) DESC
        LIMIT ?
    """
    params.append(limit)
    rows = conn.execute(sql, params).fetchall()
    out: list[dict[str, Any]] = []
    for row in rows:
        d = dict(row)
        stage = normalize_pipeline_stage(d.get("primary_pipeline_stage"))
        d["primary_pipeline_stage"] = stage
        d["primary_pipeline_label"] = pipeline_stage_label(stage)
        ls = str(d.get("lead_source") or "")
        d["lead_source_label"] = CUSTOMER_LEAD_SOURCE_LABELS_VI.get(ls, ls) if ls else ""
        cs = str(d.get("last_care_status") or "")
        d["last_care_status_label"] = CRM_CARE_STATUS_LABELS_VI.get(cs, cs) if cs else ""
        if active_only and int(d.get("cases_open") or 0) == 0:
            continue
        out.append(d)
    return out


def _crm_row_case(row: sqlite3.Row, channel_labels: dict[str, str] | None = None) -> dict[str, Any]:
    d = dict(row)
    display_name = d.pop("staff_display_name", None)
    if display_name:
        d["assigned_to"] = display_name
    aid = d.get("assigned_staff_id")
    if aid is not None:
        d["assigned_staff_id"] = int(aid)
    else:
        d["assigned_staff_id"] = None
    d["status_label"] = CRM_STATUS_LABELS_VI.get(d["status"], d["status"])
    ch_labels = channel_labels if channel_labels is not None else CRM_CHANNEL_LABELS_VI
    d["channel_label"] = ch_labels.get(d["channel"], d["channel"])
    d["priority_label"] = CRM_PRIORITY_LABELS_VI.get(d["priority"], d["priority"])
    enrich_case_row(d)
    return d


def _crm_normalize_contract_status(raw: str | None) -> str:
    s = str(raw or "").strip().lower().replace(" ", "_").replace("-", "_")
    if not s:
        return "draft"
    return s if s in CRM_CONTRACT_STATUSES else "draft"


def _crm_normalize_campaign_channel(raw: str | None) -> str:
    s = str(raw or "other").strip().lower()
    return s if s in CRM_CAMPAIGN_CHANNELS else "other"


def _crm_normalize_reminder_scope(raw: str | None) -> str:
    s = str(raw or "general").strip().lower()
    return s if s in CRM_REMINDER_SCOPES else "general"


def _crm_normalize_reminder_kind(raw: str | None) -> str:
    s = str(raw or "manual").strip().lower()
    return s if s in CRM_REMINDER_KINDS else "manual"


def _crm_normalize_reminder_status(raw: str | None) -> str:
    s = str(raw or "pending").strip().lower()
    return s if s in CRM_REMINDER_STATUSES else "pending"


def _crm_hub_date_add_days_iso(date_str: str, delta_days: int) -> str | None:
    d = date_str.strip()[:10]
    if len(d) != 10 or not _crm_validate_date_ymd(d):
        return None
    dt = datetime.strptime(d, "%Y-%m-%d") + timedelta(days=delta_days)
    return dt.strftime("%Y-%m-%d")


def _crm_hub_sync_contract_renewal_reminder(conn: sqlite3.Connection, contract_id: int) -> None:
    row = conn.execute("SELECT * FROM crm_contracts WHERE id = ?", (contract_id,)).fetchone()
    if row is None:
        return
    d = dict(row)
    conn.execute(
        """
        DELETE FROM crm_reminders
        WHERE scope = 'contract' AND ref_id = ? AND reminder_kind = 'contract_renewal'
        """,
        (contract_id,),
    )
    ends = str(d.get("ends_on") or "").strip()[:10]
    try:
        days = int(d.get("renewal_reminder_days") or 0)
    except (TypeError, ValueError):
        days = 30
    if not ends or not _crm_validate_date_ymd(ends):
        return
    days = max(1, min(366, days or 30))
    remind_on = _crm_hub_date_add_days_iso(ends, -days)
    if not remind_on:
        return
    title_v = str(d.get("title") or "Hợp đồng").strip()[:240]
    ref_code = str(d.get("reference_code") or "").strip()
    suf = f" ({ref_code})" if ref_code else ""
    ts = _crm_ts()
    short_date = datetime.now().strftime("%Y-%m-%d")
    conn.execute(
        """
        INSERT INTO crm_reminders (
            scope, ref_id, reminder_kind, title, body, remind_at,
            status, staff_id, meta_json, created_at, updated_at
        )
        VALUES ('contract', ?, 'contract_renewal', ?, ?, ?, 'pending', NULL, '{}', ?, ?)
        """,
        (
            contract_id,
            f"[Hết hạn HĐ{suf}] {title_v}".strip(),
            f"Theo dõi hết hiệu lực ({ends}).",
            remind_on,
            short_date,
            ts,
        ),
    )


def _crm_lookup_campaign_by_code(conn: sqlite3.Connection, raw: Any) -> int | None:
    code = str(raw or "").strip()
    if not code:
        return None
    hit = conn.execute(
        """
        SELECT id FROM crm_campaigns
        WHERE trim(code) != '' AND lower(trim(code)) = lower(?)
        ORDER BY active DESC, id DESC LIMIT 1
        """,
        (code,),
    ).fetchone()
    return int(hit["id"]) if hit else None


def _crm_marketing_ingest_allowed() -> bool:
    tok = os.getenv("CRM_MARKETING_INGEST_SECRET") or ""
    return bool(tok.strip())


def _crm_marketing_verify_bearer() -> bool:
    exp = (os.getenv("CRM_MARKETING_INGEST_SECRET") or "").strip()
    if not exp:
        return False
    ah = request.headers.get("Authorization", "") or ""
    prefix = "Bearer "
    got = ah[len(prefix) :].strip() if ah.startswith(prefix) else ""
    if not got:
        return False
    try:
        return secrets.compare_digest(got.encode("utf-8"), exp.encode("utf-8"))
    except AttributeError:
        return secrets.compare_digest(got.encode(), exp.encode())


def _crm_facebook_sync_secret_ok() -> bool:
    ah = request.headers.get("Authorization", "") or ""
    if not ah.startswith("Bearer "):
        return False
    got = ah[7:].strip()
    if not got:
        return False
    for env_key in ("CRM_FACEBOOK_SYNC_SECRET", "CRM_MARKETING_INGEST_SECRET"):
        exp = (os.getenv(env_key) or "").strip()
        if not exp:
            continue
        try:
            if secrets.compare_digest(got.encode("utf-8"), exp.encode("utf-8")):
                return True
        except AttributeError:
            if secrets.compare_digest(got.encode(), exp.encode()):
                return True
    return False


def _crm_facebook_sync_cron_allowed() -> bool:
    if _crm_facebook_sync_secret_ok():
        return True
    if os.getenv("CRM_FACEBOOK_SYNC_ALLOW_LOCAL", "1").strip().lower() in ("0", "false", "no", "off"):
        return False
    remote = (request.remote_addr or "").strip().lower()
    if remote in ("127.0.0.1", "::1", "localhost"):
        return True
    if remote.startswith("::ffff:127.0.0.1"):
        return True
    # Gunicorn / proxy nội bộ đôi khi để trống remote_addr
    host = (request.host or "").split(":")[0].strip().lower()
    if host in ("127.0.0.1", "localhost"):
        return True
    return False


def _crm_finance_kpi_cron_secret_ok() -> bool:
    auth = request.headers.get("Authorization") or ""
    if auth.lower().startswith("bearer "):
        got = auth[7:].strip()
        for env_key in (
            "CRM_FINANCE_KPI_CRON_SECRET",
            "CRM_FACEBOOK_SYNC_SECRET",
            "CRM_MARKETING_INGEST_SECRET",
        ):
            exp = (os.getenv(env_key) or "").strip()
            if exp and secrets.compare_digest(got.encode(), exp.encode()):
                return True
    return False


def _crm_finance_kpi_cron_allowed() -> bool:
    if _crm_finance_kpi_cron_secret_ok():
        return True
    if os.getenv("CRM_FINANCE_KPI_CRON_ALLOW_LOCAL", "1").strip().lower() in (
        "0",
        "false",
        "no",
        "off",
    ):
        return False
    remote = (request.remote_addr or "").strip().lower()
    if remote in ("127.0.0.1", "::1", "localhost"):
        return True
    if remote.startswith("::ffff:127.0.0.1"):
        return True
    host = (request.host or "").split(":")[0].strip().lower()
    if host in ("127.0.0.1", "localhost"):
        return True
    return False


def _facebook_sync_json_response(result: dict[str, Any]) -> Any:
    """JSON cho sync Facebook — luôn có error/message rõ ràng cho UI."""
    payload = dict(result)
    msg = str(payload.get("message") or "").strip()
    sync_msg = str((payload.get("sync") or {}).get("message") or "").strip()
    detail = msg or sync_msg
    if not detail and payload.get("graph_errors"):
        detail = str(payload["graph_errors"][0])
    if detail:
        payload.setdefault("message", detail)
        payload.setdefault("error", detail)
    if payload.get("ok"):
        code = 200
    elif payload.get("rate_limited"):
        code = 200
    else:
        code = 400
    return jsonify(payload), code


def _crm_find_existing_customer(conn: sqlite3.Connection, phone: str, email: str) -> int | None:
    p = "".join(ch for ch in (phone or "").strip() if ch.isdigit() or ch == "+")
    if len(p) >= 8:
        hit = conn.execute(
            """
            SELECT id FROM crm_customers
            WHERE TRIM(REPLACE(REPLACE(REPLACE(COALESCE(phone,''),' ',''),'-',''),'.',''))
              = TRIM(REPLACE(REPLACE(REPLACE(?,' ',''),'-',''),'.',''))
            ORDER BY id ASC LIMIT 1
            """,
            (phone.strip(),),
        ).fetchone()
        if hit:
            return int(hit["id"])
    em = str(email or "").strip().lower()
    if em and "@" in em:
        hit = conn.execute(
            """
            SELECT id FROM crm_customers WHERE lower(trim(email)) = ?
            ORDER BY id ASC LIMIT 1
            """,
            (em,),
        ).fetchone()
        if hit:
            return int(hit["id"])
    return None


def _crm_resolve_assignment(
    conn: sqlite3.Connection, payload: dict[str, Any], *, allow_legacy_text: bool
) -> tuple[str, int | None]:
    """Trả (assigned_to hiển thị, assigned_staff_id)."""
    if "assigned_staff_id" in payload:
        raw = payload.get("assigned_staff_id")
        if raw is None or raw == "":
            return "", None
        if raw in (0, "0"):
            return "", None
        try:
            sid = int(raw)
        except (TypeError, ValueError):
            raise ValueError("assigned_staff_id không hợp lệ") from None
        if sid <= 0:
            return "", None
        st = conn.execute(
            "SELECT name FROM crm_staff WHERE id = ? AND active = 1",
            (sid,),
        ).fetchone()
        if st is None:
            raise ValueError("Nhân viên phụ trách không tồn tại hoặc đã ngưng hoạt động")
        return (str(st["name"]), sid)
    if allow_legacy_text and "assigned_to" in payload and isinstance(payload["assigned_to"], str):
        return (payload["assigned_to"].strip()[:200], None)
    return ("", None)


def _staff_field_exists(
    conn: sqlite3.Connection,
    field: str,
    value: str,
    *,
    exclude_id: int | None,
) -> bool:
    """Trùng email hoặc mã nhân viên nội bộ (khi có giá trị)."""
    if field not in ("email", "internal_code", "attendance_pin"):
        return False
    column = {"email": "email", "internal_code": "internal_code", "attendance_pin": "attendance_pin"}[
        field
    ]
    v = (value or "").strip()
    if not v:
        return False
    q = f"SELECT 1 FROM crm_staff WHERE lower(trim({column})) = lower(?) AND trim({column}) != ''"
    p: list[Any] = [v]
    if exclude_id is not None:
        q += " AND id != ?"
        p.append(exclude_id)
    return conn.execute(q, p).fetchone() is not None


def _crm_append_event(conn: sqlite3.Connection, case_id: int, kind: str, body: str) -> None:
    conn.execute(
        """
        INSERT INTO crm_case_events (case_id, kind, body, created_at)
        VALUES (?, ?, ?, ?)
        """,
        (case_id, kind, body, _crm_ts()),
    )


def _crm_assign_label(
    conn: sqlite3.Connection, assigned_to: Any, assigned_staff_id: Any
) -> str:
    """Nhãn hiển thị người phụ trách (ưu tiên tên từ crm_staff)."""
    raw_id = assigned_staff_id
    try:
        sid = int(raw_id) if raw_id is not None else None
    except (TypeError, ValueError):
        sid = None
    if sid:
        r = conn.execute("SELECT name FROM crm_staff WHERE id = ?", (sid,)).fetchone()
        if r:
            return str(r["name"])
    t = str(assigned_to or "").strip()
    return t if t else "—"


def _opt_pos_int(val: Any) -> int | None:
    if val is None or val == "":
        return None
    try:
        n = int(val)
    except (TypeError, ValueError):
        return None
    return n if n > 0 else None


def _dept_exists_active(conn: sqlite3.Connection, dept_id: int) -> bool:
    r = conn.execute(
        "SELECT id FROM crm_departments WHERE id = ? AND active = 1",
        (dept_id,),
    ).fetchone()
    return r is not None


def _position_exists_active(conn: sqlite3.Connection, position_id: int) -> bool:
    r = conn.execute(
        "SELECT id FROM crm_positions WHERE id = ? AND active = 1",
        (position_id,),
    ).fetchone()
    return r is not None


def fetch_settings() -> dict[str, str]:
    with get_connection() as conn:
        rows = conn.execute("SELECT key, value FROM settings").fetchall()
    return {row["key"]: row["value"] for row in rows}


def fetch_service_categories() -> list[dict[str, Any]]:
    settings = fetch_settings()
    raw_value = settings.get("service_categories_json")
    if not raw_value:
        return deepcopy(DEFAULT_SERVICE_CATEGORIES)

    try:
        parsed = json.loads(raw_value)
        if isinstance(parsed, list):
            return parsed
    except json.JSONDecodeError:
        pass
    return deepcopy(DEFAULT_SERVICE_CATEGORIES)


def build_service_details(categories: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    details: dict[str, dict[str, Any]] = {}
    for category in categories:
        title = str(category.get("title", "")).strip()
        for item in category.get("items", []):
            if not isinstance(item, dict):
                continue
            slug = str(item.get("slug", "")).strip()
            if not slug:
                continue
            name = str(item.get("name", "")).strip()
            sm = str(item.get("summary", "")).strip() or (
                f"Giải pháp {name} giúp doanh nghiệp tối ưu hiệu quả truyền thông."
                if name
                else "Giải pháp digital marketing giúp doanh nghiệp tối ưu truyền thông."
            )
            h_raw = item.get("highlights", [])
            highlights: list[str] = (
                [str(x).strip() for x in h_raw if str(x).strip()] if isinstance(h_raw, list) else []
            )
            base_landing = _finalize_landing(_merge_landing_extras(item, slug), sm)
            details[slug] = {
                **base_landing,
                "slug": slug,
                "name": name,
                "category": title,
                "summary": sm,
                "highlights": highlights,
            }
    return details


def sibling_services_in_category(
    categories: list[dict[str, Any]], current_slug: str
) -> list[dict[str, str]]:
    """Các dịch vụ khác cùng nhóm (submenu) — trang chi tiết kiểu BNG: sidebar liên quan."""
    for category in categories:
        items = category.get("items")
        if not isinstance(items, list):
            continue
        slugs = {str(i.get("slug", "")).strip() for i in items if isinstance(i, dict)}
        if current_slug not in slugs:
            continue
        out: list[dict[str, str]] = []
        for i in items:
            if not isinstance(i, dict):
                continue
            s = str(i.get("slug", "")).strip()
            n = str(i.get("name", "")).strip()
            if s and s != current_slug and n:
                out.append({"slug": s, "name": n})
        return out
    return []


def validate_service_payload(payload: Any) -> tuple[bool, str]:
    if not isinstance(payload, list):
        return False, "Payload must be a JSON array of service categories."

    seen_slugs: set[str] = set()
    for category in payload:
        if not isinstance(category, dict):
            return False, "Each category must be an object."
        if not str(category.get("title", "")).strip():
            return False, "Each category must have a non-empty title."
        items = category.get("items")
        if not isinstance(items, list):
            return False, "Each category must include an items array."
        for item in items:
            if not isinstance(item, dict):
                return False, "Each service item must be an object."
            slug = str(item.get("slug", "")).strip()
            name = str(item.get("name", "")).strip()
            if not slug or not name:
                return False, "Each service item must have non-empty slug and name."
            if slug in seen_slugs:
                return False, f"Duplicate slug found: {slug}"
            seen_slugs.add(slug)
            highlights = item.get("highlights", [])
            if highlights is not None and not isinstance(highlights, list):
                return False, f"highlights must be an array for slug: {slug}"
            for opt in ("stats", "pillars", "outcomes", "faq", "overview"):
                if opt not in item:
                    continue
                val = item[opt]
                if val is not None and not isinstance(val, (list, str)):
                    return False, f"{opt} must be a string or array for slug: {slug}"
            if "stats" in item and item["stats"] is not None and isinstance(item["stats"], list):
                for st in item["stats"]:
                    if st is not None and not isinstance(st, dict):
                        return False, f"stats items must be objects for slug: {slug}"
            if "pillars" in item and item["pillars"] is not None and isinstance(item["pillars"], list):
                for p in item["pillars"]:
                    if p is not None and not isinstance(p, dict):
                        return False, f"pillars items must be objects for slug: {slug}"
            if "faq" in item and item["faq"] is not None and isinstance(item["faq"], list):
                for fa in item["faq"]:
                    if fa is not None and not isinstance(fa, dict):
                        return False, f"faq items must be objects for slug: {slug}"

    return True, ""


def rows_to_dict(rows: list[sqlite3.Row]) -> list[dict[str, Any]]:
    return [dict(row) for row in rows]


def get_or_create_consult_sheet():
    if EXCEL_PATH.exists():
        workbook = load_workbook(EXCEL_PATH)
    else:
        workbook = Workbook()

    if CONSULT_SHEET_NAME in workbook.sheetnames:
        sheet = workbook[CONSULT_SHEET_NAME]
    else:
        sheet = workbook.create_sheet(CONSULT_SHEET_NAME)
        sheet.append(
            [
                "submitted_at",
                "service_slug",
                "service_name",
                "full_name",
                "email",
                "phone",
                "budget",
                "company",
                "goal",
                "help_request",
                "additional_info",
            ]
        )
    return workbook, sheet


def append_consultation_to_excel(payload: dict[str, Any]) -> None:
    workbook, sheet = get_or_create_consult_sheet()
    sheet.append(
        [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            payload["service_slug"],
            payload["service_name"],
            payload["full_name"],
            payload["email"],
            payload["phone"],
            payload["budget"],
            payload["company"],
            payload["goal"],
            payload["help_request"],
            payload.get("additional_info", ""),
        ]
    )
    workbook.save(EXCEL_PATH)


def send_consultation_email(payload: dict[str, Any]) -> None:
    smtp_host = os.getenv("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_username = os.getenv("SMTP_USERNAME", "").strip()
    smtp_password = os.getenv("SMTP_PASSWORD", "").strip()
    sender_email = os.getenv("SMTP_FROM", smtp_username).strip()

    if not smtp_username or not smtp_password or not sender_email:
        raise RuntimeError(
            "Missing SMTP configuration. Please set SMTP_USERNAME, SMTP_PASSWORD, SMTP_FROM in .env."
        )

    body = (
        "Thong tin tu van moi:\n\n"
        f"Thoi gian: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        f"Giai phap: {payload['service_name']} ({payload['service_slug']})\n"
        f"Ho va ten: {payload['full_name']}\n"
        f"Email: {payload['email']}\n"
        f"So dien thoai: {payload['phone']}\n"
        f"Ngan sach: {payload['budget']}\n"
        f"Cong ty: {payload['company']}\n"
        f"Muc tieu: {payload['goal']}\n"
        f"Can ho tro: {payload['help_request']}\n"
        f"Thong tin khac: {payload.get('additional_info', '')}\n"
    )

    message = EmailMessage()
    message["Subject"] = f"[Tu van moi] {payload['service_name']} - {payload['full_name']}"
    message["From"] = sender_email
    message["To"] = CONSULT_EMAIL_TO
    message.set_content(body)

    with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as server:
        server.starttls()
        server.login(smtp_username, smtp_password)
        server.send_message(message)


def _cv_mime_parts(filename: str) -> tuple[str, str]:
    ext = Path(filename).suffix.lower()
    if ext == ".pdf":
        return "application", "pdf"
    if ext == ".doc":
        return "application", "msword"
    if ext == ".docx":
        return (
            "application",
            "vnd.openxmlformats-officedocument.wordprocessingml.document",
        )
    return "application", "octet-stream"


def send_job_application_email(
    job: dict[str, Any],
    full_name: str,
    email: str,
    phone: str,
    cover_letter: str,
    cv_original_name: str,
    cv_data: bytes,
) -> None:
    smtp_host = os.getenv("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_username = os.getenv("SMTP_USERNAME", "").strip()
    smtp_password = os.getenv("SMTP_PASSWORD", "").strip()
    sender_email = os.getenv("SMTP_FROM", smtp_username).strip()

    if not smtp_username or not smtp_password or not sender_email:
        raise RuntimeError(
            "Chưa cấu hình gửi email (SMTP). Vui lòng liên hệ quản trị hệ thống."
        )

    safe_attachment = secure_filename(cv_original_name)
    if not safe_attachment:
        ext = Path(cv_original_name).suffix.lower() or ".pdf"
        safe_attachment = f"cv{ext}"

    maintype, subtype = _cv_mime_parts(cv_original_name)

    body = (
        "Hồ sơ ứng tuyển\n\n"
        f"Thời gian gửi: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        f"Vị trí: {job['title']}\n"
        f"Mã vị trí (slug): {job['slug']}\n"
        f"Địa điểm (theo JD): {job['location']} · {job['employment_type']}\n\n"
        f"Họ và tên: {full_name}\n"
        f"Email liên hệ: {email}\n"
        f"Số điện thoại: {phone}\n\n"
        "Giới thiệu ngắn / thư xin việc:\n"
        f"{cover_letter or '(Không có)'}\n"
    )

    message = EmailMessage()
    message["Subject"] = f"[Ứng tuyển] {job['title']} — {full_name}"
    message["From"] = sender_email
    message["To"] = RECRUITMENT_EMAIL
    message["Reply-To"] = email
    message.set_content(body)
    message.add_attachment(
        cv_data,
        maintype=maintype,
        subtype=subtype,
        filename=safe_attachment,
    )

    with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as server:
        server.starttls()
        server.login(smtp_username, smtp_password)
        server.send_message(message)


def _career_request_wants_json() -> bool:
    return request.headers.get("X-Requested-With") == "XMLHttpRequest"


@app.errorhandler(RequestEntityTooLarge)
def handle_request_entity_too_large(_e: RequestEntityTooLarge):
    if request.path.startswith("/api/"):
        max_mb = MAX_VIDEO_UPLOAD_BYTES // (1024 * 1024)
        return jsonify({"error": f"File quá lớn (tối đa {max_mb} MB)"}), 413
    msg = "Tệp CV quá lớn (tối đa 5MB). Vui lòng chọn bản nhẹ hơn."
    if _career_request_wants_json():
        return jsonify({"ok": False, "error": msg}), 413
    flash(msg, "error")
    ref = request.referrer
    if ref:
        return redirect(ref)
    return redirect(url_for("career_list"))


@app.get("/")
def landing() -> str:
    with get_connection() as conn:
        projects = rows_to_dict(
            conn.execute("SELECT * FROM projects ORDER BY id DESC").fetchall()
        )
        news = rows_to_dict(
            conn.execute("SELECT * FROM news ORDER BY id DESC LIMIT 6").fetchall()
        )

    settings = _settings_for_public_pages()
    service_categories = fetch_service_categories()
    raw_slides = settings.get("hero_slides_json", "")
    try:
        hero_slides = json.loads(raw_slides) if raw_slides else DEFAULT_HERO_SLIDES
        if not isinstance(hero_slides, list) or not hero_slides:
            hero_slides = DEFAULT_HERO_SLIDES
    except (json.JSONDecodeError, TypeError):
        hero_slides = DEFAULT_HERO_SLIDES
    from cms_media_images import normalize_hero_slides

    hero_slides = normalize_hero_slides(hero_slides, UPLOAD_DIR)
    return render_template(
        "landing.html",
        settings=settings,
        contact_tel_href=_tel_href_from_display(settings.get("contact_phone", "")),
        projects=projects,
        news=news,
        hero_slides=hero_slides,
        service_categories=service_categories,
        partner_logos=_partner_logos_for_landing(settings),
        capabilities_items=_capabilities_items_for_landing(settings),
        recruitment_positions=get_recruitment_jobs(active_only=True),
    )


@app.get("/du-an/<int:project_id>")
def project_detail(project_id: int) -> str:
    """Trang chi tiết dự án (layout kiểu Saga Marketing work detail)."""
    with get_connection() as conn:
        row = conn.execute(
            "SELECT * FROM projects WHERE id = ?", (project_id,)
        ).fetchone()
        if row is None:
            abort(404)
        project = rows_to_dict([row])[0]
        other_projects = rows_to_dict(
            conn.execute(
                "SELECT * FROM projects WHERE id != ? ORDER BY id DESC LIMIT 12",
                (project_id,),
            ).fetchall()
        )

    settings = _settings_for_public_pages()
    return render_template(
        "project_detail.html",
        settings=settings,
        contact_tel_href=_tel_href_from_display(settings.get("contact_phone", "")),
        project=project,
        other_projects=other_projects,
    )


@app.get("/tin-tuc/<int:news_id>")
def news_detail(news_id: int) -> str:
    """Trang chi tiết tin tức (entry từ landing)."""
    with get_connection() as conn:
        row = conn.execute(
            "SELECT * FROM news WHERE id = ?", (news_id,)
        ).fetchone()
        if row is None:
            abort(404)
        item = rows_to_dict([row])[0]
        related_news = rows_to_dict(
            conn.execute(
                """
                SELECT * FROM news
                WHERE id != ?
                ORDER BY published_at DESC, id DESC
                LIMIT 6
                """,
                (news_id,),
            ).fetchall()
        )

    settings = _settings_for_public_pages()
    return render_template(
        "news_detail.html",
        settings=settings,
        contact_tel_href=_tel_href_from_display(settings.get("contact_phone", "")),
        item=item,
        related_news=related_news,
    )


@app.get("/services/<slug>")
def service_detail(slug: str) -> str:
    service_details = build_service_details(fetch_service_categories())
    service = service_details.get(slug)
    if service is None:
        abort(404)

    settings = _settings_for_public_pages()
    categories = fetch_service_categories()
    return render_template(
        "service_detail.html",
        settings=settings,
        contact_tel_href=_tel_href_from_display(settings.get("contact_phone", "")),
        service=service,
        sibling_services=sibling_services_in_category(categories, slug),
    )


@app.get("/career")
def career_list() -> str:
    settings = _settings_for_public_pages()
    return render_template(
        "career_list.html",
        settings=settings,
        contact_tel_href=_tel_href_from_display(settings.get("contact_phone", "")),
        jobs=get_recruitment_jobs(active_only=True),
    )


@app.get("/chinh-sach-bao-mat")
def privacy_policy() -> str:
    """Chính sách bảo mật — trang công khai cho Meta App Review và khách hàng."""
    settings = _settings_for_public_pages()
    brand = settings.get("brand_name") or "PTT Advertising Solutions"
    return render_template(
        "privacy_policy.html",
        settings=settings,
        brand_name=brand,
        contact_email=settings.get("contact_email") or "contact@pttadvertising.vn",
        contact_phone=settings.get("contact_phone") or "",
        contact_tel_href=_tel_href_from_display(settings.get("contact_phone", "")),
        site_url=(request.url_root or "https://pttads.vn/").rstrip("/"),
        site_host=(request.host or "pttads.vn").split(":")[0],
        legal_years=settings.get("footer_legal_years") or "2020 – 2026",
        updated_at="10/06/2026",
    )


@app.get("/career/<slug>")
def career_detail(slug: str):
    if get_recruitment_job(slug) is None:
        abort(404)
    return redirect(url_for("career_list", job=slug))


@app.post("/career/<slug>/apply")
def career_apply(slug: str):
    job = get_recruitment_job(slug)
    if job is None:
        if _career_request_wants_json():
            return jsonify({"ok": False, "error": "Không tìm thấy vị trí."}), 404
        abort(404)

    def respond_error(msg: str, code: int = 400):
        if _career_request_wants_json():
            return jsonify({"ok": False, "error": msg}), code
        flash(msg, "error")
        return redirect(url_for("career_list", job=slug))

    def respond_success(msg: str):
        if _career_request_wants_json():
            return jsonify({"ok": True, "message": msg})
        flash(msg, "success")
        return redirect(url_for("career_list", job=slug))

    full_name = (request.form.get("full_name") or "").strip()
    email = (request.form.get("email") or "").strip()
    phone = (request.form.get("phone") or "").strip()
    cover_letter = (request.form.get("cover_letter") or "").strip()

    if not full_name or not email or not phone:
        return respond_error(
            "Vui lòng điền đầy đủ họ tên, email và số điện thoại.",
        )

    if not _EMAIL_RE.match(email):
        return respond_error("Địa chỉ email không hợp lệ.")

    file = request.files.get("cv")
    if file is None or not (file.filename or "").strip():
        return respond_error("Vui lòng đính kèm file CV (PDF, DOC hoặc DOCX).")

    raw_name = file.filename or "cv.pdf"
    ext = Path(raw_name).suffix.lower()
    if ext not in ALLOWED_CV_EXTENSIONS:
        return respond_error("Chỉ chấp nhận CV dạng PDF, DOC hoặc DOCX.")

    cv_data = file.read()
    if not cv_data:
        return respond_error("File CV rỗng hoặc không đọc được.")
    if len(cv_data) > MAX_CV_UPLOAD_BYTES:
        return respond_error("File CV quá lớn (tối đa 5 MB). Vui lòng chọn bản nhẹ hơn.")

    try:
        send_job_application_email(
            job, full_name, email, phone, cover_letter, raw_name, cv_data
        )
    except RuntimeError as exc:
        return respond_error(str(exc))
    except Exception:
        return respond_error(
            "Không gửi được hồ sơ. Vui lòng thử lại sau hoặc gửi CV qua email.",
            500,
        )

    return respond_success(
        "Đã gửi hồ sơ thành công. Bộ phận nhân sự sẽ liên hệ khi phù hợp.",
    )


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    """Đăng nhập quản trị (Admin/CMS) hoặc nhân viên CSKH."""
    if _admin_logged_in():
        raw_next = (
            request.args.get("next") if request.method == "GET" else request.form.get("next")
        )
        return redirect(_safe_internal_redirect_path(raw_next, default_endpoint="admin"))
    if _staff_logged_in():
        return redirect(url_for("crm_staff_home"))

    err_msg: str | None = None
    eu, epw = _admin_expected_credentials()
    if request.method == "POST":
        u_raw = request.form.get("username", "")
        p_raw = request.form.get("password", "")
        nxt = request.form.get("next")
        ts = _crm_ts()
        with get_connection() as conn:
            hit = unified_login(
                conn,
                u_raw.strip(),
                p_raw,
                env_username=eu,
                env_password=epw,
                const_eq=_const_eq_str,
            )
            if hit:
                if hit.kind == "admin":
                    ensure_unified_password_stored(
                        conn, hit.username, p_raw, updated_at=ts
                    )
                    _start_admin_session(
                        hit.username,
                        hit.role_code or "super_admin",
                        position_id=hit.position_id,
                    )
                    return redirect(
                        _safe_internal_redirect_path(nxt, default_endpoint="admin")
                    )
                session.clear()
                session[STAFF_SESSION_KEY] = True
                session[STAFF_ID_SESSION_KEY] = int(hit.staff_id or 0)
                session[STAFF_NAME_SESSION_KEY] = str(hit.staff_name or "")
                session.permanent = True
                return redirect(url_for("crm_staff_home"))
        err_msg = "Sai tên đăng nhập hoặc mật khẩu."

    return render_template(
        "admin_login.html",
        error=err_msg,
        next_val=request.args.get("next", "") or "",
    )


@app.post("/admin/logout")
def admin_logout():
    session.clear()
    return redirect(url_for("admin_login"))


@app.get("/account/password")
def account_password_page() -> str:
    if not _internal_logged_in():
        return redirect(url_for("admin_login", next="/account/password"))
    staff_portal = _crm_staff_portal_active()
    extra = _admin_page_template_kwargs() if not staff_portal else {}
    return render_template(
        "account_password.html",
        crm_staff_portal=staff_portal,
        admin_staff_portal=staff_portal,
        crm_staff_name=_staff_session_name() if staff_portal else "",
        account_username=_cms_session_username()
        if _admin_logged_in()
        else _staff_session_name(),
        **extra,
    )


@app.post("/api/account/change-password")
def api_account_change_password() -> Any:
    if not _internal_logged_in():
        return jsonify({"error": "Chưa đăng nhập.", "login": url_for("admin_login")}), 401
    body = request.get_json(force=True) or {}
    current_pw = str(body.get("current_password") or "")
    new_pw = str(body.get("new_password") or "")
    confirm_pw = str(body.get("new_password_confirm") or body.get("confirm_password") or "")
    if not current_pw or not new_pw:
        return jsonify({"error": "Cần mật khẩu hiện tại và mật khẩu mới."}), 400
    if new_pw != confirm_pw:
        return jsonify({"error": "Mật khẩu mới và xác nhận không khớp."}), 400
    if current_pw == new_pw:
        return jsonify({"error": "Mật khẩu mới phải khác mật khẩu hiện tại."}), 400
    ts = _crm_ts()
    eu, epw = _admin_expected_credentials()

    if _staff_logged_in() and not _admin_logged_in():
        sid = _staff_session_id()
        if sid is None:
            return jsonify({"error": "Phiên đăng nhập không hợp lệ."}), 401
        with get_connection() as conn:
            row = conn.execute(
                """
                SELECT id, login_enabled, login_username
                FROM crm_staff WHERE id = ?
                """,
                (sid,),
            ).fetchone()
            if row is None:
                return jsonify({"error": "Không tìm thấy tài khoản nhân viên."}), 404
            if not int(row["login_enabled"] or 0):
                return jsonify({"error": "Tài khoản đăng nhập chưa được bật."}), 403
            uname = str(row["login_username"] or "").strip()
            if not uname:
                return jsonify({"error": "Chưa có tên đăng nhập."}), 400
            if not verify_unified_password(
                conn,
                uname,
                current_pw,
                env_username=eu,
                env_password=epw,
                const_eq=_const_eq_str,
            ):
                return jsonify({"error": "Mật khẩu hiện tại không đúng."}), 400
            try:
                set_unified_password(conn, uname, new_pw.strip(), updated_at=ts)
            except ValueError as exc:
                return jsonify({"error": str(exc)}), 400
            conn.commit()
        return jsonify({"ok": True, "message": "Đã đổi mật khẩu thành công."})

    if not _admin_logged_in():
        return jsonify({"error": "Không có quyền."}), 403

    uname = _cms_session_username()
    if not uname:
        eu_name, _ = _admin_expected_credentials()
        uname = eu_name
    with get_connection() as conn:
        if not verify_unified_password(
            conn,
            uname,
            current_pw,
            env_username=eu,
            env_password=epw,
            const_eq=_const_eq_str,
        ):
            return jsonify({"error": "Mật khẩu hiện tại không đúng."}), 400
        row = conn.execute(
            """
            SELECT id FROM cms_admin_users
            WHERE lower(trim(username)) = lower(trim(?)) AND active = 1
            """,
            (uname,),
        ).fetchone()
        try:
            new_hash = set_unified_password(conn, uname, new_pw.strip(), updated_at=ts)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        if row is None:
            conn.execute(
                """
                INSERT INTO cms_admin_users (
                    username, display_name, role_code, password_hash, active, created_at, updated_at
                )
                VALUES (?, ?, 'super_admin', ?, 1, ?, ?)
                """,
                (uname, uname, new_hash, ts[:10], ts),
            )
        conn.commit()
    return jsonify({"ok": True, "message": "Đã đổi mật khẩu thành công."})


def _admin_page_template_kwargs() -> dict[str, Any]:
    if _staff_logged_in() and not _admin_logged_in():
        return {"admin_grants_json": "{}", "admin_full_access": False, "crm_leads_only_ui": False}
    with get_connection() as conn:
        return _admin_grants_template_kwargs(conn)


_LEADS_ONLY_ADMIN_PATHS = (
    "/crm/leads",
    "/crm/hdsd",
    "/crm/test-cases",
    "/api/crm/leads",
    "/api/crm/hdsd",
    "/admin/logout",
    "/account/password",
    "/api/account/change-password",
)


def _crm_leads_only_path_allowed(path: str) -> bool:
    p = path or ""
    if p.startswith("/static/"):
        return True
    return any(p == prefix or p.startswith(prefix + "/") for prefix in _LEADS_ONLY_ADMIN_PATHS)


@app.get("/admin")
def admin() -> str:
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    if _crm_leads_only_ui():
        return redirect(url_for("crm_leads_page"))
    return render_template("admin.html", **_admin_page_template_kwargs())


@app.get("/crm/home")
def crm_staff_home() -> str:
    """Trang chủ portal nhân viên — KPI, Lead, báo cáo."""
    redir = _ensure_crm_session_html()
    if redir is not None:
        return redir
    if not _crm_staff_portal_active():
        return redirect(url_for("crm_board"))
    return render_template(
        "crm_staff_home.html",
        crm_staff_portal=True,
        crm_staff_id=_staff_session_id(),
        crm_staff_name=_staff_session_name(),
        lead_status_labels=LEAD_STATUS_LABELS,
        lead_level_labels=LEAD_LEVEL_LABELS,
        lead_source_labels=LEAD_SOURCE_LABELS,
        kpi_status_labels=CRM_KPI_STATUS_LABELS_VI,
    )


@app.get("/api/crm/staff/dashboard")
def api_crm_staff_dashboard() -> Any:
    """Tổng quan portal nhân viên — KPI, Lead phân công, báo cáo ngày."""
    if not _crm_staff_portal_active():
        return jsonify({"error": "Chỉ dành cho portal nhân viên."}), 403
    sid = _staff_session_id()
    if sid is None:
        return jsonify({"error": "Chưa đăng nhập."}), 401
    from crm_staff_dashboard import build_staff_dashboard

    with get_connection() as conn:
        payload = build_staff_dashboard(conn, int(sid), ts=_crm_ts())
    return jsonify(payload)


@app.get("/crm")
def crm_board() -> str:
    redir = _ensure_crm_session_html()
    if redir is not None:
        return redir
    staff_portal = _crm_staff_portal_active()
    return render_template(
        "crm.html",
        crm_statuses=CRM_STATUSES_ORDER,
        crm_status_labels=CRM_STATUS_LABELS_VI,
        crm_channels=CRM_CHANNELS,
        crm_channel_labels=CRM_CHANNEL_LABELS_VI,
        crm_priorities=CRM_PRIORITIES,
        crm_priority_labels=CRM_PRIORITY_LABELS_VI,
        crm_pipeline_stages=SALES_PIPELINE_STAGES,
        crm_pipeline_labels=SALES_PIPELINE_LABELS_VI,
        crm_workflow_playbook=get_crm_workflow_playbook(),
        crm_lead_intake_flow=get_crm_lead_intake_master_flow(),
        crm_marketing_channels=get_crm_marketing_ingress_channels(),
        crm_care_contact_types=CRM_CARE_CONTACT_TYPES,
        crm_care_contact_labels=CRM_CARE_CONTACT_LABELS_VI,
        crm_care_status_types=CRM_CARE_STATUS_TYPES,
        crm_care_status_labels=CRM_CARE_STATUS_LABELS_VI,
        crm_staff_portal=staff_portal,
        crm_staff_id=_staff_session_id() if staff_portal else None,
        crm_staff_name=_staff_session_name() if staff_portal else "",
        **_admin_page_template_kwargs(),
    )


@app.get("/crm/customers")
def crm_customers_page() -> str:
    redir = _ensure_crm_session_html()
    if redir is not None:
        return redir
    staff_portal = _crm_staff_portal_active()
    return render_template(
        "crm_customers.html",
        crm_pipeline_labels=SALES_PIPELINE_LABELS_VI,
        crm_care_status_labels=CRM_CARE_STATUS_LABELS_VI,
        crm_care_contact_labels=CRM_CARE_CONTACT_LABELS_VI,
        crm_contract_status_labels=CRM_CONTRACT_STATUS_LABELS_VI,
        crm_customer_lead_sources=CUSTOMER_LEAD_SOURCE_LABELS_VI,
        crm_customer_genders=CUSTOMER_GENDER_LABELS_VI,
        crm_relation_types=RELATION_TYPE_LABELS_VI,
        crm_purchase_statuses=PURCHASE_STATUS_LABELS_VI,
        crm_issue_types=ISSUE_TYPE_LABELS_VI,
        crm_issue_statuses=ISSUE_STATUS_LABELS_VI,
        crm_issue_priorities=ISSUE_PRIORITY_LABELS_VI,
        crm_staff_portal=staff_portal,
        crm_staff_id=_staff_session_id() if staff_portal else None,
        crm_staff_name=_staff_session_name() if staff_portal else "",
        **_admin_page_template_kwargs(),
    )


@app.get("/crm/customer/<int:customer_id>/meeting-brief")
def crm_customer_meeting_brief_page(customer_id: int) -> str:
    redir = _ensure_crm_session_html()
    if redir is not None:
        return redir
    with get_connection() as conn:
        from crm_customer_brief import get_customer_snapshot as _snap, get_latest_brief as _latest
        customer = conn.execute(
            "SELECT id, name, company FROM crm_customers WHERE id = ?",
            (customer_id,),
        ).fetchone()
        if customer is None:
            return "Không tìm thấy khách hàng", 404
        latest = _latest(conn, customer_id)
    return render_template(
        "crm_customer_meeting_brief.html",
        customer=dict(customer),
        latest_brief=latest,
        **_admin_page_template_kwargs(),
    )


@app.post("/api/crm/customers/<int:customer_id>/brief/generate")
def api_crm_customer_brief_generate(customer_id: int):
    from flask import request as req
    with get_connection() as conn:
        from crm_customer_brief import get_customer_snapshot as _snap, run_brief_ai as _run
        body = req.get_json(silent=True) or {}
        meeting_purpose = str(body.get("meeting_purpose", ""))[:500]
        snapshot = _snap(conn, customer_id)
        output = _run(conn, customer_id, meeting_purpose, snapshot)
    from datetime import datetime
    return jsonify({"ai_output": output, "created_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")})


@app.get("/api/crm/customers/<int:customer_id>/brief/latest")
def api_crm_customer_brief_latest(customer_id: int):
    with get_connection() as conn:
        from crm_customer_brief import get_latest_brief as _latest
        brief = _latest(conn, customer_id)
    return jsonify(brief or {})


@app.get("/crm/aeo")
def crm_aeo_page() -> Any:
    redir = _ensure_crm_session_html()
    if redir is not None:
        return redir
    from crm_aeo import list_queries as _aeo_list
    customer_id = _opt_pos_int(request.args.get("customer_id"))
    with get_connection() as conn:
        all_customers = [
            dict(r) for r in conn.execute(
                "SELECT id, name, company FROM crm_customers ORDER BY name"
            ).fetchall()
        ]
        queries = []
        selected_customer = None
        if customer_id:
            row = conn.execute(
                "SELECT id, name, company FROM crm_customers WHERE id = ?",
                (customer_id,),
            ).fetchone()
            selected_customer = dict(row) if row else None
            if selected_customer:
                queries = _aeo_list(conn, customer_id)
    return render_template(
        "crm_aeo.html",
        all_customers=all_customers,
        selected_customer=selected_customer,
        queries=queries,
        customer_id=customer_id,
        **_admin_page_template_kwargs(),
    )


@app.post("/api/crm/aeo/queries")
def api_crm_aeo_add_query() -> Any:
    from crm_aeo import add_query as _aeo_add
    body = request.get_json(silent=True) or {}
    customer_id = _opt_pos_int(body.get("customer_id"))
    query_text = str(body.get("query_text", "")).strip()[:500]
    brand_name = str(body.get("brand_name", "")).strip()[:200]
    lifecycle_id = _opt_pos_int(body.get("lifecycle_id"))
    notes = str(body.get("notes", "")).strip()[:1000]
    if not customer_id or not query_text or not brand_name:
        return jsonify({"error": "Thiếu customer_id, query_text hoặc brand_name"}), 400
    with get_connection() as conn:
        qid = _aeo_add(conn, customer_id, query_text, brand_name, lifecycle_id=lifecycle_id, notes=notes)
    return jsonify({"id": qid})


@app.delete("/api/crm/aeo/queries/<int:query_id>")
def api_crm_aeo_delete_query(query_id: int) -> Any:
    from crm_aeo import delete_query as _aeo_del
    with get_connection() as conn:
        _aeo_del(conn, query_id)
    return jsonify({})


@app.post("/api/crm/aeo/queries/<int:query_id>/scan")
def api_crm_aeo_scan(query_id: int) -> Any:
    from crm_aeo import run_scan as _aeo_scan, get_scan_history as _aeo_history
    with get_connection() as conn:
        output = _aeo_scan(conn, query_id)
        history = _aeo_history(conn, query_id)
    latest = history[0] if history else {}
    return jsonify({
        "ai_response": output,
        "brand_visible": latest.get("brand_visible", 0),
        "gap_notes": latest.get("gap_notes", ""),
        "created_at": latest.get("created_at", ""),
    })


@app.post("/api/crm/aeo/queries/<int:query_id>/content")
def api_crm_aeo_content(query_id: int) -> Any:
    from crm_aeo import generate_content as _aeo_gen
    with get_connection() as conn:
        result = _aeo_gen(conn, query_id)
    return jsonify(result)


@app.get("/crm/proposals")
def crm_proposals_page() -> Any:
    redir = _ensure_crm_session_html()
    if redir is not None:
        return redir
    from crm_proposal import list_proposals as _prop_list, SERVICE_NAMES as _svc_names
    customer_id = _opt_pos_int(request.args.get("customer_id"))
    lifecycle_id = _opt_pos_int(request.args.get("lifecycle_id"))
    prefill_service_slug = str(request.args.get("service_slug") or "").strip()
    from_consult = bool(request.args.get("from_consult"))
    with get_connection() as conn:
        all_customers = [
            dict(r) for r in conn.execute(
                "SELECT id, name, company FROM crm_customers ORDER BY name"
            ).fetchall()
        ]
        proposals = []
        selected_customer = None
        proposal_prefill_notes = ""
        if customer_id:
            row = conn.execute(
                "SELECT id, name, company FROM crm_customers WHERE id = ?",
                (customer_id,),
            ).fetchone()
            selected_customer = dict(row) if row else None
            if selected_customer:
                proposals = _prop_list(conn, customer_id)
            if lifecycle_id and from_consult:
                from crm_proposal import get_customer_context as _prop_ctx

                pctx = _prop_ctx(conn, customer_id, lifecycle_id=lifecycle_id)
                consult = pctx.get("consult") or {}
                note_parts: list[str] = []
                if consult.get("notes"):
                    note_parts.append(str(consult["notes"]))
                if consult.get("ai_output"):
                    note_parts.append(
                        "Consult AI:\n" + str(consult["ai_output"])[:1500]
                    )
                if note_parts:
                    proposal_prefill_notes = "\n\n".join(note_parts)[:2000]
    return render_template(
        "crm_proposals.html",
        all_customers=all_customers,
        selected_customer=selected_customer,
        proposals=proposals,
        customer_id=customer_id,
        lifecycle_id=lifecycle_id,
        prefill_service_slug=prefill_service_slug,
        proposal_prefill_notes=proposal_prefill_notes,
        from_consult=from_consult,
        service_names=_svc_names,
        **_admin_page_template_kwargs(),
    )


@app.get("/crm/proposals/<int:proposal_id>/preview")
def crm_proposal_preview_page(proposal_id: int) -> Any:
    redir = _ensure_crm_session_html()
    if redir is not None:
        return redir
    from crm_proposal import get_proposal as _prop_get, get_customer_context as _prop_ctx, SERVICE_NAMES as _svc_names
    with get_connection() as conn:
        proposal = _prop_get(conn, proposal_id)
        if proposal is None:
            return "Không tìm thấy đề xuất.", 404
        ctx = _prop_ctx(
            conn,
            proposal["customer_id"],
            lifecycle_id=proposal.get("lifecycle_id"),
        )
    return render_template(
        "crm_proposal_preview.html",
        proposal=proposal,
        customer=ctx["customer"],
        service_names=_svc_names,
        **_admin_page_template_kwargs(),
    )


@app.post("/api/crm/proposals")
def api_crm_proposals_create() -> Any:
    from crm_proposal import create_proposal as _prop_create
    body = request.get_json(silent=True) or {}
    customer_id = _opt_pos_int(body.get("customer_id"))
    service_slugs = [str(s).strip() for s in body.get("service_slugs", []) if str(s).strip()]
    total_vnd = _opt_pos_int(body.get("total_vnd")) or 0
    timeline_months = _opt_pos_int(body.get("timeline_months")) or 1
    notes = str(body.get("notes", "")).strip()[:2000]
    lifecycle_id = _opt_pos_int(body.get("lifecycle_id"))
    if not customer_id or not service_slugs:
        return jsonify({"error": "Thiếu customer_id hoặc service_slugs"}), 400
    with get_connection() as conn:
        pid = _prop_create(
            conn, customer_id, service_slugs, total_vnd, timeline_months, notes,
            lifecycle_id=lifecycle_id,
        )
    return jsonify({"id": pid})


@app.post("/api/crm/proposals/<int:proposal_id>/generate")
def api_crm_proposals_generate(proposal_id: int) -> Any:
    from crm_proposal import run_proposal_ai as _prop_ai, get_proposal as _prop_get
    with get_connection() as conn:
        sections = _prop_ai(conn, proposal_id)
        if not sections or not any(sections.values()):
            return jsonify({"error": "Không thể tạo nội dung AI"}), 500
        proposal = _prop_get(conn, proposal_id)
    return jsonify({**sections, "updated_at": proposal["updated_at"] if proposal else ""})


@app.delete("/api/crm/proposals/<int:proposal_id>")
def api_crm_proposals_delete(proposal_id: int) -> Any:
    from crm_proposal import delete_proposal as _prop_del
    with get_connection() as conn:
        _prop_del(conn, proposal_id)
    return jsonify({})


@app.get("/crm/staff")
def crm_staff_page() -> str:
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    return render_template(
        "crm_staff.html",
        **_admin_page_template_kwargs(),
    )


@app.get("/cms")
def cms() -> str:
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    with get_connection() as conn:
        grants = _cms_load_role_grants(conn, _cms_session_role())
        role_row = conn.execute(
            "SELECT name FROM cms_roles WHERE code = ?",
            (_cms_session_role(),),
        ).fetchone()
    if _admin_full_access():
        grants = {mid: list(CMS_ACTIONS) for mid in CMS_MODULE_IDS}
    role_name = str(role_row["name"]) if role_row else _cms_session_role()
    return render_template(
        "cms.html",
        cms_role_code=_cms_session_role(),
        cms_role_name=role_name,
        cms_username=_cms_session_username(),
        cms_grants_json=json.dumps(grants, ensure_ascii=False),
        cms_can_configure_matrix=_cms_can("permissions_matrix", "configure"),
        **_admin_page_template_kwargs(),
    )


@app.get("/api/projects")
def get_projects():
    with get_connection() as conn:
        rows = conn.execute("SELECT * FROM projects ORDER BY id DESC").fetchall()
    return jsonify(rows_to_dict(rows))


@app.post("/api/projects")
def create_project():
    if not _cms_can("projects", "create"):
        return _cms_forbidden_json("projects", "create")
    payload = request.get_json(force=True)
    title = str(payload.get("title") or "").strip()
    if not title:
        return jsonify({"error": "Missing fields: title"}), 400

    created_at = payload.get("created_at") or datetime.now().strftime("%Y-%m-%d")
    with get_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO projects (title, category, image_url, intro, description, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                title,
                str(payload.get("category") or "").strip(),
                str(payload.get("image_url") or "").strip(),
                str(payload.get("intro") or "").strip(),
                str(payload.get("description") or "").strip(),
                created_at,
            ),
        )
        new_id = cursor.lastrowid
        row = conn.execute("SELECT * FROM projects WHERE id = ?", (new_id,)).fetchone()

    return jsonify(dict(row)), 201


@app.put("/api/projects/<int:project_id>")
def update_project(project_id: int):
    if not _cms_can("projects", "edit"):
        return _cms_forbidden_json("projects", "edit")
    payload = request.get_json(force=True)
    with get_connection() as conn:
        current = conn.execute(
            "SELECT * FROM projects WHERE id = ?", (project_id,)
        ).fetchone()
        if current is None:
            return jsonify({"error": "Project not found"}), 404

        merged = dict(current)
        merged.update(payload)
        conn.execute(
            """
            UPDATE projects
            SET title = ?, category = ?, image_url = ?, intro = ?, description = ?, created_at = ?
            WHERE id = ?
            """,
            (
                merged["title"],
                merged["category"],
                merged["image_url"],
                merged.get("intro") or "",
                merged["description"],
                merged["created_at"],
                project_id,
            ),
        )

        row = conn.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
    return jsonify(dict(row))


@app.delete("/api/projects/<int:project_id>")
def delete_project(project_id: int):
    if not _cms_can("projects", "delete"):
        return _cms_forbidden_json("projects", "delete")
    with get_connection() as conn:
        cursor = conn.execute("DELETE FROM projects WHERE id = ?", (project_id,))
        if cursor.rowcount == 0:
            return jsonify({"error": "Project not found"}), 404
    return "", 204


@app.get("/api/news")
def get_news():
    with get_connection() as conn:
        rows = conn.execute("SELECT * FROM news ORDER BY id DESC").fetchall()
    return jsonify(rows_to_dict(rows))


@app.post("/api/news")
def create_news():
    if not _cms_can("news", "create"):
        return _cms_forbidden_json("news", "create")
    payload = request.get_json(force=True)
    required = ["title", "summary", "url"]
    missing = [field for field in required if not payload.get(field)]
    if missing:
        return jsonify({"error": f"Missing fields: {', '.join(missing)}"}), 400

    published_at = payload.get("published_at") or datetime.now().strftime("%Y-%m-%d")
    image_url = str(payload.get("image_url") or "")
    with get_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO news (title, summary, url, published_at, image_url)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                payload["title"],
                payload["summary"],
                payload["url"],
                published_at,
                image_url,
            ),
        )
        row = conn.execute(
            "SELECT * FROM news WHERE id = ?",
            (cursor.lastrowid,),
        ).fetchone()
    return jsonify(dict(row)), 201


@app.put("/api/news/<int:news_id>")
def update_news(news_id: int):
    if not _cms_can("news", "edit"):
        return _cms_forbidden_json("news", "edit")
    payload = request.get_json(force=True)
    with get_connection() as conn:
        current = conn.execute("SELECT * FROM news WHERE id = ?", (news_id,)).fetchone()
        if current is None:
            return jsonify({"error": "News not found"}), 404

        merged = dict(current)
        merged.update(payload)
        conn.execute(
            """
            UPDATE news
            SET title = ?, summary = ?, url = ?, published_at = ?, image_url = ?
            WHERE id = ?
            """,
            (
                merged["title"],
                merged["summary"],
                merged["url"],
                merged["published_at"],
                str(merged.get("image_url") or ""),
                news_id,
            ),
        )
        row = conn.execute("SELECT * FROM news WHERE id = ?", (news_id,)).fetchone()

    return jsonify(dict(row))


@app.delete("/api/news/<int:news_id>")
def delete_news(news_id: int):
    if not _cms_can("news", "delete"):
        return _cms_forbidden_json("news", "delete")
    with get_connection() as conn:
        cursor = conn.execute("DELETE FROM news WHERE id = ?", (news_id,))
        if cursor.rowcount == 0:
            return jsonify({"error": "News not found"}), 404
    return "", 204


@app.get("/api/crm/cases")
def api_crm_list_cases() -> Any:
    q_raw = (request.args.get("q") or "").strip().lower()
    staff_raw = (request.args.get("staff_id") or "").strip()
    staff_id: int | None = None
    portal_sid = _crm_effective_staff_id()
    if portal_sid is not None:
        staff_id = portal_sid
    elif staff_raw:
        try:
            staff_id = int(staff_raw)
            if staff_id <= 0:
                staff_id = None
        except ValueError:
            staff_id = None
    with get_connection() as conn:
        ch_labels = _crm_lead_channel_labels_map(conn)
        if staff_id is not None:
            rows = conn.execute(
                f"""
                {_CRM_CASE_SELECT}
                WHERE c.assigned_staff_id = ?
                ORDER BY datetime(c.updated_at) DESC
                """,
                (staff_id,),
            ).fetchall()
        else:
            rows = conn.execute(
                f"{_CRM_CASE_SELECT} ORDER BY datetime(c.updated_at) DESC"
            ).fetchall()
        case_ids = [int(r["id"]) for r in rows]
        last_care = fetch_last_care_reports_map(conn, case_ids)
    cases = [_crm_row_case(r, ch_labels) for r in rows]
    for c in cases:
        lc = last_care.get(int(c["id"]))
        if lc:
            c["last_care_report"] = lc
    if q_raw:

        def _hay(c: dict[str, Any]) -> str:
            return " ".join(
                str(c.get(k, "") or "").lower()
                for k in (
                    "title",
                    "description",
                    "assigned_to",
                    "customer_name",
                    "customer_phone",
                    "customer_email",
                    "customer_company",
                )
            )

        cases = [c for c in cases if q_raw in _hay(c)]
    if portal_sid is not None:
        cases = [
            c
            for c in cases
            if normalize_pipeline_stage(c.get("pipeline_stage") or c.get("status"))
            not in TERMINAL_STAGES
        ]
    return jsonify({"cases": cases, "staff_id": staff_id})


@app.get("/api/crm/cases/<int:case_id>")
def api_crm_case_detail(case_id: int) -> Any:
    with get_connection() as conn:
        row = conn.execute(
            f'{_CRM_CASE_SELECT} WHERE c.id = ?',
            (case_id,),
        ).fetchone()
        if row is None:
            return jsonify({"error": "Case not found"}), 404
        portal_sid = _crm_effective_staff_id()
        if portal_sid is not None and not _crm_case_assigned_to_staff(conn, case_id, portal_sid):
            return _crm_forbid_staff_case()
        ch_labels = _crm_lead_channel_labels_map(conn)
        events = conn.execute(
            """
            SELECT id, kind, body, created_at FROM crm_case_events
            WHERE case_id = ?
            ORDER BY id ASC
            """,
            (case_id,),
        ).fetchall()
        care_rows = conn.execute(
            """
            SELECT * FROM crm_care_reports
            WHERE case_id = ?
            ORDER BY id DESC
            LIMIT 50
            """,
            (case_id,),
        ).fetchall()
    data = _crm_row_case(row, ch_labels)
    data["events"] = [dict(e) for e in events]
    data["care_reports"] = [care_report_row_to_dict(r) for r in care_rows]
    if care_rows:
        data["last_care_report"] = care_report_row_to_dict(care_rows[0])
    return jsonify(data)


@app.post("/api/crm/cases")
def api_crm_create_case() -> Any:
    if _staff_logged_in():
        return jsonify({"error": "Nhân viên không tạo yêu cầu mới — liên hệ quản trị."}), 403
    payload = request.get_json(force=True) or {}
    cust_raw = payload.get("customer")
    customer_data = cust_raw if isinstance(cust_raw, dict) else {}
    title = str(payload.get("title", "")).strip()[:800]
    description = str(payload.get("description", "")).strip()[:8000]
    channel = str(payload.get("channel", "khac")).strip()
    priority = str(payload.get("priority", "binh_thuong")).strip()
    status = str(payload.get("status", "tiep_nhan")).strip()

    cust_name = str(customer_data.get("name", "")).strip()[:240]
    cust_phone = str(customer_data.get("phone", "")).strip()[:80]
    cust_email = str(customer_data.get("email", "")).strip()[:240]
    cust_address = str(customer_data.get("address", "")).strip()[:500]
    cust_company = str(customer_data.get("company", "")).strip()[:400]

    if not title:
        return jsonify({"error": "Thiếu tiêu đề yêu cầu (title)"}), 400
    if not cust_name:
        return jsonify({"error": "Thiếu tên khách hàng"}), 400
    if not cust_phone and not cust_email:
        return jsonify({"error": "Cần ít nhất số điện thoại hoặc email khách"}), 400
    if priority not in CRM_PRIORITIES:
        priority = "binh_thuong"
    if status not in CRM_STATUSES_ORDER:
        status = "tiep_nhan"

    ts = _crm_ts()
    short_date = datetime.now().strftime("%Y-%m-%d")

    with get_connection() as conn:
        channel = _crm_resolve_lead_channel(conn, channel)
        ch_labels = _crm_lead_channel_labels_map(conn)
        try:
            assigned_to, assigned_staff_id = _crm_resolve_assignment(
                conn, payload, allow_legacy_text=True
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

        cur_cu = conn.execute(
            """
            INSERT INTO crm_customers (name, phone, email, address, company, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (cust_name, cust_phone, cust_email, cust_address, cust_company, short_date),
        )
        customer_id = int(cur_cu.lastrowid)
        assigned_at_val = ts if assigned_staff_id else ""
        campaign_id_final: int | None = None
        if "campaign_id" in payload:
            pcamp = _opt_pos_int(payload.get("campaign_id"))
            if pcamp is not None:
                if conn.execute("SELECT id FROM crm_campaigns WHERE id = ?", (pcamp,)).fetchone():
                    campaign_id_final = pcamp
                else:
                    return jsonify({"error": "Chiến dịch (campaign_id) không tồn tại"}), 400
        cur_case = conn.execute(
            """
            INSERT INTO crm_cases (
              customer_id, title, description, channel, priority, status,
              assigned_to, assigned_staff_id, assigned_at, created_at, updated_at,
              campaign_id
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                customer_id,
                title,
                description,
                channel,
                priority,
                status,
                assigned_to,
                assigned_staff_id,
                assigned_at_val,
                ts,
                ts,
                campaign_id_final,
            ),
        )
        cid = int(cur_case.lastrowid)
        _crm_append_event(conn, cid, "ghi_chu", "Tạo yêu cầu trong quy trình chăm sóc khách hàng.")
        if status != "tiep_nhan":
            _crm_append_event(
                conn,
                cid,
                "trang_thai",
                f"Trạng thái ban đầu: {CRM_STATUS_LABELS_VI.get(status, status)}.",
            )
        lead_src = str(payload.get("lead_source") or "").strip()[:120]
        auto_assign = payload.get("auto_assign", True) is not False
        ato, aid = on_case_created(
            conn,
            cid,
            title=title,
            priority=priority,
            assigned_staff_id=assigned_staff_id,
            assigned_to=assigned_to,
            lead_source=lead_src,
            auto_assign=auto_assign,
        )
        if aid and not assigned_staff_id:
            _crm_append_event(
                conn,
                cid,
                "phan_cong",
                f"Phân công tự động (round-robin): {ato}.",
            )
        elif assigned_staff_id:
            _crm_append_event(
                conn,
                cid,
                "phan_cong",
                f"Phân công phụ trách: {assigned_to}.",
            )

        joined = conn.execute(
            f'{_CRM_CASE_SELECT} WHERE c.id = ?',
            (cid,),
        ).fetchone()

    assert joined is not None
    return jsonify(_crm_row_case(joined, ch_labels)), 201


@app.patch("/api/crm/cases/<int:case_id>")
def api_crm_patch_case(case_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    portal_sid = _crm_effective_staff_id()
    if portal_sid is not None:
        payload = dict(payload)
        payload.pop("assigned_staff_id", None)
        payload.pop("assigned_to", None)
    with get_connection() as conn:
        ch_labels = _crm_lead_channel_labels_map(conn)
        row = conn.execute("SELECT * FROM crm_cases WHERE id = ?", (case_id,)).fetchone()
        if row is None:
            return jsonify({"error": "Case not found"}), 404
        if portal_sid is not None and not _crm_case_assigned_to_staff(conn, case_id, portal_sid):
            return _crm_forbid_staff_case()
        prev = dict(row)
        merged = dict(prev)
        touched = False

        pipeline_explicit = False
        new_stage_raw = payload.get("pipeline_stage")
        if isinstance(new_stage_raw, str) and new_stage_raw.strip():
            ps = normalize_pipeline_stage(new_stage_raw.strip())
            old_ps = normalize_pipeline_stage(
                prev.get("pipeline_stage") or prev.get("status")
            )
            if ps != old_ps:
                pipeline_explicit = True
                on_pipeline_stage_change(
                    conn,
                    case_id,
                    old_stage=old_ps,
                    new_stage=ps,
                    title=str(prev.get("title") or merged.get("title") or ""),
                    assigned_staff_id=merged.get("assigned_staff_id")
                    or prev.get("assigned_staff_id"),
                    append_event=_crm_append_event,
                )
                merged["pipeline_stage"] = ps
                merged["stage_entered_at"] = _crm_ts()
                merged["status"] = legacy_status_for_stage(ps)
                touched = True

        new_status_raw = payload.get("status")
        if isinstance(new_status_raw, str) and new_status_raw.strip():
            ns = new_status_raw.strip()
            if ns not in CRM_STATUSES_ORDER:
                return jsonify({"error": "status không hợp lệ"}), 400
            if prev["status"] != ns:
                _crm_append_event(
                    conn,
                    case_id,
                    "trang_thai",
                    (
                        "Chuyển trạng thái: "
                        f"{CRM_STATUS_LABELS_VI.get(prev['status'], prev['status'])}"
                        " → "
                        f"{CRM_STATUS_LABELS_VI.get(ns, ns)}"
                    ),
                )
                merged["status"] = ns
                if not pipeline_explicit:
                    ps = LEGACY_STATUS_TO_PIPELINE.get(ns, "moi")
                    old_ps = normalize_pipeline_stage(
                        merged.get("pipeline_stage") or prev.get("pipeline_stage") or prev.get("status")
                    )
                    if ps != old_ps:
                        on_pipeline_stage_change(
                            conn,
                            case_id,
                            old_stage=old_ps,
                            new_stage=ps,
                            title=str(prev.get("title") or ""),
                            assigned_staff_id=merged.get("assigned_staff_id")
                            or prev.get("assigned_staff_id"),
                            append_event=_crm_append_event,
                        )
                        merged["pipeline_stage"] = ps
                        merged["stage_entered_at"] = _crm_ts()
                touched = True

        if "assigned_staff_id" in payload or (
            "assigned_to" in payload and isinstance(payload.get("assigned_to"), str)
        ):
            try:
                ato, aid = _crm_resolve_assignment(conn, payload, allow_legacy_text=True)
            except ValueError as exc:
                return jsonify({"error": str(exc)}), 400
            pid = prev.get("assigned_staff_id")
            pto = str(prev.get("assigned_to") or "")
            if pid != aid or pto != str(ato):
                merged["assigned_to"] = ato
                merged["assigned_staff_id"] = aid
                merged["assigned_at"] = _crm_ts() if aid else ""
                old_l = _crm_assign_label(conn, prev.get("assigned_to"), prev.get("assigned_staff_id"))
                if aid:
                    new_l = _crm_assign_label(conn, ato, aid)
                    _crm_append_event(
                        conn,
                        case_id,
                        "phan_cong",
                        f"Phân công phụ trách: {old_l} → {new_l}.",
                    )
                else:
                    _crm_append_event(
                        conn,
                        case_id,
                        "phan_cong",
                        f"Gỡ phụ trách (trước đây: {old_l}).",
                    )
                touched = True

        for key in ("title", "description"):
            if key in payload and isinstance(payload[key], str):
                stripped = payload[key].strip()
                merged[key] = stripped[:8000] if key == "description" else stripped[:800]
                touched = True

        if "priority" in payload and isinstance(payload["priority"], str):
            pr = payload["priority"].strip()
            if pr in CRM_PRIORITIES:
                merged["priority"] = pr
                touched = True

        if "channel" in payload and isinstance(payload["channel"], str):
            ch = _crm_resolve_lead_channel(conn, payload["channel"].strip())
            if ch != merged.get("channel"):
                merged["channel"] = ch
                touched = True

        if "campaign_id" in payload:
            cid_raw = payload.get("campaign_id")
            if cid_raw in (None, "", 0, "0"):
                merged["campaign_id"] = None
                touched = True
            else:
                try:
                    pc = int(cid_raw)
                except (TypeError, ValueError):
                    return jsonify({"error": "campaign_id không hợp lệ"}), 400
                if pc <= 0:
                    merged["campaign_id"] = None
                    touched = True
                elif conn.execute("SELECT id FROM crm_campaigns WHERE id = ?", (pc,)).fetchone():
                    merged["campaign_id"] = pc
                    touched = True
                else:
                    return jsonify({"error": "Không tìm thấy chiến dịch"}), 404

        if "lead_source" in payload and isinstance(payload["lead_source"], str):
            merged["lead_source"] = payload["lead_source"].strip()[:120]
            touched = True

        if "deal_value_vnd" in payload:
            raw_dv = payload.get("deal_value_vnd")
            if raw_dv in (None, "", 0, "0"):
                merged["deal_value_vnd"] = None
            else:
                try:
                    merged["deal_value_vnd"] = max(0, int(raw_dv))
                except (TypeError, ValueError):
                    return jsonify({"error": "deal_value_vnd không hợp lệ"}), 400
            touched = True

        if touched:
            merged["updated_at"] = _crm_ts()
            conn.execute(
                """
                UPDATE crm_cases
                SET title = ?, description = ?, channel = ?, priority = ?, status = ?,
                    assigned_to = ?, assigned_staff_id = ?, assigned_at = ?,
                    campaign_id = ?, pipeline_stage = ?, stage_entered_at = ?,
                    lead_source = ?, deal_value_vnd = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    merged["title"],
                    merged["description"],
                    merged["channel"],
                    merged["priority"],
                    merged["status"],
                    merged["assigned_to"],
                    merged.get("assigned_staff_id"),
                    str(merged.get("assigned_at") or ""),
                    merged.get("campaign_id"),
                    normalize_pipeline_stage(
                        merged.get("pipeline_stage") or merged.get("status")
                    ),
                    str(merged.get("stage_entered_at") or prev.get("stage_entered_at") or _crm_ts()),
                    str(merged.get("lead_source") or ""),
                    merged.get("deal_value_vnd"),
                    merged["updated_at"],
                    case_id,
                ),
            )

        joined = conn.execute(
            f'{_CRM_CASE_SELECT} WHERE c.id = ?',
            (case_id,),
        ).fetchone()

    assert joined is not None
    return jsonify(_crm_row_case(joined, ch_labels))


_CRM_STAFF_PIPELINE_SUB = (
    "(SELECT COUNT(*) FROM crm_cases c WHERE c.assigned_staff_id = s.id "
    "AND c.status != 'dong') AS pipeline_case_count"
)

_CRM_STAFF_FROM = """
FROM crm_staff s
LEFT JOIN crm_departments d ON d.id = s.department_id
LEFT JOIN crm_staff mgr ON mgr.id = s.reports_to_id
LEFT JOIN crm_positions pos ON pos.id = s.position_id
"""


def _crm_staff_params_from_request() -> tuple[bool, str, str, int | None]:
    """Tham số lọc giống GET /api/crm/staff (và export)."""
    raw_flag = (request.args.get("include_inactive") or "").strip().lower()
    include_inactive = raw_flag in ("1", "true", "yes", "all")
    q_raw = (request.args.get("q") or "").strip()
    status_filter = (request.args.get("status") or "all").strip().lower()
    if status_filter not in ("all", "active", "inactive"):
        status_filter = "all"
    dept_filter = _opt_pos_int(request.args.get("department_id"))
    return include_inactive, status_filter, q_raw, dept_filter


def _crm_staff_pagination_from_request() -> tuple[int, int]:
    """page (≥1), per_page (1…100)."""
    try:
        page = int(request.args.get("page") or 1)
    except (TypeError, ValueError):
        page = 1
    try:
        per_page = int(request.args.get("per_page") or 25)
    except (TypeError, ValueError):
        per_page = 25
    page = max(1, page)
    per_page = min(100, max(1, per_page))
    return page, per_page


def _crm_staff_where_and_order(
    *,
    include_inactive: bool,
    status_filter: str,
    q_raw: str,
    dept_filter: int | None,
) -> tuple[str, list[Any], str]:
    """Trả về mệnh đề WHERE (có thể rỗng), params, ORDER BY."""
    where_parts: list[str] = []
    params: list[Any] = []

    if not include_inactive:
        where_parts.append("s.active = 1")
    else:
        if status_filter == "active":
            where_parts.append("s.active = 1")
        elif status_filter == "inactive":
            where_parts.append("s.active = 0")

    if dept_filter is not None:
        where_parts.append("s.department_id = ?")
        params.append(dept_filter)

    if q_raw:
        pat = f"%{q_raw.lower()}%"
        where_parts.append(
            "("
            "LOWER(s.name) LIKE ? OR LOWER(COALESCE(s.email,'')) LIKE ? "
            "OR LOWER(COALESCE(s.phone,'')) LIKE ? "
            "OR LOWER(COALESCE(s.job_title,'')) LIKE ? OR LOWER(COALESCE(s.department,'')) LIKE ? "
            "OR LOWER(COALESCE(s.internal_code,'')) LIKE ? OR LOWER(COALESCE(s.notes,'')) LIKE ? "
            "OR LOWER(COALESCE(d.name,'')) LIKE ? OR LOWER(COALESCE(d.code,'')) LIKE ? "
            "OR LOWER(COALESCE(pos.name,'')) LIKE ? OR LOWER(COALESCE(pos.code,'')) LIKE ?"
            ")"
        )
        params.extend([pat] * 11)

    where_sql = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""
    order_sql = (
        "ORDER BY s.active DESC, s.sort_order ASC, s.name COLLATE NOCASE ASC"
        if include_inactive
        else "ORDER BY s.sort_order ASC, s.name COLLATE NOCASE ASC"
    )
    return where_sql, params, order_sql


def _count_crm_staff_filtered(
    conn: sqlite3.Connection,
    *,
    include_inactive: bool,
    status_filter: str,
    q_raw: str,
    dept_filter: int | None,
) -> int:
    where_sql, params, _ = _crm_staff_where_and_order(
        include_inactive=include_inactive,
        status_filter=status_filter,
        q_raw=q_raw,
        dept_filter=dept_filter,
    )
    sql = f"SELECT COUNT(*) AS n {_CRM_STAFF_FROM} {where_sql}"
    row = conn.execute(sql, params).fetchone()
    return int(row["n"]) if row else 0


def _fetch_crm_staff_rows(
    conn: sqlite3.Connection,
    *,
    include_inactive: bool,
    status_filter: str,
    q_raw: str,
    dept_filter: int | None,
    limit: int | None = None,
    offset: int | None = None,
) -> list[sqlite3.Row]:
    where_sql, params, order_sql = _crm_staff_where_and_order(
        include_inactive=include_inactive,
        status_filter=status_filter,
        q_raw=q_raw,
        dept_filter=dept_filter,
    )
    pipeline_sub = _CRM_STAFF_PIPELINE_SUB
    sql = f"""
        SELECT s.*, {pipeline_sub},
               d.code AS dept_code, d.name AS dept_name,
               mgr.name AS reports_to_name,
               pos.name AS position_catalog_name,
               pos.code AS position_catalog_code
        {_CRM_STAFF_FROM}
        {where_sql}
        {order_sql}
    """
    exec_params: list[Any] = list(params)
    if limit is not None and offset is not None:
        sql += " LIMIT ? OFFSET ?"
        exec_params.extend([limit, offset])
    return conn.execute(sql, exec_params).fetchall()


def _staff_export_row_values(row: sqlite3.Row) -> list[Any]:
    d = dict(row)
    emp_key = str(d.get("employment_type") or "").strip()
    emp_label = CRM_EMPLOYMENT_LABELS_VI.get(emp_key, emp_key or "—")
    act = d.get("active")
    active_label = "Hoạt động" if act in (1, True) else "Ngưng"
    pipeline = int(d.get("pipeline_case_count") or 0)
    dept_name = str(d.get("dept_name") or "").strip()
    dept_code = str(d.get("dept_code") or "").strip()
    return [
        str(d.get("internal_code") or "").strip(),
        str(d.get("name") or ""),
        str(d.get("job_title") or "").strip(),
        dept_name or "—",
        dept_code or "—",
        str(d.get("department") or "").strip(),
        str(d.get("phone") or "").strip(),
        str(d.get("email") or "").strip(),
        emp_label,
        str(d.get("started_on") or "").strip(),
        str(d.get("ended_on") or "").strip(),
        str(d.get("reports_to_name") or "").strip(),
        pipeline,
        active_label,
        str(d.get("notes") or "").strip(),
    ]


_STAFF_EXPORT_HEADERS = [
    "Mã NV",
    "Họ tên",
    "Chức danh",
    "Phòng ban (danh mục)",
    "Mã phòng ban",
    "Nhãn nhóm",
    "Điện thoại",
    "Email",
    "Loại hợp đồng",
    "Ngày bắt đầu",
    "Ngày kết thúc",
    "Quản lý trực tiếp",
    "Hồ sơ chưa đóng",
    "Trạng thái",
    "Ghi chú nội bộ",
]


def _staff_export_csv_response(rows: list[sqlite3.Row], *, filename: str) -> Response:
    si = StringIO()
    w = csv.writer(si)
    w.writerow(_STAFF_EXPORT_HEADERS)
    for r in rows:
        w.writerow(_staff_export_row_values(r))
    raw = si.getvalue().encode("utf-8-sig")
    safe = filename.replace('"', "")
    return Response(
        raw,
        mimetype="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": (f'attachment; filename="{safe}"; filename*=UTF-8\'\'{quote(filename)}'),
        },
    )


def _staff_export_xlsx_response(rows: list[sqlite3.Row], *, filename: str) -> Response:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Nhân viên CRM"
    ws.append(list(_STAFF_EXPORT_HEADERS))
    for r in rows:
        ws.append(_staff_export_row_values(r))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
        etag=False,
        max_age=0,
        conditional=False,
    )


@app.get("/api/crm/staff/export")
def api_crm_export_staff() -> Any:
    """Xuất danh sách nhân viên theo cùng bộ lọc với bảng (CSV hoặc Excel)."""
    fmt = (request.args.get("format") or "xlsx").strip().lower()
    if fmt not in ("csv", "xlsx"):
        fmt = "xlsx"
    inc, st, q_raw, df = _crm_staff_params_from_request()
    with get_connection() as conn:
        rows = _fetch_crm_staff_rows(
            conn, include_inactive=inc, status_filter=st, q_raw=q_raw, dept_filter=df
        )
    stamp = datetime.now().strftime("%Y-%m-%d")
    base = f"crm-nhan-vien-{stamp}"
    if fmt == "csv":
        return _staff_export_csv_response(rows, filename=f"{base}.csv")
    return _staff_export_xlsx_response(rows, filename=f"{base}.xlsx")


@app.get("/api/crm/staff/daily-work-report-template")
def api_crm_daily_work_report_template() -> Any:
    """Tải mẫu Excel báo cáo công việc hàng ngày (có thể điền sẵn thông tin NV)."""
    if not (
        _admin_section_can("crm_staff_roster", "export")
        or _admin_section_can("crm_staff_roster", "view")
    ):
        return _admin_section_forbidden_json("crm_staff_roster", "export")
    try:
        year = int(request.args.get("year") or 0)
    except (TypeError, ValueError):
        year = 0
    try:
        month = int(request.args.get("month") or 0)
    except (TypeError, ValueError):
        month = 0
    if year <= 0 or month <= 0:
        now = datetime.now()
        year = year if year > 0 else now.year
        month = month if month > 0 else now.month

    staff_raw = str(request.args.get("staff_id") or "").strip()
    staff_id: int | None = None
    if staff_raw:
        try:
            staff_id = int(staff_raw)
        except ValueError:
            return jsonify({"error": "staff_id không hợp lệ"}), 400
        if staff_id <= 0:
            staff_id = None

    portal_sid = _crm_effective_staff_id()
    if portal_sid is not None:
        staff_id = portal_sid

    settings = fetch_settings()
    brand = str(settings.get("brand_name") or "PTT Advertising Solutions").strip()
    staff_dict: dict[str, Any] | None = None
    if staff_id is not None:
        with get_connection() as conn:
            row = conn.execute(
                """
                SELECT id, name, internal_code, department, job_title
                FROM crm_staff WHERE id = ?
                """,
                (staff_id,),
            ).fetchone()
            if row is None:
                return jsonify({"error": "Không tìm thấy nhân viên"}), 404
            staff_dict = {
                "name": str(row["name"] or ""),
                "internal_code": str(row["internal_code"] or "").strip(),
                "department": str(row["department"] or "").strip(),
                "job_title": str(row["job_title"] or "").strip(),
            }

    buf, fname = daily_work_report_xlsx_response(
        brand=brand,
        staff=staff_dict,
        year=year,
        month=month,
    )
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
        etag=False,
        max_age=0,
        conditional=False,
    )


def _crm_daily_report_staff_id_from_request(payload: dict[str, Any] | None = None) -> tuple[int | None, str | None]:
    """Trả (staff_id, error). Portal luôn dùng session."""
    portal_sid = _crm_effective_staff_id()
    if portal_sid is not None:
        return int(portal_sid), None
    raw = ""
    if payload is not None:
        raw = str(payload.get("staff_id") or "").strip()
    else:
        raw = str(request.args.get("staff_id") or "").strip()
    if not raw:
        return None, "Cần chọn nhân viên."
    try:
        sid = int(raw)
    except ValueError:
        return None, "staff_id không hợp lệ"
    if sid <= 0:
        return None, "staff_id không hợp lệ"
    return sid, None


def _crm_parse_daily_report_payload(payload: dict[str, Any]) -> tuple[dict[str, Any] | None, str | None]:
    report_date = validate_report_date(payload.get("report_date"))
    if not report_date:
        return None, "report_date phải là YYYY-MM-DD"
    summary = str(payload.get("summary") or "").strip()
    tomorrow_plan = str(payload.get("tomorrow_plan") or "").strip()
    support_needed = str(payload.get("support_needed") or "").strip()
    tasks = normalize_tasks(payload.get("tasks"))
    if not summary and not tasks:
        return None, "Cần tổng kết hoặc ít nhất một công việc."
    if len(summary) > 8000:
        return None, "Tổng kết quá dài"
    hw_raw = payload.get("hours_worked")
    hours_worked: float | None = None
    if hw_raw is not None and str(hw_raw).strip() != "":
        try:
            hours_worked = float(hw_raw)
            if hours_worked < 0 or hours_worked > 24:
                return None, "Giờ làm phải từ 0 đến 24"
        except (TypeError, ValueError):
            return None, "Giờ làm không hợp lệ"
    status = str(payload.get("status") or "submitted").strip().lower()
    if status not in ("draft", "submitted"):
        status = "submitted"
    return {
        "report_date": report_date,
        "summary": summary[:8000],
        "tomorrow_plan": tomorrow_plan[:4000],
        "support_needed": support_needed[:2000],
        "hours_worked": hours_worked,
        "tasks": tasks,
        "status": status,
    }, None


@app.get("/crm/daily-reports")
def crm_daily_reports_page() -> str:
    redir = _ensure_crm_session_html()
    if redir is not None:
        return redir
    staff_portal = _crm_staff_portal_active()
    if staff_portal:
        if not _admin_section_can("crm_daily_work_report", "view"):
            return redirect(url_for("crm_board"))
        return render_template(
            "crm_daily_reports.html",
            crm_staff_portal=True,
            crm_staff_id=_staff_session_id(),
            crm_staff_name=_staff_session_name(),
        )
    if not _admin_section_can("crm_daily_work_report", "view"):
        return redirect(url_for("crm_board"))
    return render_template(
        "crm_daily_reports.html",
        crm_staff_portal=False,
        crm_staff_id=None,
        crm_staff_name="",
        **_admin_page_template_kwargs(),
    )


@app.get("/api/crm/daily-work-reports")
def api_crm_list_daily_work_reports() -> Any:
    if not _admin_section_can("crm_daily_work_report", "view"):
        return _admin_section_forbidden_json("crm_daily_work_report", "view")
    portal_sid = _crm_effective_staff_id()
    staff_id: int | None = portal_sid
    if portal_sid is None:
        staff_raw = str(request.args.get("staff_id") or "").strip()
        if staff_raw:
            try:
                staff_id = int(staff_raw)
            except ValueError:
                return jsonify({"error": "staff_id không hợp lệ"}), 400
    date_from = validate_report_date(request.args.get("from"))
    date_to = validate_report_date(request.args.get("to"))
    if request.args.get("from") and not date_from:
        return jsonify({"error": "from phải là YYYY-MM-DD"}), 400
    if request.args.get("to") and not date_to:
        return jsonify({"error": "to phải là YYYY-MM-DD"}), 400
    try:
        limit = int(request.args.get("limit") or 100)
    except ValueError:
        limit = 100
    try:
        offset = int(request.args.get("offset") or 0)
    except ValueError:
        offset = 0
    with get_connection() as conn:
        rows = fetch_daily_work_reports(
            conn,
            staff_id=staff_id,
            date_from=date_from,
            date_to=date_to,
            limit=limit,
            offset=offset,
        )
    return jsonify({"reports": [daily_work_report_row_to_dict(r) for r in rows]})


@app.get("/api/crm/daily-work-reports/by-date")
def api_crm_daily_work_report_by_date() -> Any:
    if not _admin_section_can("crm_daily_work_report", "view"):
        return _admin_section_forbidden_json("crm_daily_work_report", "view")
    report_date = validate_report_date(request.args.get("report_date"))
    if not report_date:
        return jsonify({"error": "report_date phải là YYYY-MM-DD"}), 400
    staff_id, err = _crm_daily_report_staff_id_from_request()
    if err and _crm_effective_staff_id() is None:
        staff_raw = str(request.args.get("staff_id") or "").strip()
        if not staff_raw:
            return jsonify({"report": None})
        try:
            staff_id = int(staff_raw)
        except ValueError:
            return jsonify({"error": "staff_id không hợp lệ"}), 400
    if staff_id is None:
        return jsonify({"report": None})
    with get_connection() as conn:
        row = fetch_daily_work_report_by_staff_date(
            conn, staff_id=int(staff_id), report_date=report_date
        )
    if row is None:
        return jsonify({"report": None})
    return jsonify({"report": daily_work_report_row_to_dict(row)})


@app.get("/api/crm/daily-work-reports/<int:report_id>")
def api_crm_get_daily_work_report(report_id: int) -> Any:
    if not _admin_section_can("crm_daily_work_report", "view"):
        return _admin_section_forbidden_json("crm_daily_work_report", "view")
    with get_connection() as conn:
        row = fetch_daily_work_report_by_id(conn, report_id)
        if row is None:
            return jsonify({"error": "Không tìm thấy báo cáo"}), 404
        portal_sid = _crm_effective_staff_id()
        if portal_sid is not None and int(row["staff_id"]) != int(portal_sid):
            return jsonify({"error": "Không xem được báo cáo của người khác."}), 403
    return jsonify({"report": daily_work_report_row_to_dict(row)})


@app.post("/api/crm/daily-work-reports")
def api_crm_create_daily_work_report() -> Any:
    if not _admin_section_can("crm_daily_work_report", "create"):
        return _admin_section_forbidden_json("crm_daily_work_report", "create")
    payload = request.get_json(force=True) or {}
    parsed, perr = _crm_parse_daily_report_payload(payload)
    if perr:
        return jsonify({"error": perr}), 400
    assert parsed is not None
    staff_id, serr = _crm_daily_report_staff_id_from_request(payload)
    if serr:
        return jsonify({"error": serr}), 400
    assert staff_id is not None
    ts = _crm_ts()
    with get_connection() as conn:
        srow = conn.execute(
            "SELECT id FROM crm_staff WHERE id = ? AND COALESCE(active, 1) = 1",
            (staff_id,),
        ).fetchone()
        if srow is None:
            return jsonify({"error": "Nhân viên không tồn tại hoặc đã ngưng."}), 400
        row = upsert_daily_work_report(
            conn,
            staff_id=staff_id,
            report_date=parsed["report_date"],
            summary=parsed["summary"],
            tomorrow_plan=parsed["tomorrow_plan"],
            hours_worked=parsed["hours_worked"],
            support_needed=parsed["support_needed"],
            tasks=parsed["tasks"],
            status=parsed["status"],
            ts=ts,
        )
    return jsonify({"report": daily_work_report_row_to_dict(row)}), 201


@app.patch("/api/crm/daily-work-reports/<int:report_id>")
def api_crm_patch_daily_work_report(report_id: int) -> Any:
    if not _admin_section_can("crm_daily_work_report", "edit"):
        return _admin_section_forbidden_json("crm_daily_work_report", "edit")
    payload = request.get_json(force=True) or {}
    parsed, perr = _crm_parse_daily_report_payload(payload)
    if perr:
        return jsonify({"error": perr}), 400
    assert parsed is not None
    ts = _crm_ts()
    with get_connection() as conn:
        prev = fetch_daily_work_report_by_id(conn, report_id)
        if prev is None:
            return jsonify({"error": "Không tìm thấy báo cáo"}), 404
        portal_sid = _crm_effective_staff_id()
        if portal_sid is not None and int(prev["staff_id"]) != int(portal_sid):
            return jsonify({"error": "Chỉ sửa báo cáo của mình."}), 403
        staff_id: int | None = None
        if portal_sid is None and payload.get("staff_id") not in (None, "", 0, "0"):
            sid, serr = _crm_daily_report_staff_id_from_request(payload)
            if serr:
                return jsonify({"error": serr}), 400
            staff_id = sid
        try:
            row = update_daily_work_report_by_id(
                conn,
                report_id,
                staff_id=staff_id,
                report_date=parsed["report_date"],
                summary=parsed["summary"],
                tomorrow_plan=parsed["tomorrow_plan"],
                hours_worked=parsed["hours_worked"],
                support_needed=parsed["support_needed"],
                tasks=parsed["tasks"],
                status=parsed["status"],
                ts=ts,
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 409
    return jsonify({"report": daily_work_report_row_to_dict(row)})


def _crm_audit_user() -> str:
    if _crm_staff_portal_active():
        return _staff_session_name() or "Nhân viên"
    return _cms_session_username() or "Admin"


def _crm_leads_only_ui(conn: sqlite3.Connection | None = None) -> bool:
    """Menu thu gọn «Danh sách lead» — theo ma trận chức vụ, không cứng mã KD-01."""
    if _admin_full_access() or not _admin_logged_in():
        return False
    pid = _cms_session_position_id()
    if pid is None:
        return False

    def _check(c: sqlite3.Connection) -> bool:
        grants = _cms_load_position_grants(c, int(pid))
        return position_leads_only_ui(grants)

    if conn is not None:
        return _check(conn)
    with get_connection() as c:
        return _check(c)


def _crm_lead_session_staff_id(conn: sqlite3.Connection | None = None) -> int | None:
    """Staff portal hoặc NV KD đăng nhập admin — scope danh sách lead."""
    sid = _crm_effective_staff_id()
    if sid is not None:
        return sid
    if _admin_full_access() or not _admin_logged_in():
        return None
    pid = _cms_session_position_id()
    if pid is None:
        return None

    def _lookup(c: sqlite3.Connection) -> int | None:
        pos = c.execute(
            "SELECT code FROM crm_positions WHERE id = ? AND active = 1",
            (int(pid),),
        ).fetchone()
        if not pos or str(pos["code"] or "").strip().upper() != "KD-01":
            return None
        uname = _cms_session_username()
        if not uname:
            return None
        row = c.execute(
            """
            SELECT id FROM crm_staff
            WHERE lower(trim(login_username)) = lower(trim(?))
              AND trim(login_username) != ''
              AND active = 1
            LIMIT 1
            """,
            (uname,),
        ).fetchone()
        return int(row["id"]) if row else None

    if conn is not None:
        return _lookup(conn)
    with get_connection() as c:
        return _lookup(c)


def _crm_lead_owner_filter() -> int | None:
    """Portal NV / NV KD admin — owner filter (kết hợp scope dự án qua staff id)."""
    return _crm_lead_session_staff_id()


def _crm_presales_on_lead_enabled() -> bool:
    from crm_lead_presales import presales_on_lead_enabled

    return presales_on_lead_enabled()


def _crm_presales_stage_labels() -> dict[str, str]:
    return {"lead": "Lead", "consult": "Tư vấn", "proposal": "Báo giá"}


def _crm_lead_staff_portal_id(conn: sqlite3.Connection) -> int | None:
    """ID NV có scope lead (portal hoặc KD admin) + kiểm tra dự án tham gia."""
    sid = _crm_lead_session_staff_id(conn)
    if sid is None:
        return None
    if not fetch_staff_project_ids(conn, int(sid)):
        return int(sid)
    return int(sid)


def _crm_lead_list_kwargs(conn: sqlite3.Connection, *, owner_id: int | None) -> dict[str, Any]:
    """Tham số fetch/count lead — scope portal / KD theo dự án tham gia."""
    portal_sid = _crm_lead_session_staff_id(conn)
    if portal_sid is None:
        return {"owner_id": owner_id}
    if not fetch_staff_project_ids(conn, int(portal_sid)):
        return {"staff_portal_id": int(portal_sid)}
    return {"staff_portal_id": int(portal_sid)}


def _crm_lead_can_access(conn: sqlite3.Connection, row: sqlite3.Row | None) -> bool:
    if row is None:
        return False
    portal_sid = _crm_lead_session_staff_id(conn)
    if portal_sid is None:
        return True
    return staff_can_view_lead(conn, int(portal_sid), row)


def _crm_lead_project_filter_from_request() -> int | None | object:
    """Lọc lead theo dự án BĐS từ query/body — _UNSET nếu không lọc."""
    raw = request.args.get("re_project_id")
    if raw is None and request.is_json:
        body = request.get_json(silent=True) or {}
        if isinstance(body, dict) and "re_project_id" in body:
            raw = body.get("re_project_id")
    return parse_re_project_filter(raw)


def _crm_lead_ui_filters_from_request(*, owner_id: int | None) -> dict[str, Any]:
    """Bộ lọc UI danh sách lead — dùng chung list + stats."""
    raw_owner = str(request.args.get("owner_id") or "").strip()
    if owner_id is None and raw_owner:
        try:
            owner_id = int(raw_owner)
        except ValueError:
            owner_id = None
    sla_only = str(request.args.get("sla_overdue") or "").strip() in ("1", "true", "yes")
    review_only = str(request.args.get("review_queue") or "").strip() in ("1", "true", "yes")
    return {
        "owner_id": owner_id,
        "status": request.args.get("status") or None,
        "level": request.args.get("level") or None,
        "source": request.args.get("source") or None,
        "q": request.args.get("q") or None,
        "product_line": str(request.args.get("product_line") or "").strip() or None,
        "zone": str(request.args.get("zone") or "").strip() or None,
        "sla_overdue_only": sla_only,
        "review_queue_only": review_only,
        "hide_review_queue": not review_only,
    }


def _crm_ingest_lead_from_form(
    conn: sqlite3.Connection,
    *,
    full_name: str,
    phone: str,
    email: str,
    need: str,
    source: str,
    region: str = "",
    product_interest: str = "",
    utm_campaign: str = "",
    re_project_id: int | None = None,
    re_project_code: str | None = None,
    ingest_site: str = "",
    ts: str,
    _from_worker: bool = False,
) -> int | None:
    """Tạo lead CRM từ form/API ngoài — bỏ qua nếu trùng policy reject."""
    from crm_project_webhooks import resolve_project_for_lead_ingest

    try:
        pid = resolve_project_for_lead_ingest(
            conn,
            re_project_id=re_project_id,
            re_project_code=re_project_code,
            utm_campaign=utm_campaign,
            ingest_site=ingest_site,
        )
        ingest_meta: dict[str, Any] = {"ingest_channel": "website_form"}
        if ingest_site:
            ingest_meta["ingest_site"] = str(ingest_site).strip()[:120]
        row, _dups, _dup_matches = create_lead(
            conn,
            full_name=full_name,
            phone=phone,
            email=email,
            source=source,
            region=region,
            product_interest=product_interest,
            need=need,
            utm_campaign=utm_campaign,
            re_project_id=pid,
            meta=ingest_meta,
            auto_assign=True,
            duplicate_policy=None,
            created_by="system:ingest",
            ts=ts,
        )
        lead_id = int(row["id"])

        # Trigger AI qualify brief trong background (non-blocking)
        try:
            from crm_ai_qualify import trigger_qualify_brief_async
            from crm_re_projects import fetch_project

            project_name = ""
            if pid:
                proj = fetch_project(conn, pid)
                project_name = str(proj.get("name") or "") if proj else ""

            trigger_qualify_brief_async(
                lead_id,
                full_name=full_name,
                product_interest=product_interest,
                source=source,
                need=need,
                project_name=project_name,
                db_path=str(DB_PATH),
            )
        except Exception as _ai_exc:
            app.logger.debug("AI qualify trigger bỏ qua: %s", _ai_exc)

        return lead_id
    except ValueError as exc:
        app.logger.warning("CRM ingest lead from form failed: %s", exc)
        if not _from_worker:
            _enqueue_form_ingest_failure(
                full_name=full_name,
                phone=phone,
                email=email,
                need=need,
                source=source,
                error=str(exc),
            )
        return None
    except Exception as exc:
        app.logger.exception("CRM ingest lead from form error: %s", exc)
        if not _from_worker:
            _enqueue_form_ingest_failure(
                full_name=full_name,
                phone=phone,
                email=email,
                need=need,
                source=source,
                error=str(exc),
            )
        return None


def _enqueue_form_ingest_failure(**fields: Any) -> None:
    """Queue form ingest retry / spillover — không nuốt lỗi im lặng (P0-08)."""
    from ptt_jobs.form_ingest_failure import enqueue_form_ingest_failure

    enqueue_form_ingest_failure(**fields)


def _crm_leads_template_kwargs(
    conn: sqlite3.Connection,
    *,
    staff_portal: bool,
    leads_only: bool,
    lead_id: int | None = None,
    extra: dict[str, Any] | None = None,
) -> dict[str, Any]:
    tiers = fetch_level_tiers(conn)
    tier_ids = [str(t["id"]) for t in tiers if t.get("enabled", True)]
    page_levels = list(dict.fromkeys([*tier_ids, UNCLASSIFIED_TIER_ID, *LEAD_LEVELS]))
    page_level_labels = {**LEAD_LEVEL_LABELS, **level_labels_map(conn, tiers)}
    from crm_lead_catalog import catalog_public_payload

    catalog = catalog_public_payload(conn)
    kw: dict[str, Any] = {
        "crm_staff_portal": staff_portal,
        "crm_leads_only_ui": leads_only,
        "crm_staff_id": _staff_session_id() if staff_portal else (_crm_lead_session_staff_id() if leads_only else None),
        "crm_lead_id": lead_id,
        "crm_lead_statuses": list(LEAD_STATUSES),
        "crm_lead_status_labels": LEAD_STATUS_LABELS,
        "crm_lead_levels": page_levels,
        "crm_lead_level_labels": page_level_labels,
        "crm_lead_sources": list(LEAD_SOURCES),
        "crm_lead_source_labels": LEAD_SOURCE_LABELS,
        "crm_lead_activity_types": list(ACTIVITY_TYPES),
        "crm_lead_activity_labels": ACTIVITY_TYPE_LABELS,
        "crm_scoring_rule_defaults": DEFAULT_SCORING_RULES,
        "crm_scoring_rubric_defaults": DEFAULT_LEAD_SCORING_RUBRIC,
        "crm_level_tier_defaults": DEFAULT_LEVEL_TIERS,
        "crm_assign_strategy_defaults": STRATEGY_DEFS,
        "crm_assign_tier_level_map_defaults": DEFAULT_TIER_LEVEL_MAP,
        "crm_scoring_evaluators": EVALUATOR_OPTIONS,
        "crm_scoring_conditions": SCORING_CONDITIONS,
        "crm_scoring_field_options": list(SCORING_FIELD_OPTIONS),
        "crm_leads_can_configure": (
            False if staff_portal else _admin_section_can("crm_leads", "configure")
        ),
        "crm_leads_can_delete": (
            False if staff_portal else _admin_section_can("crm_leads", "delete")
        ),
        "crm_product_line_labels": PRODUCT_LINE_LABELS,
        "crm_care_contact_types": list(CRM_CARE_CONTACT_TYPES),
        "crm_care_contact_labels": CRM_CARE_CONTACT_LABELS_VI,
        "crm_care_status_types": list(CRM_CARE_STATUS_TYPES),
        "crm_care_status_labels": CRM_CARE_STATUS_LABELS_VI,
        "crm_care_pipeline_stages": CARE_PIPELINE_STAGES_PUBLIC,
        "crm_presales_on_lead": _crm_presales_on_lead_enabled(),
        "crm_presales_stage_labels": _crm_presales_stage_labels(),
        "crm_service_slugs": catalog.get("service_slugs") or sorted(SVC_LIFECYCLE_SLUGS),
        "crm_service_labels": catalog.get("service_labels") or {},
        "crm_industry_slugs": catalog.get("industry_slugs") or [],
        "crm_industry_labels": catalog.get("industry_labels") or {},
        "crm_catalog_services": catalog.get("services") or [],
        "crm_catalog_industries": catalog.get("industries") or [],
        "crm_hide_re_lead_fields": True,
        "crm_leads_can_manage_review_queue": (
            False if staff_portal else _admin_section_can("crm_leads", "configure")
        ),
    }
    if extra:
        kw.update(extra)
    return kw


@app.get("/crm/leads")
def crm_leads_page() -> str:
    redir = _ensure_crm_session_html()
    if redir is not None:
        return redir
    staff_portal = _crm_staff_portal_active()
    with get_connection() as conn:
        if staff_portal:
            if not _admin_section_can("crm_leads", "view"):
                return redirect(url_for("crm_board"))
            return render_template(
                "crm_leads.html",
                **_crm_leads_template_kwargs(conn, staff_portal=True, leads_only=False),
            )
        if not _admin_section_can("crm_leads", "view"):
            return redirect(url_for("crm_board"))
        leads_only = _crm_leads_only_ui()
        return render_template(
            "crm_leads.html",
            **{
                **_admin_page_template_kwargs(),
                **_crm_leads_template_kwargs(conn, staff_portal=False, leads_only=leads_only),
            },
        )


@app.get("/crm/leads/<int:lead_id>")
def crm_lead_detail_page(lead_id: int) -> str:
    redir = _ensure_crm_session_html()
    if redir is not None:
        return redir
    staff_portal = _crm_staff_portal_active()
    with get_connection() as conn:
        row = fetch_lead_by_id(conn, lead_id)
        if row is None:
            abort(404)
        if not _crm_lead_can_access(conn, row):
            abort(403)
        leads_only = _crm_leads_only_ui() and not staff_portal
        if staff_portal:
            if not _admin_section_can("crm_leads", "view"):
                return redirect(url_for("crm_board"))
            return render_template(
                "crm_lead_detail.html",
                **_crm_leads_template_kwargs(
                    conn, staff_portal=True, leads_only=False, lead_id=lead_id
                ),
            )
        if not _admin_section_can("crm_leads", "view"):
            return redirect(url_for("crm_board"))
        return render_template(
            "crm_lead_detail.html",
            **{
                **_admin_page_template_kwargs(),
                **_crm_leads_template_kwargs(
                    conn, staff_portal=False, leads_only=leads_only, lead_id=lead_id
                ),
            },
        )


@app.get("/api/crm/leads/stats")
def api_crm_leads_stats() -> Any:
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    owner_id = _crm_lead_owner_filter()
    try:
        re_project_id = _crm_lead_project_filter_from_request()
    except ValueError:
        return jsonify({"error": "re_project_id không hợp lệ"}), 400
    ui_filters = _crm_lead_ui_filters_from_request(owner_id=owner_id)
    with get_connection() as conn:
        stats = fetch_lead_stats_extended(
            conn,
            **_crm_lead_list_kwargs(conn, owner_id=ui_filters["owner_id"]),
            re_project_id=re_project_id,
            status=ui_filters["status"],
            level=ui_filters["level"],
            source=ui_filters["source"],
            q=ui_filters["q"],
            product_line=ui_filters["product_line"],
            zone=ui_filters["zone"],
            sla_overdue_only=ui_filters["sla_overdue_only"],
            hide_review_queue=ui_filters["hide_review_queue"],
            review_queue_only=ui_filters["review_queue_only"],
            ts=_crm_ts(),
        )
    return jsonify({"stats": stats})


@app.get("/api/crm/leads/projects/<int:project_id>/staff")
def api_crm_leads_project_staff(project_id: int) -> Any:
    """NV tham gia dự án — dropdown phân lead."""
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    with get_connection() as conn:
        try:
            staff = list_assignable_staff_for_project(conn, project_id)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    return jsonify({"project_id": project_id, "staff": staff})


@app.get("/api/crm/leads/projects")
def api_crm_leads_projects() -> Any:
    """Danh sách dự án BĐS cho dropdown lọc / gán lead."""
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    q = str(request.args.get("q") or "").strip()
    with get_connection() as conn:
        portal_sid = _crm_lead_session_staff_id(conn)
        if portal_sid is not None:
            projects = list_lead_project_options_for_staff(conn, int(portal_sid), q=q)
        else:
            projects = list_lead_project_options(conn, q=q)
    return jsonify({"projects": projects})


@app.get("/api/crm/leads")
def api_crm_list_leads() -> Any:
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    owner_id = _crm_lead_owner_filter()
    if owner_id is None:
        raw_owner = str(request.args.get("owner_id") or "").strip()
        if raw_owner:
            try:
                owner_id = int(raw_owner)
            except ValueError:
                return jsonify({"error": "owner_id không hợp lệ"}), 400
    sla_only = str(request.args.get("sla_overdue") or "").strip() in ("1", "true", "yes")
    review_only = str(request.args.get("review_queue") or "").strip() in ("1", "true", "yes")
    try:
        lim = max(1, min(int(request.args.get("limit") or 500), 1000))
    except ValueError:
        return jsonify({"error": "limit không hợp lệ"}), 400
    try:
        off = max(0, int(request.args.get("offset") or 0))
    except ValueError:
        return jsonify({"error": "offset không hợp lệ"}), 400
    try:
        re_project_id = _crm_lead_project_filter_from_request()
    except ValueError:
        return jsonify({"error": "re_project_id không hợp lệ"}), 400
    ui_filters = _crm_lead_ui_filters_from_request(owner_id=owner_id)
    filt = {
        "status": ui_filters["status"],
        "level": ui_filters["level"],
        "source": ui_filters["source"],
        "q": ui_filters["q"],
        "product_line": ui_filters["product_line"],
        "zone": ui_filters["zone"],
        "hide_review_queue": ui_filters["hide_review_queue"],
        "review_queue_only": ui_filters["review_queue_only"],
    }
    if re_project_id is not _UNSET:
        filt["re_project_id"] = re_project_id
    with get_connection() as conn:
        if ui_filters["review_queue_only"] and _crm_lead_session_staff_id(conn) is not None:
            return jsonify({"error": "Không có quyền xem Lead Phải tra soát."}), 403
        filt.update(_crm_lead_list_kwargs(conn, owner_id=ui_filters["owner_id"]))
        total = count_leads(conn, **filt)
        rows = fetch_leads(
            conn,
            **filt,
            sla_overdue_only=ui_filters["sla_overdue_only"],
            limit=lim,
            offset=off,
        )
        shown = len(rows)
        return jsonify(
            {
                "leads": [lead_row_to_dict(r, conn) for r in rows],
                "total": total,
                "limit": lim,
                "offset": off,
                "truncated": (off + shown) < total,
            }
        )


@app.get("/api/crm/leads/notifications")
def api_crm_leads_notifications() -> Any:
    """Poll lead mới đã phân công — toast + âm thanh trên trang Quản lý Lead."""
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    bootstrap = str(request.args.get("bootstrap") or "").strip().lower() in ("1", "true", "yes")
    after_id = 0
    if not bootstrap:
        try:
            after_id = max(0, int(request.args.get("after_id") or 0))
        except ValueError:
            return jsonify({"error": "after_id không hợp lệ"}), 400
    try:
        client_revision = max(0, int(request.args.get("list_revision") or 0))
    except ValueError:
        client_revision = 0
    owner_id = _crm_lead_owner_filter()
    with get_connection() as conn:
        pending_facebook = 0
        try:
            from crm_facebook_pending import ensure_facebook_pending_schema, list_pending_facebook_leadgens

            ensure_facebook_pending_schema(conn)
            pending_facebook = len(list_pending_facebook_leadgens(conn, limit=100))
        except Exception:
            pending_facebook = 0
        if bootstrap:
            max_id = fetch_max_lead_id_any(conn)
            live = fetch_leads_live_revision(conn)
            return jsonify(
                {
                    "leads": [],
                    "after_id": max_id,
                    "count": 0,
                    "pending_facebook": pending_facebook,
                    "list_revision": live.get("revision", 0),
                    "webhook_at": live.get("webhook_at", ""),
                    "webhook_message": live.get("webhook_message", ""),
                    "webhook_created": live.get("webhook_created", 0),
                }
            )
        rows = fetch_new_assigned_leads(
            conn,
            after_id=after_id,
            limit=30,
            **_crm_lead_list_kwargs(conn, owner_id=owner_id),
        )
        leads_out = [lead_row_to_dict(r, conn) for r in rows]
        live = fetch_leads_live_revision(conn)
        next_id = after_id
        for item in leads_out:
            next_id = max(next_id, int(item.get("id") or 0))
        max_id = fetch_max_lead_id_any(conn)
    should_refresh = (
        max_id > after_id
        or bool(leads_out)
        or int(live.get("revision") or 0) > client_revision
    )
    return jsonify(
        {
            "leads": leads_out,
            "after_id": next_id,
            "max_id": max_id,
            "count": len(leads_out),
            "pending_facebook": pending_facebook,
            "should_refresh": should_refresh,
            "list_revision": live.get("revision", 0),
            "webhook_at": live.get("webhook_at", ""),
            "webhook_message": live.get("webhook_message", ""),
            "webhook_created": live.get("webhook_created", 0),
        }
    )


@app.get("/api/crm/leads/<int:lead_id>")
def api_crm_get_lead(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    with get_connection() as conn:
        row = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, row):
            return jsonify({"error": "Không có quyền xem lead này."}), 403
        if row is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        out = lead_row_to_dict(row, conn)
        if not out.get("score_breakdown"):
            ts = _crm_ts()
            apply_lead_score(conn, lead_id, updated_by=_crm_audit_user(), ts=ts)
            row = fetch_lead_by_id(conn, lead_id)
            if row is not None:
                out = lead_row_to_dict(row, conn)
        out["allowed_transitions"] = allowed_next_statuses(str(row["status"]))
        dups = fetch_lead_duplicates(conn, lead_id)
        out["duplicate_count"] = len(dups)
        if dups:
            out["duplicates"] = [lead_row_to_dict(d, conn) for d in dups[:10]]
    return jsonify({"lead": out})


@app.post("/api/crm/leads")
def api_crm_create_lead() -> Any:
    if not _admin_section_can("crm_leads", "create"):
        return _admin_section_forbidden_json("crm_leads", "create")
    payload = request.get_json(force=True) or {}
    industry_raw = str(payload.get("industry_slug") or "").strip()
    if not industry_raw:
        return jsonify({"error": "Ngành khách hàng bắt buộc — chọn từ danh mục."}), 400
    ts = _crm_ts()
    actor = _crm_audit_user()
    portal_sid = _crm_effective_staff_id()
    owner_raw = payload.get("owner_id")
    owner_id: int | None = int(portal_sid) if portal_sid is not None else None
    if portal_sid is None and owner_raw not in (None, "", 0, "0"):
        try:
            owner_id = int(owner_raw)
        except (TypeError, ValueError):
            return jsonify({"error": "owner_id không hợp lệ"}), 400
    with get_connection() as conn:
        try:
            re_project_id = parse_re_project_id(payload.get("re_project_id"))
            if re_project_id is not None:
                re_project_id = None  # P3 — bỏ gán dự án RE trên funnel lead
            if portal_sid is not None:
                assert_staff_portal_project(conn, int(portal_sid), re_project_id)
            row, dups, dup_matches = create_lead(
                conn,
                full_name=str(payload.get("full_name") or ""),
                phone=str(payload.get("phone") or ""),
                email=str(payload.get("email") or ""),
                source=str(payload.get("source") or "manual"),
                region=str(payload.get("region") or ""),
                product_interest=str(payload.get("product_interest") or ""),
                industry_slug=str(payload.get("industry_slug") or ""),
                need=str(payload.get("need") or ""),
                status=str(payload.get("status") or "new"),
                owner_id=owner_id,
                re_project_id=re_project_id,
                product_line=normalize_product_line(str(payload.get("product_line") or "")),
                zone=str(payload.get("zone") or "").strip(),
                re_product_id=int(payload["re_product_id"]) if payload.get("re_product_id") not in (None, "", 0, "0") else None,
                utm_campaign=str(payload.get("utm_campaign") or ""),
                meta=payload.get("meta") if isinstance(payload.get("meta"), dict) else None,
                auto_assign=owner_id is None,
                duplicate_policy=(
                    str(payload["duplicate_policy"])
                    if payload.get("duplicate_policy") not in (None, "")
                    else None
                ),
                created_by=actor,
                ts=ts,
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        out = lead_row_to_dict(row, conn)
        resp: dict[str, Any] = {
            "lead": out,
            "create_summary": {
                "score": out.get("lead_score"),
                "lead_level": out.get("lead_level"),
                "lead_level_label": out.get("lead_level_label"),
                "duplicate_check": "found" if dup_matches else "clear",
            },
        }
        if dup_matches:
            resp["duplicates"] = [
                {
                    "id": m["lead_id"],
                    "full_name": m["full_name"],
                    "phone": m.get("phone") or "",
                    "email": m.get("email") or "",
                    "match_type": m["match_type"],
                    "match_label": m["match_label"],
                }
                for m in dup_matches[:5]
            ]
        elif dups:
            resp["duplicates"] = [lead_row_to_dict(d, conn) for d in dups[:5]]
    return jsonify(resp), 201


@app.post("/api/crm/leads/validate")
def api_crm_validate_lead() -> Any:
    """Kiểm tra SĐT/email và trùng lead trước khi tạo."""
    if not _admin_section_can("crm_leads", "create"):
        return _admin_section_forbidden_json("crm_leads", "create")
    payload = request.get_json(force=True) or {}
    phone = str(payload.get("phone") or "")
    email = str(payload.get("email") or "")
    errors: list[str] = []
    ph_norm = ""
    em_norm = ""
    try:
        ph_norm, em_norm = validate_lead_contacts(phone=phone, email=email)
    except ValueError as exc:
        errors.append(str(exc))
    dup_matches: list[dict[str, Any]] = []
    re_project_id = parse_re_project_id(payload.get("re_project_id"))
    if not errors:
        with get_connection() as conn:
            dup_matches = find_duplicate_matches(
                conn,
                phone=phone,
                email=email,
                re_project_id=re_project_id,
            )
    return jsonify(
        {
            "valid": not errors,
            "errors": errors,
            "validated": {"phone": bool(ph_norm), "email": bool(em_norm)},
            "duplicates": [
                {
                    "id": m["lead_id"],
                    "full_name": m["full_name"],
                    "match_type": m["match_type"],
                    "match_label": m["match_label"],
                }
                for m in dup_matches[:5]
            ],
        }
    )


@app.put("/api/crm/leads/<int:lead_id>")
def api_crm_update_lead(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "edit"):
        return _admin_section_forbidden_json("crm_leads", "edit")
    payload = request.get_json(force=True) or {}
    ts = _crm_ts()
    actor = _crm_audit_user()
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền sửa lead này."}), 403
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        owner_id = payload.get("owner_id")
        if _crm_effective_staff_id() is not None:
            owner_id = None
        status_override = bool(payload.get("status_override")) and _admin_section_can(
            "crm_leads", "delete"
        )
        re_project_kw: dict[str, Any] = {}
        segment_kw: dict[str, Any] = {}
        if "re_project_id" in payload:
            new_pid = parse_re_project_id(payload.get("re_project_id"))
            re_project_kw["re_project_id"] = new_pid
            if _crm_effective_staff_id() is not None:
                assert_staff_portal_project(conn, int(_crm_effective_staff_id()), new_pid)
        if "product_line" in payload:
            segment_kw["product_line"] = normalize_product_line(str(payload.get("product_line") or ""))
        if "zone" in payload:
            segment_kw["zone"] = str(payload.get("zone") or "").strip()
        if "re_product_id" in payload:
            raw_pid = payload.get("re_product_id")
            segment_kw["re_product_id"] = (
                int(raw_pid) if raw_pid not in (None, "", 0, "0") else None
            )
        try:
            row = update_lead(
                conn,
                lead_id,
                full_name=payload.get("full_name"),
                phone=payload.get("phone"),
                email=payload.get("email"),
                source=payload.get("source"),
                region=payload.get("region"),
                product_interest=payload.get("product_interest"),
                industry_slug=payload.get("industry_slug"),
                need=payload.get("need"),
                status=payload.get("status"),
                owner_id=int(owner_id) if owner_id not in (None, "", 0, "0") else None,
                lead_level=payload.get("lead_level"),
                updated_by=actor,
                ts=ts,
                status_note=str(payload.get("status_note") or ""),
                status_override=status_override,
                **re_project_kw,
                **segment_kw,
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        out = lead_row_to_dict(row, conn)
    return jsonify({"lead": out})


@app.delete("/api/crm/leads/<int:lead_id>")
def api_crm_delete_lead(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "delete"):
        return _admin_section_forbidden_json("crm_leads", "delete")
    payload = request.get_json(silent=True) or {}
    force = str(payload.get("force") or "").strip().lower() in ("1", "true", "yes")
    ts = _crm_ts()
    actor = _crm_audit_user()
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        if _crm_effective_staff_id() is not None:
            return jsonify({"error": "Nhân viên không có quyền xóa lead."}), 403
        try:
            result = delete_lead(conn, lead_id, deleted_by=actor, force=force)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        conn.commit()
    result["deleted_at"] = ts
    return jsonify({"ok": True, **result})


@app.post("/api/crm/leads/<int:lead_id>/activities")
def api_crm_create_lead_activity(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "edit"):
        return _admin_section_forbidden_json("crm_leads", "edit")
    payload = request.get_json(force=True) or {}
    ts = _crm_ts()
    actor = _crm_audit_user()
    user_id = _crm_effective_staff_id()
    raw_contact = payload.get("contact_type")
    raw_care_status = payload.get("care_status")
    activity_type = str(payload.get("activity_type") or "").strip().lower()
    if raw_contact:
        ct = normalize_care_contact(str(raw_contact))
        activity_type = CRM_CARE_CONTACT_TO_ACTIVITY.get(ct, "note")
    elif not activity_type:
        activity_type = "note"
    content = str(payload.get("content") or payload.get("summary") or "")
    care_stage_key = str(payload.get("care_stage_key") or "").strip()
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        is_care_report = bool(raw_contact or payload.get("summary"))
        if is_care_report and activity_type != "note":
            from crm_lead_care_pipeline import CARE_STAGE_KEYS, care_stage_label

            current_stage = str(prev["care_stage_current"] or "first_contact").strip()
            if not care_stage_key:
                care_stage_key = current_stage
            if care_stage_key not in CARE_STAGE_KEYS:
                return jsonify({"error": "Bước chăm sóc không hợp lệ."}), 400
            if care_stage_key != current_stage:
                return (
                    jsonify(
                        {
                            "error": f"Báo cáo phải ghi cho bước đang thực hiện: {care_stage_label(current_stage)}.",
                        }
                    ),
                    400,
                )
        row = log_lead_activity(
            conn,
            lead_id=lead_id,
            activity_type=activity_type,
            content=content,
            result=str(payload.get("result") or ""),
            next_action=str(payload.get("next_action") or ""),
            next_action_at=str(payload.get("next_action_at") or ""),
            care_contact_type=str(raw_contact or ""),
            care_status=str(raw_care_status or ""),
            care_stage_key=care_stage_key,
            user_id=int(user_id) if user_id else None,
            created_by=actor,
            ts=ts,
        )
        lead_row = fetch_lead_by_id(conn, lead_id)
    return jsonify(
        {
            "activity": activity_row_to_dict(row),
            "lead": lead_row_to_dict(lead_row, conn) if lead_row else None,
        }
    ), 201


@app.post("/api/crm/leads/<int:lead_id>/care-stages")
def api_crm_complete_lead_care_stage(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "edit"):
        return _admin_section_forbidden_json("crm_leads", "edit")
    payload = request.get_json(force=True) or {}
    action = str(payload.get("action") or "complete").strip().lower()
    if action != "complete":
        return jsonify({"error": "Hành động không hỗ trợ."}), 400
    stage_key = str(payload.get("stage_key") or "").strip()
    note = str(payload.get("note") or "").strip()
    ts = _crm_ts()
    actor = _crm_audit_user()
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        try:
            row = complete_lead_care_stage(
                conn,
                lead_id=lead_id,
                stage_key=stage_key,
                note=note,
                created_by=actor,
                ts=ts,
            )
            conn.commit()
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    return jsonify(
        {
            "ok": True,
            "lead": lead_row_to_dict(row, conn),
        }
    )


@app.post("/api/crm/leads/<int:lead_id>/rescore")
def api_crm_rescore_lead(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "edit"):
        return _admin_section_forbidden_json("crm_leads", "edit")
    ts = _crm_ts()
    actor = _crm_audit_user()
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        try:
            score_result = apply_lead_score(conn, lead_id, updated_by=actor, ts=ts)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        row = fetch_lead_by_id(conn, lead_id)
        out = lead_row_to_dict(row, conn) if row else None
    return jsonify({"lead": out, "score_result": score_result})


@app.get("/api/crm/leads/<int:lead_id>/activities")
def api_crm_list_lead_activities(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        rows = fetch_lead_activities(conn, lead_id, limit=int(request.args.get("limit") or 100))
    return jsonify({"activities": [activity_row_to_dict(r) for r in rows]})


@app.get("/api/crm/leads/<int:lead_id>/status-logs")
def api_crm_lead_status_logs(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        logs = fetch_lead_status_logs(conn, lead_id)
    return jsonify({"status_logs": logs})


@app.get("/api/crm/leads/<int:lead_id>/assignment-logs")
def api_crm_lead_assignment_logs(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        logs = fetch_lead_assignment_logs(conn, lead_id)
    return jsonify({"assignment_logs": logs})


@app.get("/api/crm/leads/<int:lead_id>/audit")
def api_crm_lead_audit(lead_id: int) -> Any:
    """Truy xuất audit — status + assignment logs."""
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        status_logs = fetch_lead_status_logs(conn, lead_id)
        assignment_logs = fetch_lead_assignment_logs(conn, lead_id)
    return jsonify({"status_logs": status_logs, "assignment_logs": assignment_logs})


@app.post("/api/crm/leads/<int:lead_id>/assign")
def api_crm_assign_lead(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "edit"):
        return _admin_section_forbidden_json("crm_leads", "edit")
    if _crm_effective_staff_id() is not None:
        return jsonify({"error": "Chỉ quản lý mới phân lại lead."}), 403
    payload = request.get_json(force=True) or {}
    try:
        to_id = int(payload.get("to_user_id") or payload.get("owner_id") or 0)
    except (TypeError, ValueError):
        return jsonify({"error": "to_user_id không hợp lệ"}), 400
    reason = str(payload.get("reason") or "").strip()
    if not reason:
        return jsonify({"error": "Cần ghi lý do phân lại."}), 400
    ts = _crm_ts()
    assigned_by = _crm_audit_user()

    from ptt_crm.leads_write_upstream import nest_write_upstream_enabled, proxy_assign_lead

    if nest_write_upstream_enabled():
        body, status = proxy_assign_lead(
            lead_id,
            to_user_id=to_id,
            reason=reason,
            assigned_by=assigned_by,
            ts=ts,
        )
        return jsonify(body), status

    with get_connection() as conn:
        try:
            row = assign_lead(
                conn,
                lead_id,
                to_user_id=to_id,
                reason=reason,
                assigned_by=assigned_by,
                ts=ts,
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        out = lead_row_to_dict(row, conn)
    return jsonify({"lead": out})


@app.get("/api/crm/leads/config")
def api_crm_leads_config_get() -> Any:
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    with get_connection() as conn:
        cfg = fetch_lead_config(conn)
    return jsonify(
        {
            "config": cfg,
            "duplicate_policies": list(DUPLICATE_POLICIES),
            "scoring_rule_defaults": DEFAULT_SCORING_RULES,
            "scoring_rubric_defaults": default_scoring_rubric(),
            "scoring_evaluators": EVALUATOR_OPTIONS,
            "scoring_conditions": SCORING_CONDITIONS,
            "scoring_field_options": list(SCORING_FIELD_OPTIONS),
            "level_tier_defaults": DEFAULT_LEVEL_TIERS,
            "assign_strategy_defaults": STRATEGY_DEFS,
            "assign_tier_level_map_defaults": DEFAULT_TIER_LEVEL_MAP,
            "facebook_config_defaults": DEFAULT_FACEBOOK_CONFIG,
            "can_configure": _admin_section_can("crm_leads", "configure"),
        }
    )


@app.put("/api/crm/leads/config")
def api_crm_leads_config_put() -> Any:
    if not _admin_section_can("crm_leads", "configure"):
        return _admin_section_forbidden_json("crm_leads", "configure")
    payload = request.get_json(force=True) or {}
    ts = _crm_ts()
    actor = _crm_audit_user()
    with get_connection() as conn:
        try:
            cfg = save_lead_config(conn, config=payload, updated_by=actor, ts=ts)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    return jsonify({"config": cfg})


@app.get("/api/crm/leads/rules/transitions")
def api_crm_leads_transitions() -> Any:
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    return jsonify({"transitions": transitions_for_ui()})


@app.get("/api/crm/leads/<int:lead_id>/duplicates")
def api_crm_lead_duplicates(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        dups = fetch_lead_duplicates(conn, lead_id)
        dup_out = [lead_row_to_dict(d, conn) for d in dups]
    return jsonify({"duplicates": dup_out})


@app.post("/api/crm/leads/<int:lead_id>/merge")
def api_crm_lead_merge(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "delete"):
        return _admin_section_forbidden_json("crm_leads", "delete")
    payload = request.get_json(force=True) or {}
    raw_ids = payload.get("duplicate_ids") or payload.get("merge_ids") or []
    if not isinstance(raw_ids, list):
        return jsonify({"error": "duplicate_ids phải là mảng."}), 400
    try:
        dup_ids = [int(x) for x in raw_ids]
    except (TypeError, ValueError):
        return jsonify({"error": "duplicate_ids không hợp lệ."}), 400
    reason = str(payload.get("reason") or "").strip()
    ts = _crm_ts()
    actor = _crm_audit_user()
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        try:
            row = merge_leads(
                conn,
                lead_id,
                dup_ids,
                merged_by=actor,
                ts=ts,
                reason=reason,
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        out = lead_row_to_dict(row, conn)
    return jsonify({"lead": out, "merged_count": len(dup_ids)})


@app.post("/api/crm/leads/import")
def api_crm_import_leads() -> Any:
    if not _admin_section_can("crm_leads", "create"):
        return _admin_section_forbidden_json("crm_leads", "create")
    payload = request.get_json(force=True) or {}
    csv_text = str(payload.get("csv") or "")
    if not csv_text.strip():
        return jsonify({"error": "Thiếu dữ liệu CSV."}), 400
    reader = csv.DictReader(StringIO(csv_text))
    ts = _crm_ts()
    actor = _crm_audit_user()
    created = 0
    skipped = 0
    errors: list[str] = []
    with get_connection() as conn:
        for i, row in enumerate(reader, start=2):
            name = str(row.get("full_name") or row.get("name") or row.get("ho_ten") or "").strip()
            phone = str(row.get("phone") or row.get("sdt") or "").strip()
            email = str(row.get("email") or "").strip()
            if not name:
                skipped += 1
                continue
            try:
                create_lead(
                    conn,
                    full_name=name,
                    phone=phone,
                    email=email,
                    source=str(row.get("source") or "import"),
                    region=str(row.get("region") or row.get("khu_vuc") or ""),
                    product_interest=str(row.get("product_interest") or row.get("san_pham") or ""),
                    need=str(row.get("need") or row.get("nhu_cau") or ""),
                    auto_assign=True,
                    duplicate_policy=None,
                    created_by=actor,
                    ts=ts,
                )
                created += 1
            except ValueError as exc:
                skipped += 1
                if len(errors) < 5:
                    errors.append(f"Dòng {i}: {exc}")
    return jsonify({"created": created, "skipped": skipped, "errors": errors})


@app.post("/api/crm/leads/ai/search")
def api_crm_leads_ai_search() -> Any:
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    payload = request.get_json(force=True) or {}
    question = str(payload.get("question") or "").strip()
    if not question:
        return jsonify({"error": "Thiếu câu hỏi."}), 400
    try:
        re_project_id = parse_re_project_filter(payload.get("re_project_id"))
    except ValueError:
        return jsonify({"error": "re_project_id không hợp lệ"}), 400
    with get_connection() as conn:
        scope = _crm_lead_list_kwargs(conn, owner_id=_crm_lead_owner_filter())
        portal_sid = scope.get("staff_portal_id")
        out = ai_search_leads(
            conn,
            question,
            owner_id=_crm_lead_owner_filter() if portal_sid is None else None,
            staff_portal_id=int(portal_sid) if portal_sid is not None else None,
            re_project_id=re_project_id,
            created_by=_crm_audit_user(),
            ts=_crm_ts(),
        )
    return jsonify(out)


@app.post("/api/crm/leads/ai/summary")
def api_crm_leads_ai_summary() -> Any:
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    payload = request.get_json(force=True) or {}
    try:
        lead_id = int(payload.get("lead_id") or 0)
    except (TypeError, ValueError):
        return jsonify({"error": "lead_id không hợp lệ"}), 400
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        out = ai_summarize_lead(conn, lead_id, created_by=_crm_audit_user(), ts=_crm_ts())
    return jsonify(out)


@app.post("/api/crm/leads/ai/recommend")
def api_crm_leads_ai_recommend() -> Any:
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    payload = request.get_json(force=True) or {}
    lead_id_raw = payload.get("lead_id")
    lead_id: int | None = None
    if lead_id_raw not in (None, "", 0, "0"):
        try:
            lead_id = int(lead_id_raw)
        except (TypeError, ValueError):
            return jsonify({"error": "lead_id không hợp lệ"}), 400
    try:
        re_project_id = parse_re_project_filter(payload.get("re_project_id"))
    except ValueError:
        return jsonify({"error": "re_project_id không hợp lệ"}), 400
    with get_connection() as conn:
        if lead_id is not None:
            prev = fetch_lead_by_id(conn, lead_id)
            if not _crm_lead_can_access(conn, prev):
                return jsonify({"error": "Không có quyền."}), 403
        scope = _crm_lead_list_kwargs(conn, owner_id=_crm_lead_owner_filter())
        portal_sid = scope.get("staff_portal_id")
        out = ai_recommend_lead(
            conn,
            lead_id,
            owner_id=_crm_lead_owner_filter() if portal_sid is None else None,
            staff_portal_id=int(portal_sid) if portal_sid is not None else None,
            re_project_id=re_project_id,
            created_by=_crm_audit_user(),
            ts=_crm_ts(),
        )
    return jsonify(out)


@app.post("/api/crm/leads/ai/suggest-products")
def api_crm_leads_ai_suggest_products() -> Any:
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    payload = request.get_json(force=True) or {}
    try:
        lead_id = int(payload.get("lead_id") or 0)
    except (TypeError, ValueError):
        return jsonify({"error": "lead_id không hợp lệ"}), 400
    if lead_id <= 0:
        return jsonify({"error": "Thiếu lead_id."}), 400
    try:
        limit = max(1, min(int(payload.get("limit") or 5), 20))
    except (TypeError, ValueError):
        limit = 5
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        out = ai_suggest_products_for_lead(
            conn,
            lead_id,
            limit=limit,
            created_by=_crm_audit_user(),
            ts=_crm_ts(),
        )
    return jsonify(out)


@app.post("/api/crm/leads/ai/price-list")
def api_crm_leads_ai_price_list() -> Any:
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    payload = request.get_json(force=True) or {}
    question = str(payload.get("question") or "").strip()
    if not question:
        return jsonify({"error": "Thiếu câu hỏi."}), 400
    try:
        re_project_id = parse_re_project_filter(payload.get("re_project_id"))
    except ValueError:
        return jsonify({"error": "re_project_id không hợp lệ"}), 400
    if not isinstance(re_project_id, int):
        return jsonify({"error": "Cần chọn dự án BĐS để tra cứu bảng giá."}), 400
    with get_connection() as conn:
        out = ai_price_list_query(
            conn,
            question,
            re_project_id=int(re_project_id),
            created_by=_crm_audit_user(),
            ts=_crm_ts(),
        )
    return jsonify(out)


@app.post("/api/crm/leads/ai/classify")
def api_crm_leads_ai_classify() -> Any:
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    payload = request.get_json(force=True) or {}
    try:
        lead_id = int(payload.get("lead_id") or 0)
    except (TypeError, ValueError):
        return jsonify({"error": "lead_id không hợp lệ"}), 400
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        out = ai_classify_suggestion(conn, lead_id, created_by=_crm_audit_user(), ts=_crm_ts())
    return jsonify(out)


@app.post("/api/crm/leads/<int:lead_id>/convert")
def api_crm_convert_lead(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "edit"):
        return _admin_section_forbidden_json("crm_leads", "edit")
    payload = request.get_json(silent=True) or {}
    ts = _crm_ts()
    actor = _crm_audit_user()
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        try:
            result = convert_lead_to_crm(
                conn,
                lead_id,
                case_title=str(payload.get("case_title") or "").strip() or None,
                actor=actor,
                ts=ts,
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        row = fetch_lead_by_id(conn, lead_id)
        out = lead_row_to_dict(row, conn) if row else {}
    return jsonify({"result": result, "lead": out})


@app.get("/api/crm/leads/<int:lead_id>/presales")
def api_crm_lead_presales_get(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    if not _crm_presales_on_lead_enabled():
        return jsonify({"enabled": False, "presales": None}), 200
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        from crm_lead_presales import presales_payload as _presales_payload
        from crm_lead_catalog import catalog_public_payload

        payload = _presales_payload(conn, lead_id)
        lead_out = lead_row_to_dict(prev, conn)
        care_gate = lead_out.get("presales_care_gate")
        catalog = catalog_public_payload(conn)
    return jsonify({
        "enabled": True,
        "presales": payload,
        "presales_care_gate": care_gate,
        "service_labels": catalog.get("service_labels") or {},
        "service_slugs": catalog.get("service_slugs") or [],
    })


@app.post("/api/crm/leads/<int:lead_id>/presales")
def api_crm_lead_presales_create(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "edit"):
        return _admin_section_forbidden_json("crm_leads", "edit")
    if not _crm_presales_on_lead_enabled():
        return jsonify({"error": "PTT_PRESALES_ON_LEAD chưa bật"}), 400
    body = request.get_json(silent=True) or {}
    slug = str(body.get("service_slug") or "").strip()
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        from crm_lead_presales import ensure_presales as _ensure_ps
        from crm_lead_presales import presales_payload as _presales_payload

        try:
            _ensure_ps(conn, lead_id, slug, suggested_by=_crm_audit_user())
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        payload = _presales_payload(conn, lead_id)
    return jsonify({"presales": payload}), 201


@app.patch("/api/crm/leads/<int:lead_id>/presales")
def api_crm_lead_presales_patch(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "edit"):
        return _admin_section_forbidden_json("crm_leads", "edit")
    if not _crm_presales_on_lead_enabled():
        return jsonify({"error": "PTT_PRESALES_ON_LEAD chưa bật"}), 400
    body = request.get_json(silent=True) or {}
    to_stage = str(body.get("stage") or "").strip()
    notes = str(body.get("notes") or "").strip()[:2000]
    override_reason = str(body.get("override_reason") or "").strip()[:500]
    confirm = bool(body.get("confirm"))
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        from crm_lead_presales import require_presales_care_gate

        try:
            require_presales_care_gate(conn, lead_id)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        from crm_lead_presales import (
            PresalesAdvanceError,
            advance_presales_stage,
            get_by_lead,
            presales_payload as _presales_payload,
        )

        ps = get_by_lead(conn, lead_id)
        if ps is None:
            return jsonify({"error": "Chưa có pre-sales — chọn dịch vụ trước"}), 404
        if not to_stage:
            return jsonify({"error": "Thiếu stage"}), 400
        try:
            advance_presales_stage(
                conn,
                int(ps["id"]),
                to_stage,
                notes=notes,
                override_reason=override_reason,
                allow_override=_admin_full_access(),
                confirm=confirm,
            )
        except PresalesAdvanceError as exc:
            err_msg = str(exc)
            if ps.get("stage") == "lead" and to_stage == "consult":
                from crm_lead_presales_bridge import validate_presales_consult_advance

                gate = validate_presales_consult_advance(
                    conn,
                    int(ps["id"]),
                    override_reason=override_reason,
                    allow_override=_admin_full_access(),
                )
                return jsonify({
                    "error": err_msg,
                    "gate": gate,
                    "requires_confirm": bool(gate.get("requires_confirm")),
                    "requires_override": bool(gate.get("requires_override")),
                }), 400
            return jsonify({"error": err_msg}), 400
        payload = _presales_payload(conn, lead_id)
    return jsonify({"presales": payload})


@app.get("/api/crm/leads/<int:lead_id>/presales-cost-summary")
def api_crm_lead_presales_cost_summary(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    if not _crm_presales_on_lead_enabled():
        return jsonify({"enabled": False}), 200
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        from crm_lead_presales import get_by_lead
        from crm_svc_presales import get_presales_cost_summary_by_presales

        ps = get_by_lead(conn, lead_id)
        if ps is None:
            return jsonify({"error": "Chưa có pre-sales"}), 404
        summary = get_presales_cost_summary_by_presales(conn, int(ps["id"]))
    return jsonify(summary)


@app.patch("/api/crm/leads/<int:lead_id>/presales-cost-cap")
def api_crm_lead_presales_cost_cap(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "edit"):
        return _admin_section_forbidden_json("crm_leads", "edit")
    if not _crm_presales_on_lead_enabled():
        return jsonify({"error": "PTT_PRESALES_ON_LEAD chưa bật"}), 400
    body = request.get_json(silent=True) or {}
    cap_val = _opt_pos_int(body.get("presales_cost_cap_vnd"))
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        from crm_lead_presales import get_by_lead
        from crm_svc_presales import (
            get_presales_cost_summary_by_presales,
            set_presales_cost_cap_for_presales,
        )

        ps = get_by_lead(conn, lead_id)
        if ps is None:
            return jsonify({"error": "Chưa có pre-sales"}), 404
        set_presales_cost_cap_for_presales(conn, int(ps["id"]), cap_val)
        summary = get_presales_cost_summary_by_presales(conn, int(ps["id"]))
    return jsonify(summary)


@app.post("/api/crm/leads/<int:lead_id>/presales/consult-prefill")
def api_crm_lead_presales_consult_prefill(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "edit"):
        return _admin_section_forbidden_json("crm_leads", "edit")
    if not _crm_presales_on_lead_enabled():
        return jsonify({"error": "PTT_PRESALES_ON_LEAD chưa bật"}), 400
    body = request.get_json(silent=True) or {}
    overwrite = bool(body.get("overwrite"))
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        from crm_lead_presales import require_presales_care_gate

        try:
            require_presales_care_gate(conn, lead_id)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        from crm_lead_presales_bridge import prefill_presales_consult_task
        from crm_lead_presales import presales_payload as _presales_payload

        try:
            stats = prefill_presales_consult_task(
                conn, lead_id, overwrite=overwrite
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        payload = _presales_payload(conn, lead_id)
    return jsonify({"stats": stats, "presales": payload})


@app.post("/api/crm/leads/<int:lead_id>/presales/draft-contract")
def api_crm_lead_presales_draft_contract(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "edit"):
        return _admin_section_forbidden_json("crm_leads", "edit")
    if not _crm_presales_on_lead_enabled():
        return jsonify({"error": "PTT_PRESALES_ON_LEAD chưa bật"}), 400
    body = request.get_json(silent=True) or {}
    title = str(body.get("title") or "").strip()[:500] or None
    notes = str(body.get("notes") or "").strip()[:8000]
    try:
        amount_vnd = int(body.get("amount_vnd") or 0)
    except (TypeError, ValueError):
        amount_vnd = 0
    ts = _crm_ts()
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        from crm_lead_presales import require_presales_care_gate

        try:
            require_presales_care_gate(conn, lead_id)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        from crm_lead_presales_contract import (
            PresalesContractError,
            create_draft_contract_from_lead,
        )
        from crm_lead_presales import presales_payload as _presales_payload

        try:
            contract = create_draft_contract_from_lead(
                conn,
                lead_id,
                title=title,
                amount_vnd=amount_vnd,
                notes=notes,
                actor=_crm_audit_user(),
                ts=ts,
            )
        except PresalesContractError as exc:
            return jsonify({"error": str(exc)}), 400
        payload = _presales_payload(conn, lead_id)
    return jsonify({"contract": contract, "presales": payload}), 201


@app.patch("/api/crm/leads/<int:lead_id>/presales/tasks/<int:task_id>")
def api_crm_lead_presales_task_patch(lead_id: int, task_id: int) -> Any:
    if not _admin_section_can("crm_leads", "edit"):
        return _admin_section_forbidden_json("crm_leads", "edit")
    if not _crm_presales_on_lead_enabled():
        return jsonify({"error": "PTT_PRESALES_ON_LEAD chưa bật"}), 400
    body = request.get_json(silent=True) or {}
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền."}), 403
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        row = conn.execute(
            """
            SELECT t.id FROM crm_lead_presales_tasks t
            INNER JOIN crm_lead_presales p ON p.id = t.presales_id
            WHERE t.id = ? AND p.lead_id = ?
            """,
            (task_id, lead_id),
        ).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy task"}), 404
        from crm_lead_presales import require_presales_care_gate

        try:
            require_presales_care_gate(conn, lead_id)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        from crm_lead_presales import presales_payload as _presales_payload
        from crm_lead_presales import update_presales_task as _upd_ps_task

        form_data = body.get("form_data")
        kwargs: dict[str, Any] = {}
        if "is_done" in body:
            kwargs["is_done"] = bool(body.get("is_done"))
        if "notes" in body:
            kwargs["notes"] = str(body.get("notes") or "")
        if isinstance(form_data, dict):
            kwargs["form_data"] = form_data
        if kwargs:
            _upd_ps_task(conn, task_id, **kwargs)
        payload = _presales_payload(conn, lead_id)
    return jsonify({"presales": payload})


@app.get("/api/crm/leads/export")
def api_crm_export_leads() -> Any:
    if not _admin_section_can("crm_leads", "export"):
        return _admin_section_forbidden_json("crm_leads", "export")
    fmt = str(request.args.get("format") or "xlsx").strip().lower()
    if fmt not in ("xlsx", "pdf"):
        fmt = "xlsx"
    owner_id = _crm_lead_owner_filter()
    if owner_id is None:
        raw_owner = str(request.args.get("owner_id") or "").strip()
        if raw_owner:
            try:
                owner_id = int(raw_owner)
            except ValueError:
                return jsonify({"error": "owner_id không hợp lệ"}), 400
    try:
        re_project_id = _crm_lead_project_filter_from_request()
    except ValueError:
        return jsonify({"error": "re_project_id không hợp lệ"}), 400
    filt: dict[str, Any] = {
        "owner_id": owner_id,
        "status": request.args.get("status"),
        "level": request.args.get("level"),
        "source": request.args.get("source"),
        "q": request.args.get("q"),
    }
    if re_project_id is not _UNSET:
        filt["re_project_id"] = re_project_id
    with get_connection() as conn:
        rows = fetch_leads(
            conn,
            **filt,
            sla_overdue_only=str(request.args.get("sla_overdue") or "") in ("1", "true"),
            limit=2000,
        )
        stats = fetch_lead_stats_extended(
            conn,
            owner_id=owner_id,
            re_project_id=re_project_id,
            ts=_crm_ts(),
        )
    try:
        if fmt == "pdf":
            buf, fname = build_leads_pdf(rows, stats)
            return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=fname)
        buf, fname = build_leads_xlsx(rows, stats)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=fname,
        )
    except RuntimeError as exc:
        return jsonify({"error": str(exc)}), 503


@app.get("/api/crm/integration/webhooks/facebook", strict_slashes=False)
@app.get("/api/crm/integration/webhooks/facebook/<webhook_slug>", strict_slashes=False)
def api_crm_facebook_webhook_verify(webhook_slug: str | None = None) -> Any:
    """Facebook Lead Ads — xác minh webhook (hub.challenge)."""
    _log = logging.getLogger("ptt.facebook.webhook")
    mode = str(request.args.get("hub.mode") or "")
    token = str(request.args.get("hub.verify_token") or "")
    challenge = request.args.get("hub.challenge")
    expected = facebook_verify_token()
    slug = str(webhook_slug or _facebook_webhook_slug_from_path() or "").strip().lower()
    token_ok = bool(expected and _const_eq_str(token, expected))
    project_id: int | None = None
    if slug and not token_ok:
        try:
            with get_connection() as conn:
                from crm_project_webhooks import verify_project_webhook_token

                project_id = verify_project_webhook_token(conn, slug, token)
                token_ok = project_id is not None
        except Exception:
            token_ok = False
    if (
        mode == "subscribe"
        and token_ok
        and challenge is not None
    ):
        if slug and project_id:
            _log.info("Facebook webhook verify OK slug=%s project_id=%s", slug, project_id)
        return str(challenge), 200, {"Content-Type": "text/plain; charset=utf-8"}
    _log.warning(
        "Facebook webhook verify FAILED mode=%s slug=%s token_match=%s has_challenge=%s",
        mode,
        slug or "-",
        token_ok,
        challenge is not None,
    )
    return jsonify({"error": "Xác minh webhook Facebook thất bại."}), 403


def _log_facebook_webhook_rejection(reason: str) -> None:
    try:
        ts = _crm_ts()
        with get_connection() as conn:
            save_facebook_webhook_receipt(
                conn,
                ts=ts,
                event_count=0,
                created_count=0,
                message=f"REJECTED: {reason}"[:500],
            )
            conn.commit()
    except Exception:
        pass


def _log_facebook_webhook_rejection_async(reason: str) -> None:
    def _run() -> None:
        try:
            with app.app_context():
                _log_facebook_webhook_rejection(reason)
        except Exception:
            pass

    threading.Thread(target=_run, daemon=True, name="ptt-fb-webhook-reject-log").start()


@app.post("/api/crm/integration/webhooks/facebook", strict_slashes=False)
@app.post("/api/crm/integration/webhooks/facebook/<webhook_slug>", strict_slashes=False)
def api_crm_facebook_webhook_ingest(webhook_slug: str | None = None) -> Any:
    """Meta yêu cầu luôn 200 OK — trả 401/403/5xx sẽ kẹt hàng đợi leadgen (delivery.rejected)."""
    _log = logging.getLogger("ptt.facebook.webhook")
    slug = str(webhook_slug or _facebook_webhook_slug_from_path() or "").strip().lower()
    raw = _facebook_webhook_raw_body()
    sig_hdr = _facebook_webhook_signature_header()
    sig_ok = verify_facebook_signature(raw, sig_hdr)
    if not sig_ok:
        _log.warning(
            "Facebook webhook: chữ ký không hợp lệ slug=%s sig=%s len=%s — bỏ qua payload, vẫn trả 200 cho Meta",
            slug or "-",
            (sig_hdr or "")[:24],
            len(raw),
        )
        _log_facebook_webhook_rejection_async("chữ ký không hợp lệ (đã ack 200)")
        return "EVENT_RECEIVED", 200, {"Content-Type": "text/plain; charset=utf-8"}
    payload = parse_facebook_webhook_json(raw)
    if payload:
        def _run() -> None:
            with app.app_context():
                _facebook_webhook_worker(payload, webhook_slug=slug or None)

        threading.Thread(target=_run, daemon=True, name="ptt-fb-webhook").start()
    return "EVENT_RECEIVED", 200, {"Content-Type": "text/plain; charset=utf-8"}


def _facebook_webhook_worker(payload: dict[str, Any], *, webhook_slug: str | None = None) -> None:
    """Xử lý webhook sau khi đã trả 200 — commit nhanh, enrich Graph nền."""
    import time

    _log = logging.getLogger("ptt.facebook.webhook")
    from crm_facebook_leads import extract_facebook_leadgen_events
    from crm_facebook_pending import process_pending_facebook_leads
    from crm_project_webhooks import resolve_project_from_webhook

    ts = _crm_ts()
    result: dict[str, Any] = {}
    event_count = 0
    forced_project_id: int | None = None
    slug = str(webhook_slug or "").strip().lower() or None
    try:
        with get_connection() as conn:
            if slug:
                forced_project_id = resolve_project_from_webhook(conn, webhook_slug=slug)
            result = process_facebook_webhook_payload(
                conn,
                payload,
                created_by="webhook:facebook",
                ts=ts,
                webhook_slug=slug,
                forced_project_id=forced_project_id,
            )
            event_count = len(extract_facebook_leadgen_events(payload))
            msg = str(result.get("message") or "").strip()
            save_facebook_webhook_receipt(
                conn,
                ts=ts,
                event_count=event_count,
                created_count=int(result.get("created_count") or 0),
                message=msg[:500],
            )
            conn.commit()

        def _enrich_pending() -> None:
            try:
                with app.app_context():
                    pending_created = 0
                    for round_idx in range(8):
                        with get_connection() as conn2:
                            pending = process_pending_facebook_leads(
                                conn2, created_by="webhook:pending", ts=_crm_ts(), max_items=10
                            )
                            pending_created += int(pending.get("created_count") or 0)
                            if int(pending.get("processed") or 0) <= 0:
                                break
                            if pending_created or any(
                                r.get("status") == "enriched"
                                for r in (pending.get("results") or [])
                            ):
                                save_facebook_webhook_receipt(
                                    conn2,
                                    ts=_crm_ts(),
                                    event_count=event_count,
                                    created_count=int(result.get("created_count") or 0) + pending_created,
                                    message=(msg + f" · enrich+{pending_created}")[:500]
                                    if pending_created
                                    else msg[:500],
                                )
                            conn2.commit()
                        if round_idx < 7:
                            time.sleep(1.5)
            except Exception:
                _log.exception("Facebook webhook pending enrich failed")

        threading.Thread(target=_enrich_pending, daemon=True, name="ptt-fb-webhook-enrich").start()
        _log.info(
            "Facebook webhook: events=%s processed=%s created=%s skipped=%s",
            event_count,
            result.get("processed_count"),
            result.get("created_count"),
            result.get("skipped_count"),
        )
        if result.get("results"):
            for r in result.get("results") or []:
                st = r.get("status")
                if st not in ("created_assigned", "created_unassigned", "duplicate_skipped"):
                    _log.warning(
                        "Facebook webhook lead skip: status=%s msg=%s leadgen=%s",
                        st,
                        r.get("message"),
                        r.get("facebook_leadgen_id") or r.get("leadgen_id"),
                    )
    except Exception:
        _log.exception("Facebook webhook background processing failed")


@app.get("/api/crm/integration/facebook/status")
def api_crm_facebook_integration_status() -> Any:
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    with get_connection() as conn:
        status = facebook_integration_status(conn)
    return jsonify(status)


@app.post("/api/crm/integration/facebook/sync")
def api_crm_facebook_sync() -> Any:
    """Đồng bộ lead từ form Facebook đã cấu hình."""
    if not _admin_section_can("crm_leads", "create"):
        return _admin_section_forbidden_json("crm_leads", "create")
    ts = _crm_ts()
    actor = _crm_audit_user()
    with get_connection() as conn:
        # Quét N lead mới nhất trên form (không lọc theo last_sync_at) — kéo lead cũ còn thiếu
        result = run_facebook_ingest_cycle(
            conn, created_by=actor, ts=ts, recent_only=True, limit_per_form=50
        )
    return _facebook_sync_json_response(result)


@app.post("/api/crm/integration/facebook/sync-auto")
def api_crm_facebook_sync_auto() -> Any:
    """Tự động kéo lead (UI polling / nền) — quyền view."""
    if not _admin_section_can("crm_leads", "view"):
        return _admin_section_forbidden_json("crm_leads", "view")
    ts = _crm_ts()
    actor = _crm_audit_user() or "sync-auto"
    with get_connection() as conn:
        result = run_facebook_ingest_cycle(conn, created_by=actor, ts=ts, recent_only=True)
    return _facebook_sync_json_response(result)


@app.post("/api/crm/integration/facebook/sync-cron")
def api_crm_facebook_sync_cron() -> Any:
    """Cron nội bộ — tự động kéo lead Facebook (localhost hoặc Bearer secret)."""
    if not _crm_facebook_sync_cron_allowed():
        return jsonify({"error": "Unauthorized."}), 401
    result = run_facebook_autosync_once()
    code = 200 if result.get("ok") else 400
    return jsonify(result), code


@app.post("/api/crm/integration/facebook/leads")
def api_crm_facebook_leads_ingest() -> Any:
    """Nhập lead Facebook thủ công (payload chuẩn hoặc leadgen_id)."""
    if not _admin_section_can("crm_leads", "create"):
        return _admin_section_forbidden_json("crm_leads", "create")
    payload = request.get_json(force=True) or {}
    ts = _crm_ts()
    actor = _crm_audit_user()
    leadgen_id = str(payload.get("leadgen_id") or payload.get("facebook_leadgen_id") or "").strip()
    with get_connection() as conn:
        if leadgen_id:
            result = process_facebook_leadgen_id(
                conn, leadgen_id, created_by=actor, ts=ts
            )
        else:
            batch = process_facebook_webhook_payload(
                conn, payload, created_by=actor, ts=ts
            )
            result = batch["results"][0] if batch.get("results") else batch
    if result.get("status") == "error":
        return jsonify({"error": result.get("message"), "result": result}), 400
    if result.get("status") == "skipped":
        return jsonify({"error": result.get("message"), "result": result}), 400
    lead_id = result.get("lead_id")
    lead_out = None
    if lead_id:
        with get_connection() as conn:
            row = fetch_lead_by_id(conn, int(lead_id))
            if row:
                lead_out = lead_row_to_dict(row, conn)
    return jsonify({"ok": True, "result": result, "lead": lead_out}), 201


@app.post("/api/crm/integration/webhooks/zalo", strict_slashes=False)
@app.post("/api/crm/integration/webhooks/zalo/<webhook_slug>", strict_slashes=False)
def api_crm_zalo_webhook_ingest(webhook_slug: str | None = None) -> Any:
    raw = request.get_data()
    sig = request.headers.get("X-Zalo-Signature") or request.headers.get("X-Signature")
    if zalo_webhook_secret() and not verify_zalo_signature(raw, sig):
        return jsonify({"error": "Chữ ký Zalo không hợp lệ."}), 401
    payload = request.get_json(silent=True) or {}
    items = parse_zalo_webhook(payload)
    if not items:
        return jsonify({"ok": True, "message": "Không có lead trong payload.", "created_count": 0})
    slug = str(webhook_slug or "").strip().lower() or None
    ts = _crm_ts()
    with get_connection() as conn:
        result = ingest_webhook_leads(
            conn,
            items,
            default_source="zalo",
            created_by="webhook:zalo",
            ts=ts,
            webhook_slug=slug,
        )
        conn.commit()
    return jsonify({"ok": True, **result}), 201


@app.get("/api/crm/staff/me")
def api_crm_staff_me() -> Any:
    sid = _crm_effective_staff_id()
    if sid is None:
        return jsonify({"error": "Chỉ dành cho nhân viên đăng nhập portal."}), 403
    with get_connection() as conn:
        profile = _crm_staff_profile(conn, sid)
        if profile is None:
            return jsonify({"error": "Không tìm thấy nhân viên."}), 404
        position_id = _crm_staff_position_id(conn, sid)
        metric_ids = _crm_position_metric_ids(conn, position_id)
    return jsonify(
        {
            "staff": profile,
            "position_id": position_id,
            "metric_ids": metric_ids,
        }
    )


@app.get("/api/crm/staff")
def api_crm_list_staff() -> Any:
    include_inactive, status_filter, q_raw, dept_filter = _crm_staff_params_from_request()
    # Kanban / dropdown gọi API không tham số → trả toàn bộ danh sách (không phân trang).
    paginate = (request.args.get("page") is not None) or (request.args.get("per_page") is not None)

    with get_connection() as conn:
        total_filtered = _count_crm_staff_filtered(
            conn,
            include_inactive=include_inactive,
            status_filter=status_filter,
            q_raw=q_raw,
            dept_filter=dept_filter,
        )
        if paginate:
            page, per_page = _crm_staff_pagination_from_request()
            offset = (page - 1) * per_page
            rows = _fetch_crm_staff_rows(
                conn,
                include_inactive=include_inactive,
                status_filter=status_filter,
                q_raw=q_raw,
                dept_filter=dept_filter,
                limit=per_page,
                offset=offset,
            )
            total_pages = max(1, (total_filtered + per_page - 1) // per_page) if total_filtered else 1
            meta = {
                "page": page,
                "per_page": per_page,
                "total": total_filtered,
                "total_pages": total_pages,
            }
        else:
            rows = _fetch_crm_staff_rows(
                conn,
                include_inactive=include_inactive,
                status_filter=status_filter,
                q_raw=q_raw,
                dept_filter=dept_filter,
                limit=None,
                offset=None,
            )
            meta = {
                "page": 1,
                "per_page": max(total_filtered, 1),
                "total": total_filtered,
                "total_pages": 1,
            }

        sum_row = conn.execute(
            """
            SELECT
              (SELECT COUNT(*) FROM crm_staff) AS staff_total,
              (SELECT COUNT(*) FROM crm_staff WHERE active = 1) AS staff_active,
              (SELECT COUNT(*) FROM crm_staff WHERE active = 0) AS staff_inactive,
              (SELECT COUNT(*) FROM crm_cases
               WHERE assigned_staff_id IS NOT NULL AND status != 'dong') AS open_assigned_cases
            """
        ).fetchone()

    staff_out = [staff_row_for_api(r) for r in rows]
    summary = dict(sum_row) if sum_row is not None else {}
    return jsonify({"staff": staff_out, "summary": summary, "meta": meta})


@app.get("/api/crm/staff/levels")
def api_crm_staff_levels_get() -> Any:
    if not _admin_section_can("crm_staff_roster", "view"):
        return _admin_section_forbidden_json("crm_staff_roster", "view")
    with get_connection() as conn:
        cfg = fetch_staff_config(conn)
    return jsonify(
        {
            "staff_levels": cfg.get("staff_levels") or DEFAULT_STAFF_LEVELS,
            "defaults": DEFAULT_STAFF_LEVELS,
            "can_configure": _admin_section_can("crm_staff_roster", "edit"),
        }
    )


@app.put("/api/crm/staff/levels")
def api_crm_staff_levels_put() -> Any:
    if not _admin_section_can("crm_staff_roster", "edit"):
        return _admin_section_forbidden_json("crm_staff_roster", "edit")
    payload = request.get_json(force=True) or {}
    ts = _crm_ts()
    actor = _crm_audit_user()
    with get_connection() as conn:
        try:
            cfg = save_staff_config(conn, config=payload, updated_by=actor, ts=ts)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    return jsonify({"staff_levels": cfg.get("staff_levels") or DEFAULT_STAFF_LEVELS})


@app.get("/api/crm/staff/competency")
def api_crm_staff_competency_get() -> Any:
    if not _admin_section_can("crm_staff_roster", "view"):
        return _admin_section_forbidden_json("crm_staff_roster", "view")
    with get_connection() as conn:
        cfg = fetch_staff_config(conn)
    comp = cfg.get("competency") or default_competency_config()
    return jsonify(
        {
            "competency": comp,
            "defaults": default_competency_config(),
            "metric_options": METRIC_OPTIONS,
            "can_configure": _admin_section_can("crm_staff_roster", "edit"),
        }
    )


@app.put("/api/crm/staff/competency")
def api_crm_staff_competency_put() -> Any:
    if not _admin_section_can("crm_staff_roster", "edit"):
        return _admin_section_forbidden_json("crm_staff_roster", "edit")
    payload = request.get_json(force=True) or {}
    ts = _crm_ts()
    actor = _crm_audit_user()
    with get_connection() as conn:
        try:
            cfg = save_staff_config(
                conn,
                config={"competency": payload.get("competency") or payload},
                updated_by=actor,
                ts=ts,
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    return jsonify({"competency": cfg.get("competency") or default_competency_config()})


@app.post("/api/crm/staff/competency/score")
def api_crm_staff_competency_score() -> Any:
    if not _admin_section_can("crm_staff_roster", "view"):
        return _admin_section_forbidden_json("crm_staff_roster", "view")
    payload = request.get_json(force=True) or {}
    metrics = payload.get("metrics")
    if not isinstance(metrics, dict):
        return jsonify({"error": "metrics phải là object."}), 400
    parsed: dict[str, float] = {}
    for k, v in metrics.items():
        key = str(k).strip()
        if key not in {m["id"] for m in METRIC_OPTIONS}:
            continue
        try:
            parsed[key] = float(v)
        except (TypeError, ValueError):
            return jsonify({"error": f"Giá trị metric «{key}» không hợp lệ."}), 400
    staff_id = _opt_pos_int(payload.get("staff_id"))
    year = _opt_pos_int(payload.get("year"))
    month = _opt_pos_int(payload.get("month"))
    if staff_id and "close_rate_pct" not in parsed:
        from crm_lead_kpi_metrics import get_staff_close_rate_pct

        with get_connection() as conn:
            parsed["close_rate_pct"] = get_staff_close_rate_pct(conn, staff_id)
    with get_connection() as conn:
        comp = fetch_staff_competency(conn)
    out = score_staff_competency(parsed, comp)
    return jsonify(out)


@app.get("/api/crm/staff/competency/metrics")
def api_crm_staff_competency_metrics() -> Any:
    if not _admin_section_can("crm_staff_roster", "view"):
        return _admin_section_forbidden_json("crm_staff_roster", "view")
    staff_id = _opt_pos_int(request.args.get("staff_id"))
    if staff_id is None:
        return jsonify({"error": "Thiếu staff_id."}), 400
    now = datetime.utcnow()
    year = _opt_pos_int(request.args.get("year")) or now.year
    month = _opt_pos_int(request.args.get("month")) or now.month
    from crm_lead_kpi_metrics import get_staff_close_rate_pct, get_unified_lead_kpi_summary

    with get_connection() as conn:
        monthly = get_unified_lead_kpi_summary(
            conn, year=year, month=month, staff_id=staff_id, period_cohort=True
        )
        all_time = get_unified_lead_kpi_summary(
            conn, staff_id=staff_id, period_cohort=False
        )
        close_rate_pct = get_staff_close_rate_pct(conn, staff_id)
    return jsonify(
        {
            "staff_id": staff_id,
            "year": year,
            "month": month,
            "close_rate_pct": close_rate_pct,
            "monthly": monthly,
            "all_time": all_time,
        }
    )


@app.post("/api/crm/staff")
def api_crm_create_staff() -> Any:
    payload = request.get_json(force=True) or {}
    name = str(payload.get("name", "")).strip()[:240]
    if not name:
        return jsonify({"error": "Thiếu họ tên nhân viên"}), 400
    phone = str(payload.get("phone", "")).strip()[:80]
    email = str(payload.get("email", "")).strip()[:240]
    job_title = str(payload.get("job_title", "")).strip()[:200]
    department = str(payload.get("department", "")).strip()[:200]
    internal_code = str(payload.get("internal_code", "")).strip()[:80]
    attendance_pin = str(payload.get("attendance_pin", "")).strip()[:64]
    notes = str(payload.get("notes", "")).strip()[:4000]
    dept_id = _opt_pos_int(payload.get("department_id"))
    pos_id = _opt_pos_int(payload.get("position_id"))
    rep_id = _opt_pos_int(payload.get("reports_to_id"))
    employment_type = str(payload.get("employment_type") or "full_time").strip()
    if employment_type not in CRM_EMPLOYMENT_TYPES:
        employment_type = "full_time"
    started_on = str(payload.get("started_on") or "").strip()[:32]
    ended_on = str(payload.get("ended_on") or "").strip()[:32]
    if email and not _EMAIL_RE.match(email):
        return jsonify({"error": "Email không hợp lệ"}), 400
    try:
        sort_order = int(payload.get("sort_order") or 0)
    except (TypeError, ValueError):
        sort_order = 0
    try:
        base_salary_vnd = int(payload.get("base_salary_vnd") or 0)
    except (TypeError, ValueError):
        base_salary_vnd = 0
    base_salary_vnd = max(0, min(base_salary_vnd, 9_999_999_999))
    sales_level = ""
    if "sales_level" in payload:
        try:
            sales_level = normalize_sales_level(payload.get("sales_level"))
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    created = datetime.now().strftime("%Y-%m-%d")
    ts = _crm_ts()
    with get_connection() as conn:
        if dept_id is not None and not _dept_exists_active(conn, dept_id):
            return jsonify({"error": "Phòng ban không tồn tại hoặc đã ngưng hoạt động"}), 400
        if pos_id is not None and not _position_exists_active(conn, pos_id):
            return jsonify({"error": "Chức vụ không tồn tại hoặc đã ngưng hoạt động"}), 400
        if rep_id is not None:
            if conn.execute("SELECT id FROM crm_staff WHERE id = ?", (rep_id,)).fetchone() is None:
                return jsonify({"error": "Người quản lý trực tiếp không tồn tại"}), 400
        if email and _staff_field_exists(conn, "email", email, exclude_id=None):
            return jsonify({"error": "Email đã được dùng cho nhân viên khác"}), 400
        if internal_code and _staff_field_exists(conn, "internal_code", internal_code, exclude_id=None):
            return jsonify({"error": "Mã nhân viên nội bộ đã tồn tại"}), 400
        if attendance_pin and _staff_field_exists(conn, "attendance_pin", attendance_pin, exclude_id=None):
            return jsonify({"error": "Mã máy chấm công đã dùng cho nhân viên khác"}), 400
        login_username = str(payload.get("login_username") or "").strip()
        login_password = str(payload.get("login_password") or "")
        login_enabled = bool(payload.get("login_enabled"))
        merged_login: dict[str, Any] = {
            "login_username": login_username,
            "password_hash": "",
            "login_enabled": 1 if login_enabled else 0,
        }
        if login_password:
            err_login = apply_staff_login_from_payload(conn, merged_login, payload, staff_id=None)
            if err_login:
                return jsonify({"error": err_login}), 400
        elif login_username:
            err_login = apply_staff_login_from_payload(conn, merged_login, {"login_username": login_username}, staff_id=None)
            if err_login:
                return jsonify({"error": err_login}), 400
        cur = conn.execute(
            """
            INSERT INTO crm_staff (
                name, phone, email, job_title, department, internal_code, attendance_pin, notes,
                active, sort_order, created_at, updated_at,
                department_id, position_id, reports_to_id, employment_type, started_on, ended_on,
                base_salary_vnd, sales_level, login_username, password_hash, login_enabled
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                name,
                phone,
                email,
                job_title,
                department,
                internal_code,
                attendance_pin,
                notes,
                sort_order,
                created,
                ts,
                dept_id,
                pos_id,
                rep_id,
                employment_type,
                started_on,
                ended_on,
                base_salary_vnd,
                sales_level,
                merged_login.get("login_username") or "",
                merged_login.get("password_hash") or "",
                int(merged_login.get("login_enabled") or 0),
            ),
        )
        sid = int(cur.lastrowid)
        login_u = str(merged_login.get("login_username") or "").strip()
        login_ph = str(merged_login.get("password_hash") or "").strip()
        if login_u and login_ph:
            sync_password_hash(conn, login_u, login_ph, updated_at=ts)
        row = conn.execute("SELECT * FROM crm_staff WHERE id = ?", (sid,)).fetchone()
    assert row is not None
    return jsonify(staff_row_for_api(row)), 201


@app.patch("/api/crm/staff/<int:staff_id>")
def api_crm_patch_staff(staff_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM crm_staff WHERE id = ?", (staff_id,)).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy nhân viên"}), 404
        prev = dict(row)
        merged: dict[str, Any] = dict(row)

        if "name" in payload and isinstance(payload["name"], str):
            n = payload["name"].strip()[:240]
            if not n:
                return jsonify({"error": "Tên không được trống"}), 400
            merged["name"] = n
        if "phone" in payload and isinstance(payload["phone"], str):
            merged["phone"] = payload["phone"].strip()[:80]
        if "email" in payload and isinstance(payload["email"], str):
            em = payload["email"].strip()[:240]
            if em and not _EMAIL_RE.match(em):
                return jsonify({"error": "Email không hợp lệ"}), 400
            merged["email"] = em
        if "job_title" in payload and isinstance(payload["job_title"], str):
            merged["job_title"] = payload["job_title"].strip()[:200]
        if "department" in payload and isinstance(payload["department"], str):
            merged["department"] = payload["department"].strip()[:200]
        if "internal_code" in payload and isinstance(payload["internal_code"], str):
            merged["internal_code"] = payload["internal_code"].strip()[:80]
        if "attendance_pin" in payload and isinstance(payload["attendance_pin"], str):
            merged["attendance_pin"] = payload["attendance_pin"].strip()[:64]
        if "notes" in payload and isinstance(payload["notes"], str):
            merged["notes"] = payload["notes"].strip()[:4000]
        if "sort_order" in payload and payload["sort_order"] is not None:
            try:
                merged["sort_order"] = int(payload["sort_order"])
            except (TypeError, ValueError):
                pass
        if "active" in payload:
            merged["active"] = bool(payload["active"] in (True, 1, "1", "true"))

        if "department_id" in payload:
            merged["department_id"] = _opt_pos_int(payload.get("department_id"))
        if "position_id" in payload:
            merged["position_id"] = _opt_pos_int(payload.get("position_id"))
        if "reports_to_id" in payload:
            merged["reports_to_id"] = _opt_pos_int(payload.get("reports_to_id"))
        if "employment_type" in payload and isinstance(payload["employment_type"], str):
            et = payload["employment_type"].strip()
            if et in CRM_EMPLOYMENT_TYPES:
                merged["employment_type"] = et
        if "started_on" in payload and isinstance(payload["started_on"], str):
            merged["started_on"] = payload["started_on"].strip()[:32]
        if "ended_on" in payload and isinstance(payload["ended_on"], str):
            merged["ended_on"] = payload["ended_on"].strip()[:32]
        if "base_salary_vnd" in payload and payload["base_salary_vnd"] is not None:
            try:
                merged["base_salary_vnd"] = max(
                    0, min(int(payload["base_salary_vnd"]), 9_999_999_999)
                )
            except (TypeError, ValueError):
                pass
        if "sales_level" in payload:
            try:
                merged["sales_level"] = normalize_sales_level(payload.get("sales_level"))
            except ValueError as exc:
                return jsonify({"error": str(exc)}), 400

        dpid = merged.get("department_id")
        if dpid is not None and not _dept_exists_active(conn, int(dpid)):
            return jsonify({"error": "Phòng ban không tồn tại hoặc đã ngưng hoạt động"}), 400
        ppid = merged.get("position_id")
        if ppid is not None and not _position_exists_active(conn, int(ppid)):
            return jsonify({"error": "Chức vụ không tồn tại hoặc đã ngưng hoạt động"}), 400
        rtid = merged.get("reports_to_id")
        if rtid is not None:
            if int(rtid) == staff_id:
                return jsonify({"error": "Không thể báo cáo trực tiếp cho chính mình"}), 400
            if conn.execute("SELECT id FROM crm_staff WHERE id = ?", (int(rtid),)).fetchone() is None:
                return jsonify({"error": "Người quản lý trực tiếp không tồn tại"}), 400

        etn = str(merged.get("employment_type") or "full_time").strip()
        if etn not in CRM_EMPLOYMENT_TYPES:
            etn = "full_time"
        merged["employment_type"] = etn

        em = str(merged.get("email") or "").strip()
        if em and _staff_field_exists(conn, "email", em, exclude_id=staff_id):
            return jsonify({"error": "Email đã được dùng cho nhân viên khác"}), 400
        code = str(merged.get("internal_code") or "").strip()
        if code and _staff_field_exists(conn, "internal_code", code, exclude_id=staff_id):
            return jsonify({"error": "Mã nhân viên nội bộ đã tồn tại"}), 400
        ap = str(merged.get("attendance_pin") or "").strip()
        if ap and _staff_field_exists(conn, "attendance_pin", ap, exclude_id=staff_id):
            return jsonify({"error": "Mã máy chấm công đã dùng cho nhân viên khác"}), 400

        err_login = apply_staff_login_from_payload(conn, merged, payload, staff_id=staff_id)
        if err_login:
            return jsonify({"error": err_login}), 400

        ts = _crm_ts()
        conn.execute(
            """
            UPDATE crm_staff
            SET name = ?, phone = ?, email = ?, job_title = ?, department = ?, internal_code = ?,
                attendance_pin = ?, notes = ?, active = ?, sort_order = ?, updated_at = ?,
                department_id = ?, position_id = ?, reports_to_id = ?, employment_type = ?, started_on = ?, ended_on = ?,
                base_salary_vnd = ?, sales_level = ?, login_username = ?, password_hash = ?, login_enabled = ?
            WHERE id = ?
            """,
            (
                merged["name"],
                merged.get("phone") or "",
                merged.get("email") or "",
                merged.get("job_title") or "",
                merged.get("department") or "",
                merged.get("internal_code") or "",
                merged.get("attendance_pin") or "",
                merged.get("notes") or "",
                1 if merged["active"] else 0,
                merged.get("sort_order") or 0,
                ts,
                merged.get("department_id"),
                merged.get("position_id"),
                merged.get("reports_to_id"),
                merged["employment_type"],
                merged.get("started_on") or "",
                merged.get("ended_on") or "",
                int(merged.get("base_salary_vnd") or 0),
                str(merged.get("sales_level") or ""),
                str(merged.get("login_username") or ""),
                str(merged.get("password_hash") or ""),
                int(merged.get("login_enabled") or 0),
                staff_id,
            ),
        )
        login_u = str(merged.get("login_username") or "").strip()
        login_ph = str(merged.get("password_hash") or "").strip()
        if login_u and login_ph:
            sync_password_hash(conn, login_u, login_ph, updated_at=ts)

        if str(prev.get("name") or "") != str(merged.get("name") or ""):
            conn.execute(
                """
                UPDATE crm_cases SET assigned_to = ?
                WHERE assigned_staff_id = ?
                """,
                (str(merged["name"]), staff_id),
            )

        if not merged["active"]:
            conn.execute(
                """
                UPDATE crm_cases
                SET assigned_staff_id = NULL, assigned_to = '', assigned_at = ''
                WHERE assigned_staff_id = ?
                """,
                (staff_id,),
            )
        row2 = conn.execute("SELECT * FROM crm_staff WHERE id = ?", (staff_id,)).fetchone()
    assert row2 is not None
    return jsonify(staff_row_for_api(row2))


def _crm_validate_date_ymd(s: str) -> bool:
    s = str(s or "").strip()
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
    if not m:
        return False
    y, mo, d = int(m[1]), int(m[2]), int(m[3])
    try:
        datetime(y, mo, d)
    except ValueError:
        return False
    return True


def _crm_today_ymd() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def _crm_attendance_work_date_error(work_date: str) -> str | None:
    """None nếu hợp lệ; ngược lại trả thông báo lỗi tiếng Việt."""
    wd = str(work_date or "").strip()
    if not _crm_validate_date_ymd(wd):
        return "Ngày chấm công không hợp lệ (YYYY-MM-DD)"
    if wd > _crm_today_ymd():
        return "Ngày chấm công không được sau ngày hiện tại"
    return None


def _crm_normalize_hhmm_safe(s: str) -> str | None:
    m = re.match(r"^(\d{1,2}):(\d{2})$", (s or "").strip())
    if not m:
        return None
    h, mm = int(m[1]), int(m[2])
    if h > 23 or mm > 59:
        return None
    return f"{h:02d}:{mm:02d}"


def _crm_hhmm_to_minutes(s: str) -> int | None:
    m = re.match(r"^(\d{1,2}):(\d{2})$", (s or "").strip())
    if not m:
        return None
    h, mm = int(m[1]), int(m[2])
    if h > 23 or mm > 59:
        return None
    return h * 60 + mm


def _crm_minutes_to_hhmm(n: int) -> str:
    n = max(0, min(n, 24 * 60 - 1))
    return f"{n // 60:02d}:{n % 60:02d}"


def _crm_merge_hhmm_pair(a: str, b: str, *, earliest: bool) -> str:
    ma, mb = _crm_hhmm_to_minutes(a), _crm_hhmm_to_minutes(b)
    if ma is None:
        return b
    if mb is None:
        return a
    pick = min(ma, mb) if earliest else max(ma, mb)
    return _crm_minutes_to_hhmm(pick)


def _crm_attendance_device_key_expected() -> str:
    return (os.getenv("CRM_ATTENDANCE_DEVICE_KEY") or "").strip()


def _crm_attendance_device_auth_ok(req: Any) -> bool:
    exp = _crm_attendance_device_key_expected()
    if not exp:
        return False
    got = (req.headers.get("X-CRM-Device-Key") or "").strip()
    if got and secrets.compare_digest(got.encode("utf-8"), exp.encode("utf-8")):
        return True
    auth = req.headers.get("Authorization") or ""
    if auth.lower().startswith("bearer "):
        tok = auth[7:].strip()
        if tok and secrets.compare_digest(tok.encode("utf-8"), exp.encode("utf-8")):
            return True
    return False


def _crm_resolve_staff_by_device_id(
    conn: sqlite3.Connection, raw: str
) -> tuple[int | None, str | None]:
    key = (raw or "").strip()
    if not key:
        return None, "Thiếu mã nhân sự trên máy (pin / staff_code)"
    lk = key.lower()
    rows = conn.execute(
        """
        SELECT id FROM crm_staff
        WHERE active = 1 AND trim(coalesce(attendance_pin, '')) != ''
          AND lower(trim(attendance_pin)) = ?
        """,
        (lk,),
    ).fetchall()
    if len(rows) > 1:
        return None, "Trùng mã máy chấm công (attendance_pin)"
    if len(rows) == 1:
        return int(rows[0]["id"]), None
    rows2 = conn.execute(
        """
        SELECT id FROM crm_staff
        WHERE active = 1 AND trim(coalesce(internal_code, '')) != ''
          AND lower(trim(internal_code)) = ?
        """,
        (lk,),
    ).fetchall()
    if len(rows2) > 1:
        return None, "Trùng mã nhân viên nội bộ"
    if len(rows2) == 1:
        return int(rows2[0]["id"]), None
    return (
        None,
        "Không tìm thấy nhân viên hoạt động — gán attendance_pin hoặc mã NV trùng mã máy",
    )


def _crm_parse_device_punch_datetime(payload: dict[str, Any]) -> tuple[str | None, str | None, str | None]:
    """Trả (work_date YYYY-MM-DD, hhmm, lỗi)."""
    pa = payload.get("punched_at")
    if pa is not None and str(pa).strip():
        raw = str(pa).strip().replace("Z", "+00:00")
        try:
            dt = datetime.fromisoformat(raw)
        except ValueError:
            m = re.match(
                r"^(\d{4}-\d{2}-\d{2})[ T](\d{1,2}):(\d{2})(?::(\d{2}))?",
                raw,
            )
            if not m:
                return None, None, "punched_at không đọc được (ISO-8601 hoặc YYYY-MM-DD HH:MM)"
            d = m.group(1)
            h, mi = int(m.group(2)), int(m.group(3))
            if not _crm_validate_date_ymd(d) or h > 23 or mi > 59:
                return None, None, "punched_at không hợp lệ"
            return d, f"{h:02d}:{mi:02d}", None
        else:
            wd = dt.strftime("%Y-%m-%d")
            if not _crm_validate_date_ymd(wd):
                return None, None, "Ngày không hợp lệ"
            return wd, f"{dt.hour:02d}:{dt.minute:02d}", None
    wd = str(payload.get("work_date") or "").strip()
    tm = str(payload.get("time") or "").strip()
    if not _crm_validate_date_ymd(wd):
        return None, None, "Cần punched_at hoặc work_date + time"
    hm = _crm_normalize_hhmm_safe(tm)
    if not hm:
        return None, None, "time phải dạng HH:MM"
    return wd, hm, None


def _crm_zkteco_serial_allowed(serial: str) -> bool:
    """Nếu CRM_ZKTECO_ALLOWED_SN được đặt — chỉ chấp nhận SN trong danh sách."""
    raw = (os.getenv("CRM_ZKTECO_ALLOWED_SN") or "").strip()
    if not raw:
        return True
    sn = (serial or "").strip()
    if not sn:
        return False
    allowed = {s.strip() for s in raw.split(",") if s.strip()}
    return sn in allowed


def _crm_apply_device_punch(
    conn: sqlite3.Connection,
    sid: int,
    wd_e: str,
    hm_e: str,
    kind: str,
    ts: str,
) -> sqlite3.Row:
    """Ghi / cập nhật một lần chấm công (vào/ra/auto) cho nhân viên + ngày."""
    date_err = _crm_attendance_work_date_error(wd_e)
    if date_err:
        raise ValueError(date_err)
    prev = conn.execute(
        """
        SELECT check_in, check_out, break_minutes, note
        FROM crm_attendance WHERE staff_id = ? AND work_date = ?
        """,
        (sid, wd_e),
    ).fetchone()

    check_in = str(prev["check_in"] or "").strip() if prev else ""
    check_out = str(prev["check_out"] or "").strip() if prev else ""
    brk = int(prev["break_minutes"] or 0) if prev else 0
    note = str(prev["note"] or "") if prev else ""

    ci, co = check_in, check_out
    if kind == "in":
        if not ci.strip():
            ci = hm_e
        else:
            ci = _crm_merge_hhmm_pair(ci, hm_e, earliest=True)
    elif kind == "out":
        if not co.strip():
            co = hm_e
        else:
            co = _crm_merge_hhmm_pair(co, hm_e, earliest=False)
    else:
        if not ci.strip():
            ci = hm_e
        elif not co.strip():
            co = hm_e
        else:
            co = _crm_merge_hhmm_pair(co, hm_e, earliest=False)

    conn.execute(
        """
        INSERT INTO crm_attendance (
            staff_id, work_date, check_in, check_out, break_minutes, note, created_at, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(staff_id, work_date) DO UPDATE SET
            check_in = excluded.check_in,
            check_out = excluded.check_out,
            break_minutes = excluded.break_minutes,
            note = excluded.note,
            updated_at = excluded.updated_at
        """,
        (sid, wd_e, ci, co, brk, note, ts, ts),
    )
    row = conn.execute(
        """
        SELECT a.*, s.name AS staff_name, s.internal_code AS staff_code
        FROM crm_attendance a
        JOIN crm_staff s ON s.id = a.staff_id
        WHERE a.staff_id = ? AND a.work_date = ?
        """,
        (sid, wd_e),
    ).fetchone()
    assert row is not None
    return row


def _crm_count_present_days(conn: sqlite3.Connection, staff_id: int, year: int, month: int) -> int:
    d0, d1 = _crm_month_bounds(year, month)
    row = conn.execute(
        """
        SELECT COUNT(*) AS n FROM crm_attendance
        WHERE staff_id = ? AND work_date >= ? AND work_date <= ?
          AND TRIM(check_in) != '' AND TRIM(check_out) != ''
        """,
        (staff_id, d0, d1),
    ).fetchone()
    return int(row["n"]) if row else 0


@app.get("/crm/payroll")
def crm_payroll_page() -> str:
    redir = _ensure_crm_session_html()
    if redir is not None:
        return redir
    staff_portal = _crm_staff_portal_active()
    if staff_portal:
        return render_template(
            "crm_payroll.html",
            attendance_device_url="",
            zkteco_iclock_url="",
            device_key_configured=False,
            crm_staff_portal=True,
            crm_staff_id=_staff_session_id(),
            crm_staff_name=_staff_session_name(),
            attendance_readonly=True,
        )
    au = (request.host_url or "").rstrip("/") + url_for("api_crm_attendance_device")
    iu = (request.host_url or "").rstrip("/") + url_for("iclock_cdata")
    return render_template(
        "crm_payroll.html",
        attendance_device_url=au,
        zkteco_iclock_url=iu,
        device_key_configured=bool(_crm_attendance_device_key_expected()),
        crm_staff_portal=False,
        crm_staff_id=None,
        crm_staff_name="",
        attendance_readonly=False,
        **_admin_page_template_kwargs(),
    )


@app.get("/crm/attendance")
def crm_attendance_redirect() -> Any:
    return redirect(url_for("crm_payroll_page"))


@app.get("/api/crm/attendance")
def api_crm_list_attendance() -> Any:
    staff_raw = (request.args.get("staff_id") or "").strip()
    date_from = (request.args.get("from") or "").strip()
    date_to = (request.args.get("to") or "").strip()
    staff_id: int | None = None
    if staff_raw:
        try:
            staff_id = int(staff_raw)
        except ValueError:
            return jsonify({"error": "staff_id không hợp lệ"}), 400
        if staff_id <= 0:
            staff_id = None
    eff = _crm_effective_staff_id()
    if eff is not None:
        if staff_id is not None and staff_id != eff:
            return jsonify({"error": "Chỉ xem chấm công của mình."}), 403
        staff_id = eff
    if date_from and not _crm_validate_date_ymd(date_from):
        return jsonify({"error": "from phải là YYYY-MM-DD"}), 400
    if date_to and not _crm_validate_date_ymd(date_to):
        return jsonify({"error": "to phải là YYYY-MM-DD"}), 400
    clauses: list[str] = []
    params: list[Any] = []
    if staff_id is not None:
        clauses.append("a.staff_id = ?")
        params.append(staff_id)
    if date_from:
        clauses.append("a.work_date >= ?")
        params.append(date_from)
    if date_to:
        clauses.append("a.work_date <= ?")
        params.append(date_to)
    where_sql = (" WHERE " + " AND ".join(clauses)) if clauses else ""
    with get_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT a.*, s.name AS staff_name, s.internal_code AS staff_code
            FROM crm_attendance a
            JOIN crm_staff s ON s.id = a.staff_id
            {where_sql}
            ORDER BY a.work_date DESC, s.name COLLATE NOCASE ASC
            """,
            params,
        ).fetchall()
        policy = load_policy(conn)
    enriched = [enrich_attendance_row(dict(r), policy) for r in rows]
    return jsonify({"attendance": enriched})


@app.post("/api/crm/attendance")
def api_crm_upsert_attendance() -> Any:
    if _staff_logged_in():
        return jsonify({"error": "Nhân viên chỉ được xem bảng chấm công."}), 403
    if not _admin_section_can("crm_payroll_attendance", "create"):
        return _admin_section_forbidden_json("crm_payroll_attendance", "create")
    payload = request.get_json(force=True) or {}
    try:
        sid = int(payload.get("staff_id") or 0)
    except (TypeError, ValueError):
        sid = 0
    work_date = str(payload.get("work_date") or "").strip()
    if sid <= 0 or not work_date:
        return jsonify({"error": "Cần staff_id và work_date (YYYY-MM-DD)"}), 400
    date_err = _crm_attendance_work_date_error(work_date)
    if date_err:
        return jsonify({"error": date_err}), 400
    check_in = str(payload.get("check_in") or "").strip()[:16]
    check_out = str(payload.get("check_out") or "").strip()[:16]
    try:
        brk = int(payload.get("break_minutes") or 0)
    except (TypeError, ValueError):
        brk = 0
    brk = max(0, min(brk, 24 * 60))
    note = str(payload.get("note") or "").strip()[:2000]
    ts = _crm_ts()
    with get_connection() as conn:
        if conn.execute("SELECT id FROM crm_staff WHERE id = ?", (sid,)).fetchone() is None:
            return jsonify({"error": "Không tìm thấy nhân viên"}), 404
        conn.execute(
            """
            INSERT INTO crm_attendance (
                staff_id, work_date, check_in, check_out, break_minutes, note, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(staff_id, work_date) DO UPDATE SET
                check_in = excluded.check_in,
                check_out = excluded.check_out,
                break_minutes = excluded.break_minutes,
                note = excluded.note,
                updated_at = excluded.updated_at
            """,
            (sid, work_date, check_in, check_out, brk, note, ts, ts),
        )
        row = conn.execute(
            """
            SELECT a.*, s.name AS staff_name, s.internal_code AS staff_code
            FROM crm_attendance a
            JOIN crm_staff s ON s.id = a.staff_id
            WHERE a.staff_id = ? AND a.work_date = ?
            """,
            (sid, work_date),
        ).fetchone()
        policy = load_policy(conn)
    assert row is not None
    return jsonify(enrich_attendance_row(dict(row), policy)), 201


def _crm_upsert_attendance_import_row(
    conn: sqlite3.Connection,
    *,
    staff_id: int,
    work_date: str,
    check_in: str,
    check_out: str,
    ts: str,
) -> None:
    prev = conn.execute(
        """
        SELECT check_in, check_out, break_minutes, note
        FROM crm_attendance WHERE staff_id = ? AND work_date = ?
        """,
        (staff_id, work_date),
    ).fetchone()
    ci = check_in or (str(prev["check_in"] or "").strip() if prev else "")
    co = check_out or (str(prev["check_out"] or "").strip() if prev else "")
    brk = int(prev["break_minutes"] or 0) if prev else 0
    note = str(prev["note"] or "") if prev else ""
    conn.execute(
        """
        INSERT INTO crm_attendance (
            staff_id, work_date, check_in, check_out, break_minutes, note, created_at, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(staff_id, work_date) DO UPDATE SET
            check_in = excluded.check_in,
            check_out = excluded.check_out,
            break_minutes = excluded.break_minutes,
            note = excluded.note,
            updated_at = excluded.updated_at
        """,
        (staff_id, work_date, ci, co, brk, note, ts, ts),
    )


@app.post("/api/crm/attendance/import")
def api_crm_import_attendance() -> Any:
    """
    Nhập chấm công từ file Excel (.xlsx) — mẫu «Bảng chấm công chi tiết».
    Cột: Mã PIN → attendance_pin; Ngày; Vào; Ra.
    """
    if _staff_logged_in():
        return jsonify({"error": "Nhân viên không được nhập file chấm công."}), 403
    if not _admin_section_can("crm_payroll_attendance", "create"):
        return _admin_section_forbidden_json("crm_payroll_attendance", "create")

    upload = request.files.get("file")
    if upload is None or not upload.filename:
        return jsonify({"error": "Cần chọn file Excel (.xlsx)"}), 400
    fn = str(upload.filename or "").lower()
    if not fn.endswith(".xlsx"):
        return jsonify({"error": "Chỉ hỗ trợ file .xlsx"}), 400

    raw = upload.read()
    if not raw:
        return jsonify({"error": "File rỗng"}), 400
    if len(raw) > 8 * 1024 * 1024:
        return jsonify({"error": "File quá lớn (tối đa 8 MB)"}), 400

    records, parse_errors = parse_timesheet_xlsx(raw)
    if not records and parse_errors:
        return jsonify({"error": parse_errors[0], "errors": parse_errors}), 400

    ts = _crm_ts()
    imported = 0
    skipped = 0
    errors: list[str] = list(parse_errors)
    applied_keys: set[tuple[int, str]] = set()

    with get_connection() as conn:
        for rec in records:
            pin = str(rec.get("pin") or "").strip()
            wd = str(rec.get("work_date") or "").strip()
            row_no = int(rec.get("row") or 0)
            sid, err_s = _crm_resolve_staff_by_device_id(conn, pin)
            if sid is None:
                errors.append(f"Dòng {row_no}: PIN «{pin}» — {err_s or 'không khớp nhân viên'}")
                skipped += 1
                continue
            date_err = _crm_attendance_work_date_error(wd)
            if date_err:
                errors.append(f"Dòng {row_no}: PIN «{pin}» — {date_err}")
                skipped += 1
                continue
            key = (sid, wd)
            if key in applied_keys:
                skipped += 1
                continue
            applied_keys.add(key)
            _crm_upsert_attendance_import_row(
                conn,
                staff_id=sid,
                work_date=wd,
                check_in=str(rec.get("check_in") or "").strip(),
                check_out=str(rec.get("check_out") or "").strip(),
                ts=ts,
            )
            imported += 1

    return jsonify(
        {
            "ok": True,
            "imported": imported,
            "skipped": skipped,
            "parsed_rows": len(records),
            "errors": errors[:50],
            "message": (
                f"Đã nhập {imported} dòng chấm công"
                + (f", bỏ qua {skipped}" if skipped else "")
                + (f" ({len(errors)} cảnh báo)" if errors else "")
                + "."
            ),
        }
    )


@app.get("/api/crm/payroll/policy")
def api_crm_get_payroll_policy() -> Any:
    if not _admin_section_can("crm_payroll_salary", "view"):
        return _admin_section_forbidden_json("crm_payroll_salary", "view")
    with get_connection() as conn:
        policy = load_policy(conn)
    return jsonify({"policy": policy_for_api(policy)})


@app.put("/api/crm/payroll/policy")
def api_crm_put_payroll_policy() -> Any:
    if not _admin_section_can("crm_payroll_salary", "edit"):
        return _admin_section_forbidden_json("crm_payroll_salary", "edit")
    payload = request.get_json(force=True) or {}

    def _int_field(key: str, default: int, lo: int, hi: int) -> int:
        try:
            v = int(payload.get(key, default))
        except (TypeError, ValueError):
            v = default
        return max(lo, min(v, hi))

    def _float_field(key: str, default: float, lo: float, hi: float) -> float:
        try:
            v = float(payload.get(key, default))
        except (TypeError, ValueError):
            v = default
        return max(lo, min(v, hi))

    raw_shifts = payload.get("weekday_shifts")
    if isinstance(raw_shifts, list) and raw_shifts:
        shifts = normalize_weekday_shifts(raw_shifts)
    else:
        weekdays_raw = str(payload.get("work_weekdays") or "0,1,2,3,4").strip()
        work_set = parse_work_weekdays(weekdays_raw)
        shifts = default_weekday_shifts(
            work_weekdays=work_set,
            shift_start=str(payload.get("shift_start") or "08:30").strip()[:5],
            shift_end=str(payload.get("shift_end") or "17:30").strip()[:5],
            break_minutes=_int_field("break_minutes_default", 60, 0, 24 * 60),
            standard_hours=_float_field("standard_hours_per_day", 8.0, 0.5, 24.0),
        )
    weekdays_raw = work_weekdays_from_shifts(shifts)
    shifts_json = weekday_shifts_json(shifts)
    first_work = next((s for s in shifts if s.get("work")), shifts[0])
    shift_start = str(first_work.get("shift_start") or "08:30").strip()[:5]
    shift_end = str(first_work.get("shift_end") or "17:30").strip()[:5]
    break_default = max(0, min(int(first_work.get("break_minutes") or 60), 24 * 60))
    std_hours_day = max(0.5, min(float(first_work.get("standard_hours") or 8), 24.0))
    bonus_mode = str(payload.get("bonus_mode") or "attendance").strip().lower()
    if bonus_mode not in ("attendance", "none"):
        bonus_mode = "attendance"

    ts = _crm_ts()
    with get_connection() as conn:
        ensure_payroll_policy_schema(conn)
        conn.execute(
            """
            UPDATE crm_payroll_policy SET
                work_weekdays = ?,
                shift_start = ?,
                shift_end = ?,
                break_minutes_default = ?,
                late_grace_minutes = ?,
                late_penalty_vnd_per_min = ?,
                late_penalty_max_vnd = ?,
                standard_hours_per_day = ?,
                bonus_mode = ?,
                bonus_pct = ?,
                bonus_min_days = ?,
                overtime_multiplier = ?,
                weekday_shifts = ?,
                updated_at = ?
            WHERE id = 1
            """,
            (
                weekdays_raw,
                shift_start,
                shift_end,
                break_default,
                _int_field("late_grace_minutes", 5, 0, 120),
                _int_field("late_penalty_vnd_per_min", 5000, 0, 50_000_000),
                _int_field("late_penalty_max_vnd", 200_000, 0, 500_000_000),
                std_hours_day,
                bonus_mode,
                _float_field("bonus_pct", 5.0, 0.0, 100.0),
                _int_field("bonus_min_days", 20, 0, 31),
                _float_field("overtime_multiplier", 1.5, 1.0, 3.0),
                shifts_json,
                ts,
            ),
        )
        policy = load_policy(conn)
    return jsonify({"policy": policy_for_api(policy)})


@app.get("/api/crm/payroll/position-rates")
def api_crm_get_position_rates() -> Any:
    if not _admin_section_can("crm_payroll_salary", "view"):
        return _admin_section_forbidden_json("crm_payroll_salary", "view")
    with get_connection() as conn:
        pos_map = load_position_payroll_map(conn)
    rows = sorted(pos_map.values(), key=lambda r: (int(r.get("rank_level") or 0), str(r.get("position_code") or "")))
    return jsonify({"positions": rows})


@app.put("/api/crm/payroll/position-rates")
def api_crm_put_position_rates() -> Any:
    if not _admin_section_can("crm_payroll_salary", "edit"):
        return _admin_section_forbidden_json("crm_payroll_salary", "edit")
    payload = request.get_json(force=True) or {}
    items = payload.get("positions")
    if not isinstance(items, list):
        return jsonify({"error": "Cần mảng positions"}), 400
    ts = _crm_ts()
    with get_connection() as conn:
        ensure_payroll_policy_schema(conn)
        for item in items:
            if not isinstance(item, dict):
                continue
            try:
                pid = int(item.get("position_id") or 0)
            except (TypeError, ValueError):
                continue
            if pid <= 0:
                continue
            if conn.execute("SELECT id FROM crm_positions WHERE id = ?", (pid,)).fetchone() is None:
                continue
            try:
                rank = max(1, min(int(item.get("rank_level") or 1), 99))
            except (TypeError, ValueError):
                rank = 1
            try:
                allow = max(0, min(int(item.get("allowance_vnd") or 0), 999_999_999))
            except (TypeError, ValueError):
                allow = 0
            try:
                bp = max(0.0, min(float(item.get("bonus_pct") or 0), 100.0))
            except (TypeError, ValueError):
                bp = 0.0
            conn.execute(
                """
                INSERT INTO crm_position_payroll (position_id, rank_level, allowance_vnd, bonus_pct, updated_at)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(position_id) DO UPDATE SET
                    rank_level = excluded.rank_level,
                    allowance_vnd = excluded.allowance_vnd,
                    bonus_pct = excluded.bonus_pct,
                    updated_at = excluded.updated_at
                """,
                (pid, rank, allow, bp, ts),
            )
        pos_map = load_position_payroll_map(conn)
    rows = sorted(pos_map.values(), key=lambda r: (int(r.get("rank_level") or 0), str(r.get("position_code") or "")))
    return jsonify({"positions": rows})


@app.get("/api/crm/payroll/dashboard")
def api_crm_payroll_dashboard() -> Any:
    if not _admin_section_can("crm_payroll_salary", "view"):
        return _admin_section_forbidden_json("crm_payroll_salary", "view")
    try:
        year = int(request.args.get("year") or 0)
        month = int(request.args.get("month") or 0)
    except (TypeError, ValueError):
        return jsonify({"error": "year/month không hợp lệ"}), 400
    if year < 2000 or year > 2100 or month < 1 or month > 12:
        return jsonify({"error": "Kỳ không hợp lệ"}), 400
    with get_connection() as conn:
        policy = load_policy(conn)
        summary = dashboard_summary(conn, year=year, month=month, policy=policy)
        pos_map = load_position_payroll_map(conn)
    summary["position_rates"] = sorted(
        pos_map.values(),
        key=lambda r: (int(r.get("rank_level") or 0), str(r.get("position_code") or "")),
    )
    return jsonify(summary)


@app.post("/api/crm/attendance/device")
def api_crm_attendance_device() -> Any:
    """
    Chấm công từ máy vân tay / nhận diện khuôn mặt / phần mềm trung gian (HTTP, không cần cookie admin).
    Đặt biến môi trường CRM_ATTENDANCE_DEVICE_KEY; gửi header X-CRM-Device-Key hoặc Authorization: Bearer.

    JSON: pin hoặc staff_code (khớp attendance_pin ưu tiên, sau đó internal_code);
    punched_at (ISO) hoặc work_date + time (HH:MM);
    kind: in | out | auto (mặc định auto: lần đầu = vào, sau = ra; cả hai có rồi thì cập nhật giờ ra muộn nhất).
    verify (tuỳ chọn): fingerprint | face | card — chỉ trả lại trong response.
    """
    if not _crm_attendance_device_key_expected():
        return jsonify({"error": "Chưa cấu hình CRM_ATTENDANCE_DEVICE_KEY trên server"}), 503
    if not _crm_attendance_device_auth_ok(request):
        return jsonify({"error": "Không hợp lệ hoặc thiếu khóa thiết bị"}), 401

    payload = request.get_json(force=True) or {}
    pid_raw = str(payload.get("pin") or payload.get("staff_code") or "").strip()
    kind = str(payload.get("kind") or "auto").strip().lower()
    if kind not in ("in", "out", "auto"):
        kind = "auto"
    verify = str(payload.get("verify") or "").strip().lower()[:32]

    wd_e, hm_e, err_t = _crm_parse_device_punch_datetime(payload)
    if err_t or not wd_e or not hm_e:
        return jsonify({"error": err_t or "Thiếu thời gian chấm"}), 400

    ts = _crm_ts()
    with get_connection() as conn:
        sid, err_s = _crm_resolve_staff_by_device_id(conn, pid_raw)
        if sid is None:
            return jsonify({"error": err_s or "Không xác định nhân viên"}), 404

        try:
            row = _crm_apply_device_punch(conn, sid, wd_e, hm_e, kind, ts)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    out: dict[str, Any] = dict(row)
    out["kind_applied"] = kind
    if verify:
        out["verify"] = verify
    return jsonify(out), 200


@app.get("/iclock/getrequest")
def iclock_getrequest() -> Response:
    """ZKTeco iClock — máy hỏi lệnh (MB20-VL, MB560-VL, ZMM…)."""
    sn = (request.args.get("SN") or request.args.get("sn") or "").strip()
    if not _crm_zkteco_serial_allowed(sn):
        return Response("ERROR", status=403, mimetype="text/plain")
    return Response(iclock_get_response(), mimetype="text/plain")


@app.route("/iclock/cdata", methods=["GET", "POST"])
def iclock_cdata() -> Response:
    """
    ZKTeco iClock PUSH — nhận ATTLOG từ máy chấm công (MB20-VL nhận diện khuôn mặt).
    Cấu hình trên máy: Server URL = http://<IP-máy-chủ>:5050/iclock/cdata
    Tuỳ chọn: CRM_ZKTECO_ALLOWED_SN=SN1,SN2 (serial in trên máy).
    """
    sn = (request.args.get("SN") or request.args.get("sn") or "").strip()
    if not _crm_zkteco_serial_allowed(sn):
        return Response("ERROR", status=403, mimetype="text/plain")

    if request.method == "GET":
        if request.args.get("options") or request.args.get("table") == "options":
            return Response(iclock_options_response(sn), mimetype="text/plain")
        return Response(iclock_get_response(), mimetype="text/plain")

    table = (request.args.get("table") or "").strip().upper()
    body = request.get_data(as_text=True) or ""

    if table and table not in ("ATTLOG", "OPERLOG", "ATTPHOTO", "OPTIONS"):
        return Response(iclock_get_response(), mimetype="text/plain")

    if table == "OPERLOG" or (not table and "USER PIN=" in body.upper()):
        return Response(iclock_get_response(), mimetype="text/plain")

    logs = parse_attlog_body(body)
    if not logs:
        return Response(iclock_get_response(), mimetype="text/plain")

    ts = _crm_ts()
    applied = 0
    skipped = 0
    with get_connection() as conn:
        for entry in logs:
            sid, _err = _crm_resolve_staff_by_device_id(conn, entry.pin)
            if sid is None:
                skipped += 1
                continue
            if _crm_attendance_work_date_error(entry.work_date):
                skipped += 1
                continue
            _crm_apply_device_punch(conn, sid, entry.work_date, entry.time_hm, entry.kind, ts)
            applied += 1

    return Response(iclock_get_response(), mimetype="text/plain")


@app.get("/api/crm/payroll")
def api_crm_get_payroll() -> Any:
    try:
        year = int(request.args.get("year") or 0)
        month = int(request.args.get("month") or 0)
    except (TypeError, ValueError):
        return jsonify({"error": "year/month không hợp lệ"}), 400
    if year < 2000 or year > 2100 or month < 1 or month > 12:
        return jsonify({"error": "Kỳ không hợp lệ"}), 400
    with get_connection() as conn:
        pr = conn.execute(
            "SELECT * FROM crm_payroll WHERE year = ? AND month = ?",
            (year, month),
        ).fetchone()
        if pr is None:
            return jsonify({"payroll": None, "lines": []})
        pid = int(pr["id"])
        lines = conn.execute(
            """
            SELECT pl.*, s.name AS staff_name, s.internal_code AS staff_code
            FROM crm_payroll_line pl
            JOIN crm_staff s ON s.id = pl.staff_id
            WHERE pl.payroll_id = ?
            ORDER BY s.name COLLATE NOCASE ASC
            """,
            (pid,),
        ).fetchall()
    return jsonify({"payroll": dict(pr), "lines": rows_to_dict(lines)})


_PAYROLL_EXPORT_HEADERS = [
    "Kỳ",
    "Quý",
    "Mã NV",
    "Họ tên",
    "Ngày công",
    "Giờ làm",
    "Trễ (phút)",
    "Lương CB",
    "Lương theo giờ",
    "PC cấp bậc",
    "Thưởng",
    "Phạt trễ",
    "Tổng phụ cấp",
    "Tổng khấu trừ",
    "Thực lĩnh",
    "Trạng thái kỳ",
    "Ghi chú",
]

_PAYROLL_EXPORT_SUMMARY_HEADERS = [
    "Mã NV",
    "Họ tên",
    "Số tháng",
    "Tổng ngày công",
    "Tổng giờ làm",
    "Tổng trễ (phút)",
    "Tổng lương theo giờ",
    "Tổng PC cấp bậc",
    "Tổng thưởng",
    "Tổng phạt trễ",
    "Tổng phụ cấp",
    "Tổng khấu trừ",
    "Tổng thực lĩnh",
]


def _crm_payroll_status_label(raw: str | None) -> str:
    s = str(raw or "draft").strip().lower()
    return "Đã khóa" if s == "final" else "Nháp"


def _crm_quarter_label(year: int, month: int) -> str:
    q = (month - 1) // 3 + 1
    return f"Q{q}/{year}"


def _crm_parse_payroll_export_period() -> tuple[tuple[int, int, int, int], str, str | None]:
    """
    Trả về (y0, m0, y1, m1), period_kind, error.
    period: month | quarter | range
    """
    period = str(request.args.get("period") or "month").strip().lower()
    if period not in ("month", "quarter", "range"):
        period = "month"

    try:
        year = int(request.args.get("year") or 0)
    except (TypeError, ValueError):
        year = 0

    if period == "month":
        try:
            month = int(request.args.get("month") or 0)
        except (TypeError, ValueError):
            month = 0
        if year < 2000 or year > 2100 or month < 1 or month > 12:
            return (0, 0, 0, 0), period, "Cần year và month hợp lệ (kỳ tháng)"
        return (year, month, year, month), period, None

    if period == "quarter":
        try:
            quarter = int(request.args.get("quarter") or 0)
        except (TypeError, ValueError):
            quarter = 0
        if year < 2000 or year > 2100 or quarter < 1 or quarter > 4:
            return (0, 0, 0, 0), period, "Cần year và quarter (1–4) hợp lệ"
        m0 = (quarter - 1) * 3 + 1
        m1 = m0 + 2
        return (year, m0, year, m1), period, None

    date_from = str(request.args.get("from") or "").strip()
    date_to = str(request.args.get("to") or "").strip()
    if not _crm_validate_date_ymd(date_from) or not _crm_validate_date_ymd(date_to):
        return (0, 0, 0, 0), period, "from và to phải là YYYY-MM-DD"
    if date_from > date_to:
        return (0, 0, 0, 0), period, "from phải ≤ to"
    y0, m0 = int(date_from[:4]), int(date_from[5:7])
    y1, m1 = int(date_to[:4]), int(date_to[5:7])
    if y0 < 2000 or y1 > 2100:
        return (0, 0, 0, 0), period, "Khoảng năm không hợp lệ"
    return (y0, m0, y1, m1), period, None


def _crm_fetch_payroll_export_rows(
    conn: sqlite3.Connection,
    *,
    y0: int,
    m0: int,
    y1: int,
    m1: int,
    staff_id: int | None = None,
    staff_q: str | None = None,
) -> list[sqlite3.Row]:
    clauses = ["(p.year > ? OR (p.year = ? AND p.month >= ?))", "(p.year < ? OR (p.year = ? AND p.month <= ?))"]
    params: list[Any] = [y0, y0, m0, y1, y1, m1]
    if staff_id is not None:
        clauses.append("pl.staff_id = ?")
        params.append(staff_id)
    elif staff_q:
        staff_ids = find_payroll_staff_ids_by_query(conn, staff_q)
        if not staff_ids:
            return []
        placeholders = ",".join("?" * len(staff_ids))
        clauses.append(f"pl.staff_id IN ({placeholders})")
        params.extend(staff_ids)
    where_sql = " AND ".join(clauses)
    return conn.execute(
        f"""
        SELECT pl.*,
               s.name AS staff_name, s.internal_code AS staff_code,
               p.year AS payroll_year, p.month AS payroll_month,
               p.status AS payroll_status, p.workdays_standard
        FROM crm_payroll_line pl
        JOIN crm_payroll p ON p.id = pl.payroll_id
        JOIN crm_staff s ON s.id = pl.staff_id
        WHERE {where_sql}
        ORDER BY p.year ASC, p.month ASC, s.name COLLATE NOCASE ASC
        """,
        params,
    ).fetchall()


def _crm_payroll_export_row_values(row: sqlite3.Row) -> list[Any]:
    d = dict(row)
    py = int(d.get("payroll_year") or 0)
    pm = int(d.get("payroll_month") or 0)
    period_label = f"{pm:02d}/{py}" if py and pm else "—"
    pos_allow = int(d.get("position_allowance_vnd") or 0)
    bonus = int(d.get("bonus_vnd") or 0)
    late_ded = int(d.get("late_deduction_vnd") or 0)
    return [
        period_label,
        _crm_quarter_label(py, pm) if py and pm else "—",
        str(d.get("staff_code") or "").strip(),
        str(d.get("staff_name") or "").strip(),
        int(d.get("days_present") or 0),
        float(d.get("hours_worked_total") or 0),
        int(d.get("late_minutes_total") or 0),
        int(d.get("base_salary_vnd") or 0),
        int(d.get("salary_from_attendance_vnd") or 0),
        pos_allow,
        bonus,
        late_ded,
        int(d.get("allowances_vnd") or 0),
        int(d.get("deductions_vnd") or 0),
        int(d.get("net_salary_vnd") or 0),
        _crm_payroll_status_label(str(d.get("payroll_status") or "")),
        str(d.get("note") or "").strip(),
    ]


def _crm_payroll_export_summary_rows(rows: list[sqlite3.Row]) -> list[list[Any]]:
    agg: dict[int, dict[str, Any]] = {}
    for r in rows:
        d = dict(r)
        sid = int(d.get("staff_id") or 0)
        if sid <= 0:
            continue
        if sid not in agg:
            agg[sid] = {
                "staff_code": str(d.get("staff_code") or "").strip(),
                "staff_name": str(d.get("staff_name") or "").strip(),
                "months": set(),
                "days_present": 0,
                "hours_worked_total": 0.0,
                "late_minutes_total": 0,
                "salary_from_attendance_vnd": 0,
                "position_allowance_vnd": 0,
                "bonus_vnd": 0,
                "late_deduction_vnd": 0,
                "allowances_vnd": 0,
                "deductions_vnd": 0,
                "net_salary_vnd": 0,
            }
        a = agg[sid]
        py = int(d.get("payroll_year") or 0)
        pm = int(d.get("payroll_month") or 0)
        if py and pm:
            a["months"].add(f"{py}-{pm:02d}")
        a["days_present"] += int(d.get("days_present") or 0)
        a["hours_worked_total"] += float(d.get("hours_worked_total") or 0)
        a["late_minutes_total"] += int(d.get("late_minutes_total") or 0)
        a["salary_from_attendance_vnd"] += int(d.get("salary_from_attendance_vnd") or 0)
        a["position_allowance_vnd"] += int(d.get("position_allowance_vnd") or 0)
        a["bonus_vnd"] += int(d.get("bonus_vnd") or 0)
        a["late_deduction_vnd"] += int(d.get("late_deduction_vnd") or 0)
        a["allowances_vnd"] += int(d.get("allowances_vnd") or 0)
        a["deductions_vnd"] += int(d.get("deductions_vnd") or 0)
        a["net_salary_vnd"] += int(d.get("net_salary_vnd") or 0)

    out: list[list[Any]] = []
    for sid in sorted(agg.keys(), key=lambda k: agg[k]["staff_name"].casefold()):
        a = agg[sid]
        out.append(
            [
                a["staff_code"],
                a["staff_name"],
                len(a["months"]),
                a["days_present"],
                round(a["hours_worked_total"], 2),
                a["late_minutes_total"],
                a["salary_from_attendance_vnd"],
                a["position_allowance_vnd"],
                a["bonus_vnd"],
                a["late_deduction_vnd"],
                a["allowances_vnd"],
                a["deductions_vnd"],
                a["net_salary_vnd"],
            ]
        )
    return out


def _crm_payroll_export_filename(period: str, y0: int, m0: int, y1: int, m1: int) -> str:
    if period == "month":
        return f"crm-luong-{y0}-{m0:02d}"
    if period == "quarter":
        q = (m0 - 1) // 3 + 1
        return f"crm-luong-Q{q}-{y0}"
    return f"crm-luong-{y0}-{m0:02d}_{y1}-{m1:02d}"


def _crm_payroll_export_csv(rows: list[sqlite3.Row], *, filename: str, include_summary: bool) -> Response:
    si = StringIO()
    w = csv.writer(si)
    w.writerow(_PAYROLL_EXPORT_HEADERS)
    for r in rows:
        w.writerow(_crm_payroll_export_row_values(r))
    if include_summary and rows:
        w.writerow([])
        w.writerow(["--- Tổng hợp ---"])
        w.writerow(_PAYROLL_EXPORT_SUMMARY_HEADERS)
        for line in _crm_payroll_export_summary_rows(rows):
            w.writerow(line)
    raw = si.getvalue().encode("utf-8-sig")
    safe = filename.replace('"', "")
    return Response(
        raw,
        mimetype="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": (f'attachment; filename="{safe}"; filename*=UTF-8\'\'{quote(filename)}'),
        },
    )


def _crm_safe_xlsx_sheet_title(title: str) -> str:
    """Excel cấm \\ / ? * [ ] : trong tên sheet."""
    out = str(title or "Sheet")
    for ch in ("\\", "/", "?", "*", "[", "]", ":"):
        out = out.replace(ch, "-")
    out = out.strip() or "Sheet"
    return out[:31]


def _crm_payroll_export_xlsx(rows: list[sqlite3.Row], *, filename: str, include_summary: bool, sheet_title: str) -> Response:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = _crm_safe_xlsx_sheet_title(sheet_title)
    ws.append(list(_PAYROLL_EXPORT_HEADERS))
    for r in rows:
        ws.append(_crm_payroll_export_row_values(r))
    if include_summary and rows:
        ws2 = wb.create_sheet("Tổng hợp")
        ws2.append(list(_PAYROLL_EXPORT_SUMMARY_HEADERS))
        for line in _crm_payroll_export_summary_rows(rows):
            ws2.append(line)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
        etag=False,
        max_age=0,
        conditional=False,
    )


@app.get("/api/crm/payroll/export")
def api_crm_export_payroll() -> Any:
    """Xuất bảng lương theo tháng, quý hoặc khoảng thời gian (CSV / Excel)."""
    if not (
        _admin_section_can("crm_payroll_salary", "export")
        or _admin_section_can("crm_payroll_salary", "view")
        or _admin_section_can("crm_payroll_salary", "edit")
    ):
        return _admin_section_forbidden_json("crm_payroll_salary", "export")
    fmt = str(request.args.get("format") or "xlsx").strip().lower()
    if fmt not in ("csv", "xlsx"):
        fmt = "xlsx"

    (y0, m0, y1, m1), period, err = _crm_parse_payroll_export_period()
    if err:
        return jsonify({"error": err}), 400

    staff_raw = str(request.args.get("staff_id") or "").strip()
    staff_id: int | None = None
    if staff_raw:
        try:
            staff_id = int(staff_raw)
        except ValueError:
            return jsonify({"error": "staff_id không hợp lệ"}), 400
        if staff_id <= 0:
            staff_id = None

    staff_q = str(request.args.get("q") or "").strip()
    if staff_id is not None and staff_q:
        staff_q = ""

    portal_sid = _crm_effective_staff_id()
    if portal_sid is not None:
        staff_id = portal_sid
        staff_q = ""

    with get_connection() as conn:
        rows = _crm_fetch_payroll_export_rows(
            conn, y0=y0, m0=m0, y1=y1, m1=m1, staff_id=staff_id, staff_q=staff_q or None
        )

    include_summary = period in ("quarter", "range") or (y0, m0) != (y1, m1)
    base = _crm_payroll_export_filename(period, y0, m0, y1, m1)
    if period == "month":
        sheet_title = f"Lương {m0:02d}-{y0}"
    elif period == "quarter":
        q = (m0 - 1) // 3 + 1
        sheet_title = f"Lương Q{q}-{y0}"
    else:
        sheet_title = f"Lương {m0:02d}-{y0} đến {m1:02d}-{y1}"

    if not rows:
        period_hint = f"{m0:02d}/{y0}" if period == "month" else f"{m0:02d}/{y0}–{m1:02d}/{y1}"
        return (
            jsonify(
                {
                    "error": (
                        f"Không có dữ liệu lương cho kỳ {period_hint}. "
                        "Hãy chọn đúng Năm/Tháng, bấm «Tính / cập nhật lương» rồi xuất lại."
                    )
                }
            ),
            400,
        )

    try:
        if fmt == "csv":
            return _crm_payroll_export_csv(rows, filename=f"{base}.csv", include_summary=include_summary)
        return _crm_payroll_export_xlsx(
            rows, filename=f"{base}.xlsx", include_summary=include_summary, sheet_title=sheet_title
        )
    except Exception as exc:
        return jsonify({"error": f"Không tạo được file xuất: {exc}"}), 500


@app.post("/api/crm/payroll/compute")
def api_crm_compute_payroll() -> Any:
    if not _admin_section_can("crm_payroll_salary", "edit"):
        return _admin_section_forbidden_json("crm_payroll_salary", "edit")
    payload = request.get_json(force=True) or {}
    try:
        year = int(payload.get("year") or 0)
        month = int(payload.get("month") or 0)
    except (TypeError, ValueError):
        return jsonify({"error": "year/month không hợp lệ"}), 400
    if year < 2000 or year > 2100 or month < 1 or month > 12:
        return jsonify({"error": "Kỳ không hợp lệ"}), 400
    standard = _crm_weekdays_in_month(year, month)
    ts = _crm_ts()
    with get_connection() as conn:
        prev = conn.execute(
            "SELECT * FROM crm_payroll WHERE year = ? AND month = ?",
            (year, month),
        ).fetchone()
        if prev is not None and str(prev["status"] or "").strip() == "final":
            return jsonify({"error": "Kỳ lương đã khóa. Đặt về nháp (PATCH) để tính lại."}), 409
        if prev is None:
            conn.execute(
                """
                INSERT INTO crm_payroll (year, month, workdays_standard, status, created_at, updated_at)
                VALUES (?, ?, ?, 'draft', ?, ?)
                """,
                (year, month, standard, ts, ts),
            )
            pr_row = conn.execute(
                "SELECT * FROM crm_payroll WHERE year = ? AND month = ?",
                (year, month),
            ).fetchone()
        else:
            conn.execute(
                "UPDATE crm_payroll SET workdays_standard = ?, updated_at = ? WHERE id = ?",
                (standard, ts, int(prev["id"])),
            )
            pr_row = conn.execute(
                "SELECT * FROM crm_payroll WHERE id = ?", (int(prev["id"]),)
            ).fetchone()
        assert pr_row is not None
        pid = int(pr_row["id"])

        policy = load_policy(conn)
        position_map = load_position_payroll_map(conn)

        staff_rows = conn.execute(
            """
            SELECT id, name, base_salary_vnd, position_id
            FROM crm_staff WHERE active = 1 ORDER BY name COLLATE NOCASE
            """
        ).fetchall()

        for sr in staff_rows:
            st_id = int(sr["id"])
            base = int(sr["base_salary_vnd"] or 0)
            pos_id = int(sr["position_id"]) if sr["position_id"] is not None else None
            computed = compute_staff_payroll(
                conn,
                staff_id=st_id,
                base_salary_vnd=base,
                position_id=pos_id,
                year=year,
                month=month,
                policy=policy,
                position_map=position_map,
            )
            existing = conn.execute(
                """
                SELECT id, allowances_vnd, deductions_vnd, note,
                       position_allowance_vnd, bonus_vnd, late_deduction_vnd
                FROM crm_payroll_line WHERE payroll_id = ? AND staff_id = ?
                """,
                (pid, st_id),
            ).fetchone()

            auto_allow = int(computed["position_allowance_vnd"]) + int(computed["bonus_vnd"])
            auto_ded = int(computed["late_deduction_vnd"])
            manual_allow = 0
            manual_ded = 0
            note = ""
            if existing:
                note = str(existing["note"] or "")
                prev_auto_allow = int(existing["position_allowance_vnd"] or 0) + int(
                    existing["bonus_vnd"] or 0
                )
                prev_auto_ded = int(existing["late_deduction_vnd"] or 0)
                manual_allow = max(0, int(existing["allowances_vnd"] or 0) - prev_auto_allow)
                manual_ded = max(0, int(existing["deductions_vnd"] or 0) - prev_auto_ded)

            allow = auto_allow + manual_allow
            ded = auto_ded + manual_ded
            salary_att = int(computed["salary_from_attendance_vnd"])
            net = salary_att + allow - ded
            days = int(computed["days_present"])

            if existing:
                conn.execute(
                    """
                    UPDATE crm_payroll_line SET
                      days_present = ?, base_salary_vnd = ?, salary_from_attendance_vnd = ?,
                      hours_worked_total = ?, late_minutes_total = ?, late_deduction_vnd = ?,
                      position_allowance_vnd = ?, bonus_vnd = ?,
                      allowances_vnd = ?, deductions_vnd = ?, net_salary_vnd = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        days,
                        base,
                        salary_att,
                        computed["hours_worked_total"],
                        computed["late_minutes_total"],
                        computed["late_deduction_vnd"],
                        computed["position_allowance_vnd"],
                        computed["bonus_vnd"],
                        allow,
                        ded,
                        net,
                        ts,
                        int(existing["id"]),
                    ),
                )
            else:
                conn.execute(
                    """
                    INSERT INTO crm_payroll_line (
                      payroll_id, staff_id, days_present, base_salary_vnd,
                      salary_from_attendance_vnd, hours_worked_total, late_minutes_total,
                      late_deduction_vnd, position_allowance_vnd, bonus_vnd,
                      allowances_vnd, deductions_vnd, net_salary_vnd,
                      note, created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        pid,
                        st_id,
                        days,
                        base,
                        salary_att,
                        computed["hours_worked_total"],
                        computed["late_minutes_total"],
                        computed["late_deduction_vnd"],
                        computed["position_allowance_vnd"],
                        computed["bonus_vnd"],
                        allow,
                        ded,
                        net,
                        note,
                        ts,
                        ts,
                    ),
                )

        lines = conn.execute(
            """
            SELECT pl.*, s.name AS staff_name, s.internal_code AS staff_code
            FROM crm_payroll_line pl
            JOIN crm_staff s ON s.id = pl.staff_id
            WHERE pl.payroll_id = ?
            ORDER BY s.name COLLATE NOCASE ASC
            """,
            (pid,),
        ).fetchall()
        pr_final = conn.execute("SELECT * FROM crm_payroll WHERE id = ?", (pid,)).fetchone()

    return jsonify({"payroll": dict(pr_final), "lines": rows_to_dict(lines)})


@app.patch("/api/crm/payroll/<int:payroll_id>")
def api_crm_patch_payroll(payroll_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM crm_payroll WHERE id = ?", (payroll_id,)).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy kỳ lương"}), 404
        status = str(row["status"] or "draft")
        if "status" in payload:
            s = str(payload.get("status") or "").strip().lower()
            if s in ("draft", "final"):
                status = s
        ts = _crm_ts()
        conn.execute(
            "UPDATE crm_payroll SET status = ?, updated_at = ? WHERE id = ?",
            (status, ts, payroll_id),
        )
        row2 = conn.execute("SELECT * FROM crm_payroll WHERE id = ?", (payroll_id,)).fetchone()
    assert row2 is not None
    return jsonify(dict(row2))


@app.patch("/api/crm/payroll/line/<int:line_id>")
def api_crm_patch_payroll_line(line_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        line = conn.execute("SELECT * FROM crm_payroll_line WHERE id = ?", (line_id,)).fetchone()
        if line is None:
            return jsonify({"error": "Không tìm thấy dòng lương"}), 404
        pr = conn.execute(
            "SELECT * FROM crm_payroll WHERE id = ?",
            (int(line["payroll_id"]),),
        ).fetchone()
        if pr is not None and str(pr["status"] or "").strip() == "final":
            return jsonify({"error": "Kỳ lương đã khóa"}), 409
        allow = int(line["allowances_vnd"] or 0)
        ded = int(line["deductions_vnd"] or 0)
        note = str(line["note"] or "")
        if "allowances_vnd" in payload:
            try:
                allow = max(0, min(int(payload["allowances_vnd"]), 9_999_999_999))
            except (TypeError, ValueError):
                pass
        if "deductions_vnd" in payload:
            try:
                ded = max(0, min(int(payload["deductions_vnd"]), 9_999_999_999))
            except (TypeError, ValueError):
                pass
        if "note" in payload and isinstance(payload["note"], str):
            note = payload["note"].strip()[:2000]
        sat = int(line["salary_from_attendance_vnd"] or 0)
        net = sat + allow - ded
        ts = _crm_ts()
        conn.execute(
            """
            UPDATE crm_payroll_line
            SET allowances_vnd = ?, deductions_vnd = ?, net_salary_vnd = ?, note = ?, updated_at = ?
            WHERE id = ?
            """,
            (allow, ded, net, note, ts, line_id),
        )
        row2 = conn.execute(
            """
            SELECT pl.*, s.name AS staff_name, s.internal_code AS staff_code
            FROM crm_payroll_line pl
            JOIN crm_staff s ON s.id = pl.staff_id
            WHERE pl.id = ?
            """,
            (line_id,),
        ).fetchone()
    assert row2 is not None
    return jsonify(dict(row2))


def _crm_kpi_normalize_status(raw: str | None) -> str:
    s = str(raw or "draft").strip().lower()
    return s if s in CRM_KPI_STATUSES else "draft"


def _crm_kpi_staff_progress_fields(payload: dict[str, Any]) -> dict[str, Any]:
    """Nhân viên chỉ cập nhật tiến độ / ghi chú."""
    out: dict[str, Any] = {}
    if "actual_value" in payload:
        av = payload["actual_value"]
        try:
            out["actual_value"] = None if av is None or av == "" else float(av)
        except (TypeError, ValueError):
            pass
    if "status" in payload:
        out["status"] = _crm_kpi_normalize_status(str(payload.get("status")))
    if "note" in payload and isinstance(payload["note"], str):
        out["note"] = payload["note"].strip()[:2000]
    return out


@app.get("/crm/kpi")
def crm_kpi_page() -> str:
    redir = _ensure_crm_session_html()
    if redir is not None:
        return redir
    staff_portal = _crm_staff_portal_active()
    if not staff_portal and not _admin_logged_in() and _crm_kpi_access_mode() == "off":
        abort(404)
    mode = _crm_kpi_access_mode()
    if staff_portal:
        mode = "staff"
    elif _admin_logged_in():
        mode = "full"
    return render_template(
        "crm_kpi.html",
        kpi_status_labels=CRM_KPI_STATUS_LABELS_VI,
        kpi_readonly=not staff_portal and mode == "view",
        kpi_access_mode=mode,
        kpi_progress_only=staff_portal,
        crm_staff_portal=staff_portal,
        crm_staff_id=_staff_session_id() if staff_portal else None,
        crm_staff_name=_staff_session_name() if staff_portal else "",
        **({} if staff_portal else _admin_page_template_kwargs()),
    )


@app.get("/crm/hub")
def crm_hub_page() -> str:
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    return render_template(
        "crm_hub.html",
        contract_status_labels=CRM_CONTRACT_STATUS_LABELS_VI,
        contract_statuses=list(CRM_CONTRACT_STATUSES),
        campaign_channel_labels=CRM_CAMPAIGN_CHANNEL_LABELS_VI,
        campaign_channels=list(CRM_CAMPAIGN_CHANNELS),
        reminder_scope_labels=CRM_REMINDER_SCOPE_LABELS_VI,
        reminder_kind_labels=CRM_REMINDER_KIND_LABELS_VI,
        reminder_status_labels=CRM_REMINDER_STATUS_LABELS_VI,
        reminder_scopes=list(CRM_REMINDER_SCOPES),
        reminder_kinds=list(CRM_REMINDER_KINDS),
        reminder_statuses_edit=list(CRM_REMINDER_STATUSES),
        crm_presales_on_lead=_crm_presales_on_lead_enabled(),
        **_admin_page_template_kwargs(),
    )


def _crm_sales_json_guard(section_id: str, action: str = "view") -> Any | None:
    if not _internal_logged_in():
        return jsonify({"error": "Chưa đăng nhập", "login": url_for("admin_login")}), 401
    if not _admin_section_can(section_id, action):
        return _admin_section_forbidden_json(section_id, action)
    return None


@app.get("/crm/sales")
def crm_sales_page() -> str:
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    return render_template(
        "crm_sales.html",
        crm_sales_plan_statuses=list(PLAN_STATUSES),
        crm_sales_plan_status_labels=PLAN_STATUS_LABELS_VI,
        crm_sales_partner_types=list(PARTNER_TYPES),
        crm_sales_partner_type_labels=PARTNER_TYPE_LABELS_VI,
        crm_sales_partner_statuses=list(PARTNER_STATUSES),
        crm_sales_partner_status_labels=PARTNER_STATUS_LABELS_VI,
        crm_sales_training_statuses=list(TRAINING_STATUSES),
        crm_sales_training_status_labels=TRAINING_STATUS_LABELS_VI,
        crm_sales_market_statuses=list(MARKET_STATUSES),
        crm_sales_market_status_labels=MARKET_STATUS_LABELS_VI,
        crm_sales_tx_types=list(TX_TYPES),
        crm_sales_tx_type_labels=TX_TYPE_LABELS_VI,
        crm_sales_tx_stages=list(TX_STAGES),
        crm_sales_tx_stage_labels=TX_STAGE_LABELS_VI,
        crm_sales_target_types=list(TARGET_TYPES),
        crm_sales_target_type_labels=TARGET_TYPE_LABELS_VI,
        crm_sales_pipeline_labels=SALES_PIPELINE_LABELS_VI,
        crm_sales_pipeline_stages=list(SALES_PIPELINE_STAGES),
        **_admin_page_template_kwargs(),
    )


@app.get("/api/crm/sales/summary")
def api_crm_sales_summary() -> Any:
    deny = _crm_sales_json_guard("crm_sales_overview", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        return jsonify(fetch_sales_summary(conn))


@app.get("/api/crm/sales/plans")
def api_crm_sales_list_plans() -> Any:
    deny = _crm_sales_json_guard("crm_sales_plans", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        return jsonify({"plans": list_plans(conn)})


@app.post("/api/crm/sales/plans")
def api_crm_sales_create_plan() -> Any:
    deny = _crm_sales_json_guard("crm_sales_plans", "create")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    title = str(payload.get("title", "")).strip()[:400]
    if not title:
        return jsonify({"error": "Thiếu tên kế hoạch"}), 400
    ts = _crm_ts()
    with get_connection() as conn:
        cur = conn.execute(
            """
            INSERT INTO crm_sales_plans (
                title, fiscal_year, period_start, period_end, revenue_target_vnd,
                status, summary, strategy_notes, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                title,
                int(payload.get("fiscal_year") or datetime.now().year),
                str(payload.get("period_start") or "")[:10],
                str(payload.get("period_end") or "")[:10],
                int(payload.get("revenue_target_vnd") or 0),
                str(payload.get("status") or "draft")[:32],
                str(payload.get("summary") or "")[:4000],
                str(payload.get("strategy_notes") or "")[:8000],
                ts,
                ts,
            ),
        )
        conn.commit()
        return jsonify({"id": int(cur.lastrowid)})


@app.get("/api/crm/sales/targets")
def api_crm_sales_list_targets() -> Any:
    deny = _crm_sales_json_guard("crm_sales_plans", "view")
    if deny is not None:
        return deny
    plan_raw = (request.args.get("plan_id") or "").strip()
    plan_id = int(plan_raw) if plan_raw.isdigit() else None
    with get_connection() as conn:
        return jsonify({"targets": list_targets(conn, plan_id)})


@app.post("/api/crm/sales/targets")
def api_crm_sales_create_target() -> Any:
    deny = _crm_sales_json_guard("crm_sales_plans", "create")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    metric = str(payload.get("metric_name", "")).strip()[:240]
    if not metric:
        return jsonify({"error": "Thiếu tên chỉ tiêu"}), 400
    ts = _crm_ts()
    plan_id = payload.get("plan_id")
    with get_connection() as conn:
        cur = conn.execute(
            """
            INSERT INTO crm_sales_targets (
                plan_id, staff_id, department_id, target_type, metric_name,
                target_value, actual_value, unit, period_month, notes, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                int(plan_id) if plan_id else None,
                int(payload["staff_id"]) if payload.get("staff_id") else None,
                int(payload["department_id"]) if payload.get("department_id") else None,
                str(payload.get("target_type") or "revenue")[:32],
                metric,
                float(payload.get("target_value") or 0),
                float(payload.get("actual_value") or 0),
                str(payload.get("unit") or "vnd")[:16],
                str(payload.get("period_month") or "")[:7],
                str(payload.get("notes") or "")[:500],
                ts,
                ts,
            ),
        )
        conn.commit()
        return jsonify({"id": int(cur.lastrowid)})


@app.get("/api/crm/sales/partners")
def api_crm_sales_list_partners() -> Any:
    deny = _crm_sales_json_guard("crm_sales_prospects", "view")
    if deny is not None:
        return deny
    q = str(request.args.get("q") or "")
    with get_connection() as conn:
        return jsonify({"partners": list_partners(conn, q)})


@app.post("/api/crm/sales/partners")
def api_crm_sales_create_partner() -> Any:
    deny = _crm_sales_json_guard("crm_sales_prospects", "create")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    name = str(payload.get("name", "")).strip()[:240]
    if not name:
        return jsonify({"error": "Thiếu tên đối tác"}), 400
    ts = _crm_ts()
    with get_connection() as conn:
        cur = conn.execute(
            """
            INSERT INTO crm_sales_partners (
                partner_type, name, phone, email, company, territory,
                commission_pct, status, assigned_staff_id, notes, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                str(payload.get("partner_type") or "ctv")[:32],
                name,
                str(payload.get("phone") or "")[:64],
                str(payload.get("email") or "")[:240],
                str(payload.get("company") or "")[:240],
                str(payload.get("territory") or "")[:240],
                float(payload["commission_pct"]) if payload.get("commission_pct") not in (None, "") else None,
                str(payload.get("status") or "active")[:32],
                int(payload["assigned_staff_id"]) if payload.get("assigned_staff_id") else None,
                str(payload.get("notes") or "")[:2000],
                ts,
                ts,
            ),
        )
        conn.commit()
        return jsonify({"id": int(cur.lastrowid)})


@app.get("/api/crm/sales/trainings")
def api_crm_sales_list_trainings() -> Any:
    deny = _crm_sales_json_guard("crm_sales_training", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        return jsonify({"trainings": list_trainings(conn)})


@app.post("/api/crm/sales/trainings")
def api_crm_sales_create_training() -> Any:
    deny = _crm_sales_json_guard("crm_sales_training", "create")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    title = str(payload.get("title", "")).strip()[:400]
    if not title:
        return jsonify({"error": "Thiếu tiêu đề"}), 400
    ts = _crm_ts()
    with get_connection() as conn:
        cur = conn.execute(
            """
            INSERT INTO crm_sales_trainings (
                title, training_date, trainer_name, topic, content_summary,
                materials_url, attendee_staff_ids, status, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                title,
                str(payload.get("training_date") or "")[:10],
                str(payload.get("trainer_name") or "")[:240],
                str(payload.get("topic") or "")[:400],
                str(payload.get("content_summary") or "")[:8000],
                str(payload.get("materials_url") or "")[:500],
                "[]",
                str(payload.get("status") or "planned")[:32],
                ts,
                ts,
            ),
        )
        conn.commit()
        return jsonify({"id": int(cur.lastrowid)})


@app.get("/api/crm/sales/market")
def api_crm_sales_list_market() -> Any:
    deny = _crm_sales_json_guard("crm_sales_market", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        return jsonify({"research": list_market_research(conn)})


@app.post("/api/crm/sales/market")
def api_crm_sales_create_market() -> Any:
    deny = _crm_sales_json_guard("crm_sales_market", "create")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    title = str(payload.get("title", "")).strip()[:400]
    if not title:
        return jsonify({"error": "Thiếu tiêu đề"}), 400
    ts = _crm_ts()
    with get_connection() as conn:
        cur = conn.execute(
            """
            INSERT INTO crm_sales_market_research (
                title, research_date, area, property_type, competitor_notes,
                price_analysis, strategy_proposal, status, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                title,
                str(payload.get("research_date") or "")[:10],
                str(payload.get("area") or "")[:240],
                str(payload.get("property_type") or "")[:240],
                str(payload.get("competitor_notes") or "")[:8000],
                str(payload.get("price_analysis") or "")[:8000],
                str(payload.get("strategy_proposal") or "")[:8000],
                str(payload.get("status") or "draft")[:32],
                ts,
                ts,
            ),
        )
        conn.commit()
        return jsonify({"id": int(cur.lastrowid)})


@app.get("/api/crm/sales/transactions")
def api_crm_sales_list_transactions() -> Any:
    deny = _crm_sales_json_guard("crm_sales_deals", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        return jsonify({"transactions": list_transactions(conn)})


@app.post("/api/crm/sales/transactions")
def api_crm_sales_create_transaction() -> Any:
    deny = _crm_sales_json_guard("crm_sales_deals", "create")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    prop = str(payload.get("property_ref", "")).strip()[:400]
    if not prop:
        return jsonify({"error": "Thiếu mã/mô tả BĐS"}), 400
    ts = _crm_ts()
    with get_connection() as conn:
        cur = conn.execute(
            """
            INSERT INTO crm_sales_transactions (
                case_id, contract_id, customer_id, transaction_type, property_ref,
                stage, deal_value_vnd, assigned_staff_id, notes, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                int(payload["case_id"]) if payload.get("case_id") else None,
                int(payload["contract_id"]) if payload.get("contract_id") else None,
                int(payload["customer_id"]) if payload.get("customer_id") else None,
                str(payload.get("transaction_type") or "ban")[:32],
                prop,
                str(payload.get("stage") or "tu_van")[:32],
                int(payload.get("deal_value_vnd") or 0),
                int(payload["assigned_staff_id"]) if payload.get("assigned_staff_id") else None,
                str(payload.get("notes") or "")[:4000],
                ts,
                ts,
            ),
        )
        conn.commit()
        return jsonify({"id": int(cur.lastrowid)})


@app.get("/api/crm/sales/pipeline-cases")
def api_crm_sales_pipeline_cases() -> Any:
    deny = _crm_sales_json_guard("crm_sales_funnel", "view")
    if deny is not None:
        return deny
    stage = str(request.args.get("stage") or "").strip() or None
    with get_connection() as conn:
        return jsonify({"cases": list_pipeline_cases(conn, stage)})


@app.get("/api/crm/sales/reports")
def api_crm_sales_reports() -> Any:
    deny = _crm_sales_json_guard("crm_sales_reports", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        return jsonify(sales_report_data(conn))


def _crm_re_projects_json_guard(section_id: str, action: str = "view") -> Any | None:
    return _crm_sales_json_guard(section_id, action)


@app.get("/crm/re-projects")
def crm_re_projects_page() -> str:
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    with get_connection() as conn:
        type_rows = list_project_types(conn, include_inactive=True)
        type_labels = project_type_label_map(conn, include_inactive=True)
        active_codes = [str(t["code"]) for t in type_rows if t.get("active")]
    return render_template(
        "crm_re_projects.html",
        crm_re_project_types=active_codes or list(PROJECT_TYPES),
        crm_re_project_type_labels=type_labels or PROJECT_TYPE_LABELS,
        crm_re_project_statuses=list(PROJECT_STATUSES),
        crm_re_project_status_labels=PROJECT_STATUS_LABELS,
        crm_re_product_statuses=list(PRODUCT_STATUSES),
        crm_re_product_status_labels=PRODUCT_STATUS_LABELS,
        crm_re_product_lines=list(PRODUCT_LINES),
        crm_re_product_line_labels=PRODUCT_LINE_LABELS,
        crm_re_product_typologies=list(PRODUCT_TYPOLOGIES),
        crm_re_product_typology_labels=PRODUCT_TYPOLOGY_LABELS,
        crm_re_kpi_categories=list(KPI_CATEGORIES),
        crm_re_kpi_category_labels=KPI_CATEGORY_LABELS,
        crm_re_kpi_track_statuses=list(KPI_TRACK_STATUSES),
        crm_re_kpi_track_status_labels=KPI_TRACK_STATUS_LABELS,
        crm_re_kpi_metric_templates=list(KPI_METRIC_TEMPLATES),
        crm_re_risk_categories=list(RISK_CATEGORIES),
        crm_re_risk_category_labels=RISK_CATEGORY_LABELS,
        crm_re_risk_levels=list(RISK_LEVELS),
        crm_re_risk_level_labels=RISK_LEVEL_LABELS,
        crm_re_budget_categories=list(BUDGET_CATEGORIES),
        crm_re_budget_category_labels=BUDGET_CATEGORY_LABELS,
        crm_re_cash_flow_type_labels=CASH_FLOW_TYPE_LABELS,
        crm_re_cash_flow_status_labels=CASH_FLOW_STATUS_LABELS,
        crm_re_cash_flow_source_labels=CASH_FLOW_SOURCE_LABELS,
        crm_re_marketing_sub_category_labels=MARKETING_SUB_CATEGORY_LABELS,
        crm_re_price_list_status_labels=PRICE_LIST_STATUS_LABELS,
        **_admin_page_template_kwargs(),
    )


@app.get("/api/crm/re-projects/types")
def api_crm_re_project_types_list() -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects", "view")
    if deny is not None:
        return deny
    raw = (request.args.get("include_inactive") or "").strip().lower()
    include_inactive = raw in ("1", "true", "yes", "all")
    with get_connection() as conn:
        types = list_project_types(conn, include_inactive=include_inactive)
        labels = project_type_label_map(conn, include_inactive=include_inactive)
    return jsonify({"types": types, "labels": labels})


@app.post("/api/crm/re-projects/types")
def api_crm_re_project_types_create() -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects", "create")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    try:
        with get_connection() as conn:
            row = save_project_type(conn, payload, ts=_crm_ts())
            conn.commit()
            return jsonify(row), 201
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.put("/api/crm/re-projects/types/<int:type_id>")
def api_crm_re_project_types_update(type_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects", "edit")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    try:
        with get_connection() as conn:
            row = save_project_type(conn, payload, type_id=type_id, ts=_crm_ts())
            conn.commit()
            return jsonify(row)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.delete("/api/crm/re-projects/types/<int:type_id>")
def api_crm_re_project_types_delete(type_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects", "delete")
    if deny is not None:
        return deny
    try:
        with get_connection() as conn:
            delete_project_type(conn, type_id)
            conn.commit()
        return jsonify({"ok": True})
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/crm/re-projects")
def api_crm_re_projects_list() -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects", "view")
    if deny is not None:
        return deny
    q = (request.args.get("q") or "").strip()
    with get_connection() as conn:
        return jsonify({"projects": list_projects(conn, q=q)})


@app.post("/api/crm/re-projects")
def api_crm_re_projects_create() -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects", "create")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    try:
        with get_connection() as conn:
            proj = re_create_project(conn, payload, ts=_crm_ts())
            conn.commit()
            return jsonify(proj), 201
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/crm/re-projects/<int:project_id>")
def api_crm_re_projects_get(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        proj = fetch_project(conn, project_id)
        if proj is None:
            return jsonify({"error": "Không tìm thấy dự án."}), 404
        return jsonify(proj)


@app.put("/api/crm/re-projects/<int:project_id>")
def api_crm_re_projects_update(project_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    section = "crm_re_projects"
    if "business_plan" in payload:
        section = "crm_re_projects_business"
    elif "marketing_plan" in payload:
        section = "crm_re_projects_marketing"
    elif "sales_plan" in payload:
        section = "crm_re_projects_sales"
    deny = _crm_re_projects_json_guard(section, "edit")
    if deny is not None:
        return deny
    try:
        with get_connection() as conn:
            proj = re_update_project(conn, project_id, payload, ts=_crm_ts())
            conn.commit()
            return jsonify(proj)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.delete("/api/crm/re-projects/<int:project_id>")
def api_crm_re_projects_delete(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects", "delete")
    if deny is not None:
        return deny
    with get_connection() as conn:
        re_delete_project(conn, project_id)
        conn.commit()
    return jsonify({"ok": True})


@app.get("/api/crm/re-projects/<int:project_id>/summary")
def api_crm_re_projects_summary(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects", "view")
    if deny is not None:
        return deny
    try:
        with get_connection() as conn:
            return jsonify(fetch_project_summary(conn, project_id))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 404


@app.get("/api/crm/re-projects/<int:project_id>/products")
def api_crm_re_projects_list_products(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_products", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        products = list_products(conn, project_id)
        return jsonify({
            "products": products,
            "inventory": compute_product_inventory_stats(products),
        })


@app.post("/api/crm/re-projects/<int:project_id>/products")
def api_crm_re_projects_create_product(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_products", "create")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    try:
        with get_connection() as conn:
            row = save_product(conn, project_id, payload, ts=_crm_ts())
            conn.commit()
            return jsonify(row), 201
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.put("/api/crm/re-projects/<int:project_id>/products/<int:product_id>")
def api_crm_re_projects_update_product(project_id: int, product_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_products", "edit")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = save_product(conn, project_id, payload, product_id=product_id, ts=_crm_ts())
        conn.commit()
        return jsonify(row)


@app.delete("/api/crm/re-projects/<int:project_id>/products/<int:product_id>")
def api_crm_re_projects_delete_product(project_id: int, product_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_products", "delete")
    if deny is not None:
        return deny
    with get_connection() as conn:
        delete_product(conn, project_id, product_id)
        conn.commit()
    return jsonify({"ok": True})


@app.get("/api/crm/re-projects/<int:project_id>/products/search")
def api_crm_re_projects_search_products(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_products", "view")
    if deny is not None:
        return deny
    q = str(request.args.get("q") or "").strip()
    line = normalize_product_line(str(request.args.get("product_line") or ""))
    zone = str(request.args.get("zone") or "").strip()
    price_batch = str(request.args.get("price_batch") or "").strip()
    status = str(request.args.get("status") or "available").strip() or "available"
    try:
        lim = max(1, min(int(request.args.get("limit") or 50), 200))
    except ValueError:
        lim = 50
    with get_connection() as conn:
        products = search_available_products(
            conn,
            project_id,
            q=q,
            product_line=line,
            zone=zone,
            price_batch=price_batch,
            status=status,
            limit=lim,
        )
    return jsonify({"products": products})


@app.get("/api/crm/re-projects/<int:project_id>/zones")
def api_crm_re_projects_list_zones(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        zones = list_project_zones(conn, project_id)
    return jsonify({"zones": zones})


@app.get("/api/crm/re-projects/<int:project_id>/inventory-by-zone")
def api_crm_re_projects_inventory_by_zone(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_products", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        rows = inventory_by_zone_summary(conn, project_id)
    return jsonify({"zones": rows})


@app.get("/api/crm/re-projects/<int:project_id>/price-batches")
def api_crm_re_projects_list_price_batches(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_products", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        batches = list_price_batches(conn, project_id)
        summary = inventory_by_price_batch_summary(conn, project_id)
    return jsonify({"batches": batches, "summary": summary})


@app.get("/api/crm/re-projects/<int:project_id>/price-lists")
def api_crm_re_projects_list_price_lists(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_products", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        rows = list_price_lists(conn, project_id)
        versions = list_all_version_codes(conn, project_id)
    return jsonify({"price_lists": rows, "version_codes": versions})


@app.post("/api/crm/re-projects/<int:project_id>/price-lists")
def api_crm_re_projects_create_price_list(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_products", "create")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    try:
        with get_connection() as conn:
            row = save_price_list(
                conn,
                project_id,
                payload,
                created_by=_crm_audit_user(),
                ts=_crm_ts(),
            )
            conn.commit()
            return jsonify(row), 201
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/crm/re-projects/<int:project_id>/price-lists/<int:list_id>")
def api_crm_re_projects_get_price_list(project_id: int, list_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_products", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        row = fetch_price_list(conn, project_id, list_id)
        if row is None:
            return jsonify({"error": "Không tìm thấy bảng giá."}), 404
        items, total = list_price_list_items(conn, list_id, limit=500)
    return jsonify({"price_list": row, "items": items, "items_total": total})


@app.put("/api/crm/re-projects/<int:project_id>/price-lists/<int:list_id>")
def api_crm_re_projects_update_price_list(project_id: int, list_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_products", "edit")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    try:
        with get_connection() as conn:
            row = save_price_list(
                conn,
                project_id,
                payload,
                list_id=list_id,
                created_by=_crm_audit_user(),
                ts=_crm_ts(),
            )
            conn.commit()
            return jsonify(row)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.delete("/api/crm/re-projects/<int:project_id>/price-lists/<int:list_id>")
def api_crm_re_projects_delete_price_list(project_id: int, list_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_products", "delete")
    if deny is not None:
        return deny
    try:
        with get_connection() as conn:
            delete_price_list(conn, project_id, list_id)
            conn.commit()
        return jsonify({"ok": True})
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/crm/re-projects/<int:project_id>/price-lists/<int:list_id>/items/import")
def api_crm_re_projects_import_price_list_items(project_id: int, list_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_products", "create")
    if deny is not None:
        return deny
    pl_check = None
    with get_connection() as conn:
        pl_check = fetch_price_list(conn, project_id, list_id)
    if pl_check is None:
        return jsonify({"error": "Không tìm thấy bảng giá."}), 404
    csv_text = ""
    if request.content_type and "multipart/form-data" in request.content_type:
        f = request.files.get("file")
        if f:
            csv_text = f.read().decode("utf-8-sig", errors="replace")
    if not csv_text:
        payload = request.get_json(silent=True) or {}
        csv_text = str(payload.get("csv") or payload.get("csv_text") or "")
    if not csv_text.strip():
        return jsonify({"error": "Thiếu nội dung CSV."}), 400
    try:
        with get_connection() as conn:
            result = import_price_list_items_csv(conn, list_id, csv_text, ts=_crm_ts())
            conn.commit()
        return jsonify(result)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/crm/re-projects/<int:project_id>/price-lists/<int:list_id>/apply")
def api_crm_re_projects_apply_price_list(project_id: int, list_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_products", "edit")
    if deny is not None:
        return deny
    try:
        with get_connection() as conn:
            result = apply_price_list(
                conn,
                project_id,
                list_id,
                updated_by=_crm_audit_user(),
                ts=_crm_ts(),
            )
            conn.commit()
        return jsonify(result)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/crm/re-projects/<int:project_id>/price-lists/compare")
def api_crm_re_projects_compare_price_lists(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_products", "view")
    if deny is not None:
        return deny
    va = str(request.args.get("a") or request.args.get("version_a") or "").strip()
    vb = str(request.args.get("b") or request.args.get("version_b") or "").strip()
    if not va or not vb:
        return jsonify({"error": "Cần tham số a và b (mã version)."}), 400
    try:
        with get_connection() as conn:
            out = compare_price_lists(conn, project_id, va, vb)
        return jsonify(out)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/crm/re-projects/<int:project_id>/price-lists/products")
def api_crm_re_projects_products_on_version(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_products", "view")
    if deny is not None:
        return deny
    version = str(request.args.get("version") or request.args.get("version_code") or "").strip()
    if not version:
        return jsonify({"error": "Thiếu version."}), 400
    try:
        lim = max(1, min(int(request.args.get("limit") or 100), 500))
    except ValueError:
        lim = 100
    with get_connection() as conn:
        products = products_on_price_version(conn, project_id, version, limit=lim)
    return jsonify({"version_code": version, "products": products, "total": len(products)})


@app.post("/api/crm/re-projects/<int:project_id>/products/import")
def api_crm_re_projects_import_products(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_products", "create")
    if deny is not None:
        return deny
    ts = _crm_ts()
    actor = _crm_audit_user()
    csv_text = ""
    if request.content_type and "multipart/form-data" in request.content_type:
        f = request.files.get("file")
        if f:
            csv_text = f.read().decode("utf-8-sig", errors="replace")
    if not csv_text:
        payload = request.get_json(silent=True) or {}
        csv_text = str(payload.get("csv") or payload.get("csv_text") or "")
    if not csv_text.strip():
        return jsonify({"error": "Thiếu nội dung CSV (file hoặc trường csv)."}), 400
    with get_connection() as conn:
        try:
            result = import_products_csv(conn, project_id, csv_text, updated_by=actor, ts=ts)
            conn.commit()
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    return jsonify(result)


@app.post("/api/crm/leads/<int:lead_id>/hold-product")
def api_crm_lead_hold_product(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "edit"):
        return _admin_section_forbidden_json("crm_leads", "edit")
    payload = request.get_json(force=True) or {}
    try:
        product_id = int(payload.get("product_id") or 0)
    except (TypeError, ValueError):
        return jsonify({"error": "product_id không hợp lệ"}), 400
    if product_id <= 0:
        return jsonify({"error": "Thiếu product_id."}), 400
    ts = _crm_ts()
    actor = _crm_audit_user()
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền sửa lead này."}), 403
        try:
            hold_product_for_lead(conn, lead_id, product_id, updated_by=actor, ts=ts)
            conn.commit()
            row = fetch_lead_by_id(conn, lead_id)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        out = lead_row_to_dict(row, conn) if row else None
    return jsonify({"ok": True, "lead": out})


@app.post("/api/crm/leads/<int:lead_id>/release-product")
def api_crm_lead_release_product(lead_id: int) -> Any:
    if not _admin_section_can("crm_leads", "edit"):
        return _admin_section_forbidden_json("crm_leads", "edit")
    ts = _crm_ts()
    actor = _crm_audit_user()
    with get_connection() as conn:
        prev = fetch_lead_by_id(conn, lead_id)
        if prev is None:
            return jsonify({"error": "Không tìm thấy lead."}), 404
        if not _crm_lead_can_access(conn, prev):
            return jsonify({"error": "Không có quyền sửa lead này."}), 403
        try:
            release_product_hold(conn, lead_id, updated_by=actor, ts=ts)
            conn.commit()
            row = fetch_lead_by_id(conn, lead_id)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        out = lead_row_to_dict(row, conn) if row else None
    return jsonify({"ok": True, "lead": out})


@app.get("/api/crm/re-projects/<int:project_id>/kpis")
def api_crm_re_projects_list_kpis(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_kpi", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        kpis = list_kpis(conn, project_id)
        return jsonify({
            "kpis": kpis,
            "board": compute_kpi_board_stats(kpis),
        })


@app.post("/api/crm/re-projects/<int:project_id>/kpis")
def api_crm_re_projects_create_kpi(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_kpi", "create")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    try:
        with get_connection() as conn:
            row = save_kpi(conn, project_id, payload, ts=_crm_ts())
            conn.commit()
            return jsonify(row), 201
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.put("/api/crm/re-projects/<int:project_id>/kpis/<int:kpi_id>")
def api_crm_re_projects_update_kpi(project_id: int, kpi_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_kpi", "edit")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    try:
        with get_connection() as conn:
            row = save_kpi(conn, project_id, payload, kpi_id=kpi_id, ts=_crm_ts())
            conn.commit()
            return jsonify(row)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.delete("/api/crm/re-projects/<int:project_id>/kpis/<int:kpi_id>")
def api_crm_re_projects_delete_kpi(project_id: int, kpi_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_kpi", "delete")
    if deny is not None:
        return deny
    with get_connection() as conn:
        delete_kpi(conn, project_id, kpi_id)
        conn.commit()
    return jsonify({"ok": True})


@app.get("/api/crm/re-projects/kpi-metrics")
def api_crm_re_projects_kpi_metrics() -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_kpi", "view")
    if deny is not None:
        return deny
    re_only = (request.args.get("re_only") or "1").strip().lower() in ("1", "true", "yes")
    with get_connection() as conn:
        metrics = list_crm_kpi_metrics(conn, re_only=re_only)
    return jsonify({"metrics": metrics})


@app.post("/api/crm/re-projects/<int:project_id>/kpis/sync-to-staff")
def api_crm_re_projects_sync_kpis_to_staff(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_kpi", "edit")
    if deny is not None:
        return deny
    with get_connection() as conn:
        if fetch_project(conn, project_id) is None:
            return jsonify({"error": "Không tìm thấy dự án."}), 404
        result = sync_project_kpis_to_staff(conn, project_id, ts=_crm_ts())
        conn.commit()
    return jsonify(result)


@app.post("/api/crm/re-projects/<int:project_id>/kpis/pull-from-staff")
def api_crm_re_projects_pull_kpis_from_staff(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_kpi", "edit")
    if deny is not None:
        return deny
    with get_connection() as conn:
        if fetch_project(conn, project_id) is None:
            return jsonify({"error": "Không tìm thấy dự án."}), 404
        result = pull_project_kpis_from_staff(conn, project_id, ts=_crm_ts())
        conn.commit()
    return jsonify(result)


@app.post("/api/crm/re-projects/<int:project_id>/kpis/refresh-leads-new")
def api_crm_re_projects_refresh_leads_new_kpi(project_id: int) -> Any:
    """Đồng bộ KPI RE_LEADS_NEW từ crm_leads (theo tháng)."""
    deny = _crm_re_projects_json_guard("crm_re_projects_kpi", "edit")
    if deny is not None:
        return deny
    payload = request.get_json(silent=True) or {}
    period_month = str(payload.get("period_month") or "").strip()[:7]
    with get_connection() as conn:
        if fetch_project(conn, project_id) is None:
            return jsonify({"error": "Không tìm thấy dự án."}), 404
        result = refresh_project_re_leads_new_kpi(
            conn,
            project_id,
            period_month=period_month,
            ts=_crm_ts(),
        )
        conn.commit()
    return jsonify(result)


@app.get("/api/crm/re-projects/<int:project_id>/risks")
def api_crm_re_projects_list_risks(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_risks", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        return jsonify({"risks": list_risks(conn, project_id)})


@app.post("/api/crm/re-projects/<int:project_id>/risks")
def api_crm_re_projects_create_risk(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_risks", "create")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    try:
        with get_connection() as conn:
            row = save_risk(conn, project_id, payload, ts=_crm_ts())
            conn.commit()
            return jsonify(row), 201
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.put("/api/crm/re-projects/<int:project_id>/risks/<int:risk_id>")
def api_crm_re_projects_update_risk(project_id: int, risk_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_risks", "edit")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    try:
        with get_connection() as conn:
            row = save_risk(conn, project_id, payload, risk_id=risk_id, ts=_crm_ts())
            conn.commit()
            return jsonify(row)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.delete("/api/crm/re-projects/<int:project_id>/risks/<int:risk_id>")
def api_crm_re_projects_delete_risk(project_id: int, risk_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_risks", "delete")
    if deny is not None:
        return deny
    with get_connection() as conn:
        delete_risk(conn, project_id, risk_id)
        conn.commit()
    return jsonify({"ok": True})


@app.get("/api/crm/re-projects/<int:project_id>/budget")
def api_crm_re_projects_list_budget(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_budget", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        return jsonify({"lines": list_budget_lines(conn, project_id)})


@app.post("/api/crm/re-projects/<int:project_id>/budget")
def api_crm_re_projects_create_budget(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_budget", "create")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    try:
        with get_connection() as conn:
            row = save_budget_line(conn, project_id, payload, ts=_crm_ts())
            conn.commit()
            return jsonify(row), 201
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.put("/api/crm/re-projects/<int:project_id>/budget/<int:line_id>")
def api_crm_re_projects_update_budget(project_id: int, line_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_budget", "edit")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    try:
        with get_connection() as conn:
            row = save_budget_line(conn, project_id, payload, line_id=line_id, ts=_crm_ts())
            conn.commit()
            return jsonify(row)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.delete("/api/crm/re-projects/<int:project_id>/budget/<int:line_id>")
def api_crm_re_projects_delete_budget(project_id: int, line_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_budget", "delete")
    if deny is not None:
        return deny
    with get_connection() as conn:
        delete_budget_line(conn, project_id, line_id)
        conn.commit()
    return jsonify({"ok": True})


@app.get("/api/crm/re-projects/<int:project_id>/accounting/dashboard")
def api_crm_re_projects_accounting_dashboard(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_budget", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        return jsonify(compute_accounting_dashboard(conn, project_id))


@app.get("/api/crm/re-projects/<int:project_id>/accounting/cash-flow")
def api_crm_re_projects_list_cash_flow(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_budget", "view")
    if deny is not None:
        return deny
    flow_type = (request.args.get("flow_type") or "").strip() or None
    category = (request.args.get("category") or "").strip() or None
    status = (request.args.get("status") or "").strip() or None
    with get_connection() as conn:
        lines = list_cash_flow_lines(conn, project_id, flow_type=flow_type, category=category, status=status)
    return jsonify({"lines": lines})


@app.post("/api/crm/re-projects/<int:project_id>/accounting/cash-flow")
def api_crm_re_projects_create_cash_flow(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_budget", "create")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    try:
        with get_connection() as conn:
            row = save_cash_flow_line(
                conn,
                project_id,
                payload,
                created_by=_crm_audit_user(),
                ts=_crm_ts(),
            )
            conn.commit()
            return jsonify(row), 201
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.put("/api/crm/re-projects/<int:project_id>/accounting/cash-flow/<int:line_id>")
def api_crm_re_projects_update_cash_flow(project_id: int, line_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_budget", "edit")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    try:
        with get_connection() as conn:
            row = save_cash_flow_line(
                conn,
                project_id,
                payload,
                line_id=line_id,
                created_by=_crm_audit_user(),
                ts=_crm_ts(),
            )
            conn.commit()
            return jsonify(row)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.delete("/api/crm/re-projects/<int:project_id>/accounting/cash-flow/<int:line_id>")
def api_crm_re_projects_delete_cash_flow(project_id: int, line_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_budget", "delete")
    if deny is not None:
        return deny
    with get_connection() as conn:
        delete_cash_flow_line(conn, project_id, line_id)
        conn.commit()
    return jsonify({"ok": True})


@app.post("/api/crm/re-projects/<int:project_id>/accounting/cash-flow/import")
def api_crm_re_projects_import_cash_flow(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_budget", "create")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    csv_text = str(payload.get("csv") or "")
    if not csv_text.strip():
        return jsonify({"error": "Thiếu nội dung CSV."}), 400
    with get_connection() as conn:
        result = import_cash_flow_csv(conn, project_id, csv_text, created_by=_crm_audit_user(), ts=_crm_ts())
        conn.commit()
    return jsonify(result)


@app.post("/api/crm/re-projects/<int:project_id>/accounting/sync-from-plans")
def api_crm_re_projects_accounting_sync_plans(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_budget", "edit")
    if deny is not None:
        return deny
    try:
        with get_connection() as conn:
            result = sync_budget_from_plans(conn, project_id, ts=_crm_ts())
            conn.commit()
            return jsonify(result)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/crm/re-projects/<int:project_id>/accounting/sync-inventory-revenue")
def api_crm_re_projects_accounting_sync_inventory(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_budget", "edit")
    if deny is not None:
        return deny
    try:
        with get_connection() as conn:
            result = sync_revenue_from_inventory(conn, project_id, ts=_crm_ts(), created_by=_crm_audit_user())
            conn.commit()
            return jsonify(result)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/crm/re-projects/<int:project_id>/accounting/ai/ask")
def api_crm_re_projects_accounting_ai_ask(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_budget", "view")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    question = str(payload.get("question") or payload.get("q") or "").strip()
    if not question:
        return jsonify({"error": "Thiếu câu hỏi."}), 400
    with get_connection() as conn:
        out = ai_project_finance_query(
            conn,
            question,
            re_project_id=project_id,
            created_by=_crm_audit_user(),
            ts=_crm_ts(),
        )
        conn.commit()
    return jsonify(out)


@app.get("/api/crm/re-projects/<int:project_id>/accounting/export")
def api_crm_re_projects_accounting_export(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_budget", "export")
    if deny is not None:
        return deny
    fmt = (request.args.get("format") or "xlsx").strip().lower()
    if fmt not in ("csv", "xlsx"):
        fmt = "xlsx"
    with get_connection() as conn:
        try:
            proj = fetch_project(conn, project_id)
            if proj is None:
                raise ValueError("Không tìm thấy dự án.")
            sheets = build_accounting_export_sheets(conn, project_id)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 404
    code = str(proj.get("code") or f"du-an-{project_id}").strip() or f"du-an-{project_id}"
    stamp = datetime.now().strftime("%Y-%m-%d")
    base = f"ke-toan-{code}-{stamp}".replace(" ", "-")
    if fmt == "xlsx":
        return _crm_re_export_xlsx_sheets(sheets, filename=f"{base}.xlsx")
    summary = sheets[0][2] if sheets else []
    return _crm_re_export_csv(["Trường", "Giá trị"], summary, filename=f"{base}.csv")


@app.get("/api/crm/re-projects/<int:project_id>/accounting/risk-predictions")
def api_crm_re_projects_accounting_risk_predictions(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_budget", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        try:
            out = predict_financial_risks(conn, project_id)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 404
    return jsonify(out)


@app.get("/api/crm/re-projects/<int:project_id>/accounting/forecast")
def api_crm_re_projects_accounting_forecast(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_budget", "view")
    if deny is not None:
        return deny
    months = request.args.get("months") or "3"
    try:
        months_ahead = max(1, min(12, int(months)))
    except ValueError:
        months_ahead = 3
    with get_connection() as conn:
        try:
            out = forecast_financial_outlook(conn, project_id, months_ahead=months_ahead)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 404
    return jsonify(out)


@app.post("/api/crm/re-projects/<int:project_id>/accounting/risk-predictions/apply")
def api_crm_re_projects_accounting_apply_risks(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects_budget", "edit")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    codes = payload.get("codes")
    code_list = [str(c) for c in codes] if isinstance(codes, list) else None
    with get_connection() as conn:
        try:
            result = apply_predicted_risks_to_register(conn, project_id, codes=code_list, ts=_crm_ts())
            conn.commit()
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    return jsonify(result)


@app.get("/api/crm/re-projects/<int:project_id>/staff")
def api_crm_re_projects_list_staff(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        try:
            staff = list_project_staff(conn, project_id, active_only=True)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    return jsonify({"project_id": project_id, "staff": staff})


@app.post("/api/crm/re-projects/<int:project_id>/staff")
def api_crm_re_projects_add_staff(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects", "edit")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    try:
        staff_id = int(payload.get("staff_id") or 0)
    except (TypeError, ValueError):
        return jsonify({"error": "staff_id không hợp lệ"}), 400
    if staff_id <= 0:
        return jsonify({"error": "Thiếu staff_id."}), 400
    ts = _crm_ts()
    with get_connection() as conn:
        try:
            row = add_project_staff(
                conn,
                project_id,
                staff_id=staff_id,
                role=str(payload.get("role") or "sales"),
                assign_enabled=bool(payload.get("assign_enabled", True)),
                sort_order=int(payload.get("sort_order") or 0),
                scope_product_lines=payload.get("scope_product_lines"),
                scope_zones=payload.get("scope_zones"),
                ts=ts,
            )
            conn.commit()
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    return jsonify({"staff": row}), 201


@app.put("/api/crm/re-projects/<int:project_id>/staff/<int:staff_id>")
def api_crm_re_projects_update_staff(project_id: int, staff_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects", "edit")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    ts = _crm_ts()
    kwargs: dict[str, Any] = {"ts": ts}
    if "role" in payload:
        kwargs["role"] = str(payload.get("role") or "sales")
    if "assign_enabled" in payload:
        kwargs["assign_enabled"] = bool(payload.get("assign_enabled"))
    if "sort_order" in payload:
        try:
            kwargs["sort_order"] = int(payload.get("sort_order") or 0)
        except (TypeError, ValueError):
            return jsonify({"error": "sort_order không hợp lệ"}), 400
    if "scope_product_lines" in payload:
        raw = payload.get("scope_product_lines")
        kwargs["scope_product_lines"] = raw if isinstance(raw, list) else []
    if "scope_zones" in payload:
        raw = payload.get("scope_zones")
        kwargs["scope_zones"] = raw if isinstance(raw, list) else []
    with get_connection() as conn:
        try:
            row = update_project_staff(conn, project_id, staff_id, **kwargs)
            conn.commit()
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    return jsonify({"staff": row})


@app.delete("/api/crm/re-projects/<int:project_id>/staff/<int:staff_id>")
def api_crm_re_projects_remove_staff(project_id: int, staff_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects", "edit")
    if deny is not None:
        return deny
    ts = _crm_ts()
    with get_connection() as conn:
        try:
            remove_project_staff(conn, project_id, staff_id, ts=ts)
            conn.commit()
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    return jsonify({"ok": True})


@app.get("/api/crm/re-projects/<int:project_id>/lead-config")
def api_crm_re_projects_get_lead_config(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        try:
            from crm_project_webhooks import get_project_lead_config

            cfg = get_project_lead_config(conn, project_id)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    return jsonify({"config": cfg})


@app.put("/api/crm/re-projects/<int:project_id>/lead-config")
def api_crm_re_projects_save_lead_config(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects", "edit")
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    ts = _crm_ts()
    with get_connection() as conn:
        try:
            from crm_project_webhooks import save_project_lead_config

            cfg = save_project_lead_config(
                conn,
                project_id,
                payload,
                updated_by=_crm_audit_user(),
                ts=ts,
            )
            conn.commit()
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    return jsonify({"config": cfg})


def _crm_re_export_csv(headers: list[str], rows: list[list[Any]], *, filename: str) -> Response:
    si = StringIO()
    w = csv.writer(si)
    if headers:
        w.writerow(headers)
    for row in rows:
        w.writerow(row)
    raw = si.getvalue().encode("utf-8-sig")
    safe = filename.replace('"', "")
    return Response(
        raw,
        mimetype="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": (f'attachment; filename="{safe}"; filename*=UTF-8\'\'{quote(filename)}'),
        },
    )


def _crm_re_export_xlsx_sheets(sheets: list[tuple[str, list[str], list[list[Any]]]], *, filename: str) -> Response:
    wb = Workbook()
    first = True
    for title, headers, rows in sheets:
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet(title=title[:31])
        assert ws is not None
        ws.title = title[:31]
        if headers:
            ws.append(headers)
        for row in rows:
            ws.append(list(row))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
        etag=False,
        max_age=0,
        conditional=False,
    )


@app.get("/api/crm/re-projects/<int:project_id>/workflow")
def api_crm_re_projects_workflow(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects", "view")
    if deny is not None:
        return deny
    with get_connection() as conn:
        try:
            data = compute_project_workflow(conn, project_id)
        except ValueError as e:
            return jsonify({"error": str(e)}), 404
    return jsonify(data)


@app.get("/api/crm/re-projects/<int:project_id>/export")
def api_crm_re_projects_export(project_id: int) -> Any:
    deny = _crm_re_projects_json_guard("crm_re_projects", "export")
    if deny is not None:
        return deny
    fmt = (request.args.get("format") or "xlsx").strip().lower()
    report = (request.args.get("report") or "full").strip().lower()
    if fmt not in ("csv", "xlsx"):
        fmt = "xlsx"
    with get_connection() as conn:
        try:
            pack = fetch_project_export_data(conn, project_id)
        except ValueError as e:
            return jsonify({"error": str(e)}), 404
    proj = pack["project"]
    code = str(proj.get("code") or f"du-an-{project_id}").strip() or f"du-an-{project_id}"
    stamp = datetime.now().strftime("%Y-%m-%d")
    base = f"re-{code}-{stamp}".replace(" ", "-")

    if report == "summary":
        rows = project_export_summary_rows(proj, pack["summary"], pack["workflow"])
        if fmt == "xlsx":
            return _crm_re_export_xlsx_sheets([("Tóm tắt", ["Trường", "Giá trị"], rows)], filename=f"{base}-tom-tat.xlsx")
        return _crm_re_export_csv(["Trường", "Giá trị"], rows, filename=f"{base}-tom-tat.csv")

    if report == "workflow":
        headers, rows = project_export_workflow_rows(pack["workflow"])
        if fmt == "xlsx":
            return _crm_re_export_xlsx_sheets([("Quy trình", headers, rows)], filename=f"{base}-quy-trinh.xlsx")
        return _crm_re_export_csv(headers, rows, filename=f"{base}-quy-trinh.csv")

    if report == "kpis":
        headers, rows = project_export_kpi_rows(pack["kpis"])
        fname = f"{base}-kpi.{'xlsx' if fmt == 'xlsx' else 'csv'}"
        if fmt == "xlsx":
            return _crm_re_export_xlsx_sheets([("KPI", headers, rows)], filename=fname)
        return _crm_re_export_csv(headers, rows, filename=fname)

    if report == "products":
        headers, rows = project_export_product_rows(pack["products"])
        fname = f"{base}-ton-kho.{'xlsx' if fmt == 'xlsx' else 'csv'}"
        if fmt == "xlsx":
            return _crm_re_export_xlsx_sheets([("Tồn kho", headers, rows)], filename=fname)
        return _crm_re_export_csv(headers, rows, filename=fname)

    if report == "risks":
        headers, rows = project_export_risk_rows(pack["risks"])
        fname = f"{base}-rui-ro.{'xlsx' if fmt == 'xlsx' else 'csv'}"
        if fmt == "xlsx":
            return _crm_re_export_xlsx_sheets([("Rủi ro", headers, rows)], filename=fname)
        return _crm_re_export_csv(headers, rows, filename=fname)

    if report == "budget":
        headers, rows = project_export_budget_rows(pack["budget"])
        fname = f"{base}-ngan-sach.{'xlsx' if fmt == 'xlsx' else 'csv'}"
        if fmt == "xlsx":
            return _crm_re_export_xlsx_sheets([("Ngân sách", headers, rows)], filename=fname)
        return _crm_re_export_csv(headers, rows, filename=fname)

    if report == "accounting":
        with get_connection() as conn:
            try:
                sheets = build_accounting_export_sheets(conn, project_id)
            except ValueError as e:
                return jsonify({"error": str(e)}), 404
        fname = f"{base}-ke-toan.xlsx" if fmt == "xlsx" else f"{base}-ke-toan.csv"
        if fmt == "xlsx":
            return _crm_re_export_xlsx_sheets(sheets, filename=fname)
        summary = sheets[0][2] if sheets else []
        return _crm_re_export_csv(["Trường", "Giá trị"], summary, filename=fname)

    if report == "plans":
        headers, rows = project_export_plan_rows(proj)
        fname = f"{base}-ke-hoach.{'xlsx' if fmt == 'xlsx' else 'csv'}"
        if fmt == "xlsx":
            return _crm_re_export_xlsx_sheets([("Kế hoạch", headers, rows)], filename=fname)
        return _crm_re_export_csv(headers, rows, filename=fname)

    # full — Excel đa sheet
    summary_rows = project_export_summary_rows(proj, pack["summary"], pack["workflow"])
    wf_h, wf_r = project_export_workflow_rows(pack["workflow"])
    plan_h, plan_r = project_export_plan_rows(proj)
    kpi_h, kpi_r = project_export_kpi_rows(pack["kpis"])
    prod_h, prod_r = project_export_product_rows(pack["products"])
    risk_h, risk_r = project_export_risk_rows(pack["risks"])
    bud_h, bud_r = project_export_budget_rows(pack["budget"])
    sheets = [
        ("Tóm tắt", ["Trường", "Giá trị"], summary_rows),
        ("Quy trình", wf_h, wf_r),
        ("Kế hoạch", plan_h, plan_r),
        ("KPI", kpi_h, kpi_r),
        ("Tồn kho", prod_h, prod_r),
        ("Rủi ro", risk_h, risk_r),
        ("Ngân sách", bud_h, bud_r),
    ]
    return _crm_re_export_xlsx_sheets(sheets, filename=f"{base}-tong-hop.xlsx")


@app.get("/api/crm/kpi/metrics")
def api_crm_list_kpi_metrics() -> Any:
    deny = _crm_kpi_access_json_error()
    if deny is not None:
        return deny
    raw = (request.args.get("include_inactive") or "").strip().lower()
    include_inactive = raw in ("1", "true", "yes", "all")
    with get_connection() as conn:
        eff = _crm_effective_staff_id()
        if eff is not None:
            position_id = _crm_staff_position_id(conn, eff)
            metric_ids = _crm_position_metric_ids(conn, position_id)
            if not metric_ids:
                return jsonify({"metrics": []})
            placeholders = ",".join("?" * len(metric_ids))
            rows = conn.execute(
                f"""
                SELECT * FROM crm_kpi_metrics
                WHERE id IN ({placeholders}) AND active = 1
                ORDER BY sort_order ASC, name COLLATE NOCASE ASC
                """,
                metric_ids,
            ).fetchall()
            return jsonify({"metrics": rows_to_dict(rows)})
        if include_inactive:
            rows = conn.execute(
                """
                SELECT * FROM crm_kpi_metrics
                ORDER BY active DESC, sort_order ASC, name COLLATE NOCASE ASC
                """
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT * FROM crm_kpi_metrics
                WHERE active = 1
                ORDER BY sort_order ASC, name COLLATE NOCASE ASC
                """
            ).fetchall()
    return jsonify({"metrics": rows_to_dict(rows)})


@app.post("/api/crm/kpi/metrics")
def api_crm_create_kpi_metric() -> Any:
    deny = _crm_kpi_require_edit_json()
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    code = str(payload.get("code", "")).strip()[:64]
    name = str(payload.get("name", "")).strip()[:240]
    if not name:
        return jsonify({"error": "Thiếu tên chỉ tiêu"}), 400
    unit = str(payload.get("unit", "")).strip()[:64]
    desc = str(payload.get("description", "")).strip()[:2000]
    try:
        so = int(payload.get("sort_order") or 0)
    except (TypeError, ValueError):
        so = 0
    hi = 1
    if "higher_is_better" in payload:
        hi = 1 if payload["higher_is_better"] in (True, 1, "1", "true", "yes") else 0
    wr: float | None = None
    if "warn_ratio" in payload:
        wrv = payload["warn_ratio"]
        if wrv is None or wrv == "":
            wr = None
        else:
            try:
                wr = float(wrv)
            except (TypeError, ValueError):
                wr = None
    ts_d = datetime.now().strftime("%Y-%m-%d")
    ts = _crm_ts()
    with get_connection() as conn:
        if code and conn.execute(
            "SELECT 1 FROM crm_kpi_metrics WHERE lower(trim(code)) = lower(?) AND trim(code) != ''",
            (code,),
        ).fetchone():
            return jsonify({"error": "Mã chỉ tiêu đã tồn tại"}), 400
        cur = conn.execute(
            """
            INSERT INTO crm_kpi_metrics (
                code, name, unit, description, sort_order, active,
                created_at, updated_at, higher_is_better, warn_ratio
            )
            VALUES (?, ?, ?, ?, ?, 1, ?, ?, ?, ?)
            """,
            (code, name, unit, desc, so, ts_d, ts, hi, wr),
        )
        mid = int(cur.lastrowid)
        row = conn.execute("SELECT * FROM crm_kpi_metrics WHERE id = ?", (mid,)).fetchone()
    assert row is not None
    return jsonify(dict(row)), 201


@app.patch("/api/crm/kpi/metrics/<int:metric_id>")
def api_crm_patch_kpi_metric(metric_id: int) -> Any:
    deny = _crm_kpi_require_edit_json()
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM crm_kpi_metrics WHERE id = ?", (metric_id,)).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy chỉ tiêu"}), 404
        merged: dict[str, Any] = dict(row)
        if "code" in payload and isinstance(payload["code"], str):
            merged["code"] = payload["code"].strip()[:64]
        if "name" in payload and isinstance(payload["name"], str):
            n = payload["name"].strip()[:240]
            if not n:
                return jsonify({"error": "Tên không được trống"}), 400
            merged["name"] = n
        if "unit" in payload and isinstance(payload["unit"], str):
            merged["unit"] = payload["unit"].strip()[:64]
        if "description" in payload and isinstance(payload["description"], str):
            merged["description"] = payload["description"].strip()[:2000]
        if "sort_order" in payload and payload["sort_order"] is not None:
            try:
                merged["sort_order"] = int(payload["sort_order"])
            except (TypeError, ValueError):
                pass
        if "active" in payload:
            merged["active"] = 1 if payload["active"] in (True, 1, "1", "true") else 0
        if "higher_is_better" in payload:
            merged["higher_is_better"] = (
                1 if payload["higher_is_better"] in (True, 1, "1", "true", "yes") else 0
            )
        if "warn_ratio" in payload:
            wrv = payload["warn_ratio"]
            if wrv is None or wrv == "":
                merged["warn_ratio"] = None
            else:
                try:
                    merged["warn_ratio"] = float(wrv)
                except (TypeError, ValueError):
                    pass

        c = str(merged.get("code") or "").strip()
        if c and conn.execute(
            """
            SELECT 1 FROM crm_kpi_metrics
            WHERE lower(trim(code)) = lower(?) AND trim(code) != '' AND id != ?
            """,
            (c, metric_id),
        ).fetchone():
            return jsonify({"error": "Mã chỉ tiêu đã tồn tại"}), 400

        ts = _crm_ts()
        hi_m = int(merged.get("higher_is_better") or 1)
        wr_m = merged.get("warn_ratio")
        conn.execute(
            """
            UPDATE crm_kpi_metrics
            SET code = ?, name = ?, unit = ?, description = ?, sort_order = ?, active = ?,
                higher_is_better = ?, warn_ratio = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                str(merged.get("code") or ""),
                merged["name"],
                str(merged.get("unit") or ""),
                str(merged.get("description") or ""),
                int(merged.get("sort_order") or 0),
                int(merged.get("active") or 0),
                hi_m,
                wr_m,
                ts,
                metric_id,
            ),
        )
        row2 = conn.execute("SELECT * FROM crm_kpi_metrics WHERE id = ?", (metric_id,)).fetchone()
    assert row2 is not None
    return jsonify(dict(row2))


_KPI_EXPORT_HEADERS = [
    "Năm",
    "Tháng",
    "Mã NV",
    "Họ tên",
    "Mã chỉ tiêu",
    "Chỉ tiêu",
    "Đơn vị",
    "Hướng",
    "Ngưỡng cảnh báo",
    "Mục tiêu",
    "Thực tế",
    "% đạt",
    "Trạng thái",
    "Mức cảnh báo",
    "Chi tiết cảnh báo",
    "Ghi chú",
]


def _crm_kpi_export_workbook(rows: list[sqlite3.Row]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "KPI nhân viên"
    ws.append(list(_KPI_EXPORT_HEADERS))
    for r in rows:
        d = dict(r)
        st = str(d.get("status") or "draft")
        st_lab = CRM_KPI_STATUS_LABELS_VI.get(st, st)
        hi = int(d.get("higher_is_better") or 1)
        dir_lab = "Cao hơn tốt" if hi == 1 else "Thấp hơn tốt"
        wrv = d.get("warn_ratio")
        wr_s = ""
        if wrv is not None:
            try:
                wr_s = str(float(wrv))
            except (TypeError, ValueError):
                wr_s = ""
        ach = _crm_kpi_achievement_pct(hi, d.get("target_value"), d.get("actual_value"))
        ach_s = "" if ach is None else str(ach)
        lev, reason = _crm_kpi_derive_alert(
            status=st,
            higher_is_better=hi,
            warn_ratio=wrv,
            target_value=d.get("target_value"),
            actual_value=d.get("actual_value"),
        )
        alert_lv = ""
        if lev == "critical":
            alert_lv = "Nghiêm trọng"
        elif lev == "warn":
            alert_lv = "Cảnh báo"
        msg = _crm_kpi_alert_label_vi(lev, reason) if lev else ""
        ws.append(
            [
                int(d.get("year") or 0),
                int(d.get("month") or 0),
                str(d.get("staff_code") or ""),
                str(d.get("staff_name") or ""),
                str(d.get("metric_code") or ""),
                str(d.get("metric_name") or ""),
                str(d.get("metric_unit") or ""),
                dir_lab,
                wr_s,
                d.get("target_value"),
                d.get("actual_value"),
                ach_s,
                st_lab,
                alert_lv,
                msg,
                str(d.get("note") or ""),
            ]
        )
    return wb


@app.get("/api/crm/kpi/alerts")
def api_crm_kpi_alerts() -> Any:
    deny = _crm_kpi_access_json_error()
    if deny is not None:
        return deny
    year, month, staff_id, qerr = _crm_parse_staff_kpi_query_args()
    if qerr or year is None or month is None:
        return jsonify({"error": qerr or "Tham số không hợp lệ"}), 400
    clauses = ["k.year = ?", "k.month = ?"]
    params: list[Any] = [year, month]
    if staff_id is not None:
        clauses.append("k.staff_id = ?")
        params.append(staff_id)
    where_sql = " AND ".join(clauses)
    with get_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT k.*,
                   m.name AS metric_name, m.code AS metric_code,
                   m.higher_is_better AS metric_higher_is_better,
                   m.warn_ratio AS metric_warn_ratio,
                   s.name AS staff_name, s.internal_code AS staff_code
            FROM crm_staff_kpi k
            JOIN crm_kpi_metrics m ON m.id = k.metric_id
            JOIN crm_staff s ON s.id = k.staff_id
            WHERE {where_sql}
            ORDER BY s.name COLLATE NOCASE ASC, m.sort_order ASC, m.name COLLATE NOCASE ASC
            """,
            params,
        ).fetchall()
    alerts_out: list[dict[str, Any]] = []
    crit = 0
    wrn = 0
    for r in rows:
        d = dict(r)
        st = str(d.get("status") or "draft")
        hi = int(d.get("metric_higher_is_better") or 1)
        wr = d.get("metric_warn_ratio")
        lev, reason = _crm_kpi_derive_alert(
            status=st,
            higher_is_better=hi,
            warn_ratio=wr,
            target_value=d.get("target_value"),
            actual_value=d.get("actual_value"),
        )
        if not lev:
            continue
        if lev == "critical":
            crit += 1
        elif lev == "warn":
            wrn += 1
        alerts_out.append(
            {
                "level": lev,
                "reason": reason,
                "message": _crm_kpi_alert_label_vi(lev, reason),
                "kpi_id": d.get("id"),
                "staff_id": d.get("staff_id"),
                "staff_name": d.get("staff_name"),
                "staff_code": d.get("staff_code"),
                "metric_id": d.get("metric_id"),
                "metric_name": d.get("metric_name"),
                "metric_code": d.get("metric_code"),
                "target_value": d.get("target_value"),
                "actual_value": d.get("actual_value"),
                "status": st,
            }
        )
    return jsonify(
        {
            "alerts": alerts_out,
            "summary": {"critical": crit, "warn": wrn},
            "year": year,
            "month": month,
        }
    )


@app.get("/api/crm/kpi/chart")
def api_crm_kpi_chart() -> Any:
    deny = _crm_kpi_access_json_error()
    if deny is not None:
        return deny
    try:
        metric_id = int((request.args.get("metric_id") or "0").strip())
    except (TypeError, ValueError):
        return jsonify({"error": "metric_id không hợp lệ"}), 400
    if metric_id <= 0:
        return jsonify({"error": "Cần metric_id (chỉ tiêu để vẽ biểu đồ)"}), 400
    year, month, staff_id, qerr = _crm_parse_staff_kpi_query_args()
    if qerr or year is None or month is None:
        return jsonify({"error": qerr or "Tham số không hợp lệ"}), 400
    clauses = ["k.year = ?", "k.month = ?", "k.metric_id = ?"]
    params: list[Any] = [year, month, metric_id]
    if staff_id is not None:
        clauses.append("k.staff_id = ?")
        params.append(staff_id)
    where_sql = " AND ".join(clauses)
    with get_connection() as conn:
        mrow = conn.execute(
            "SELECT id, name, code, unit, higher_is_better FROM crm_kpi_metrics WHERE id = ?",
            (metric_id,),
        ).fetchone()
        if mrow is None:
            return jsonify({"error": "Không tìm thấy chỉ tiêu"}), 404
        hi = int(mrow["higher_is_better"] or 1)
        rows = conn.execute(
            f"""
            SELECT k.*, s.name AS staff_name, s.internal_code AS staff_code
            FROM crm_staff_kpi k
            JOIN crm_staff s ON s.id = k.staff_id
            WHERE {where_sql}
            ORDER BY s.name COLLATE NOCASE ASC
            """,
            params,
        ).fetchall()
    labels: list[str] = []
    achievement: list[float | None] = []
    staff_ids: list[int] = []
    for r in rows:
        d = dict(r)
        sid = int(d["staff_id"])
        sn = str(d.get("staff_name") or "").strip()
        sc = str(d.get("staff_code") or "").strip()
        lab = f"{sn}" + (f" ({sc})" if sc else "")
        labels.append(lab)
        staff_ids.append(sid)
        ach = _crm_kpi_achievement_pct(hi, d.get("target_value"), d.get("actual_value"))
        achievement.append(ach)
    return jsonify(
        {
            "metric": dict(mrow),
            "higher_is_better": hi,
            "year": year,
            "month": month,
            "labels": labels,
            "achievement_pct": achievement,
            "staff_ids": staff_ids,
        }
    )


@app.get("/api/crm/staff/kpi/export")
def api_crm_export_staff_kpi() -> Any:
    deny = _crm_kpi_access_json_error()
    if deny is not None:
        return deny
    year, month, staff_id, qerr = _crm_parse_staff_kpi_query_args()
    if qerr or year is None or month is None:
        return jsonify({"error": qerr or "Tham số không hợp lệ"}), 400
    clauses = ["k.year = ?", "k.month = ?"]
    params: list[Any] = [year, month]
    if staff_id is not None:
        clauses.append("k.staff_id = ?")
        params.append(staff_id)
    where_sql = " AND ".join(clauses)
    with get_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT k.*,
                   m.name AS metric_name, m.code AS metric_code, m.unit AS metric_unit,
                   m.higher_is_better AS higher_is_better, m.warn_ratio AS warn_ratio,
                   s.name AS staff_name, s.internal_code AS staff_code
            FROM crm_staff_kpi k
            JOIN crm_kpi_metrics m ON m.id = k.metric_id
            JOIN crm_staff s ON s.id = k.staff_id
            WHERE {where_sql}
            ORDER BY s.name COLLATE NOCASE ASC, m.sort_order ASC, m.name COLLATE NOCASE ASC
            """,
            params,
        ).fetchall()
    wb = _crm_kpi_export_workbook(rows)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = datetime.now().strftime("%Y%m%d-%H%M")
    fn = f"crm-kpi-{year}-{month:02d}-{stamp}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fn,
        etag=False,
        max_age=0,
        conditional=False,
    )


@app.get("/api/crm/staff/kpi")
def api_crm_list_staff_kpi() -> Any:
    deny = _crm_kpi_access_json_error()
    if deny is not None:
        return deny
    year, month, staff_id, qerr = _crm_parse_staff_kpi_query_args()
    if qerr or year is None or month is None:
        return jsonify({"error": qerr or "Tham số không hợp lệ"}), 400
    clauses = ["k.year = ?", "k.month = ?"]
    params: list[Any] = [year, month]
    if staff_id is not None:
        clauses.append("k.staff_id = ?")
        params.append(staff_id)
    with get_connection() as conn:
        eff = _crm_effective_staff_id()
        if eff is not None:
            position_id = _crm_staff_position_id(conn, eff)
            _crm_ensure_staff_kpi_rows(conn, eff, position_id, year, month)
            metric_ids = _crm_position_metric_ids(conn, position_id)
            if metric_ids:
                placeholders = ",".join("?" * len(metric_ids))
                clauses.append(f"k.metric_id IN ({placeholders})")
                params.extend(metric_ids)
        where_sql = " AND ".join(clauses)
        rows = conn.execute(
            f"""
            SELECT k.*,
                   m.name AS metric_name, m.code AS metric_code, m.unit AS metric_unit,
                   m.higher_is_better AS metric_higher_is_better,
                   m.warn_ratio AS metric_warn_ratio,
                   s.name AS staff_name, s.internal_code AS staff_code
            FROM crm_staff_kpi k
            JOIN crm_kpi_metrics m ON m.id = k.metric_id
            JOIN crm_staff s ON s.id = k.staff_id
            WHERE {where_sql}
            ORDER BY s.name COLLATE NOCASE ASC, m.sort_order ASC, m.name COLLATE NOCASE ASC
            """,
            params,
        ).fetchall()
    return jsonify({"staff_kpi": rows_to_dict(rows)})


@app.post("/api/crm/staff/kpi")
def api_crm_upsert_staff_kpi() -> Any:
    if _staff_logged_in():
        return jsonify({"error": "Nhân viên cập nhật tiến độ qua nút Lưu trên từng dòng KPI."}), 403
    deny = _crm_kpi_require_edit_json()
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    try:
        sid = int(payload.get("staff_id") or 0)
        mid = int(payload.get("metric_id") or 0)
        year = int(payload.get("year") or 0)
        month = int(payload.get("month") or 0)
    except (TypeError, ValueError):
        return jsonify({"error": "staff_id, metric_id, year, month không hợp lệ"}), 400
    if sid <= 0 or mid <= 0 or year < 2000 or year > 2100 or month < 1 or month > 12:
        return jsonify({"error": "Tham số KPI không hợp lệ"}), 400
    tv = payload.get("target_value")
    av = payload.get("actual_value")
    try:
        target_v = None if tv is None or tv == "" else float(tv)
    except (TypeError, ValueError):
        return jsonify({"error": "target_value không hợp lệ"}), 400
    try:
        actual_v = None if av is None or av == "" else float(av)
    except (TypeError, ValueError):
        return jsonify({"error": "actual_value không hợp lệ"}), 400
    status = _crm_kpi_normalize_status(str(payload.get("status") or "draft"))
    note = str(payload.get("note") or "").strip()[:2000]
    ts = _crm_ts()
    ts_d = datetime.now().strftime("%Y-%m-%d")
    with get_connection() as conn:
        if conn.execute("SELECT id FROM crm_staff WHERE id = ?", (sid,)).fetchone() is None:
            return jsonify({"error": "Không tìm thấy nhân viên"}), 404
        mrow = conn.execute("SELECT active FROM crm_kpi_metrics WHERE id = ?", (mid,)).fetchone()
        if mrow is None:
            return jsonify({"error": "Không tìm thấy chỉ tiêu"}), 404
        if int(mrow["active"] or 0) != 1:
            return jsonify({"error": "Chỉ tiêu đã ngưng — bật lại hoặc chọn chỉ tiêu khác"}), 400
        conn.execute(
            """
            INSERT INTO crm_staff_kpi (
                staff_id, metric_id, year, month, target_value, actual_value, status, note, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(staff_id, metric_id, year, month) DO UPDATE SET
                target_value = excluded.target_value,
                actual_value = excluded.actual_value,
                status = excluded.status,
                note = excluded.note,
                updated_at = excluded.updated_at
            """,
            (sid, mid, year, month, target_v, actual_v, status, note, ts_d, ts),
        )
        row = conn.execute(
            """
            SELECT k.*,
                   m.name AS metric_name, m.code AS metric_code, m.unit AS metric_unit,
                   s.name AS staff_name, s.internal_code AS staff_code
            FROM crm_staff_kpi k
            JOIN crm_kpi_metrics m ON m.id = k.metric_id
            JOIN crm_staff s ON s.id = k.staff_id
            WHERE k.staff_id = ? AND k.metric_id = ? AND k.year = ? AND k.month = ?
            """,
            (sid, mid, year, month),
        ).fetchone()
    assert row is not None
    return jsonify(dict(row)), 201


@app.patch("/api/crm/staff/kpi/<int:kpi_id>")
def api_crm_patch_staff_kpi(kpi_id: int) -> Any:
    deny = _crm_kpi_require_edit_json()
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM crm_staff_kpi WHERE id = ?", (kpi_id,)).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy KPI"}), 404
        eff = _crm_effective_staff_id()
        if eff is not None:
            if int(row["staff_id"]) != eff:
                return jsonify({"error": "Không sửa KPI của nhân viên khác."}), 403
            if not _crm_staff_metric_allowed(conn, eff, int(row["metric_id"])):
                return jsonify({"error": "Chỉ tiêu không thuộc chức vụ của bạn."}), 403
            merged = dict(row)
            merged.update(_crm_kpi_staff_progress_fields(payload))
        else:
            merged = dict(row)
            if "target_value" in payload:
                tv = payload["target_value"]
                try:
                    merged["target_value"] = None if tv is None or tv == "" else float(tv)
                except (TypeError, ValueError):
                    return jsonify({"error": "target_value không hợp lệ"}), 400
            if "actual_value" in payload:
                av = payload["actual_value"]
                try:
                    merged["actual_value"] = None if av is None or av == "" else float(av)
                except (TypeError, ValueError):
                    return jsonify({"error": "actual_value không hợp lệ"}), 400
            if "status" in payload:
                merged["status"] = _crm_kpi_normalize_status(str(payload.get("status")))
            if "note" in payload and isinstance(payload["note"], str):
                merged["note"] = payload["note"].strip()[:2000]
        ts = _crm_ts()
        conn.execute(
            """
            UPDATE crm_staff_kpi
            SET target_value = ?, actual_value = ?, status = ?, note = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                merged.get("target_value"),
                merged.get("actual_value"),
                merged["status"],
                merged.get("note") or "",
                ts,
                kpi_id,
            ),
        )
        row2 = conn.execute(
            """
            SELECT k.*,
                   m.name AS metric_name, m.code AS metric_code, m.unit AS metric_unit,
                   s.name AS staff_name, s.internal_code AS staff_code
            FROM crm_staff_kpi k
            JOIN crm_kpi_metrics m ON m.id = k.metric_id
            JOIN crm_staff s ON s.id = k.staff_id
            WHERE k.id = ?
            """,
            (kpi_id,),
        ).fetchone()
    assert row2 is not None
    return jsonify(dict(row2))


@app.delete("/api/crm/staff/kpi/<int:kpi_id>")
def api_crm_delete_staff_kpi(kpi_id: int) -> Any:
    if _staff_logged_in():
        return jsonify({"error": "Nhân viên không được xóa bản ghi KPI."}), 403
    deny = _crm_kpi_require_edit_json()
    if deny is not None:
        return deny
    with get_connection() as conn:
        cur = conn.execute("DELETE FROM crm_staff_kpi WHERE id = ?", (kpi_id,))
        if cur.rowcount == 0:
            return jsonify({"error": "Không tìm thấy KPI"}), 404
    return jsonify({"ok": True})


@app.get("/api/crm/customers")
def api_crm_list_customers() -> Any:
    q_raw = (request.args.get("q") or "").strip().lower()
    overview = (request.args.get("overview") or "").strip().lower() in ("1", "true", "yes")
    active_only = (request.args.get("active_only") or "").strip().lower() in ("1", "true", "yes")
    try:
        lim = min(500, max(1, int(request.args.get("limit") or 200)))
    except (TypeError, ValueError):
        lim = 200
    portal_sid = _crm_effective_staff_id()
    if overview or portal_sid is not None:
        if portal_sid is not None and not _admin_section_can("crm_board_customers", "view"):
            return _admin_section_forbidden_json("crm_board_customers", "view")
        with get_connection() as conn:
            customers = _crm_customers_overview_rows(
                conn,
                portal_sid=portal_sid,
                q_raw=q_raw,
                limit=lim,
                active_only=active_only or portal_sid is not None,
            )
        total = len(customers)
        stats = {
            "total": total,
            "active_care": sum(1 for c in customers if int(c.get("cases_open") or 0) > 0),
            "no_care_report": sum(1 for c in customers if not c.get("last_care_status")),
            "with_contracts": sum(1 for c in customers if int(c.get("contracts_total") or 0) > 0),
            "issues_open": sum(int(c.get("issues_open") or 0) for c in customers),
        }
        return jsonify({"customers": customers, "stats": stats, "staff_id": portal_sid})
    with get_connection() as conn:
        hide_placeholder = _crm_presales_on_lead_enabled()
        ph_clause = " AND COALESCE(is_placeholder, 0) = 0" if hide_placeholder else ""
        if q_raw:
            like = f"%{q_raw}%"
            rows = conn.execute(
                f"""
                SELECT id, name, phone, email, address, company, created_at
                FROM crm_customers
                WHERE (
                    lower(coalesce(trim(name), '')) LIKE ?
                   OR lower(coalesce(trim(phone), '')) LIKE ?
                   OR lower(coalesce(trim(email), '')) LIKE ?
                   OR lower(coalesce(trim(address), '')) LIKE ?
                   OR lower(coalesce(trim(company), '')) LIKE ?
                ){ph_clause}
                ORDER BY id DESC
                LIMIT ?
                """,
                (like, like, like, like, like, lim),
            ).fetchall()
        else:
            rows = conn.execute(
                f"""
                SELECT id, name, phone, email, address, company, created_at
                FROM crm_customers
                WHERE 1=1{ph_clause}
                ORDER BY id DESC
                LIMIT ?
                """,
                (lim,),
            ).fetchall()
    return jsonify({"customers": rows_to_dict(rows)})


@app.get("/api/crm/customers/<int:customer_id>")
def api_crm_customer_detail(customer_id: int) -> Any:
    portal_sid = _crm_effective_staff_id()
    if not _admin_section_can("crm_board_customers", "view"):
        return _admin_section_forbidden_json("crm_board_customers", "view")
    with get_connection() as conn:
        cu = conn.execute(
            "SELECT * FROM crm_customers WHERE id = ?",
            (customer_id,),
        ).fetchone()
        if cu is None:
            return jsonify({"error": "Không tìm thấy khách hàng"}), 404
        if portal_sid is not None and not _crm_customer_staff_can_access(
            conn, customer_id, portal_sid
        ):
            return _crm_forbid_staff_case()
        ch_labels = _crm_lead_channel_labels_map(conn)
        case_sql = f"{_CRM_CASE_SELECT} WHERE c.customer_id = ?"
        case_params: list[Any] = [customer_id]
        if portal_sid is not None:
            case_sql += " AND c.assigned_staff_id = ?"
            case_params.append(portal_sid)
        case_sql += " ORDER BY datetime(c.updated_at) DESC"
        case_rows = conn.execute(case_sql, case_params).fetchall()
        cases = [_crm_row_case(r, ch_labels) for r in case_rows]
        if portal_sid is not None:
            cases = [
                c
                for c in cases
                if normalize_pipeline_stage(c.get("pipeline_stage") or c.get("status"))
                not in TERMINAL_STAGES
            ]
        case_ids = [int(c["id"]) for c in cases]
        last_care = fetch_last_care_reports_map(conn, case_ids) if case_ids else {}
        for c in cases:
            lc = last_care.get(int(c["id"]))
            if lc:
                c["last_care_report"] = lc
        care_sql = """
            SELECT r.*, c.title AS case_title
            FROM crm_care_reports r
            JOIN crm_cases c ON c.id = r.case_id
            WHERE c.customer_id = ?
        """
        care_params: list[Any] = [customer_id]
        if portal_sid is not None:
            care_sql += " AND c.assigned_staff_id = ?"
            care_params.append(portal_sid)
        care_sql += " ORDER BY r.id DESC LIMIT 40"
        care_reports = conn.execute(care_sql, care_params).fetchall()
        contracts: list[dict[str, Any]] = []
        if portal_sid is None:
            ct_rows = conn.execute(
                """
                SELECT ct.*, cu.name AS customer_name
                FROM crm_contracts ct
                JOIN crm_customers cu ON cu.id = ct.customer_id
                WHERE ct.customer_id = ?
                ORDER BY ct.id DESC
                LIMIT 20
                """,
                (customer_id,),
            ).fetchall()
            contracts = [dict(r) for r in ct_rows]
            for ct in contracts:
                st = str(ct.get("status") or "")
                ct["status_label"] = CRM_CONTRACT_STATUS_LABELS_VI.get(st, st)
        open_cases = [
            c
            for c in cases
            if normalize_pipeline_stage(c.get("pipeline_stage") or c.get("status"))
            not in TERMINAL_STAGES
        ]
        cu_dict = enrich_customer_row(cu)
        lead_sources = compute_lead_sources(conn, customer_id, cu_dict, ch_labels)
        relations = fetch_customer_relations(conn, customer_id)
        purchases = fetch_customer_purchases(conn, customer_id)
        issues = fetch_customer_issues(conn, customer_id, portal_staff_id=portal_sid)
        issues_open = sum(
            1 for i in issues if str(i.get("status") or "") not in ("da_xu_ly", "dong")
        )
    return jsonify(
        {
            "customer": cu_dict,
            "lead_sources": lead_sources,
            "relations": relations,
            "purchases": purchases,
            "issues": issues,
            "cases": cases,
            "care_reports": [care_report_row_to_dict(r) for r in care_reports],
            "contracts": contracts,
            "stats": {
                "cases_total": len(cases),
                "cases_open": len(open_cases),
                "contracts_total": len(contracts),
                "care_reports_total": len(care_reports),
                "purchases_total": len(purchases),
                "relations_total": len(relations),
                "issues_open": issues_open,
                "issues_total": len(issues),
            },
        }
    )


@app.post("/api/crm/customers")
def api_crm_create_customer() -> Any:
    if not _admin_section_can_customer_write("create"):
        return _admin_section_forbidden_json("crm_board_customers", "create")
    if _crm_effective_staff_id() is not None:
        return jsonify({"error": "Nhân viên không thể tạo khách hàng mới."}), 403
    payload = request.get_json(force=True) or {}
    name = str(payload.get("name") or "").strip()[:240]
    phone = str(payload.get("phone") or "").strip()[:64]
    email = str(payload.get("email") or "").strip()[:240]
    address = str(payload.get("address") or "").strip()[:500]
    company = str(payload.get("company") or "").strip()[:240]
    lead_source = normalize_lead_source(str(payload.get("lead_source") or ""))
    if not name:
        return jsonify({"error": "Cần tên khách hàng"}), 400
    if not phone and not email:
        return jsonify({"error": "Cần ít nhất số điện thoại hoặc email"}), 400
    ts = _crm_ts()
    extra = {
        "lead_source": lead_source,
        "lead_source_note": str(payload.get("lead_source_note") or "").strip()[:4000],
        "date_of_birth": str(payload.get("date_of_birth") or "").strip()[:32],
        "gender": normalize_gender(str(payload.get("gender") or "")),
        "id_number": str(payload.get("id_number") or "").strip()[:32],
        "occupation": str(payload.get("occupation") or "").strip()[:240],
        "interests": str(payload.get("interests") or "").strip()[:4000],
        "profile_notes": str(payload.get("profile_notes") or "").strip()[:4000],
    }
    with get_connection() as conn:
        cur = conn.execute(
            f"""
            INSERT INTO crm_customers ({PROFILE_INSERT_COLS})
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                name,
                phone,
                email,
                address,
                company,
                extra["lead_source"],
                extra["lead_source_note"],
                extra["date_of_birth"],
                extra["gender"],
                extra["id_number"],
                extra["occupation"],
                extra["interests"],
                extra["profile_notes"],
                ts,
            ),
        )
        cid = int(cur.lastrowid)
        row = conn.execute("SELECT * FROM crm_customers WHERE id = ?", (cid,)).fetchone()
    assert row is not None
    return jsonify(enrich_customer_row(row)), 201


@app.patch("/api/crm/customers/<int:customer_id>")
def api_crm_patch_customer(customer_id: int) -> Any:
    if not _admin_section_can("crm_board_customers", "edit"):
        return _admin_section_forbidden_json("crm_board_customers", "edit")
    if _crm_effective_staff_id() is not None:
        return jsonify({"error": "Nhân viên không thể sửa hồ sơ khách hàng."}), 403
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute(
            "SELECT * FROM crm_customers WHERE id = ?",
            (customer_id,),
        ).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy khách hàng"}), 404
        merged = dict(row)
        apply_profile_patch(merged, payload)
        if not str(merged.get("name") or "").strip():
            return jsonify({"error": "Tên không được trống"}), 400
        if not str(merged.get("phone") or "").strip() and not str(merged.get("email") or "").strip():
            return jsonify({"error": "Cần ít nhất SĐT hoặc email"}), 400
        vals = profile_update_sql_values(merged)
        conn.execute(
            """
            UPDATE crm_customers
            SET name = ?, phone = ?, email = ?, address = ?, company = ?,
                lead_source = ?, lead_source_note = ?, date_of_birth = ?, gender = ?,
                id_number = ?, occupation = ?, interests = ?, profile_notes = ?
            WHERE id = ?
            """,
            (*vals, customer_id),
        )
        out = conn.execute("SELECT * FROM crm_customers WHERE id = ?", (customer_id,)).fetchone()
    assert out is not None
    return jsonify(enrich_customer_row(out))


def _crm_customer_access_or_403(
    conn: sqlite3.Connection, customer_id: int
) -> Any | None:
    if not _admin_section_can("crm_board_customers", "view"):
        return _admin_section_forbidden_json("crm_board_customers", "view")
    portal_sid = _crm_effective_staff_id()
    row = conn.execute("SELECT id FROM crm_customers WHERE id = ?", (customer_id,)).fetchone()
    if row is None:
        return jsonify({"error": "Không tìm thấy khách hàng"}), 404
    if portal_sid is not None and not _crm_customer_staff_can_access(
        conn, customer_id, portal_sid
    ):
        return _crm_forbid_staff_case()
    return None


def _crm_customer_admin_write_only() -> Any | None:
    if _crm_effective_staff_id() is not None:
        return jsonify({"error": "Chỉ quản trị viên mới thực hiện thao tác này."}), 403
    if not _admin_section_can("crm_board_customers", "edit"):
        return _admin_section_forbidden_json("crm_board_customers", "edit")
    return None


@app.post("/api/crm/customers/<int:customer_id>/relations")
def api_crm_customer_relation_create(customer_id: int) -> Any:
    deny = _crm_customer_admin_write_only()
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    ts = _crm_ts()
    with get_connection() as conn:
        err = _crm_customer_access_or_403(conn, customer_id)
        if err is not None:
            return err
        full_name = str(payload.get("full_name") or "").strip()[:240]
        if not full_name:
            return jsonify({"error": "Cần họ tên người liên quan"}), 400
        cur = conn.execute(
            """
            INSERT INTO crm_customer_relations (
                customer_id, relation_type, full_name, phone, email, notes, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                customer_id,
                normalize_relation_type(str(payload.get("relation_type") or "")),
                full_name,
                str(payload.get("phone") or "").strip()[:64],
                str(payload.get("email") or "").strip()[:240],
                str(payload.get("notes") or "").strip()[:2000],
                ts,
                ts,
            ),
        )
        rid = int(cur.lastrowid)
        row = conn.execute(
            "SELECT * FROM crm_customer_relations WHERE id = ?", (rid,)
        ).fetchone()
    assert row is not None
    return jsonify(_relation_row(row)), 201


@app.patch("/api/crm/customers/<int:customer_id>/relations/<int:relation_id>")
def api_crm_customer_relation_patch(customer_id: int, relation_id: int) -> Any:
    deny = _crm_customer_admin_write_only()
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    ts = _crm_ts()
    with get_connection() as conn:
        err = _crm_customer_access_or_403(conn, customer_id)
        if err is not None:
            return err
        row = conn.execute(
            """
            SELECT * FROM crm_customer_relations
            WHERE id = ? AND customer_id = ?
            """,
            (relation_id, customer_id),
        ).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy quan hệ"}), 404
        merged = dict(row)
        if "relation_type" in payload:
            merged["relation_type"] = normalize_relation_type(str(payload.get("relation_type") or ""))
        for key in ("full_name", "phone", "email", "notes"):
            if key in payload and isinstance(payload[key], str):
                merged[key] = payload[key].strip()[:240 if key != "notes" else 2000]
        if not str(merged.get("full_name") or "").strip():
            return jsonify({"error": "Họ tên không được trống"}), 400
        conn.execute(
            """
            UPDATE crm_customer_relations
            SET relation_type = ?, full_name = ?, phone = ?, email = ?, notes = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                merged["relation_type"],
                merged["full_name"],
                merged["phone"],
                merged["email"],
                merged["notes"],
                ts,
                relation_id,
            ),
        )
        out = conn.execute(
            "SELECT * FROM crm_customer_relations WHERE id = ?", (relation_id,)
        ).fetchone()
    assert out is not None
    return jsonify(_relation_row(out))


@app.delete("/api/crm/customers/<int:customer_id>/relations/<int:relation_id>")
def api_crm_customer_relation_delete(customer_id: int, relation_id: int) -> Any:
    deny = _crm_customer_admin_write_only()
    if deny is not None:
        return deny
    with get_connection() as conn:
        err = _crm_customer_access_or_403(conn, customer_id)
        if err is not None:
            return err
        cur = conn.execute(
            "DELETE FROM crm_customer_relations WHERE id = ? AND customer_id = ?",
            (relation_id, customer_id),
        )
        if cur.rowcount == 0:
            return jsonify({"error": "Không tìm thấy quan hệ"}), 404
    return jsonify({"ok": True})


@app.post("/api/crm/customers/<int:customer_id>/purchases")
def api_crm_customer_purchase_create(customer_id: int) -> Any:
    deny = _crm_customer_admin_write_only()
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    ts = _crm_ts()
    with get_connection() as conn:
        err = _crm_customer_access_or_403(conn, customer_id)
        if err is not None:
            return err
        product = str(payload.get("product_name") or "").strip()[:400]
        if not product:
            return jsonify({"error": "Cần tên sản phẩm / dịch vụ"}), 400
        try:
            amount = int(payload.get("amount_vnd") or 0)
        except (TypeError, ValueError):
            amount = 0
        try:
            qty = max(1, int(payload.get("quantity") or 1))
        except (TypeError, ValueError):
            qty = 1
        contract_id = None
        if payload.get("contract_id") not in (None, "", 0, "0"):
            try:
                contract_id = int(payload.get("contract_id"))
            except (TypeError, ValueError):
                contract_id = None
        cur = conn.execute(
            """
            INSERT INTO crm_customer_purchases (
                customer_id, order_date, product_name, amount_vnd, quantity, status,
                reference_code, notes, contract_id, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                customer_id,
                str(payload.get("order_date") or ts[:10]).strip()[:32],
                product,
                max(0, amount),
                qty,
                normalize_purchase_status(str(payload.get("status") or "")),
                str(payload.get("reference_code") or "").strip()[:120],
                str(payload.get("notes") or "").strip()[:2000],
                contract_id,
                ts,
                ts,
            ),
        )
        pid = int(cur.lastrowid)
        row = conn.execute(
            "SELECT * FROM crm_customer_purchases WHERE id = ?", (pid,)
        ).fetchone()
    assert row is not None
    return jsonify(_purchase_row(row)), 201


@app.patch("/api/crm/customers/<int:customer_id>/purchases/<int:purchase_id>")
def api_crm_customer_purchase_patch(customer_id: int, purchase_id: int) -> Any:
    deny = _crm_customer_admin_write_only()
    if deny is not None:
        return deny
    payload = request.get_json(force=True) or {}
    ts = _crm_ts()
    with get_connection() as conn:
        err = _crm_customer_access_or_403(conn, customer_id)
        if err is not None:
            return err
        row = conn.execute(
            "SELECT * FROM crm_customer_purchases WHERE id = ? AND customer_id = ?",
            (purchase_id, customer_id),
        ).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy giao dịch"}), 404
        merged = dict(row)
        for key in ("product_name", "order_date", "reference_code", "notes"):
            if key in payload and isinstance(payload[key], str):
                merged[key] = payload[key].strip()[:400 if key == "product_name" else 2000]
        if "status" in payload:
            merged["status"] = normalize_purchase_status(str(payload.get("status") or ""))
        if "amount_vnd" in payload:
            try:
                merged["amount_vnd"] = max(0, int(payload.get("amount_vnd") or 0))
            except (TypeError, ValueError):
                pass
        if "quantity" in payload:
            try:
                merged["quantity"] = max(1, int(payload.get("quantity") or 1))
            except (TypeError, ValueError):
                pass
        conn.execute(
            """
            UPDATE crm_customer_purchases
            SET order_date = ?, product_name = ?, amount_vnd = ?, quantity = ?, status = ?,
                reference_code = ?, notes = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                merged["order_date"],
                merged["product_name"],
                merged["amount_vnd"],
                merged["quantity"],
                merged["status"],
                merged["reference_code"],
                merged["notes"],
                ts,
                purchase_id,
            ),
        )
        out = conn.execute(
            "SELECT * FROM crm_customer_purchases WHERE id = ?", (purchase_id,)
        ).fetchone()
    assert out is not None
    return jsonify(_purchase_row(out))


@app.delete("/api/crm/customers/<int:customer_id>/purchases/<int:purchase_id>")
def api_crm_customer_purchase_delete(customer_id: int, purchase_id: int) -> Any:
    deny = _crm_customer_admin_write_only()
    if deny is not None:
        return deny
    with get_connection() as conn:
        err = _crm_customer_access_or_403(conn, customer_id)
        if err is not None:
            return err
        cur = conn.execute(
            "DELETE FROM crm_customer_purchases WHERE id = ? AND customer_id = ?",
            (purchase_id, customer_id),
        )
        if cur.rowcount == 0:
            return jsonify({"error": "Không tìm thấy giao dịch"}), 404
    return jsonify({"ok": True})


@app.post("/api/crm/customers/<int:customer_id>/issues")
def api_crm_customer_issue_create(customer_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    ts = _crm_ts()
    portal_sid = _crm_effective_staff_id()
    with get_connection() as conn:
        err = _crm_customer_access_or_403(conn, customer_id)
        if err is not None:
            return err
        if portal_sid is None and not _admin_section_can("crm_board_customers", "edit"):
            return _admin_section_forbidden_json("crm_board_customers", "edit")
        title = str(payload.get("title") or "").strip()[:400]
        if not title:
            return jsonify({"error": "Cần tiêu đề vấn đề"}), 400
        assigned_staff_id = portal_sid
        if portal_sid is None and payload.get("assigned_staff_id") not in (None, "", 0, "0"):
            try:
                assigned_staff_id = int(payload.get("assigned_staff_id"))
            except (TypeError, ValueError):
                assigned_staff_id = None
        case_id = None
        if payload.get("case_id") not in (None, "", 0, "0"):
            try:
                case_id = int(payload.get("case_id"))
            except (TypeError, ValueError):
                case_id = None
        cur = conn.execute(
            """
            INSERT INTO crm_customer_issues (
                customer_id, case_id, issue_type, priority, status, title, description,
                resolution, assigned_staff_id, created_at, updated_at, resolved_at
            ) VALUES (?, ?, ?, ?, 'moi', ?, ?, '', ?, ?, ?, '')
            """,
            (
                customer_id,
                case_id,
                normalize_issue_type(str(payload.get("issue_type") or "")),
                normalize_issue_priority(str(payload.get("priority") or "")),
                title,
                str(payload.get("description") or "").strip()[:8000],
                assigned_staff_id,
                ts,
                ts,
            ),
        )
        iid = int(cur.lastrowid)
        row = conn.execute(
            """
            SELECT i.*, st.name AS assigned_staff_name
            FROM crm_customer_issues i
            LEFT JOIN crm_staff st ON st.id = i.assigned_staff_id
            WHERE i.id = ?
            """,
            (iid,),
        ).fetchone()
    assert row is not None
    return jsonify(_issue_row(row)), 201


@app.patch("/api/crm/customers/<int:customer_id>/issues/<int:issue_id>")
def api_crm_customer_issue_patch(customer_id: int, issue_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    ts = _crm_ts()
    portal_sid = _crm_effective_staff_id()
    with get_connection() as conn:
        err = _crm_customer_access_or_403(conn, customer_id)
        if err is not None:
            return err
        row = conn.execute(
            "SELECT * FROM crm_customer_issues WHERE id = ? AND customer_id = ?",
            (issue_id, customer_id),
        ).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy vấn đề"}), 404
        if portal_sid is not None:
            aid = row["assigned_staff_id"]
            try:
                if aid is not None and int(aid) != int(portal_sid):
                    return jsonify({"error": "Không có quyền cập nhật vấn đề này."}), 403
            except (TypeError, ValueError):
                pass
        elif not _admin_section_can("crm_board_customers", "edit"):
            return _admin_section_forbidden_json("crm_board_customers", "edit")
        merged = dict(row)
        for key in ("title", "description", "resolution"):
            if key in payload and isinstance(payload[key], str):
                merged[key] = payload[key].strip()[:8000 if key != "title" else 400]
        if "issue_type" in payload:
            merged["issue_type"] = normalize_issue_type(str(payload.get("issue_type") or ""))
        if "priority" in payload:
            merged["priority"] = normalize_issue_priority(str(payload.get("priority") or ""))
        if "status" in payload:
            merged["status"] = normalize_issue_status(str(payload.get("status") or ""))
        if portal_sid is None and "assigned_staff_id" in payload:
            raw = payload.get("assigned_staff_id")
            if raw in (None, "", 0, "0"):
                merged["assigned_staff_id"] = None
            else:
                try:
                    merged["assigned_staff_id"] = int(raw)
                except (TypeError, ValueError):
                    merged["assigned_staff_id"] = None
        resolved_at = str(merged.get("resolved_at") or "")
        if merged["status"] in ("da_xu_ly", "dong") and not resolved_at:
            resolved_at = ts
        elif merged["status"] not in ("da_xu_ly", "dong"):
            resolved_at = ""
        conn.execute(
            """
            UPDATE crm_customer_issues
            SET issue_type = ?, priority = ?, status = ?, title = ?, description = ?,
                resolution = ?, assigned_staff_id = ?, updated_at = ?, resolved_at = ?
            WHERE id = ?
            """,
            (
                merged["issue_type"],
                merged["priority"],
                merged["status"],
                merged["title"],
                merged["description"],
                merged["resolution"],
                merged["assigned_staff_id"],
                ts,
                resolved_at,
                issue_id,
            ),
        )
        out = conn.execute(
            """
            SELECT i.*, st.name AS assigned_staff_name
            FROM crm_customer_issues i
            LEFT JOIN crm_staff st ON st.id = i.assigned_staff_id
            WHERE i.id = ?
            """,
            (issue_id,),
        ).fetchone()
    assert out is not None
    return jsonify(_issue_row(out))


@app.get("/api/crm/campaigns")
def api_crm_list_campaigns() -> Any:
    raw = (request.args.get("include_inactive") or "").strip().lower()
    incl = raw in ("1", "true", "yes", "all")
    try:
        from ptt_crm.config import hub_read_source_pg
        from ptt_crm.hub_pg_read import list_hub_campaigns

        if hub_read_source_pg():
            campaigns = list_hub_campaigns(active_only=not incl)
            if campaigns:
                try:
                    from ptt_agency.hub_campaign_sync import enrich_campaigns_with_client_codes

                    campaigns = enrich_campaigns_with_client_codes(campaigns)
                except Exception:
                    pass
                return jsonify({"campaigns": campaigns, "read_source": "pg"})
    except Exception:
        pass
    with get_connection() as conn:
        _ensure_crm_hub_schema(conn)
        if incl:
            rows = conn.execute(
                """
                SELECT * FROM crm_campaigns
                ORDER BY active DESC, id DESC
                """
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT * FROM crm_campaigns WHERE active = 1
                ORDER BY name COLLATE NOCASE ASC
                """
            ).fetchall()
    campaigns = rows_to_dict(rows)
    try:
        from ptt_agency.hub_campaign_sync import enrich_campaigns_with_client_codes

        campaigns = enrich_campaigns_with_client_codes(campaigns)
    except Exception:
        pass
    return jsonify({"campaigns": campaigns})


@app.post("/api/crm/campaigns")
def api_crm_create_campaign() -> Any:
    if not _admin_section_can("crm_hub_campaigns", "create"):
        return _admin_section_forbidden_json("crm_hub_campaigns", "create")
    payload = request.get_json(force=True) or {}
    name = str(payload.get("name", "")).strip()[:240]
    if not name:
        return jsonify({"error": "Thiếu tên chiến dịch"}), 400
    code = str(payload.get("code", "")).strip()[:64]
    ch = _crm_normalize_campaign_channel(str(payload.get("channel") or "other"))
    ext = str(payload.get("external_ref", "")).strip()[:240]
    utm = str(payload.get("utm_campaign", "")).strip()[:240]
    notes = str(payload.get("notes", "")).strip()[:4000]
    agency_client_id = str(payload.get("agency_client_id") or "").strip()[:64]
    try:
        target_cpl_vnd = max(0, int(payload.get("target_cpl_vnd") or 0))
    except (TypeError, ValueError):
        target_cpl_vnd = 0
    ts_d = datetime.now().strftime("%Y-%m-%d")
    ts = _crm_ts()
    with get_connection() as conn:
        _ensure_crm_hub_schema(conn)
        if code and conn.execute(
            """
            SELECT 1 FROM crm_campaigns WHERE trim(code) != '' AND lower(trim(code)) = lower(?)
            """,
            (code,),
        ).fetchone():
            return jsonify({"error": "Mã chiến dịch đã tồn tại"}), 400
        cur = conn.execute(
            """
            INSERT INTO crm_campaigns (
                code, name, channel, external_ref, utm_campaign, notes, active,
                agency_client_id, target_cpl_vnd, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, 1, ?, ?, ?, ?)
            """,
            (code, name, ch, ext, utm, notes, agency_client_id, target_cpl_vnd, ts_d, ts),
        )
        mid = int(cur.lastrowid)
        _crm_sync_hub_campaign_to_pg(mid, conn)
        row = conn.execute("SELECT * FROM crm_campaigns WHERE id = ?", (mid,)).fetchone()
    assert row is not None
    out = dict(row)
    try:
        from ptt_crm.hub_pg_write import upsert_hub_campaign_from_sqlite

        upsert_hub_campaign_from_sqlite(out)
    except Exception:
        pass
    try:
        from ptt_agency.hub_campaign_sync import enrich_campaigns_with_client_codes

        enriched = enrich_campaigns_with_client_codes([out])
        out = enriched[0] if enriched else out
    except Exception:
        pass
    return jsonify(out), 201


@app.patch("/api/crm/campaigns/<int:campaign_id>")
def api_crm_patch_campaign(campaign_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        _ensure_crm_hub_schema(conn)
        row = conn.execute("SELECT * FROM crm_campaigns WHERE id = ?", (campaign_id,)).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy chiến dịch"}), 404
        merged: dict[str, Any] = dict(row)
        if "code" in payload and isinstance(payload["code"], str):
            merged["code"] = payload["code"].strip()[:64]
        if "name" in payload and isinstance(payload["name"], str):
            nm = payload["name"].strip()[:240]
            if not nm:
                return jsonify({"error": "Tên không được trống"}), 400
            merged["name"] = nm
        if "channel" in payload:
            merged["channel"] = _crm_normalize_campaign_channel(payload.get("channel"))
        if "external_ref" in payload and isinstance(payload["external_ref"], str):
            merged["external_ref"] = payload["external_ref"].strip()[:240]
        if "utm_campaign" in payload and isinstance(payload["utm_campaign"], str):
            merged["utm_campaign"] = payload["utm_campaign"].strip()[:240]
        if "notes" in payload and isinstance(payload["notes"], str):
            merged["notes"] = payload["notes"].strip()[:4000]
        if "active" in payload:
            merged["active"] = 1 if payload["active"] in (True, 1, "1", "true") else 0
        _crm_apply_campaign_hub_fields(merged, payload)
        cc = str(merged.get("code") or "").strip()
        if cc and conn.execute(
            """
            SELECT 1 FROM crm_campaigns
            WHERE trim(code) != '' AND lower(trim(code)) = lower(?) AND id != ?
            """,
            (cc, campaign_id),
        ).fetchone():
            return jsonify({"error": "Mã chiến dịch đã dùng cho bản khác"}), 400
        ts = _crm_ts()
        conn.execute(
            """
            UPDATE crm_campaigns
            SET code = ?, name = ?, channel = ?, external_ref = ?, utm_campaign = ?,
                notes = ?, active = ?, agency_client_id = ?, target_cpl_vnd = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (
                str(merged.get("code") or ""),
                merged["name"],
                merged["channel"],
                merged["external_ref"],
                merged["utm_campaign"],
                merged["notes"],
                int(merged.get("active") or 0),
                str(merged.get("agency_client_id") or ""),
                int(merged.get("target_cpl_vnd") or 0),
                ts,
                campaign_id,
            ),
        )
        _crm_sync_hub_campaign_to_pg(campaign_id, conn)
        row2 = conn.execute("SELECT * FROM crm_campaigns WHERE id = ?", (campaign_id,)).fetchone()
    assert row2 is not None
    out = dict(row2)
    try:
        from ptt_crm.hub_pg_write import upsert_hub_campaign_from_sqlite

        upsert_hub_campaign_from_sqlite(out)
    except Exception:
        pass
    try:
        from ptt_agency.hub_campaign_sync import enrich_campaigns_with_client_codes

        enriched = enrich_campaigns_with_client_codes([out])
        out = enriched[0] if enriched else out
    except Exception:
        pass
    return jsonify(out)


@app.post("/api/crm/campaigns/<int:campaign_id>/sync-hub-map")
def api_crm_sync_campaign_hub_map(campaign_id: int) -> Any:
    if not _admin_section_can("crm_hub_campaigns", "edit"):
        return _admin_section_forbidden_json("crm_hub_campaigns", "edit")
    with get_connection() as conn:
        _ensure_crm_hub_schema(conn)
        row = conn.execute("SELECT * FROM crm_campaigns WHERE id = ?", (campaign_id,)).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy chiến dịch"}), 404
        from ptt_agency.hub_campaign_sync import sync_campaign_row

        out = sync_campaign_row(dict(row), sqlite_conn=conn)
        conn.commit()
        refreshed = conn.execute("SELECT * FROM crm_campaigns WHERE id = ?", (campaign_id,)).fetchone()
    payload = dict(refreshed) if refreshed else {}
    try:
        from ptt_agency.hub_campaign_sync import enrich_campaigns_with_client_codes

        enriched = enrich_campaigns_with_client_codes([payload])
        payload = enriched[0] if enriched else payload
    except Exception:
        pass
    status = 200 if out.get("ok") or out.get("skipped") else 503
    return jsonify({"sync": out, "campaign": payload}), status


@app.post("/api/crm/campaigns/sync-hub-map-all")
def api_crm_sync_all_campaign_hub_maps() -> Any:
    if not _admin_section_can("crm_hub_campaigns", "edit"):
        return _admin_section_forbidden_json("crm_hub_campaigns", "edit")
    secret = (request.headers.get("X-Cron-Secret") or request.args.get("secret") or "").strip()
    expected = os.environ.get("CRM_AGENCY_HUB_MAP_CRON_SECRET", "").strip()
    if expected and secret != expected:
        return jsonify({"error": "Unauthorized"}), 401
    incl = (request.args.get("include_inactive") or "").strip().lower() in ("1", "true", "yes")
    try:
        from ptt_agency.hub_campaign_sync import sync_all_from_sqlite
        from ptt_jobs.config import sqlite_db_path

        out = sync_all_from_sqlite(sqlite_path=sqlite_db_path(), include_inactive=incl)
        return jsonify(out), 200 if out.get("ok", True) else 503
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.get("/api/crm/contracts")
def api_crm_list_contracts() -> Any:
    cid = _opt_pos_int(request.args.get("customer_id"))
    sid_case = _opt_pos_int(request.args.get("case_id"))
    clauses: list[str] = []
    params: list[Any] = []
    if cid is not None:
        clauses.append("ct.customer_id = ?")
        params.append(cid)
    if sid_case is not None:
        clauses.append("ct.case_id = ?")
        params.append(sid_case)
    where_sql = (" WHERE " + " AND ".join(clauses)) if clauses else ""
    with get_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT ct.*,
                   cu.name AS customer_name, cu.company AS customer_company,
                   cs.title AS case_title,
                   camp.name AS campaign_name, camp.code AS campaign_code,
                   l.full_name AS lead_name
            FROM crm_contracts ct
            JOIN crm_customers cu ON cu.id = ct.customer_id
            LEFT JOIN crm_cases cs ON cs.id = ct.case_id
            LEFT JOIN crm_campaigns camp ON camp.id = ct.campaign_id
            LEFT JOIN crm_leads l ON l.id = ct.lead_id
            {where_sql}
            ORDER BY datetime(ct.updated_at) DESC, ct.id DESC
            """,
            params,
        ).fetchall()
    return jsonify({"contracts": rows_to_dict(rows)})


@app.post("/api/crm/contracts")
def api_crm_create_contract() -> Any:
    payload = request.get_json(force=True) or {}
    lead_id = _opt_pos_int(payload.get("lead_id"))
    cust_id = _opt_pos_int(payload.get("customer_id"))
    title = str(payload.get("title", "")).strip()[:500]
    case_id = _opt_pos_int(payload.get("case_id"))
    camp_id = _opt_pos_int(payload.get("campaign_id"))
    ref = str(payload.get("reference_code", "")).strip()[:120]
    status = _crm_normalize_contract_status(str(payload.get("status") or "draft"))
    signed_on = str(payload.get("signed_on") or "").strip()[:32]
    starts_on = str(payload.get("starts_on") or "").strip()[:32]
    ends_on = str(payload.get("ends_on") or "").strip()[:32]
    for dlab, dval in (("signed_on", signed_on), ("starts_on", starts_on), ("ends_on", ends_on)):
        if dval and not _crm_validate_date_ymd(dval):
            return jsonify({"error": f"{dlab} phải YYYY-MM-DD"}), 400
    try:
        amount_vnd = int(payload.get("amount_vnd") or 0)
    except (TypeError, ValueError):
        amount_vnd = 0
    amount_vnd = max(0, min(amount_vnd, 9_999_999_999_999))
    try:
        rdays = int(payload.get("renewal_reminder_days") or 30)
    except (TypeError, ValueError):
        rdays = 30
    rdays = max(0, min(366, rdays))
    notes = str(payload.get("notes", "")).strip()[:8000]
    from crm_svc_finance import infer_billing_type_from_service_slug, normalize_billing_type
    from crm_svc_finance import normalize_billing_cycle

    billing_type = normalize_billing_type(payload.get("billing_type"))
    billing_cycle = normalize_billing_cycle(payload.get("billing_cycle"))
    service_slug_for_billing = str(payload.get("service_slug") or "").strip()
    if not payload.get("billing_type") and service_slug_for_billing:
        billing_type = infer_billing_type_from_service_slug(service_slug_for_billing)
    ts_d = datetime.now().strftime("%Y-%m-%d")
    ts = _crm_ts()

    if lead_id and _crm_presales_on_lead_enabled():
        if status != "draft":
            return jsonify({"error": "HĐ từ Lead pre-sales chỉ tạo ở trạng thái draft"}), 400
        with get_connection() as conn:
            from crm_lead_presales_contract import (
                PresalesContractError,
                create_draft_contract_from_lead,
            )

            try:
                ct = create_draft_contract_from_lead(
                    conn,
                    lead_id,
                    title=title or None,
                    amount_vnd=amount_vnd,
                    notes=notes,
                    actor=_crm_audit_user(),
                    ts=ts,
                )
            except PresalesContractError as exc:
                return jsonify({"error": str(exc)}), 400
            nid = int(ct["id"])
            _crm_hub_sync_contract_renewal_reminder(conn, nid)
            row = conn.execute(
                """
                SELECT ct.*,
                       cu.name AS customer_name, cu.company AS customer_company,
                       cs.title AS case_title,
                       camp.name AS campaign_name, camp.code AS campaign_code,
                       l.full_name AS lead_name
                FROM crm_contracts ct
                JOIN crm_customers cu ON cu.id = ct.customer_id
                LEFT JOIN crm_cases cs ON cs.id = ct.case_id
                LEFT JOIN crm_campaigns camp ON camp.id = ct.campaign_id
                LEFT JOIN crm_leads l ON l.id = ct.lead_id
                WHERE ct.id = ?
                """,
                (nid,),
            ).fetchone()
        assert row is not None
        return jsonify(dict(row)), 201

    if cust_id is None:
        return jsonify({"error": "Cần customer_id hoặc lead_id (pre-sales)"}), 400
    if not title:
        return jsonify({"error": "Thiếu tiêu đề hợp đồng"}), 400
    with get_connection() as conn:
        if conn.execute("SELECT id FROM crm_customers WHERE id = ?", (cust_id,)).fetchone() is None:
            return jsonify({"error": "Khách hàng không tồn tại"}), 404
        if case_id and conn.execute("SELECT id FROM crm_cases WHERE id = ?", (case_id,)).fetchone() is None:
            return jsonify({"error": "Hồ sơ CRM không tồn tại"}), 404
        if camp_id and conn.execute("SELECT id FROM crm_campaigns WHERE id = ?", (camp_id,)).fetchone() is None:
            return jsonify({"error": "Chiến dịch không tồn tại"}), 404
        cur = conn.execute(
            """
            INSERT INTO crm_contracts (
                customer_id, case_id, campaign_id, reference_code, title, status,
                signed_on, starts_on, ends_on, amount_vnd, renewal_reminder_days, notes,
                billing_type, billing_cycle, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                cust_id,
                case_id,
                camp_id,
                ref,
                title,
                status,
                signed_on,
                starts_on,
                ends_on,
                amount_vnd,
                rdays,
                notes,
                billing_type,
                billing_cycle,
                ts_d,
                ts,
            ),
        )
        nid = int(cur.lastrowid)
        _crm_hub_sync_contract_renewal_reminder(conn, nid)
        row = conn.execute(
            """
            SELECT ct.*,
                   cu.name AS customer_name, cu.company AS customer_company,
                   cs.title AS case_title,
                   camp.name AS campaign_name, camp.code AS campaign_code,
                   l.full_name AS lead_name
            FROM crm_contracts ct
            JOIN crm_customers cu ON cu.id = ct.customer_id
            LEFT JOIN crm_cases cs ON cs.id = ct.case_id
            LEFT JOIN crm_campaigns camp ON camp.id = ct.campaign_id
            LEFT JOIN crm_leads l ON l.id = ct.lead_id
            WHERE ct.id = ?
            """,
            (nid,),
        ).fetchone()
    assert row is not None
    return jsonify(dict(row)), 201


@app.patch("/api/crm/contracts/<int:contract_id>")
def api_crm_patch_contract(contract_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM crm_contracts WHERE id = ?", (contract_id,)).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy hợp đồng"}), 404
        old_status = str(row["status"] or "")
        merged = dict(row)
        if "title" in payload and isinstance(payload["title"], str):
            tl = payload["title"].strip()[:500]
            if not tl:
                return jsonify({"error": "Tiêu đề không được trống"}), 400
            merged["title"] = tl
        if "reference_code" in payload and isinstance(payload["reference_code"], str):
            merged["reference_code"] = payload["reference_code"].strip()[:120]
        if "case_id" in payload:
            raw_c = payload.get("case_id")
            if raw_c in (None, "", 0, "0"):
                merged["case_id"] = None
            else:
                ex = _opt_pos_int(raw_c)
                if ex and conn.execute("SELECT id FROM crm_cases WHERE id = ?", (ex,)).fetchone() is None:
                    return jsonify({"error": "Hồ sơ không tồn tại"}), 404
                merged["case_id"] = ex
        if "campaign_id" in payload:
            raw_p = payload.get("campaign_id")
            if raw_p in (None, "", 0, "0"):
                merged["campaign_id"] = None
            else:
                ex = _opt_pos_int(raw_p)
                if ex and conn.execute("SELECT id FROM crm_campaigns WHERE id = ?", (ex,)).fetchone() is None:
                    return jsonify({"error": "Chiến dịch không tồn tại"}), 404
                merged["campaign_id"] = ex
        if "status" in payload:
            merged["status"] = _crm_normalize_contract_status(str(payload.get("status")))
        for fld in ("signed_on", "starts_on", "ends_on"):
            if fld in payload and isinstance(payload[fld], str):
                v = payload[fld].strip()[:32]
                if v and not _crm_validate_date_ymd(v):
                    return jsonify({"error": f"{fld} phải YYYY-MM-DD"}), 400
                merged[fld] = v
        if "amount_vnd" in payload:
            try:
                merged["amount_vnd"] = max(0, min(int(payload.get("amount_vnd") or 0), 9_999_999_999_999))
            except (TypeError, ValueError):
                pass
        if "renewal_reminder_days" in payload:
            try:
                rd = int(payload.get("renewal_reminder_days") or 0)
                merged["renewal_reminder_days"] = max(0, min(366, rd))
            except (TypeError, ValueError):
                pass
        if "notes" in payload and isinstance(payload["notes"], str):
            merged["notes"] = payload["notes"].strip()[:8000]
        if "billing_type" in payload:
            from crm_svc_finance import normalize_billing_type

            merged["billing_type"] = normalize_billing_type(payload.get("billing_type"))
        if "billing_cycle" in payload:
            from crm_svc_finance import normalize_billing_cycle

            merged["billing_cycle"] = normalize_billing_cycle(payload.get("billing_cycle"))
        ts = _crm_ts()
        lead_id_val = merged.get("lead_id")
        signing = merged.get("status") == "active" and old_status != "active"
        presales_sign = (
            signing
            and _crm_presales_on_lead_enabled()
            and lead_id_val
        )
        sign_result: dict[str, Any] | None = None
        if presales_sign:
            from crm_lead_presales_contract import (
                PresalesContractError,
                on_presales_contract_signed,
            )

            try:
                sign_result = on_presales_contract_signed(
                    conn,
                    contract_id,
                    actor=_crm_audit_user(),
                    ts=ts,
                )
            except PresalesContractError as exc:
                return jsonify({"error": str(exc)}), 400
            except Exception as exc:
                logger.warning(
                    "on_presales_contract_signed lỗi contract=%s: %s",
                    contract_id,
                    exc,
                )
                return jsonify({"error": "Không kích hoạt lifecycle từ pre-sales"}), 500
            merged["customer_id"] = sign_result["customer_id"]
            if sign_result.get("case_id"):
                merged["case_id"] = sign_result["case_id"]
        conn.execute(
            """
            UPDATE crm_contracts
            SET customer_id = ?, case_id = ?, campaign_id = ?, reference_code = ?, title = ?,
                status = ?, signed_on = ?, starts_on = ?, ends_on = ?, amount_vnd = ?,
                renewal_reminder_days = ?, notes = ?, billing_type = ?, billing_cycle = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (
                merged["customer_id"],
                merged.get("case_id"),
                merged.get("campaign_id"),
                merged["reference_code"],
                merged["title"],
                merged["status"],
                merged.get("signed_on") or "",
                merged.get("starts_on") or "",
                merged.get("ends_on") or "",
                int(merged.get("amount_vnd") or 0),
                int(merged.get("renewal_reminder_days") or 0),
                merged.get("notes") or "",
                str(merged.get("billing_type") or "one_off"),
                str(merged.get("billing_cycle") or "monthly"),
                ts,
                contract_id,
            ),
        )
        _crm_hub_sync_contract_renewal_reminder(conn, contract_id)
        if signing and not presales_sign:
            try:
                activate_lifecycle(conn, contract_id)
            except Exception as _lc_exc:
                logger.warning("activate_lifecycle lỗi contract=%s: %s", contract_id, _lc_exc)
        row2 = conn.execute(
            """
            SELECT ct.*,
                   cu.name AS customer_name, cu.company AS customer_company,
                   cs.title AS case_title,
                   camp.name AS campaign_name, camp.code AS campaign_code,
                   l.full_name AS lead_name
            FROM crm_contracts ct
            JOIN crm_customers cu ON cu.id = ct.customer_id
            LEFT JOIN crm_cases cs ON cs.id = ct.case_id
            LEFT JOIN crm_campaigns camp ON camp.id = ct.campaign_id
            LEFT JOIN crm_leads l ON l.id = ct.lead_id
            WHERE ct.id = ?
            """,
            (contract_id,),
        ).fetchone()
    assert row2 is not None
    return jsonify(dict(row2))


@app.delete("/api/crm/contracts/<int:contract_id>")
def api_crm_delete_contract(contract_id: int) -> Any:
    with get_connection() as conn:
        cur = conn.execute("DELETE FROM crm_contracts WHERE id = ?", (contract_id,))
        if cur.rowcount == 0:
            return jsonify({"error": "Không tìm thấy hợp đồng"}), 404
        conn.execute(
            "DELETE FROM crm_reminders WHERE scope = 'contract' AND ref_id = ?",
            (contract_id,),
        )
    return jsonify({"ok": True})


@app.get("/api/crm/reminders")
def api_crm_list_reminders() -> Any:
    st = (request.args.get("status") or "pending").strip().lower()
    if st not in CRM_REMINDER_STATUSES and st != "all":
        st = "pending"
    scope = str(request.args.get("scope") or "").strip().lower()
    clauses: list[str] = []
    params: list[Any] = []
    if st != "all":
        clauses.append("r.status = ?")
        params.append(st)
    if scope and scope in CRM_REMINDER_SCOPES:
        clauses.append("r.scope = ?")
        params.append(scope)
    where_sql = (" WHERE " + " AND ".join(clauses)) if clauses else ""
    with get_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT r.* FROM crm_reminders r
            {where_sql}
            ORDER BY r.remind_at ASC, r.id ASC
            LIMIT 500
            """,
            params,
        ).fetchall()
    return jsonify({"reminders": rows_to_dict(rows)})


@app.get("/api/crm/reminders/summary")
def api_crm_reminders_summary() -> Any:
    today = datetime.now().strftime("%Y-%m-%d")
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT * FROM crm_reminders
            WHERE status = 'pending'
            ORDER BY remind_at ASC
            LIMIT 500
            """
        ).fetchall()
    overdue: list[dict[str, Any]] = []
    today_l: list[dict[str, Any]] = []
    upcoming: list[dict[str, Any]] = []
    for r in rows:
        d = dict(r)
        ra = str(d.get("remind_at") or "").strip()[:10]
        if len(ra) == 10 and _crm_validate_date_ymd(ra):
            if ra < today:
                overdue.append(d)
            elif ra == today:
                today_l.append(d)
            else:
                upcoming.append(d)
        else:
            upcoming.append(d)
    return jsonify(
        {
            "today_iso": today,
            "overdue": overdue,
            "today": today_l,
            "upcoming": upcoming,
            "counts": {
                "overdue": len(overdue),
                "today": len(today_l),
                "upcoming": len(upcoming),
            },
        }
    )


@app.post("/api/crm/reminders")
def api_crm_create_reminder() -> Any:
    payload = request.get_json(force=True) or {}
    title = str(payload.get("title", "")).strip()[:500]
    if not title:
        return jsonify({"error": "Thiếu tiêu đề nhắc nhở"}), 400
    scope = _crm_normalize_reminder_scope(payload.get("scope"))
    kind = _crm_normalize_reminder_kind(payload.get("reminder_kind"))
    remind_at_raw = str(payload.get("remind_at") or "").strip()[:32]
    if len(remind_at_raw) >= 10:
        rd = remind_at_raw[:10]
        if not _crm_validate_date_ymd(rd):
            return jsonify({"error": "remind_at cần YYYY-MM-DD"}), 400
        remind_at = rd
    else:
        return jsonify({"error": "Thiếu remind_at (YYYY-MM-DD)"}), 400
    try:
        ref_id = int(payload.get("ref_id") or 0)
    except (TypeError, ValueError):
        ref_id = 0
    body = str(payload.get("body", "")).strip()[:8000]
    staff_id = _opt_pos_int(payload.get("staff_id"))
    ts_d = datetime.now().strftime("%Y-%m-%d")
    ts = _crm_ts()
    with get_connection() as conn:
        if staff_id and conn.execute("SELECT id FROM crm_staff WHERE id = ?", (staff_id,)).fetchone() is None:
            return jsonify({"error": "Nhân viên không tồn tại"}), 400
        cur = conn.execute(
            """
            INSERT INTO crm_reminders (
                scope, ref_id, reminder_kind, title, body, remind_at, status, staff_id, meta_json,
                created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, 'pending', ?, '{}', ?, ?)
            """,
            (scope, ref_id, kind, title, body, remind_at, staff_id, ts_d, ts),
        )
        rid = int(cur.lastrowid)
        row = conn.execute("SELECT * FROM crm_reminders WHERE id = ?", (rid,)).fetchone()
    assert row is not None
    return jsonify(dict(row)), 201


@app.patch("/api/crm/reminders/<int:reminder_id>")
def api_crm_patch_reminder(reminder_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM crm_reminders WHERE id = ?", (reminder_id,)).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy nhắc nhở"}), 404
        merged = dict(row)
        if "title" in payload and isinstance(payload["title"], str):
            merged["title"] = payload["title"].strip()[:500]
        if "body" in payload and isinstance(payload["body"], str):
            merged["body"] = payload["body"].strip()[:8000]
        if "remind_at" in payload and isinstance(payload["remind_at"], str):
            rd = payload["remind_at"].strip()[:10]
            if rd and not _crm_validate_date_ymd(rd):
                return jsonify({"error": "remind_at không hợp lệ"}), 400
            merged["remind_at"] = rd
        if "status" in payload:
            merged["status"] = _crm_normalize_reminder_status(str(payload.get("status")))
        if "staff_id" in payload:
            ex = _opt_pos_int(payload.get("staff_id"))
            if ex and conn.execute("SELECT id FROM crm_staff WHERE id = ?", (ex,)).fetchone() is None:
                return jsonify({"error": "Nhân viên không tồn tại"}), 400
            merged["staff_id"] = ex
        if "scope" in payload:
            merged["scope"] = _crm_normalize_reminder_scope(payload.get("scope"))
        if "reminder_kind" in payload:
            merged["reminder_kind"] = _crm_normalize_reminder_kind(payload.get("reminder_kind"))
        if "ref_id" in payload:
            try:
                merged["ref_id"] = max(0, int(payload.get("ref_id") or 0))
            except (TypeError, ValueError):
                return jsonify({"error": "ref_id không hợp lệ"}), 400
        ts = _crm_ts()
        conn.execute(
            """
            UPDATE crm_reminders
            SET title = ?, body = ?, remind_at = ?, status = ?, staff_id = ?,
                scope = ?, ref_id = ?, reminder_kind = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                merged["title"],
                merged["body"],
                merged["remind_at"],
                merged["status"],
                merged.get("staff_id"),
                merged["scope"],
                int(merged.get("ref_id") or 0),
                merged["reminder_kind"],
                ts,
                reminder_id,
            ),
        )
        row2 = conn.execute("SELECT * FROM crm_reminders WHERE id = ?", (reminder_id,)).fetchone()
    assert row2 is not None
    return jsonify(dict(row2))


def _mp_normalize_plan_status(raw: str | None) -> str:
    s = str(raw or "draft").strip().lower()
    return s if s in CRM_MARKETING_PLAN_STATUSES else "draft"


def _mp_normalize_plan_priority(raw: str | None) -> str:
    s = str(raw or "normal").strip().lower()
    return s if s in CRM_MARKETING_PLAN_PRIORITIES else "normal"


def _mp_normalize_milestone_status(raw: str | None) -> str:
    s = str(raw or "pending").strip().lower()
    return s if s in CRM_MARKETING_PLAN_MS_STATUSES else "pending"


def _mp_json_array_str(payload: dict[str, Any], key: str, *, max_bytes: int = 12000) -> str:
    raw = payload.get(key)
    if raw is None:
        return "[]"
    try:
        if isinstance(raw, str):
            parsed = json.loads(raw)
            if not isinstance(parsed, list):
                return "[]"
            out = parsed
        elif isinstance(raw, list):
            out = raw
        else:
            return "[]"
        s = json.dumps(out, ensure_ascii=False)
        return s[:max_bytes] if len(s) <= max_bytes else json.dumps([], ensure_ascii=False)
    except (json.JSONDecodeError, TypeError, ValueError):
        return "[]"


def _mp_meta_json(payload: dict[str, Any]) -> tuple[str, str, str]:
    """pillars_json, channels_focus_json, success_metrics_json — truncated if needed."""
    return (
        _mp_json_array_str(payload, "pillars_json", max_bytes=16000),
        _mp_channels_focus_json(payload.get("channels_focus")),
        _mp_json_array_str(payload, "success_metrics_json", max_bytes=16000),
    )


def _mp_channels_focus_json(raw: Any) -> str:
    out: list[str] = []
    if isinstance(raw, list):
        for x in raw:
            sx = _crm_normalize_campaign_channel(str(x))
            if sx not in out:
                out.append(sx)
    elif isinstance(raw, str) and raw.strip():
        try:
            arr = json.loads(raw)
            if isinstance(arr, list):
                raw = arr
            else:
                return "[]"
        except json.JSONDecodeError:
            return "[]"
        for x in raw:
            sx = _crm_normalize_campaign_channel(str(x))
            if sx not in out:
                out.append(sx)
    else:
        return "[]"
    return json.dumps(out[:32], ensure_ascii=False)


def _mp_strategy_framework_json(raw: Any) -> str:
    """Khung 9 nghiệp vụ marketing — object JSON với key cố định."""
    default: dict[str, str] = {k: "" for k in CRM_MP_STRATEGY_FRAMEWORK_KEYS}
    obj: dict[str, Any] = {}
    if raw is None:
        return json.dumps(default, ensure_ascii=False)
    if isinstance(raw, str) and raw.strip():
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                obj = parsed
            else:
                return json.dumps(default, ensure_ascii=False)
        except (json.JSONDecodeError, TypeError, ValueError):
            return json.dumps(default, ensure_ascii=False)
    elif isinstance(raw, dict):
        obj = raw
    else:
        return json.dumps(default, ensure_ascii=False)
    out = default.copy()
    for k in CRM_MP_STRATEGY_FRAMEWORK_KEYS:
        v = obj.get(k)
        if isinstance(v, str):
            out[k] = v.strip()[:12000]
        elif v is not None:
            out[k] = str(v).strip()[:12000]
    try:
        s = json.dumps(out, ensure_ascii=False)
    except (TypeError, ValueError):
        return json.dumps(default, ensure_ascii=False)
    max_bytes = 62000
    if len(s) <= max_bytes:
        return s
    for kk in CRM_MP_STRATEGY_FRAMEWORK_KEYS:
        if len(out[kk]) > 5000:
            out[kk] = out[kk][:5000]
    try:
        return json.dumps(out, ensure_ascii=False)[:max_bytes]
    except (TypeError, ValueError):
        return json.dumps(default, ensure_ascii=False)


def _mp_target_market_prof_json(raw: Any) -> str:
    """Chương trình xác định TMMT (12 trụ) — object JSON."""
    default: dict[str, str] = {k: "" for k in CRM_MP_TARGET_MARKET_PROF_KEYS}
    obj: dict[str, Any] = {}
    if raw is None:
        return json.dumps(default, ensure_ascii=False)
    if isinstance(raw, str) and raw.strip():
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                obj = parsed
            else:
                return json.dumps(default, ensure_ascii=False)
        except (json.JSONDecodeError, TypeError, ValueError):
            return json.dumps(default, ensure_ascii=False)
    elif isinstance(raw, dict):
        obj = raw
    else:
        return json.dumps(default, ensure_ascii=False)
    out = default.copy()
    for k in CRM_MP_TARGET_MARKET_PROF_KEYS:
        v = obj.get(k)
        if isinstance(v, str):
            out[k] = v.strip()[:8000]
        elif v is not None:
            out[k] = str(v).strip()[:8000]
    try:
        s = json.dumps(out, ensure_ascii=False)
    except (TypeError, ValueError):
        return json.dumps(default, ensure_ascii=False)
    max_bytes = 98000
    if len(s) <= max_bytes:
        return s
    for kk in CRM_MP_TARGET_MARKET_PROF_KEYS:
        if len(out[kk]) > 6000:
            out[kk] = out[kk][:6000]
    try:
        return json.dumps(out, ensure_ascii=False)[:max_bytes]
    except (TypeError, ValueError):
        return json.dumps(default, ensure_ascii=False)


def _mp_target_market_steps4_json(raw: Any) -> str:
    """TMMT — quy trình 4 bước (14 trường) — JSON object."""
    default: dict[str, str] = {k: "" for k in CRM_MP_TARGET_MARKET_STEPS4_KEYS}
    obj: dict[str, Any] = {}
    if raw is None:
        return json.dumps(default, ensure_ascii=False)
    if isinstance(raw, str) and raw.strip():
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                obj = parsed
            else:
                return json.dumps(default, ensure_ascii=False)
        except (json.JSONDecodeError, TypeError, ValueError):
            return json.dumps(default, ensure_ascii=False)
    elif isinstance(raw, dict):
        obj = raw
    else:
        return json.dumps(default, ensure_ascii=False)
    out = default.copy()
    for k in CRM_MP_TARGET_MARKET_STEPS4_KEYS:
        v = obj.get(k)
        if isinstance(v, str):
            out[k] = v.strip()[:10000]
        elif v is not None:
            out[k] = str(v).strip()[:10000]
    try:
        s = json.dumps(out, ensure_ascii=False)
    except (TypeError, ValueError):
        return json.dumps(default, ensure_ascii=False)
    max_bytes = 120000
    if len(s) <= max_bytes:
        return s
    for kk in CRM_MP_TARGET_MARKET_STEPS4_KEYS:
        if len(out[kk]) > 6500:
            out[kk] = out[kk][:6500]
    try:
        return json.dumps(out, ensure_ascii=False)[:max_bytes]
    except (TypeError, ValueError):
        return json.dumps(default, ensure_ascii=False)


CRM_KHTN_MARKET_RESEARCH_KEYS: tuple[str, ...] = (
    "market_size",
    "market_trends",
    "competitors",
    "segment_demographics_age",
    "segment_demographics_gender",
    "segment_demographics_income",
    "segment_demographics_education",
    "segment_demographics_occupation",
    "segment_demographics",
    "segment_geography_location",
    "segment_geography_climate",
    "segment_geography_density",
    "segment_geography",
    "segment_psychographics_lifestyle",
    "segment_psychographics_interests",
    "segment_psychographics_concerns",
    "segment_psychographics_values",
    "segment_psychographics_opinions",
    "segment_psychographics",
    "segment_consumption_frequency",
    "segment_consumption_channels",
    "segment_consumption_quantity",
    "segment_consumption_decision_process",
    "segment_consumption_behavior",
    "icp_current_customer_analysis",
    "icp_prospect_market_research",
    "icp_competitive_customer_analysis",
    "icp_data_analytics_tools",
    "approach_multi_segment_marketing",
    "approach_concentrated_marketing",
    "approach_micro_targeting",
    "approach_product_specialization",
    "pilot_focus_hypothesis_validation",
    "pilot_focus_cost_efficiency",
    "pilot_focus_roas",
    "pilot_focus_segment_discovery",
    "pilot_prog_channels_platforms",
    "pilot_prog_methodology",
    "pilot_prog_variants_matrix",
    "pilot_prog_budget_schedule",
    "pilot_prog_tracking_kpis",
    "pilot_prog_results_decisions",
    "pilot_practice_tools",
    "pilot_practice_methods",
    "pilot_practice_statistics",
    "pilot_analysis_quant_significance",
    "metric_ctr_target",
    "metric_ctr_ideal",
    "metric_ctr_meaning",
    "metric_ctr_actual",
    "metric_ctr_note",
    "metric_cpc_target",
    "metric_cpc_ideal",
    "metric_cpc_meaning",
    "metric_cpc_actual",
    "metric_cpc_note",
    "metric_cpl_target",
    "metric_cpl_ideal",
    "metric_cpl_meaning",
    "metric_cpl_actual",
    "metric_cpl_note",
    "metric_cvr_target",
    "metric_cvr_ideal",
    "metric_cvr_meaning",
    "metric_cvr_actual",
    "metric_cvr_note",
    "metric_roas_target",
    "metric_roas_ideal",
    "metric_roas_meaning",
    "metric_roas_actual",
    "metric_roas_note",
    "metric_ql_target",
    "metric_ql_ideal",
    "metric_ql_meaning",
    "metric_ql_actual",
    "metric_ql_note",
    "pilot_segment_groups_json",
    "pilot_analysis_segment_breakdown",
    "breakdown_win_text",
    "breakdown_win_budget_pct",
    "breakdown_tie_text",
    "breakdown_tie_budget_pct",
    "breakdown_lose_text",
    "breakdown_lose_budget_pct",
    "surprise_insights_json",
    "next_test_plan",
    "forecast_3months",
    "final_target_market_confirmation",
    "pilot_analysis_creative_funnel",
    "pilot_analysis_qualitative_sentiment",
    "pilot_analysis_comparison",
    "pilot_analysis_learnings_risks",
    "pilot_analysis_decision_framework",
    "pilot_experiment_notes",
)

# Cờ checkbox bước 4 & 5 — chỉ nhận "1" hoặc "" sau chuẩn hóa.
CRM_KHTN_CHECKBOX_FLAG_KEYS: frozenset[str] = frozenset(
    (
        "approach_multi_segment_marketing",
        "approach_concentrated_marketing",
        "approach_micro_targeting",
        "approach_product_specialization",
        "pilot_focus_hypothesis_validation",
        "pilot_focus_cost_efficiency",
        "pilot_focus_roas",
        "pilot_focus_segment_discovery",
    )
)


def _mp_khtn_market_research_json(raw: Any) -> str:
    """KHTN — bước 1–5 (`khtn_market_research_json`): thị trường đến thử nghiệm quảng cáo."""
    default = {k: "" for k in CRM_KHTN_MARKET_RESEARCH_KEYS}
    obj: dict[str, Any] = {}
    if raw is None:
        return json.dumps(default, ensure_ascii=False)
    if isinstance(raw, str) and raw.strip():
        try:
            parsed = json.loads(raw)
            obj = parsed if isinstance(parsed, dict) else {}
        except (json.JSONDecodeError, TypeError, ValueError):
            obj = {}
    elif isinstance(raw, dict):
        obj = raw
    else:
        return json.dumps(default, ensure_ascii=False)
    out = default.copy()
    for k in CRM_KHTN_MARKET_RESEARCH_KEYS:
        v = obj.get(k)
        if isinstance(v, str):
            out[k] = v.strip()[:24000]
        elif v is not None:
            out[k] = str(v).strip()[:24000]
    for fk in CRM_KHTN_CHECKBOX_FLAG_KEYS:
        if fk not in out:
            continue
        sv = (out[fk] or "").strip().lower()
        out[fk] = "1" if sv in ("1", "true", "yes", "on") else ""
    try:
        s = json.dumps(out, ensure_ascii=False)
        if len(s) <= 180000:
            return s
        for kk in CRM_KHTN_MARKET_RESEARCH_KEYS:
            if len(out[kk]) > 32000:
                out[kk] = out[kk][:32000]
        return json.dumps(out, ensure_ascii=False)[:180000]
    except (TypeError, ValueError):
        return json.dumps(default, ensure_ascii=False)


@app.get("/crm/marketing-plan")
def crm_marketing_plan_page() -> str:
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    return render_template(
        "crm_marketing_plan.html",
        plan_status_labels=CRM_MARKETING_PLAN_STATUS_LABELS_VI,
        plan_statuses=list(CRM_MARKETING_PLAN_STATUSES),
        plan_priority_labels=CRM_MARKETING_PLAN_PRIORITY_LABELS_VI,
        plan_priorities=list(CRM_MARKETING_PLAN_PRIORITIES),
        ms_status_labels=CRM_MARKETING_PLAN_MS_STATUS_LABELS_VI,
        ms_statuses=list(CRM_MARKETING_PLAN_MS_STATUSES),
        campaign_channel_labels=CRM_CAMPAIGN_CHANNEL_LABELS_VI,
        campaign_channels=list(CRM_CAMPAIGN_CHANNELS),
        strategy_framework_labels=CRM_MP_STRATEGY_FRAMEWORK_LABELS_VI,
        strategy_framework_keys=list(CRM_MP_STRATEGY_FRAMEWORK_KEYS),
        strategy_framework_other_keys=list(CRM_MP_STRATEGY_FRAMEWORK_KEYS_WITHOUT_TMMT),
        target_market_prof_labels=CRM_MP_TARGET_MARKET_PROF_LABELS_VI,
        target_market_prof_keys=list(CRM_MP_TARGET_MARKET_PROF_KEYS),
        target_market_steps4_sections=list(CRM_MP_TARGET_MARKET_STEPS4_SECTIONS),
        target_market_steps4_labels=CRM_MP_TARGET_MARKET_STEPS4_LABELS_VI,
        target_market_steps4_keys=list(CRM_MP_TARGET_MARKET_STEPS4_KEYS),
        marketing_segment_cards=[
            {"slug": s, **MARKETING_LIFECYCLE_PAGES[s]}
            for s in MARKETING_LIFECYCLE_SLUGS
        ],
        **_admin_page_template_kwargs(),
    )


@app.get("/crm/marketing-plan/segment/<slug>")
def crm_marketing_plan_segment_page(slug: str) -> str:
    """Tham chiếu nghiệp vụ KHTN / KHQT / CSKH (cùng khối Kế hoạch marketing CRM)."""
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    key = (slug or "").strip().lower()
    page = MARKETING_LIFECYCLE_PAGES.get(key)
    if not page:
        abort(404)
    others = [
        {"slug": s, **MARKETING_LIFECYCLE_PAGES[s]}
        for s in MARKETING_LIFECYCLE_SLUGS
        if s != key
    ]
    return render_template(
        "crm_marketing_segment.html",
        page_key=key,
        page=page,
        sibling_pages=others,
        khtn_pipeline_steps=KHTN_PIPELINE_STEPS if key == "khtn" else [],
    )


@app.get("/api/crm/marketing-plans")
def api_crm_mp_list() -> Any:
    fy_raw = (request.args.get("fiscal_year") or "").strip()
    fy: int | None = None
    if fy_raw:
        try:
            fy = max(1990, min(2120, int(fy_raw)))
        except ValueError:
            fy = None
    st_raw = (request.args.get("status") or "").strip().lower()
    st = st_raw if st_raw in CRM_MARKETING_PLAN_STATUSES or st_raw == "all" else "all"
    q_raw = (request.args.get("q") or "").strip().lower()

    clauses: list[str] = []
    params: list[Any] = []
    if fy is not None:
        clauses.append("p.fiscal_year = ?")
        params.append(fy)
    if st != "all":
        clauses.append("p.status = ?")
        params.append(st)
    if q_raw:
        clauses.append("(lower(p.name) LIKE ? OR lower(p.code) LIKE ? OR lower(p.period_label) LIKE ?)")
        like = f"%{q_raw}%"
        params.extend((like, like, like))

    where_sql = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    with get_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT p.*,
                   st.name AS owner_name,
                   (SELECT COUNT(*) FROM crm_marketing_plan_campaigns mpc WHERE mpc.plan_id = p.id)
                     AS linked_campaign_count,
                   (SELECT COUNT(*) FROM crm_marketing_plan_milestones mm WHERE mm.plan_id = p.id)
                     AS milestone_total,
                   (SELECT COUNT(*) FROM crm_marketing_plan_milestones mm
                    WHERE mm.plan_id = p.id AND mm.status = 'done') AS milestone_done
            FROM crm_marketing_plans p
            LEFT JOIN crm_staff st ON st.id = p.owner_staff_id
            {where_sql}
            ORDER BY p.fiscal_year DESC, datetime(p.updated_at) DESC, p.id DESC
            LIMIT 300
            """,
            params,
        ).fetchall()
    return jsonify({"plans": rows_to_dict(rows)})


@app.get("/api/crm/marketing-plans/<int:plan_id>")
def api_crm_mp_get(plan_id: int) -> Any:
    with get_connection() as conn:
        plan = conn.execute(
            """
            SELECT p.*, st.name AS owner_name
            FROM crm_marketing_plans p
            LEFT JOIN crm_staff st ON st.id = p.owner_staff_id
            WHERE p.id = ?
            """,
            (plan_id,),
        ).fetchone()
        if plan is None:
            return jsonify({"error": "Không tìm thấy kế hoạch"}), 404
        milestones = conn.execute(
            """
            SELECT * FROM crm_marketing_plan_milestones
            WHERE plan_id = ?
            ORDER BY position ASC, id ASC
            """,
            (plan_id,),
        ).fetchall()
        c_rows = conn.execute(
            """
            SELECT c.*
            FROM crm_marketing_plan_campaigns l
            JOIN crm_campaigns c ON c.id = l.campaign_id
            WHERE l.plan_id = ?
            ORDER BY c.name COLLATE NOCASE ASC
            """,
            (plan_id,),
        ).fetchall()
        data = dict(plan)
        data["milestones"] = rows_to_dict(milestones)
        data["campaigns"] = rows_to_dict(c_rows)
        return jsonify(data)


@app.post("/api/crm/marketing-plans")
def api_crm_mp_create() -> Any:
    payload = request.get_json(force=True) or {}
    name = str(payload.get("name", "")).strip()[:400]
    if not name:
        return jsonify({"error": "Thiếu tên kế hoạch"}), 400
    code = str(payload.get("code", "")).strip()[:64]
    status = _mp_normalize_plan_status(payload.get("status"))
    prio = _mp_normalize_plan_priority(payload.get("priority"))
    try:
        fiscal_year = int(payload.get("fiscal_year") or datetime.now().year)
    except (TypeError, ValueError):
        fiscal_year = datetime.now().year
    fiscal_year = max(1990, min(2120, fiscal_year))
    period_label = str(payload.get("period_label", "")).strip()[:120]
    north_star = str(payload.get("north_star", "")).strip()[:2000]
    objectives = str(payload.get("objectives", "")).strip()[:32000]
    audiences = str(payload.get("audiences", "")).strip()[:32000]
    risks_notes = str(payload.get("risks_notes", "")).strip()[:32000]
    notes_field = str(payload.get("notes", "")).strip()[:32000]
    pillars_j, chan_j, met_j = _mp_meta_json(payload)
    strat_j = _mp_strategy_framework_json(
        payload.get("strategy_framework_json") if "strategy_framework_json" in payload else payload.get("strategy_framework")
    )
    tm_prof_j = _mp_target_market_prof_json(
        payload.get("target_market_prof_json")
        if "target_market_prof_json" in payload
        else payload.get("target_market_prof")
    )
    tm_steps4_j = _mp_target_market_steps4_json(
        payload.get("target_market_steps4_json")
        if "target_market_steps4_json" in payload
        else payload.get("target_market_steps4")
    )
    khtn_mr_j = _mp_khtn_market_research_json(
        payload.get("khtn_market_research_json")
        if "khtn_market_research_json" in payload
        else payload.get("khtn_market_research")
    )
    try:
        bp = max(0, min(int(payload.get("budget_planned_vnd") or 0), 9_999_999_999_999))
    except (TypeError, ValueError):
        bp = 0
    try:
        ba = max(0, min(int(payload.get("budget_actual_vnd") or 0), 9_999_999_999_999))
    except (TypeError, ValueError):
        ba = 0
    sd = str(payload.get("start_date") or "").strip()[:32]
    ed = str(payload.get("end_date") or "").strip()[:32]
    if sd and not _crm_validate_date_ymd(sd):
        return jsonify({"error": "start_date không hợp lệ"}), 400
    if ed and not _crm_validate_date_ymd(ed):
        return jsonify({"error": "end_date không hợp lệ"}), 400
    oid = _opt_pos_int(payload.get("owner_staff_id"))
    ts_d = datetime.now().strftime("%Y-%m-%d")
    ts = _crm_ts()
    with get_connection() as conn:
        if oid and conn.execute("SELECT id FROM crm_staff WHERE id = ?", (oid,)).fetchone() is None:
            return jsonify({"error": "Nhân viên owner không tồn tại"}), 400
        cur = conn.execute(
            """
            INSERT INTO crm_marketing_plans (
                code, name, status, priority, fiscal_year, period_label, north_star, objectives,
                pillars_json, audiences, channels_focus_json, budget_planned_vnd, budget_actual_vnd,
                success_metrics_json, risks_notes, owner_staff_id, start_date, end_date, notes,
                strategy_framework_json, target_market_prof_json, target_market_steps4_json, khtn_market_research_json, created_at, updated_at
            ) VALUES (
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
            )
            """,
            (
                code, name, status, prio, fiscal_year, period_label, north_star, objectives,
                pillars_j, audiences, chan_j, bp, ba,
                met_j, risks_notes, oid,
                sd, ed, notes_field, strat_j, tm_prof_j, tm_steps4_j, khtn_mr_j,
                ts_d, ts,
            ),
        )
        nid = int(cur.lastrowid)
        camp_ids_raw = payload.get("campaign_ids")
        if isinstance(camp_ids_raw, list):
            for cid in camp_ids_raw:
                cex = _opt_pos_int(cid)
                if not cex:
                    continue
                if conn.execute("SELECT id FROM crm_campaigns WHERE id = ?", (cex,)).fetchone():
                    conn.execute(
                        "INSERT OR IGNORE INTO crm_marketing_plan_campaigns (plan_id, campaign_id) VALUES (?, ?)",
                        (nid, cex),
                    )
        row = conn.execute(
            """
            SELECT p.*, st.name AS owner_name FROM crm_marketing_plans p
            LEFT JOIN crm_staff st ON st.id = p.owner_staff_id WHERE p.id = ?
            """,
            (nid,),
        ).fetchone()
    assert row is not None
    return jsonify(dict(row)), 201


@app.patch("/api/crm/marketing-plans/<int:plan_id>")
def api_crm_mp_patch(plan_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM crm_marketing_plans WHERE id = ?", (plan_id,)).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy kế hoạch"}), 404
        merged: dict[str, Any] = dict(row)
        if "code" in payload and isinstance(payload["code"], str):
            merged["code"] = payload["code"].strip()[:64]
        if "name" in payload and isinstance(payload["name"], str):
            nm = payload["name"].strip()[:400]
            if not nm:
                return jsonify({"error": "Tên không được trống"}), 400
            merged["name"] = nm
        if "status" in payload:
            merged["status"] = _mp_normalize_plan_status(str(payload.get("status")))
        if "priority" in payload:
            merged["priority"] = _mp_normalize_plan_priority(payload.get("priority"))
        if "fiscal_year" in payload:
            try:
                merged["fiscal_year"] = max(
                    1990, min(2120, int(payload.get("fiscal_year")))
                )
            except (TypeError, ValueError):
                pass
        if "period_label" in payload and isinstance(payload["period_label"], str):
            merged["period_label"] = payload["period_label"].strip()[:120]
        if "north_star" in payload and isinstance(payload["north_star"], str):
            merged["north_star"] = payload["north_star"].strip()[:2000]
        if "objectives" in payload and isinstance(payload["objectives"], str):
            merged["objectives"] = payload["objectives"].strip()[:32000]
        if "pillars_json" in payload:
            pj, _, _ = _mp_meta_json({"pillars_json": payload.get("pillars_json")})
            merged["pillars_json"] = pj
        if "audiences" in payload and isinstance(payload["audiences"], str):
            merged["audiences"] = payload["audiences"].strip()[:32000]
        if "channels_focus" in payload:
            merged["channels_focus_json"] = _mp_channels_focus_json(payload.get("channels_focus"))
        elif "channels_focus_json" in payload:
            merged["channels_focus_json"] = _mp_channels_focus_json(
                payload.get("channels_focus_json")
            )

        if "strategy_framework_json" in payload or "strategy_framework" in payload:
            merged["strategy_framework_json"] = _mp_strategy_framework_json(
                payload.get("strategy_framework_json")
                if "strategy_framework_json" in payload
                else payload.get("strategy_framework")
            )

        if "target_market_prof_json" in payload or "target_market_prof" in payload:
            merged["target_market_prof_json"] = _mp_target_market_prof_json(
                payload.get("target_market_prof_json")
                if "target_market_prof_json" in payload
                else payload.get("target_market_prof")
            )

        if "target_market_steps4_json" in payload or "target_market_steps4" in payload:
            merged["target_market_steps4_json"] = _mp_target_market_steps4_json(
                payload.get("target_market_steps4_json")
                if "target_market_steps4_json" in payload
                else payload.get("target_market_steps4")
            )

        if "khtn_market_research_json" in payload or "khtn_market_research" in payload:
            merged["khtn_market_research_json"] = _mp_khtn_market_research_json(
                payload.get("khtn_market_research_json")
                if "khtn_market_research_json" in payload
                else payload.get("khtn_market_research")
            )

        if "budget_planned_vnd" in payload:
            try:
                merged["budget_planned_vnd"] = max(
                    0,
                    min(int(payload.get("budget_planned_vnd") or 0), 9_999_999_999_999),
                )
            except (TypeError, ValueError):
                pass
        if "budget_actual_vnd" in payload:
            try:
                merged["budget_actual_vnd"] = max(
                    0,
                    min(int(payload.get("budget_actual_vnd") or 0), 9_999_999_999_999),
                )
            except (TypeError, ValueError):
                pass
        if "success_metrics_json" in payload:
            _, _, mj = _mp_meta_json(payload)
            merged["success_metrics_json"] = mj
        if "risks_notes" in payload and isinstance(payload["risks_notes"], str):
            merged["risks_notes"] = payload["risks_notes"].strip()[:32000]
        if "notes" in payload and isinstance(payload["notes"], str):
            merged["notes"] = payload["notes"].strip()[:32000]
        if "owner_staff_id" in payload:
            raw_o = payload.get("owner_staff_id")
            if raw_o in (None, "", 0, "0"):
                merged["owner_staff_id"] = None
            else:
                oid = _opt_pos_int(raw_o)
                if oid and conn.execute(
                    "SELECT id FROM crm_staff WHERE id = ?", (oid,)
                ).fetchone():
                    merged["owner_staff_id"] = oid
                elif oid:
                    return jsonify({"error": "Nhân viên không tồn tại"}), 404
                else:
                    merged["owner_staff_id"] = None

        if "start_date" in payload and isinstance(payload["start_date"], str):
            sd = payload["start_date"].strip()[:32]
            if sd and not _crm_validate_date_ymd(sd):
                return jsonify({"error": "start_date không hợp lệ"}), 400
            merged["start_date"] = sd
        if "end_date" in payload and isinstance(payload["end_date"], str):
            ed = payload["end_date"].strip()[:32]
            if ed and not _crm_validate_date_ymd(ed):
                return jsonify({"error": "end_date không hợp lệ"}), 400
            merged["end_date"] = ed

        oid_final = merged.get("owner_staff_id")
        ts = _crm_ts()
        conn.execute(
            """
            UPDATE crm_marketing_plans SET
              code = ?, name = ?, status = ?, priority = ?, fiscal_year = ?, period_label = ?,
              north_star = ?, objectives = ?, pillars_json = ?, audiences = ?, channels_focus_json = ?,
              budget_planned_vnd = ?, budget_actual_vnd = ?, success_metrics_json = ?,
              risks_notes = ?, owner_staff_id = ?, start_date = ?, end_date = ?, notes = ?,
              strategy_framework_json = ?, target_market_prof_json = ?, target_market_steps4_json = ?, khtn_market_research_json = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                merged.get("code") or "",
                merged["name"],
                merged["status"],
                merged["priority"],
                int(merged.get("fiscal_year") or 2026),
                merged.get("period_label") or "",
                merged.get("north_star") or "",
                merged.get("objectives") or "",
                merged.get("pillars_json") or "[]",
                merged.get("audiences") or "",
                merged.get("channels_focus_json") or "[]",
                int(merged.get("budget_planned_vnd") or 0),
                int(merged.get("budget_actual_vnd") or 0),
                merged.get("success_metrics_json") or "[]",
                merged.get("risks_notes") or "",
                oid_final,
                str(merged.get("start_date") or ""),
                str(merged.get("end_date") or ""),
                merged.get("notes") or "",
                merged.get("strategy_framework_json") or "{}",
                merged.get("target_market_prof_json") or "{}",
                merged.get("target_market_steps4_json") or "{}",
                merged.get("khtn_market_research_json") or "{}",
                ts,
                plan_id,
            ),
        )
        row_out = conn.execute(
            """
            SELECT p.*, st.name AS owner_name
            FROM crm_marketing_plans p
            LEFT JOIN crm_staff st ON st.id = p.owner_staff_id WHERE p.id = ?
            """,
            (plan_id,),
        ).fetchone()
    assert row_out is not None
    return jsonify(dict(row_out))


@app.delete("/api/crm/marketing-plans/<int:plan_id>")
def api_crm_mp_delete(plan_id: int) -> Any:
    with get_connection() as conn:
        cur = conn.execute("DELETE FROM crm_marketing_plans WHERE id = ?", (plan_id,))
        if cur.rowcount == 0:
            return jsonify({"error": "Không tìm thấy kế hoạch"}), 404
    return jsonify({"ok": True})


@app.put("/api/crm/marketing-plans/<int:plan_id>/campaigns")
def api_crm_mp_set_campaigns(plan_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    raw = payload.get("campaign_ids")
    if raw is None or not isinstance(raw, list):
        return jsonify({"error": "Cần campaign_ids: [...]"}), 400
    with get_connection() as conn:
        if conn.execute("SELECT id FROM crm_marketing_plans WHERE id = ?", (plan_id,)).fetchone() is None:
            return jsonify({"error": "Không tìm thấy kế hoạch"}), 404
        conn.execute("DELETE FROM crm_marketing_plan_campaigns WHERE plan_id = ?", (plan_id,))
        for cid_any in raw:
            cid = _opt_pos_int(cid_any)
            if not cid:
                continue
            if conn.execute("SELECT id FROM crm_campaigns WHERE id = ?", (cid,)).fetchone():
                conn.execute(
                    "INSERT OR IGNORE INTO crm_marketing_plan_campaigns (plan_id, campaign_id) VALUES (?, ?)",
                    (plan_id, cid),
                )
        c_rows = conn.execute(
            """
            SELECT c.*
            FROM crm_marketing_plan_campaigns l
            JOIN crm_campaigns c ON c.id = l.campaign_id
            WHERE l.plan_id = ?
            ORDER BY c.name COLLATE NOCASE ASC
            """,
            (plan_id,),
        ).fetchall()
    return jsonify({"campaigns": rows_to_dict(c_rows)})


@app.post("/api/crm/marketing-plans/<int:plan_id>/milestones")
def api_crm_mp_add_milestone(plan_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    title = str(payload.get("title", "")).strip()[:500]
    if not title:
        return jsonify({"error": "Thiếu tiêu đề mốc"}), 400
    due_date = str(payload.get("due_date") or "").strip()[:32]
    if due_date and not _crm_validate_date_ymd(due_date):
        return jsonify({"error": "due_date phải YYYY-MM-DD"}), 400
    status_ms = _mp_normalize_milestone_status(payload.get("status"))
    notes = str(payload.get("notes", "")).strip()[:8000]
    ts_d = datetime.now().strftime("%Y-%m-%d")
    ts = _crm_ts()
    with get_connection() as conn:
        if conn.execute("SELECT id FROM crm_marketing_plans WHERE id = ?", (plan_id,)).fetchone() is None:
            return jsonify({"error": "Không tìm thấy kế hoạch"}), 404
        mx = conn.execute(
            "SELECT COALESCE(MAX(position), -1) AS m FROM crm_marketing_plan_milestones WHERE plan_id = ?",
            (plan_id,),
        ).fetchone()
        pos = int(mx["m"]) + 1 if mx else 0
        if "position" in payload:
            try:
                pos = int(payload["position"])
            except (TypeError, ValueError):
                pass
        cur = conn.execute(
            """
            INSERT INTO crm_marketing_plan_milestones (
              plan_id, position, title, due_date, status, notes, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (plan_id, pos, title, due_date, status_ms, notes, ts_d, ts),
        )
        nid = int(cur.lastrowid)
        row = conn.execute("SELECT * FROM crm_marketing_plan_milestones WHERE id = ?", (nid,)).fetchone()
    assert row is not None
    return jsonify(dict(row)), 201


@app.patch("/api/crm/marketing-plan-milestones/<int:mid>")
def api_crm_mp_patch_milestone(mid: int) -> Any:
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute(
            "SELECT * FROM crm_marketing_plan_milestones WHERE id = ?", (mid,)
        ).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy mốc"}), 404
        merged = dict(row)
        if "title" in payload and isinstance(payload["title"], str):
            tl = payload["title"].strip()[:500]
            if not tl:
                return jsonify({"error": "Tiêu đề không được trống"}), 400
            merged["title"] = tl
        if "due_date" in payload and isinstance(payload["due_date"], str):
            dd = payload["due_date"].strip()[:32]
            if dd and not _crm_validate_date_ymd(dd):
                return jsonify({"error": "due_date không hợp lệ"}), 400
            merged["due_date"] = dd
        if "status" in payload:
            merged["status"] = _mp_normalize_milestone_status(str(payload.get("status")))
        if "notes" in payload and isinstance(payload["notes"], str):
            merged["notes"] = payload["notes"].strip()[:8000]
        if "position" in payload:
            try:
                merged["position"] = int(payload["position"])
            except (TypeError, ValueError):
                pass
        ts = _crm_ts()
        conn.execute(
            """
            UPDATE crm_marketing_plan_milestones
            SET title = ?, due_date = ?, status = ?, notes = ?, position = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                merged["title"],
                merged.get("due_date") or "",
                merged["status"],
                merged.get("notes") or "",
                int(merged.get("position") or 0),
                ts,
                mid,
            ),
        )
        row2 = conn.execute("SELECT * FROM crm_marketing_plan_milestones WHERE id = ?", (mid,)).fetchone()
    assert row2 is not None
    return jsonify(dict(row2))


@app.delete("/api/crm/marketing-plan-milestones/<int:mid>")
def api_crm_mp_delete_milestone(mid: int) -> Any:
    with get_connection() as conn:
        cur = conn.execute("DELETE FROM crm_marketing_plan_milestones WHERE id = ?", (mid,))
        if cur.rowcount == 0:
            return jsonify({"error": "Không tìm thấy mốc"}), 404
    return jsonify({"ok": True})


@app.get("/crm/sop")
def crm_sop_page() -> str:
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    return render_template(
        "crm_sop.html",
        sop_task_status_labels=CRM_SOP_TASK_STATUS_LABELS_VI,
        sop_task_statuses=list(CRM_SOP_TASK_STATUSES),
        sop_run_status_labels=CRM_SOP_RUN_STATUS_LABELS_VI,
        sop_run_statuses=list(CRM_SOP_RUN_STATUSES),
        sop_step_roles=list(CRM_SOP_STEP_ROLES),
        sop_step_role_labels=CRM_SOP_STEP_ROLE_LABELS_VI,
        campaign_channel_labels=CRM_CAMPAIGN_CHANNEL_LABELS_VI,
        campaign_channels=list(CRM_CAMPAIGN_CHANNELS),
        **_admin_page_template_kwargs(),
    )


@app.get("/crm/hdsd")
@app.get("/crm/hdsd/<section>/<slug>")
def crm_hdsd_page(section: str | None = None, slug: str | None = None) -> Any:
    """HDSD — đọc tài liệu Markdown từ docs/ trên hệ thống."""
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    if not _admin_section_can("crm_hdsd", "view"):
        return redirect(url_for("crm_board"))
    from crm_hdsd_docs import list_hdsd_catalog, read_hdsd_doc

    catalog = list_hdsd_catalog()
    view_doc = None
    doc_markdown = ""
    if section and slug:
        loaded = read_hdsd_doc(section, slug)
        if loaded is None:
            abort(404)
        doc_markdown, view_doc = loaded
    return render_template(
        "crm_hdsd.html",
        catalog=catalog,
        view_doc=view_doc,
        doc_markdown=doc_markdown,
        can_export=_admin_section_can("crm_hdsd", "export"),
        **_admin_page_template_kwargs(),
    )


@app.get("/crm/hdsd/<section>/<slug>/download")
def crm_hdsd_doc_download(section: str, slug: str) -> Any:
    if not _admin_section_can("crm_hdsd", "export"):
        return _admin_section_forbidden_json("crm_hdsd", "export")
    from crm_hdsd_docs import resolve_hdsd_doc

    resolved = resolve_hdsd_doc(section, slug)
    if resolved is None:
        abort(404)
    path, meta = resolved
    return send_file(
        path,
        mimetype="text/markdown; charset=utf-8",
        as_attachment=True,
        download_name=meta["filename"],
    )


@app.get("/crm/hdsd/<section>/<slug>/download.xlsx")
def crm_hdsd_doc_download_xlsx(section: str, slug: str) -> Any:
    if not _admin_section_can("crm_hdsd", "export"):
        return _admin_section_forbidden_json("crm_hdsd", "export")
    from crm_hdsd_docs import read_hdsd_doc
    from crm_hdsd_excel import build_hdsd_doc_xlsx

    loaded = read_hdsd_doc(section, slug)
    if loaded is None:
        abort(404)
    text, meta = loaded
    buf = build_hdsd_doc_xlsx(text, meta)
    stem = str(meta.get("filename") or slug).rsplit(".", 1)[0]
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{stem}.xlsx",
        etag=False,
        max_age=0,
        conditional=False,
    )


@app.get("/crm/hdsd/export-all.zip")
def crm_hdsd_export_all_zip() -> Any:
    if not _admin_section_can("crm_hdsd", "export"):
        return _admin_section_forbidden_json("crm_hdsd", "export")
    from crm_hdsd_docs import list_hdsd_catalog, read_hdsd_doc
    from crm_hdsd_excel import build_hdsd_all_zip

    catalog = list_hdsd_catalog()
    buf = build_hdsd_all_zip(catalog, read_hdsd_doc)
    return send_file(
        buf,
        mimetype="application/zip",
        as_attachment=True,
        download_name="hdsd-tai-lieu-excel.zip",
        etag=False,
        max_age=0,
        conditional=False,
    )


@app.get("/crm/test-cases/download.xlsx")
def crm_test_cases_download_xlsx() -> Any:
    """Bộ test case CRM — Excel đầy đủ cột + sơ đồ cho tester."""
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    if not _admin_section_can("crm_hdsd", "view"):
        return _admin_section_forbidden_json("crm_hdsd", "view")
    from crm_test_cases_workbook import build_crm_test_cases_workbook

    buf = build_crm_test_cases_workbook()
    stamp = date.today().strftime("%Y%m%d")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"CRM-test-cases-{stamp}.xlsx",
        etag=False,
        max_age=0,
        conditional=False,
    )


@app.get("/api/crm/hdsd/catalog")
def api_crm_hdsd_catalog() -> Any:
    if not _admin_section_can("crm_hdsd", "view"):
        return _admin_section_forbidden_json("crm_hdsd", "view")
    from crm_hdsd_docs import list_hdsd_catalog

    return jsonify({"catalog": list_hdsd_catalog()})


# ── helpers ──────────────────────────────────────────────────────────────────

def _sop_normalize_role(raw: str | None) -> str:
    s = str(raw or "any").strip().lower()
    return s if s in CRM_SOP_STEP_ROLES else "any"


def _sop_normalize_task_status(raw: str | None) -> str:
    s = str(raw or "todo").strip().lower()
    return s if s in CRM_SOP_TASK_STATUSES else "todo"


def _sop_normalize_run_status(raw: str | None) -> str:
    s = str(raw or "active").strip().lower()
    return s if s in CRM_SOP_RUN_STATUSES else "active"


def _sop_run_stats(conn: sqlite3.Connection, run_id: int) -> dict[str, int]:
    rows = conn.execute(
        "SELECT status FROM crm_sop_run_tasks WHERE run_id = ?", (run_id,)
    ).fetchall()
    total = len(rows)
    done = sum(1 for r in rows if r["status"] == "done")
    skipped = sum(1 for r in rows if r["status"] == "skipped")
    in_prog = sum(1 for r in rows if r["status"] == "in_progress")
    overdue_count = 0
    today = datetime.now().strftime("%Y-%m-%d")
    for r in rows:
        task_row = conn.execute(
            "SELECT due_date FROM crm_sop_run_tasks WHERE run_id = ? AND status NOT IN ('done','skipped')",
            (run_id,),
        )
    overdue_rows = conn.execute(
        """
        SELECT count(*) AS cnt FROM crm_sop_run_tasks
        WHERE run_id = ? AND status NOT IN ('done','skipped')
          AND due_date != '' AND due_date < ?
        """,
        (run_id, today),
    ).fetchone()
    overdue_count = int((overdue_rows["cnt"] if overdue_rows else 0))
    return {
        "total": total,
        "done": done,
        "skipped": skipped,
        "in_progress": in_prog,
        "todo": total - done - skipped - in_prog,
        "overdue": overdue_count,
    }


def _sop_row_run(row: sqlite3.Row, conn: sqlite3.Connection) -> dict[str, Any]:
    d = dict(row)
    d["stats"] = _sop_run_stats(conn, int(d["id"]))
    return d


def _sop_generate_tasks(
    conn: sqlite3.Connection, run_id: int, template_id: int, start_date: str
) -> None:
    """Instantiate SOP steps into run_tasks for the given run."""
    steps = conn.execute(
        """
        SELECT * FROM crm_sop_steps
        WHERE template_id = ?
        ORDER BY position ASC, id ASC
        """,
        (template_id,),
    ).fetchall()
    ts = _crm_ts()
    today = datetime.now().strftime("%Y-%m-%d")
    for step in steps:
        d = dict(step)
        if start_date and _crm_validate_date_ymd(start_date):
            due = _crm_hub_date_add_days_iso(start_date, int(d.get("offset_days") or 0)) or ""
        else:
            due = ""
        conn.execute(
            """
            INSERT INTO crm_sop_run_tasks (
                run_id, step_id, position, title, description,
                role, due_date, status, checklist_json, notes, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, 'todo', ?, '', ?, ?)
            """,
            (
                run_id,
                int(d["id"]),
                int(d.get("position") or 0),
                str(d["title"]),
                str(d.get("description") or ""),
                _sop_normalize_role(d.get("role")),
                due,
                str(d.get("checklist_json") or "[]"),
                today,
                ts,
            ),
        )


# ── SOP Template APIs ────────────────────────────────────────────────────────

@app.get("/api/crm/sop/templates")
def api_crm_sop_list_templates() -> Any:
    raw = (request.args.get("include_inactive") or "").strip().lower()
    incl = raw in ("1", "true", "yes", "all")
    try:
        from ptt_crm.config import sop_read_source_pg
        from ptt_crm.sop_pg_read import list_sop_templates

        if sop_read_source_pg():
            templates = list_sop_templates(active_only=not incl)
            if templates:
                return jsonify({"templates": templates, "read_source": "pg"})
    except Exception:
        pass
    with get_connection() as conn:
        if incl:
            rows = conn.execute(
                "SELECT * FROM crm_sop_templates ORDER BY active DESC, name COLLATE NOCASE ASC"
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM crm_sop_templates WHERE active = 1 ORDER BY name COLLATE NOCASE ASC"
            ).fetchall()
    return jsonify({"templates": rows_to_dict(rows)})


@app.get("/api/crm/sop/templates/<int:tpl_id>")
def api_crm_sop_get_template(tpl_id: int) -> Any:
    with get_connection() as conn:
        tpl = conn.execute(
            "SELECT * FROM crm_sop_templates WHERE id = ?", (tpl_id,)
        ).fetchone()
        if tpl is None:
            return jsonify({"error": "Không tìm thấy template"}), 404
        steps = conn.execute(
            "SELECT * FROM crm_sop_steps WHERE template_id = ? ORDER BY position ASC, id ASC",
            (tpl_id,),
        ).fetchall()
    return jsonify({"template": dict(tpl), "steps": rows_to_dict(steps)})


@app.post("/api/crm/sop/templates")
def api_crm_sop_create_template() -> Any:
    payload = request.get_json(force=True) or {}
    name = str(payload.get("name", "")).strip()[:300]
    if not name:
        return jsonify({"error": "Thiếu tên SOP"}), 400
    code = str(payload.get("code", "")).strip()[:64]
    ch = _crm_normalize_campaign_channel(str(payload.get("channel") or "other"))
    desc = str(payload.get("description", "")).strip()[:4000]
    notes = str(payload.get("notes", "")).strip()[:8000]
    ts_d = datetime.now().strftime("%Y-%m-%d")
    ts = _crm_ts()
    with get_connection() as conn:
        cur = conn.execute(
            """
            INSERT INTO crm_sop_templates (code, name, channel, description, notes, active, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, 1, ?, ?)
            """,
            (code, name, ch, desc, notes, ts_d, ts),
        )
        nid = int(cur.lastrowid)
        row = conn.execute("SELECT * FROM crm_sop_templates WHERE id = ?", (nid,)).fetchone()
    assert row is not None
    return jsonify(dict(row)), 201


@app.patch("/api/crm/sop/templates/<int:tpl_id>")
def api_crm_sop_patch_template(tpl_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM crm_sop_templates WHERE id = ?", (tpl_id,)).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy template"}), 404
        merged = dict(row)
        if "name" in payload and isinstance(payload["name"], str):
            nm = payload["name"].strip()[:300]
            if not nm:
                return jsonify({"error": "Tên không được trống"}), 400
            merged["name"] = nm
        if "code" in payload and isinstance(payload["code"], str):
            merged["code"] = payload["code"].strip()[:64]
        if "channel" in payload:
            merged["channel"] = _crm_normalize_campaign_channel(payload.get("channel"))
        if "description" in payload and isinstance(payload["description"], str):
            merged["description"] = payload["description"].strip()[:4000]
        if "notes" in payload and isinstance(payload["notes"], str):
            merged["notes"] = payload["notes"].strip()[:8000]
        if "active" in payload:
            merged["active"] = 1 if payload["active"] in (True, 1, "1", "true") else 0
        ts = _crm_ts()
        conn.execute(
            """
            UPDATE crm_sop_templates
            SET code = ?, name = ?, channel = ?, description = ?, notes = ?, active = ?, updated_at = ?
            WHERE id = ?
            """,
            (merged["code"], merged["name"], merged["channel"], merged["description"],
             merged["notes"], int(merged.get("active") or 0), ts, tpl_id),
        )
        row2 = conn.execute("SELECT * FROM crm_sop_templates WHERE id = ?", (tpl_id,)).fetchone()
    assert row2 is not None
    return jsonify(dict(row2))


@app.delete("/api/crm/sop/templates/<int:tpl_id>")
def api_crm_sop_delete_template(tpl_id: int) -> Any:
    with get_connection() as conn:
        runs = conn.execute(
            "SELECT count(*) AS c FROM crm_sop_runs WHERE template_id = ?", (tpl_id,)
        ).fetchone()
        if runs and int(runs["c"]) > 0:
            return jsonify({"error": "Không xoá được — đã có SOP Run tham chiếu template này. Hãy lưu trữ thay vì xoá."}), 400
        cur = conn.execute("DELETE FROM crm_sop_templates WHERE id = ?", (tpl_id,))
        if cur.rowcount == 0:
            return jsonify({"error": "Không tìm thấy template"}), 404
    return jsonify({"ok": True})


# ── SOP Step APIs ─────────────────────────────────────────────────────────────

@app.get("/api/crm/sop/templates/<int:tpl_id>/steps")
def api_crm_sop_list_steps(tpl_id: int) -> Any:
    with get_connection() as conn:
        if conn.execute("SELECT id FROM crm_sop_templates WHERE id = ?", (tpl_id,)).fetchone() is None:
            return jsonify({"error": "Không tìm thấy template"}), 404
        rows = conn.execute(
            "SELECT * FROM crm_sop_steps WHERE template_id = ? ORDER BY position ASC, id ASC",
            (tpl_id,),
        ).fetchall()
    return jsonify({"steps": rows_to_dict(rows)})


@app.post("/api/crm/sop/templates/<int:tpl_id>/steps")
def api_crm_sop_add_step(tpl_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    title = str(payload.get("title", "")).strip()[:500]
    if not title:
        return jsonify({"error": "Thiếu tiêu đề bước"}), 400
    desc = str(payload.get("description", "")).strip()[:8000]
    try:
        offset_days = int(payload.get("offset_days") or 0)
    except (TypeError, ValueError):
        offset_days = 0
    try:
        duration_days = max(1, int(payload.get("duration_days") or 1))
    except (TypeError, ValueError):
        duration_days = 1
    role = _sop_normalize_role(payload.get("role"))
    required = 1 if payload.get("required") not in (False, 0, "0", "false") else 0
    checklist_raw = payload.get("checklist_json")
    if isinstance(checklist_raw, list):
        import json as _json
        checklist = _json.dumps(checklist_raw, ensure_ascii=False)[:4000]
    else:
        checklist = "[]"
    ts = _crm_ts()
    today = datetime.now().strftime("%Y-%m-%d")
    with get_connection() as conn:
        if conn.execute("SELECT id FROM crm_sop_templates WHERE id = ?", (tpl_id,)).fetchone() is None:
            return jsonify({"error": "Không tìm thấy template"}), 404
        max_pos = conn.execute(
            "SELECT coalesce(max(position), -1) AS m FROM crm_sop_steps WHERE template_id = ?", (tpl_id,)
        ).fetchone()
        pos = int(max_pos["m"]) + 1 if max_pos else 0
        if "position" in payload:
            try:
                pos = int(payload["position"])
            except (TypeError, ValueError):
                pass
        cur = conn.execute(
            """
            INSERT INTO crm_sop_steps (
                template_id, position, title, description, offset_days, duration_days,
                role, required, checklist_json, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (tpl_id, pos, title, desc, offset_days, duration_days, role, required, checklist, today, ts),
        )
        nid = int(cur.lastrowid)
        row = conn.execute("SELECT * FROM crm_sop_steps WHERE id = ?", (nid,)).fetchone()
    assert row is not None
    return jsonify(dict(row)), 201


@app.patch("/api/crm/sop/steps/<int:step_id>")
def api_crm_sop_patch_step(step_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM crm_sop_steps WHERE id = ?", (step_id,)).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy bước"}), 404
        merged = dict(row)
        if "title" in payload and isinstance(payload["title"], str):
            t = payload["title"].strip()[:500]
            if not t:
                return jsonify({"error": "Tiêu đề không được trống"}), 400
            merged["title"] = t
        if "description" in payload and isinstance(payload["description"], str):
            merged["description"] = payload["description"].strip()[:8000]
        if "offset_days" in payload:
            try:
                merged["offset_days"] = int(payload["offset_days"])
            except (TypeError, ValueError):
                pass
        if "duration_days" in payload:
            try:
                merged["duration_days"] = max(1, int(payload["duration_days"] or 1))
            except (TypeError, ValueError):
                pass
        if "role" in payload:
            merged["role"] = _sop_normalize_role(payload.get("role"))
        if "required" in payload:
            merged["required"] = 1 if payload["required"] not in (False, 0, "0", "false") else 0
        if "position" in payload:
            try:
                merged["position"] = int(payload["position"])
            except (TypeError, ValueError):
                pass
        if "checklist_json" in payload and isinstance(payload["checklist_json"], list):
            import json as _json
            merged["checklist_json"] = _json.dumps(payload["checklist_json"], ensure_ascii=False)[:4000]
        ts = _crm_ts()
        conn.execute(
            """
            UPDATE crm_sop_steps
            SET title = ?, description = ?, offset_days = ?, duration_days = ?,
                role = ?, required = ?, position = ?, checklist_json = ?, updated_at = ?
            WHERE id = ?
            """,
            (merged["title"], merged["description"], int(merged.get("offset_days") or 0),
             int(merged.get("duration_days") or 1), merged["role"], int(merged.get("required") or 1),
             int(merged.get("position") or 0), merged["checklist_json"], ts, step_id),
        )
        row2 = conn.execute("SELECT * FROM crm_sop_steps WHERE id = ?", (step_id,)).fetchone()
    assert row2 is not None
    return jsonify(dict(row2))


@app.delete("/api/crm/sop/steps/<int:step_id>")
def api_crm_sop_delete_step(step_id: int) -> Any:
    with get_connection() as conn:
        cur = conn.execute("DELETE FROM crm_sop_steps WHERE id = ?", (step_id,))
        if cur.rowcount == 0:
            return jsonify({"error": "Không tìm thấy bước"}), 404
    return jsonify({"ok": True})


@app.post("/api/crm/sop/templates/<int:tpl_id>/steps/reorder")
def api_crm_sop_reorder_steps(tpl_id: int) -> Any:
    """Body: {"order": [step_id, step_id, ...]}"""
    payload = request.get_json(force=True) or {}
    order = payload.get("order")
    if not isinstance(order, list):
        return jsonify({"error": "Cần order: [id, id, ...]"}), 400
    ts = _crm_ts()
    with get_connection() as conn:
        if conn.execute("SELECT id FROM crm_sop_templates WHERE id = ?", (tpl_id,)).fetchone() is None:
            return jsonify({"error": "Không tìm thấy template"}), 404
        for pos, sid in enumerate(order):
            try:
                conn.execute(
                    "UPDATE crm_sop_steps SET position = ?, updated_at = ? WHERE id = ? AND template_id = ?",
                    (pos, ts, int(sid), tpl_id),
                )
            except (TypeError, ValueError):
                pass
    return jsonify({"ok": True})


# ── SOP Run APIs ──────────────────────────────────────────────────────────────

@app.get("/api/crm/sop/runs")
def api_crm_sop_list_runs() -> Any:
    status_filter = (request.args.get("status") or "active").strip().lower()
    if status_filter not in CRM_SOP_RUN_STATUSES and status_filter != "all":
        status_filter = "active"
    try:
        from ptt_crm.config import sop_read_source_pg
        from ptt_crm.sop_pg_read import list_sop_runs

        if sop_read_source_pg():
            st = None if status_filter == "all" else status_filter
            runs = list_sop_runs(status=st)
            if runs:
                return jsonify({"runs": runs, "read_source": "pg"})
    except Exception:
        pass
    with get_connection() as conn:
        if status_filter == "all":
            rows = conn.execute(
                """
                SELECT r.*, t.name AS template_name, t.channel AS template_channel,
                       c.name AS campaign_name, c.code AS campaign_code
                FROM crm_sop_runs r
                LEFT JOIN crm_sop_templates t ON t.id = r.template_id
                LEFT JOIN crm_campaigns c ON c.id = r.campaign_id
                ORDER BY datetime(r.updated_at) DESC, r.id DESC
                LIMIT 300
                """
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT r.*, t.name AS template_name, t.channel AS template_channel,
                       c.name AS campaign_name, c.code AS campaign_code
                FROM crm_sop_runs r
                LEFT JOIN crm_sop_templates t ON t.id = r.template_id
                LEFT JOIN crm_campaigns c ON c.id = r.campaign_id
                WHERE r.status = ?
                ORDER BY r.start_date ASC, r.id ASC
                LIMIT 300
                """,
                (status_filter,),
            ).fetchall()
        result = []
        for r in rows:
            d = _sop_row_run(r, conn)
            result.append(d)
    return jsonify({"runs": result})


@app.post("/api/crm/sop/runs")
def api_crm_sop_create_run() -> Any:
    """Create a run (optionally launch from a template → auto-generate tasks)."""
    payload = request.get_json(force=True) or {}
    name = str(payload.get("name", "")).strip()[:400]
    if not name:
        return jsonify({"error": "Thiếu tên SOP Run"}), 400
    campaign_id = _opt_pos_int(payload.get("campaign_id"))
    template_id = _opt_pos_int(payload.get("template_id"))
    start_date = str(payload.get("start_date") or "").strip()[:32]
    if start_date and not _crm_validate_date_ymd(start_date):
        return jsonify({"error": "start_date phải YYYY-MM-DD"}), 400
    status = _sop_normalize_run_status(payload.get("status"))
    notes = str(payload.get("notes", "")).strip()[:8000]
    ts_d = datetime.now().strftime("%Y-%m-%d")
    ts = _crm_ts()
    with get_connection() as conn:
        if campaign_id and conn.execute(
            "SELECT id FROM crm_campaigns WHERE id = ?", (campaign_id,)
        ).fetchone() is None:
            return jsonify({"error": "Chiến dịch không tồn tại"}), 404
        if template_id and conn.execute(
            "SELECT id FROM crm_sop_templates WHERE id = ?", (template_id,)
        ).fetchone() is None:
            return jsonify({"error": "Template SOP không tồn tại"}), 404
        cur = conn.execute(
            """
            INSERT INTO crm_sop_runs (campaign_id, template_id, name, status, start_date, notes, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (campaign_id, template_id, name, status, start_date, notes, ts_d, ts),
        )
        run_id = int(cur.lastrowid)
        if template_id and payload.get("generate_tasks", True) is not False:
            _sop_generate_tasks(conn, run_id, template_id, start_date)
        run_row = conn.execute(
            """
            SELECT r.*, t.name AS template_name, t.channel AS template_channel,
                   c.name AS campaign_name, c.code AS campaign_code
            FROM crm_sop_runs r
            LEFT JOIN crm_sop_templates t ON t.id = r.template_id
            LEFT JOIN crm_campaigns c ON c.id = r.campaign_id
            WHERE r.id = ?
            """,
            (run_id,),
        ).fetchone()
    assert run_row is not None
    with get_connection() as conn2:
        d = _sop_row_run(run_row, conn2)
    return jsonify(d), 201


@app.patch("/api/crm/sop/runs/<int:run_id>")
def api_crm_sop_patch_run(run_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM crm_sop_runs WHERE id = ?", (run_id,)).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy SOP Run"}), 404
        merged = dict(row)
        if "name" in payload and isinstance(payload["name"], str):
            nm = payload["name"].strip()[:400]
            if not nm:
                return jsonify({"error": "Tên không được trống"}), 400
            merged["name"] = nm
        if "status" in payload:
            merged["status"] = _sop_normalize_run_status(payload.get("status"))
        if "start_date" in payload and isinstance(payload["start_date"], str):
            sd = payload["start_date"].strip()[:32]
            if sd and not _crm_validate_date_ymd(sd):
                return jsonify({"error": "start_date phải YYYY-MM-DD"}), 400
            merged["start_date"] = sd
        if "notes" in payload and isinstance(payload["notes"], str):
            merged["notes"] = payload["notes"].strip()[:8000]
        if "campaign_id" in payload:
            raw_c = payload.get("campaign_id")
            if raw_c in (None, "", 0, "0"):
                merged["campaign_id"] = None
            else:
                nc = _opt_pos_int(raw_c)
                if nc and conn.execute(
                    "SELECT id FROM crm_campaigns WHERE id = ?", (nc,)
                ).fetchone() is None:
                    return jsonify({"error": "Chiến dịch không tồn tại"}), 404
                merged["campaign_id"] = nc
        ts = _crm_ts()
        conn.execute(
            """
            UPDATE crm_sop_runs
            SET name = ?, status = ?, start_date = ?, notes = ?, campaign_id = ?, updated_at = ?
            WHERE id = ?
            """,
            (merged["name"], merged["status"], merged.get("start_date") or "",
             merged.get("notes") or "", merged.get("campaign_id"), ts, run_id),
        )
        run_row = conn.execute(
            """
            SELECT r.*, t.name AS template_name, t.channel AS template_channel,
                   c.name AS campaign_name, c.code AS campaign_code
            FROM crm_sop_runs r
            LEFT JOIN crm_sop_templates t ON t.id = r.template_id
            LEFT JOIN crm_campaigns c ON c.id = r.campaign_id
            WHERE r.id = ?
            """,
            (run_id,),
        ).fetchone()
    assert run_row is not None
    with get_connection() as conn2:
        d = _sop_row_run(run_row, conn2)
    return jsonify(d)


# ── SOP Run Task APIs ─────────────────────────────────────────────────────────

@app.get("/api/crm/sop/runs/<int:run_id>/tasks")
def api_crm_sop_run_tasks(run_id: int) -> Any:
    with get_connection() as conn:
        if conn.execute("SELECT id FROM crm_sop_runs WHERE id = ?", (run_id,)).fetchone() is None:
            return jsonify({"error": "Không tìm thấy SOP Run"}), 404
        rows = conn.execute(
            """
            SELECT rt.*, st.name AS assigned_staff_name
            FROM crm_sop_run_tasks rt
            LEFT JOIN crm_staff st ON st.id = rt.assigned_staff_id
            WHERE rt.run_id = ?
            ORDER BY rt.position ASC, rt.id ASC
            """,
            (run_id,),
        ).fetchall()
    return jsonify({"tasks": rows_to_dict(rows)})


@app.post("/api/crm/sop/runs/<int:run_id>/tasks")
def api_crm_sop_add_run_task(run_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    title = str(payload.get("title", "")).strip()[:500]
    if not title:
        return jsonify({"error": "Thiếu tiêu đề task"}), 400
    desc = str(payload.get("description", "")).strip()[:8000]
    role = _sop_normalize_role(payload.get("role"))
    due_date = str(payload.get("due_date") or "").strip()[:32]
    if due_date and not _crm_validate_date_ymd(due_date):
        return jsonify({"error": "due_date phải YYYY-MM-DD"}), 400
    status = _sop_normalize_task_status(payload.get("status"))
    staff_id = _opt_pos_int(payload.get("assigned_staff_id"))
    notes = str(payload.get("notes", "")).strip()[:8000]
    ts_d = datetime.now().strftime("%Y-%m-%d")
    ts = _crm_ts()
    with get_connection() as conn:
        if conn.execute("SELECT id FROM crm_sop_runs WHERE id = ?", (run_id,)).fetchone() is None:
            return jsonify({"error": "Không tìm thấy SOP Run"}), 404
        if staff_id and conn.execute("SELECT id FROM crm_staff WHERE id = ?", (staff_id,)).fetchone() is None:
            return jsonify({"error": "Nhân viên không tồn tại"}), 404
        max_pos = conn.execute(
            "SELECT coalesce(max(position), -1) AS m FROM crm_sop_run_tasks WHERE run_id = ?", (run_id,)
        ).fetchone()
        pos = int(max_pos["m"]) + 1 if max_pos else 0
        cur = conn.execute(
            """
            INSERT INTO crm_sop_run_tasks (
                run_id, step_id, position, title, description, role,
                due_date, status, assigned_staff_id, notes, checklist_json, created_at, updated_at
            ) VALUES (?, NULL, ?, ?, ?, ?, ?, ?, ?, ?, '[]', ?, ?)
            """,
            (run_id, pos, title, desc, role, due_date, status, staff_id, notes, ts_d, ts),
        )
        nid = int(cur.lastrowid)
        row = conn.execute(
            """
            SELECT rt.*, st.name AS assigned_staff_name
            FROM crm_sop_run_tasks rt
            LEFT JOIN crm_staff st ON st.id = rt.assigned_staff_id
            WHERE rt.id = ?
            """,
            (nid,),
        ).fetchone()
    assert row is not None
    return jsonify(dict(row)), 201


@app.patch("/api/crm/sop/run_tasks/<int:task_id>")
def api_crm_sop_patch_run_task(task_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT rt.*, st.name AS assigned_staff_name
            FROM crm_sop_run_tasks rt
            LEFT JOIN crm_staff st ON st.id = rt.assigned_staff_id
            WHERE rt.id = ?
            """,
            (task_id,),
        ).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy task"}), 404
        merged = dict(row)
        if "title" in payload and isinstance(payload["title"], str):
            t = payload["title"].strip()[:500]
            if not t:
                return jsonify({"error": "Tiêu đề không được trống"}), 400
            merged["title"] = t
        if "description" in payload and isinstance(payload["description"], str):
            merged["description"] = payload["description"].strip()[:8000]
        if "role" in payload:
            merged["role"] = _sop_normalize_role(payload.get("role"))
        if "due_date" in payload and isinstance(payload["due_date"], str):
            dd = payload["due_date"].strip()[:32]
            if dd and not _crm_validate_date_ymd(dd):
                return jsonify({"error": "due_date phải YYYY-MM-DD"}), 400
            merged["due_date"] = dd
        if "status" in payload:
            merged["status"] = _sop_normalize_task_status(payload.get("status"))
        if "assigned_staff_id" in payload:
            ex = _opt_pos_int(payload.get("assigned_staff_id"))
            if ex and conn.execute("SELECT id FROM crm_staff WHERE id = ?", (ex,)).fetchone() is None:
                return jsonify({"error": "Nhân viên không tồn tại"}), 404
            merged["assigned_staff_id"] = ex
        if "notes" in payload and isinstance(payload["notes"], str):
            merged["notes"] = payload["notes"].strip()[:8000]
        if "checklist_json" in payload and isinstance(payload["checklist_json"], list):
            import json as _json
            merged["checklist_json"] = _json.dumps(payload["checklist_json"], ensure_ascii=False)[:4000]
        if "position" in payload:
            try:
                merged["position"] = int(payload["position"])
            except (TypeError, ValueError):
                pass
        ts = _crm_ts()
        conn.execute(
            """
            UPDATE crm_sop_run_tasks
            SET title = ?, description = ?, role = ?, due_date = ?, status = ?,
                assigned_staff_id = ?, notes = ?, checklist_json = ?, position = ?, updated_at = ?
            WHERE id = ?
            """,
            (merged["title"], merged["description"], merged["role"],
             merged.get("due_date") or "", merged["status"],
             merged.get("assigned_staff_id"), merged.get("notes") or "",
             merged.get("checklist_json") or "[]",
             int(merged.get("position") or 0), ts, task_id),
        )
        row2 = conn.execute(
            """
            SELECT rt.*, st.name AS assigned_staff_name
            FROM crm_sop_run_tasks rt
            LEFT JOIN crm_staff st ON st.id = rt.assigned_staff_id
            WHERE rt.id = ?
            """,
            (task_id,),
        ).fetchone()
    assert row2 is not None
    return jsonify(dict(row2))


@app.delete("/api/crm/sop/run_tasks/<int:task_id>")
def api_crm_sop_delete_run_task(task_id: int) -> Any:
    with get_connection() as conn:
        cur = conn.execute("DELETE FROM crm_sop_run_tasks WHERE id = ?", (task_id,))
        if cur.rowcount == 0:
            return jsonify({"error": "Không tìm thấy task"}), 404
    return jsonify({"ok": True})


@app.get("/api/crm/sop/overdue_tasks")
def api_crm_sop_overdue_tasks() -> Any:
    today = datetime.now().strftime("%Y-%m-%d")
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT rt.id, rt.title, rt.due_date, rt.status, rt.run_id,
                   r.name AS run_name, st.name AS assigned_staff_name
            FROM crm_sop_run_tasks rt
            JOIN crm_sop_runs r ON r.id = rt.run_id
            LEFT JOIN crm_staff st ON st.id = rt.assigned_staff_id
            WHERE rt.status NOT IN ('done','skipped')
              AND r.status = 'active'
              AND rt.due_date != ''
              AND rt.due_date < ?
            ORDER BY rt.due_date ASC
            LIMIT 100
            """,
            (today,),
        ).fetchall()
    return jsonify({"overdue": rows_to_dict(rows), "today": today})


@app.get("/api/crm/funnel")
def api_crm_funnel() -> Any:
    with get_connection() as conn:
        stats = compute_funnel_stats(conn)
    return jsonify(stats)


@app.get("/api/crm/funnel/live")
def api_crm_funnel_live() -> Any:
    """Endpoint nhẹ cho polling real-time trên Bảng CSKH."""
    with get_connection() as conn:
        stats = compute_funnel_stats(conn)
    return jsonify(
        {
            "generated_at": stats["generated_at"],
            "totals": stats["totals"],
            "stages": stats["stages"],
            "bottlenecks": stats["bottlenecks"],
        }
    )


@app.post("/api/crm/integration/marketing/ingest")
def api_crm_marketing_ingest() -> Any:
    if not _crm_marketing_ingest_allowed():
        return (
            jsonify(
                {
                    "error": "Chưa cấu hình CRM_MARKETING_INGEST_SECRET — không nhận lead từ ngoài.",
                }
            ),
            503,
        )
    if not _crm_marketing_verify_bearer():
        return jsonify({"error": "Không xác thực (Bearer token)."}), 401
    payload = request.get_json(force=True) or {}
    name = str(payload.get("name", "")).strip()[:240]
    phone = str(payload.get("phone", "")).strip()[:80]
    email = str(payload.get("email", "")).strip()[:240]
    address = str(payload.get("address", "")).strip()[:500]
    company = str(payload.get("company", "")).strip()[:400]
    if not name:
        return jsonify({"error": "Thiếu name"}), 400
    if not phone and not email:
        return jsonify({"error": "Cần phone hoặc email"}), 400
    title = str(payload.get("title", "")).strip()[:800] or f"Lead marketing — {name}"
    desc = str(payload.get("message") or payload.get("notes") or payload.get("body") or "").strip()[:8000]
    utm = str(payload.get("utm_campaign") or payload.get("campaign_code") or "").strip()
    ingest_site = str(payload.get("ingest_site") or payload.get("site") or "").strip()
    re_project_id_raw = payload.get("re_project_id")
    re_project_id: int | None = None
    if re_project_id_raw is not None and str(re_project_id_raw).strip() != "":
        try:
            re_project_id = int(re_project_id_raw)
        except (TypeError, ValueError):
            return jsonify({"error": "re_project_id không hợp lệ"}), 400
    re_project_code = str(payload.get("re_project_code") or payload.get("project_code") or "").strip()
    channel = str(payload.get("channel") or "khac").strip()
    ts = _crm_ts()
    short_date = datetime.now().strftime("%Y-%m-%d")
    with get_connection() as conn:
        channel = _crm_resolve_lead_channel(conn, channel)
        ch_labels = _crm_lead_channel_labels_map(conn)
        campaign_id_final: int | None = _crm_lookup_campaign_by_code(conn, utm) if utm else None
        cust_id = _crm_find_existing_customer(conn, phone, email)
        if cust_id is None:
            cur_cu = conn.execute(
                """
                INSERT INTO crm_customers (name, phone, email, address, company, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (name, phone, email, address, company, short_date),
            )
            cust_id = int(cur_cu.lastrowid)
        else:
            ex_row = conn.execute("SELECT * FROM crm_customers WHERE id = ?", (cust_id,)).fetchone()
            assert ex_row is not None
            exd = dict(ex_row)
            nm = str(exd.get("name") or "").strip() or name
            ph = str(exd.get("phone") or "").strip() or phone
            em = str(exd.get("email") or "").strip() or email
            ad = str(exd.get("address") or "").strip() or address
            co = str(exd.get("company") or "").strip() or company
            conn.execute(
                """
                UPDATE crm_customers
                SET name = ?, phone = ?, email = ?, address = ?, company = ?
                WHERE id = ?
                """,
                (nm[:240], ph[:80], em[:240], ad[:500], co[:400], cust_id),
            )
        cur_case = conn.execute(
            """
            INSERT INTO crm_cases (
                customer_id, title, description, channel, priority, status,
                assigned_to, assigned_staff_id, assigned_at, created_at, updated_at, campaign_id
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                cust_id,
                title,
                desc,
                channel,
                "binh_thuong",
                "tiep_nhan",
                "",
                None,
                "",
                ts,
                ts,
                campaign_id_final,
            ),
        )
        cid = int(cur_case.lastrowid)
        src = utm or "—"
        lead_src = str(payload.get("lead_source") or payload.get("source") or "marketing").strip()[:120]
        _crm_append_event(
            conn,
            cid,
            "ghi_chu",
            f"Lead đồng bộ từ marketing (UTM/campaign: {src}).",
        )
        ato, aid = on_case_created(
            conn,
            cid,
            title=title,
            priority=str(payload.get("priority") or "binh_thuong"),
            assigned_staff_id=_opt_pos_int(payload.get("assigned_staff_id")),
            assigned_to=str(payload.get("assigned_to") or "").strip(),
            lead_source=lead_src,
            auto_assign=payload.get("auto_assign", True) is not False,
        )
        if aid:
            _crm_append_event(
                conn,
                cid,
                "phan_cong",
                f"Phân công tự động: {ato}.",
            )
        joined = conn.execute(f"{_CRM_CASE_SELECT} WHERE c.id = ?", (cid,)).fetchone()
        lead_id = _crm_ingest_lead_from_form(
            conn,
            full_name=name,
            phone=phone,
            email=email,
            need=desc,
            source=str(payload.get("lead_source") or payload.get("source") or "api"),
            region=str(payload.get("region") or ""),
            product_interest=str(payload.get("product_interest") or payload.get("product") or ""),
            utm_campaign=utm,
            re_project_id=re_project_id,
            re_project_code=re_project_code or None,
            ingest_site=ingest_site,
            ts=ts,
        )
        conn.commit()
    assert joined is not None
    out = _crm_row_case(joined, ch_labels)
    if lead_id is not None:
        out["lead_id"] = int(lead_id)
    return jsonify(out), 201


def _staff_csv_cell(norm: dict[str, str], header: str) -> str:
    return str(norm.get(header) or "").strip()


def _staff_import_resolve_department_id(
    conn: sqlite3.Connection, dept_name: str, dept_code: str
) -> int | None:
    dc = dept_code.strip()
    if dc and dc != "—":
        hit = conn.execute(
            "SELECT id FROM crm_departments WHERE lower(trim(code)) = lower(?) AND active = 1",
            (dc,),
        ).fetchone()
        if hit is not None:
            return int(hit["id"])
    dn = dept_name.strip()
    if dn and dn != "—":
        hit = conn.execute(
            "SELECT id FROM crm_departments WHERE lower(trim(name)) = lower(?) AND active = 1",
            (dn,),
        ).fetchone()
        if hit is not None:
            return int(hit["id"])
    return None


def _staff_import_find_existing_id(
    conn: sqlite3.Connection, internal_code: str, email: str
) -> int | None:
    ic = internal_code.strip()
    if ic and ic != "—":
        hit = conn.execute(
            """
            SELECT id FROM crm_staff
            WHERE lower(trim(internal_code)) = lower(?) AND trim(internal_code) != ''
            """,
            (ic,),
        ).fetchone()
        if hit is not None:
            return int(hit["id"])
    em = email.strip()
    if em and em != "—":
        hit = conn.execute(
            """
            SELECT id FROM crm_staff
            WHERE lower(trim(email)) = lower(?) AND trim(email) != ''
            """,
            (em,),
        ).fetchone()
        if hit is not None:
            return int(hit["id"])
    return None


def _staff_import_parse_employment(label: str) -> str:
    t = label.strip()
    if not t or t == "—":
        return "full_time"
    for k, v in CRM_EMPLOYMENT_LABELS_VI.items():
        if v == t:
            return k
    if t in CRM_EMPLOYMENT_TYPES:
        return t
    return "full_time"


def _staff_import_parse_active(label: str) -> bool:
    return str(label).strip() != "Ngưng"


def _staff_import_resolve_manager_id(conn: sqlite3.Connection, mgr_name: str) -> int | None:
    n = str(mgr_name).strip()
    if not n or n == "—":
        return None
    hit = conn.execute(
        "SELECT id FROM crm_staff WHERE active = 1 AND trim(name) = ? LIMIT 1",
        (n,),
    ).fetchone()
    return int(hit["id"]) if hit is not None else None


@app.post("/api/crm/staff/import")
def api_crm_import_staff() -> Any:
    """Nhập CSV trùng cấu trúc file xuất. Cập nhật nếu trùng Mã NV hoặc Email."""
    uf = request.files.get("file")
    if uf is None or not (getattr(uf, "filename", "") or "").strip():
        return jsonify({"error": "Thiếu file (multipart field: file)"}), 400
    raw = uf.read()
    if len(raw) > 2_000_000:
        return jsonify({"error": "File quá lớn (tối đa 2MB)"}), 400
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        return jsonify({"error": "Không đọc được — dùng CSV UTF-8"}), 400

    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames:
        return jsonify({"error": "CSV không có dòng tiêu đề cột"}), 400

    norm_headers = [str(h).strip() for h in reader.fieldnames if h is not None and str(h).strip()]
    if "Họ tên" not in norm_headers:
        return jsonify({"error": "Thiếu cột «Họ tên» — nên xuất CSV từ hệ thống rồi chỉnh."}), 400

    created = 0
    updated = 0
    skipped = 0
    errors: list[dict[str, Any]] = []
    row_num = 1
    ts = _crm_ts()
    short_date = datetime.now().strftime("%Y-%m-%d")

    with get_connection() as conn:
        for raw_row in reader:
            row_num += 1
            if row_num > 2002:
                errors.append({"row": row_num, "error": "Giới hạn 2000 dữ liệu — dừng."})
                break
            norm = {
                str(k).strip(): ("" if v is None else str(v)).strip()
                for k, v in raw_row.items()
                if k is not None and str(k).strip()
            }
            name = _staff_csv_cell(norm, "Họ tên")
            if not name:
                skipped += 1
                continue
            try:
                conn.execute("SAVEPOINT staff_csv_row")
                internal_code = _staff_csv_cell(norm, "Mã NV")[:80]
                email = _staff_csv_cell(norm, "Email")[:240]
                phone = _staff_csv_cell(norm, "Điện thoại")[:80]
                job_title = _staff_csv_cell(norm, "Chức danh")[:200]
                dept_label = _staff_csv_cell(norm, "Nhãn nhóm")[:200]
                notes = _staff_csv_cell(norm, "Ghi chú nội bộ")[:4000]
                started_on = _staff_csv_cell(norm, "Ngày bắt đầu")[:32]
                ended_on = _staff_csv_cell(norm, "Ngày kết thúc")[:32]
                emp = _staff_import_parse_employment(_staff_csv_cell(norm, "Loại hợp đồng"))
                active = _staff_import_parse_active(_staff_csv_cell(norm, "Trạng thái"))
                dept_id = _staff_import_resolve_department_id(
                    conn,
                    _staff_csv_cell(norm, "Phòng ban (danh mục)"),
                    _staff_csv_cell(norm, "Mã phòng ban"),
                )
                mgr_id = _staff_import_resolve_manager_id(
                    conn, _staff_csv_cell(norm, "Quản lý trực tiếp")
                )

                if email and not _EMAIL_RE.match(email):
                    raise ValueError("Email không hợp lệ")

                ex_id = _staff_import_find_existing_id(conn, internal_code, email)
                if mgr_id is not None and ex_id is not None and mgr_id == ex_id:
                    mgr_id = None

                if ex_id is not None:
                    prev = conn.execute("SELECT * FROM crm_staff WHERE id = ?", (ex_id,)).fetchone()
                    if prev is None:
                        raise ValueError("Bản ghi không tồn tại")
                    prev_d = dict(prev)
                    if dept_id is not None and not _dept_exists_active(conn, dept_id):
                        raise ValueError("Phòng ban không hợp lệ hoặc đã ngưng")
                    if mgr_id is not None:
                        if conn.execute("SELECT id FROM crm_staff WHERE id = ?", (mgr_id,)).fetchone() is None:
                            mgr_id = None
                        elif mgr_id == ex_id:
                            mgr_id = None
                    if email and _staff_field_exists(conn, "email", email, exclude_id=ex_id):
                        raise ValueError("Email trùng nhân viên khác")
                    if internal_code and _staff_field_exists(conn, "internal_code", internal_code, exclude_id=ex_id):
                        raise ValueError("Mã NV trùng nhân viên khác")

                    conn.execute(
                        """
                        UPDATE crm_staff
                        SET name = ?, phone = ?, email = ?, job_title = ?, department = ?,
                            internal_code = ?, notes = ?, active = ?, updated_at = ?,
                            department_id = ?, reports_to_id = ?, employment_type = ?,
                            started_on = ?, ended_on = ?
                        WHERE id = ?
                        """,
                        (
                            name[:240],
                            phone,
                            email,
                            job_title,
                            dept_label,
                            internal_code,
                            notes,
                            1 if active else 0,
                            ts,
                            dept_id,
                            mgr_id,
                            emp,
                            started_on,
                            ended_on,
                            ex_id,
                        ),
                    )
                    if str(prev_d.get("name") or "") != name[:240]:
                        conn.execute(
                            """
                            UPDATE crm_cases SET assigned_to = ?
                            WHERE assigned_staff_id = ?
                            """,
                            (name[:240], ex_id),
                        )
                    if not active:
                        conn.execute(
                            """
                            UPDATE crm_cases
                            SET assigned_staff_id = NULL, assigned_to = '', assigned_at = ''
                            WHERE assigned_staff_id = ?
                            """,
                            (ex_id,),
                        )
                    updated += 1
                else:
                    if email and _staff_field_exists(conn, "email", email, exclude_id=None):
                        raise ValueError("Email đã thuộc nhân viên khác")
                    if internal_code and _staff_field_exists(conn, "internal_code", internal_code, exclude_id=None):
                        raise ValueError("Mã NV đã tồn tại")
                    if dept_id is not None and not _dept_exists_active(conn, dept_id):
                        dept_id = None

                    cur = conn.execute(
                        """
                        INSERT INTO crm_staff (
                            name, phone, email, job_title, department, internal_code, notes,
                            active, sort_order, created_at, updated_at,
                            department_id, reports_to_id, employment_type, started_on, ended_on
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            name[:240],
                            phone,
                            email,
                            job_title,
                            dept_label,
                            internal_code,
                            notes,
                            1 if active else 0,
                            short_date,
                            ts,
                            dept_id,
                            mgr_id,
                            emp,
                            started_on,
                            ended_on,
                        ),
                    )
                    new_id = int(cur.lastrowid)
                    if mgr_id is not None and mgr_id == new_id:
                        conn.execute(
                            "UPDATE crm_staff SET reports_to_id = NULL WHERE id = ?",
                            (new_id,),
                        )
                    created += 1

                conn.execute("RELEASE staff_csv_row")
            except (sqlite3.Error, ValueError) as exc:
                try:
                    conn.execute("ROLLBACK TO staff_csv_row")
                except sqlite3.Error:
                    pass
                errors.append({"row": row_num, "error": str(exc)})

    return jsonify(
        {
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors,
        }
    )


@app.get("/api/crm/channels")
def api_crm_list_lead_channels() -> Any:
    """Danh mục kênh tiếp nhận — dùng dropdown Yêu cầu mới trên Bảng CSKH."""
    raw = (request.args.get("include_inactive") or "").strip().lower()
    include_inactive = raw in ("1", "true", "yes", "all")
    with get_connection() as conn:
        if include_inactive:
            rows = conn.execute(
                """
                SELECT * FROM crm_lead_channels
                ORDER BY active DESC, sort_order ASC, name COLLATE NOCASE ASC
                """
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT * FROM crm_lead_channels
                WHERE active = 1
                ORDER BY sort_order ASC, name COLLATE NOCASE ASC
                """
            ).fetchall()
    return jsonify({"channels": rows_to_dict(rows)})


@app.post("/api/crm/channels")
def api_crm_create_lead_channel() -> Any:
    if not _cms_can("crm_lead_channels", "create"):
        return _cms_forbidden_json("crm_lead_channels", "create")
    payload = request.get_json(force=True) or {}
    name = str(payload.get("name", "")).strip()[:240]
    if not name:
        return jsonify({"error": "Thiếu tên kênh"}), 400
    code = str(payload.get("code") or "").strip()[:32]
    if not code:
        code = _crm_slug_lead_channel_code(name)
    desc = str(payload.get("description", "")).strip()[:2000]
    try:
        so = int(payload.get("sort_order") or 0)
    except (TypeError, ValueError):
        so = 0
    ts_date = datetime.now().strftime("%Y-%m-%d")
    ts = _crm_ts()
    with get_connection() as conn:
        dup = conn.execute(
            """
            SELECT id FROM crm_lead_channels
            WHERE lower(trim(code)) = lower(?) AND trim(code) != ''
            LIMIT 1
            """,
            (code,),
        ).fetchone()
        if dup:
            return jsonify({"error": f"Mã kênh «{code}» đã tồn tại"}), 409
        cur = conn.execute(
            """
            INSERT INTO crm_lead_channels (code, name, description, sort_order, active, created_at, updated_at)
            VALUES (?, ?, ?, ?, 1, ?, ?)
            """,
            (code, name, desc, so, ts_date, ts),
        )
        cid = int(cur.lastrowid)
        row = conn.execute("SELECT * FROM crm_lead_channels WHERE id = ?", (cid,)).fetchone()
    assert row is not None
    return jsonify(dict(row)), 201


@app.patch("/api/crm/channels/<int:channel_id>")
def api_crm_patch_lead_channel(channel_id: int) -> Any:
    if not _cms_can("crm_lead_channels", "edit"):
        return _cms_forbidden_json("crm_lead_channels", "edit")
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute(
            "SELECT * FROM crm_lead_channels WHERE id = ?", (channel_id,)
        ).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy kênh"}), 404
        merged: dict[str, Any] = dict(row)
        if "code" in payload and isinstance(payload["code"], str):
            nc = payload["code"].strip()[:32]
            if nc:
                dup = conn.execute(
                    """
                    SELECT id FROM crm_lead_channels
                    WHERE lower(trim(code)) = lower(?) AND id != ?
                    LIMIT 1
                    """,
                    (nc, channel_id),
                ).fetchone()
                if dup:
                    return jsonify({"error": f"Mã kênh «{nc}» đã tồn tại"}), 409
                merged["code"] = nc
        if "name" in payload and isinstance(payload["name"], str):
            n = payload["name"].strip()[:240]
            if not n:
                return jsonify({"error": "Tên kênh không được trống"}), 400
            merged["name"] = n
        if "description" in payload and isinstance(payload["description"], str):
            merged["description"] = payload["description"].strip()[:2000]
        if "sort_order" in payload and payload["sort_order"] is not None:
            try:
                merged["sort_order"] = int(payload["sort_order"])
            except (TypeError, ValueError):
                pass
        if "active" in payload:
            merged["active"] = 1 if payload["active"] in (True, 1, "1", "true") else 0
        ts = _crm_ts()
        conn.execute(
            """
            UPDATE crm_lead_channels
            SET code = ?, name = ?, description = ?, sort_order = ?, active = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                str(merged.get("code") or ""),
                merged["name"],
                str(merged.get("description") or ""),
                int(merged.get("sort_order") or 0),
                1 if merged.get("active") else 0,
                ts,
                channel_id,
            ),
        )
        row2 = conn.execute(
            "SELECT * FROM crm_lead_channels WHERE id = ?", (channel_id,)
        ).fetchone()
    assert row2 is not None
    return jsonify(dict(row2))


@app.get("/api/crm/departments")
def api_crm_list_departments() -> Any:
    raw = (request.args.get("include_inactive") or "").strip().lower()
    include_inactive = raw in ("1", "true", "yes", "all")
    with get_connection() as conn:
        if include_inactive:
            rows = conn.execute(
                """
                SELECT * FROM crm_departments
                ORDER BY active DESC, sort_order ASC, name COLLATE NOCASE ASC
                """
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT * FROM crm_departments
                WHERE active = 1
                ORDER BY sort_order ASC, name COLLATE NOCASE ASC
                """
            ).fetchall()
    return jsonify({"departments": rows_to_dict(rows)})


@app.post("/api/crm/departments")
def api_crm_create_department() -> Any:
    payload = request.get_json(force=True) or {}
    code = str(payload.get("code", "")).strip()[:32]
    name = str(payload.get("name", "")).strip()[:240]
    if not name:
        return jsonify({"error": "Thiếu tên phòng ban"}), 400
    desc = str(payload.get("description", "")).strip()[:2000]
    try:
        so = int(payload.get("sort_order") or 0)
    except (TypeError, ValueError):
        so = 0
    ts_date = datetime.now().strftime("%Y-%m-%d")
    ts = _crm_ts()
    with get_connection() as conn:
        cur = conn.execute(
            """
            INSERT INTO crm_departments (code, name, description, sort_order, active, created_at, updated_at)
            VALUES (?, ?, ?, ?, 1, ?, ?)
            """,
            (code, name, desc, so, ts_date, ts),
        )
        did = int(cur.lastrowid)
        row = conn.execute("SELECT * FROM crm_departments WHERE id = ?", (did,)).fetchone()
    assert row is not None
    return jsonify(dict(row)), 201


@app.patch("/api/crm/departments/<int:dep_id>")
def api_crm_patch_department(dep_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM crm_departments WHERE id = ?", (dep_id,)).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy phòng ban"}), 404
        prev = dict(row)
        merged: dict[str, Any] = dict(row)
        if "code" in payload and isinstance(payload["code"], str):
            merged["code"] = payload["code"].strip()[:32]
        if "name" in payload and isinstance(payload["name"], str):
            n = payload["name"].strip()[:240]
            if not n:
                return jsonify({"error": "Tên phòng ban không được trống"}), 400
            merged["name"] = n
        if "description" in payload and isinstance(payload["description"], str):
            merged["description"] = payload["description"].strip()[:2000]
        if "sort_order" in payload and payload["sort_order"] is not None:
            try:
                merged["sort_order"] = int(payload["sort_order"])
            except (TypeError, ValueError):
                pass
        if "active" in payload:
            merged["active"] = 1 if payload["active"] in (True, 1, "1", "true") else 0

        ts = _crm_ts()

        if merged.get("active") == 0 and int(prev.get("active") or 0) == 1:
            conn.execute(
                "UPDATE crm_staff SET department_id = NULL WHERE department_id = ?",
                (dep_id,),
            )

        conn.execute(
            """
            UPDATE crm_departments
            SET code = ?, name = ?, description = ?, sort_order = ?, active = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                str(merged.get("code") or ""),
                merged["name"],
                str(merged.get("description") or ""),
                int(merged.get("sort_order") or 0),
                1 if merged.get("active") else 0,
                ts,
                dep_id,
            ),
        )
        row2 = conn.execute("SELECT * FROM crm_departments WHERE id = ?", (dep_id,)).fetchone()
    assert row2 is not None
    return jsonify(dict(row2))


@app.get("/api/crm/positions")
def api_crm_list_positions() -> Any:
    raw = (request.args.get("include_inactive") or "").strip().lower()
    include_inactive = raw in ("1", "true", "yes", "all")
    with get_connection() as conn:
        if include_inactive:
            rows = conn.execute(
                """
                SELECT * FROM crm_positions
                ORDER BY active DESC, sort_order ASC, name COLLATE NOCASE ASC
                """
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT * FROM crm_positions
                WHERE active = 1
                ORDER BY sort_order ASC, name COLLATE NOCASE ASC
                """
            ).fetchall()
    return jsonify({"positions": rows_to_dict(rows)})


@app.post("/api/crm/positions")
def api_crm_create_position() -> Any:
    payload = request.get_json(force=True) or {}
    code = str(payload.get("code", "")).strip()[:32]
    name = str(payload.get("name", "")).strip()[:240]
    if not name:
        return jsonify({"error": "Thiếu tên chức vụ"}), 400
    desc = str(payload.get("description", "")).strip()[:2000]
    try:
        so = int(payload.get("sort_order") or 0)
    except (TypeError, ValueError):
        so = 0
    ts_date = datetime.now().strftime("%Y-%m-%d")
    ts = _crm_ts()
    with get_connection() as conn:
        cur = conn.execute(
            """
            INSERT INTO crm_positions (code, name, description, sort_order, active, created_at, updated_at)
            VALUES (?, ?, ?, ?, 1, ?, ?)
            """,
            (code, name, desc, so, ts_date, ts),
        )
        pid = int(cur.lastrowid)
        row = conn.execute("SELECT * FROM crm_positions WHERE id = ?", (pid,)).fetchone()
    assert row is not None
    return jsonify(dict(row)), 201


@app.patch("/api/crm/positions/<int:position_id>")
def api_crm_patch_position(position_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM crm_positions WHERE id = ?", (position_id,)).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy chức vụ"}), 404
        prev = dict(row)
        merged: dict[str, Any] = dict(row)
        if "code" in payload and isinstance(payload["code"], str):
            merged["code"] = payload["code"].strip()[:32]
        if "name" in payload and isinstance(payload["name"], str):
            n = payload["name"].strip()[:240]
            if not n:
                return jsonify({"error": "Tên chức vụ không được trống"}), 400
            merged["name"] = n
        if "description" in payload and isinstance(payload["description"], str):
            merged["description"] = payload["description"].strip()[:2000]
        if "sort_order" in payload and payload["sort_order"] is not None:
            try:
                merged["sort_order"] = int(payload["sort_order"])
            except (TypeError, ValueError):
                pass
        if "active" in payload:
            merged["active"] = 1 if payload["active"] in (True, 1, "1", "true") else 0

        ts = _crm_ts()

        if merged.get("active") == 0 and int(prev.get("active") or 0) == 1:
            conn.execute(
                "UPDATE crm_staff SET position_id = NULL WHERE position_id = ?",
                (position_id,),
            )

        conn.execute(
            """
            UPDATE crm_positions
            SET code = ?, name = ?, description = ?, sort_order = ?, active = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                str(merged.get("code") or ""),
                merged["name"],
                str(merged.get("description") or ""),
                int(merged.get("sort_order") or 0),
                1 if merged.get("active") else 0,
                ts,
                position_id,
            ),
        )
        row2 = conn.execute("SELECT * FROM crm_positions WHERE id = ?", (position_id,)).fetchone()
    assert row2 is not None
    return jsonify(dict(row2))

@app.get("/api/crm/staff/<int:staff_id>/workspace")
def api_crm_staff_workspace(staff_id: int) -> Any:
    """Tổng quan khách được gán và báo cáo CSKH của một nhân viên."""
    portal_sid = _crm_effective_staff_id()
    if portal_sid is not None and portal_sid != staff_id:
        return _crm_forbid_staff_case()
    with get_connection() as conn:
        staff = conn.execute(
            "SELECT id, name, phone, email, job_title, department, active FROM crm_staff WHERE id = ?",
            (staff_id,),
        ).fetchone()
        if staff is None:
            return jsonify({"error": "Không tìm thấy nhân viên"}), 404
        ch_labels = _crm_lead_channel_labels_map(conn)
        rows = conn.execute(
            f"""
            {_CRM_CASE_SELECT}
            WHERE c.assigned_staff_id = ?
            ORDER BY datetime(c.updated_at) DESC
            """,
            (staff_id,),
        ).fetchall()
        case_ids = [int(r["id"]) for r in rows]
        last_care = fetch_last_care_reports_map(conn, case_ids)
        recent_reports = conn.execute(
            """
            SELECT r.*, c.title AS case_title, cu.name AS customer_name
            FROM crm_care_reports r
            JOIN crm_cases c ON c.id = r.case_id
            JOIN crm_customers cu ON cu.id = c.customer_id
            WHERE r.staff_id = ? OR c.assigned_staff_id = ?
            ORDER BY r.id DESC
            LIMIT 15
            """,
            (staff_id, staff_id),
        ).fetchall()

    cases = [_crm_row_case(r, ch_labels) for r in rows]
    open_cases = [
        c
        for c in cases
        if normalize_pipeline_stage(c.get("pipeline_stage") or c.get("status"))
        not in TERMINAL_STAGES
    ]
    today_prefix = datetime.now().strftime("%Y-%m-%d")
    stats = {
        "total_assigned": len(open_cases),
        "open": len(open_cases),
        "high_priority": sum(1 for c in open_cases if c.get("priority") == "cao"),
        "sla_overdue": sum(1 for c in open_cases if c.get("sla_overdue")),
        "new_today": sum(
            1 for c in open_cases if str(c.get("assigned_at") or "").startswith(today_prefix)
        ),
        "no_care_report": sum(1 for c in open_cases if int(c["id"]) not in last_care),
    }
    return jsonify(
        {
            "staff": dict(staff),
            "stats": stats,
            "cases": open_cases if portal_sid is not None else cases,
            "recent_care_reports": [care_report_row_to_dict(r) for r in recent_reports],
        }
    )


@app.post("/api/crm/cases/<int:case_id>/care-reports")
def api_crm_add_care_report(case_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    summary = str(payload.get("summary") or "").strip()
    if not summary:
        return jsonify({"error": "Nội dung báo cáo không được để trống"}), 400
    if len(summary) > 4000:
        return jsonify({"error": "Báo cáo quá dài"}), 400
    contact_type = normalize_care_contact(payload.get("contact_type"))
    care_status = normalize_care_status(payload.get("care_status"))
    next_action = str(payload.get("next_action") or "").strip()[:800]
    portal_sid = _crm_effective_staff_id()
    staff_id = portal_sid or _opt_pos_int(payload.get("staff_id"))
    ts = _crm_ts()
    with get_connection() as conn:
        case_row = conn.execute(
            f"{_CRM_CASE_SELECT} WHERE c.id = ?",
            (case_id,),
        ).fetchone()
        if case_row is None:
            return jsonify({"error": "Case not found"}), 404
        if portal_sid is not None and not _crm_case_assigned_to_staff(conn, case_id, portal_sid):
            return _crm_forbid_staff_case()
        staff_name = ""
        if staff_id:
            srow = conn.execute(
                "SELECT name FROM crm_staff WHERE id = ? AND active = 1",
                (staff_id,),
            ).fetchone()
            if srow:
                staff_name = str(srow["name"])
            else:
                staff_id = None
        if not staff_id:
            aid = case_row["assigned_staff_id"]
            if aid:
                staff_id = int(aid)
                staff_name = str(case_row["staff_display_name"] or case_row["assigned_to"] or "")

        cur = conn.execute(
            """
            INSERT INTO crm_care_reports (
                case_id, staff_id, staff_name, contact_type, care_status,
                summary, next_action, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                case_id,
                staff_id,
                staff_name,
                contact_type,
                care_status,
                summary,
                next_action,
                ts,
            ),
        )
        rid = int(cur.lastrowid)
        event_body = format_care_event_body(
            contact_type=contact_type,
            care_status=care_status,
            summary=summary,
            next_action=next_action,
            staff_name=staff_name,
        )
        _crm_append_event(conn, case_id, "bao_cao_cskh", event_body)
        conn.execute("UPDATE crm_cases SET updated_at = ? WHERE id = ?", (ts, case_id))
        row = conn.execute(
            "SELECT * FROM crm_care_reports WHERE id = ?",
            (rid,),
        ).fetchone()
    assert row is not None
    return jsonify(care_report_row_to_dict(row)), 201


@app.post("/api/crm/cases/<int:case_id>/events")
def api_crm_add_event(case_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    body = str(payload.get("body", "")).strip()
    if not body:
        return jsonify({"error": "Nội dung ghi chú không được để trống"}), 400
    if len(body) > 8000:
        return jsonify({"error": "Ghi chú quá dài"}), 400
    portal_sid = _crm_effective_staff_id()
    ts = _crm_ts()
    with get_connection() as conn:
        if portal_sid is not None and not _crm_case_assigned_to_staff(conn, case_id, portal_sid):
            return _crm_forbid_staff_case()
        found = conn.execute("SELECT id FROM crm_cases WHERE id = ?", (case_id,)).fetchone()
        if found is None:
            return jsonify({"error": "Case not found"}), 404
        cur = conn.execute(
            """
            INSERT INTO crm_case_events (case_id, kind, body, created_at)
            VALUES (?, ?, ?, ?)
            """,
            (case_id, "ghi_chu", body, ts),
        )
        rid = int(cur.lastrowid)
        conn.execute(
            "UPDATE crm_cases SET updated_at = ? WHERE id = ?",
            (ts, case_id),
        )
        evt = conn.execute(
            "SELECT id, kind, body, created_at FROM crm_case_events WHERE id = ?",
            (rid,),
        ).fetchone()
    assert evt is not None
    return jsonify(dict(evt)), 201


@app.post("/api/landing-contact")
def landing_contact() -> Any:
    """Gửi form liên hệ từ trang landing — lưu Excel + email (cùng pipeline tư vấn)."""
    data = request.get_json(silent=True) or {}
    form_type = str(data.get("form_type") or "").strip().lower()
    full_name = (data.get("name") or data.get("full_name") or "").strip()
    email = (data.get("email") or "").strip()
    phone = (data.get("phone") or "").strip()
    budget = (data.get("budget") or "").strip()
    company = (data.get("company") or "").strip()
    goal = (data.get("goal") or "").strip()
    help_request = (data.get("help_request") or "").strip()
    additional_info = (data.get("additional_info") or "").strip()
    message = (data.get("message") or "").strip()

    if len(full_name) > 200:
        return jsonify({"ok": False, "error": "Họ tên quá dài."}), 400
    if email and not _EMAIL_RE.match(email):
        return jsonify({"ok": False, "error": "Email không hợp lệ."}), 400

    is_quick = form_type == "quick" or (
        form_type != "full"
        and not budget
        and not company
        and not goal
        and not help_request
        and not additional_info
    )
    if is_quick:
        if not full_name or not email:
            return jsonify({"ok": False, "error": "Vui lòng điền họ tên và email."}), 400
        if len(message) > 2000:
            return jsonify({"ok": False, "error": "Nội dung không quá 2000 ký tự."}), 400
        short_goal = (message[:200] + "…") if len(message) > 200 else (message or "—")
        payload = {
            "service_slug": "landing",
            "service_name": "Form liên hệ nhanh (FAB)",
            "full_name": full_name,
            "email": email,
            "phone": phone or "—",
            "budget": "—",
            "company": "—",
            "goal": short_goal,
            "help_request": message or "—",
            "additional_info": "",
        }
        crm_need = message or short_goal
    else:
        missing = [
            label
            for label, val in (
                ("họ tên", full_name),
                ("email", email),
                ("số điện thoại", phone),
                ("ngân sách", budget),
                ("tên công ty", company),
            )
            if not val
        ]
        if missing:
            return jsonify(
                {"ok": False, "error": f"Vui lòng điền đủ: {', '.join(missing)}."}
            ), 400
        for field_name, field_val, limit in (
            ("goal", goal, 500),
            ("help_request", help_request, 500),
            ("additional_info", additional_info, 2000),
        ):
            if len(field_val) > limit:
                return jsonify(
                    {"ok": False, "error": f"{field_name} không quá {limit} ký tự."}
                ), 400
        payload = {
            "service_slug": "landing",
            "service_name": "Form liên hệ (Landing)",
            "full_name": full_name,
            "email": email,
            "phone": phone,
            "budget": budget,
            "company": company,
            "goal": goal or "—",
            "help_request": help_request or "—",
            "additional_info": additional_info,
        }
        crm_need = " | ".join(
            part
            for part in (
                f"Mục tiêu: {goal}" if goal else "",
                f"Hỗ trợ: {help_request}" if help_request else "",
                f"Thêm: {additional_info}" if additional_info else "",
            )
            if part
        ) or "—"
    try:
        append_consultation_to_excel(payload)
    except OSError as exc:
        return jsonify({"ok": False, "error": f"Không lưu được dữ liệu. Thử lại sau. ({exc})"}), 500

    try:
        with get_connection() as conn:
            utm = str(data.get("utm_campaign") or data.get("campaign_code") or "").strip()
            ingest_site = str(data.get("ingest_site") or data.get("site") or "landing-ptt").strip()
            re_project_code = str(data.get("re_project_code") or data.get("project_code") or "").strip()
            re_project_id_raw = data.get("re_project_id")
            re_project_id: int | None = None
            if re_project_id_raw is not None and str(re_project_id_raw).strip() != "":
                try:
                    re_project_id = int(re_project_id_raw)
                except (TypeError, ValueError):
                    re_project_id = None
            _crm_ingest_lead_from_form(
                conn,
                full_name=full_name,
                phone=phone,
                email=email,
                need=crm_need,
                source="website",
                utm_campaign=utm,
                re_project_id=re_project_id,
                re_project_code=re_project_code or None,
                ingest_site=ingest_site,
                ts=_crm_ts(),
            )
            conn.commit()
    except Exception:
        pass

    try:
        send_consultation_email(payload)
    except Exception:
        return jsonify(
            {
                "ok": True,
                "emailed": False,
                "message": (
                    "Đã ghi nhận thông tin. Gửi email thông báo thất bại; "
                    "chúng tôi sẽ liên hệ theo thông tin bạn đã gửi."
                ),
            }
        )
    return jsonify({"ok": True, "emailed": True})


@app.post("/api/consultations")
def create_consultation():
    payload = request.get_json(force=True)

    required_fields = [
        "service_slug",
        "service_name",
        "full_name",
        "email",
        "phone",
        "budget",
        "company",
        "goal",
        "help_request",
    ]
    missing = [field for field in required_fields if not str(payload.get(field, "")).strip()]
    if missing:
        return jsonify({"error": f"Missing fields: {', '.join(missing)}"}), 400

    consultation = {
        key: str(payload.get(key, "")).strip()
        for key in required_fields + ["additional_info"]
    }

    append_consultation_to_excel(consultation)

    try:
        send_consultation_email(consultation)
    except Exception as exc:
        return (
            jsonify(
                {
                    "error": f"Saved to data.xlsx but failed to send email: {exc}",
                    "saved": True,
                    "emailed": False,
                }
            ),
            500,
        )

    return jsonify({"saved": True, "emailed": True})


@app.get("/api/cms/permissions/me")
def cms_permissions_me_api() -> Any:
    with get_connection() as conn:
        grants = _cms_load_role_grants(conn, _cms_session_role())
        role_row = conn.execute(
            "SELECT code, name, description FROM cms_roles WHERE code = ?",
            (_cms_session_role(),),
        ).fetchone()
    return jsonify(
        {
            "username": _cms_session_username(),
            "role": dict(role_row) if role_row else {"code": _cms_session_role()},
            "grants": grants,
            "is_super_admin": _cms_is_super_admin(),
            "can_configure_matrix": _cms_can("permissions_matrix", "configure"),
        }
    )


def _cms_all_role_grants(conn: sqlite3.Connection) -> dict[str, dict[str, list[str]]]:
    out: dict[str, dict[str, list[str]]] = {}
    for role in CMS_ROLES:
        code = role["code"]
        out[code] = _cms_load_role_grants(conn, code)
    return out


def _cms_persist_role_grants(
    conn: sqlite3.Connection, role_code: str, grants: dict[str, list[str]]
) -> None:
    conn.execute("DELETE FROM cms_role_permissions WHERE role_code = ?", (role_code,))
    for mid, acts in grants.items():
        if mid not in CMS_MODULE_IDS:
            continue
        for act in acts:
            if act in CMS_ACTIONS:
                conn.execute(
                    """
                    INSERT INTO cms_role_permissions (role_code, module_id, action)
                    VALUES (?, ?, ?)
                    """,
                    (role_code, mid, act),
                )
    ts = _crm_ts()
    conn.execute(
        "UPDATE cms_roles SET updated_at = ?, grants_customized = 1 WHERE code = ?",
        (ts, role_code),
    )


@app.get("/api/cms/permissions/matrix")
def cms_permissions_matrix_api() -> Any:
    if not _cms_can("permissions_matrix", "view"):
        return _cms_forbidden_json("permissions_matrix", "view")
    with get_connection() as conn:
        matrix = build_permission_matrix(_cms_all_role_grants(conn))
    matrix["current_role"] = _cms_session_role()
    matrix["can_configure"] = _cms_can("permissions_matrix", "configure")
    return jsonify(matrix)


@app.patch("/api/cms/permissions/roles/<role_code>")
def cms_permissions_patch_role_api(role_code: str) -> Any:
    if not _cms_can("permissions_matrix", "configure"):
        return _cms_forbidden_json("permissions_matrix", "configure")
    code = str(role_code or "").strip()
    if code not in {r["code"] for r in CMS_ROLES}:
        return jsonify({"error": "Vai trò không tồn tại."}), 404
    if code == "super_admin" and not _cms_is_super_admin():
        return jsonify({"error": "Chỉ Quản trị hệ thống mới chỉnh vai trò super_admin."}), 403
    body = request.get_json(force=True) or {}
    raw_grants = body.get("grants") if isinstance(body, dict) else body
    grants = parse_grants_payload(raw_grants)
    if grants is None:
        return jsonify({"error": "Payload grants không hợp lệ."}), 400
    if code != "super_admin":
        pm = grants.get("permissions_matrix") or []
        if "configure" in pm:
            return jsonify({"error": "Không thể gán quyền cấu hình ma trận cho vai trò khác super_admin."}), 400
    with get_connection() as conn:
        _cms_persist_role_grants(conn, code, grants)
        conn.commit()
        saved = _cms_load_role_grants(conn, code)
    return jsonify({"role_code": code, "grants": saved, "ok": True})


def _cms_all_position_grants(conn: sqlite3.Connection) -> dict[int, dict[str, list[str]]]:
    out: dict[int, dict[str, list[str]]] = {}
    rows = conn.execute(
        "SELECT id, code FROM crm_positions WHERE active = 1 ORDER BY sort_order ASC, name COLLATE NOCASE ASC"
    ).fetchall()
    for prow in rows:
        pid = int(prow["id"])
        out[pid] = _cms_load_position_grants(conn, pid)
    return out


def _cms_persist_position_grants(
    conn: sqlite3.Connection, position_id: int, grants: dict[str, list[str]]
) -> None:
    conn.execute(
        "DELETE FROM crm_position_section_permissions WHERE position_id = ?",
        (position_id,),
    )
    for sid, acts in grants.items():
        if sid not in ADMIN_CRM_PERMISSION_IDS:
            continue
        for act in acts:
            if act in CMS_ACTIONS:
                conn.execute(
                    """
                    INSERT INTO crm_position_section_permissions
                    (position_id, section_id, action)
                    VALUES (?, ?, ?)
                    """,
                    (position_id, sid, act),
                )


@app.get("/api/cms/permissions/positions/matrix")
def cms_position_permissions_matrix_api() -> Any:
    if not _cms_can("permissions_matrix", "view"):
        return _cms_forbidden_json("permissions_matrix", "view")
    with get_connection() as conn:
        pos_rows = conn.execute(
            """
            SELECT id, code, name, description FROM crm_positions
            WHERE active = 1
            ORDER BY sort_order ASC, name COLLATE NOCASE ASC
            """
        ).fetchall()
        matrix = build_position_permission_matrix(
            [dict(r) for r in pos_rows],
            _cms_all_position_grants(conn),
        )
    matrix["can_configure"] = _cms_can("permissions_matrix", "configure")
    return jsonify(matrix)


@app.patch("/api/cms/permissions/positions/<int:position_id>")
def cms_position_permissions_patch_api(position_id: int) -> Any:
    if not _cms_can("permissions_matrix", "configure"):
        return _cms_forbidden_json("permissions_matrix", "configure")
    if position_id <= 0:
        return jsonify({"error": "position_id không hợp lệ"}), 400
    body = request.get_json(force=True) or {}
    raw_grants = body.get("grants") if isinstance(body, dict) else body
    grants = parse_position_grants_payload(raw_grants)
    if grants is None:
        return jsonify({"error": "Payload grants không hợp lệ."}), 400
    with get_connection() as conn:
        if conn.execute(
            "SELECT 1 FROM crm_positions WHERE id = ? AND active = 1",
            (position_id,),
        ).fetchone() is None:
            return jsonify({"error": "Không tìm thấy chức vụ."}), 404
        _cms_persist_position_grants(conn, position_id, grants)
        mark_position_grants_customized(conn, position_id, ts=_crm_ts())
        conn.commit()
        saved = _cms_load_position_grants(conn, position_id)
    return jsonify(
        {
            "position_id": position_id,
            "grants": saved,
            "grants_rows": grants_map_to_rows(saved),
            "ok": True,
        }
    )


@app.get("/api/cms/admin-users")
def cms_admin_users_list_api() -> Any:
    if not _cms_can("permissions_matrix", "view"):
        return _cms_forbidden_json("permissions_matrix", "view")
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT u.id, u.username, u.display_name, u.role_code, u.active,
                   u.position_id, u.created_at, u.updated_at,
                   r.name AS role_name, p.name AS position_name, p.code AS position_code
            FROM cms_admin_users u
            LEFT JOIN cms_roles r ON r.code = u.role_code
            LEFT JOIN crm_positions p ON p.id = u.position_id
            ORDER BY u.active DESC, u.username COLLATE NOCASE ASC
            """
        ).fetchall()
        roles = conn.execute(
            "SELECT code, name FROM cms_roles ORDER BY name COLLATE NOCASE ASC"
        ).fetchall()
        positions = conn.execute(
            """
            SELECT id, code, name FROM crm_positions
            WHERE active = 1 ORDER BY sort_order ASC, name COLLATE NOCASE ASC
            """
        ).fetchall()
    return jsonify({
        "users": rows_to_dict(rows),
        "roles": rows_to_dict(roles),
        "positions": rows_to_dict(positions),
    })


@app.post("/api/cms/admin-users")
def cms_admin_users_create_api() -> Any:
    if not _cms_can("permissions_matrix", "configure"):
        return _cms_forbidden_json("permissions_matrix", "configure")
    body = request.get_json(force=True) or {}
    username = str(body.get("username") or "").strip()
    if not username:
        return jsonify({"error": "Thiếu username."}), 400
    role_code = str(body.get("role_code") or "viewer").strip()
    if role_code not in {r["code"] for r in CMS_ROLES}:
        return jsonify({"error": "Vai trò không hợp lệ."}), 400
    if role_code == "super_admin" and not _cms_is_super_admin():
        return jsonify({"error": "Chỉ super_admin mới gán vai trò Quản trị hệ thống."}), 403
    display_name = str(body.get("display_name") or username).strip()[:120]
    position_id: int | None = None
    if "position_id" in body:
        try:
            pid = int(body.get("position_id") or 0)
            position_id = pid if pid > 0 else None
        except (TypeError, ValueError):
            return jsonify({"error": "position_id không hợp lệ."}), 400
    ts_date = datetime.now().strftime("%Y-%m-%d")
    ts = _crm_ts()
    with get_connection() as conn:
        if position_id is not None and conn.execute(
            "SELECT 1 FROM crm_positions WHERE id = ? AND active = 1", (position_id,)
        ).fetchone() is None:
            return jsonify({"error": "Chức vụ không hợp lệ."}), 400
        try:
            cur = conn.execute(
                """
                INSERT INTO cms_admin_users (
                    username, display_name, role_code, position_id, active, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, 1, ?, ?)
                """,
                (username, display_name, role_code, position_id, ts_date, ts),
            )
        except sqlite3.IntegrityError:
            return jsonify({"error": f"Username «{username}» đã tồn tại."}), 409
        uid = int(cur.lastrowid)
        pw = str(body.get("password") or body.get("login_password") or "").strip()
        if pw:
            try:
                set_unified_password(conn, username, pw, updated_at=ts)
            except ValueError as exc:
                return jsonify({"error": str(exc)}), 400
        row = conn.execute(
            """
            SELECT u.id, u.username, u.display_name, u.role_code, u.active,
                   u.created_at, u.updated_at, r.name AS role_name
            FROM cms_admin_users u
            LEFT JOIN cms_roles r ON r.code = u.role_code
            WHERE u.id = ?
            """,
            (uid,),
        ).fetchone()
    return jsonify(dict(row)), 201


@app.patch("/api/cms/admin-users/<int:user_id>")
def cms_admin_users_patch_api(user_id: int) -> Any:
    if not _cms_can("permissions_matrix", "configure"):
        return _cms_forbidden_json("permissions_matrix", "configure")
    body = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute(
            "SELECT * FROM cms_admin_users WHERE id = ?", (user_id,)
        ).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy user."}), 404
        role_code = str(body.get("role_code") or row["role_code"]).strip()
        if role_code not in {r["code"] for r in CMS_ROLES}:
            return jsonify({"error": "Vai trò không hợp lệ."}), 400
        if role_code == "super_admin" and not _cms_is_super_admin():
            return jsonify({"error": "Chỉ super_admin mới gán vai trò Quản trị hệ thống."}), 403
        display_name = str(body.get("display_name") or row["display_name"]).strip()[:120]
        active = row["active"]
        if "active" in body:
            active = 1 if body.get("active") else 0
        position_id = row["position_id"]
        if "position_id" in body:
            try:
                pid = int(body.get("position_id") or 0)
                position_id = pid if pid > 0 else None
            except (TypeError, ValueError):
                return jsonify({"error": "position_id không hợp lệ."}), 400
            if position_id is not None and conn.execute(
                "SELECT 1 FROM crm_positions WHERE id = ? AND active = 1", (position_id,)
            ).fetchone() is None:
                return jsonify({"error": "Chức vụ không hợp lệ."}), 400
        ts = _crm_ts()
        pw = str(body.get("password") or body.get("login_password") or "").strip()
        if pw:
            try:
                set_unified_password(conn, str(row["username"]), pw, updated_at=ts)
            except ValueError as exc:
                return jsonify({"error": str(exc)}), 400
        conn.execute(
            """
            UPDATE cms_admin_users
            SET display_name = ?, role_code = ?, position_id = ?, active = ?, updated_at = ?
            WHERE id = ?
            """,
            (display_name, role_code, position_id, active, ts, user_id),
        )
        updated = conn.execute(
            """
            SELECT u.id, u.username, u.display_name, u.role_code, u.active,
                   u.position_id, u.created_at, u.updated_at,
                   r.name AS role_name, p.name AS position_name
            FROM cms_admin_users u
            LEFT JOIN cms_roles r ON r.code = u.role_code
            LEFT JOIN crm_positions p ON p.id = u.position_id
            WHERE u.id = ?
            """,
            (user_id,),
        ).fetchone()
    return jsonify(dict(updated))


@app.get("/api/crm/assistant/config")
def crm_assistant_config_api():
    if not _admin_section_can("crm_assistant", "view"):
        return _admin_section_forbidden_json("crm_assistant", "view")
    settings = fetch_settings()
    staff_id = _crm_effective_staff_id()
    with get_connection() as conn:
        ctx = fetch_crm_context(conn, staff_id=staff_id)
    resp = jsonify(build_crm_assistant_config(settings, context=ctx))
    resp.headers["Cache-Control"] = "no-store"
    return resp


@app.post("/api/crm/assistant/send")
def crm_assistant_send_api():
    if not _admin_section_can("crm_assistant", "create"):
        return _admin_section_forbidden_json("crm_assistant", "create")
    body = request.get_json(silent=True) or {}
    text = str(body.get("text", "")).strip()
    if not text:
        return jsonify({"error": "Thiếu nội dung tin nhắn."}), 400
    raw_history = body.get("messages")
    history: list[dict[str, Any]] = []
    if isinstance(raw_history, list):
        for item in raw_history[-20:]:
            if not isinstance(item, dict):
                continue
            role = str(item.get("role") or "")
            msg_text = str(item.get("text") or "").strip()
            if role in ("user", "assistant", "bot") and msg_text:
                history.append({"role": role, "text": msg_text[:4000]})
    case_id_raw = body.get("case_id")
    case_id: int | None = None
    if case_id_raw is not None:
        try:
            case_id = int(case_id_raw)
        except (TypeError, ValueError):
            case_id = None
    settings = fetch_settings()
    if not build_crm_assistant_config(settings).get("enabled"):
        return jsonify({"error": "Trợ lý CRM đang tắt trong cấu hình."}), 403
    staff_id = _crm_effective_staff_id()
    can_payroll_file = (
        _admin_section_can("crm_payroll_salary", "export")
        or _admin_section_can("crm_payroll_salary", "view")
        or _admin_section_can("crm_payroll_salary", "edit")
    )
    with get_connection() as conn:
        if case_id is not None and staff_id is not None:
            if not _crm_case_assigned_to_staff(conn, case_id, staff_id):
                return jsonify({"error": "Bạn chỉ hỏi trợ lý về case được gán cho mình."}), 403
        ctx = fetch_crm_context(conn, staff_id=staff_id, case_id=case_id)
        result = build_crm_assistant_response(
            text,
            history,
            settings,
            ctx,
            conn,
            portal_staff_id=staff_id,
            can_payroll_file=can_payroll_file,
        )
    return jsonify({"ok": True, "reply": result["reply"], "files": result.get("files") or []})


@app.post("/api/crm/assistant/export")
def crm_assistant_export_api():
    if not _admin_section_can("crm_assistant", "export"):
        return _admin_section_forbidden_json("crm_assistant", "export")
    body = request.get_json(silent=True) or {}
    fmt = str(body.get("format") or "md").strip().lower()
    raw = body.get("messages")
    messages: list[dict[str, Any]] = []
    if isinstance(raw, list):
        for item in raw[-80:]:
            if not isinstance(item, dict):
                continue
            role = str(item.get("role") or "")
            msg_text = str(item.get("text") or "").strip()
            if role in ("user", "assistant", "bot") and msg_text:
                messages.append({"role": role, "text": msg_text})
    if not messages:
        return jsonify({"error": "Không có nội dung để xuất."}), 400
    settings = fetch_settings()
    brand = str(settings.get("brand_name") or "PTT Advertising Solutions").strip()
    stamp = datetime.now().strftime("%Y%m%d-%H%M")
    if fmt == "json":
        payload = json.dumps({"brand": brand, "messages": messages}, ensure_ascii=False, indent=2)
        return Response(
            payload,
            mimetype="application/json; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="ptt-crm-assistant-{stamp}.json"'},
        )
    if fmt == "html":
        payload = build_crm_export_html(messages, brand=brand)
        return Response(
            payload,
            mimetype="text/html; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="ptt-crm-assistant-{stamp}.html"'},
        )
    payload = build_crm_export_markdown(messages, brand=brand)
    return Response(
        payload,
        mimetype="text/markdown; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="ptt-crm-assistant-{stamp}.md"'},
    )


@app.get("/api/cms/marketing-chat/config")
def cms_marketing_chat_config_api():
    if not _cms_can("mk_chat_config", "view"):
        return _cms_forbidden_json("mk_chat_config", "view")
    settings = {**DEFAULT_SETTINGS_NAV_MEGA_BANNER, **DEFAULT_SETTINGS_FOOTER, **fetch_settings()}
    resp = jsonify(build_marketing_chat_config(settings))
    resp.headers["Cache-Control"] = "no-store"
    return resp


@app.post("/api/cms/marketing-chat/send")
def cms_marketing_chat_send_api():
    if not _cms_can("mk_chat_conversation", "create"):
        return _cms_forbidden_json("mk_chat_conversation", "create")
    body = request.get_json(silent=True) or {}
    text = str(body.get("text", "")).strip()
    if not text:
        return jsonify({"error": "Thiếu nội dung tin nhắn."}), 400
    raw_history = body.get("messages")
    history: list[dict[str, Any]] = []
    if isinstance(raw_history, list):
        for item in raw_history[-20:]:
            if not isinstance(item, dict):
                continue
            role = str(item.get("role") or "")
            msg_text = str(item.get("text") or "").strip()
            if role in ("user", "assistant", "bot") and msg_text:
                history.append({"role": role, "text": msg_text[:4000]})
    settings = {**DEFAULT_SETTINGS_NAV_MEGA_BANNER, **DEFAULT_SETTINGS_FOOTER, **fetch_settings()}
    if not build_marketing_chat_config(settings).get("enabled"):
        return jsonify({"error": "Chatbox marketing đang tắt trong cấu hình CMS."}), 403
    reply = build_marketing_strategy_reply(text, history, settings)
    return jsonify({"ok": True, "reply": reply})


@app.post("/api/cms/marketing-chat/export")
def cms_marketing_chat_export_api():
    if not _cms_can("mk_chat_export", "export"):
        return _cms_forbidden_json("mk_chat_export", "export")
    body = request.get_json(silent=True) or {}
    fmt = str(body.get("format") or "md").strip().lower()
    raw = body.get("messages")
    messages: list[dict[str, Any]] = []
    if isinstance(raw, list):
        for item in raw[-80:]:
            if not isinstance(item, dict):
                continue
            role = str(item.get("role") or "")
            msg_text = str(item.get("text") or "").strip()
            if role in ("user", "assistant", "bot") and msg_text:
                messages.append({"role": role, "text": msg_text})
    if not messages:
        return jsonify({"error": "Không có nội dung để xuất."}), 400
    settings = {**DEFAULT_SETTINGS_NAV_MEGA_BANNER, **DEFAULT_SETTINGS_FOOTER, **fetch_settings()}
    brand = str(settings.get("brand_name") or "PTT Advertising Solutions").strip()
    stamp = datetime.now().strftime("%Y%m%d-%H%M")

    if fmt == "json":
        payload = json.dumps({"brand": brand, "messages": messages}, ensure_ascii=False, indent=2)
        return Response(
            payload,
            mimetype="application/json; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="ptt-marketing-chat-{stamp}.json"'},
        )
    if fmt == "html":
        payload = build_export_html(messages, brand=brand)
        return Response(
            payload,
            mimetype="text/html; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="ptt-marketing-chat-{stamp}.html"'},
        )
    payload = build_export_markdown(messages, brand=brand)
    return Response(
        payload,
        mimetype="text/markdown; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="ptt-marketing-chat-{stamp}.md"'},
    )


@app.post("/api/cms/marketing-chat/campaign-kit")
def cms_marketing_campaign_kit_api():
    if not _cms_can("mk_chat_campaign_kit", "create"):
        return _cms_forbidden_json("mk_chat_campaign_kit", "create")
    body = request.get_json(silent=True) or {}
    brief = str(body.get("brief") or "").strip()
    if not brief:
        return jsonify({"error": "Thiếu mô tả dự án / chiến dịch."}), 400
    if len(brief) > 4000:
        brief = brief[:4000]
    force = bool(body.get("force"))
    settings = {**DEFAULT_SETTINGS_NAV_MEGA_BANNER, **DEFAULT_SETTINGS_FOOTER, **fetch_settings()}
    if not build_marketing_chat_config(settings).get("enabled"):
        return jsonify({"error": "Chatbox marketing đang tắt trong cấu hình CMS."}), 403
    with get_connection() as conn:
        ptt_projects = rows_to_dict(
            conn.execute("SELECT id, title, category, description FROM projects ORDER BY id DESC").fetchall()
        )
    result = generate_campaign_kit(brief, settings, ptt_projects=ptt_projects, force=force)
    return jsonify(result)


@app.get("/api/cms/marketing-chat/campaign-kit/<kit_id>/khmkt.xlsx")
def cms_marketing_campaign_kit_khmkt(kit_id: str):
    if not _cms_can("mk_chat_campaign_kit", "export"):
        return _cms_forbidden_json("mk_chat_campaign_kit", "export")
    plan = get_campaign_kit(kit_id)
    if not plan:
        return jsonify({"error": "Bộ tài liệu đã hết hạn hoặc không tồn tại. Tạo lại từ chatbox."}), 404
    settings = {**DEFAULT_SETTINGS_NAV_MEGA_BANNER, **DEFAULT_SETTINGS_FOOTER, **fetch_settings()}
    brand = str(settings.get("brand_name") or "PTT Advertising Solutions").strip()
    slug = str(plan.get("slug") or "du-an")
    stamp = datetime.now().strftime("%Y%m%d")
    data = build_khmkt_xlsx(plan, brand=brand)
    return Response(
        data,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="KHMKT-{slug}-{stamp}.xlsx"'
        },
    )


@app.get("/api/cms/marketing-chat/campaign-kit/<kit_id>/kpi.xlsx")
def cms_marketing_campaign_kit_kpi(kit_id: str):
    if not _cms_can("mk_chat_campaign_kit", "export"):
        return _cms_forbidden_json("mk_chat_campaign_kit", "export")
    plan = get_campaign_kit(kit_id)
    if not plan:
        return jsonify({"error": "Bộ tài liệu đã hết hạn hoặc không tồn tại. Tạo lại từ chatbox."}), 404
    settings = {**DEFAULT_SETTINGS_NAV_MEGA_BANNER, **DEFAULT_SETTINGS_FOOTER, **fetch_settings()}
    brand = str(settings.get("brand_name") or "PTT Advertising Solutions").strip()
    slug = str(plan.get("slug") or "du-an")
    stamp = datetime.now().strftime("%Y%m%d")
    data = build_campaign_kpi_xlsx(plan, brand=brand)
    return Response(
        data,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="KPI-{slug}-{stamp}.xlsx"'
        },
    )


@app.get("/api/cms/marketing-chat/kpi-strategy.xlsx")
def cms_marketing_chat_kpi_strategy_xlsx():
    if not _cms_can("mk_chat_excel", "export"):
        return _cms_forbidden_json("mk_chat_excel", "export")
    settings = {**DEFAULT_SETTINGS_NAV_MEGA_BANNER, **DEFAULT_SETTINGS_FOOTER, **fetch_settings()}
    brand = str(settings.get("brand_name") or "PTT Advertising Solutions").strip()
    stamp = datetime.now().strftime("%Y%m%d")
    data = build_kpi_strategy_xlsx(brand=brand)
    return Response(
        data,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="ptt-kpi-chien-luoc-marketing-{stamp}.xlsx"'
        },
    )


@app.get("/api/cms/marketing-chat/multichannel-plan.xlsx")
def cms_marketing_chat_multichannel_plan_xlsx():
    if not _cms_can("mk_chat_excel", "export"):
        return _cms_forbidden_json("mk_chat_excel", "export")
    settings = {**DEFAULT_SETTINGS_NAV_MEGA_BANNER, **DEFAULT_SETTINGS_FOOTER, **fetch_settings()}
    brand = str(settings.get("brand_name") or "PTT Advertising Solutions").strip()
    stamp = datetime.now().strftime("%Y%m%d")
    data = build_multichannel_plan_xlsx(brand=brand)
    return Response(
        data,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="ptt-ke-hoach-truyen-thong-da-kenh-{stamp}.xlsx"'
        },
    )


@app.get("/api/cms/marketing-chat/weekly-plan.xlsx")
def cms_marketing_chat_weekly_plan_xlsx():
    if not _cms_can("mk_chat_excel", "export"):
        return _cms_forbidden_json("mk_chat_excel", "export")
    settings = {**DEFAULT_SETTINGS_NAV_MEGA_BANNER, **DEFAULT_SETTINGS_FOOTER, **fetch_settings()}
    brand = str(settings.get("brand_name") or "PTT Advertising Solutions").strip()
    stamp = datetime.now().strftime("%Y%m%d")
    data = build_weekly_marketing_plan_xlsx(brand=brand)
    return Response(
        data,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="ptt-ke-hoach-marketing-tuan-{stamp}.xlsx"'
        },
    )


@app.get("/api/settings")
def get_settings():
    # Gộp default (nav, footer) + DB để form CMS luôn có giá trị ban đầu cho key mới.
    merged = _merged_landing_settings()
    merged["partner_logos_effective_json"] = _partner_logos_effective_json(merged)
    merged["capabilities_items_effective_json"] = _capabilities_items_effective_json(merged)
    return jsonify(merged)


@app.get("/api/services")
def get_services():
    return jsonify(fetch_service_categories())


@app.put("/api/services")
def update_services():
    if not _cms_can("services_builder", "edit"):
        return _cms_forbidden_json("services_builder", "edit")
    payload = request.get_json(force=True)
    is_valid, error_message = validate_service_payload(payload)
    if not is_valid:
        return jsonify({"error": error_message}), 400

    with get_connection() as conn:
        conn.execute(
            """
            INSERT INTO settings (key, value) VALUES (?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """,
            ("service_categories_json", json.dumps(payload, ensure_ascii=False)),
        )
    return jsonify(payload)


@app.put("/api/settings")
def update_settings():
    payload = request.get_json(force=True)
    if not isinstance(payload, dict):
        return jsonify({"error": "Payload must be a JSON object"}), 400

    mk_keys = {k for k in payload if str(k).startswith("cms_mk_chat_")}
    land_keys = set(payload.keys()) - mk_keys
    if land_keys and not _cms_can("landing_settings", "edit"):
        return _cms_forbidden_json("landing_settings", "edit")
    if mk_keys and not _cms_can("mk_chat_config", "edit"):
        return _cms_forbidden_json("mk_chat_config", "edit")

    with get_connection() as conn:
        for key, value in payload.items():
            if str(key) in READONLY_SETTINGS_KEYS:
                continue
            conn.execute(
                """
                INSERT INTO settings (key, value) VALUES (?, ?)
                ON CONFLICT(key) DO UPDATE SET value = excluded.value
                """,
                (str(key), json.dumps(value) if isinstance(value, (dict, list)) else str(value)),
            )

    merged = _merged_landing_settings()
    merged["partner_logos_effective_json"] = _partner_logos_effective_json(merged)
    merged["capabilities_items_effective_json"] = _capabilities_items_effective_json(merged)
    return jsonify(merged)

def _allowed_upload(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_UPLOAD_EXTENSIONS


@app.post("/api/cms/media/upload")
def cms_media_upload():
    if not _cms_can("landing_media", "create"):
        return _cms_forbidden_json("landing_media", "create")
    if "file" not in request.files:
        return jsonify({"error": "Không tìm thấy file"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "Tên file rỗng"}), 400
    if not _allowed_upload(f.filename):
        return jsonify({"error": "Định dạng không hỗ trợ. Dùng: jpg, png, webp, gif, svg, mp4, webm, mov"}), 400
    ext = f.filename.rsplit(".", 1)[1].lower()
    data = f.read()
    max_bytes = MAX_VIDEO_UPLOAD_BYTES if ext in ("mp4", "webm", "mov") else MAX_UPLOAD_BYTES
    if len(data) > max_bytes:
        cap_mb = max_bytes // (1024 * 1024)
        kind = "Video" if ext in ("mp4", "webm", "mov") else "File"
        return jsonify({"error": f"{kind} quá lớn (tối đa {cap_mb} MB)"}), 413
    purpose = (request.form.get("purpose") or request.args.get("purpose") or "").strip().lower()
    img_meta: dict = {}

    if purpose == "hero" and ext in ("jpg", "jpeg", "png", "webp"):
        from cms_media_images import process_hero_variants

        hero_result = process_hero_variants(data, ext)
        if hero_result is not None:
            desktop_bytes, mobile_bytes, img_meta = hero_result
            from cms_media_images import write_hero_upload_files

            paths = write_hero_upload_files(UPLOAD_DIR, desktop_bytes, mobile_bytes)
            payload = {
                **paths,
                "size": len(desktop_bytes),
                "size_mobile": len(mobile_bytes),
                "ext": "webp",
            }
            payload.update(img_meta)
            return jsonify(payload), 201

    if ext in ("jpg", "jpeg", "png", "webp"):
        from cms_media_images import process_image_upload

        data, ext, img_meta = process_image_upload(data, ext, purpose or None)

    if ext in ("mp4", "webm", "mov"):
        from cms_media_video import process_video_upload

        video_result = process_video_upload(data, ext, purpose or None)
        if video_result is not None:
            data, ext, video_meta = video_result
            img_meta.update(video_meta)

    filename = f"{uuid.uuid4().hex}.{ext}"
    dest = UPLOAD_DIR / filename
    dest.write_bytes(data)
    url = f"/static/uploads/{filename}"
    payload = {"url": url, "filename": filename, "size": len(data), "ext": ext}
    if img_meta:
        payload.update(img_meta)
    return jsonify(payload), 201


@app.get("/api/cms/media")
def cms_media_list():
    if not _cms_can("landing_media", "view"):
        return _cms_forbidden_json("landing_media", "view")
    files = []
    for p in sorted(UPLOAD_DIR.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
        if p.is_file() and _allowed_upload(p.name):
            files.append({
                "filename": p.name,
                "url": f"/static/uploads/{p.name}",
                "size": p.stat().st_size,
                "ext": p.suffix.lstrip("."),
            })
    return jsonify(files)


@app.delete("/api/cms/media/<filename>")
def cms_media_delete(filename: str):
    if not _cms_can("landing_media", "delete"):
        return _cms_forbidden_json("landing_media", "delete")
    safe = secure_filename(filename)
    if not safe or not _allowed_upload(safe):
        return jsonify({"error": "Tên file không hợp lệ"}), 400
    dest = UPLOAD_DIR / safe
    if not dest.exists():
        return jsonify({"error": "File không tồn tại"}), 404
    dest.unlink()
    return "", 204


# ── CMS Landing page ─────────────────────────────────────────────────────────

@app.get("/cms/landing")
def cms_landing() -> str:
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    with get_connection() as conn:
        grants = _cms_load_role_grants(conn, _cms_session_role())
        role_row = conn.execute(
            "SELECT name FROM cms_roles WHERE code = ?",
            (_cms_session_role(),),
        ).fetchone()
    if _admin_full_access():
        grants = {mid: list(CMS_ACTIONS) for mid in CMS_MODULE_IDS}
    role_name = str(role_row["name"]) if role_row else _cms_session_role()
    cms_settings = fetch_settings()
    partner_seed = _partner_logos_seed_for_cms(cms_settings)
    capabilities_seed = _capabilities_items_seed_for_cms(cms_settings)
    return render_template(
        "cms_landing.html",
        cms_role_code=_cms_session_role(),
        cms_role_name=role_name,
        cms_username=_cms_session_username(),
        cms_grants_json=json.dumps(grants, ensure_ascii=False),
        partner_logos_seed_json=json.dumps(partner_seed, ensure_ascii=False),
        capabilities_items_seed_json=json.dumps(capabilities_seed, ensure_ascii=False),
        **_admin_page_template_kwargs(),
    )


# ── CMS Recruitment ──────────────────────────────────────────────────────────

@app.get("/api/cms/recruitment")
def api_recruitment_list():
    if not _cms_can("recruitment_jobs", "view"):
        return _cms_forbidden_json("recruitment_jobs", "view")
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT * FROM recruitment_jobs ORDER BY sort_order, id"
        ).fetchall()
    return jsonify([_job_row_to_dict(r) for r in rows])


@app.post("/api/cms/recruitment")
def api_recruitment_create():
    if not _cms_can("recruitment_jobs", "create"):
        return _cms_forbidden_json("recruitment_jobs", "create")
    data = request.get_json(force=True) or {}
    slug = str(data.get("slug", "")).strip()
    title = str(data.get("title", "")).strip()
    if not slug or not title:
        return jsonify({"error": "slug và title là bắt buộc"}), 400
    with get_connection() as conn:
        try:
            conn.execute(
                """INSERT INTO recruitment_jobs
                   (slug, title, location, employment_type, description, intro,
                    responsibilities, requirements, benefits, is_active, sort_order)
                   VALUES (?,?,?,?,?,?,?,?,?,?,
                     COALESCE((SELECT MAX(sort_order)+1 FROM recruitment_jobs),0))""",
                (
                    slug, title,
                    str(data.get("location", "")),
                    str(data.get("employment_type", "Toàn thời gian")),
                    str(data.get("description", "")),
                    str(data.get("intro", "")),
                    json.dumps(data.get("responsibilities", []), ensure_ascii=False),
                    json.dumps(data.get("requirements", []), ensure_ascii=False),
                    json.dumps(data.get("benefits", []), ensure_ascii=False),
                    1 if data.get("is_active", True) else 0,
                ),
            )
            row = conn.execute(
                "SELECT * FROM recruitment_jobs WHERE slug=?", (slug,)
            ).fetchone()
        except Exception as exc:
            return jsonify({"error": str(exc)}), 409
    return jsonify(_job_row_to_dict(row)), 201


@app.put("/api/cms/recruitment/<int:job_id>")
def api_recruitment_update(job_id: int):
    if not _cms_can("recruitment_jobs", "edit"):
        return _cms_forbidden_json("recruitment_jobs", "edit")
    data = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute(
            "SELECT id FROM recruitment_jobs WHERE id=?", (job_id,)
        ).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy"}), 404
        fields: list[str] = []
        values: list[Any] = []
        str_fields = ("slug", "title", "location", "employment_type", "description", "intro")
        for f in str_fields:
            if f in data:
                fields.append(f"{f}=?")
                values.append(str(data[f]))
        for f in ("responsibilities", "requirements", "benefits"):
            if f in data:
                fields.append(f"{f}=?")
                values.append(json.dumps(data[f], ensure_ascii=False))
        if "is_active" in data:
            fields.append("is_active=?")
            values.append(1 if data["is_active"] else 0)
        if "sort_order" in data:
            fields.append("sort_order=?")
            values.append(int(data["sort_order"]))
        if not fields:
            return jsonify({"error": "Không có trường nào để cập nhật"}), 400
        values.append(job_id)
        conn.execute(f"UPDATE recruitment_jobs SET {', '.join(fields)} WHERE id=?", values)
        row = conn.execute("SELECT * FROM recruitment_jobs WHERE id=?", (job_id,)).fetchone()
    return jsonify(_job_row_to_dict(row))


@app.delete("/api/cms/recruitment/<int:job_id>")
def api_recruitment_delete(job_id: int):
    if not _cms_can("recruitment_jobs", "delete"):
        return _cms_forbidden_json("recruitment_jobs", "delete")
    with get_connection() as conn:
        conn.execute("DELETE FROM recruitment_jobs WHERE id=?", (job_id,))
    return "", 204


@app.get("/cms/recruitment")
def cms_recruitment() -> str:
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    with get_connection() as conn:
        grants = _cms_load_role_grants(conn, _cms_session_role())
        role_row = conn.execute(
            "SELECT name FROM cms_roles WHERE code = ?",
            (_cms_session_role(),),
        ).fetchone()
    if _admin_full_access():
        grants = {mid: list(CMS_ACTIONS) for mid in CMS_MODULE_IDS}
    role_name = str(role_row["name"]) if role_row else _cms_session_role()
    return render_template(
        "cms_recruitment.html",
        cms_role_code=_cms_session_role(),
        cms_role_name=role_name,
        cms_username=_cms_session_username(),
        cms_grants_json=json.dumps(grants, ensure_ascii=False),
        **_admin_page_template_kwargs(),
    )


def _sitemap_url_entries() -> list[tuple[str, str, str]]:
    out: list[tuple[str, str, str]] = [
        ("/", "1.0", "weekly"),
        ("/career", "0.8", "weekly"),
        ("/chinh-sach-bao-mat", "0.5", "yearly"),
    ]
    for cat in fetch_service_categories():
        for it in cat.get("items") or []:
            if not isinstance(it, dict):
                continue
            slug = str(it.get("slug", "")).strip()
            if slug:
                out.append((f"/services/{slug}", "0.7", "monthly"))
    return out


@app.get("/robots.txt")
def robots_txt() -> Any:
    from flask import Response

    base = (request.url_root or "").rstrip("/")
    body = (
        "User-agent: *\n"
        "Disallow: /admin\n"
        "Disallow: /cms\n"
        "Disallow: /crm\n"
        "Disallow: /api/\n"
        "\n"
        f"Sitemap: {base}/sitemap.xml\n"
    )
    return Response(body, mimetype="text/plain; charset=utf-8")


@app.get("/sitemap.xml")
def sitemap_xml() -> Any:
    from datetime import date
    from html import escape

    from flask import Response

    base = (request.url_root or "").rstrip("/")
    if not base.startswith("http"):
        base = f"{request.scheme}://{request.host}"
    lastmod = date.today().isoformat()
    parts = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">',
    ]
    for path, priority, changefreq in _sitemap_url_entries():
        loc = f"{base}{path}" if path != "/" else f"{base}/"
        loc_esc = escape(loc, quote=True)
        parts.append(
            "  <url>"
            f"<loc>{loc_esc}</loc><lastmod>{lastmod}</lastmod>"
            f"<changefreq>{changefreq}</changefreq><priority>{priority}</priority>"
            "</url>"
        )
    parts.append("</urlset>")
    return Response("\n".join(parts) + "\n", mimetype="application/xml; charset=utf-8")


@app.get("/healthz")
def healthz_ptt() -> Any:
    """Kiểm tra nhanh đúng tiến trình PTT (tránh nhầm cổng với app Flask khác)."""
    return jsonify({"ok": True, "app": "ptt-landing-flask"})


# ── Live Chat — Public API ────────────────────────────────────────────────────

@app.post("/api/chat/start")
def chat_start() -> Any:
    data = request.get_json(silent=True) or {}
    session_id = str(data.get("session_id") or "").strip() or uuid.uuid4().hex
    page = str(data.get("page") or "").strip()[:200]
    settings_d = fetch_settings()
    welcome = settings_d.get("live_chat_welcome") or "Xin chào! Tôi có thể giúp gì cho bạn?"
    mode = settings_d.get("live_chat_mode") or "ai"
    with get_connection() as conn:
        row = conn.execute(
            "SELECT id FROM chat_conversations WHERE session_id=?", (session_id,)
        ).fetchone()
        if not row:
            conn.execute(
                "INSERT INTO chat_conversations (session_id, visitor_page) VALUES (?,?)",
                (session_id, page),
            )
            conv_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        else:
            conv_id = row[0]
    return jsonify({"ok": True, "session_id": session_id, "conv_id": conv_id,
                    "welcome": welcome, "mode": mode})


@app.post("/api/chat/send")
def chat_send() -> Any:
    data = request.get_json(silent=True) or {}
    session_id = str(data.get("session_id") or "").strip()
    content = str(data.get("message") or "").strip()
    if not session_id or not content:
        return jsonify({"ok": False, "error": "Thiếu session_id hoặc message"}), 400
    if len(content) > 2000:
        return jsonify({"ok": False, "error": "Tin nhắn tối đa 2000 ký tự"}), 400
    settings_d = fetch_settings()
    mode = settings_d.get("live_chat_mode") or "ai"
    with get_connection() as conn:
        row = conn.execute(
            "SELECT id FROM chat_conversations WHERE session_id=?", (session_id,)
        ).fetchone()
        if not row:
            return jsonify({"ok": False, "error": "Phiên không hợp lệ"}), 404
        conv_id = row[0]
        conn.execute(
            "INSERT INTO chat_messages (conversation_id, sender, content) VALUES (?,?,?)",
            (conv_id, "visitor", content),
        )
        conn.execute(
            "UPDATE chat_conversations SET updated_at=datetime('now'), status='open' WHERE id=?",
            (conv_id,),
        )
        last_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        if mode == "ai":
            sys_prompt = settings_d.get("live_chat_ai_prompt") or (
                "Bạn là trợ lý tư vấn của PTT Advertising Solutions — Creative Martech Platform. "
                "Hỗ trợ khách hàng về dịch vụ marketing automation, AI content, data analytics, CRM và paid media. "
                "Trả lời ngắn gọn, thân thiện, chuyên nghiệp bằng tiếng Việt. Không quá 3 đoạn."
            )
            history = conn.execute(
                "SELECT sender, content FROM chat_messages WHERE conversation_id=? ORDER BY id DESC LIMIT 20",
                (conv_id,),
            ).fetchall()
            history = list(reversed(history))
            msgs: list[dict[str, Any]] = [{"role": "system", "content": sys_prompt}]
            for h in history[:-1]:
                msgs.append({"role": "user" if h[0] == "visitor" else "assistant", "content": h[1]})
            msgs.append({"role": "user", "content": content})
            ai_reply = _chat_openai_reply(msgs)
            if ai_reply:
                conn.execute(
                    "INSERT INTO chat_messages (conversation_id, sender, content) VALUES (?,?,?)",
                    (conv_id, "ai", ai_reply),
                )
                reply_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                return jsonify({"ok": True, "reply": ai_reply, "sender": "ai",
                                "msg_id": last_id, "reply_id": reply_id})
    return jsonify({"ok": True, "reply": None, "queued": True, "msg_id": last_id})


@app.get("/api/chat/poll/<session_id>")
def chat_poll(session_id: str) -> Any:
    try:
        since_id = int(request.args.get("since", "0"))
    except (ValueError, TypeError):
        since_id = 0
    with get_connection() as conn:
        row = conn.execute(
            "SELECT id FROM chat_conversations WHERE session_id=?", (session_id,)
        ).fetchone()
        if not row:
            return jsonify({"ok": False, "error": "Not found"}), 404
        conv_id = row[0]
        msgs = conn.execute(
            "SELECT id, sender, content, created_at FROM chat_messages "
            "WHERE conversation_id=? AND id>? ORDER BY id",
            (conv_id, since_id),
        ).fetchall()
    return jsonify({"ok": True, "messages": [dict(m) for m in msgs]})


# ── Live Chat — CMS API ───────────────────────────────────────────────────────

@app.get("/cms/live-chat")
def cms_live_chat() -> Any:
    redir = _ensure_admin_session_html()
    if redir:
        return redir
    if not _cms_can("live_chat", "view"):
        return _cms_forbidden_html("live_chat", "view")
    with get_connection() as conn:
        grants = _cms_load_role_grants(conn, _cms_session_role())
    if _admin_full_access():
        grants = {mid: list(CMS_ACTIONS) for mid in CMS_MODULE_IDS}
    return render_template(
        "cms_live_chat.html",
        settings=fetch_settings(),
        cms_grants_json=json.dumps(grants, ensure_ascii=False),
        can_reply_live_chat=_cms_can("live_chat", "create"),
        **_admin_page_template_kwargs(),
    )


@app.get("/api/cms/live-chat/conversations")
def cms_live_chat_conversations() -> Any:
    if not _cms_can("live_chat", "view"):
        return _cms_forbidden_json("live_chat", "view")
    status_filter = request.args.get("status", "open")
    with get_connection() as conn:
        rows = conn.execute(
            """SELECT c.id, c.session_id, c.visitor_name, c.visitor_page, c.status,
                      c.created_at, c.updated_at,
                      (SELECT COUNT(*) FROM chat_messages m WHERE m.conversation_id=c.id) AS msg_count,
                      (SELECT content FROM chat_messages m WHERE m.conversation_id=c.id ORDER BY m.id DESC LIMIT 1) AS last_msg,
                      (SELECT sender  FROM chat_messages m WHERE m.conversation_id=c.id ORDER BY m.id DESC LIMIT 1) AS last_sender
               FROM chat_conversations c WHERE c.status=? ORDER BY c.updated_at DESC LIMIT 100""",
            (status_filter,),
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.get("/api/cms/live-chat/messages/<int:conv_id>")
def cms_live_chat_messages(conv_id: int) -> Any:
    if not _cms_can("live_chat", "view"):
        return _cms_forbidden_json("live_chat", "view")
    try:
        since_id = int(request.args.get("since", "0"))
    except (ValueError, TypeError):
        since_id = 0
    with get_connection() as conn:
        msgs = conn.execute(
            "SELECT id, sender, content, created_at FROM chat_messages "
            "WHERE conversation_id=? AND id>? ORDER BY id",
            (conv_id, since_id),
        ).fetchall()
    return jsonify([dict(m) for m in msgs])


@app.post("/api/cms/live-chat/reply")
def cms_live_chat_reply() -> Any:
    if not _cms_can("live_chat", "create"):
        return _cms_forbidden_json("live_chat", "create")
    data = request.get_json(silent=True) or {}
    conv_id = int(data.get("conv_id") or 0)
    content = str(data.get("content") or "").strip()
    if not conv_id or not content:
        return jsonify({"ok": False, "error": "conv_id và content bắt buộc"}), 400
    with get_connection() as conn:
        if not conn.execute("SELECT id FROM chat_conversations WHERE id=?", (conv_id,)).fetchone():
            return jsonify({"ok": False, "error": "Conversation not found"}), 404
        conn.execute(
            "INSERT INTO chat_messages (conversation_id, sender, content) VALUES (?,?,?)",
            (conv_id, "staff", content),
        )
        conn.execute(
            "UPDATE chat_conversations SET updated_at=datetime('now') WHERE id=?", (conv_id,)
        )
        msg_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    return jsonify({"ok": True, "msg_id": msg_id})


@app.put("/api/cms/live-chat/conversation/<int:conv_id>")
def cms_live_chat_update_conv(conv_id: int) -> Any:
    if not _cms_can("live_chat", "edit"):
        return _cms_forbidden_json("live_chat", "edit")
    data = request.get_json(silent=True) or {}
    with get_connection() as conn:
        if "status" in data:
            conn.execute(
                "UPDATE chat_conversations SET status=?, updated_at=datetime('now') WHERE id=?",
                (str(data["status"])[:20], conv_id),
            )
    return jsonify({"ok": True})


@app.get("/api/cms/live-chat/unread")
def cms_live_chat_unread() -> Any:
    if not _cms_can("live_chat", "view"):
        return jsonify({"count": 0})
    with get_connection() as conn:
        count = conn.execute(
            "SELECT COUNT(*) FROM chat_conversations WHERE status='open'",
        ).fetchone()[0]
    return jsonify({"count": count})


@app.put("/api/cms/live-chat/settings")
def cms_live_chat_save_settings() -> Any:
    if not _cms_can("live_chat", "edit"):
        return _cms_forbidden_json("live_chat", "edit")
    data = request.get_json(silent=True) or {}
    allowed = {"live_chat_enabled", "live_chat_mode", "live_chat_welcome", "live_chat_ai_prompt"}
    with get_connection() as conn:
        for key, value in data.items():
            if key in allowed:
                conn.execute(
                    "INSERT INTO settings (key, value) VALUES (?, ?) "
                    "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
                    (key, str(value)[:2000]),
                )
    return jsonify({"ok": True})


# ── Lead intake forms (HTML in docs/forms/lead-intake) ─────────────────────

_LEAD_INTAKE_FORMS_DIR = Path(__file__).resolve().parent / "docs" / "forms" / "lead-intake"
_LEAD_INTAKE_FORM_FILES = frozenset({
    "00-form-chung.html",
    "dich-vu-aeo.html",
    "dich-vu-quan-tri-website.html",
    "dich-vu-seo-audit.html",
    "dich-vu-seo-local.html",
    "dich-vu-seo-tong-the.html",
    "quang-cao-facebook.html",
    "quang-cao-google.html",
    "thiet-ke-landing-page.html",
    "thiet-ke-website-tron-goi.html",
    "thiet-ke-website.html",
    "thue-tai-khoan-quang-cao.html",
    "tiep-thi-noi-dung.html",
})


@app.get("/crm/forms/lead-intake/<filename>")
def crm_lead_intake_form_file(filename: str) -> Any:
    """Phục vụ form HTML tiếp nhận lead (in / điền offline)."""
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    safe = (filename or "").strip()
    if safe not in _LEAD_INTAKE_FORM_FILES:
        abort(404)
    path = _LEAD_INTAKE_FORMS_DIR / safe
    if not path.is_file():
        abort(404)
    return send_file(path, mimetype="text/html; charset=utf-8")


# ── Lead intake CRM sessions (form nhập & chỉnh sửa trên hệ thống) ─────────

@app.get("/crm/intake")
def crm_lead_intake_page() -> Any:
    """Trang nhập/chỉnh sửa form Lead Intake khi gặp KH."""
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    from crm_lead_intake import get_session as _intake_get, list_sessions as _intake_list
    from crm_lead_intake_definitions import (
        COMMON_FORM_SLUG,
        get_crm_fields_for_slug,
        get_ui_definition,
        resolve_definition_slug,
    )
    from crm_svc_tasks import SERVICE_LABELS as _svc_labels

    lifecycle_id = _opt_pos_int(request.args.get("lifecycle_id"))
    lead_id = _opt_pos_int(request.args.get("lead_id"))
    session_id = _opt_pos_int(request.args.get("session_id"))
    auto_create_mode = str(request.args.get("mode") or "").strip()
    if auto_create_mode not in ("phone", "in_person"):
        auto_create_mode = ""
    auto_create = str(request.args.get("auto_create") or "").strip() in ("1", "true", "yes")
    lifecycle = None
    customer = None
    lead = None
    service_slug = str(request.args.get("service_slug") or "").strip()
    sessions: list[dict] = []
    active_session = None
    _intake_print_file = "00-form-chung.html"

    with get_connection() as conn:
        if lifecycle_id:
            row = conn.execute(
                "SELECT * FROM crm_service_lifecycle WHERE id = ?", (lifecycle_id,)
            ).fetchone()
            if row is None:
                return "Không tìm thấy lifecycle", 404
            lifecycle = dict(row)
            service_slug = lifecycle.get("service_slug") or service_slug
            if not lead_id and lifecycle.get("lead_id"):
                lead_id = int(lifecycle["lead_id"])
            if lifecycle.get("customer_id"):
                crow = conn.execute(
                    "SELECT * FROM crm_customers WHERE id = ?",
                    (lifecycle["customer_id"],),
                ).fetchone()
                customer = dict(crow) if crow else None
            sessions = _intake_list(conn, lifecycle_id=lifecycle_id)
            def_slug = resolve_definition_slug(service_slug)
            _intake_print_file = (
                "00-form-chung.html"
                if def_slug == COMMON_FORM_SLUG
                else f"{service_slug}.html"
                if f"{service_slug}.html" in _LEAD_INTAKE_FORM_FILES
                else "00-form-chung.html"
            )
        elif lead_id:
            lrow = conn.execute(
                "SELECT id, full_name FROM crm_leads WHERE id = ?", (lead_id,)
            ).fetchone()
            if lrow is None:
                return "Không tìm thấy lead", 404
            lead = dict(lrow)
            if not service_slug:
                service_slug = COMMON_FORM_SLUG
            sessions = _intake_list(conn, lead_id=lead_id)
            _intake_print_file = "00-form-chung.html"

        if session_id:
            active_session = _intake_get(conn, session_id)
            if active_session is None:
                return "Không tìm thấy phiên intake", 404
            if lifecycle_id and active_session.get("lifecycle_id") not in (None, lifecycle_id):
                return "Phiên không thuộc lifecycle này", 400
            if lead_id and active_session.get("lead_id") not in (None, lead_id):
                if not lifecycle_id:
                    return "Phiên không thuộc lead này", 400
            if not lifecycle_id and active_session.get("lifecycle_id"):
                lifecycle_id = int(active_session["lifecycle_id"])
                row = conn.execute(
                    "SELECT * FROM crm_service_lifecycle WHERE id = ?", (lifecycle_id,)
                ).fetchone()
                lifecycle = dict(row) if row else None
                service_slug = (lifecycle or {}).get("service_slug") or service_slug
                sessions = _intake_list(conn, lifecycle_id=lifecycle_id)
            if not lead_id and active_session.get("lead_id"):
                lead_id = int(active_session["lead_id"])
                if not sessions:
                    sessions = _intake_list(conn, lead_id=lead_id)
            service_slug = active_session.get("service_slug") or service_slug

    def_slug = resolve_definition_slug(service_slug)
    is_common_form = def_slug == COMMON_FORM_SLUG
    definition = get_ui_definition(service_slug)
    form_fields = get_crm_fields_for_slug(service_slug)

    recap_info: dict[str, Any] | None = None
    if active_session and active_session.get("mode") == "in_person":
        ans = active_session.get("answers_json") or {}
        meta = ans.get("meta") if isinstance(ans.get("meta"), dict) else {}
        if meta.get("recap") or meta.get("phone_completed_at"):
            recap_info = {
                "text": meta.get("recap") or ans.get("recap") or "",
                "phone_session_id": meta.get("phone_session_id"),
                "phone_completed_at": meta.get("phone_completed_at") or "",
            }

    service_label = _svc_labels.get(def_slug, definition.get("title") or service_slug)
    can_auto_create = auto_create and not active_session and bool(lifecycle_id or lead_id)

    page_meta = {
        "lifecycle_id": lifecycle_id,
        "lead_id": lead_id,
        "service_slug": def_slug if is_common_form else service_slug,
        "is_common_form": is_common_form,
        "service_label": service_label,
        "customer_name": (customer or {}).get("name") or "",
        "lead_name": (lead or {}).get("full_name") or "",
        "definition": definition,
        "form_fields": form_fields,
        "sessions": sessions,
        "active_session": active_session,
        "auto_create": can_auto_create,
        "auto_create_mode": auto_create_mode or "phone",
        "workflow_url": (
            url_for("crm_service_workflow_page", lifecycle_id=lifecycle_id)
            if lifecycle_id else ""
        ),
        "leads_url": url_for("crm_leads_page") if lead_id and not lifecycle_id else "",
        "print_form_url": url_for(
            "crm_lead_intake_form_file", filename=_intake_print_file
        ),
        "recap": recap_info,
        "go_thresholds": {"go": 24, "nurture_min": 18},
    }
    return render_template(
        "crm_lead_intake.html",
        lifecycle=lifecycle,
        customer=customer,
        lead=lead,
        service_label=service_label,
        is_common_form=is_common_form,
        page_meta=page_meta,
        active_session=active_session,
        **_admin_page_template_kwargs(),
    )


@app.get("/api/crm/intake/definitions")
def api_intake_definitions() -> Any:
    from crm_lead_intake_definitions import (
        COMMON_FORM_SLUG,
        bant_rows,
        get_common_form_definition,
        service_slugs,
    )
    common = get_common_form_definition()
    return jsonify({
        "slugs": list(service_slugs()),
        "common_slug": COMMON_FORM_SLUG,
        "common": {
            "title": common.get("title"),
            "phone_questions_count": len(common.get("phone_qs") or []),
            "inperson_questions_count": len(common.get("inperson_qs") or []),
        },
        "bant_rows": [{"label": l, "hint": h} for l, h in bant_rows()],
    })


@app.get("/api/crm/intake/definitions/<slug>")
def api_intake_definition(slug: str) -> Any:
    from crm_lead_intake_definitions import get_ui_definition
    return jsonify(get_ui_definition(slug))


@app.get("/api/crm/intake/sessions")
def api_intake_sessions_list() -> Any:
    from crm_lead_intake import list_sessions as _intake_list
    lifecycle_id = _opt_pos_int(request.args.get("lifecycle_id"))
    lead_id = _opt_pos_int(request.args.get("lead_id"))
    if not lifecycle_id and not lead_id:
        return jsonify({"error": "Cần lifecycle_id hoặc lead_id"}), 400
    with get_connection() as conn:
        sessions = _intake_list(
            conn, lifecycle_id=lifecycle_id, lead_id=lead_id
        )
    return jsonify({"sessions": sessions})


@app.post("/api/crm/intake/sessions")
def api_intake_session_create() -> Any:
    from crm_lead_intake import create_session as _intake_create
    payload = request.get_json(force=True) or {}
    lifecycle_id = _opt_pos_int(payload.get("lifecycle_id"))
    lead_id = _opt_pos_int(payload.get("lead_id"))
    service_slug = str(payload.get("service_slug") or "").strip()
    mode = str(payload.get("mode") or "phone").strip()
    if not lifecycle_id and not lead_id:
        return jsonify({"error": "Cần lifecycle_id hoặc lead_id"}), 400
    from crm_lead_intake_definitions import COMMON_FORM_SLUG, normalize_intake_slug
    if not service_slug and lifecycle_id:
        with get_connection() as conn:
            row = conn.execute(
                "SELECT service_slug FROM crm_service_lifecycle WHERE id = ?",
                (lifecycle_id,),
            ).fetchone()
            if row:
                service_slug = str(row["service_slug"] or "")
    if not service_slug:
        service_slug = COMMON_FORM_SLUG
    service_slug = normalize_intake_slug(service_slug) or COMMON_FORM_SLUG
    with get_connection() as conn:
        try:
            sid = _intake_create(
                conn,
                lifecycle_id=lifecycle_id,
                lead_id=lead_id,
                service_slug=service_slug,
                mode=mode,
                am_id=_opt_pos_int(session.get("admin_id")),
                contact_name=str(payload.get("contact_name") or "")[:500],
                contact_role=str(payload.get("contact_role") or "")[:200],
                company_name=str(payload.get("company_name") or "")[:500],
                source=str(payload.get("source") or "")[:200],
            )
            from crm_lead_intake import get_session as _intake_get
            created = _intake_get(conn, sid)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    return jsonify(created), 201


@app.get("/api/crm/intake/sessions/<int:session_id>")
def api_intake_session_get(session_id: int) -> Any:
    from crm_lead_intake import get_session as _intake_get
    with get_connection() as conn:
        row = _intake_get(conn, session_id)
    if row is None:
        return jsonify({"error": "Không tìm thấy phiên"}), 404
    return jsonify(row)


@app.patch("/api/crm/intake/sessions/<int:session_id>")
def api_intake_session_patch(session_id: int) -> Any:
    from crm_lead_intake import update_session as _intake_update
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        updated = _intake_update(conn, session_id, payload)
    if updated is None:
        return jsonify({"error": "Không tìm thấy phiên"}), 404
    return jsonify(updated)


@app.post("/api/crm/intake/sessions/<int:session_id>/complete")
def api_intake_session_complete(session_id: int) -> Any:
    from crm_lead_intake import complete_session as _intake_complete
    from crm_lead_intake import trigger_intake_summary_async as _intake_ai_async
    with get_connection() as conn:
        try:
            updated = _intake_complete(
                conn, session_id, actor_id=_opt_pos_int(session.get("admin_id"))
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    if updated is None:
        return jsonify({"error": "Không tìm thấy phiên"}), 404
    try:
        _intake_ai_async(session_id, db_path=str(DB_PATH))
    except Exception as _ai_exc:
        app.logger.debug("Intake AI summary trigger skipped: %s", _ai_exc)
    return jsonify(updated)


@app.post("/api/crm/intake/sessions/<int:session_id>/ai-summary")
def api_intake_session_ai_summary(session_id: int) -> Any:
    """Tạo lại AI summary cho phiên intake (sync — dùng khi cần refresh)."""
    from crm_lead_intake import (
        generate_intake_summary as _gen_summary,
        get_session as _intake_get,
        save_intake_ai_result as _save_ai,
    )
    with get_connection() as conn:
        session = _intake_get(conn, session_id)
        if session is None:
            return jsonify({"error": "Không tìm thấy phiên"}), 404
        result = _gen_summary(session)
        if result is None:
            return jsonify({
                "error": "Không tạo được AI summary (thiếu API key hoặc lỗi model)",
            }), 503
        updated = _save_ai(conn, session_id, result)
    return jsonify(updated)


@app.get("/api/crm/intake/stats")
def api_intake_stats() -> Any:
    from crm_lead_intake import get_intake_stats as _intake_stats
    am_id = _opt_pos_int(request.args.get("am_id"))
    by_am = str(request.args.get("by_am", "")).lower() in ("1", "true", "yes")
    with get_connection() as conn:
        stats = _intake_stats(conn, am_id=am_id, by_am=by_am)
    return jsonify(stats)


@app.get("/api/crm/intake/entry")
def api_intake_entry() -> Any:
    """Resolve lifecycle + redirect URL cho entry từ Lead UI."""
    from crm_lead_intake import resolve_intake_entry as _intake_entry
    lead_id = _opt_pos_int(request.args.get("lead_id"))
    mode = str(request.args.get("mode") or "phone").strip()
    form = str(request.args.get("form") or "").strip()
    if not lead_id:
        return jsonify({"ok": False, "error": "Cần lead_id"}), 400
    with get_connection() as conn:
        result = _intake_entry(conn, lead_id=lead_id, mode=mode, form=form)
    if not result.get("ok"):
        return jsonify(result), 404
    return jsonify(result)


@app.post("/api/crm/intake/sessions/<int:session_id>/reopen")
def api_intake_session_reopen(session_id: int) -> Any:
    from crm_lead_intake import reopen_session as _intake_reopen
    with get_connection() as conn:
        updated = _intake_reopen(conn, session_id)
    if updated is None:
        return jsonify({"error": "Không tìm thấy phiên"}), 404
    return jsonify(updated)


# ── Service Delivery Dashboard ─────────────────────────────────────────────

@app.get("/crm/service-delivery")
def crm_service_delivery_page() -> Any:
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    from collections import defaultdict
    from crm_svc_presales import get_funnel_stats as _funnel_stats
    now = datetime.utcnow()
    period_start = f"{now.year:04d}-{now.month:02d}-01"
    if now.month == 12:
        period_end = f"{now.year:04d}-12-31"
    else:
        from datetime import timedelta
        next_month = datetime(now.year, now.month + 1, 1)
        period_end = (next_month - timedelta(days=1)).strftime("%Y-%m-%d")
    with get_connection() as conn:
        include_draft = not _crm_presales_on_lead_enabled()
        lifecycles = _svc_list_active(conn, include_draft=include_draft)
        from crm_lead_intake import get_intake_stats as _intake_stats
        intake_stats = _intake_stats(conn)
        funnel_stats = _funnel_stats(
            conn,
            period_start=period_start,
            period_end=period_end,
        )
        am_staff = [
            dict(r)
            for r in conn.execute(
                """
                SELECT DISTINCT s.id, s.name
                FROM crm_staff s
                WHERE s.active = 1
                  AND (
                    s.id IN (
                        SELECT assigned_am FROM crm_service_lifecycle
                        WHERE assigned_am IS NOT NULL
                    )
                    OR s.id IN (
                        SELECT assigned_am FROM crm_lead_presales
                        WHERE assigned_am IS NOT NULL
                    )
                    OR s.id IN (
                        SELECT owner_id FROM crm_leads
                        WHERE owner_id IS NOT NULL
                    )
                  )
                ORDER BY s.name
                """
            ).fetchall()
        ]
    by_stage: dict = defaultdict(list)
    for lc in lifecycles:
        by_stage[lc["stage"]].append(lc)
    return render_template(
        "crm_service_delivery.html",
        by_stage=by_stage,
        stages=SVC_LIFECYCLE_STAGES,
        valid_slugs=sorted(SVC_LIFECYCLE_SLUGS),
        intake_stats=intake_stats,
        funnel_stats=funnel_stats,
        am_staff=am_staff,
        funnel_period_start=period_start,
        funnel_period_end=period_end,
        **_admin_page_template_kwargs(),
    )


# ── Service Workflow Detail ─────────────────────────────────────────────────

@app.get("/crm/service-delivery/<int:lifecycle_id>")
def crm_service_workflow_page(lifecycle_id: int) -> Any:
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    from crm_svc_tasks import (
        SERVICE_LABELS as _svc_labels,
        get_progress as _svc_progress,
        list_tasks as _svc_list_tasks,
        seed_tasks as _svc_seed,
    )
    from crm_svc_workflow_steps import SERVICE_WORKFLOW_STEPS as _svc_steps
    with get_connection() as conn:
        lc = conn.execute(
            "SELECT * FROM crm_service_lifecycle WHERE id = ?", (lifecycle_id,)
        ).fetchone()
        if lc is None:
            return "Không tìm thấy lifecycle", 404
        lc = dict(lc)
        _svc_seed(conn, lifecycle_id=lifecycle_id, service_slug=lc["service_slug"])
        from crm_svc_tasks import ensure_recurring_deliver_tasks as _svc_ensure_deliver

        _svc_ensure_deliver(conn, lifecycle_id=lifecycle_id, service_slug=lc["service_slug"])
        from crm_service_lifecycle import get_stage_advance_info as _svc_advance_info

        advance_info = _svc_advance_info(conn, lifecycle_id=lifecycle_id)
        # Seed và load risks
        from crm_svc_risk import (
            seed_risks as _risk_seed,
            list_risks as _risk_list,
            get_latest_scan as _risk_latest_scan,
        )
        _risk_seed(conn, lifecycle_id=lifecycle_id, service_slug=lc["service_slug"])
        risks = _risk_list(conn, lifecycle_id=lifecycle_id)
        latest_risk_scan = _risk_latest_scan(conn, lifecycle_id=lifecycle_id)
        tasks_by_stage = _svc_list_tasks(conn, lifecycle_id=lifecycle_id)
        progress = _svc_progress(conn, lifecycle_id=lifecycle_id)
        customer = None
        if lc.get("customer_id"):
            row = conn.execute(
                "SELECT * FROM crm_customers WHERE id = ?", (lc["customer_id"],)
            ).fetchone()
            customer = dict(row) if row else None
        # Finance data
        from crm_svc_finance import (
            get_summary as _fin_summary,
            list_payments as _fin_payments,
            list_expenses as _fin_expenses,
            get_latest_finance_scan as _fin_scan,
        )
        contract_amount_vnd = 0
        if lc.get("contract_id"):
            c_row = conn.execute(
                "SELECT amount_vnd FROM crm_contracts WHERE id = ?",
                (lc["contract_id"],),
            ).fetchone()
            if c_row:
                contract_amount_vnd = int(c_row["amount_vnd"] or 0)
        finance_summary = _fin_summary(conn, lifecycle_id, contract_amount_vnd)
        payments = _fin_payments(conn, lifecycle_id)
        expenses = _fin_expenses(conn, lifecycle_id, cost_phase="delivery")
        from crm_svc_presales import (
            get_presales_cost_summary as _presales_summary,
            show_presales_panel as _show_presales_panel,
        )
        presales_summary = _presales_summary(conn, lifecycle_id)
        presales_expenses = _fin_expenses(conn, lifecycle_id, cost_phase="presales")
        show_presales_panel = _show_presales_panel(lc)
        latest_health_scan = _fin_scan(conn, lifecycle_id, "health")
        latest_forecast_scan = _fin_scan(conn, lifecycle_id, "forecast")
        from crm_svc_kpi import get_lifecycle_staff_metrics as _kpi_staff
        lifecycle_staff = _kpi_staff(conn, lifecycle_id)
        crm_staff_list = [
            dict(r)
            for r in conn.execute(
                "SELECT id, name FROM crm_staff WHERE COALESCE(active, 1) = 1 ORDER BY name"
            ).fetchall()
        ]
        aeo_stats = None
        if lc["service_slug"] == "dich-vu-aeo":
            from crm_aeo import list_queries as _aeo_list
            aeo_qs = _aeo_list(conn, lc["customer_id"]) if lc.get("customer_id") else []
            total_q = len(aeo_qs)
            visible_q = sum(1 for q in aeo_qs if q.get("brand_visible") == 1)
            aeo_stats = {"total": total_q, "visible": visible_q}
        from crm_lead_intake import list_sessions as _intake_list_sessions
        intake_sessions = _intake_list_sessions(conn, lifecycle_id=lifecycle_id)
        consult_brief = None
        consult_advance_gate = None
        if lc.get("stage") == "consult":
            from crm_svc_consult_bridge import get_consult_brief as _consult_brief
            try:
                consult_brief = _consult_brief(conn, lifecycle_id)
            except ValueError:
                consult_brief = None
        elif lc.get("stage") == "lead":
            from crm_svc_consult_bridge import validate_consult_advance as _consult_gate
            try:
                consult_advance_gate = _consult_gate(conn, lifecycle_id)
            except ValueError:
                consult_advance_gate = None
        consult_proposal_url = None
        consult_stage_complete = False
        if lc.get("stage") == "consult":
            from crm_svc_tasks import is_stage_complete as _svc_stage_done

            consult_stage_complete = _svc_stage_done(conn, lifecycle_id, "consult")
            prop_customer_id = lc.get("customer_id")
            if not prop_customer_id and lc.get("lead_id"):
                lr = conn.execute(
                    "SELECT converted_customer_id FROM crm_leads WHERE id = ?",
                    (lc["lead_id"],),
                ).fetchone()
                if lr and lr[0]:
                    prop_customer_id = int(lr[0])
            if consult_stage_complete and prop_customer_id:
                consult_proposal_url = (
                    f"/crm/proposals?customer_id={prop_customer_id}"
                    f"&lifecycle_id={lifecycle_id}"
                    f"&service_slug={lc['service_slug']}&from_consult=1"
                )
        _intake_file = (
            f'{lc["service_slug"]}.html'
            if f'{lc["service_slug"]}.html' in _LEAD_INTAKE_FORM_FILES
            else "00-form-chung.html"
        )
        from crm_svc_consult_bridge import get_lifecycle_funnel_progress as _lc_funnel

        try:
            lifecycle_funnel = _lc_funnel(conn, lifecycle_id)
        except ValueError:
            lifecycle_funnel = None
        official_marketing_plan = None
        tmmt_deliver_gate = None
        if lc.get("marketing_plan_id") or lc.get("stage") in ("onboard", "deliver", "handover", "retain"):
            from crm_lead_presales_marketing_plan import (
                official_plan_payload as _official_mp_payload,
                validate_lifecycle_deliver_advance as _tmmt_deliver_gate,
            )

            try:
                official_marketing_plan = _official_mp_payload(conn, lifecycle_id)
            except Exception:
                official_marketing_plan = None
            if lc.get("stage") == "onboard":
                try:
                    tmmt_deliver_gate = _tmmt_deliver_gate(conn, lifecycle_id)
                except Exception:
                    tmmt_deliver_gate = None
    return render_template(
        "crm_service_workflow.html",
        lifecycle=lc,
        tasks_by_stage=tasks_by_stage,
        progress=progress,
        stages=SVC_LIFECYCLE_STAGES,
        customer=customer,
        service_label=_svc_labels.get(lc["service_slug"], lc["service_slug"]),
        service_steps=_svc_steps.get(lc["service_slug"], {}),
        stage_labels={
            "lead": "Lead", "consult": "Tư vấn", "proposal": "Báo giá",
            "onboard": "Onboarding", "deliver": "Triển khai",
            "handover": "Nghiệm thu", "retain": "Chăm sóc",
        },
        advance_info=advance_info,
        intake_form_file=_intake_file,
        intake_form_common="00-form-chung.html",
        intake_sessions=intake_sessions,
        risks=risks,
        latest_risk_scan=latest_risk_scan,
        finance_summary=finance_summary,
        payments=payments,
        expenses=expenses,
        latest_health_scan=latest_health_scan,
        latest_forecast_scan=latest_forecast_scan,
        lifecycle_staff=lifecycle_staff,
        crm_staff_list=crm_staff_list,
        presales_summary=presales_summary,
        presales_expenses=presales_expenses,
        show_presales_panel=show_presales_panel,
        consult_brief=consult_brief,
        consult_advance_gate=consult_advance_gate,
        consult_proposal_url=consult_proposal_url,
        consult_stage_complete=consult_stage_complete,
        lifecycle_funnel=lifecycle_funnel,
        aeo_stats=aeo_stats,
        official_marketing_plan=official_marketing_plan,
        tmmt_deliver_gate=tmmt_deliver_gate,
        today_iso=datetime.now().strftime("%Y-%m-%d"),
        tmmt_strategy_labels=CRM_MP_STRATEGY_FRAMEWORK_LABELS_VI,
        tmmt_prof_labels=CRM_MP_TARGET_MARKET_PROF_LABELS_VI,
        tmmt_core_keys=[
            "market_context",
            "segmentation_icp",
            "personas_roles",
            "pains_desired_outcomes",
        ],
        **_admin_page_template_kwargs(),
    )


@app.patch("/api/crm/svc-tasks/<int:task_id>")
def api_svc_task_patch(task_id: int) -> Any:
    from crm_svc_tasks import update_task as _svc_update
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute(
            "SELECT id, lifecycle_id FROM crm_svc_tasks WHERE id = ?", (task_id,)
        ).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy task"}), 404
        lifecycle_id = int(row["lifecycle_id"])
        is_done = payload.get("is_done")
        notes = payload.get("notes")
        form_data = payload.get("form_data")
        done_by = _opt_pos_int(payload.get("done_by"))
        _svc_update(
            conn, task_id,
            is_done=bool(is_done) if is_done is not None else None,
            notes=str(notes)[:4000] if notes is not None else None,
            form_data=form_data if isinstance(form_data, dict) else None,
            done_by=done_by,
        )
        if isinstance(form_data, dict) and any(
            k in form_data for k in ("assigned_sp", "seo_specialist", "content_specialist")
        ):
            try:
                from crm_service_lifecycle import sync_assigned_sp_from_tasks

                sync_assigned_sp_from_tasks(conn, lifecycle_id, overwrite=False)
            except Exception as exc:
                logger.warning(
                    "sync_assigned_sp_from_tasks lifecycle=%s task=%s: %s",
                    lifecycle_id,
                    task_id,
                    exc,
                )
        updated = conn.execute(
            "SELECT * FROM crm_svc_tasks WHERE id = ?", (task_id,)
        ).fetchone()
    return jsonify(dict(updated))


@app.post("/api/crm/svc-tasks")
def api_svc_task_create() -> Any:
    from crm_svc_tasks import create_custom_task as _svc_create
    payload = request.get_json(force=True) or {}
    lifecycle_id = _opt_pos_int(payload.get("lifecycle_id"))
    stage = str(payload.get("stage", "")).strip()
    title = str(payload.get("title", "")).strip()[:500]
    description = str(payload.get("description", "")).strip()[:2000]
    if not lifecycle_id or not stage or not title:
        return jsonify({"error": "Cần lifecycle_id, stage và title"}), 400
    with get_connection() as conn:
        tid = _svc_create(
            conn, lifecycle_id=lifecycle_id,
            stage=stage, title=title, description=description,
        )
        row = conn.execute(
            "SELECT * FROM crm_svc_tasks WHERE id = ?", (tid,)
        ).fetchone()
    return jsonify(dict(row)), 201


@app.delete("/api/crm/svc-tasks/<int:task_id>")
def api_svc_task_delete(task_id: int) -> Any:
    from crm_svc_tasks import delete_task as _svc_delete
    with get_connection() as conn:
        ok = _svc_delete(conn, task_id)
    if not ok:
        return jsonify({"error": "Không thể xoá — không phải custom task"}), 404
    return jsonify({"ok": True})


@app.post("/api/crm/svc-tasks/<int:task_id>/ai-assist")
def api_svc_task_ai_assist(task_id: int) -> Any:
    from crm_svc_tasks import SERVICE_LABELS as _svc_lbl, run_ai_assist as _svc_ai
    payload = request.get_json(force=True) or {}
    ctx: dict = payload.get("context") or {}
    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT t.*, lc.service_slug, lc.customer_id
            FROM crm_svc_tasks t
            JOIN crm_service_lifecycle lc ON lc.id = t.lifecycle_id
            WHERE t.id = ?
            """,
            (task_id,),
        ).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy task"}), 404
        if row["customer_id"]:
            cust = conn.execute(
                "SELECT name FROM crm_customers WHERE id = ?",
                (row["customer_id"],),
            ).fetchone()
            if cust:
                ctx.setdefault("customer_name", cust["name"] or "KH")
        ctx.setdefault(
            "service_name", _svc_lbl.get(row["service_slug"], row["service_slug"])
        )
        if str(row["stage"] or "") == "consult":
            from crm_svc_consult_bridge import build_ai_context_for_consult

            ctx = build_ai_context_for_consult(
                conn,
                int(row["lifecycle_id"]),
                task_id,
                ctx,
            )
            ctx.setdefault(
                "service_name", _svc_lbl.get(row["service_slug"], row["service_slug"])
            )
            if row["customer_id"] and not ctx.get("customer_name"):
                cust = conn.execute(
                    "SELECT name FROM crm_customers WHERE id = ?",
                    (row["customer_id"],),
                ).fetchone()
                if cust:
                    ctx["customer_name"] = cust["name"] or "KH"
        output = _svc_ai(conn, task_id=task_id, customer_context=ctx)
    return jsonify({"ai_output": output, "task_id": task_id})


# ── Service Risk Management ──────────────────────────────────────────────────

@app.get("/api/crm/svc-risks/<int:lifecycle_id>")
def api_svc_risks_list(lifecycle_id: int) -> Any:
    from crm_svc_risk import list_risks as _risk_list, get_latest_scan as _risk_scan
    with get_connection() as conn:
        risks = _risk_list(conn, lifecycle_id)
        latest_scan = _risk_scan(conn, lifecycle_id)
    return jsonify({"risks": risks, "latest_scan": latest_scan})


@app.patch("/api/crm/svc-risks/<int:risk_id>")
def api_svc_risk_patch(risk_id: int) -> Any:
    from crm_svc_risk import update_risk as _risk_update
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute(
            "SELECT id FROM crm_svc_risks WHERE id = ?", (risk_id,)
        ).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy risk"}), 404
        _risk_update(
            conn, risk_id,
            probability=str(payload["probability"]) if "probability" in payload else None,
            impact=str(payload["impact"]) if "impact" in payload else None,
            mitigation=str(payload["mitigation"])[:2000] if "mitigation" in payload else None,
            is_active=bool(payload["is_active"]) if "is_active" in payload else None,
        )
        updated = conn.execute(
            "SELECT * FROM crm_svc_risks WHERE id = ?", (risk_id,)
        ).fetchone()
    return jsonify(dict(updated))


@app.post("/api/crm/svc-risks")
def api_svc_risk_create() -> Any:
    from crm_svc_risk import create_custom_risk as _risk_create
    payload = request.get_json(force=True) or {}
    lifecycle_id = _opt_pos_int(payload.get("lifecycle_id"))
    title = str(payload.get("title", "")).strip()[:500]
    stage = str(payload.get("stage", "")).strip()
    category = str(payload.get("category", "")).strip()[:100]
    if not lifecycle_id or not title:
        return jsonify({"error": "Cần lifecycle_id và title"}), 400
    with get_connection() as conn:
        rid = _risk_create(
            conn, lifecycle_id=lifecycle_id, stage=stage, title=title, category=category
        )
        row = conn.execute("SELECT * FROM crm_svc_risks WHERE id = ?", (rid,)).fetchone()
    return jsonify(dict(row)), 201


@app.delete("/api/crm/svc-risks/<int:risk_id>")
def api_svc_risk_delete(risk_id: int) -> Any:
    from crm_svc_risk import delete_risk as _risk_delete
    with get_connection() as conn:
        ok = _risk_delete(conn, risk_id)
    if not ok:
        return jsonify({"error": "Không thể xoá — không phải custom risk"}), 404
    return jsonify({"ok": True})


@app.post("/api/crm/svc-risks/<int:lifecycle_id>/ai-scan")
def api_svc_risk_ai_scan(lifecycle_id: int) -> Any:
    from crm_svc_risk import run_ai_risk_scan as _risk_scan_fn
    from crm_svc_tasks import SERVICE_LABELS as _svc_labels
    with get_connection() as conn:
        lc = conn.execute(
            "SELECT * FROM crm_service_lifecycle WHERE id = ?", (lifecycle_id,)
        ).fetchone()
        if lc is None:
            return jsonify({"error": "Không tìm thấy lifecycle"}), 404
        lc = dict(lc)
        ctx: dict = {
            "service_name": _svc_labels.get(lc["service_slug"], lc["service_slug"]),
            "current_stage": lc["stage"],
            "progress_summary": "",
            "customer_name": "KH",
        }
        if lc.get("customer_id"):
            cust = conn.execute(
                "SELECT name FROM crm_customers WHERE id = ?", (lc["customer_id"],)
            ).fetchone()
            if cust:
                ctx["customer_name"] = cust["name"] or "KH"
        output = _risk_scan_fn(conn, lifecycle_id=lifecycle_id, customer_context=ctx)
    return jsonify({"ai_output": output, "lifecycle_id": lifecycle_id})



# ── Business Dashboard + Service Finance Tracking ──────────────────────────────

@app.get("/crm/business-dashboard")
def crm_business_dashboard_page() -> Any:
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    from crm_svc_finance_kpi import (
        collect_finance_kpi_alerts,
        get_alert_thresholds,
        get_finance_kpi_trends,
        load_finance_kpi_bundle,
    )
    from crm_svc_finance_kpi_inbox import get_finance_kpi_inbox_summary

    now = datetime.utcnow()
    year = _opt_pos_int(request.args.get("year")) or now.year
    month = _opt_pos_int(request.args.get("month")) or now.month
    trend_months = _opt_pos_int(request.args.get("trend_months")) or 6
    trend_months = max(3, min(trend_months, 12))

    with get_connection() as conn:
        bundle = load_finance_kpi_bundle(conn, year=year, month=month)
        kpi_alerts = collect_finance_kpi_alerts(
            conn, year=year, month=month, bundle=bundle
        )
        trends = get_finance_kpi_trends(
            conn, year=year, month=month, months=trend_months
        )
        thresholds = get_alert_thresholds(conn)
        kpi_inbox = get_finance_kpi_inbox_summary(conn)

    exec_m = bundle["exec_metrics"]
    conc = bundle["portfolio_metrics"]["concentration"]
    lead = bundle["lead_kpi"]
    ret = bundle["retention_metrics"]
    ar = bundle["ar_aging"]

    return render_template(
        "crm_business_dashboard.html",
        year=year,
        month=month,
        trend_months=trend_months,
        kpi_alerts=kpi_alerts,
        trends=trends,
        thresholds=thresholds,
        exec_metrics=exec_m,
        concentration=conc,
        lead_kpi=lead,
        retention_metrics=ret,
        ar_aging=ar,
        kpi_inbox=kpi_inbox,
        **_admin_page_template_kwargs(),
    )


@app.get("/crm/owner-weekly")
def crm_owner_weekly_dashboard_page() -> Any:
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    from crm_owner_weekly_dashboard import (
        OWNER_WEEKLY_TARGET_GROUPS,
        OWNER_WEEKLY_TARGET_LABELS,
        get_owner_weekly_dashboard,
        get_owner_weekly_targets,
    )
    from crm_owner_weekly_inbox import get_owner_weekly_inbox_summary

    year = _opt_pos_int(request.args.get("year"))
    week = _opt_pos_int(request.args.get("week"))
    trend_weeks = _opt_pos_int(request.args.get("trend_weeks")) or 8
    week_end_raw = str(request.args.get("week_end") or "").strip()[:10]
    week_end = None
    if week_end_raw:
        try:
            week_end = date.fromisoformat(week_end_raw)
        except ValueError:
            week_end = None

    with get_connection() as conn:
        dashboard = get_owner_weekly_dashboard(
            conn,
            week_end=week_end,
            year=year,
            iso_week=week,
            trend_weeks=trend_weeks,
        )
        targets = get_owner_weekly_targets(conn)
        weekly_inbox = get_owner_weekly_inbox_summary(conn)

    return render_template(
        "crm_owner_weekly_dashboard.html",
        dashboard=dashboard,
        targets=targets,
        target_labels=OWNER_WEEKLY_TARGET_LABELS,
        target_groups=OWNER_WEEKLY_TARGET_GROUPS,
        weekly_inbox=weekly_inbox,
        can_configure=_admin_section_can("crm_owner_weekly_dashboard", "configure"),
        can_export=_admin_section_can("crm_owner_weekly_dashboard", "export"),
        **_admin_page_template_kwargs(),
    )


@app.get("/api/crm/owner-weekly")
def api_crm_owner_weekly_dashboard() -> Any:
    from crm_owner_weekly_dashboard import get_owner_weekly_dashboard

    year = _opt_pos_int(request.args.get("year"))
    week = _opt_pos_int(request.args.get("week"))
    trend_weeks = _opt_pos_int(request.args.get("trend_weeks")) or 8
    week_end_raw = str(request.args.get("week_end") or "").strip()[:10]
    week_end = None
    if week_end_raw:
        try:
            week_end = date.fromisoformat(week_end_raw)
        except ValueError:
            week_end = None

    with get_connection() as conn:
        dashboard = get_owner_weekly_dashboard(
            conn,
            week_end=week_end,
            year=year,
            iso_week=week,
            trend_weeks=trend_weeks,
        )
    return jsonify(dashboard)


@app.get("/api/crm/owner-weekly/config")
def api_crm_owner_weekly_config_get() -> Any:
    from crm_owner_weekly_dashboard import (
        OWNER_WEEKLY_ENV_KEYS,
        OWNER_WEEKLY_TARGET_DEFAULTS,
        OWNER_WEEKLY_TARGET_LABELS,
        get_owner_weekly_targets,
    )

    with get_connection() as conn:
        values = get_owner_weekly_targets(conn)
    return jsonify(
        {
            "targets": values,
            "defaults": OWNER_WEEKLY_TARGET_DEFAULTS,
            "labels": OWNER_WEEKLY_TARGET_LABELS,
            "env_keys": OWNER_WEEKLY_ENV_KEYS,
        }
    )


@app.patch("/api/crm/owner-weekly/config")
def api_crm_owner_weekly_config_patch() -> Any:
    if not _admin_section_can("crm_owner_weekly_dashboard", "configure"):
        return _admin_section_forbidden_json("crm_owner_weekly_dashboard", "configure")
    from crm_owner_weekly_dashboard import set_owner_weekly_targets

    payload = request.get_json(force=True) or {}
    updates = payload.get("targets") or payload.get("thresholds") or payload
    if not isinstance(updates, dict):
        return jsonify({"error": "targets phải là object."}), 400
    with get_connection() as conn:
        values = set_owner_weekly_targets(conn, updates)
    return jsonify({"ok": True, "targets": values})


@app.get("/api/crm/owner-weekly/cash-snapshots")
def api_crm_owner_weekly_cash_snapshots_get() -> Any:
    from crm_owner_cash_ledger import list_cash_snapshots

    limit = _opt_pos_int(request.args.get("limit")) or 24
    with get_connection() as conn:
        snapshots = list_cash_snapshots(conn, limit=limit)
    return jsonify({"snapshots": snapshots})


@app.post("/api/crm/owner-weekly/cash-snapshots")
def api_crm_owner_weekly_cash_snapshots_upsert() -> Any:
    if not _admin_section_can("crm_owner_weekly_dashboard", "configure"):
        return _admin_section_forbidden_json("crm_owner_weekly_dashboard", "configure")
    from crm_owner_cash_ledger import upsert_cash_snapshot

    payload = request.get_json(force=True) or {}
    snap_raw = str(payload.get("snapshot_on") or "").strip()[:10]
    if not snap_raw:
        return jsonify({"error": "snapshot_on bắt buộc (YYYY-MM-DD)."}), 400
    try:
        snap_on = date.fromisoformat(snap_raw)
    except ValueError:
        return jsonify({"error": "snapshot_on không hợp lệ."}), 400
    try:
        balance = int(payload.get("balance_vnd"))
    except (TypeError, ValueError):
        return jsonify({"error": "balance_vnd phải là số nguyên."}), 400
    source = str(payload.get("source") or "manual").strip().lower()
    notes = str(payload.get("notes") or "").strip()
    with get_connection() as conn:
        row = upsert_cash_snapshot(
            conn,
            snapshot_on=snap_on,
            balance_vnd=balance,
            source=source,
            notes=notes,
        )
    return jsonify({"ok": True, "snapshot": row})


@app.delete("/api/crm/owner-weekly/cash-snapshots")
def api_crm_owner_weekly_cash_snapshots_delete() -> Any:
    if not _admin_section_can("crm_owner_weekly_dashboard", "configure"):
        return _admin_section_forbidden_json("crm_owner_weekly_dashboard", "configure")
    from crm_owner_cash_ledger import delete_cash_snapshot

    snap_raw = str(request.args.get("snapshot_on") or "").strip()[:10]
    if not snap_raw:
        payload = request.get_json(silent=True) or {}
        snap_raw = str(payload.get("snapshot_on") or "").strip()[:10]
    if not snap_raw:
        return jsonify({"error": "snapshot_on bắt buộc."}), 400
    try:
        snap_on = date.fromisoformat(snap_raw)
    except ValueError:
        return jsonify({"error": "snapshot_on không hợp lệ."}), 400
    with get_connection() as conn:
        deleted = delete_cash_snapshot(conn, snap_on)
    return jsonify({"ok": True, "deleted": deleted})


@app.get("/api/crm/owner-weekly/export")
def api_crm_owner_weekly_export() -> Any:
    if not _admin_section_can("crm_owner_weekly_dashboard", "export"):
        return _admin_section_forbidden_json("crm_owner_weekly_dashboard", "export")
    from crm_owner_weekly_dashboard import (
        build_owner_weekly_export_sheets,
        get_owner_weekly_dashboard,
    )

    year = _opt_pos_int(request.args.get("year"))
    week = _opt_pos_int(request.args.get("week"))
    week_end_raw = str(request.args.get("week_end") or "").strip()[:10]
    week_end = None
    if week_end_raw:
        try:
            week_end = date.fromisoformat(week_end_raw)
        except ValueError:
            week_end = None
    fmt = str(request.args.get("format") or "xlsx").strip().lower()
    if fmt not in ("csv", "xlsx"):
        fmt = "xlsx"

    with get_connection() as conn:
        dashboard = get_owner_weekly_dashboard(
            conn,
            week_end=week_end,
            year=year,
            iso_week=week,
        )
    sheets = build_owner_weekly_export_sheets(dashboard)
    wk = dashboard.get("week") or {}
    iso_year = int(wk.get("iso_year") or datetime.utcnow().year)
    iso_week = int(wk.get("iso_week") or 1)
    stamp = datetime.now().strftime("%Y-%m-%d")
    base = f"owner-weekly-{iso_year:04d}-W{iso_week:02d}-{stamp}"
    if fmt == "xlsx":
        return _crm_re_export_xlsx_sheets(sheets, filename=f"{base}.xlsx")
    headers, rows = sheets[0][1], list(sheets[0][2])
    for title, _hdr, sheet_rows in sheets[1:]:
        rows.append([])
        rows.append([f"=== {title.upper()} ==="])
        if _hdr:
            rows.append(list(_hdr))
        rows.extend(sheet_rows)
    return _crm_re_export_csv(headers, rows, filename=f"{base}.csv")


@app.post("/api/crm/owner-weekly/alert-cron")
def api_crm_owner_weekly_alert_cron() -> Any:
    if not _crm_finance_kpi_cron_allowed():
        return jsonify({"error": "Unauthorized"}), 403
    from crm_owner_weekly_notify import dispatch_owner_weekly_alerts

    payload = request.get_json(silent=True) or {}
    year = _opt_pos_int(payload.get("year") or request.args.get("year"))
    week = _opt_pos_int(payload.get("week") or request.args.get("iso_week") or request.args.get("week"))
    only_red_raw = payload.get("only_red", request.args.get("only_red", "0"))
    only_red = str(only_red_raw).strip().lower() not in ("0", "false", "no")
    public_base = (os.getenv("PTT_PUBLIC_BASE_URL") or request.host_url or "").rstrip("/")
    dashboard_url = ""
    if year and week:
        dashboard_url = f"{public_base}/crm/owner-weekly?year={year}&week={week}"

    with get_connection() as conn:
        result = dispatch_owner_weekly_alerts(
            conn,
            iso_year=year,
            iso_week=week,
            only_red=only_red,
            dashboard_url=dashboard_url,
        )
    code = 200 if result.get("ok") else 400
    return jsonify(result), code


@app.post("/api/crm/owner-weekly/inbox/sync")
def api_crm_owner_weekly_inbox_sync() -> Any:
    from crm_owner_weekly_dashboard import get_owner_weekly_dashboard
    from crm_owner_weekly_inbox import sync_owner_weekly_inbox

    payload = request.get_json(silent=True) or {}
    year = _opt_pos_int(payload.get("year") or request.args.get("year"))
    week = _opt_pos_int(payload.get("week") or request.args.get("week"))
    public_base = (os.getenv("PTT_PUBLIC_BASE_URL") or request.host_url or "").rstrip("/")

    with get_connection() as conn:
        if year and week:
            dashboard = get_owner_weekly_dashboard(conn, year=year, iso_week=week)
            y = int(dashboard["week"]["iso_year"])
            w = int(dashboard["week"]["iso_week"])
        else:
            from crm_owner_weekly_dashboard import resolve_week_bounds

            _s, _e, y, w = resolve_week_bounds()
            dashboard = get_owner_weekly_dashboard(conn, year=y, iso_week=w)
        dash_url = f"{public_base}/crm/owner-weekly?year={y}&week={w}"
        inbox = sync_owner_weekly_inbox(
            conn,
            iso_year=y,
            iso_week=w,
            dashboard=dashboard,
            dashboard_url=dash_url,
        )
    return jsonify({"ok": True, "inbox": inbox})


@app.get("/api/crm/owner-weekly/inbox/summary")
def api_crm_owner_weekly_inbox_summary() -> Any:
    from crm_owner_weekly_inbox import get_owner_weekly_inbox_summary

    with get_connection() as conn:
        summary = get_owner_weekly_inbox_summary(conn)
    return jsonify(summary)


@app.get("/crm/financials")
def crm_financials_page() -> Any:
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    from crm_svc_finance import (
        BILLING_TYPE_LABELS,
        get_ar_aging as _ar_aging,
        get_recurring_revenue_summary as _recurring_sum,
        get_service_package_rollup as _pkg_rollup,
        get_summary as _fin_sum,
    )
    from crm_svc_retention import get_retention_metrics as _retention_metrics
    from crm_lead_kpi_metrics import get_unified_lead_kpi_summary as _lead_kpi
    from crm_svc_portfolio import get_portfolio_metrics as _portfolio_metrics
    from crm_svc_exec_metrics import get_exec_metrics as _exec_metrics
    from crm_svc_finance_kpi import collect_finance_kpi_alerts
    from crm_svc_tasks import SERVICE_LABELS as _svc_labels
    now = datetime.utcnow()
    year = _opt_pos_int(request.args.get("year")) or now.year
    month = _opt_pos_int(request.args.get("month")) or now.month
    with get_connection() as conn:
        lcs = conn.execute(
            """
            SELECT lc.id, lc.service_slug, lc.stage, lc.contract_id, lc.customer_id,
                   cu.name AS customer_name
            FROM crm_service_lifecycle lc
            LEFT JOIN crm_customers cu ON cu.id = lc.customer_id
            WHERE lc.status = 'active'
            ORDER BY lc.id
            """
        ).fetchall()
        rows = []
        for lc in lcs:
            lc = dict(lc)
            contract_amount_vnd = 0
            if lc.get("contract_id"):
                c_row = conn.execute(
                    "SELECT amount_vnd FROM crm_contracts WHERE id = ?",
                    (lc["contract_id"],),
                ).fetchone()
                if c_row:
                    contract_amount_vnd = int(c_row["amount_vnd"] or 0)
            summary = _fin_sum(conn, lc["id"], contract_amount_vnd)
            rows.append({
                "lifecycle_id": lc["id"],
                "service_slug": lc["service_slug"],
                "service_label": _svc_labels.get(lc["service_slug"], lc["service_slug"]),
                "stage": lc["stage"],
                "customer_name": lc.get("customer_name") or "—",
                **summary,
            })
        ar_aging = _ar_aging(conn)
        recurring_summary = _recurring_sum(conn, year=year, month=month)
        package_rollup = _pkg_rollup(conn, year=year, month=month)
        for pkg in package_rollup["packages"]:
            pkg["service_label"] = _svc_labels.get(
                pkg["service_slug"], pkg["service_slug"]
            )
        retention_metrics = _retention_metrics(conn, year=year, month=month)
        lead_kpi = _lead_kpi(conn, year=year, month=month, period_cohort=True)
        portfolio_metrics = _portfolio_metrics(conn, year=year, month=month)
        exec_metrics = _exec_metrics(conn, year=year, month=month)
        kpi_bundle = {
            "year": year,
            "month": month,
            "ar_aging": ar_aging,
            "recurring_summary": recurring_summary,
            "package_rollup": package_rollup,
            "retention_metrics": retention_metrics,
            "lead_kpi": lead_kpi,
            "portfolio_metrics": portfolio_metrics,
            "exec_metrics": exec_metrics,
        }
        kpi_alerts = collect_finance_kpi_alerts(
            conn, year=year, month=month, bundle=kpi_bundle
        )
    rows.sort(key=lambda r: r["margin_pct"])
    return render_template(
        "crm_financials.html",
        rows=rows,
        ar_aging=ar_aging,
        recurring_summary=recurring_summary,
        package_rollup=package_rollup,
        retention_metrics=retention_metrics,
        lead_kpi=lead_kpi,
        portfolio_metrics=portfolio_metrics,
        exec_metrics=exec_metrics,
        kpi_alerts=kpi_alerts,
        billing_type_labels=BILLING_TYPE_LABELS,
        year=year,
        month=month,
        **_admin_page_template_kwargs(),
    )


@app.get("/api/crm/finance/ar-aging")
def api_crm_finance_ar_aging() -> Any:
    from crm_svc_finance import get_ar_aging as _ar_aging
    as_of = str(request.args.get("as_of") or "").strip()[:10]
    am_id = _opt_pos_int(request.args.get("am_id"))
    with get_connection() as conn:
        result = _ar_aging(conn, as_of=as_of or None, am_id=am_id)
    return jsonify(result)


@app.get("/api/crm/finance/recurring-summary")
def api_crm_finance_recurring_summary() -> Any:
    from crm_svc_finance import get_recurring_revenue_summary as _recurring_sum
    now = datetime.utcnow()
    year = _opt_pos_int(request.args.get("year")) or now.year
    month = _opt_pos_int(request.args.get("month")) or now.month
    am_id = _opt_pos_int(request.args.get("am_id"))
    with get_connection() as conn:
        result = _recurring_sum(conn, year=year, month=month, am_id=am_id)
    return jsonify(result)


@app.get("/api/crm/finance/lead-kpi")
def api_crm_finance_lead_kpi() -> Any:
    from crm_lead_kpi_metrics import get_unified_lead_kpi_summary as _lead_kpi

    now = datetime.utcnow()
    year = _opt_pos_int(request.args.get("year")) or now.year
    month = _opt_pos_int(request.args.get("month")) or now.month
    staff_id = _opt_pos_int(request.args.get("staff_id"))
    with get_connection() as conn:
        result = _lead_kpi(
            conn,
            year=year,
            month=month,
            staff_id=staff_id,
            period_cohort=True,
        )
    return jsonify(result)


@app.post("/api/crm/finance/period-inputs")
def api_crm_finance_period_inputs() -> Any:
    payload = request.get_json(force=True) or {}
    year = _opt_pos_int(payload.get("year"))
    month = _opt_pos_int(payload.get("month"))
    if year is None or month is None or month < 1 or month > 12:
        return jsonify({"error": "year và month không hợp lệ."}), 400
    try:
        amount = max(0, int(payload.get("marketing_spend_vnd") or 0))
    except (TypeError, ValueError):
        return jsonify({"error": "marketing_spend_vnd không hợp lệ."}), 400
    from crm_svc_exec_metrics import get_cac_metrics, set_marketing_spend_vnd

    with get_connection() as conn:
        set_marketing_spend_vnd(conn, year=year, month=month, amount_vnd=amount)
        cac = get_cac_metrics(conn, year=year, month=month)
    return jsonify({"ok": True, "cac": cac})


@app.get("/api/crm/finance/kpi-alerts")
def api_crm_finance_kpi_alerts() -> Any:
    from crm_svc_finance_kpi import collect_finance_kpi_alerts as _kpi_alerts

    now = datetime.utcnow()
    year = _opt_pos_int(request.args.get("year")) or now.year
    month = _opt_pos_int(request.args.get("month")) or now.month
    with get_connection() as conn:
        result = _kpi_alerts(conn, year=year, month=month)
    return jsonify(result)


@app.get("/api/crm/finance/kpi-export")
def api_crm_finance_kpi_export() -> Any:
    from crm_svc_finance_kpi import (
        build_finance_kpi_export_sheets as _export_sheets,
        load_finance_kpi_bundle as _kpi_bundle,
    )

    now = datetime.utcnow()
    year = _opt_pos_int(request.args.get("year")) or now.year
    month = _opt_pos_int(request.args.get("month")) or now.month
    fmt = str(request.args.get("format") or "csv").strip().lower()
    if fmt not in ("csv", "xlsx"):
        fmt = "csv"
    with get_connection() as conn:
        bundle = _kpi_bundle(conn, year=year, month=month)
    sheets = _export_sheets(bundle)
    stamp = datetime.now().strftime("%Y-%m-%d")
    base = f"crm-finance-kpi-{year:04d}-{month:02d}-{stamp}"
    if fmt == "xlsx":
        return _crm_re_export_xlsx_sheets(sheets, filename=f"{base}.xlsx")
    headers, rows = sheets[0][1], list(sheets[0][2])
    for title, _hdr, sheet_rows in sheets[1:]:
        rows.append([])
        rows.append([f"=== {title.upper()} ==="])
        if _hdr:
            rows.append(list(_hdr))
        rows.extend(sheet_rows)
    return _crm_re_export_csv(headers, rows, filename=f"{base}.csv")


@app.get("/api/crm/finance/kpi-trends")
def api_crm_finance_kpi_trends() -> Any:
    from crm_svc_finance_kpi import get_finance_kpi_trends as _kpi_trends

    now = datetime.utcnow()
    year = _opt_pos_int(request.args.get("year")) or now.year
    month = _opt_pos_int(request.args.get("month")) or now.month
    months = _opt_pos_int(request.args.get("months")) or 6
    with get_connection() as conn:
        result = _kpi_trends(conn, year=year, month=month, months=months)
    return jsonify(result)


@app.get("/api/crm/finance/kpi-config")
def api_crm_finance_kpi_config_get() -> Any:
    from crm_svc_finance_kpi import THRESHOLD_DEFAULTS, THRESHOLD_ENV_KEYS
    from crm_svc_finance_kpi import get_alert_thresholds as _thresholds

    with get_connection() as conn:
        values = _thresholds(conn)
    return jsonify(
        {
            "thresholds": values,
            "defaults": THRESHOLD_DEFAULTS,
            "env_keys": THRESHOLD_ENV_KEYS,
        }
    )


@app.patch("/api/crm/finance/kpi-config")
def api_crm_finance_kpi_config_patch() -> Any:
    from crm_svc_finance_kpi import set_alert_thresholds as _set_thresholds

    payload = request.get_json(force=True) or {}
    updates = payload.get("thresholds") or payload
    if not isinstance(updates, dict):
        return jsonify({"error": "thresholds phải là object."}), 400
    with get_connection() as conn:
        values = _set_thresholds(conn, updates)
    return jsonify({"ok": True, "thresholds": values})


@app.post("/api/crm/finance/kpi-alert-cron")
def api_crm_finance_kpi_alert_cron() -> Any:
    if not _crm_finance_kpi_cron_allowed():
        return jsonify({"error": "Unauthorized"}), 403
    from crm_svc_finance_kpi_notify import dispatch_finance_kpi_alerts

    now = datetime.utcnow()
    payload = request.get_json(silent=True) or {}
    year = _opt_pos_int(payload.get("year") or request.args.get("year")) or now.year
    month = _opt_pos_int(payload.get("month") or request.args.get("month")) or now.month
    only_critical_raw = payload.get("only_critical", request.args.get("only_critical", "1"))
    only_critical = str(only_critical_raw).strip().lower() not in ("0", "false", "no")
    public_base = (os.getenv("PTT_PUBLIC_BASE_URL") or request.host_url or "").rstrip("/")
    dashboard_url = f"{public_base}/crm/business-dashboard?year={year}&month={month}"

    with get_connection() as conn:
        result = dispatch_finance_kpi_alerts(
            conn,
            year=year,
            month=month,
            only_critical=only_critical,
            dashboard_url=dashboard_url,
        )
    code = 200 if result.get("ok") else 400
    return jsonify(result), code


@app.post("/api/crm/finance/kpi-inbox/sync")
def api_crm_finance_kpi_inbox_sync() -> Any:
    from crm_svc_finance_kpi_inbox import sync_finance_kpi_inbox

    now = datetime.utcnow()
    payload = request.get_json(silent=True) or {}
    year = _opt_pos_int(payload.get("year") or request.args.get("year")) or now.year
    month = _opt_pos_int(payload.get("month") or request.args.get("month")) or now.month
    public_base = (os.getenv("PTT_PUBLIC_BASE_URL") or request.host_url or "").rstrip("/")
    dashboard_url = f"{public_base}/crm/business-dashboard?year={year}&month={month}"
    with get_connection() as conn:
        inbox = sync_finance_kpi_inbox(
            conn, year=year, month=month, dashboard_url=dashboard_url
        )
    return jsonify({"ok": True, "inbox": inbox})


@app.get("/api/crm/finance/kpi-inbox/summary")
def api_crm_finance_kpi_inbox_summary() -> Any:
    from crm_svc_finance_kpi_inbox import get_finance_kpi_inbox_summary

    with get_connection() as conn:
        summary = get_finance_kpi_inbox_summary(conn)
    return jsonify(summary)


@app.get("/api/crm/finance/exec-metrics")
def api_crm_finance_exec_metrics() -> Any:
    from crm_svc_exec_metrics import get_exec_metrics as _exec_metrics

    now = datetime.utcnow()
    year = _opt_pos_int(request.args.get("year")) or now.year
    month = _opt_pos_int(request.args.get("month")) or now.month
    am_id = _opt_pos_int(request.args.get("am_id"))
    with get_connection() as conn:
        result = _exec_metrics(conn, year=year, month=month, am_id=am_id)
    return jsonify(result)


@app.get("/api/crm/finance/portfolio-metrics")
def api_crm_finance_portfolio_metrics() -> Any:
    from crm_svc_portfolio import get_portfolio_metrics as _portfolio_metrics

    now = datetime.utcnow()
    year = _opt_pos_int(request.args.get("year")) or now.year
    month = _opt_pos_int(request.args.get("month")) or now.month
    with get_connection() as conn:
        result = _portfolio_metrics(conn, year=year, month=month)
    return jsonify(result)


@app.get("/api/crm/finance/retention-metrics")
def api_crm_finance_retention_metrics() -> Any:
    from crm_svc_retention import get_retention_metrics as _retention_metrics

    now = datetime.utcnow()
    year = _opt_pos_int(request.args.get("year")) or now.year
    month = _opt_pos_int(request.args.get("month")) or now.month
    with get_connection() as conn:
        result = _retention_metrics(conn, year=year, month=month)
    return jsonify(result)


@app.get("/api/crm/finance/service-package-rollup")
def api_crm_finance_service_package_rollup() -> Any:
    from crm_svc_finance import get_service_package_rollup as _pkg_rollup
    from crm_svc_tasks import SERVICE_LABELS as _svc_labels

    now = datetime.utcnow()
    year = _opt_pos_int(request.args.get("year")) or now.year
    month = _opt_pos_int(request.args.get("month")) or now.month
    status = str(request.args.get("status") or "active").strip() or "active"
    with get_connection() as conn:
        result = _pkg_rollup(conn, year=year, month=month, lifecycle_status=status)
        for pkg in result["packages"]:
            pkg["service_label"] = _svc_labels.get(
                pkg["service_slug"], pkg["service_slug"]
            )
    return jsonify(result)


@app.get("/crm/staff-kpi")
def crm_staff_kpi_page() -> Any:
    redir = _ensure_admin_session_html()
    if redir is not None:
        return redir
    from crm_svc_kpi import (
        get_am_metrics as _am_met,
        get_sp_metrics as _sp_met,
        get_targets as _get_tgt,
        get_latest_kpi_scan as _latest_scan,
    )
    from crm_svc_presales import (
        AM_LEAD_METRIC_LABELS as _am_lead_labels,
        get_am_lead_metrics as _am_lead_met,
    )
    now = datetime.utcnow()
    staff_id = _opt_pos_int(request.args.get("staff_id"))
    year = _opt_pos_int(request.args.get("year")) or now.year
    month = _opt_pos_int(request.args.get("month")) or now.month

    with get_connection() as conn:
        all_staff = [
            dict(r) for r in conn.execute(
                "SELECT id, name FROM crm_staff WHERE active = 1 ORDER BY name"
            ).fetchall()
        ]
        selected_staff = None
        am_metrics = None
        am_lead_metrics = None
        sp_metrics = None
        targets: dict = {}
        latest_am_scan = ""
        latest_sp_scan = ""
        latest_am_lead_scan = ""
        am_lead_cap_alerts: dict[str, Any] = {"over_cap_count": 0, "alerts": []}

        kpi_readiness: dict[str, Any] | None = None
        am_backfill_result: dict[str, int] | None = None

        if staff_id:
            row = conn.execute(
                "SELECT id, name FROM crm_staff WHERE id = ?", (staff_id,)
            ).fetchone()
            if row:
                selected_staff = dict(row)
                from crm_service_lifecycle import (
                    backfill_assigned_am_for_staff as _backfill_am,
                )
                from crm_svc_kpi import get_staff_kpi_readiness as _kpi_ready

                am_backfill_result = _backfill_am(conn, staff_id)
                kpi_readiness = _kpi_ready(conn, staff_id)
                am_metrics = _am_met(conn, staff_id, year, month)
                am_lead_metrics = _am_lead_met(conn, staff_id, year, month)
                from crm_svc_presales import get_am_presales_cap_alerts as _cap_alerts

                am_lead_cap_alerts = _cap_alerts(conn, staff_id)
                sp_metrics = _sp_met(conn, staff_id, year, month)
                targets = _get_tgt(conn, staff_id, year, month)
                latest_am_scan = _latest_scan(conn, staff_id, "am", year, month)
                latest_sp_scan = _latest_scan(conn, staff_id, "sp", year, month)
                latest_am_lead_scan = _latest_scan(conn, staff_id, "am_lead", year, month)

    return render_template(
        "crm_staff_kpi.html",
        all_staff=all_staff,
        selected_staff=selected_staff,
        selected_staff_id=staff_id,
        year=year,
        month=month,
        am_metrics=am_metrics,
        am_lead_metrics=am_lead_metrics,
        am_lead_metric_labels=_am_lead_labels,
        sp_metrics=sp_metrics,
        targets=targets,
        latest_am_scan=latest_am_scan,
        latest_sp_scan=latest_sp_scan,
        latest_am_lead_scan=latest_am_lead_scan,
        am_lead_cap_alerts=am_lead_cap_alerts,
        kpi_readiness=kpi_readiness,
        am_backfill_result=am_backfill_result,
        **_admin_page_template_kwargs(),
    )


@app.get("/api/crm/svc-finance/<int:lifecycle_id>/summary")
def api_svc_finance_summary(lifecycle_id: int) -> Any:
    from crm_svc_finance import get_summary as _fin_sum
    with get_connection() as conn:
        lc_row = conn.execute(
            "SELECT contract_id FROM crm_service_lifecycle WHERE id = ?", (lifecycle_id,)
        ).fetchone()
        if lc_row is None:
            return jsonify({"error": "Không tìm thấy lifecycle"}), 404
        contract_amount_vnd = 0
        if lc_row["contract_id"]:
            c_row = conn.execute(
                "SELECT amount_vnd FROM crm_contracts WHERE id = ?",
                (lc_row["contract_id"],),
            ).fetchone()
            if c_row:
                contract_amount_vnd = int(c_row["amount_vnd"] or 0)
        summary = _fin_sum(conn, lifecycle_id, contract_amount_vnd)
    return jsonify(summary)


@app.post("/api/crm/svc-payments")
def api_svc_payment_create() -> Any:
    from crm_svc_finance import create_payment as _pay_create
    payload = request.get_json(force=True) or {}
    lifecycle_id = _opt_pos_int(payload.get("lifecycle_id"))
    amount_vnd = _opt_pos_int(payload.get("amount_vnd"))
    received_on = str(payload.get("received_on", "")).strip()[:10]
    due_on = str(payload.get("due_on", "")).strip()[:10]
    status = str(payload.get("status", "pending")).strip()
    notes = str(payload.get("notes", "")).strip()
    if not lifecycle_id or amount_vnd is None or not received_on:
        return jsonify({"error": "Cần lifecycle_id, amount_vnd, received_on"}), 400
    with get_connection() as conn:
        pid = _pay_create(
            conn,
            lifecycle_id,
            amount_vnd,
            received_on,
            status,
            notes,
            due_on=due_on,
        )
        row = conn.execute(
            "SELECT * FROM crm_svc_payments WHERE id = ?", (pid,)
        ).fetchone()
    return jsonify(dict(row)), 201


@app.patch("/api/crm/svc-payments/<int:payment_id>")
def api_svc_payment_patch(payment_id: int) -> Any:
    from crm_svc_finance import update_payment as _pay_update
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute(
            "SELECT id FROM crm_svc_payments WHERE id = ?", (payment_id,)
        ).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy payment"}), 404
        amount_vnd = _opt_pos_int(payload.get("amount_vnd")) if "amount_vnd" in payload else None
        _pay_update(
            conn, payment_id,
            amount_vnd=amount_vnd,
            received_on=str(payload["received_on"])[:10] if "received_on" in payload else None,
            due_on=str(payload["due_on"])[:10] if "due_on" in payload else None,
            status=str(payload["status"]) if "status" in payload else None,
            notes=str(payload["notes"]) if "notes" in payload else None,
        )
        updated = conn.execute(
            "SELECT * FROM crm_svc_payments WHERE id = ?", (payment_id,)
        ).fetchone()
    return jsonify(dict(updated))


@app.delete("/api/crm/svc-payments/<int:payment_id>")
def api_svc_payment_delete(payment_id: int) -> Any:
    from crm_svc_finance import delete_payment as _pay_del
    with get_connection() as conn:
        ok = _pay_del(conn, payment_id)
    if not ok:
        return jsonify({"error": "Không tìm thấy payment"}), 404
    return jsonify({"ok": True})


@app.post("/api/crm/svc-expenses")
def api_svc_expense_create() -> Any:
    from crm_svc_finance import ExpenseValidationError, create_expense as _exp_create
    from crm_svc_finance import create_presales_expense as _presales_exp_create
    payload = request.get_json(force=True) or {}
    lifecycle_id = _opt_pos_int(payload.get("lifecycle_id"))
    presales_id = _opt_pos_int(payload.get("presales_id"))
    lead_id = _opt_pos_int(payload.get("lead_id"))
    title = str(payload.get("title", "")).strip()[:500]
    category = str(payload.get("category", "khac")).strip()
    amount_vnd = _opt_pos_int(payload.get("amount_vnd"))
    expense_on = str(payload.get("expense_on", "")).strip()[:10]
    notes = str(payload.get("notes", "")).strip()
    cost_phase = str(payload.get("cost_phase", "")).strip() or None
    lifecycle_stage = str(payload.get("lifecycle_stage", "")).strip() or None
    if not title or amount_vnd is None or not expense_on:
        return jsonify({"error": "Cần title, amount_vnd, expense_on"}), 400
    with get_connection() as conn:
        try:
            if presales_id or lead_id:
                if not _crm_presales_on_lead_enabled():
                    return jsonify({"error": "PTT_PRESALES_ON_LEAD chưa bật"}), 400
                if lifecycle_id:
                    return jsonify({
                        "error": "Dùng presales_id hoặc lead_id — không gửi lifecycle_id"
                    }), 400
                if lead_id:
                    prev = fetch_lead_by_id(conn, lead_id)
                    if not _crm_lead_can_access(conn, prev):
                        return jsonify({"error": "Không có quyền."}), 403
                    if prev is None:
                        return jsonify({"error": "Không tìm thấy lead."}), 404
                eid = _presales_exp_create(
                    conn,
                    title=title,
                    category=category,
                    amount_vnd=amount_vnd,
                    expense_on=expense_on,
                    notes=notes,
                    presales_id=presales_id,
                    lead_id=lead_id,
                    lifecycle_stage=lifecycle_stage,
                )
            else:
                if not lifecycle_id:
                    return jsonify({
                        "error": "Cần lifecycle_id hoặc presales_id/lead_id"
                    }), 400
                eid = _exp_create(
                    conn,
                    lifecycle_id,
                    title,
                    category,
                    amount_vnd,
                    expense_on,
                    notes,
                    cost_phase=cost_phase,
                    lifecycle_stage=lifecycle_stage,
                )
        except ExpenseValidationError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            from crm_svc_presales import PresalesCapExceededError

            if isinstance(exc, PresalesCapExceededError):
                return jsonify({"error": str(exc)}), 400
            raise
        row = conn.execute(
            "SELECT * FROM crm_svc_expenses WHERE id = ?", (eid,)
        ).fetchone()
    return jsonify(dict(row)), 201


@app.get("/api/crm/service-lifecycle/<int:lifecycle_id>/presales-summary")
def api_svc_presales_summary(lifecycle_id: int) -> Any:
    from crm_svc_presales import get_presales_cost_summary as _presales_sum
    with get_connection() as conn:
        row = conn.execute(
            "SELECT id FROM crm_service_lifecycle WHERE id = ?", (lifecycle_id,)
        ).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy lifecycle"}), 404
        summary = _presales_sum(conn, lifecycle_id)
    return jsonify(summary)


@app.patch("/api/crm/svc-expenses/<int:expense_id>")
def api_svc_expense_patch(expense_id: int) -> Any:
    from crm_svc_finance import update_expense as _exp_update
    payload = request.get_json(force=True) or {}
    with get_connection() as conn:
        row = conn.execute(
            "SELECT id FROM crm_svc_expenses WHERE id = ?", (expense_id,)
        ).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy expense"}), 404
        amount_vnd = _opt_pos_int(payload.get("amount_vnd")) if "amount_vnd" in payload else None
        _exp_update(
            conn, expense_id,
            title=str(payload["title"])[:500] if "title" in payload else None,
            category=str(payload["category"]) if "category" in payload else None,
            amount_vnd=amount_vnd,
            expense_on=str(payload["expense_on"])[:10] if "expense_on" in payload else None,
            notes=str(payload["notes"]) if "notes" in payload else None,
        )
        updated = conn.execute(
            "SELECT * FROM crm_svc_expenses WHERE id = ?", (expense_id,)
        ).fetchone()
    return jsonify(dict(updated))


@app.delete("/api/crm/svc-expenses/<int:expense_id>")
def api_svc_expense_delete(expense_id: int) -> Any:
    from crm_svc_finance import delete_expense as _exp_del
    with get_connection() as conn:
        ok = _exp_del(conn, expense_id)
    if not ok:
        return jsonify({"error": "Không tìm thấy expense"}), 404
    return jsonify({"ok": True})


@app.post("/api/crm/svc-finance/<int:lifecycle_id>/ai-scan")
def api_svc_finance_ai_scan(lifecycle_id: int) -> Any:
    from crm_svc_finance import (
        get_summary as _fin_sum,
        run_ai_finance_scan as _fin_scan_fn,
    )
    from crm_svc_tasks import SERVICE_LABELS as _svc_labels
    payload = request.get_json(force=True) or {}
    scan_type = str(payload.get("scan_type", "health")).strip()
    if scan_type not in ("health", "forecast"):
        return jsonify({"error": "scan_type phải là 'health' hoặc 'forecast'"}), 400
    with get_connection() as conn:
        lc = conn.execute(
            "SELECT * FROM crm_service_lifecycle WHERE id = ?", (lifecycle_id,)
        ).fetchone()
        if lc is None:
            return jsonify({"error": "Không tìm thấy lifecycle"}), 404
        lc = dict(lc)
        contract_amount_vnd = 0
        days_elapsed = 0
        contract_days = 0
        if lc.get("contract_id"):
            c_row = conn.execute(
                "SELECT amount_vnd, starts_on, ends_on FROM crm_contracts WHERE id = ?",
                (lc["contract_id"],),
            ).fetchone()
            if c_row:
                contract_amount_vnd = int(c_row["amount_vnd"] or 0)
                try:
                    from datetime import date
                    today = date.today()
                    if c_row["starts_on"]:
                        start = date.fromisoformat(c_row["starts_on"][:10])
                        days_elapsed = max(0, (today - start).days)
                        if c_row["ends_on"]:
                            end = date.fromisoformat(c_row["ends_on"][:10])
                            contract_days = max(0, (end - start).days)
                except (ValueError, TypeError):
                    pass
        summary = _fin_sum(conn, lifecycle_id, contract_amount_vnd)
        customer_name = "KH"
        if lc.get("customer_id"):
            cust = conn.execute(
                "SELECT name FROM crm_customers WHERE id = ?", (lc["customer_id"],)
            ).fetchone()
            if cust:
                customer_name = cust["name"] or "KH"
        ctx = {
            "service_name": _svc_labels.get(lc["service_slug"], lc["service_slug"]),
            "customer_name": customer_name,
            "contract_amount_vnd": summary["expected_revenue"],
            "received_revenue": summary["received_revenue"],
            "total_expenses": summary["delivery_expenses"],
            "profit": summary["profit"],
            "margin_pct": summary["margin_pct"],
            "days_elapsed": days_elapsed,
            "contract_days": contract_days,
        }
        output = _fin_scan_fn(conn, lifecycle_id, scan_type, ctx)
    return jsonify({"ai_output": output, "scan_type": scan_type, "lifecycle_id": lifecycle_id})


@app.get("/api/crm/staff-kpi/<int:staff_id>/metrics")
def api_staff_kpi_metrics(staff_id: int) -> Any:
    from crm_svc_kpi import get_am_metrics as _am_met, get_sp_metrics as _sp_met
    role = request.args.get("role", "am")
    year = _opt_pos_int(request.args.get("year")) or datetime.utcnow().year
    month = _opt_pos_int(request.args.get("month")) or datetime.utcnow().month
    with get_connection() as conn:
        staff = conn.execute(
            "SELECT id, name FROM crm_staff WHERE id = ?", (staff_id,)
        ).fetchone()
        if staff is None:
            return jsonify({"error": "Không tìm thấy staff"}), 404
        if role == "am":
            metrics = _am_met(conn, staff_id, year, month)
        else:
            metrics = _sp_met(conn, staff_id, year, month)
    return jsonify({"staff_id": staff_id, "role": role, "year": year, "month": month, **metrics})


@app.get("/api/crm/staff-kpi/<int:staff_id>/lead-metrics")
def api_staff_kpi_lead_metrics(staff_id: int) -> Any:
    from crm_svc_presales import get_am_lead_metrics as _am_lead_met
    year = _opt_pos_int(request.args.get("year")) or datetime.utcnow().year
    month = _opt_pos_int(request.args.get("month")) or datetime.utcnow().month
    with get_connection() as conn:
        staff = conn.execute(
            "SELECT id, name FROM crm_staff WHERE id = ?", (staff_id,)
        ).fetchone()
        if staff is None:
            return jsonify({"error": "Không tìm thấy staff"}), 404
        metrics = _am_lead_met(conn, staff_id, year, month)
    return jsonify({"staff_id": staff_id, "year": year, "month": month, **metrics})


@app.get("/api/crm/service-lifecycle/funnel-stats")
def api_service_lifecycle_funnel_stats() -> Any:
    from crm_svc_presales import get_funnel_stats as _funnel_stats
    am_id = _opt_pos_int(request.args.get("am_id"))
    service_slug = str(request.args.get("service_slug") or "").strip() or None
    period_start = str(request.args.get("from") or "").strip()[:10] or None
    period_end = str(request.args.get("to") or "").strip()[:10] or None
    with get_connection() as conn:
        stats = _funnel_stats(
            conn,
            am_id=am_id,
            service_slug=service_slug,
            period_start=period_start,
            period_end=period_end,
        )
    return jsonify(stats)


@app.get("/api/crm/service-lifecycle/<int:lifecycle_id>/consult-brief")
def api_svc_consult_brief(lifecycle_id: int) -> Any:
    from crm_svc_consult_bridge import get_consult_brief as _consult_brief
    with get_connection() as conn:
        try:
            brief = _consult_brief(conn, lifecycle_id)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 404
    return jsonify(brief)


@app.post("/api/crm/service-lifecycle/<int:lifecycle_id>/consult-prefill")
def api_svc_consult_prefill(lifecycle_id: int) -> Any:
    from crm_svc_consult_bridge import prefill_consult_task as _prefill
    payload = request.get_json(force=True) or {}
    overwrite = bool(payload.get("overwrite"))
    with get_connection() as conn:
        try:
            result = _prefill(conn, lifecycle_id, overwrite=overwrite)
            conn.commit()
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 404
    return jsonify(result)


@app.get("/api/crm/service-lifecycle/<int:lifecycle_id>/funnel-progress")
def api_svc_lifecycle_funnel_progress(lifecycle_id: int) -> Any:
    from crm_svc_consult_bridge import get_lifecycle_funnel_progress as _lc_funnel
    with get_connection() as conn:
        try:
            progress = _lc_funnel(conn, lifecycle_id)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 404
    return jsonify(progress)


@app.post("/api/crm/staff-kpi/<int:staff_id>/backfill-am")
def api_staff_kpi_backfill_am(staff_id: int) -> Any:
    from crm_service_lifecycle import backfill_assigned_am_for_staff as _backfill_am
    from crm_svc_kpi import get_staff_kpi_readiness as _kpi_ready

    with get_connection() as conn:
        staff = conn.execute(
            "SELECT id, name FROM crm_staff WHERE id = ?", (staff_id,)
        ).fetchone()
        if staff is None:
            return jsonify({"error": "Không tìm thấy staff"}), 404
        result = _backfill_am(conn, staff_id)
        readiness = _kpi_ready(conn, staff_id)
    return jsonify({"ok": True, **result, "readiness": readiness})


@app.post("/api/crm/staff-kpi/<int:staff_id>/targets")
def api_staff_kpi_set_target(staff_id: int) -> Any:
    from crm_svc_kpi import set_target as _set_tgt
    payload = request.get_json(force=True) or {}
    role = str(payload.get("role", "am")).strip()
    metric_key = str(payload.get("metric_key", "")).strip()
    year = _opt_pos_int(payload.get("year"))
    month = _opt_pos_int(payload.get("month"))
    target_value = payload.get("target_value")
    if not metric_key or not year or not month or target_value is None:
        return jsonify({"error": "Cần metric_key, year, month, target_value"}), 400
    try:
        target_value = float(target_value)
    except (TypeError, ValueError):
        return jsonify({"error": "target_value phải là số"}), 400
    with get_connection() as conn:
        _set_tgt(conn, staff_id, role, metric_key, year, month, target_value)
    return jsonify({
        "ok": True, "staff_id": staff_id, "metric_key": metric_key,
        "year": year, "month": month, "target_value": target_value,
    })


@app.post("/api/crm/staff-kpi/<int:staff_id>/ai-scan")
def api_staff_kpi_ai_scan(staff_id: int) -> Any:
    from crm_svc_kpi import (
        get_am_metrics as _am_met,
        get_sp_metrics as _sp_met,
        get_targets as _get_tgt,
        run_ai_kpi_scan as _ai_scan,
    )
    payload = request.get_json(force=True) or {}
    role = str(payload.get("role", "am")).strip()
    year = _opt_pos_int(payload.get("year")) or datetime.utcnow().year
    month = _opt_pos_int(payload.get("month")) or datetime.utcnow().month
    if role not in ("am", "sp", "am_lead"):
        return jsonify({"error": "role phải là 'am', 'sp' hoặc 'am_lead'"}), 400
    with get_connection() as conn:
        staff = conn.execute(
            "SELECT name FROM crm_staff WHERE id = ?", (staff_id,)
        ).fetchone()
        if staff is None:
            return jsonify({"error": "Không tìm thấy staff"}), 404
        targets = _get_tgt(conn, staff_id, year, month)
        if role == "am_lead":
            from crm_svc_kpi import run_ai_lead_kpi_scan as _lead_ai_scan
            from crm_svc_presales import (
                get_am_lead_metrics as _am_lead_met,
                get_am_presales_cap_alerts as _cap_alerts,
            )
            metrics = _am_lead_met(conn, staff_id, year, month)
            cap_alerts = _cap_alerts(conn, staff_id)
            go_n = int(metrics.get("lead_go_decisions") or 0)
            cost = int(metrics.get("presales_cost_vnd") or 0)
            ctx = {
                "staff_name": staff["name"],
                "month": month,
                "year": year,
                "lead_intake_completed": int(metrics.get("lead_intake_completed") or 0),
                "lead_phone_within_48h_pct": float(
                    metrics.get("lead_phone_within_48h_pct") or 0
                ),
                "lead_phone_within_48h_num": int(
                    metrics.get("lead_phone_within_48h_num") or 0
                ),
                "lead_phone_within_48h_denom": int(
                    metrics.get("lead_phone_within_48h_denom") or 0
                ),
                "lead_go_decisions": go_n,
                "lead_to_consult_pct": float(metrics.get("lead_to_consult_pct") or 0),
                "lead_to_consult_num": int(metrics.get("lead_to_consult_num") or 0),
                "lead_to_consult_denom": int(metrics.get("lead_to_consult_denom") or 0),
                "presales_cost_vnd": cost,
                "presales_cost_per_go_vnd": int(cost / go_n) if go_n > 0 else 0,
                "lead_avg_phone_minutes": float(
                    metrics.get("lead_avg_phone_minutes") or 0
                ),
                "target_lead_intake_completed": int(
                    targets.get("lead_intake_completed", 0)
                ),
                "target_lead_phone_within_48h_pct": float(
                    targets.get("lead_phone_within_48h_pct", 0)
                ),
                "target_lead_to_consult_pct": float(
                    targets.get("lead_to_consult_pct", 0)
                ),
                "target_presales_cost_vnd": int(targets.get("presales_cost_vnd", 0)),
                "presales_over_cap_count": int(cap_alerts.get("over_cap_count") or 0),
            }
            output = _lead_ai_scan(conn, staff_id, year, month, ctx)
        elif role == "am":
            metrics = _am_met(conn, staff_id, year, month)
            ctx = {
                "staff_name": staff["name"],
                "month": month,
                "year": year,
                "received_revenue": metrics["received_revenue"],
                "active_services": metrics["active_services"],
                "avg_margin_pct": metrics["avg_margin_pct"],
                "outstanding": metrics["outstanding"],
                "target_received_revenue": int(targets.get("received_revenue", 0)),
                "target_active_services": int(targets.get("active_services", 0)),
                "target_avg_margin_pct": float(targets.get("avg_margin_pct", 0)),
            }
            output = _ai_scan(conn, staff_id, role, year, month, ctx)
        else:
            metrics = _sp_met(conn, staff_id, year, month)
            ctx = {
                "staff_name": staff["name"],
                "month": month,
                "year": year,
                "tasks_completed": metrics["tasks_completed"],
                "tasks_pending": metrics["tasks_pending"],
                "risks_resolved": metrics["risks_resolved"],
                "target_tasks_completed": int(targets.get("tasks_completed", 0)),
                "target_risks_resolved": int(targets.get("risks_resolved", 0)),
            }
            output = _ai_scan(conn, staff_id, role, year, month, ctx)
    return jsonify({"ai_output": output, "staff_id": staff_id, "role": role})


@app.get("/api/crm/svc-lifecycle/<int:lifecycle_id>/staff-metrics")
def api_svc_lifecycle_staff_metrics(lifecycle_id: int) -> Any:
    from crm_svc_kpi import get_lifecycle_staff_metrics as _lc_staff
    with get_connection() as conn:
        result = _lc_staff(conn, lifecycle_id)
    return jsonify(result)


@app.get("/api/crm/service-lifecycle")
def api_svc_lifecycle_list() -> Any:
    service_slug = request.args.get("service_slug") or None
    am_id = _opt_pos_int(request.args.get("am_id"))
    include_draft = request.args.get("include_draft", "0") == "1"
    with get_connection() as conn:
        rows = _svc_list_active(conn, service_slug=service_slug, am_id=am_id, include_draft=include_draft)
    return jsonify(rows)


@app.post("/api/crm/service-lifecycle")
def api_svc_lifecycle_create() -> Any:
    payload = request.get_json(force=True) or {}
    lead_id = _opt_pos_int(payload.get("lead_id"))
    service_slug = str(payload.get("service_slug", "")).strip()
    if not lead_id or not service_slug:
        return jsonify({"error": "Cần lead_id và service_slug"}), 400
    from crm_service_lifecycle import create_draft_lifecycle
    with get_connection() as conn:
        lid = create_draft_lifecycle(conn, lead_id=lead_id, service_slug=service_slug, suggested_by="human")
        row = conn.execute("SELECT * FROM crm_service_lifecycle WHERE id = ?", (lid,)).fetchone()
    return jsonify(dict(row)), 201


@app.get("/crm/customers/<int:customer_id>/lifecycle/new")
def crm_customer_lifecycle_new(customer_id: int) -> Any:
    with get_connection() as conn:
        cu = conn.execute("SELECT id, name FROM crm_customers WHERE id = ?", (customer_id,)).fetchone()
        if cu is None:
            return "Không tìm thấy khách hàng", 404
    from crm_service_lifecycle import VALID_SLUGS
    slugs = sorted(VALID_SLUGS)
    return render_template("crm_lifecycle_new.html", customer=dict(cu), slugs=slugs)


@app.post("/crm/customers/<int:customer_id>/lifecycle/new")
def crm_customer_lifecycle_new_post(customer_id: int) -> Any:
    service_slug = request.form.get("service_slug", "").strip()
    from crm_service_lifecycle import create_draft_lifecycle, VALID_SLUGS
    if not service_slug or service_slug not in VALID_SLUGS:
        return "service_slug không hợp lệ", 400
    with get_connection() as conn:
        cu = conn.execute("SELECT id FROM crm_customers WHERE id = ?", (customer_id,)).fetchone()
        if cu is None:
            return "Không tìm thấy khách hàng", 404
        lid = create_draft_lifecycle(conn, lead_id=None, service_slug=service_slug,
                                     suggested_by="human", customer_id=customer_id)
    return redirect(url_for("crm_service_workflow_page", lifecycle_id=lid))


@app.post("/api/crm/customers/<int:customer_id>/lifecycle")
def api_customer_lifecycle_create(customer_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    service_slug = str(payload.get("service_slug", "")).strip()
    if not service_slug:
        return jsonify({"error": "Cần service_slug"}), 400
    from crm_service_lifecycle import create_draft_lifecycle
    with get_connection() as conn:
        cu = conn.execute("SELECT id FROM crm_customers WHERE id = ?", (customer_id,)).fetchone()
        if cu is None:
            return jsonify({"error": "Không tìm thấy khách hàng"}), 404
        lid = create_draft_lifecycle(
            conn, lead_id=None, service_slug=service_slug,
            suggested_by="human", customer_id=customer_id,
        )
        row = conn.execute("SELECT * FROM crm_service_lifecycle WHERE id = ?", (lid,)).fetchone()
    return jsonify(dict(row)), 201


@app.patch("/api/crm/service-lifecycle/<int:lifecycle_id>")
def api_svc_lifecycle_patch(lifecycle_id: int) -> Any:
    payload = request.get_json(force=True) or {}
    to_stage = str(payload.get("stage", "")).strip()
    notes = str(payload.get("notes", "")).strip()[:2000]
    actor_id = _opt_pos_int(payload.get("actor_id"))
    override_reason = str(payload.get("override_reason") or "").strip()[:500]
    confirm = bool(payload.get("confirm"))
    gate_warnings: list[str] = []
    with get_connection() as conn:
        row = conn.execute(
            "SELECT id, stage FROM crm_service_lifecycle WHERE id = ?", (lifecycle_id,)
        ).fetchone()
        if row is None:
            return jsonify({"error": "Không tìm thấy lifecycle"}), 404
        if to_stage:
            if to_stage not in SVC_LIFECYCLE_STAGES:
                return jsonify({"error": f"Stage không hợp lệ: {to_stage}"}), 400
            if to_stage == "consult" and str(row["stage"] or "") == "lead":
                from crm_svc_consult_bridge import validate_consult_advance as _consult_gate

                gate = _consult_gate(
                    conn,
                    lifecycle_id,
                    override_reason=override_reason,
                    allow_override=_admin_full_access(),
                )
                if not gate.get("ok"):
                    return jsonify({
                        "error": (gate.get("messages") or ["Không thể chuyển Consult"])[0],
                        "gate": gate,
                        "requires_override": bool(gate.get("requires_override")),
                    }), 400
                if gate.get("requires_confirm") and not confirm:
                    return jsonify({
                        "error": (gate.get("messages") or ["Cần xác nhận"])[0],
                        "gate": gate,
                        "requires_confirm": True,
                    }), 400
                gate_warnings = list(gate.get("messages") or [])
                if override_reason:
                    notes = f"{notes}\nDirector override: {override_reason}".strip()[:2000]
            try:
                from crm_service_lifecycle import StageAdvanceError

                _svc_advance_stage(conn, lifecycle_id, to_stage, actor_id=actor_id, notes=notes)
            except StageAdvanceError as exc:
                return jsonify({"error": str(exc)}), 400
            if to_stage == "retain":
                try:
                    from crm_service_lifecycle import check_kpi_alert_async
                    check_kpi_alert_async(lifecycle_id=lifecycle_id, db_path=str(DB_PATH))
                except Exception as _ka_exc:
                    logger.warning("KPI alert trigger lỗi: %s", _ka_exc)
        if "service_slug" in payload:
            slug = str(payload["service_slug"]).strip()
            ts = _crm_ts()
            conn.execute(
                "UPDATE crm_service_lifecycle SET service_slug = ?, updated_at = ? WHERE id = ?",
                (slug, ts, lifecycle_id),
            )
            conn.commit()
        if "presales_cost_cap_vnd" in payload:
            from crm_svc_presales import set_presales_cost_cap as _set_presales_cap
            cap_val = _opt_pos_int(payload.get("presales_cost_cap_vnd"))
            _set_presales_cap(conn, lifecycle_id, cap_val)
        if payload.get("suggest_sp_from_tasks"):
            from crm_service_lifecycle import sync_assigned_sp_from_tasks

            sync_assigned_sp_from_tasks(conn, lifecycle_id, overwrite=False)
        if "assigned_sp_id" in payload:
            from crm_service_lifecycle import set_assigned_sp as _set_assigned_sp

            raw_sp = payload.get("assigned_sp_id")
            sp_id = _opt_pos_int(raw_sp) if raw_sp not in (None, "", 0) else None
            try:
                _set_assigned_sp(conn, lifecycle_id, sp_id, overwrite=True)
            except ValueError as exc:
                return jsonify({"error": str(exc)}), 400
        updated = conn.execute(
            "SELECT * FROM crm_service_lifecycle WHERE id = ?", (lifecycle_id,)
        ).fetchone()
    out = dict(updated)
    if gate_warnings:
        out["gate_warnings"] = gate_warnings
    return jsonify(out)


@app.get("/api/crm/service-lifecycle/<int:lifecycle_id>/events")
def api_svc_lifecycle_events(lifecycle_id: int) -> Any:
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT * FROM crm_service_lifecycle_events WHERE lifecycle_id = ? ORDER BY id ASC",
            (lifecycle_id,),
        ).fetchall()
    return jsonify([dict(r) for r in rows])


from blueprints import register_crm_product_blueprints

register_crm_product_blueprints(app)

init_db()

if __name__ == "__main__":
    start_facebook_autosync_worker(app)
    _port = int(os.environ.get("PORT", "5050"))
    _debug = os.environ.get("FLASK_DEBUG", "1").strip().lower() in {"1", "true", "yes"}
    print("\n" + "=" * 62)
    print(f"  PTT Advertising — trang chủ: http://127.0.0.1:{_port}/")
    print(f"  Thiết bị khác (Wi‑Fi/LAN): http://<IP-máy-này>:{_port}/")
    print(f"  Kiểm tra đúng app: curl -s http://127.0.0.1:{_port}/healthz")
    print("=" * 62 + "\n")
    # host 0.0.0.0: truy cập từ điện thoại / máy khác trong mạng
    app.run(debug=_debug, host="0.0.0.0", port=_port, use_reloader=False)
