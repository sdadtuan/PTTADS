"""Tests: HDSD Markdown → Excel."""
from __future__ import annotations

import unittest
import zipfile
from io import BytesIO

from crm_hdsd_docs import list_hdsd_catalog, read_hdsd_doc
from crm_hdsd_excel import build_hdsd_all_zip, build_hdsd_doc_xlsx, parse_markdown_blocks
from openpyxl import load_workbook


SAMPLE_MD = """# Tiêu đề test

## Mục A

Đoạn văn **đậm** và [link](/crm/leads).

| Cột 1 | Cột 2 |
|-------|-------|
| a | b |

- Bullet 1
- Bullet 2

```mermaid
flowchart LR
  A --> B
```
"""


class TestHdsdExcel(unittest.TestCase):
    def test_parse_blocks(self) -> None:
        blocks = parse_markdown_blocks(SAMPLE_MD)
        types = [b["type"] for b in blocks]
        self.assertIn("heading", types)
        self.assertIn("table", types)
        self.assertIn("list", types)
        self.assertIn("code", types)

    def test_build_workbook_sheets(self) -> None:
        buf = build_hdsd_doc_xlsx(
            SAMPLE_MD,
            {"title": "Test", "section": "crm", "slug": "test", "filename": "test.md", "section_label": "CRM"},
        )
        wb = load_workbook(BytesIO(buf.getvalue()))
        self.assertEqual({"Thong_tin", "Muc_luc", "Noi_dung", "Bang"}, set(wb.sheetnames))
        ws = wb["Noi_dung"]
        self.assertGreater(ws.max_row, 2)

    def test_export_all_zip(self) -> None:
        catalog = list_hdsd_catalog()
        buf = build_hdsd_all_zip(catalog, read_hdsd_doc)
        with zipfile.ZipFile(BytesIO(buf.getvalue())) as zf:
            names = zf.namelist()
        self.assertTrue(any(n.endswith(".xlsx") for n in names))
        self.assertTrue(any(n.startswith("crm/") for n in names))


if __name__ == "__main__":
    unittest.main()
