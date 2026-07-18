"""Tests: CRM test cases Excel workbook."""
from __future__ import annotations

import unittest
from io import BytesIO

from crm_test_cases_workbook import build_crm_test_cases_workbook
from openpyxl import load_workbook


class TestCrmTestCasesWorkbook(unittest.TestCase):
    def test_build_workbook_sheets(self) -> None:
        buf = build_crm_test_cases_workbook()
        wb = load_workbook(BytesIO(buf.getvalue()))
        expected = {
            "Huong_dan_tester",
            "Flow_Index",
            "Test_Cases",
            "CRM_Lifecycle",
            "Flow_Nguon_Lead",
            "Flow_Service_Delivery",
            "Flow_CSKH",
            "Flow_Hub_MKT",
            "So_do_Lead_Retain",
            "So_do_Presales_Gate",
            "So_do_Delivery_Gate",
            "So_do_KPI",
            "So_do_Nguon_Lead",
            "So_do_Hub_MKT",
            "So_do_CSKH",
            "So_do_Finance",
            "So_do_Portal",
            "Hinh_minh_hoa",
            "Smoke_P0",
            "Tai_khoan_mau",
            "Tong_quan",
        }
        self.assertTrue(expected.issubset(set(wb.sheetnames)))

    def test_main_sheet_columns(self) -> None:
        buf = build_crm_test_cases_workbook()
        wb = load_workbook(BytesIO(buf.getvalue()))
        ws = wb["Test_Cases"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 24)]
        self.assertEqual(headers[0], "STT")
        self.assertTrue(any("Hình minh họa" in str(h) for h in headers if h))
        self.assertTrue(any("Evidence" in str(h) for h in headers if h))
        self.assertEqual(len([h for h in headers if h]), 23)
        self.assertGreater(ws.max_row, 150)

    def test_lifecycle_rows(self) -> None:
        buf = build_crm_test_cases_workbook()
        wb = load_workbook(BytesIO(buf.getvalue()))
        ws = wb["CRM_Lifecycle"]
        ids = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
        self.assertIn("TC-CRM-L01", ids)
        self.assertIn("TC-CRM-L26", ids)

    def test_system_flow_rows(self) -> None:
        buf = build_crm_test_cases_workbook()
        wb = load_workbook(BytesIO(buf.getvalue()))
        ws = wb["Flow_Nguon_Lead"]
        ids = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
        self.assertIn("TC-FLOW-L01", ids)
        self.assertIn("TC-FLOW-L06", ids)

    def test_flow_index_counts(self) -> None:
        buf = build_crm_test_cases_workbook()
        wb = load_workbook(BytesIO(buf.getvalue()))
        ws = wb["Flow_Index"]
        modules = {ws.cell(row=r, column=1).value for r in range(4, ws.max_row + 1) if ws.cell(row=r, column=1).value}
        self.assertIn("CRM Lifecycle", modules)
        self.assertIn("Nguồn Lead", modules)


if __name__ == "__main__":
    unittest.main()
