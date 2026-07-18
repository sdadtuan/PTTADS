"""Bộ Excel test case CRM — đầy đủ cột + sơ đồ cho tester."""
from __future__ import annotations

import csv
import json
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR = Path(__file__).resolve().parent
FIXTURES = BASE_DIR / "tests" / "fixtures" / "test_data"
REGISTRY = FIXTURES / "test_cases_registry.csv"
CRM_LIFECYCLE = FIXTURES / "crm_lifecycle_test_cases.csv"
SYSTEM_FLOWS = FIXTURES / "system_flows_test_cases.csv"
MANIFEST = FIXTURES / "manifest.json"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SUB_FILL = PatternFill("solid", fgColor="D6E4F0")
P0_FILL = PatternFill("solid", fgColor="FCE4D6")
P1_FILL = PatternFill("solid", fgColor="FFF2CC")
P2_FILL = PatternFill("solid", fgColor="E2EFDA")
BOX_FILL = PatternFill("solid", fgColor="E8F4EA")
ARROW_FILL = PatternFill("solid", fgColor="F5F5F5")
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

TEST_CASE_HEADERS = [
    "STT",
    "TC-ID",
    "Ưu tiên",
    "Module",
    "Nhóm / Luồng",
    "Loại test",
    "Môi trường",
    "URL / Màn hình",
    "Vai trò",
    "Tiền điều kiện",
    "Dữ liệu test (file · key)",
    "Bước thực hiện (chi tiết)",
    "Kết quả mong đợi",
    "Hình minh họa (chụp màn hình)",
    "Sơ đồ tham chiếu",
    "Test tự động (pytest)",
    "Kết quả thực tế",
    "Trạng thái",
    "Tester",
    "Ngày test",
    "Evidence (ảnh/log)",
    "Bug ID / Ticket",
    "Ghi chú",
]

STATUS_DV = DataValidation(
    type="list",
    formula1='"Pass,Fail,Blocked,Skip,Not Run"',
    allow_blank=True,
)
PRIORITY_DV = DataValidation(
    type="list",
    formula1='"P0,P1,P2"',
    allow_blank=True,
)
ENV_DV = DataValidation(
    type="list",
    formula1='"Staging,UAT,Production,Local"',
    allow_blank=True,
)
TYPE_DV = DataValidation(
    type="list",
    formula1='"Manual,Automated,Smoke,Regression"',
    allow_blank=True,
)

# Gợi ý bước + hình minh họa theo TC-ID (bổ sung registry)
ENRICHMENT: dict[str, dict[str, str]] = {
    "TC-AUTH-01": {
        "pre": "Tài khoản admin tồn tại (.env ADMIN_USERNAME/PASSWORD)",
        "steps": "1. Mở /admin/login\n2. Nhập user/pass\n3. Submit",
        "screenshot": "Chụp sidebar đầy đủ sau login",
        "diagram": "—",
        "type": "Smoke",
        "role": "Admin",
        "url": "/admin/login",
    },
    "TC-LEAD-01": {
        "pre": "Quyền create lead; flag PTT_PRESALES_ON_LEAD=1 (nếu test presales)",
        "steps": "1. /crm/leads → + Lead\n2. Điền form\n3. Lưu\n4. Mở chi tiết",
        "screenshot": "Form tạo lead + timeline activities",
        "diagram": "So_do_Lead_Retain",
        "type": "Manual",
        "role": "AM / MKT",
        "url": "/crm/leads",
    },
    "TC-CRM-L02": {
        "pre": "Lead mới; care pipeline B2 chưa done",
        "steps": "1. Mở lead\n2. Tab Chăm sóc / care\n3. Hoàn thành B2 Liên hệ lần đầu",
        "screenshot": "Banner gate + trạng thái B2 ✓",
        "diagram": "So_do_Lead_Retain",
        "type": "Manual",
        "role": "AM",
        "url": "/crm/leads/{id}",
    },
    "TC-CRM-L07": {
        "pre": "Pre-sales 3 tab ✓; HĐ draft",
        "steps": "1. /crm/hub\n2. Ký HĐ Active\n3. Kiểm tra /crm/service-delivery",
        "screenshot": "HĐ Active + workflow Onboard",
        "diagram": "So_do_Lead_Retain",
        "type": "Manual",
        "role": "AM",
        "url": "/crm/hub",
    },
    "TC-CRM-L16": {
        "pre": "Lead mới; care B2 chưa tick",
        "steps": "1. Mở lead\n2. Thử mở tab Pre-sales\n3. Quan sát banner gate",
        "screenshot": "Banner B2 + tab PS disabled",
        "diagram": "So_do_Presales_Gate",
        "type": "Manual",
        "role": "AM",
        "url": "/crm/leads/{id}",
    },
    "TC-CRM-L17": {
        "pre": "Intake form; lead đủ điều kiện No-Go",
        "steps": "1. /crm/intake?lead_id=\n2. Chọn No-Go + lý do\n3. Thử sang Consult",
        "screenshot": "Form No-Go + block Consult",
        "diagram": "So_do_Presales_Gate",
        "type": "Manual",
        "role": "AM",
        "url": "/crm/intake?lead_id=",
    },
    "TC-CRM-L19": {
        "pre": "Workflow Onboard/Deliver; task stage < 100%",
        "steps": "1. Mở /crm/service-delivery/{id}\n2. Bấm Chuyển →\n3. Đọc message gate",
        "screenshot": "Task chưa ✓ + toast/block",
        "diagram": "So_do_Delivery_Gate",
        "type": "Manual",
        "role": "AM / SP",
        "url": "/crm/service-delivery/{id}",
    },
    "TC-FLOW-L01": {
        "pre": "Webhook secret Zalo cấu hình; payload mẫu",
        "steps": "1. POST /api/crm/integration/webhooks/zalo\n2. Kiểm tra lead mới\n3. Verify source=zalo",
        "screenshot": "Response 201 + lead list",
        "diagram": "So_do_Nguon_Lead",
        "type": "Automated",
        "role": "MKT / System",
        "url": "POST /api/crm/integration/webhooks/zalo",
    },
    "TC-FLOW-L06": {
        "pre": "Rule auto-assign bật trong cấu hình lead",
        "steps": "1. Tạo lead mới (web/API)\n2. Refresh /crm/leads\n3. Kiểm tra owner_id",
        "screenshot": "Lead + owner đúng rule",
        "diagram": "So_do_Nguon_Lead",
        "type": "Automated",
        "role": "System",
        "url": "/crm/leads",
    },
    "TC-FLOW-C01": {
        "pre": "Đăng nhập CSKH; quyền /crm",
        "steps": "1. /crm kanban\n2. + Case mới\n3. Kiểm tra cột Mới",
        "screenshot": "Kanban + card case",
        "diagram": "So_do_CSKH",
        "type": "Manual",
        "role": "CSKH",
        "url": "/crm",
    },
    "TC-FLOW-M01": {
        "pre": "Lead pre-sales ✓; HĐ draft trong Hub",
        "steps": "1. /crm/hub\n2. Active HĐ\n3. Kiểm tra lifecycle promote",
        "screenshot": "HĐ Active + stage Onboard",
        "diagram": "So_do_Hub_MKT",
        "type": "Manual",
        "role": "AM",
        "url": "/crm/hub",
    },
    "TC-FLOW-M04": {
        "pre": "Template SOP Launch 14 ngày có sẵn",
        "steps": "1. /crm/sop\n2. Chọn template\n3. Chạy checklist 14 bước",
        "screenshot": "SOP progress + ngày",
        "diagram": "So_do_Hub_MKT",
        "type": "Manual",
        "role": "MKT",
        "url": "/crm/sop",
    },
    "TC-FLOW-P01": {
        "pre": "NV portal đăng nhập; ngày chưa báo cáo",
        "steps": "1. /crm/daily-reports\n2. Điền form\n3. Submit",
        "screenshot": "Form + status submitted",
        "diagram": "So_do_Portal",
        "type": "Manual",
        "role": "NV",
        "url": "/crm/daily-reports",
    },
    "TC-FLOW-Q03": {
        "pre": "Quyền admin hoặc QA",
        "steps": "1. /crm/hdsd hoặc topbar Test case\n2. Tải download.xlsx\n3. Mở Excel kiểm tra sheets",
        "screenshot": "File tải + sheet Flow_Index",
        "diagram": "So_do_QA",
        "type": "Smoke",
        "role": "QA",
        "url": "/crm/test-cases/download.xlsx",
    },
    "TC-FLOW-N01": {
        "pre": "Lifecycle thiếu assigned_am; KPI = 0",
        "steps": "1. /crm/staff-kpi\n2. Quan sát banner vàng\n3. Bấm đồng bộ AM",
        "screenshot": "Banner gap + sau sync",
        "diagram": "So_do_KPI",
        "type": "Manual",
        "role": "AM / Admin",
        "url": "/crm/staff-kpi",
    },
}

FLOW_INDEX: list[tuple[str, str, str, str]] = [
    ("CRM Lifecycle", "CRM_Lifecycle", "So_do_Lead_Retain", "Lead → Pre-sales → HĐ → Onboard → Retain"),
    ("Nguồn Lead", "Flow_Nguon_Lead", "So_do_Nguon_Lead", "Webhook, webform, API, phân lead"),
    ("Service Delivery", "Flow_Service_Delivery", "So_do_Delivery_Gate", "Kanban workflow + task gate"),
    ("CSKH", "Flow_CSKH", "So_do_CSKH", "Case kanban + care report"),
    ("Hub · MKT · Sales", "Flow_Hub_MKT", "So_do_Hub_MKT", "HĐ, chiến dịch, SOP, sales pipeline"),
    ("Tài chính", "Flow_Tai_chinh", "So_do_Finance", "KPI alert, RE project, chi phí lifecycle"),
    ("Portal · Payroll", "Flow_Portal", "So_do_Portal", "Báo cáo ngày, KPI NV, chấm công"),
    ("Product Model", "Flow_Product", "So_do_Product", "Catalog, addon, review queue"),
    ("HDSD · QA", "Flow_QA", "So_do_QA", "Tài liệu + bộ test case Excel"),
    ("Xử lý sự cố", "Flow_Xu_ly_su_co", "So_do_KPI", "KPI 0, orphan, flag, quyền"),
    ("Auth", "Flow_Auth", "So_do_Auth", "Login admin + portal NV"),
]

FLOW_SHEET_MODULES: dict[str, list[str]] = {
    "Flow_Nguon_Lead": ["Nguồn Lead"],
    "Flow_Service_Delivery": ["Service Delivery"],
    "Flow_CSKH": ["CSKH"],
    "Flow_Hub_MKT": ["Hub", "MKT Plan", "Kinh doanh"],
    "Flow_Tai_chinh": ["Tài chính"],
    "Flow_Portal": ["Portal", "Payroll"],
    "Flow_Product": ["Product Model"],
    "Flow_QA": ["HDSD", "QA"],
    "Flow_Xu_ly_su_co": ["Xử lý sự cố"],
    "Flow_Auth": ["Auth"],
}

FLOW_SMOKE_P0 = [
    "TC-CRM-L01",
    "TC-CRM-L07",
    "TC-FLOW-L06",
    "TC-FLOW-C01",
    "TC-FLOW-M01",
    "TC-FLOW-P01",
    "TC-FLOW-Q03",
]


def _load_json(name: str) -> dict[str, Any]:
    path = FIXTURES / name
    if not path.is_file():
        return {}
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def _read_csv(path: Path) -> list[dict[str, str]]:
    if not path.is_file():
        return []
    with path.open(encoding="utf-8", newline="") as f:
        return [dict(row) for row in csv.DictReader(f)]


def _dedupe_registry(registry: list[dict[str, str]], seen: set[str]) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    for row in registry:
        tc_id = row.get("TC-ID", "")
        if not tc_id or tc_id in seen:
            continue
        seen.add(tc_id)
        out.append(row)
    return out


def _all_tc_rows(*sources: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    merged: dict[str, dict[str, str]] = {}
    for rows in sources:
        for row in rows:
            tc_id = row.get("TC-ID", "")
            if tc_id:
                merged[tc_id] = row
    return merged


def _style_header(ws, ncol: int, row: int = 1) -> None:
    for col in range(1, ncol + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _priority_fill(priority: str) -> PatternFill | None:
    return {"P0": P0_FILL, "P1": P1_FILL, "P2": P2_FILL}.get(priority)


def _merge_row_style(ws, row: int, ncol: int) -> None:
    for col in range(1, ncol + 1):
        c = ws.cell(row=row, column=col)
        c.alignment = WRAP
        c.border = BORDER


def _enrich_row(row: dict[str, str]) -> dict[str, str]:
    tc_id = row.get("TC-ID", "")
    hint = ENRICHMENT.get(tc_id, {})
    fixture = row.get("Fixture File", "")
    key = row.get("Fixture Key", "")
    data = f"{fixture} → {key}".strip(" →") if fixture or key else ""
    return {
        "pre": hint.get("pre", ""),
        "steps": hint.get("steps", ""),
        "screenshot": hint.get("screenshot", "Chụp màn hình bước cuối + lỗi (nếu Fail)"),
        "diagram": hint.get("diagram", row.get("Diagram Ref", "")),
        "type": hint.get("type", "Manual"),
        "role": hint.get("role", row.get("Role", "")),
        "url": hint.get("url", row.get("URL", "")),
        "auto": row.get("Automated Test", ""),
        "data": data,
    }


def _append_test_case_rows(ws, rows: list[dict[str, str]], *, start_stt: int = 1) -> int:
    stt = start_stt
    status_col = TEST_CASE_HEADERS.index("Trạng thái") + 1
    for row in rows:
        info = _enrich_row(row)
        tc_id = row.get("TC-ID", "")
        ws.append(
            [
                stt,
                tc_id,
                row.get("Priority", ""),
                row.get("Module", ""),
                row.get("Flow", ""),
                info["type"],
                "Staging",
                info["url"],
                info["role"],
                info["pre"],
                info["data"],
                info["steps"],
                row.get("Expected Result Summary", ""),
                info["screenshot"],
                info["diagram"],
                info["auto"],
                "",
                "Not Run",
                "",
                "",
                "",
                "",
                "",
            ]
        )
        r = ws.max_row
        fill = _priority_fill(row.get("Priority", ""))
        if fill:
            ws.cell(row=r, column=3).fill = fill
        _merge_row_style(ws, r, len(TEST_CASE_HEADERS))
        STATUS_DV.add(ws.cell(row=r, column=status_col))
        stt += 1
    return stt


def _build_tester_guide_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Huong_dan_tester", 0)
    ws["A1"] = "Hướng dẫn sử dụng bộ Test Case CRM — PTT"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    lines = [
        ("Cột quan trọng", "Cách điền"),
        ("Trạng thái", "Pass / Fail / Blocked / Skip / Not Run — dùng dropdown"),
        ("Hình minh họa", "Ghi tên file ảnh chụp màn hình (vd. TC-LEAD-01_step3.png)"),
        ("Evidence", "Đường dẫn folder hoặc link Drive chứa ảnh + log"),
        ("Bug ID", "Mã Jira/Linear/GitHub issue"),
        ("Sơ đồ tham chiếu", "Xem sheet So_do_* tương ứng"),
        ("", ""),
        ("Quy trình test Lead → Retain", "1. Smoke_P0 → 2. CRM_Lifecycle → 3. Flow_* theo module → 4. Test_Cases đầy đủ"),
        ("Tài liệu HDSD", "/crm/hdsd → huong-dan-day-du-lead-den-cham-soc-khach-hang"),
        ("Tải lại file mới", "/crm/test-cases/download.xlsx"),
        ("Ngày xuất", date.today().isoformat()),
    ]
    for i, (a, b) in enumerate(lines, start=3):
        ws.cell(row=i, column=1, value=a).font = Font(bold=True) if a and not b.startswith("1.") else Font()
        ws.cell(row=i, column=2, value=b)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 64


def _draw_flow_box(ws, row: int, col: int, text: str, *, width: int = 3) -> None:
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + width - 1)
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = BOX_FILL
    cell.font = Font(bold=True, size=9)
    cell.alignment = CENTER
    cell.border = BORDER


def _draw_arrow_cell(ws, row: int, col: int) -> None:
    cell = ws.cell(row=row, column=col, value="→")
    cell.alignment = CENTER
    cell.font = Font(bold=True, size=12, color="2F7238")


def _build_diagram_lead_retain(wb: Workbook) -> None:
    ws = wb.create_sheet("So_do_Lead_Retain")
    ws["A1"] = "Sơ đồ luồng Lead → Retain (Pre-sales trên Lead)"
    ws["A1"].font = Font(bold=True, size=12)
    r = 3
    boxes = [
        "Lead vào\n/crm/leads",
        "B2 Liên hệ\nCare gate",
        "Pre-sales\nLead·Consult·Proposal",
        "Ký HĐ\n/crm/hub",
        "Onboard",
        "Deliver",
        "Handover",
        "Retain",
    ]
    col = 1
    for i, label in enumerate(boxes):
        if i > 0:
            _draw_arrow_cell(ws, r, col)
            col += 1
        _draw_flow_box(ws, r, col, label, width=2)
        col += 2
    ws.row_dimensions[r].height = 36
    ws["A5"] = "Ghi chú: Chỉ chuyển tuần tự 1 bước; task giai đoạn 100% mới Chuyển →"
    ws["A5"].font = Font(italic=True, color="555555")
    ws["A7"] = "Test case liên quan"
    ws["A7"].font = Font(bold=True)
    refs = [
        "TC-CRM-L01 → TC-CRM-L02 → TC-CRM-L03..L06 → TC-CRM-L07 → TC-CRM-L08..L10",
        "Gates: TC-CRM-L16 (B2), L17-L18 (Intake), L19-L23 (Delivery/Payment)",
        "TC-LEAD-01, TC-LEAD-06, TC-LEAD-13",
    ]
    for i, t in enumerate(refs, start=8):
        ws.cell(row=i, column=1, value=t)
    ws.column_dimensions["A"].width = 18
    for c in range(2, 20):
        ws.column_dimensions[get_column_letter(c)].width = 10


def _build_diagram_kpi(wb: Workbook) -> None:
    ws = wb.create_sheet("So_do_KPI")
    ws["A1"] = "Sơ đồ dữ liệu KPI AM/SP"
    ws["A1"].font = Font(bold=True, size=12)
    _draw_flow_box(ws, 3, 1, "Lead owner\nassigned_am", width=3)
    _draw_arrow_cell(ws, 3, 5)
    _draw_flow_box(ws, 3, 6, "Lifecycle\nassigned_am / SP", width=3)
    _draw_arrow_cell(ws, 3, 10)
    _draw_flow_box(ws, 3, 11, "/crm/staff-kpi\nMetric tự động", width=3)
    ws["A5"] = "Nếu KPI = 0: banner vàng → Đồng bộ AM từ lead owner (TC-CRM-L12)"
    ws.column_dimensions["A"].width = 20


def _build_diagram_lead_sources(wb: Workbook) -> None:
    ws = wb.create_sheet("So_do_Nguon_Lead")
    ws["A1"] = "Nguồn lead vào CRM"
    ws["A1"].font = Font(bold=True, size=12)
    sources = [
        ("Facebook", "/api/crm/integration/webhooks/facebook", "TC-CRM-L13"),
        ("Zalo", "POST /api/crm/integration/webhooks/zalo", "TC-FLOW-L01"),
        ("Webform", "POST /api/landing-contact", "TC-FLOW-L02"),
        ("Marketing API", "POST /api/crm/integration/marketing/ingest", "TC-FLOW-L03"),
        ("Tư vấn", "POST /api/consultations", "TC-FLOW-L04"),
        ("Nhập tay", "/crm/leads + Lead", "TC-FLOW-L05"),
        ("Auto-assign", "Rule /crm/leads config", "TC-FLOW-L06"),
    ]
    headers = ["Nguồn", "Endpoint / URL", "TC tham chiếu"]
    ws.append(headers)
    _style_header(ws, len(headers), row=3)
    for src in sources:
        ws.append(list(src))
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=3):
        for c in row:
            c.border = BORDER
            c.alignment = WRAP
    _auto_width(ws, max_col=3)


def _build_screenshot_guide(wb: Workbook) -> None:
    ws = wb.create_sheet("Hinh_minh_hoa")
    ws["A1"] = "Hướng dẫn chụp hình minh họa (Evidence)"
    ws["A1"].font = Font(bold=True, size=12)
    headers = ["Màn hình", "URL", "Chụp gì", "Đặt tên file gợi ý"]
    ws.append(headers)
    _style_header(ws, len(headers), row=3)
    shots = [
        ("/crm/leads", "/crm/leads", "Danh sách + filter + nút + Lead", "TC-LEAD-01_list.png"),
        ("Lead detail", "/crm/leads/{id}", "3 tab Pre-sales + care B2", "TC-CRM-L03_presales.png"),
        ("Intake", "/crm/intake?lead_id=", "Form BANT + nút Go", "TC-CRM-L04_intake.png"),
        ("Hub HĐ", "/crm/hub", "HĐ Active + promote", "TC-CRM-L07_contract.png"),
        ("Service Delivery", "/crm/service-delivery/{id}", "Task checklist + Chuyển →", "TC-CRM-L09_workflow.png"),
        ("Staff KPI", "/crm/staff-kpi", "Banner + bảng metric", "TC-CRM-L11_kpi.png"),
        ("HDSD", "/crm/hdsd", "Danh mục tài liệu", "TC-FLOW-Q01_hdsd.png"),
        ("Hub", "/crm/hub", "HĐ draft → Active", "TC-FLOW-M01_hub.png"),
        ("CSKH Kanban", "/crm", "Case mới + pipeline", "TC-FLOW-C01_kanban.png"),
        ("Portal KPI", "/crm/kpi", "Chỉ metric NV đăng nhập", "TC-FLOW-P02_kpi.png"),
        ("SOP Launch", "/crm/sop", "14 bước checklist", "TC-FLOW-M04_sop.png"),
        ("RE Projects", "/crm/re-projects", "Cash flow + export", "TC-FLOW-F03_re.png"),
        ("Test case Excel", "/crm/test-cases/download.xlsx", "Sheet Flow_Index + dropdown", "TC-FLOW-Q03_xlsx.png"),
    ]
    for row in shots:
        ws.append(list(row))
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=4):
        for c in row:
            c.border = BORDER
            c.alignment = WRAP
    _auto_width(ws, max_col=4)


def _auto_width(ws, max_col: int, *, min_w: int = 10, max_w: int = 42) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        best = min_w
        for row in ws.iter_rows(min_col=col, max_col=col, max_row=min(ws.max_row, 150)):
            val = row[0].value
            if val:
                best = max(best, min(len(str(val)) + 2, max_w))
        ws.column_dimensions[letter].width = best


def _build_diagram_presales_gate(wb: Workbook) -> None:
    ws = wb.create_sheet("So_do_Presales_Gate")
    ws["A1"] = "Gate Pre-sales trên Lead (B2 · Intake · Cap chi phí)"
    ws["A1"].font = Font(bold=True, size=12)
    r = 3
    boxes = ["Lead mới", "B2 ✓\n(gate)", "Tab Lead\n100%", "Intake\nGo/No-Go", "Consult\n+ TMMT", "Proposal\n+ cap", "Hub HĐ"]
    col = 1
    for i, label in enumerate(boxes):
        if i > 0:
            _draw_arrow_cell(ws, r, col)
            col += 1
        _draw_flow_box(ws, r, col, label, width=2)
        col += 2
    ws["A5"] = "TC: L02, L03, L04, L05, L06, L16, L17, L18, L21, N03"
    ws.column_dimensions["A"].width = 18


def _build_diagram_delivery_gate(wb: Workbook) -> None:
    ws = wb.create_sheet("So_do_Delivery_Gate")
    ws["A1"] = "Gate Service Delivery (task · TMMT · payment)"
    ws["A1"].font = Font(bold=True, size=12)
    r = 3
    boxes = ["Onboard", "Deliver\n(task 100%)", "TMMT R5", "Handover", "Payment ✓", "Retain"]
    col = 1
    for i, label in enumerate(boxes):
        if i > 0:
            _draw_arrow_cell(ws, r, col)
            col += 1
        _draw_flow_box(ws, r, col, label, width=2)
        col += 2
    ws["A5"] = "Chỉ advance 1 stage; rollback 1 bước. TC: L08-L10, L19-L24, FLOW-D*"
    ws.column_dimensions["A"].width = 18


def _build_diagram_hub_mkt(wb: Workbook) -> None:
    ws = wb.create_sheet("So_do_Hub_MKT")
    ws["A1"] = "Hub · Marketing Plan · SOP · Sales"
    ws["A1"].font = Font(bold=True, size=12)
    _draw_flow_box(ws, 3, 1, "HĐ draft", width=2)
    _draw_arrow_cell(ws, 3, 4)
    _draw_flow_box(ws, 3, 5, "Active\n→ Onboard", width=2)
    _draw_arrow_cell(ws, 3, 8)
    _draw_flow_box(ws, 3, 9, "Campaign\n+ reminder", width=2)
    _draw_flow_box(ws, 5, 1, "MKT Plan\nsegment", width=2)
    _draw_flow_box(ws, 5, 5, "SOP 14 ngày", width=2)
    _draw_flow_box(ws, 5, 9, "Sales\npipeline", width=2)
    ws["A7"] = "TC: FLOW-M01..M06, TC-CRM-L07"
    ws.column_dimensions["A"].width = 16


def _build_diagram_cskh(wb: Workbook) -> None:
    ws = wb.create_sheet("So_do_CSKH")
    ws["A1"] = "Luồng CSKH — Case → Care → 360"
    ws["A1"].font = Font(bold=True, size=12)
    boxes = ["Case Mới", "Đang xử lý", "Care report", "Khách 360"]
    col = 1
    for i, label in enumerate(boxes):
        if i > 0:
            _draw_arrow_cell(ws, 3, col)
            col += 1
        _draw_flow_box(ws, 3, col, label, width=2)
        col += 2
    ws["A5"] = "TC: FLOW-C01..C04, TC-CSKH-* (registry)"
    ws.column_dimensions["A"].width = 14


def _build_diagram_finance(wb: Workbook) -> None:
    ws = wb.create_sheet("So_do_Finance")
    ws["A1"] = "Tài chính · Dashboard · RE"
    ws["A1"].font = Font(bold=True, size=12)
    _draw_flow_box(ws, 3, 1, "Business\nDashboard", width=2)
    _draw_flow_box(ws, 3, 5, "Owner\nWeekly", width=2)
    _draw_flow_box(ws, 3, 9, "RE Project\n7-sheet export", width=2)
    ws["A5"] = "TC: L25, L26, FLOW-F01..F04"
    ws.column_dimensions["A"].width = 16


def _build_diagram_portal(wb: Workbook) -> None:
    ws = wb.create_sheet("So_do_Portal")
    ws["A1"] = "Portal NV · Báo cáo · Payroll"
    ws["A1"].font = Font(bold=True, size=12)
    _draw_flow_box(ws, 3, 1, "Login NV", width=2)
    _draw_arrow_cell(ws, 3, 4)
    _draw_flow_box(ws, 3, 5, "Daily report", width=2)
    _draw_arrow_cell(ws, 3, 8)
    _draw_flow_box(ws, 3, 9, "KPI cá nhân", width=2)
    _draw_flow_box(ws, 5, 5, "Payroll\nimport", width=2)
    ws["A7"] = "TC: FLOW-P01..P04, TC-PORTAL-*"
    ws.column_dimensions["A"].width = 14


def _build_diagram_product(wb: Workbook) -> None:
    ws = wb.create_sheet("So_do_Product")
    ws["A1"] = "Product Model — Catalog · Addon"
    ws["A1"].font = Font(bold=True, size=12)
    headers = ["Thành phần", "URL", "TC"]
    ws.append(headers)
    _style_header(ws, len(headers), row=3)
    for row in [
        ("Catalog dịch vụ", "/crm/catalog", "FLOW-PR01"),
        ("Lead industry addon", "/crm/leads/{id}", "FLOW-PR02"),
        ("Review queue", "/crm/leads", "FLOW-PR03"),
    ]:
        ws.append(list(row))
    _auto_width(ws, max_col=3)


def _build_diagram_qa(wb: Workbook) -> None:
    ws = wb.create_sheet("So_do_QA")
    ws["A1"] = "HDSD + Bộ test case QA"
    ws["A1"].font = Font(bold=True, size=12)
    _draw_flow_box(ws, 3, 1, "/crm/hdsd\nđọc MD", width=2)
    _draw_arrow_cell(ws, 3, 4)
    _draw_flow_box(ws, 3, 5, "Tải HDSD\n.xlsx", width=2)
    _draw_arrow_cell(ws, 3, 8)
    _draw_flow_box(ws, 3, 9, "Test case\nworkbook", width=2)
    ws["A5"] = "TC: FLOW-Q01..Q03, TC-CRM-L14, L15"
    ws.column_dimensions["A"].width = 14


def _build_diagram_auth(wb: Workbook) -> None:
    ws = wb.create_sheet("So_do_Auth")
    ws["A1"] = "Đăng nhập Admin vs Portal NV"
    ws["A1"].font = Font(bold=True, size=12)
    _draw_flow_box(ws, 3, 1, "/admin/login\nAdmin", width=2)
    _draw_arrow_cell(ws, 3, 4)
    _draw_flow_box(ws, 3, 5, "→ /admin", width=2)
    _draw_flow_box(ws, 5, 1, "/admin/login\nNV", width=2)
    _draw_arrow_cell(ws, 5, 4)
    _draw_flow_box(ws, 5, 5, "→ /crm/home", width=2)
    ws["A7"] = "TC: FLOW-A01, A02, TC-AUTH-*"
    ws.column_dimensions["A"].width = 14


def _build_diagram_perm(wb: Workbook) -> None:
    ws = wb.create_sheet("So_do_Perm")
    ws["A1"] = "Ma trận quyền — menu ẩn / API 403"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A3"] = "Fixture: permissions_scenarios.json · TC-FLOW-N04, TC-PERM-*"
    ws.column_dimensions["A"].width = 48


SHORT_FLOW_HEADERS = [
    "TC-ID",
    "Ưu tiên",
    "Bước công đoạn",
    "URL",
    "Vai trò",
    "Kết quả mong đợi",
    "Trạng thái",
    "Tester",
    "Ngày",
    "Screenshot",
]


def _build_short_flow_sheet(wb: Workbook, sheet_name: str, rows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.append(SHORT_FLOW_HEADERS)
    _style_header(ws, len(SHORT_FLOW_HEADERS))
    ws.add_data_validation(STATUS_DV)
    status_col = SHORT_FLOW_HEADERS.index("Trạng thái") + 1
    for row in rows:
        info = _enrich_row(row)
        ws.append(
            [
                row.get("TC-ID", ""),
                row.get("Priority", ""),
                row.get("Flow", ""),
                info["url"],
                info["role"],
                row.get("Expected Result Summary", ""),
                "Not Run",
                "",
                "",
                info["screenshot"],
            ]
        )
        r = ws.max_row
        fill = _priority_fill(row.get("Priority", ""))
        if fill:
            ws.cell(row=r, column=2).fill = fill
        _merge_row_style(ws, r, len(SHORT_FLOW_HEADERS))
        STATUS_DV.add(ws.cell(row=r, column=status_col))
    _auto_width(ws, max_col=len(SHORT_FLOW_HEADERS))


def _build_flow_index_sheet(wb: Workbook, lifecycle: list[dict[str, str]], system_flows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet("Flow_Index")
    ws["A1"] = "Chỉ mục luồng hệ thống — chọn sheet để test theo module"
    ws["A1"].font = Font(bold=True, size=12)
    headers = ["Module / Luồng", "Sheet test", "Sơ đồ", "Mô tả", "Số TC"]
    ws.append(headers)
    _style_header(ws, len(headers), row=3)
    for module, sheet, diagram, desc in FLOW_INDEX:
        if sheet == "CRM_Lifecycle":
            count = len(lifecycle)
        else:
            modules = FLOW_SHEET_MODULES.get(sheet, [])
            count = sum(1 for r in system_flows if r.get("Module") in modules)
        ws.append([module, sheet, diagram, desc, count])
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=5):
        for c in row:
            c.border = BORDER
            c.alignment = WRAP
    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1, value=f"Tổng TC lifecycle: {len(lifecycle)}")
    ws.cell(row=summary_row + 1, column=1, value=f"Tổng TC luồng hệ thống: {len(system_flows)}")
    _auto_width(ws, max_col=5)


def _build_module_flow_sheets(wb: Workbook, system_flows: list[dict[str, str]]) -> None:
    for sheet_name, modules in FLOW_SHEET_MODULES.items():
        rows = [r for r in system_flows if r.get("Module") in modules]
        if rows:
            _build_short_flow_sheet(wb, sheet_name, rows)


def _build_main_test_sheet(
    wb: Workbook,
    registry: list[dict[str, str]],
    lifecycle: list[dict[str, str]],
    system_flows: list[dict[str, str]],
) -> None:
    ws = wb.create_sheet("Test_Cases")
    ws.append(TEST_CASE_HEADERS)
    _style_header(ws, len(TEST_CASE_HEADERS))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(TEST_CASE_HEADERS))}1"
    ws.add_data_validation(STATUS_DV)
    ws.add_data_validation(PRIORITY_DV)
    ws.add_data_validation(ENV_DV)
    ws.add_data_validation(TYPE_DV)

    stt = _append_test_case_rows(ws, lifecycle, start_stt=1)
    stt = _append_test_case_rows(ws, system_flows, start_stt=stt)
    seen = {r.get("TC-ID", "") for r in lifecycle + system_flows}
    registry_unique = _dedupe_registry(registry, seen)
    _append_test_case_rows(ws, registry_unique, start_stt=stt)

    widths = [5, 14, 8, 14, 28, 10, 10, 22, 12, 24, 22, 32, 28, 24, 16, 22, 20, 10, 12, 11, 18, 14, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_crm_lifecycle_sheet(wb: Workbook, lifecycle: list[dict[str, str]]) -> None:
    _build_short_flow_sheet(wb, "CRM_Lifecycle", lifecycle)


def _build_smoke_sheet(
    wb: Workbook,
    manifest: dict[str, Any],
    tc_lookup: dict[str, dict[str, str]],
) -> None:
    ws = wb.create_sheet("Smoke_P0")
    headers = ["TC-ID", "Luồng", "Trạng thái", "Tester", "Ngày", "Screenshot", "Ghi chú"]
    ws.append(headers)
    _style_header(ws, len(headers))
    ws.add_data_validation(STATUS_DV)
    status_col = 3
    smoke_ids: list[str] = []
    for tc_id in manifest.get("smoke_p0_checklist", []):
        if tc_id not in smoke_ids:
            smoke_ids.append(tc_id)
    for tc_id in FLOW_SMOKE_P0:
        if tc_id not in smoke_ids:
            smoke_ids.append(tc_id)
    for tc_id in smoke_ids:
        row = tc_lookup.get(tc_id, {})
        info = _enrich_row(row) if row else {"screenshot": ""}
        ws.append([tc_id, row.get("Flow", ""), "Not Run", "", "", info.get("screenshot", ""), ""])
        r = ws.max_row
        ws.cell(row=r, column=1).fill = P0_FILL
        _merge_row_style(ws, r, len(headers))
        STATUS_DV.add(ws.cell(row=r, column=status_col))
    _auto_width(ws, max_col=len(headers))


def _build_data_sheets(wb: Workbook, lifecycle: list[dict[str, str]], system_flows: list[dict[str, str]]) -> None:
    accounts = _load_json("accounts.json")
    ws = wb.create_sheet("Tai_khoan_mau")
    ws.append(["Key", "Username", "Password", "Mô tả"])
    _style_header(ws, 4)
    for key, val in accounts.items():
        if key.startswith("_") or not isinstance(val, dict):
            continue
        ws.append([key, val.get("username", ""), val.get("password", ""), val.get("description", "")])
    _auto_width(ws, max_col=4)

    ws2 = wb.create_sheet("Tong_quan")
    registry = _read_csv(REGISTRY)
    ws2["A1"] = "Bộ Test Case CRM — PTT Advertising"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A3"] = "Ngày xuất"
    ws2["B3"] = date.today().isoformat()
    ws2["A4"] = "Tổng TC (registry gốc)"
    ws2["B4"] = len(registry)
    ws2["A5"] = "TC Lead→Retain (lifecycle)"
    ws2["B5"] = len(lifecycle)
    ws2["A6"] = "TC luồng hệ thống (mới)"
    ws2["B6"] = len(system_flows)
    seen = {r.get("TC-ID", "") for r in lifecycle + system_flows if r.get("TC-ID")}
    _dedupe_registry(registry, seen)
    ws2["A7"] = "Tổng TC unique (Test_Cases sheet)"
    ws2["B7"] = len(seen)
    ws2["A8"] = "Tải từ CRM"
    ws2["B8"] = "/crm/test-cases/download.xlsx"
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 48


def build_crm_test_cases_workbook() -> BytesIO:
    """Tạo workbook Excel đầy đủ cho tester CRM."""
    registry = _read_csv(REGISTRY)
    lifecycle = _read_csv(CRM_LIFECYCLE)
    system_flows = _read_csv(SYSTEM_FLOWS)
    manifest = _load_json("manifest.json")
    tc_lookup = _all_tc_rows(registry, lifecycle, system_flows)

    wb = Workbook()
    wb.remove(wb.active)

    _build_tester_guide_sheet(wb)
    _build_flow_index_sheet(wb, lifecycle, system_flows)
    _build_main_test_sheet(wb, registry, lifecycle, system_flows)
    _build_crm_lifecycle_sheet(wb, lifecycle)
    _build_module_flow_sheets(wb, system_flows)
    _build_diagram_lead_retain(wb)
    _build_diagram_presales_gate(wb)
    _build_diagram_delivery_gate(wb)
    _build_diagram_kpi(wb)
    _build_diagram_lead_sources(wb)
    _build_diagram_hub_mkt(wb)
    _build_diagram_cskh(wb)
    _build_diagram_finance(wb)
    _build_diagram_portal(wb)
    _build_diagram_product(wb)
    _build_diagram_qa(wb)
    _build_diagram_auth(wb)
    _build_diagram_perm(wb)
    _build_screenshot_guide(wb)
    _build_smoke_sheet(wb, manifest, tc_lookup)
    _build_data_sheets(wb, lifecycle, system_flows)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
