#!/usr/bin/env python3
"""Generate docs/TEST_CASES_PTT.xlsx from registry CSV + fixture JSON files."""
from __future__ import annotations

import csv
import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / "tests" / "fixtures" / "test_data"
REGISTRY = FIXTURES / "test_cases_registry.csv"
OUTPUT = ROOT / "docs" / "TEST_CASES_PTT.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
P0_FILL = PatternFill("solid", fgColor="FCE4D6")
P1_FILL = PatternFill("solid", fgColor="FFF2CC")
P2_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def _load_json(name: str) -> dict:
    path = FIXTURES / name
    if not path.is_file():
        return {}
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def _style_header_row(ws, ncol: int, row: int = 1) -> None:
    for col in range(1, ncol + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _auto_width(ws, min_w: int = 10, max_w: int = 48) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = max(len(str(c.value or "")) for c in col_cells[:120])
        ws.column_dimensions[letter].width = min(max(length + 2, min_w), max_w)


def _read_registry() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    with REGISTRY.open(encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            rows.append(dict(row))
    return rows


def _priority_fill(priority: str) -> PatternFill | None:
    return {"P0": P0_FILL, "P1": P1_FILL, "P2": P2_FILL}.get(priority)


def _build_test_cases_sheet(wb: Workbook, registry: list[dict[str, str]]) -> None:
    ws = wb.active
    ws.title = "Test Cases"

    headers = [
        "TC-ID",
        "Ưu tiên",
        "Module",
        "Luồng / Tên test",
        "Tiền điều kiện",
        "Bước thực hiện",
        "File dữ liệu",
        "Key dữ liệu",
        "Dữ liệu test (tóm tắt)",
        "Kết quả mong muốn",
        "Kết quả thực tế",
        "Trạng thái",
        "Tester",
        "Ngày test",
        "Ghi chú / Bug ID",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Enrichment hints for common TCs (steps / preconditions)
    hints: dict[str, tuple[str, str, str]] = {
        "TC-AUTH-01": (
            "User admin tồn tại trên staging",
            "1. Mở /admin/login\n2. Nhập username/password\n3. Nhấn Đăng nhập",
            "username=admin; password xem sheet Tài khoản mẫu",
        ),
        "TC-AUTH-02": (
            "NV portal enabled, có crm_staff",
            "1. Mở /admin/login\n2. Nhập credentials portal\n3. Submit",
            "username=sales_a",
        ),
        "TC-PROJ-08": (
            "Dự án DA-A có webhook + pool NV + form map",
            "1. POST payload leadgen tới slug\n2. Kiểm tra lead trong /crm/leads\n3. Kiểm tra owner + KPI",
            "form_id=2814926042203269; phone=0907000001",
        ),
        "TC-PROJ-18": (
            "Full setup Phase 1–4 cho DA-A",
            "1. Chạy TC-PROJ-08\n2. Login sales_a → /crm/leads\n3. Kiểm tra notification + KPI RE_LEADS_NEW",
            "E2E phone 0907000001",
        ),
        "TC-IO-01": (
            "Quyền import lead",
            "1. /crm/leads → Import\n2. Chọn leads_import_sample.csv\n3. Xác nhận",
            "4 dòng CSV; dòng 4 trùng SĐT dòng 1",
        ),
    }

    status_col = headers.index("Trạng thái") + 1
    dv = DataValidation(
        type="list",
        formula1='"Pass,Fail,Blocked,Skip,Not Run"',
        allow_blank=True,
    )
    dv.error = "Chọn Pass, Fail, Blocked, Skip hoặc Not Run"
    dv.errorTitle = "Trạng thái không hợp lệ"
    ws.add_data_validation(dv)

    for i, row in enumerate(registry, start=2):
        tc_id = row.get("TC-ID", "")
        pre, steps, data_hint = hints.get(tc_id, ("", "", ""))
        fixture = row.get("Fixture File", "")
        key = row.get("Fixture Key", "")
        if not data_hint and fixture:
            data_hint = f"{fixture} → {key}" if key else fixture

        ws.append(
            [
                tc_id,
                row.get("Priority", ""),
                row.get("Module", ""),
                row.get("Flow", ""),
                pre,
                steps,
                fixture,
                key,
                data_hint,
                row.get("Expected Result Summary", ""),
                "",
                "Not Run",
                "",
                "",
                "",
            ]
        )
        fill = _priority_fill(row.get("Priority", ""))
        if fill:
            ws.cell(row=i, column=2).fill = fill
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=i, column=col)
            c.alignment = WRAP
            c.border = BORDER
        dv.add(ws.cell(row=i, column=status_col))

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 32
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 28
    ws.column_dimensions["J"].width = 32
    ws.column_dimensions["K"].width = 28
    ws.column_dimensions["L"].width = 12
    ws.column_dimensions["M"].width = 14
    ws.column_dimensions["N"].width = 12
    ws.column_dimensions["O"].width = 24


def _build_accounts_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Tài khoản mẫu")
    data = _load_json("accounts.json")
    headers = ["Key", "Username", "Password (staging)", "Mô tả", "Landing mong đợi"]
    ws.append(headers)
    _style_header_row(ws, len(headers))
    for key, val in data.items():
        if key.startswith("_") or not isinstance(val, dict):
            continue
        ws.append(
            [
                key,
                val.get("username", ""),
                val.get("password", ""),
                val.get("description", val.get("expected", "")),
                val.get("expected_landing", ""),
            ]
        )
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for c in row:
            c.alignment = WRAP
            c.border = BORDER
    _auto_width(ws)


def _build_projects_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Dự án mẫu")
    data = _load_json("re_projects_setup.json")
    headers = [
        "Mã dự án",
        "Tên",
        "Webhook slug",
        "Page ID",
        "Form ID",
        "NV pool (staff_id / role)",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))
    for proj in data.get("projects", []):
        forms = ", ".join(f.get("form_id", "") for f in proj.get("facebook_forms", []))
        staff = ", ".join(
            f"{s.get('staff_id')} ({s.get('role')})" for s in proj.get("staff", [])
        )
        ws.append(
            [
                proj.get("code"),
                proj.get("name"),
                proj.get("webhook_slug"),
                proj.get("facebook_page_id"),
                forms,
                staff,
            ]
        )
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for c in row:
            c.alignment = WRAP
            c.border = BORDER
    _auto_width(ws)


def _build_leads_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Lead mẫu")
    manual = _load_json("leads_manual.json")
    headers = ["Scenario", "full_name", "phone", "email", "source", "re_project", "Ghi chú"]
    ws.append(headers)
    _style_header_row(ws, len(headers))
    scenarios = [
        ("create_valid", manual.get("create_valid", {})),
        ("create_high_score", manual.get("create_high_score", {})),
        ("create_invalid_phone", manual.get("create_invalid_phone", {})),
        ("duplicate_phone (1st)", manual.get("duplicate_phone", {}).get("first", {})),
        ("duplicate_phone (2nd)", manual.get("duplicate_phone", {}).get("second", {})),
    ]
    def _cell_note(item: dict) -> str:
        if item.get("expected_error"):
            return str(item["expected_error"])
        exp = item.get("expected")
        if isinstance(exp, dict):
            return json.dumps(exp, ensure_ascii=False)
        if exp:
            return str(exp)
        return ""

    for name, item in scenarios:
        if not item:
            continue
        ws.append(
            [
                name,
                item.get("full_name", ""),
                item.get("phone", ""),
                item.get("email", ""),
                item.get("source", ""),
                item.get("re_project_code", ""),
                _cell_note(item),
            ]
        )
    _auto_width(ws)


def _build_smoke_sheet(wb: Workbook, manifest: dict) -> None:
    ws = wb.create_sheet("Smoke P0")
    headers = ["TC-ID", "Mô tả ngắn", "Trạng thái", "Tester", "Ngày", "Ghi chú"]
    ws.append(headers)
    _style_header_row(ws, len(headers))
    registry = {r["TC-ID"]: r for r in _read_registry()}
    dv = DataValidation(type="list", formula1='"Pass,Fail,Blocked,Skip,Not Run"', allow_blank=True)
    ws.add_data_validation(dv)
    status_col = 3
    for i, tc_id in enumerate(manifest.get("smoke_p0_checklist", []), start=2):
        row = registry.get(tc_id, {})
        ws.append(
            [
                tc_id,
                row.get("Flow", ""),
                "Not Run",
                "",
                "",
                "",
            ]
        )
        ws.cell(row=i, column=2).fill = P0_FILL
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=i, column=col)
            c.border = BORDER
            c.alignment = WRAP
        dv.add(ws.cell(row=i, column=status_col))
    _auto_width(ws)


def _build_summary_sheet(wb: Workbook, registry: list[dict[str, str]], manifest: dict) -> None:
    ws = wb.create_sheet("Tổng quan")
    ws["A1"] = "Bộ Test Case — Hệ thống PTT"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Phiên bản"
    ws["B3"] = manifest.get("version", "2026-05")
    ws["A4"] = "Ngày xuất file"
    ws["B4"] = date.today().isoformat()
    ws["A5"] = "Tài liệu chi tiết"
    ws["B5"] = "docs/TEST_CASES_PTT.md"
    ws["A6"] = "Fixtures"
    ws["B6"] = "tests/fixtures/test_data/"
    ws["A8"] = "Thống kê"
    ws["A8"].font = Font(bold=True, size=12)
    total = len(registry)
    by_p = {}
    by_mod = {}
    for r in registry:
        by_p[r.get("Priority", "?")] = by_p.get(r.get("Priority", "?"), 0) + 1
        by_mod[r.get("Module", "?")] = by_mod.get(r.get("Module", "?"), 0) + 1
    ws["A9"] = "Tổng test case"
    ws["B9"] = total
    row = 10
    for p in ("P0", "P1", "P2"):
        ws.cell(row=row, column=1, value=f"  {p}")
        ws.cell(row=row, column=2, value=by_p.get(p, 0))
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="Theo module").font = Font(bold=True)
    row += 1
    for mod, cnt in sorted(by_mod.items(), key=lambda x: (-x[1], x[0])):
        ws.cell(row=row, column=1, value=f"  {mod}")
        ws.cell(row=row, column=2, value=cnt)
        row += 1
    row += 2
    ws.cell(row=row, column=1, value="Automated tests").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Lệnh chạy")
    ws.cell(row=row, column=2, value=manifest.get("automated_test_suite", {}).get("command", ""))
    row += 1
    ws.cell(row=row, column=1, value="Kỳ vọng pass")
    ws.cell(row=row, column=2, value=manifest.get("automated_test_suite", {}).get("expected_pass", ""))
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 56


def main() -> None:
    registry = _read_registry()
    manifest = _load_json("manifest.json")
    wb = Workbook()
    _build_test_cases_sheet(wb, registry)
    _build_accounts_sheet(wb)
    _build_projects_sheet(wb)
    _build_leads_sheet(wb)
    _build_smoke_sheet(wb, manifest)
    _build_summary_sheet(wb, registry, manifest)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT} ({len(registry)} test cases, {len(wb.sheetnames)} sheets)")


if __name__ == "__main__":
    main()
