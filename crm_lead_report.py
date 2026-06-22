"""Báo cáo Lead — Excel & PDF."""
from __future__ import annotations

import sqlite3
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from crm_lead_store import LEAD_LEVEL_LABELS, LEAD_SOURCE_LABELS, LEAD_STATUS_LABELS, lead_row_to_dict


def _lead_export_rows(rows: list[sqlite3.Row | dict[str, Any]]) -> list[list[Any]]:
    out: list[list[Any]] = []
    for r in rows:
        d = lead_row_to_dict(r)
        out.append(
            [
                d["id"],
                d["full_name"],
                d["phone"],
                d["email"],
                d["source_label"],
                d["region"],
                d["product_interest"],
                d["need"][:200] if d["need"] else "",
                d["lead_score"],
                d["lead_level_label"],
                d["status_label"],
                d["owner_name"],
                "Có" if d["sla_overdue"] else "",
                d["activity_count"],
                d.get("converted_case_id") or "",
                d.get("converted_customer_id") or "",
                d["created_at"],
                d["updated_at"],
            ]
        )
    return out


LEAD_EXPORT_HEADERS = [
    "ID",
    "Họ tên",
    "SĐT",
    "Email",
    "Nguồn",
    "Khu vực",
    "Sản phẩm QT",
    "Nhu cầu",
    "Điểm",
    "Phân loại",
    "Trạng thái",
    "Owner",
    "Quá SLA",
    "Số activity",
    "Case ID",
    "KH ID",
    "Tạo lúc",
    "Cập nhật",
]


def build_leads_xlsx(
    rows: list[sqlite3.Row | dict[str, Any]],
    stats: dict[str, Any] | None = None,
) -> tuple[BytesIO, str]:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Danh sach Lead"

    if stats:
        ws.append(["BÁO CÁO LEAD CRM"])
        ws.append([f"Xuất lúc: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
        ws.append([f"Tổng lead: {stats.get('total', 0)} | Hot: {(stats.get('by_level') or {}).get('hot', 0)} | Quá SLA: {stats.get('sla_overdue', 0)} | Chuyển đổi: {stats.get('conversion_rate', 0)}%"])
        ws.append([])

    header_font = Font(bold=True)
    ws.append(LEAD_EXPORT_HEADERS)
    for cell in ws[ws.max_row]:
        cell.font = header_font
    for row in _lead_export_rows(rows):
        ws.append(row)

    if stats:
        ws2 = wb.create_sheet("Thong ke")
        ws2.append(["Chỉ số", "Giá trị"])
        ws2.append(["Tổng", stats.get("total", 0)])
        ws2.append(["Chốt (won)", stats.get("won", 0)])
        ws2.append(["Tỷ lệ chuyển đổi %", stats.get("conversion_rate", 0)])
        ws2.append(["Quá SLA", stats.get("sla_overdue", 0)])
        ws2.append([])
        ws2.append(["Theo trạng thái", ""])
        for k, v in (stats.get("by_status") or {}).items():
            ws2.append([LEAD_STATUS_LABELS.get(k, k), v])
        ws2.append([])
        ws2.append(["Theo nguồn", ""])
        for k, v in (stats.get("by_source") or {}).items():
            ws2.append([LEAD_SOURCE_LABELS.get(k, k), v])
        ws2.append([])
        ws2.append(["Theo phân loại", ""])
        for k, v in (stats.get("by_level") or {}).items():
            ws2.append([LEAD_LEVEL_LABELS.get(k, k), v])
        by_owner = stats.get("by_owner") or []
        if by_owner:
            ws3 = wb.create_sheet("Theo Owner")
            from crm_lead_care_pipeline import CARE_PIPELINE_STAGES

            stage_headers = [f"B{i + 1}" for i in range(len(CARE_PIPELINE_STAGES))]
            ws3.append(
                [
                    "Owner",
                    "Mã NV",
                    "Tổng lead",
                    *stage_headers,
                    "Đang mở",
                    "Mất",
                    "Chốt",
                    "Tỷ lệ %",
                    "Quá SLA",
                    "Điểm TB",
                ]
            )
            for cell in ws3[1]:
                cell.font = header_font
            for o in by_owner:
                stage_counts = {
                    s["key"]: s.get("count", 0)
                    for s in (o.get("care_stages") or [])
                }
                ws3.append(
                    [
                        o.get("owner_name") or "—",
                        o.get("owner_code") or "",
                        o.get("total") or 0,
                        *[stage_counts.get(st["key"], 0) for st in CARE_PIPELINE_STAGES],
                        o.get("open") or 0,
                        o.get("lost") or 0,
                        o.get("won") or 0,
                        o.get("conversion_rate") or 0,
                        o.get("sla_overdue") or 0,
                        o.get("avg_score") or 0,
                    ]
                )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = datetime.now().strftime("%Y-%m-%d")
    return buf, f"bao-cao-lead-{stamp}.xlsx"


def build_leads_pdf(
    rows: list[sqlite3.Row | dict[str, Any]],
    stats: dict[str, Any] | None = None,
) -> tuple[BytesIO, str]:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError as exc:
        raise RuntimeError("Cần cài reportlab: pip install reportlab") from exc

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=12 * mm, rightMargin=12 * mm)
    styles = getSampleStyleSheet()
    story: list[Any] = []
    story.append(Paragraph("Báo cáo Lead CRM", styles["Title"]))
    story.append(Spacer(1, 6))
    if stats:
        story.append(
            Paragraph(
                f"Tổng: <b>{stats.get('total', 0)}</b> | "
                f"Quá SLA: <b>{stats.get('sla_overdue', 0)}</b> | "
                f"Chuyển đổi: <b>{stats.get('conversion_rate', 0)}%</b>",
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 8))

    table_data = [LEAD_EXPORT_HEADERS[:13]]
    for row in _lead_export_rows(rows)[:200]:
        table_data.append(row[:13])

    tbl = Table(table_data, repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ]
        )
    )
    story.append(tbl)

    by_owner = stats.get("by_owner") or [] if stats else []
    if by_owner:
        story.append(Spacer(1, 12))
        story.append(Paragraph("Hiệu suất theo Owner", styles["Heading2"]))
        from crm_lead_care_pipeline import CARE_PIPELINE_STAGES

        stage_headers = [f"B{i + 1}" for i in range(len(CARE_PIPELINE_STAGES))]
        owner_data = [["Owner", "Tổng", *stage_headers, "Mở", "Mất", "Chốt", "%", "SLA"]]
        for o in by_owner[:25]:
            stage_counts = {s["key"]: s.get("count", 0) for s in (o.get("care_stages") or [])}
            owner_data.append(
                [
                    str(o.get("owner_name") or "—")[:20],
                    str(o.get("total") or 0),
                    *[str(stage_counts.get(st["key"], 0)) for st in CARE_PIPELINE_STAGES],
                    str(o.get("open") or 0),
                    str(o.get("lost") or 0),
                    str(o.get("won") or 0),
                    str(o.get("conversion_rate") or 0),
                    str(o.get("sla_overdue") or 0),
                ]
            )
        ot = Table(owner_data, repeatRows=1)
        ot.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ]
            )
        )
        story.append(ot)

    doc.build(story)
    buf.seek(0)
    stamp = datetime.now().strftime("%Y-%m-%d")
    return buf, f"bao-cao-lead-{stamp}.pdf"
