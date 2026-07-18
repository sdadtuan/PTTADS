"""Chuyển tài liệu HDSD (Markdown) sang Excel."""
from __future__ import annotations

import re
import zipfile
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_TABLE_SEP_RE = re.compile(r"^\s*\|?\s*:?-{3,}:?\s*(\|\s*:?-{3,}:?\s*)+\|?\s*$")
_HEADING_RE = re.compile(r"^(#{1,6})\s+(.+?)\s*$")
_HR_RE = re.compile(r"^-{3,}\s*$")
_LIST_RE = re.compile(r"^(\s*)[-*+]\s+(.+)$")
_OLIST_RE = re.compile(r"^\d+\.\s+(.+)$")
_MD_LINK_RE = re.compile(r"\[([^\]]+)\]\(([^)]+)\)")
_MD_BOLD_RE = re.compile(r"\*\*([^*]+)\*\*")
_MD_CODE_INLINE_RE = re.compile(r"`([^`]+)`")


def _thin_border() -> Border:
    thin = Side(style="thin", color="B4B4B4")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _strip_md_inline(text: str) -> str:
    s = str(text or "")
    s = _MD_LINK_RE.sub(r"\1 (\2)", s)
    s = _MD_BOLD_RE.sub(r"\1", s)
    s = _MD_CODE_INLINE_RE.sub(r"\1", s)
    return s.strip()


def _is_table_row(line: str) -> bool:
    return "|" in line


def _is_table_separator(line: str) -> bool:
    return bool(_TABLE_SEP_RE.match(line.strip()))


def _parse_table_row(line: str) -> list[str]:
    raw = line.strip()
    if raw.startswith("|"):
        raw = raw[1:]
    if raw.endswith("|"):
        raw = raw[:-1]
    return [_strip_md_inline(c.strip()) for c in raw.split("|")]


def _is_block_start(line: str) -> bool:
    s = line.strip()
    if not s:
        return False
    if s.startswith("```"):
        return True
    if _HEADING_RE.match(s):
        return True
    if _HR_RE.match(s):
        return True
    if _LIST_RE.match(line):
        return True
    if _OLIST_RE.match(s):
        return True
    if _is_table_row(s):
        return True
    return False


def parse_markdown_blocks(text: str) -> list[dict[str, Any]]:
    """Tách Markdown thành block heading / đoạn / bảng / list / code."""
    lines = text.splitlines()
    blocks: list[dict[str, Any]] = []
    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        if stripped.startswith("```"):
            lang = stripped[3:].strip()
            i += 1
            code_lines: list[str] = []
            while i < len(lines) and not lines[i].strip().startswith("```"):
                code_lines.append(lines[i])
                i += 1
            blocks.append({"type": "code", "lang": lang, "text": "\n".join(code_lines).strip()})
            if i < len(lines):
                i += 1
            continue

        hm = _HEADING_RE.match(stripped)
        if hm:
            blocks.append({"type": "heading", "level": len(hm.group(1)), "text": _strip_md_inline(hm.group(2))})
            i += 1
            continue

        if _HR_RE.match(stripped):
            blocks.append({"type": "hr"})
            i += 1
            continue

        if _is_table_row(stripped) and i + 1 < len(lines) and _is_table_separator(lines[i + 1]):
            table_lines = [stripped, lines[i + 1].strip()]
            i += 2
            while i < len(lines) and _is_table_row(lines[i].strip()) and not _is_table_separator(lines[i]):
                table_lines.append(lines[i].strip())
                i += 1
            rows = [_parse_table_row(r) for r in table_lines if not _is_table_separator(r)]
            blocks.append({"type": "table", "rows": rows})
            continue

        lm = _LIST_RE.match(line)
        if lm:
            blocks.append(
                {
                    "type": "list",
                    "text": _strip_md_inline(lm.group(2)),
                    "indent": len(lm.group(1)),
                }
            )
            i += 1
            continue

        om = _OLIST_RE.match(stripped)
        if om:
            blocks.append({"type": "olist", "text": _strip_md_inline(om.group(1))})
            i += 1
            continue

        if stripped:
            para_lines = [stripped]
            i += 1
            while i < len(lines) and lines[i].strip() and not _is_block_start(lines[i]):
                para_lines.append(lines[i].strip())
                i += 1
            blocks.append({"type": "para", "text": _strip_md_inline(" ".join(para_lines))})
            continue

        i += 1
    return blocks


def _heading_path(headings: list[tuple[int, str]], level: int, title: str) -> str:
    while headings and headings[-1][0] >= level:
        headings.pop()
    headings.append((level, title))
    return " › ".join(h for _, h in headings)


def _current_path(headings: list[tuple[int, str]]) -> str:
    return " › ".join(h for _, h in headings)


def _write_header_row(ws, headers: list[str], row: int) -> None:
    border = _thin_border()
    fill = PatternFill("solid", fgColor="2F7238")
    font = Font(bold=True, color="FFFFFF")
    for col, label in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def _autosize_columns(ws, max_col: int, *, max_width: int = 60) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        best = 12
        for row in ws.iter_rows(min_col=col, max_col=col):
            val = row[0].value
            if val is None:
                continue
            best = max(best, min(max_width, len(str(val)) + 2))
        ws.column_dimensions[letter].width = best


def build_hdsd_doc_xlsx(text: str, meta: dict[str, Any]) -> BytesIO:
    """Tạo workbook Excel từ nội dung Markdown."""
    blocks = parse_markdown_blocks(text)
    headings: list[tuple[int, str]] = []
    toc_rows: list[list[Any]] = []
    content_rows: list[list[Any]] = []
    table_rows: list[list[Any]] = []
    table_idx = 0
    content_stt = 0
    toc_stt = 0

    for block in blocks:
        btype = block["type"]
        if btype == "heading":
            level = int(block["level"])
            title = str(block["text"])
            path = _heading_path(headings, level, title)
            toc_stt += 1
            toc_rows.append([toc_stt, level, title, path])
            content_stt += 1
            content_rows.append([content_stt, path, "Tiêu đề", title])
            continue

        path = _current_path(headings)

        if btype == "para":
            content_stt += 1
            content_rows.append([content_stt, path, "Đoạn văn", block["text"]])
        elif btype in ("list", "olist"):
            content_stt += 1
            prefix = "• " if btype == "list" else "1. "
            content_rows.append([content_stt, path, "Danh sách", prefix + str(block["text"])])
        elif btype == "code":
            lang = str(block.get("lang") or "")
            label = "Sơ đồ" if lang == "mermaid" else "Mã / block"
            content_stt += 1
            content_rows.append([content_stt, path, label, str(block.get("text") or "")[:8000]])
        elif btype == "table":
            table_idx += 1
            rows: list[list[str]] = block.get("rows") or []
            for r_i, row in enumerate(rows):
                padded = list(row) + [""] * max(0, 8 - len(row))
                table_rows.append([table_idx, r_i + 1] + padded[:8])
        elif btype == "hr":
            content_stt += 1
            content_rows.append([content_stt, path, "Phân cách", "—"])

    wb = Workbook()

    ws_info = wb.active
    assert ws_info is not None
    ws_info.title = "Thong_tin"
    info_pairs = [
        ("Tiêu đề", meta.get("title", "")),
        ("Nhóm", meta.get("section_label", meta.get("section", ""))),
        ("File gốc", meta.get("filename", "")),
        ("Section key", meta.get("section", "")),
        ("Slug", meta.get("slug", "")),
    ]
    ws_info["A1"] = "Trường"
    ws_info["B1"] = "Giá trị"
    _write_header_row(ws_info, ["Trường", "Giá trị"], 1)
    for r, (k, v) in enumerate(info_pairs, 2):
        ws_info.cell(row=r, column=1, value=k)
        ws_info.cell(row=r, column=2, value=v)
    _autosize_columns(ws_info, 2)

    ws_toc = wb.create_sheet("Muc_luc")
    toc_headers = ["STT", "Cấp", "Tiêu đề", "Đường dẫn mục"]
    _write_header_row(ws_toc, toc_headers, 1)
    for r, row in enumerate(toc_rows, 2):
        for c, val in enumerate(row, 1):
            ws_toc.cell(row=r, column=c, value=val)
    _autosize_columns(ws_toc, len(toc_headers))

    ws_content = wb.create_sheet("Noi_dung")
    content_headers = ["STT", "Mục", "Loại", "Nội dung"]
    _write_header_row(ws_content, content_headers, 1)
    border = _thin_border()
    for r, row in enumerate(content_rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws_content.cell(row=r, column=c, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws_content.freeze_panes = "A2"
    _autosize_columns(ws_content, len(content_headers))

    ws_tables = wb.create_sheet("Bang")
    table_headers = ["Bảng #", "Hàng #", "C1", "C2", "C3", "C4", "C5", "C6", "C7", "C8"]
    _write_header_row(ws_tables, table_headers, 1)
    for r, row in enumerate(table_rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws_tables.cell(row=r, column=c, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws_tables.freeze_panes = "A2"
    _autosize_columns(ws_tables, len(table_headers))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_hdsd_all_zip(catalog: list[dict[str, Any]], read_doc) -> BytesIO:
    """Gói tất cả tài liệu HDSD thành ZIP chứa file .xlsx."""
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for group in catalog:
            section_key = str(group.get("key") or "")
            for doc in group.get("docs") or []:
                slug = str(doc.get("slug") or "")
                loaded = read_doc(section_key, slug)
                if loaded is None:
                    continue
                text, meta = loaded
                xlsx = build_hdsd_doc_xlsx(text, meta)
                stem = str(doc.get("filename") or slug).rsplit(".", 1)[0]
                arcname = f"{section_key}/{stem}.xlsx"
                zf.writestr(arcname, xlsx.getvalue())
    buf.seek(0)
    return buf
