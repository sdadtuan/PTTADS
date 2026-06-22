"""Import chấm công từ file Excel (mẫu Bảng chấm công chi tiết)."""
from __future__ import annotations

import re
from datetime import date, datetime, time
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

_PIN_HEADERS = frozenset({"mã pin", "ma pin", "pin", "mã pin máy"})
_DATE_HEADERS = frozenset({"ngày", "ngay", "date", "work date"})
_IN_HEADERS = frozenset({"vào", "vao", "check in", "giờ vào", "gio vao", "in"})
_OUT_HEADERS = frozenset({"ra", "check out", "giờ ra", "gio ra", "out"})


def _norm_header(val: Any) -> str:
    return re.sub(r"\s+", " ", str(val or "").strip().lower())


def _cell_to_work_date(val: Any) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s or s in ("-", "—"):
        return None
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(y, mo, d).strftime("%Y-%m-%d")
        except ValueError:
            return None
    m2 = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m2:
        return f"{m2.group(1)}-{m2.group(2)}-{m2.group(3)}"
    return None


def _cell_to_hhmm(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return f"{val.hour:02d}:{val.minute:02d}"
    if isinstance(val, time):
        return f"{val.hour:02d}:{val.minute:02d}"
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        try:
            seconds = int(round(float(val) * 86400))
            seconds = max(0, min(seconds, 24 * 3600 - 60))
            h, rem = divmod(seconds, 3600)
            m = rem // 60
            return f"{h:02d}:{m:02d}"
        except (TypeError, ValueError, OverflowError):
            return ""
    s = str(val).strip()
    if not s or s in ("-", "—"):
        return ""
    m = re.match(r"^(\d{1,2}):(\d{2})(?::\d{2})?$", s)
    if not m:
        return ""
    h, mm = int(m.group(1)), int(m.group(2))
    if h > 23 or mm > 59:
        return ""
    return f"{h:02d}:{mm:02d}"


def _find_header_row(ws, *, max_scan: int = 25) -> tuple[int, dict[str, int]] | None:
    for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        if not row:
            continue
        col_map: dict[str, int] = {}
        for cidx, cell in enumerate(row):
            h = _norm_header(cell)
            if not h:
                continue
            if h in _PIN_HEADERS or "mã pin" in h or h.endswith(" pin"):
                col_map.setdefault("pin", cidx)
            elif h in _DATE_HEADERS:
                col_map.setdefault("date", cidx)
            elif h in _IN_HEADERS:
                col_map.setdefault("check_in", cidx)
            elif h in _OUT_HEADERS:
                col_map.setdefault("check_out", cidx)
        if "pin" in col_map and "date" in col_map:
            if "check_in" not in col_map and len(row) > 8:
                col_map.setdefault("check_in", 8)
            if "check_out" not in col_map and len(row) > 9:
                col_map.setdefault("check_out", 9)
            if "check_in" in col_map and "check_out" in col_map:
                return ridx, col_map
    return None


def parse_timesheet_xlsx(data: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """
    Đọc file .xlsx kiểu «Bảng chấm công chi tiết».
    Trả (danh sách bản ghi hợp lệ, lỗi theo dòng).
    """
    records: list[dict[str, Any]] = []
    errors: list[str] = []
    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        return [], [f"Không đọc được file Excel: {exc}"]

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            found = _find_header_row(ws)
            if not found:
                continue
            header_row, cols = found
            pin_i = cols["pin"]
            date_i = cols["date"]
            in_i = cols["check_in"]
            out_i = cols["check_out"]

            for ridx, row in enumerate(
                ws.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                if not row:
                    continue
                pin_raw = row[pin_i] if pin_i < len(row) else None
                pin = str(pin_raw or "").strip()
                if not pin or pin.lower() in ("-", "—"):
                    continue

                wd = _cell_to_work_date(row[date_i] if date_i < len(row) else None)
                if not wd:
                    if any(row[c] not in (None, "", "-", "—") for c in (pin_i, date_i) if c < len(row)):
                        errors.append(f"Dòng {ridx} ({sheet_name}): ngày không hợp lệ")
                    continue

                check_in = _cell_to_hhmm(row[in_i] if in_i < len(row) else None)
                check_out = _cell_to_hhmm(row[out_i] if out_i < len(row) else None)
                if not check_in and not check_out:
                    continue

                records.append(
                    {
                        "pin": pin,
                        "work_date": wd,
                        "check_in": check_in,
                        "check_out": check_out,
                        "sheet": sheet_name,
                        "row": ridx,
                    }
                )
            if records or errors:
                break
    finally:
        wb.close()

    if not records and not errors:
        errors.append(
            "Không tìm thấy dữ liệu — cần sheet có cột «Mã PIN», «Ngày», «Vào», «Ra» (mẫu Bảng chấm công chi tiết)."
        )
    return records, errors
