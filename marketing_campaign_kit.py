"""Tạo bộ KHMKT + KPI theo chiến dịch/dự án — chatbox CMS PTT."""
from __future__ import annotations

import json
import os
import re
import time
import unicodedata
import urllib.request
import uuid
from io import BytesIO
from typing import Any

from marketing_execution import _parse_budget_vnd, build_kpi_strategy_xlsx, normalize_query
from marketing_project_intel import analyze_campaign_brief, apply_intel_to_plan

# Cache tạm bộ file đã sinh (uuid → plan)
_campaign_kit_cache: dict[str, dict[str, Any]] = {}
_CACHE_TTL_SEC = 3600

STEP_BLUEPRINTS: tuple[dict[str, str], ...] = (
    {
        "id": "ads_copy",
        "num": "1",
        "title": "Mẫu FB/Google Ads",
        "weeks": "W1–W2",
        "goal": "Soạn ≥3 biến thể copy, UTM + pixel/CAPI, brief creative 14 ngày launch",
        "inputs": "ICP, pain/offer, landing URL, ngân sách test 10–15M/kênh",
        "process": "Brief 1 trang → 3 góc tiếp cận → Meta/Google structure → QA tracking → A/B 7 ngày",
        "deliverable": "Brief ads + 3 variant + UTM doc + checklist pixel",
        "kpis": "CPL, CTR, CPC, CVR landing",
        "risks": "Tracking lỗi; creative không khớp ICP",
    },
    {
        "id": "tvc_kol",
        "num": "2",
        "title": "TVC & video KOL",
        "weeks": "W2–W5",
        "goal": "Sản xuất video/KOL 21 ngày, shot list, legal hợp đồng",
        "inputs": "Key message, persona KOL, budget sản xuất, timeline",
        "process": "Brief → casting KOL → kịch bản/beat sheet → quay/dựng → QC legal",
        "deliverable": "Video 15–30s + b-roll + usage rights doc",
        "kpis": "VTR, CPV, CPL ref từ video",
        "risks": "KOL scandal; delay sản xuất",
    },
    {
        "id": "excel_weekly",
        "num": "3",
        "title": "Excel kế hoạch tuần",
        "weeks": "W1–W12",
        "goal": "Ritual daily/T4/T6, báo cáo KPI, ma trận rủi ro P×I",
        "inputs": "Template 12 tuần, owner RACI, ngưỡng cảnh báo",
        "process": "Daily spend → T4 mid-week → T6 họp KPI → reforecast",
        "deliverable": "File Excel tuần + email báo cáo T6",
        "kpis": "Pacing spend, CPL trend, MQL/SQL",
        "risks": "Thiếu data; không họp T6 đều",
    },
    {
        "id": "funnel",
        "num": "4",
        "title": "Funnel lead → khách hàng",
        "weeks": "W1–W3",
        "goal": "CRM 6 ngày setup, email nurture D0–D14, SLA SDR ≤4h",
        "inputs": "CRM fields, scoring rules, email template, landing thank-you",
        "process": "Map touchpoint → automation email → SLA alert → dashboard funnel",
        "deliverable": "CRM pipeline + 5 email nurture + dashboard",
        "kpis": "MQL→SQL, contact rate, time-to-first-touch",
        "risks": "Lead rơi do SLA chậm; scoring sai",
    },
    {
        "id": "telesales",
        "num": "5",
        "title": "Script telesales / SDR",
        "weeks": "W1–W12",
        "goal": "SOP SDR hàng ngày, script phản đối, log CRM chuẩn",
        "inputs": "ICP, objection list, CRM disposition codes",
        "process": "SPIN script → 4 lần follow-up → handoff AE",
        "deliverable": "SOP 1 trang + script + template log call",
        "kpis": "Contact rate, demo book rate, SQL/meeting",
        "risks": "Script cứng; không log CRM",
    },
    {
        "id": "multichannel_plan",
        "num": "6",
        "title": "Kế hoạch truyền thông đa kênh",
        "weeks": "W1 (workshop)",
        "goal": "Workshop 3h, key message funnel, phân bổ budget base/test/reserve",
        "inputs": "Media mix, ICP journey, ngân sách quý",
        "process": "Workshop → chốt 2–3 kênh trọng tâm → lịch phát → UTM thống nhất",
        "deliverable": "Plan đa kênh 3 tháng + calendar nội dung",
        "kpis": "CPL theo kênh, SOV, brand search",
        "risks": "Phân tán budget; message không nhất quán",
    },
    {
        "id": "channel_test",
        "num": "7",
        "title": "Phương pháp test kênh",
        "weeks": "W3–W8",
        "goal": "TEST-ID register, min sample, scale +20%/tuần, post-mortem",
        "inputs": "Giả thuyết test, KPI success, ngân sách test 10–20%",
        "process": "Standup T3 → chạy test 2–4 tuần → scale/kill → post-mortem",
        "deliverable": "TEST-ID sheet + quyết định scale/pause",
        "kpis": "CPL/ROAS test vs control, sample size",
        "risks": "Scale sớm; không đủ sample",
    },
)

OBJECTIVE_CHANNELS: dict[str, list[dict[str, Any]]] = {
    "lead": [
        {"name": "Google Search", "goal": "Conversion", "kpi": "CPL ≤250k", "budget_pct": 35},
        {"name": "Meta Ads", "goal": "Lead form", "kpi": "CPL ≤180k", "budget_pct": 25},
        {"name": "LinkedIn", "goal": "B2B Lead", "kpi": "CPL ≤350k", "budget_pct": 15},
        {"name": "Email/CRM", "goal": "Nurture", "kpi": "Open ≥25%", "budget_pct": 5},
        {"name": "Landing/CRO", "goal": "Conversion", "kpi": "CVR ≥4%", "budget_pct": 5},
        {"name": "Telesales", "goal": "SQL", "kpi": "Contact ≥60%", "budget_pct": 0},
        {"name": "Dự phòng", "goal": "Test", "kpi": "—", "budget_pct": 15},
    ],
    "awareness": [
        {"name": "TikTok/Video", "goal": "Awareness", "kpi": "CPV ≤800", "budget_pct": 30},
        {"name": "Meta Reach", "goal": "Reach", "kpi": "CPM, Freq ≤3", "budget_pct": 25},
        {"name": "YouTube", "goal": "Consideration", "kpi": "VTR ≥25%", "budget_pct": 20},
        {"name": "PR/Content", "goal": "SOV", "kpi": "Branded search +15%", "budget_pct": 15},
        {"name": "Google Display", "goal": "Remarketing", "kpi": "CTR ≥0.5%", "budget_pct": 5},
        {"name": "Dự phòng", "goal": "Test", "kpi": "—", "budget_pct": 5},
    ],
    "retention": [
        {"name": "Email/CRM", "goal": "Retention", "kpi": "Churn ≤5%", "budget_pct": 35},
        {"name": "In-app/SMS", "goal": "Activation", "kpi": "DAU/MAU", "budget_pct": 20},
        {"name": "Meta Retarget", "goal": "Upsell", "kpi": "ROAS ≥4", "budget_pct": 20},
        {"name": "CS/NPS", "goal": "Advocacy", "kpi": "NPS ≥40", "budget_pct": 10},
        {"name": "Referral", "goal": "Growth", "kpi": "Referral rate", "budget_pct": 10},
        {"name": "Dự phòng", "goal": "Test", "kpi": "—", "budget_pct": 5},
    ],
}

DEFAULT_KPIS: list[dict[str, str]] = [
    {"name": "CPL", "target": "≤250k VNĐ", "freq": "Daily", "owner": "Ads Lead"},
    {"name": "MQL", "target": "Theo quota tháng", "freq": "Weekly", "owner": "MKT"},
    {"name": "MQL→SQL", "target": "≥25%", "freq": "Weekly", "owner": "Sales Lead"},
    {"name": "ROAS", "target": "≥3.0", "freq": "Weekly", "owner": "MKT Lead"},
    {"name": "ROMI", "target": "≥150%", "freq": "Monthly", "owner": "Finance"},
]

DEFAULT_RISKS: list[dict[str, str]] = [
    {"id": "R01", "name": "Overspend ngân sách", "prob": "3", "impact": "4", "mitigation": "Daily cap, pacing rules"},
    {"id": "R02", "name": "Tracking/pixel lỗi", "prob": "3", "impact": "5", "mitigation": "QA checklist weekly"},
    {"id": "R03", "name": "MQL→SQL sụt", "prob": "4", "impact": "5", "mitigation": "SLA 4h + review scoring"},
    {"id": "R04", "name": "Creative fatigue", "prob": "4", "impact": "3", "mitigation": "Refresh 2–4 tuần"},
    {"id": "R05", "name": "Phụ thuộc 1 kênh", "prob": "3", "impact": "4", "mitigation": "Media mix đa kênh"},
]


def _slugify(name: str) -> str:
    s = normalize_query(name)
    s = re.sub(r"[^\w\s-]", "", s)
    s = re.sub(r"[\s_]+", "-", s).strip("-")
    return (s[:48] or "du-an").lower()


def _detect_objective(text: str) -> str:
    q = normalize_query(text)
    if any(k in q for k in ("nhan dien", "brand", "awareness", "thuong hieu", "top of mind")):
        return "awareness"
    if any(k in q for k in ("giu chan", "retention", "upsell", "cskh", "loyalty", "churn")):
        return "retention"
    return "lead"


def _extract_project_name(text: str) -> str:
    raw = str(text or "").strip()
    for pat in (
        r"(?:ten|tên)\s*(?:du\s*an|dự\s*án|chien\s*dich|chiến\s*dịch)\s*[:\-–]\s*(.+)",
        r"(?:du\s*an|dự\s*án|chien\s*dich|chiến\s*dịch)\s*[:\-–]\s*(.+)",
        r"^(.+?)\s*[\-–—]\s*(?:lead|marketing|quang cao|quảng cáo)",
    ):
        m = re.search(pat, raw, re.IGNORECASE)
        if m:
            return m.group(1).strip()[:120]
    first = raw.split("\n")[0].strip()
    if len(first) <= 100:
        return first
    return first[:80].strip() + "…"


def _extract_budget(text: str) -> str:
    m = re.search(
        r"(\d[\d.,\s]*)\s*(?:ty|tỷ|tr|triệu|m\b|k\b|vnd|vnđ|dong|đồng)?",
        str(text or ""),
        re.IGNORECASE,
    )
    if not m:
        return "Chưa xác định — cập nhật sau workshop"
    val = m.group(0).strip()
    if "ty" in val.lower() or "tỷ" in val.lower():
        return val
    if "tr" in val.lower() or "triệu" in val.lower():
        return val + " VNĐ"
    return val + " VNĐ (ước tính)"


def _extract_duration(text: str) -> str:
    raw = str(text or "")
    for pat in (
        r"(Q[1-4]\s*/?\s*20\d{2})",
        r"(\d+\s*tuần)",
        r"(\d+\s*tháng)",
        r"(tháng\s*\d+\s*/?\s*20\d{2})",
    ):
        m = re.search(pat, raw, re.IGNORECASE)
        if m:
            return m.group(1)
    return "12 tuần (quý hiện tại)"


def _extract_icp(text: str) -> str:
    raw = str(text or "")
    for pat in (
        r"(?:icp|doi tuong|đối tượng|khach hang|khách hàng|phan khuc|phân khúc)\s*[:\-–]\s*(.+)",
        r"(B2B[^.\n]{5,80})",
        r"(B2C[^.\n]{5,80})",
    ):
        m = re.search(pat, raw, re.IGNORECASE)
        if m:
            return m.group(1).strip()[:200]
    return "Cập nhật theo brief — ngành, quy mô, địa lý, pain chính"


def build_campaign_plan_rule(brief: str, *, brand: str = "PTT") -> dict[str, Any]:
    objective = _detect_objective(brief)
    project = _extract_project_name(brief)
    obj_labels = {
        "lead": "Lead gen / chuyển đổi",
        "awareness": "Nhận diện thương hiệu",
        "retention": "Giữ chân & upsell",
    }
    obj_label = obj_labels.get(objective, "Lead gen")
    smart = f"Mục tiêu {obj_label} cho {project} — CPL/ROAS theo ngưỡng PTT, funnel MQL→SQL≥25%"
    channels = [dict(c) for c in OBJECTIVE_CHANNELS.get(objective, OBJECTIVE_CHANNELS["lead"])]
    steps = []
    for bp in STEP_BLUEPRINTS:
        steps.append({**bp, "custom_note": ""})
    timeline = _build_timeline(steps, project)
    return {
        "project_name": project,
        "campaign_name": f"{project} — {obj_label}",
        "brand": brand,
        "objective": objective,
        "objective_label": obj_label,
        "smart_goal": smart,
        "icp": _extract_icp(brief),
        "duration": _extract_duration(brief),
        "budget": _extract_budget(brief),
        "brief_raw": brief[:4000],
        "channels": channels,
        "steps": steps,
        "kpis": [dict(k) for k in DEFAULT_KPIS],
        "risks": [dict(r) for r in DEFAULT_RISKS],
        "timeline": timeline,
        "source": "rule",
    }


def _build_timeline(steps: list[dict[str, str]], project: str) -> list[dict[str, str]]:
    rows = []
    week_map = {
        "1": list(range(1, 3)),
        "2": list(range(2, 6)),
        "3": list(range(1, 13)),
        "4": list(range(1, 4)),
        "5": list(range(1, 13)),
        "6": [1],
        "7": list(range(3, 9)),
    }
    for bp in steps:
        num = bp["num"]
        weeks = week_map.get(num, [1])
        for w in weeks[:4]:
            rows.append(
                {
                    "week": f"W{w}",
                    "step": f"B{num}",
                    "activity": f"{bp['title']} — {bp['goal'][:60]}",
                    "deliverable": bp["deliverable"],
                    "owner": "MKT Lead",
                }
            )
    return rows[:24]


def _openai_campaign_plan(brief: str, *, brand: str, api_key: str) -> dict[str, Any] | None:
    system = (
        "Bạn là chuyên gia chiến lược marketing PTT. Phân tích brief dự án và trả về JSON duy nhất "
        "(không markdown) với các key: project_name, campaign_name, objective (lead|awareness|retention), "
        "objective_label, smart_goal, icp, duration, budget, channels (array of {name, goal, kpi, budget_pct}), "
        "steps (array 7 phần tử: {num, title, weeks, goal, inputs, process, deliverable, kpis, risks}), "
        "kpis (array: {name, target, freq, owner}), risks (array: {id, name, prob, impact, mitigation}). "
        "Giữ đúng 7 bước thực thi marketing PTT."
    )
    payload = json.dumps(
        {
            "model": os.environ.get("OPENAI_MODEL") or os.environ.get("AI_CHAT_MODEL") or "gpt-4o-mini",
            "messages": [
                {"role": "system", "content": system},
                {"role": "user", "content": f"Brand: {brand}\n\nBrief:\n{brief[:3500]}"},
            ],
            "max_tokens": 2800,
            "temperature": 0.4,
            "response_format": {"type": "json_object"},
        },
        ensure_ascii=False,
    ).encode("utf-8")
    req = urllib.request.Request(
        "https://api.openai.com/v1/chat/completions",
        data=payload,
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=45) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        raw = str(data["choices"][0]["message"]["content"]).strip()
        parsed = json.loads(raw)
        if not isinstance(parsed, dict):
            return None
        base = build_campaign_plan_rule(brief, brand=brand)
        for key in (
            "project_name",
            "campaign_name",
            "objective",
            "objective_label",
            "smart_goal",
            "icp",
            "duration",
            "budget",
        ):
            if parsed.get(key):
                base[key] = str(parsed[key]).strip()
        if isinstance(parsed.get("channels"), list) and parsed["channels"]:
            base["channels"] = parsed["channels"]
        if isinstance(parsed.get("steps"), list) and len(parsed["steps"]) >= 5:
            base["steps"] = parsed["steps"]
        if isinstance(parsed.get("kpis"), list) and parsed["kpis"]:
            base["kpis"] = parsed["kpis"]
        if isinstance(parsed.get("risks"), list) and parsed["risks"]:
            base["risks"] = parsed["risks"]
        base["timeline"] = _build_timeline(
            [s if isinstance(s, dict) else {} for s in base.get("steps", [])],
            base["project_name"],
        )
        base["source"] = "openai"
        return base
    except Exception:
        return None


def build_campaign_plan(brief: str, settings: dict[str, Any] | None = None) -> dict[str, Any]:
    s = settings or {}
    brand = str(s.get("brand_name") or "PTT Advertising Solutions").strip()
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    plan = None
    if api_key:
        plan = _openai_campaign_plan(brief, brand=brand, api_key=api_key)
    if not plan:
        plan = build_campaign_plan_rule(brief, brand=brand)
    plan["brand"] = brand
    plan["slug"] = _slugify(plan.get("project_name") or "du-an")
    return plan


def build_campaign_reply_markdown(plan: dict[str, Any]) -> str:
    proj = plan.get("project_name") or "Dự án"
    lines = [
        f"## Phân tích chiến dịch — **{proj}**",
        "",
    ]
    ptt_pf = plan.get("ptt_portfolio")
    if isinstance(ptt_pf, dict) and ptt_pf.get("title"):
        lines.extend(
            [
                f"📁 **Dự án PTT portfolio:** {ptt_pf.get('title')} ({ptt_pf.get('category', '')})",
                "",
            ]
        )
    market = plan.get("market_research")
    if isinstance(market, dict) and market.get("project_name"):
        lines.extend(
            [
                "### Nghiên cứu thị trường",
                "",
                f"- **Chủ đầu tư:** {market.get('developer', '—')}",
                f"- **Vị trí:** {market.get('location', '—')}",
                f"- **Phân khúc:** {market.get('segment', '—')}",
                f"- **USP:** {market.get('usp', '—')}",
                "",
            ]
        )
    lines.extend(
        [
            f"**Mục tiêu:** {plan.get('objective_label', '')} · **Thời gian:** {plan.get('duration', '')} · **Ngân sách:** {plan.get('budget', '')}",
            "",
            f"**SMART:** {plan.get('smart_goal', '')}",
            "",
            f"**ICP:** {plan.get('icp', '')}",
            "",
            "### 7 bước triển khai (tổng hợp vào KHMKT.xlsx)",
            "",
            "| Bước | Sheet | Nội dung |",
            "|---|---|---|",
            "| 1 | Buoc1_Ads | Mẫu FB/Google Ads |",
            "| 2 | Buoc2_TVC_KOL | TVC & video KOL |",
            "| 3 | Buoc3_Excel_tuan | Excel KH tuần |",
            "| 4 | Buoc4_Funnel | Funnel lead→KH |",
            "| 5 | Buoc5_Telesales | Script telesales |",
            "| 6 | Buoc6_Da_kenh | KH truyền thông đa kênh |",
            "| 7 | Buoc7_Test_kenh | Phương pháp test kênh |",
            "| + | KPI_do_luong | KPI & đo lường |",
            "| + | Chi_so_quan_tri | Chỉ số quản lý |",
            "| + | Quan_ly_rui_ro | Quản lý rủi ro |",
            "| + | Quan_ly_ngan_sach | Quản lý ngân sách |",
            "| + | Loi_nhuan_ROMI | Lợi nhuận & ROMI |",
            "",
            "| Bước | Nội dung | Tuần | Deliverable chính |",
            "|---|---|---|---|",
        ]
    )
    for st in plan.get("steps", [])[:7]:
        if not isinstance(st, dict):
            continue
        lines.append(
            f"| {st.get('num', '')} | {st.get('title', '')} | {st.get('weeks', '')} | {st.get('deliverable', '')} |"
        )
    lines.extend(
        [
            "",
            "### KPI trọng tâm",
            "",
            "| Chỉ số | Mục tiêu | Tần suất | Owner |",
            "|---|---|---|---|",
        ]
    )
    for k in plan.get("kpis", [])[:8]:
        if isinstance(k, dict):
            lines.append(
                f"| {k.get('name', '')} | {k.get('target', '')} | {k.get('freq', '')} | {k.get('owner', '')} |"
            )
    lines.extend(
        [
            "",
            "### Kênh trọng tâm",
            "",
        ]
    )
    for ch in plan.get("channels", [])[:6]:
        if isinstance(ch, dict):
            lines.append(
                f"- **{ch.get('name', '')}** ({ch.get('budget_pct', '')}%): {ch.get('goal', '')} — KPI {ch.get('kpi', '')}"
            )
    lines.extend(
        [
            "",
            "### Rủi ro cần theo dõi",
            "",
        ]
    )
    for r in plan.get("risks", [])[:5]:
        if isinstance(r, dict):
            lines.append(f"- **{r.get('id', '')}** {r.get('name', '')}: {r.get('mitigation', '')}")
    src = plan.get("source") or "rule"
    lines.extend(
        [
            "",
            f"_Phân tích bằng {'OpenAI + khung PTT' if src == 'openai' else 'khung playbook PTT'}._",
            "",
            "📥 **File đã sẵn sàng:** **KHMKT.xlsx** (12 module playbook + lịch + ngân sách) + **KPI.xlsx** (dashboard vùng RAG).",
            "Hệ thống sẽ tự tải 2 file — hoặc bấm nút bên dưới tin nhắn.",
        ]
    )
    return "\n".join(lines)


def _khmkt_ctx(plan: dict[str, Any]) -> dict[str, str]:
    """Ngữ cảnh dự án dùng chung cho các sheet playbook."""
    project = str(plan.get("project_name") or "Dự án")
    slug = str(plan.get("slug") or "du-an")
    budget_num = _parse_budget_vnd(str(plan.get("budget") or ""))
    return {
        "project": project,
        "campaign": str(plan.get("campaign_name") or project),
        "icp": str(plan.get("icp") or ""),
        "objective": str(plan.get("objective_label") or "Lead gen"),
        "duration": str(plan.get("duration") or "12 tuần"),
        "budget": str(plan.get("budget") or ""),
        "budget_num": str(int(budget_num)),
        "smart": str(plan.get("smart_goal") or ""),
        "slug": slug,
    }


def _write_playbook_sheet(
    ws,
    *,
    module_title: str,
    ctx: dict[str, str],
    sections: list[tuple[str, list[str], list[list[Any]] | None]],
    Font,
    Alignment,
) -> None:
    """Ghi 1 sheet playbook: sections = (tiêu đề section, headers, rows)."""
    ws.cell(row=1, column=1, value=f"{module_title} — {ctx['project']}").font = Font(bold=True, size=12)
    ws.cell(row=2, column=1, value=f"Chiến dịch: {ctx['campaign']} · ICP: {ctx['icp'][:80]}").font = Font(
        italic=True, color="666666"
    )
    row = 4
    for sec_title, headers, rows in sections:
        ws.cell(row=row, column=1, value=sec_title).font = Font(bold=True, size=11, color="2F7238")
        row += 1
        if headers:
            for c, h in enumerate(headers, 1):
                ws.cell(row=row, column=c, value=h).font = Font(bold=True)
            row += 1
        if rows:
            for data_row in rows:
                for c, val in enumerate(data_row, 1):
                    ws.cell(row=row, column=c, value=val)
                row += 1
        row += 1
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 22


def _build_khmkt_playbook_modules(wb, plan: dict[str, Any], *, write_header, Font, Alignment) -> None:
    """12 sheet playbook: 7 bước thực thi + KPI + quản trị."""
    ctx = _khmkt_ctx(plan)
    slug = ctx["slug"]
    channels = plan.get("channels") or []
    ch_names = ", ".join(str(c.get("name", "")) for c in channels[:4] if isinstance(c, dict))

    # --- 1. FB/Google Ads ---
    ws1 = wb.create_sheet("Buoc1_Ads")
    _write_playbook_sheet(
        ws1,
        module_title="BƯỚC 1 — Mẫu FB/Google Ads",
        ctx=ctx,
        sections=[
            ("A. Mục tiêu", [], [[ctx["objective"], "≥3 variant copy/kênh", "UTM + pixel/CAPI", "Launch 14 ngày"]]),
            ("B. Input dự án", ["Hạng mục", "Giá trị / ghi chú"], [
                ["ICP", ctx["icp"]], ["Offer / pain", ""], ["Landing URL", ""],
                ["Ngân sách test/kênh", f"~{int(_parse_budget_vnd(ctx['budget']) * 0.1 / max(len(channels), 1)):,} VNĐ"],
                ["Kênh ưu tiên", ch_names or "Google + Meta"],
            ]),
            ("C. Quy trình 7 bước con", ["#", "Bước"], [
                ["1", "Chốt 1 pain + 1 offer"], ["2", "3 góc: Pain · Proof · Offer"],
                ["3", "Meta: primary + 3 headline + CTA Lead"], ["4", "Google RSA + extensions"],
                ["5", "Brief visual 1:1 + 9:16"], ["6", f"UTM utm_campaign={slug}"], ["7", "A/B 7 ngày — review CPL"],
            ]),
            ("D. Template Meta (3 variant)", ["Variant", "Primary text (hook)", "Headline", "CTA"], [
                ["A — Pain", f"CPL cao với {ctx['project']}? Audit funnel miễn phí 30 phút.", "Giảm CPL — Audit free", "Lead"],
                ["B — Proof", "Case thực tế — CPL giảm 35% sau 4 tuần pilot.", "200+ dự án đa kênh", "Lead"],
                ["C — Offer", "Pilot 4 tuần: creative + landing + QA tracking.", "Pilot KPI minh bạch", "Lead"],
            ]),
            ("E. Google RSA", ["Thành phần", "Nội dung"], [
                ["Headline 1–3", f"{ctx['project']} · Tối ưu CPL · Audit miễn phí"],
                ["Description 1–2", f"{ctx['smart'][:120]}"],
                ["Extensions", "Sitelink · Callout · Snippet ngành"],
            ]),
            ("F. KPI & ngưỡng", ["Metric", "Mục tiêu", "Hành động nếu lệch"], [
                ["CTR Search", "≥3%", "Đổi headline RSA"], ["CTR Meta", "≥1%", "Refresh hook 3s"],
                ["CPL", "≤ target brief", "Pause variant >+30% CPL"], ["Landing CVR", "≥3%", "→ Bước 4 funnel"],
            ]),
            ("G. Checklist", ["#", "Hạng mục", "Done"], [
                ["1", "3 variant Meta + RSA duyệt", ""], ["2", "Pixel/CAPI test OK", ""],
                ["3", "UTM chuẩn gắn landing", ""], ["4", "Brief visual gửi designer", ""],
            ]),
        ],
        Font=Font,
        Alignment=Alignment,
    )

    # --- 2. TVC/KOL ---
    ws2 = wb.create_sheet("Buoc2_TVC_KOL")
    _write_playbook_sheet(
        ws2,
        module_title="BƯỚC 2 — TVC & Video KOL",
        ctx=ctx,
        sections=[
            ("A. Mục tiêu", [], [["Sản xuất video 15–30s + KOL", "21 ngày", "Legal + usage rights"]]),
            ("B. Lộ trình 21 ngày", ["Ngày", "Hoạt động", "Deliverable"], [
                ["D1–3", "Brief + casting KOL", "Shortlist 3 KOL"], ["D4–7", "Kịch bản / beat sheet", "Script v1"],
                ["D8–12", "Quay + b-roll", "Raw footage"], ["D13–17", "Dựng + subtitle", "Video draft"],
                ["D18–19", "Legal review + QC", "Signed release"], ["D20–21", "Launch + UTM", "Live ads"],
            ]),
            ("C. Beat sheet template", ["Beat", "Thời lượng", "Visual", "VO/Text"], [
                ["Hook", "0–3s", "Pain ICP", "Câu hỏi mở"], ["Problem", "3–10s", "B-roll", "Pain cụ thể"],
                ["Solution", "10–20s", "Product/demo", f"USP {ctx['project']}"], ["CTA", "20–30s", "Logo + form", "CTA rõ"],
            ]),
            ("D. KPI", ["Chỉ số", "Mục tiêu"], [["VTR", "≥25%"], ["CPV", "≤800 VNĐ"], ["CPL ref", "So sánh static ads"]]),
            ("E. Checklist legal", ["#", "Hạng mục"], [
                ["1", "Usage rights 6–12 tháng"], ["2", "Whitelisting Meta/TikTok"], ["3", "Không claim sai quy định ngành"],
            ]),
        ],
        Font=Font,
        Alignment=Alignment,
    )

    # --- 3. Excel KH tuần ---
    ws3 = wb.create_sheet("Buoc3_Excel_tuan")
    _write_playbook_sheet(
        ws3,
        module_title="BƯỚC 3 — Excel kế hoạch marketing tuần",
        ctx=ctx,
        sections=[
            ("A. Ritual vận hành", ["Tần suất", "Hoạt động", "Output"], [
                ["Daily", "Spend + CPL + tracking", "Slack 5 phút Ads"], ["T4", "Mid-week pacing", "Điều chỉnh cap nếu lệch"],
                ["T6", "Họp KPI 60 phút", "Email báo cáo + action items"], ["Cuối tháng", "Scorecard + ROMI", "Reforecast quý"],
            ]),
            ("B. Template 12 tuần (cột)", ["Cột", "Mô tả"], [
                ["Tuần", "W1–W12"], ["Kênh", "Google/Meta/..."], ["Hạng mục/Creative", ""],
                ["Ngân sách", "VNĐ"], ["KPI mục tiêu", ""], ["KPI thực tế", ""], ["Owner", ""], ["Trạng thái", ""],
            ]),
            ("C. Mẫu email báo cáo T6", ["Mục", "Nội dung"], [
                ["1", "Spend vs budget MTD"], ["2", "CPL / MQL / SQL tuần"], ["3", "Top 3 learnings"],
                ["4", "Quyết định tuần tới (scale/pause/test)"], ["5", "Rủi ro mở"],
            ]),
            ("D. RACI", ["Vai trò", "Trách nhiệm"], [
                ["Ads Lead", "Daily spend/CPL"], ["MKT Lead", "Họp T6 + reforecast"],
                ["Sales Lead", "MQL→SQL"], ["Finance", "Pacing ngân sách"],
            ]),
        ],
        Font=Font,
        Alignment=Alignment,
    )

    # --- 4. Funnel ---
    ws4 = wb.create_sheet("Buoc4_Funnel")
    _write_playbook_sheet(
        ws4,
        module_title="BƯỚC 4 — Funnel Lead → Khách hàng",
        ctx=ctx,
        sections=[
            ("A. CRM setup (6 ngày)", ["Ngày", "Task"], [
                ["D1", "Map pipeline: MQL→SQL→Demo→Win"], ["D2", "Fields + scoring rules"],
                ["D3", "UTM + source tracking"], ["D4", "SLA alert ≤4h"], ["D5", "Dashboard funnel"],
                ["D6", "UAT + sales training"],
            ]),
            ("B. Email nurture D0–D14", ["Ngày", "Email", "Mục tiêu"], [
                ["D0", "Welcome + case study", "Confirm + proof"], ["D2", "Pain + solution", "Educate"],
                ["D5", "Social proof / webinar", "Trust"], ["D9", "Offer demo/pilot", "Conversion"],
                ["D14", "Break-up / last chance", "Re-engage"],
            ]),
            ("C. SLA & touchpoint", ["Giai đoạn", "SLA", "Owner"], [
                ["MQL mới", "Phản hồi ≤4h", "SDR"], ["Connected", "Demo book ≤48h", "SDR"],
                ["SQL", "Handoff AE ≤24h", "Sales Lead"],
            ]),
            ("D. KPI funnel", ["Chỉ số", "Mục tiêu"], [
                ["MQL→SQL", "≥25%"], ["Contact rate", "≥60%"], ["Time to first touch", "≤4h"],
            ]),
        ],
        Font=Font,
        Alignment=Alignment,
    )

    # --- 5. Telesales ---
    ws5 = wb.create_sheet("Buoc5_Telesales")
    _write_playbook_sheet(
        ws5,
        module_title="BƯỚC 5 — Script Telesales / SDR",
        ctx=ctx,
        sections=[
            ("A. SOP hàng ngày SDR", ["Khung giờ", "Hoạt động"], [
                ["9:00", "Review MQL mới + SLA"], ["9:30–12:00", "Block gọi outbound"],
                ["14:00–17:00", "Follow-up + CRM log"], ["17:00", "Báo cáo contact rate"],
            ]),
            ("B. Script SPIN (mẫu)", ["Giai đoạn", "Câu hỏi / thoại"], [
                ["Situation", f"Anh/chị đang triển khai marketing {ctx['objective'].lower()} thế nào?"],
                ["Problem", "Thách thức lớn nhất về CPL / lead quality?"], ["Implication", "Nếu không cải thiện, ảnh hưởng pipeline?"],
                ["Need-payoff", "Nếu giảm CPL 30% trong 4 tuần, giá trị thế nào?"], ["Close", "Đặt demo 30 phút thứ …"],
            ]),
            ("C. Phản đối thường gặp", ["Phản đối", "Trả lời"], [
                ["Chưa có ngân sách", "Pilot nhỏ đo ROI trước — case …"], ["Đã có agency", "Audit second opinion miễn phí"],
                ["Gửi email", "Confirm email + 1 câu tóm value + hẹn 10 phút"],
            ]),
            ("D. CRM log bắt buộc", ["Field", "Giá trị"], [
                ["Disposition", "Connected / No answer / Wrong number"], ["Next step + date", ""],
                ["Objection tag", ""], ["SQL criteria met?", "Y/N"],
            ]),
        ],
        Font=Font,
        Alignment=Alignment,
    )

    # --- 6. Đa kênh ---
    ch_rows: list[list[Any]] = [
        [c.get("name", ""), c.get("goal", ""), c.get("kpi", ""), c.get("budget_pct", "")]
        for c in channels[:8] if isinstance(c, dict)
    ] or [["Google", "Conversion", "CPL", "35%"], ["Meta", "Lead", "CPL", "25%"]]
    ws6 = wb.create_sheet("Buoc6_Da_kenh")
    _write_playbook_sheet(
        ws6,
        module_title="BƯỚC 6 — KH truyền thông đa kênh",
        ctx=ctx,
        sections=[
            ("A. Workshop 3h — agenda", ["Phút", "Nội dung"], [
                ["0–30", "Mục tiêu SMART + ICP recap"], ["30–90", "Customer journey + key message funnel"],
                ["90–150", "Media mix + phân bổ budget"], ["150–180", "Calendar + owner + KPI/kênh"],
            ]),
            ("B. Phân bổ budget", ["Loại", "%", "Mục đích"], [
                ["Base (validate)", "65–70%", "Kênh đã chứng minh CPL"], ["Test", "20–25%", "Creative/audience mới"],
                ["Reserve", "10%", "Dự phòng / brand"],
            ]),
            ("C. Kênh chiến dịch", ["Kênh", "Mục tiêu", "KPI", "% NS"], ch_rows),
            ("D. Key message funnel", ["Giai đoạn", "Message", "Kênh chính"], [
                ["Awareness", "Problem + brand", "Video/TikTok"], ["Consideration", "Proof + case", "Meta/Google"],
                ["Conversion", "Offer + urgency", "Search + retarget"],
            ]),
        ],
        Font=Font,
        Alignment=Alignment,
    )

    # --- 7. Test kênh ---
    ws7 = wb.create_sheet("Buoc7_Test_kenh")
    _write_playbook_sheet(
        ws7,
        module_title="BƯỚC 7 — Phương pháp test kênh",
        ctx=ctx,
        sections=[
            ("A. TEST-ID register", ["TEST-ID", "Giả thuyết", "Kênh", "KPI success", "Budget", "Trạng thái"], [
                [f"TEST-{slug[:6].upper()}-01", "Hook video mới", "Meta", "CPL ≤ target", "10M", "Planned"],
                [f"TEST-{slug[:6].upper()}-02", "Long-tail KW", "Google", "CPL ≤ target", "8M", "Planned"],
                ["", "", "", "", "", ""],
            ]),
            ("B. Quy tắc sample & scale", ["Quy tắc", "Chi tiết"], [
                ["Min sample", "≥30 leads hoặc ≥7 ngày data"], ["Scale", "+20% budget/tuần nếu CPL ≤ target"],
                ["Kill", "Pause nếu CPL >+30% sau min sample"], ["Post-mortem", "Template 1 trang FAIL/PASS"],
            ]),
            ("C. Standup T3 (15 phút)", ["#", "Câu hỏi"], [
                ["1", "Test nào đang chạy?"], ["2", "CPL vs control?"], ["3", "Scale / pause / iterate?"],
            ]),
        ],
        Font=Font,
        Alignment=Alignment,
    )

    # --- 8. KPI & đo lường ---
    kpis = plan.get("kpis") or DEFAULT_KPIS
    ws8 = wb.create_sheet("KPI_do_luong")
    _write_playbook_sheet(
        ws8,
        module_title="KPI & ĐO LƯỜNG",
        ctx=ctx,
        sections=[
            ("A. KPI theo funnel", ["Giai đoạn", "Leading", "Lagging", "Tần suất"], [
                ["KHTN", "CTR, CPC", "CPL, MQL rate", "Daily/Weekly"],
                ["KHQT", "Contact rate", "MQL→SQL, demo book", "Weekly"],
                ["CSKH", "NPS response", "Retention, LTV", "Monthly/Quarterly"],
                ["Tài chính", "Spend pacing", "ROAS, ROMI, LTV/CAC", "Weekly/Monthly"],
            ]),
            ("B. KPI chiến dịch (dự án)", ["Chỉ số", "Mục tiêu", "Tần suất", "Owner"], [
                [k.get("name", ""), k.get("target", ""), k.get("freq", ""), k.get("owner", "")]
                for k in kpis if isinstance(k, dict)
            ]),
            ("C. Cadence review", ["Tần suất", "Focus"], [
                ["Daily", "Spend, CPL, tracking"], ["Weekly T6", "Funnel + kênh"], ["Monthly", "Scorecard + ROMI"],
            ]),
        ],
        Font=Font,
        Alignment=Alignment,
    )

    # --- 9. Chỉ số quản lý ---
    ws9 = wb.create_sheet("Chi_so_quan_tri")
    _write_playbook_sheet(
        ws9,
        module_title="CHỈ SỐ QUẢN LÝ (Scorecard / BSC)",
        ctx=ctx,
        sections=[
            ("A. 5 trụ scorecard", ["Trụ", "Chỉ số", "Công thức", "Nguồn"], [
                ["Thu hút", "Reach, branded search", "Trend YoY", "GA4/Search Console"],
                ["Chuyển đổi", "CPL, MQL→SQL, win rate", "CRM stages", "CRM"],
                ["Hiệu quả CP", "CAC, cost/MQL", "Spend/conversions", "Finance+Ads"],
                ["Giá trị KH", "LTV, NPS, churn", "Cohort", "CRM+Survey"],
                ["Vận hành", "SLA lead, tracking accuracy", "CRM vs Ads", "Dev+CRM"],
            ]),
            ("B. Weighted scorecard tháng", ["Chỉ số", "Trọng số", "T1 MT", "T1 TT", "Owner"], [
                ["CPL", "20%", "", "", "Ads"], ["MQL", "15%", "", "", "MKT"],
                ["SQL", "15%", "", "", "Sales"], ["ROAS", "15%", "", "", "MKT Lead"],
                ["ROMI", "10%", "", "", "Finance"], ["NPS", "5%", "", "", "CS"],
            ]),
        ],
        Font=Font,
        Alignment=Alignment,
    )

    # --- 10. Quản lý rủi ro ---
    risks = plan.get("risks") or DEFAULT_RISKS
    ws10 = wb.create_sheet("Quan_ly_rui_ro")
    _write_playbook_sheet(
        ws10,
        module_title="QUẢN LÝ RỦI RO",
        ctx=ctx,
        sections=[
            ("A. Ma trận rủi ro", ["ID", "Rủi ro", "P(1-5)", "I(1-5)", "Điểm P×I", "Giảm thiểu", "Owner"], [
                [r.get("id", ""), r.get("name", ""), r.get("prob", ""), r.get("impact", ""),
                 (int(r.get("prob", 0) or 0) * int(r.get("impact", 0) or 0)) if str(r.get("prob", "")).isdigit() else "",
                 r.get("mitigation", ""), "MKT Lead"]
                for r in risks if isinstance(r, dict)
            ]),
            ("B. Quy trình", ["Bước", "Hoạt động"], [
                ["1", "Identify — review tuần + Canh_bao"], ["2", "Assess — P×I ≥10 vàng, ≥15 đỏ"],
                ["3", "Mitigate — owner + deadline 48h"], ["4", "Monitor — monthly review"],
            ]),
            ("C. Rủi ro đặc thù ngành", ["Rủi ro", "Trigger", "Hành động"], [
                ["Policy ads", "Ad rejected", "Pause + pre-approval"], ["Tracking lỗi", "CRM ≠ Ads >10%", "QA ngay"],
                ["Overspend", ">110% MTD", "Giảm cap 15%"],
            ]),
        ],
        Font=Font,
        Alignment=Alignment,
    )

    # --- 11. Quản lý ngân sách ---
    bnum = int(_parse_budget_vnd(ctx["budget"]))
    ws11 = wb.create_sheet("Quan_ly_ngan_sach")
    _write_playbook_sheet(
        ws11,
        module_title="QUẢN LÝ NGÂN SÁCH",
        ctx=ctx,
        sections=[
            ("A. Tổng quan ngân sách quý", ["Hạng mục", "Giá trị (VNĐ)", "Ghi chú"], [
                ["Ngân sách quý", bnum, ctx["duration"]],
                ["Dự phòng 10%", int(bnum * 0.1), "Không dùng trừ khi approved"],
                ["Ngân sách ads (90%)", int(bnum * 0.9), "Base + Test + Reserve"],
                ["Đã chi (nhập tay)", "", "Cập nhật weekly"],
                ["Còn lại", "", "= Ngân sách − Đã chi"],
            ]),
            ("B. Phân bổ theo kênh", ["Kênh", "%", "Ngân sách KH", "Đã chi", "Còn lại"], [
                [c.get("name", ""), c.get("budget_pct", ""), "", "", ""]
                for c in channels[:8] if isinstance(c, dict)
            ] or [["Google Search", "35%", int(bnum * 0.35), "", ""]]),
            ("C. Quy tắc reallocate", ["Quy tắc", "Chi tiết"], [
                ["Checkpoint T6", "Review pacing vs MTD"], ["Reallocate", "Max 15%/tuần giữa kênh cùng mục tiêu"],
                ["Kill switch", "CPL >350k 5 ngày → pause scale"], ["Dự phòng", "CEO approve nếu >10%"],
            ]),
        ],
        Font=Font,
        Alignment=Alignment,
    )

    # --- 12. Lợi nhuận & ROMI ---
    ws12 = wb.create_sheet("Loi_nhuan_ROMI")
    _write_playbook_sheet(
        ws12,
        module_title="LỢI NHUẬN & ROMI",
        ctx=ctx,
        sections=[
            ("A. Công thức cốt lõi", ["Chỉ số", "Công thức", "Mục tiêu"], [
                ["CAC", "Marketing spend / New customers", "Giảm QoQ"],
                ["LTV", "ARPU × Gross margin × Lifetime", "≥3× CAC"],
                ["LTV/CAC", "LTV / CAC", "≥3"], ["ROAS", "Revenue / Ad spend", "≥3"],
                ["ROMI", "(Revenue − Cost) / Cost", "≥150%"], ["Payback", "CAC / Monthly gross profit", "≤12 tháng"],
            ]),
            ("B. ROMI theo kênh (nhập TT)", ["Kênh", "Chi (VNĐ)", "Doanh thu", "ROAS", "ROMI"], [
                [c.get("name", ""), "", "", "", ""] for c in channels[:6] if isinstance(c, dict)
            ] or [["Google", "", "", "", ""], ["Meta", "", "", "", ""]]),
            ("C. Giả định minh họa", ["Chỉ số", "Giá trị"], [
                ["AOV / contract value", ""], ["Gross margin", "40–60%"], ["Win rate", "≥20%"],
                ["Retention 12 tháng", "≥85%"],
            ]),
            ("D. Khuyến nghị tối ưu", ["#", "Hành động"], [
                ["1", "Scale kênh ROAS/ROMI cao nhất (+20%/tuần)"], ["2", "Pause kênh ROAS <2 qua 2 tuần"],
                ["3", "Tăng LTV: nurture + upsell (Bước 4)"], ["4", "Review cohort monthly với Finance"],
            ]),
        ],
        Font=Font,
        Alignment=Alignment,
    )


def build_khmkt_xlsx(plan: dict[str, Any], *, brand: str | None = None) -> bytes:
    """KHMKT.xlsx — Kế hoạch marketing chiến dịch theo 7 bước PTT."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    brand = brand or str(plan.get("brand") or "PTT Advertising Solutions")
    project = str(plan.get("project_name") or "Dự án")
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="2F7238")
    header_font = Font(bold=True, color="FFFFFF")
    sub_fill = PatternFill("solid", fgColor="398B43")

    def write_header(ws, headers: list[str], row: int = 1) -> None:
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Sheet 1 — Tổng quan
    ws = wb.active
    ws.title = "Tong_quan"
    ws.cell(row=1, column=1, value=f"KẾ HOẠCH MARKETING — {project}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Thương hiệu: {brand} · File: KHMKT.xlsx").font = Font(italic=True)
    overview = [
        ("Tên dự án / chiến dịch", plan.get("campaign_name") or project),
        ("Mục tiêu chiến lược", plan.get("objective_label", "")),
        ("Mục tiêu SMART", plan.get("smart_goal", "")),
        ("ICP / phân khúc", plan.get("icp", "")),
        ("Thời gian triển khai", plan.get("duration", "")),
        ("Ngân sách", plan.get("budget", "")),
        ("Marketing Lead", ""),
        ("Sales Lead", ""),
        ("Cadence", "Daily ads · Weekly KPI T6 · Monthly scorecard"),
    ]
    for i, (k, v) in enumerate(overview, 4):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 56

    # Mục lục 12 module
    ws_muc = wb.create_sheet("Muc_luc")
    write_header(ws_muc, ["#", "Module", "Sheet Excel", "Mô tả"])
    muc_rows = [
        ("1", "Mẫu FB/Google Ads", "Buoc1_Ads", "Copy Meta/Google, UTM, KPI, checklist"),
        ("2", "TVC & video KOL", "Buoc2_TVC_KOL", "Lộ trình 21 ngày, beat sheet, legal"),
        ("3", "Excel KH tuần", "Buoc3_Excel_tuan", "Ritual T6, template 12 tuần, RACI"),
        ("4", "Funnel lead→KH", "Buoc4_Funnel", "CRM 6 ngày, email D0–D14, SLA"),
        ("5", "Script telesales", "Buoc5_Telesales", "SPIN, phản đối, CRM log"),
        ("6", "KH truyền thông đa kênh", "Buoc6_Da_kenh", "Workshop 3h, media mix, calendar"),
        ("7", "Phương pháp test kênh", "Buoc7_Test_kenh", "TEST-ID, scale/kill protocol"),
        ("8", "KPI & đo lường", "KPI_do_luong", "Funnel KPI, cadence review"),
        ("9", "Chỉ số quản lý", "Chi_so_quan_tri", "BSC 5 trụ, scorecard weighted"),
        ("10", "Quản lý rủi ro", "Quan_ly_rui_ro", "Ma trận P×I, quy trình"),
        ("11", "Quản lý ngân sách", "Quan_ly_ngan_sach", "Phân bổ, pacing, reallocate"),
        ("12", "Lợi nhuận & ROMI", "Loi_nhuan_ROMI", "CAC, LTV, ROAS, ROMI theo kênh"),
    ]
    for r, row in enumerate(muc_rows, 2):
        for c, val in enumerate(row, 1):
            ws_muc.cell(row=r, column=c, value=val)
    ws_muc.column_dimensions["A"].width = 6
    ws_muc.column_dimensions["B"].width = 28
    ws_muc.column_dimensions["C"].width = 18
    ws_muc.column_dimensions["D"].width = 42

    # Sheet — 7 bước tóm tắt
    ws7 = wb.create_sheet("7_Buoc_trien_khai")
    h7 = [
        "Bước",
        "Tên bước",
        "Tuần",
        "Mục tiêu",
        "Input cần có",
        "Quy trình tóm tắt",
        "Deliverable",
        "KPI bước",
        "Rủi ro",
        "Trạng thái",
        "Owner",
    ]
    write_header(ws7, h7)
    for r, st in enumerate(plan.get("steps", [])[:7], 2):
        if not isinstance(st, dict):
            continue
        row = (
            st.get("num", ""),
            st.get("title", ""),
            st.get("weeks", ""),
            st.get("goal", ""),
            st.get("inputs", ""),
            st.get("process", ""),
            st.get("deliverable", ""),
            st.get("kpis", ""),
            st.get("risks", ""),
            "Planned",
            "MKT Lead",
        )
        for c, val in enumerate(row, 1):
            ws7.cell(row=r, column=c, value=val)
    for i, w in enumerate([6, 18, 10, 28, 22, 28, 22, 16, 18, 10, 12], 1):
        ws7.column_dimensions[get_column_letter(i)].width = w

    # Sheet 3 — Lịch 12 tuần
    ws_t = wb.create_sheet("Lich_12_tuan")
    ht = ["Tuần", "Bước", "Hoạt động", "Deliverable", "Owner", "Trạng thái", "Ghi chú"]
    write_header(ws_t, ht)
    for r, item in enumerate(plan.get("timeline", [])[:24], 2):
        if not isinstance(item, dict):
            continue
        ws_t.cell(row=r, column=1, value=item.get("week", ""))
        ws_t.cell(row=r, column=2, value=item.get("step", ""))
        ws_t.cell(row=r, column=3, value=item.get("activity", ""))
        ws_t.cell(row=r, column=4, value=item.get("deliverable", ""))
        ws_t.cell(row=r, column=5, value=item.get("owner", ""))
        ws_t.cell(row=r, column=6, value="Planned")
        ws_t.cell(row=r, column=7, value="")
    for i, w in enumerate([8, 8, 36, 28, 12, 12, 20], 1):
        ws_t.column_dimensions[get_column_letter(i)].width = w

    # Sheet 4 — Kênh & ngân sách
    ws_ch = wb.create_sheet("Kenh_ngan_sach")
    hc = ["Kênh", "Mục tiêu", "KPI", "% ngân sách", "Ngân sách (VNĐ)", "Owner", "Ghi chú"]
    write_header(ws_ch, hc)
    for r, ch in enumerate(plan.get("channels", []), 2):
        if not isinstance(ch, dict):
            continue
        ws_ch.cell(row=r, column=1, value=ch.get("name", ""))
        ws_ch.cell(row=r, column=2, value=ch.get("goal", ""))
        ws_ch.cell(row=r, column=3, value=ch.get("kpi", ""))
        ws_ch.cell(row=r, column=4, value=ch.get("budget_pct", ""))
        ws_ch.cell(row=r, column=5, value="")
        ws_ch.cell(row=r, column=6, value="MKT")
        ws_ch.cell(row=r, column=7, value="")
    for i, w in enumerate([16, 16, 14, 12, 16, 12, 20], 1):
        ws_ch.column_dimensions[get_column_letter(i)].width = w

    # Sheet 5 — KPI tóm tắt
    ws_k = wb.create_sheet("KPI_tom_tat")
    hk = ["Chỉ số", "Mục tiêu", "Tần suất đo", "Owner", "Nguồn dữ liệu", "Ghi chú"]
    write_header(ws_k, hk)
    for r, k in enumerate(plan.get("kpis", []), 2):
        if not isinstance(k, dict):
            continue
        ws_k.cell(row=r, column=1, value=k.get("name", ""))
        ws_k.cell(row=r, column=2, value=k.get("target", ""))
        ws_k.cell(row=r, column=3, value=k.get("freq", ""))
        ws_k.cell(row=r, column=4, value=k.get("owner", ""))
        ws_k.cell(row=r, column=5, value="CRM / Ads Manager")
        ws_k.cell(row=r, column=6, value="")
    for i, w in enumerate([16, 16, 12, 14, 18, 20], 1):
        ws_k.column_dimensions[get_column_letter(i)].width = w

    # Sheet 6 — Rủi ro
    ws_r = wb.create_sheet("Rui_ro")
    hr = ["ID", "Rủi ro", "Xác suất", "Tác động", "Biện pháp", "Owner", "Trạng thái"]
    write_header(ws_r, hr)
    for r, risk in enumerate(plan.get("risks", []), 2):
        if not isinstance(risk, dict):
            continue
        ws_r.cell(row=r, column=1, value=risk.get("id", ""))
        ws_r.cell(row=r, column=2, value=risk.get("name", ""))
        ws_r.cell(row=r, column=3, value=risk.get("prob", ""))
        ws_r.cell(row=r, column=4, value=risk.get("impact", ""))
        ws_r.cell(row=r, column=5, value=risk.get("mitigation", ""))
        ws_r.cell(row=r, column=6, value="MKT Lead")
        ws_r.cell(row=r, column=7, value="Open")
    for i, w in enumerate([8, 28, 10, 10, 32, 12, 10], 1):
        ws_r.column_dimensions[get_column_letter(i)].width = w

    # Sheet 7 — Checklist
    ws_ck = wb.create_sheet("Checklist")
    ws_ck.cell(row=1, column=1, value="CHECKLIST TRIỂN KHAI — tick khi hoàn thành").font = Font(bold=True, size=12)
    checks = [
        "B1 — Brief ads + UTM + pixel QA (Buoc1_Ads)",
        "B2 — TVC/KOL script + legal (Buoc2_TVC_KOL)",
        "B3 — Excel ritual T6 setup (Buoc3_Excel_tuan)",
        "B4 — CRM pipeline + SLA 4h (Buoc4_Funnel)",
        "B5 — SOP telesales + script (Buoc5_Telesales)",
        "B6 — Workshop đa kênh + media mix (Buoc6_Da_kenh)",
        "B7 — TEST-ID đăng ký trước scale (Buoc7_Test_kenh)",
        "KPI — Bộ KPI funnel + cadence (KPI_do_luong)",
        "Scorecard — 5 trụ chỉ số (Chi_so_quan_tri)",
        "Rủi ro — Ma trận P×I review (Quan_ly_rui_ro)",
        "Ngân sách — Phân bổ + pacing (Quan_ly_ngan_sach)",
        "ROMI — Model CAC/LTV/ROAS (Loi_nhuan_ROMI)",
        "File KPI.xlsx dashboard vùng RAG",
        "Họp kickoff Marketing + Sales + Finance",
    ]
    write_header(ws_ck, ["#", "Hạng mục", "Done", "Ngày", "Owner"], row=3)
    for i, ck in enumerate(checks, 4):
        ws_ck.cell(row=i, column=1, value=i - 3)
        ws_ck.cell(row=i, column=2, value=ck)
        ws_ck.cell(row=i, column=3, value="")
        ws_ck.cell(row=i, column=4, value="")
        ws_ck.cell(row=i, column=5, value="")
    ws_ck.column_dimensions["A"].width = 6
    ws_ck.column_dimensions["B"].width = 52

    # Sheet 8 — Brief gốc
    ws_b = wb.create_sheet("Brief_goc")
    ws_b.cell(row=1, column=1, value="Brief dự án (người dùng cung cấp)").font = Font(bold=True)
    ws_b.cell(row=3, column=1, value=str(plan.get("brief_raw") or ""))
    ws_b.column_dimensions["A"].width = 90

    _build_khmkt_playbook_modules(wb, plan, write_header=write_header, Font=Font, Alignment=Alignment)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_campaign_kpi_xlsx(plan: dict[str, Any], *, brand: str | None = None) -> bytes:
    brand = brand or str(plan.get("brand") or "PTT Advertising Solutions")
    return build_kpi_strategy_xlsx(brand=brand, campaign_plan=plan)


def _purge_cache() -> None:
    now = time.time()
    expired = [k for k, v in _campaign_kit_cache.items() if now - float(v.get("ts") or 0) > _CACHE_TTL_SEC]
    for k in expired:
        _campaign_kit_cache.pop(k, None)


def store_campaign_kit(plan: dict[str, Any]) -> str:
    _purge_cache()
    kit_id = uuid.uuid4().hex
    _campaign_kit_cache[kit_id] = {"plan": plan, "ts": time.time()}
    return kit_id


def get_campaign_kit(kit_id: str) -> dict[str, Any] | None:
    _purge_cache()
    entry = _campaign_kit_cache.get(str(kit_id or "").strip())
    if not entry:
        return None
    return entry.get("plan")


def generate_campaign_kit(
    brief: str,
    settings: dict[str, Any] | None = None,
    *,
    ptt_projects: list[dict[str, Any]] | None = None,
    force: bool = False,
) -> dict[str, Any]:
    analysis = analyze_campaign_brief(brief, ptt_projects=ptt_projects, force=force)
    if analysis["status"] == "needs_info":
        return {
            "ok": True,
            "status": "needs_info",
            "reply": analysis["clarification_reply"],
            "missing_fields": analysis["missing_fields"],
            "ptt_match_score": analysis["ptt_match_score"],
            "project_name": analysis.get("project_name"),
            "has_market_research": bool(analysis.get("market_intel")),
        }

    enriched = str(analysis.get("enriched_brief") or brief)
    plan = build_campaign_plan(enriched, settings)
    apply_intel_to_plan(plan, analysis)
    plan["brief_raw"] = brief[:4000]
    plan["slug"] = _slugify(plan.get("project_name") or "du-an")
    kit_id = store_campaign_kit(plan)
    slug = plan.get("slug") or "du-an"
    reply = build_campaign_reply_markdown(plan)
    return {
        "ok": True,
        "status": "ready",
        "kit_id": kit_id,
        "slug": slug,
        "project_name": plan.get("project_name"),
        "reply": reply,
        "khmkt_url": f"/api/cms/marketing-chat/campaign-kit/{kit_id}/khmkt.xlsx",
        "kpi_url": f"/api/cms/marketing-chat/campaign-kit/{kit_id}/kpi.xlsx",
        "source": plan.get("source"),
        "ptt_portfolio": plan.get("ptt_portfolio"),
        "has_market_research": bool(plan.get("market_research")),
    }


KPI_BRIEF_PROMPT = (
    "📋 **Tạo bộ KPI & KHMKT theo chiến dịch**\n\n"
    "Trợ lý sẽ **tra cứu dự án trong portfolio PTT** và **nghiên cứu thị trường** (nếu là dự án công khai) "
    "trước khi tạo file.\n\n"
    "Vui lòng cho biết **dự án / chiến dịch marketing**:\n\n"
    "- **Tên dự án / chiến dịch**\n"
    "- **Mục tiêu** (lead gen / giữ chỗ pre-launch / nhận diện thương hiệu…)\n"
    "- **Đối tượng khách hàng** (ICP)\n"
    "- **Thời gian** (quý, số tuần/tháng)\n"
    "- **Ngân sách** (nếu có)\n"
    "- **Kênh ưu tiên** hoặc bối cảnh thêm\n\n"
    "_Nếu thiếu thông tin, trợ lý sẽ hỏi lại. Gõ **「tạo file」** để tạo ngay với dữ liệu hiện có._\n\n"
    "_Ví dụ: Vinhomes Saigon Park — lead gen giữ chỗ pre-launch GĐ1, ICP gia đình trẻ TP.HCM, Q2/2026, ngân sách 800 triệu, ưu tiên Meta + Google._"
)
