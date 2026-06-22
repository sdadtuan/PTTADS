"""Mẫu thực thi marketing — Ads, TVC/KOL, Excel tuần, Funnel, Telesales."""
from __future__ import annotations

import re
import unicodedata
from io import BytesIO
from typing import Any

from marketing_step_templates import (
    ADS_TEMPLATE,
    CHANNEL_TEST_TEMPLATE,
    EXCEL_WEEKLY_TEMPLATE,
    FUNNEL_TEMPLATE,
    MULTICHANNEL_PLAN_TEMPLATE,
    STEP_NUMBER_MAP,
    STEP_TEMPLATES,
    TELESALES_TEMPLATE,
    TVC_KOL_TEMPLATE,
    extract_step_id,
    step_trigger,
)


def normalize_query(text: str) -> str:
    """Chuẩn hóa câu hỏi: lowercase, bỏ dấu tiếng Việt, gom khoảng trắng."""
    s = str(text or "").lower().replace("đ", "d").replace("Đ", "d")
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^\w\s]", " ", s, flags=re.UNICODE)
    return re.sub(r"\s+", " ", s).strip()

EXECUTION_MODULES: tuple[dict[str, str], ...] = (
    {
        "id": "ads_copy",
        "label": "1. Mẫu FB/Google Ads",
        "prompt": (
            "Playbook thực chiến Bước 1 — FB/Google Ads: brief 1 trang, lộ trình 14 ngày, "
            "cấu trúc campaign Meta, QA tracking, công thức CPL/CVR, lỗi thường gặp."
        ),
        "action": "prompt",
    },
    {
        "id": "tvc_kol",
        "label": "2. TVC & video KOL",
        "prompt": (
            "Playbook thực chiến Bước 2 — TVC/KOL: lộ trình sản xuất 21 ngày, brief creative, "
            "shot list, KPI VTR/CPL, checklist hợp đồng KOL pháp lý."
        ),
        "action": "prompt",
    },
    {
        "id": "excel_weekly",
        "label": "3. Excel KH tuần",
        "prompt": (
            "Playbook thực chiến Bước 3 — Excel 12 tuần: ritual hàng ngày/T4/T6, sheet rủi ro, "
            "công thức Excel, mẫu báo cáo email, RACI. Nhắc tải file nút XLS."
        ),
        "action": "download_xlsx",
    },
    {
        "id": "funnel",
        "label": "4. Funnel lead→KH",
        "prompt": (
            "Playbook thực chiến Bước 4 — Funnel: triển khai CRM 6 ngày, email D0-D14, "
            "SLA SDR, dashboard funnel, công thức conversion + mermaid."
        ),
        "action": "prompt",
    },
    {
        "id": "telesales",
        "label": "5. Script telesales",
        "prompt": (
            "Playbook thực chiến Bước 5 — Telesales: SOP hàng ngày SDR, script phản đối đầy đủ, "
            "template log CRM, KPI SDR, cadence follow-up 4 lần."
        ),
        "action": "prompt",
    },
    {
        "id": "multichannel_plan",
        "label": "6. KH truyền thông đa kênh",
        "prompt": (
            "Playbook thực chiến Bước 6 — Đa kênh 3 tháng: workshop 3h, key message funnel, "
            "quy tắc phân bổ budget, báo cáo T6, UTM doc. Nhắc Excel nút ĐK."
        ),
        "action": "download_mc_xlsx",
    },
    {
        "id": "channel_test",
        "label": "7. Phương pháp test kênh",
        "prompt": (
            "Playbook thực chiến Bước 7 — Test kênh: test standup T3, mẫu đăng ký TEST-ID, "
            "cách tính mẫu, scale protocol +20%, post-mortem FAIL."
        ),
        "action": "prompt",
    },
)

EXECUTION_KNOWLEDGE = """
=== THỰC THI CHIẾN DỊCH (7 BƯỚC — PLAYBOOK THỰC CHIẾN) ===

Khung A→G + H→L thực hành: mục tiêu, input, quy trình, deliverable, KPI, rủi ro, checklist,
lộ trình ngày/tuần, template copy-paste, công thức, RACI, lỗi thường gặp.

1) ADS — brief 1 trang, 14 ngày launch, Meta/Google structure, QA pixel, CPL pacing.
2) TVC/KOL — 21 ngày sản xuất, shot list, legal KOL, VTR/CPL ref.
3) EXCEL — ritual daily/T4/T6, rủi ro P×I, báo cáo email, RACI.
4) FUNNEL — CRM 6 ngày setup, email D0-D14, SLA SDR 4h, dashboard conversion.
5) TELESALES — SOP SDR, script phản đối, CRM log, follow-up cadence.
6) ĐA KÊNH — workshop 3h, budget base/test/reserve, báo cáo cross-channel T6.
7) TEST — standup T3, TEST-ID register, min sample, scale +20%/tuần, post-mortem.
""".strip()


def build_weekly_marketing_plan_xlsx(*, brand: str = "PTT Advertising Solutions") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Ke_hoach_tuan"

    headers = [
        "Tuần",
        "Kênh",
        "Chiến dịch",
        "Hạng mục / Creative",
        "Ngân sách (VNĐ)",
        "KPI mục tiêu",
        "KPI thực tế",
        "Owner",
        "Trạng thái",
        "Ghi chú",
    ]
    header_fill = PatternFill("solid", fgColor="398B43")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sample_rows = [
        ("W1", "Google Search", "Lead gen Q2", "RSA 3 headline + extensions", 15_000_000, "CPL ≤ 250k", "", "Ads Lead", "Planned", ""),
        ("W1", "Meta Ads", "Lead gen Q2", "Carousel case study", 12_000_000, "CPL ≤ 180k", "", "Social", "Planned", ""),
        ("W1", "Content/SEO", "Blog pillar", "2 bài SEO + internal link", 5_000_000, "500 visit organic", "", "Content", "Planned", ""),
        ("W1", "Email", "Nurture MQL", "Sequence D0-D7", 2_000_000, "Open ≥ 25%", "", "CRM", "Planned", ""),
        ("W2", "Google Display", "Remarketing", "7-day audience", 8_000_000, "CPA ≤ 300k", "", "Ads Lead", "Planned", ""),
        ("W2", "Meta", "Retargeting", "Video 15s + lead form", 10_000_000, "CPL ≤ 200k", "", "Social", "Planned", ""),
        ("W2", "Telesales", "SQL outreach", "Gọi MQL ≤4h SLA", 0, "Contact rate ≥ 60%", "", "Sales", "Planned", ""),
        ("W3", "Google Search", "Scale winner", "Tăng 20% ad set thắng", 18_000_000, "ROAS ≥ 3", "", "Ads Lead", "Planned", ""),
        ("W3", "Landing", "CRO test", "A/B headline + form", 3_000_000, "CVR +15%", "", "CRO", "Planned", ""),
        ("W4", "Review", "Monthly", "Báo cáo ROMI + reforecast", 0, "Hoàn thành báo cáo", "", "Marketing Lead", "Planned", "Họp review cuối tuần"),
    ]
    for r, row in enumerate(sample_rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    widths = [8, 14, 16, 28, 16, 16, 14, 14, 12, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws_kpi = wb.create_sheet("KPI_tong_hop")
    kpi_headers = ["Chỉ số", "Mục tiêu tháng", "Thực tế", "Đơn vị", "Ghi chú"]
    for col, h in enumerate(kpi_headers, 1):
        ws_kpi.cell(row=1, column=col, value=h).font = Font(bold=True)
    kpi_rows = [
        ("CPL trung bình", "", "", "VNĐ", ""),
        ("MQL", "", "", "lead", ""),
        ("SQL", "", "", "lead", ""),
        ("Win rate", "", "", "%", ""),
        ("ROAS", "", "", "x", ""),
        ("ROMI", "", "", "%", ""),
    ]
    for r, row in enumerate(kpi_rows, 2):
        for c, val in enumerate(row, 1):
            ws_kpi.cell(row=r, column=c, value=val)

    ws_budget = wb.create_sheet("Ngan_sach")
    ws_budget.cell(row=1, column=1, value=f"Kế hoạch ngân sách — {brand}").font = Font(bold=True, size=12)
    budget_headers = ["Kênh", "Ngân sách kế hoạch", "Đã chi", "Còn lại", "% dùng"]
    for col, h in enumerate(budget_headers, 1):
        ws_budget.cell(row=3, column=col, value=h).font = Font(bold=True)
    for r, ch in enumerate(["Google Ads", "Meta Ads", "Content/SEO", "Email/CRM", "Dự phòng"], 4):
        ws_budget.cell(row=r, column=1, value=ch)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_multichannel_plan_xlsx(*, brand: str = "PTT Advertising Solutions") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="2F7238")
    header_font = Font(bold=True, color="FFFFFF")

    def write_header(ws, headers: list[str], row: int = 1) -> None:
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws = wb.active
    ws.title = "Ke_hoach_da_kenh"
    mc_headers = [
        "Kênh",
        "Loại P/O/E",
        "Vai trò funnel",
        "Thông điệp",
        "Format / Asset",
        "Ngân sách/tháng (VNĐ)",
        "KPI mục tiêu",
        "KPI thực tế",
        "Owner",
        "Ghi chú",
    ]
    write_header(ws, mc_headers)
    mc_rows = [
        ("Google Search", "Paid", "Conversion", "Audit funnel miễn phí", "RSA + extensions", 45_000_000, "CPL ≤250k", "", "Ads", ""),
        ("Meta Ads", "Paid", "Consideration", "Case study + proof", "Carousel / Video", 36_000_000, "CPL ≤180k", "", "Social", ""),
        ("TikTok", "Paid", "Awareness", "Hook pain CPL cao", "Video 15s", 15_000_000, "CPV, brand search", "", "Social", ""),
        ("LinkedIn", "Paid", "B2B Lead", "Whitepaper martech", "Lead gen form", 20_000_000, "CPL ≤350k", "", "Ads", ""),
        ("Website/Landing", "Owned", "Conversion", "USP + form", "Landing A/B", 8_000_000, "CVR ≥4%", "", "CRO", ""),
        ("Email/CRM", "Owned", "Nurture", "Sequence D0-D14", "Automation", 3_000_000, "Open ≥25%", "", "CRM", ""),
        ("SEO/Content", "Owned", "Consideration", "Pillar AEO/SEO", "4 bài/tháng", 12_000_000, "Organic lead", "", "Content", ""),
        ("KOL/PR", "Earned", "Awareness", "Review dịch vụ", "Video 60s", 10_000_000, "Reach, referral", "", "PR", ""),
    ]
    for r, row in enumerate(mc_rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    for i, w in enumerate([14, 10, 14, 28, 22, 18, 14, 14, 12, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws_cal = wb.create_sheet("Lich_phat_sinh")
    cal_headers = ["Tuần", "Tháng", "Kênh", "Asset / Deliverable", "Deadline", "Owner", "Trạng thái", "Ghi chú"]
    write_header(ws_cal, cal_headers)
    cal_rows = [
        ("W1", "T1", "Google + Meta", "Launch campaign Q2", "Thứ 2", "Ads Lead", "Planned", ""),
        ("W1", "T1", "Email", "Nurture sequence v1", "Thứ 4", "CRM", "Draft", ""),
        ("W2", "T1", "TikTok", "3 video hook test", "Thứ 3", "Social", "Production", ""),
        ("W3", "T1", "Landing", "A/B form fields", "Thứ 5", "CRO", "Dev", ""),
        ("W4", "T1", "All", "Báo cáo cross-channel", "Thứ 6", "Marketing Lead", "Review", ""),
    ]
    for r, row in enumerate(cal_rows, 2):
        for c, val in enumerate(row, 1):
            ws_cal.cell(row=r, column=c, value=val)

    ws_test = wb.create_sheet("Test_matrix")
    test_headers = [
        "Kênh",
        "Giả thuyết",
        "Biến số test",
        "Control",
        "Variant",
        "Thời gian",
        "Mẫu tối thiểu",
        "Metric chính",
        "Ngưỡng thắng",
        "Kết quả",
        "Quyết định",
    ]
    write_header(ws_test, test_headers)
    test_rows = [
        ("Google Search", "Headline mới tăng CTR", "RSA headline", "H1 cũ", "H1 pain hook", "14 ngày", "≥300 click", "CPL", "↓15% vs control", "", ""),
        ("Meta", "Video beat static", "Creative type", "Static carousel", "Video 15s", "14 ngày", "≥50 conv", "CPL", "↓12% vs control", "", ""),
        ("TikTok", "Hook 3s A/B/C", "Opening 3s", "Hook A", "Hook B/C", "7 ngày", "≥50k views", "CPV/CTR", "CTR cao nhất", "", ""),
        ("Email", "Subject urgency", "Subject line", "Subject A", "Subject B", "1 send", "≥1000/branch", "Open rate", "↑10% open", "", ""),
        ("Landing", "Form ngắn hơn", "Form fields", "5 fields", "3 fields", "21 ngày", "≥200 clicks", "CVR", "↑10% CVR", "", ""),
    ]
    for r, row in enumerate(test_rows, 2):
        for c, val in enumerate(row, 1):
            ws_test.cell(row=r, column=c, value=val)

    ws_info = wb.create_sheet("Huong_dan")
    ws_info.cell(row=1, column=1, value=f"Kế hoạch truyền thông đa kênh — {brand}").font = Font(bold=True, size=12)
    notes = [
        "Sheet Ke_hoach_da_kenh: ma trận kênh Paid/Owned/Earned theo funnel.",
        "Sheet Lich_phat_sinh: lịch asset và deadline theo tuần.",
        "Sheet Test_matrix: đăng ký giả thuyết, variant, ngưỡng quyết định scale/kill.",
        "P=Paid, O=Owned, E=Earned. Đồng bộ UTM + CRM trước khi chạy.",
    ]
    for i, line in enumerate(notes, 3):
        ws_info.cell(row=i, column=1, value=line)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_budget_vnd(text: str, default: float = 350_000_000) -> float:
    """Trích số ngân sách VNĐ từ chuỗi mô tả (vd. 350.000.000 VNĐ)."""
    raw = str(text or "")
    if re.search(r"[\d]{1,3}(?:[.,]\d{3})+", raw):
        digits = re.sub(r"[^\d]", "", raw.split("VND")[0].split("VNĐ")[0].split("đ")[0])
    else:
        digits = re.sub(r"[^\d]", "", raw)
    if not digits:
        return default
    try:
        val = float(digits)
        return val if val > 0 else default
    except ValueError:
        return default


def _rag_zone_formula(pct_cell: str, *, spend_pacing: bool = False) -> str:
    """Công thức vùng RAG từ ô % đạt (0–1+) hoặc pacing chi (spend/budget)."""
    if spend_pacing:
        return (
            f'=IF({pct_cell}="","",IF({pct_cell}<=0.95,"An toàn",'
            f'IF({pct_cell}<=1.1,"Cảnh báo","Nguy hiểm")))'
        )
    return (
        f'=IF({pct_cell}="","",IF({pct_cell}>=0.9,"An toàn",'
        f'IF({pct_cell}>=0.7,"Cảnh báo","Nguy hiểm")))'
    )


def _create_tham_so_sheet(wb, *, budget: float, plan: dict[str, Any]) -> Any:
    """Sheet tham số đầu vào — chỉnh ở đây, các sheet khác tự cập nhật."""
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb["Tham_so"] if "Tham_so" in wb.sheetnames else wb.create_sheet("Tham_so", 0)
    ws.cell(row=1, column=1, value="THAM SỐ ĐẦU VÀO — chỉnh số liệu tại đây, KPI tự cập nhật").font = Font(
        bold=True, size=12, color="1F4E79"
    )
    ws.cell(row=2, column=1, value="Ngân sách · funnel · ROAS/ROMI · kênh đều tham chiếu sheet này").font = Font(
        italic=True, color="666666"
    )
    header_fill = PatternFill("solid", fgColor="1F4E79")
    ws.cell(row=3, column=1, value="Tham số").fill = header_fill
    ws.cell(row=3, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=3, column=2, value="Giá trị").fill = header_fill
    ws.cell(row=3, column=2).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=3, column=3, value="Ghi chú").fill = header_fill
    ws.cell(row=3, column=3).font = Font(bold=True, color="FFFFFF")

    params: list[tuple[str, Any, str, str]] = [
        ("Ngân sách quý (VNĐ)", budget, "#,##0", "Sửa → pacing + phân bổ kênh tự nhảy"),
        ("Tỷ lệ dự phòng", 0.1, "0%", "Dự phòng 10% — ảnh hưởng ngân sách ads"),
        ("Ngân sách ads (VNĐ)", "=B4*(1-B5)", "#,##0", "= Ngân sách quý × (1 − dự phòng)"),
        ("CPL mục tiêu (VNĐ)", 250_000, "#,##0", "So sánh CPL trung bình KPI_tuan"),
        ("MQL mục tiêu / tháng", 180, "#,##0", "Dashboard + scorecard"),
        ("MQL→SQL mục tiêu", 0.25, "0.0%", "Funnel conversion"),
        ("SQL→Win mục tiêu", 0.2, "0.0%", "Win rate funnel"),
        ("ROAS mục tiêu", 3, "0.0", "Doanh thu / chi ads"),
        ("ROMI mục tiêu", 1.5, "0%", "150% = 1.5"),
        ("Doanh thu từ ads (VNĐ)", max(budget * 2.5, 1), "#,##0", "Nhập doanh thu quảng cáo"),
        ("Tổng chi ads (thực tế)", "=SUM(KPI_tuan!$B$2:$B$13)", "#,##0", "Tự tính từ KPI tuần"),
        ("Tổng Leads", "=SUM(KPI_tuan!$C$2:$C$13)", "#,##0", ""),
        ("Tổng MQL", "=SUM(KPI_tuan!$D$2:$D$13)", "#,##0", ""),
        ("Tổng SQL", "=SUM(KPI_tuan!$E$2:$E$13)", "#,##0", ""),
        ("Tổng Win", "=SUM(KPI_tuan!$F$2:$F$13)", "#,##0", ""),
        ("CPL trung bình (TT)", '=IF(B15>0,B14/B15,"")', "#,##0", "= Chi / Leads"),
        ("MQL→SQL (TT)", '=IF(B16>0,B17/B16,"")', "0.0%", ""),
        ("SQL→Win (TT)", '=IF(B17>0,B18/B17,"")', "0.0%", ""),
        ("ROAS (TT)", '=IF(B14>0,B13/B14,"")', "0.0", ""),
        ("ROMI (TT)", '=IF(B14>0,(B13-B14)/B14,"")', "0%", ""),
    ]
    for i, (label, val, fmt, note) in enumerate(params, 4):
        ws.cell(row=i, column=1, value=label)
        cell = ws.cell(row=i, column=2, value=val)
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="right")
        ws.cell(row=i, column=3, value=note)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 42
    return ws


def _parse_metric_number(val: Any) -> float | None:
    """Chuyển '250k', '250.000', 250000 → số."""
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    raw = str(val).strip().lower()
    raw = raw.replace("≤", "").replace("≥", "").replace("đ", "").replace("vnđ", "").replace("vnd", "").strip()
    m = re.search(r"([\d.,]+)\s*k\b", raw)
    if m:
        num = m.group(1).replace(",", "")
        if num.count(".") == 1 and len(num.split(".")[-1]) == 3:
            num = num.replace(".", "")
        return float(num) * 1000
    cleaned = raw.replace(",", "")
    if re.fullmatch(r"[\d.]+", cleaned):
        if cleaned.count(".") == 1 and len(cleaned.split(".")[-1]) == 3:
            cleaned = cleaned.replace(".", "")
        return float(cleaned)
    digits = re.sub(r"[^\d]", "", raw)
    if digits:
        return float(digits)
    return None


def _wire_kpi_workbook_formulas(wb) -> None:
    """Gắn công thức liên sheet — số liệu nhảy khi chỉnh Tham_so / KPI_tuan."""
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    fill_safe = PatternFill("solid", fgColor="C6EFCE")
    fill_warn = PatternFill("solid", fgColor="FFEB9C")
    fill_danger = PatternFill("solid", fgColor="FFC7CE")
    zone_fills = {"An toàn": fill_safe, "Cảnh báo": fill_warn, "Nguy hiểm": fill_danger}

    ws_dash = wb["Tong_quan"]
    ws_w = wb["KPI_tuan"]
    ws_sc = wb["Scorecard_thang"]
    ws_ch = wb["KPI_kenh"]
    ws_risk = wb["Rui_ro"]
    ws_rag = wb["Vung_RAG"]
    ws_st = wb["Chien_luoc"]

    # --- KPI tuần: CPL, MQL→SQL, vùng RAG ---
    w_last = 1 + ws_w.max_row - 1
    if w_last < 2:
        w_last = 13
    for r in range(2, w_last + 1):
        ws_w.cell(row=r, column=7, value=f'=IF(C{r}>0,B{r}/C{r},"")')
        ws_w.cell(row=r, column=7).number_format = "#,##0"
        ws_w.cell(row=r, column=8, value=f'=IF(D{r}>0,E{r}/D{r},"")')
        ws_w.cell(row=r, column=8).number_format = "0.0%"
        ws_w.cell(row=r, column=9, value=(
            f'=IF(G{r}="","",IF(G{r}<=Tham_so!$B$7,"An toàn",'
            f'IF(G{r}<=Tham_so!$B$7*1.2,"Cảnh báo","Nguy hiểm")))'
        ))
        ws_w.cell(row=r, column=10, value=(
            f'=IF(H{r}="","",IF(H{r}>=Tham_so!$B$9,"An toàn",'
            f'IF(H{r}>=Tham_so!$B$9*0.6,"Cảnh báo","Nguy hiểm")))'
        ))
        for c in (9, 10):
            ws_w.cell(row=r, column=c).font = Font(bold=True)

    # --- Dashboard tổng quan (liên kết Tham_so + KPI_tuan) ---
    dash_rows = [
        (5, "=Tham_so!$B$4", "=Tham_so!$B$14", "=IF(B5>0,C5/B5,\"\")", True, "#,##0"),
        (6, "=Tham_so!$B$7", "=Tham_so!$B$19", "=IF(C6>0,B6/C6,\"\")", False, "#,##0"),
        (7, "=Tham_so!$B$8", '=IF(COUNT(KPI_tuan!$D$2:$D$13)>0,SUM(KPI_tuan!$D$2:$D$13)/3,"")', "=IF(B7>0,C7/B7,\"\")", False, "#,##0"),
        (8, "=Tham_so!$B$9", "=Tham_so!$B$20", "=IF(B8>0,C8/B8,\"\")", False, "0.0%"),
        (9, "=Tham_so!$B$10", "=Tham_so!$B$21", "=IF(B9>0,C9/B9,\"\")", False, "0.0%"),
        (10, "=Tham_so!$B$11", "=Tham_so!$B$22", "=IF(B10>0,C10/B10,\"\")", False, "0.0"),
        (11, "=Tham_so!$B$12", "=Tham_so!$B$23", "=IF(B11>0,C11/B11,\"\")", False, "0%"),
        (12, 85, "=Scorecard_thang!$E$13", "=IF(B12>0,C12/B12,\"\")", False, "0"),
    ]
    for row, tgt, act, pct, pacing, fmt in dash_rows:
        ws_dash.cell(row=row, column=2, value=tgt)
        ws_dash.cell(row=row, column=3, value=act)
        ws_dash.cell(row=row, column=3).number_format = fmt
        ws_dash.cell(row=row, column=5, value=pct)
        ws_dash.cell(row=row, column=5).number_format = "0%"
        ws_dash.cell(row=row, column=6, value=_rag_zone_formula(f"E{row}", spend_pacing=pacing))
        ws_dash.cell(row=row, column=6).font = Font(bold=True)

    # Biểu đồ cột — cột M (#) là trục số; cột L là tên chỉ số (không dùng làm trục chart)
    chart_rows = [(22, "CPL", "E6"), (23, "MQL", "E7"), (24, "MQL→SQL", "E8"), (25, "ROAS", "E10"), (26, "ROMI", "E11"), (27, "Scorecard", "E12")]
    for i, (r, lbl, ref) in enumerate(chart_rows):
        ws_dash.cell(row=r, column=12, value=lbl)
        ws_dash.cell(row=r, column=13, value=i + 1)
        ws_dash.cell(row=r, column=14, value=1)
        ws_dash.cell(row=r, column=15, value=f"={ref}")
        ws_dash.cell(row=r, column=15).number_format = "0%"

    # --- Chiến lược: ngân sách liên kết ---
    ws_st.cell(row=7, column=2, value="=Tham_so!$B$4")
    ws_st.cell(row=7, column=2).number_format = "#,##0"
    ws_st.cell(row=8, column=2, value="=Tham_so!$B$4*Tham_so!$B$5")
    ws_st.cell(row=8, column=2).number_format = "#,##0"

    # --- Scorecard: % đạt + vùng + tổng weighted ---
    # B=weight, C/D/E MT/TT month1, E=% F=zone — higher-is-better except row 4 CPL
    score_cfg = [(4, True), (5, False), (6, False), (7, False), (8, False), (9, False), (10, False), (11, False)]
    weight_map = {"20%": 0.2, "15%": 0.15, "10%": 0.1, "5%": 0.05}
    for r, lower_better in score_cfg:
        w = ws_sc.cell(row=r, column=2).value
        if isinstance(w, str) and w in weight_map:
            ws_sc.cell(row=r, column=2, value=weight_map[w])
        elif isinstance(w, str) and "%" in str(w):
            ws_sc.cell(row=r, column=2, value=float(str(w).strip("%")) / 100)
        ws_sc.cell(row=r, column=2).number_format = "0%"
        c_mt, d_tt = ws_sc.cell(row=r, column=3).value, ws_sc.cell(row=r, column=4).value
        if lower_better and r == 4:
            n_mt, n_tt = _parse_metric_number(c_mt), _parse_metric_number(d_tt)
            if n_mt is not None:
                ws_sc.cell(row=r, column=3, value=n_mt)
            if n_tt is not None:
                ws_sc.cell(row=r, column=4, value=n_tt)
            ws_sc.cell(row=r, column=3).number_format = "#,##0"
            ws_sc.cell(row=r, column=4).number_format = "#,##0"
        if lower_better:
            ws_sc.cell(row=r, column=5, value=f'=IF(D{r}>0,C{r}/D{r},"")')
        else:
            ws_sc.cell(row=r, column=5, value=f'=IF(C{r}>0,D{r}/C{r},"")')
        ws_sc.cell(row=r, column=5).number_format = "0%"
        ws_sc.cell(row=r, column=6, value=_rag_zone_formula(f"E{r}"))
        ws_sc.cell(row=r, column=6).font = Font(bold=True)
    ws_sc.cell(row=13, column=5, value="=ROUND(SUMPRODUCT($B$4:$B$11,E4:E11)*100,0)")
    ws_sc.cell(row=13, column=6, value='=IF(E13>=90,"An toàn",IF(E13>=70,"Cảnh báo","Nguy hiểm"))')
    ws_sc.cell(row=13, column=6).font = Font(bold=True)

    # --- KPI kênh: % ngân sách → VNĐ tự tính (cột M) ---
    ws_ch.cell(row=1, column=13, value="Ngân sách (VNĐ)")
    ch_last = ws_ch.max_row
    for r in range(2, ch_last + 1):
        pct_cell = ws_ch.cell(row=r, column=5)
        if pct_cell.value is not None:
            s = str(pct_cell.value).replace("%", "").strip()
            if s in ("—", "-", ""):
                pct_cell.value = 0
            elif s:
                try:
                    v = float(s)
                    pct_cell.value = v / 100 if v > 1 else v
                except ValueError:
                    pct_cell.value = 0
        pct_cell.number_format = "0%"
        ws_ch.cell(row=r, column=13, value=f'=IF(E{r}>0,Tham_so!$B$6*E{r},"")')
        ws_ch.cell(row=r, column=13).number_format = "#,##0"

    # --- Rủi ro: điểm = P×I, vùng tự tính ---
    for r in range(4, ws_risk.max_row + 1):
        ws_risk.cell(row=r, column=6, value=f"=D{r}*E{r}")
        ws_risk.cell(row=r, column=7, value=(
            f'=IF(F{r}>=15,"Nguy hiểm",IF(F{r}>=10,"Cảnh báo","An toàn"))'
        ))
        ws_risk.cell(row=r, column=7).font = Font(bold=True)

    rag_zone_ref = {4: ("=Tong_quan!E6", "F6"), 5: ("=Tong_quan!E7", "F7"), 6: ("=Tong_quan!E8", "F8"), 7: ("=Tong_quan!E10", "F10"), 8: ("=Tong_quan!E5", "F5")}
    for r, (pct_ref, zone_ref) in rag_zone_ref.items():
        ws_rag.cell(row=r, column=6, value=pct_ref)
        ws_rag.cell(row=r, column=6).number_format = "0%"
        ws_rag.cell(row=r, column=7, value=f"=Tong_quan!{zone_ref}")
        ws_rag.cell(row=r, column=7).font = Font(bold=True)


def _add_kpi_workbook_charts(
    wb,
    *,
    n_weekly_rows: int,
    n_chart_kpis: int,
) -> None:
    """Thêm biểu đồ sau khi gắn công thức — chỉ dùng trục số để Excel không báo lỗi repair."""
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList

    ws_dash = wb["Tong_quan"]
    ws_w = wb["KPI_tuan"]

    bar = BarChart()
    bar.type = "col"
    bar.title = "KPI muc tieu vs thuc te (% dat)"
    bar.y_axis.title = "% dat"
    bar.x_axis.title = "Chi so (#)"
    bar.style = 10
    bar.width = 18
    bar.height = 10
    bar_data = Reference(ws_dash, min_col=14, min_row=21, max_col=15, max_row=21 + n_chart_kpis)
    bar_cats = Reference(ws_dash, min_col=13, min_row=22, max_row=21 + n_chart_kpis)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(bar_cats)
    bar.dataLabels = DataLabelList()
    bar.dataLabels.showVal = True
    ws_dash.add_chart(bar, "A17")
    bar.x_axis.axPos = "b"
    bar.y_axis.axPos = "l"

    if n_weekly_rows > 0:
        weeks = Reference(ws_w, min_col=11, min_row=2, max_row=1 + n_weekly_rows)

        line = LineChart()
        line.title = "Xu huong funnel theo tuan (MQL / SQL / Win)"
        line.y_axis.title = "So luong"
        line.x_axis.title = "Tuan (#)"
        line.style = 10
        line.width = 20
        line.height = 10
        for col in (4, 5, 6):
            line.add_data(
                Reference(ws_w, min_col=col, min_row=1, max_row=1 + n_weekly_rows),
                titles_from_data=True,
            )
        line.set_categories(weeks)
        ws_w.add_chart(line, "M2")
        line.x_axis.axPos = "b"
        line.y_axis.axPos = "l"

        line_cpl = LineChart()
        line_cpl.title = "Xu huong CPL theo tuan"
        line_cpl.y_axis.title = "CPL (VND)"
        line_cpl.style = 10
        line_cpl.width = 18
        line_cpl.height = 10
        line_cpl.add_data(
            Reference(ws_w, min_col=7, min_row=1, max_row=1 + n_weekly_rows),
            titles_from_data=True,
        )
        line_cpl.set_categories(weeks)
        ws_dash.add_chart(line_cpl, "A32")
        line_cpl.x_axis.axPos = "b"
        line_cpl.y_axis.axPos = "l"


def build_kpi_strategy_xlsx(
    *,
    brand: str = "PTT Advertising Solutions",
    campaign_plan: dict[str, Any] | None = None,
) -> bytes:
    """Excel quản lý KPI chiến lược marketing — dashboard, biểu đồ, vùng RAG, rủi ro."""
    plan = campaign_plan or {}
    project_label = str(plan.get("project_name") or "").strip()
    campaign_label = str(plan.get("campaign_name") or project_label or "").strip()
    smart_goal = str(
        plan.get("smart_goal") or "Tăng 40% MQL chất lượng, SQL rate ≥25%, CPL ≤220k"
    ).strip()
    icp_val = str(plan.get("icp") or "B2B dịch vụ/FMCG, 20–200 NV, HCM/HN").strip()
    duration_val = str(plan.get("duration") or "Q2/2026 (12 tuần)").strip()
    budget_val = str(plan.get("budget") or "350.000.000 VNĐ/quý").strip()
    dash_title = campaign_label or brand
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_tham = wb.active
    ws_tham.title = "Tham_so"
    thin = Side(style="thin", color="B4B4B4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    sub_fill = PatternFill("solid", fgColor="398B43")
    sub_font = Font(bold=True, color="FFFFFF")
    fill_safe = PatternFill("solid", fgColor="C6EFCE")
    fill_warn = PatternFill("solid", fgColor="FFEB9C")
    fill_danger = PatternFill("solid", fgColor="FFC7CE")

    def write_header(ws, headers: list[str], row: int = 1, *, fill=None, font=None) -> None:
        hfill = fill or header_fill
        hfont = font or header_font
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    def style_data_block(ws, min_row: int, max_row: int, max_col: int) -> None:
        for r in range(min_row, max_row + 1):
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).border = border
                ws.cell(row=r, column=c).alignment = Alignment(vertical="center", wrap_text=True)

    # --- Sheet: Tổng quan (dashboard + biểu đồ) ---
    ws_dash = wb.create_sheet("Tong_quan")
    ws_dash.cell(row=1, column=1, value=f"DASHBOARD KPI CHIẾN LƯỢC MARKETING — {dash_title}").font = Font(
        bold=True, size=14, color="1F4E79"
    )
    ws_dash.cell(row=2, column=1, value="Chỉnh số tại sheet Tham_so & KPI_tuan — dashboard/biểu đồ tự cập nhật").font = Font(
        italic=True, color="666666"
    )

    dash_headers = [
        "Chỉ số tổng hợp",
        "Mục tiêu",
        "Thực tế",
        "Đơn vị",
        "% đạt",
        "Vùng quản lý",
        "Xu hướng",
        "Owner",
        "Hành động tiếp theo",
    ]
    write_header(ws_dash, dash_headers, row=4, fill=sub_fill, font=sub_font)
    dash_rows = [
        ("Ngân sách đã chi (Q)", "350.000.000", "287.500.000", "VNĐ", "82%", "An toàn", "↗", "MKT Lead", "Pacing ổn — giữ nhịp"),
        ("CPL trung bình", "≤250.000", "235.000", "VNĐ", "106%", "An toàn", "↘ tốt", "Ads Lead", "Scale variant top"),
        ("MQL / tháng", "≥180", "156", "lead", "87%", "Cảnh báo", "→", "MKT", "Tăng budget Search 10%"),
        ("MQL → SQL", "≥25%", "19%", "%", "76%", "Nguy hiểm", "↘", "Sales Lead", "Review scoring + SLA 4h"),
        ("SQL → Win", "≥20%", "18%", "%", "90%", "An toàn", "↗", "Sales Lead", "Giữ script demo"),
        ("ROAS", "≥3.0", "2.6", "x", "87%", "Cảnh báo", "→", "MKT Lead", "Audit landing + offer"),
        ("ROMI", "≥150%", "128%", "%", "85%", "Cảnh báo", "↘", "Finance", "Tối ưu kênh thấp ROAS"),
        ("Scorecard tổng", "≥85", "78", "điểm", "92%", "Cảnh báo", "→", "MKT Lead", "Họp review tháng"),
        ("Rủi ro đang mở", "≤3", "5", "mục", "—", "Nguy hiểm", "↑", "MKT Lead", "Xem sheet Rui_ro"),
    ]
    zone_fills = {"An toàn": fill_safe, "Cảnh báo": fill_warn, "Nguy hiểm": fill_danger}
    for r, row in enumerate(dash_rows, 5):
        for c, val in enumerate(row, 1):
            cell = ws_dash.cell(row=r, column=c, value=val)
            if c == 6:
                cell.fill = zone_fills.get(str(val), fill_warn)
                cell.font = Font(bold=True)
    style_data_block(ws_dash, 5, 5 + len(dash_rows) - 1, len(dash_headers))
    for i, w in enumerate([22, 14, 14, 10, 10, 14, 10, 12, 28], 1):
        ws_dash.column_dimensions[get_column_letter(i)].width = w

    # Legend vùng RAG
    ws_dash.cell(row=15, column=1, value="Chú giải vùng quản lý:").font = Font(bold=True)
    legend = [
        ("An toàn (Xanh)", "Đạt ≥90% mục tiêu hoặc trong ngưỡng xanh — duy trì"),
        ("Cảnh báo (Vàng)", "70–89% hoặc chạm ngưỡng vàng — theo dõi sát, hành động trong 48h"),
        ("Nguy hiểm (Đỏ)", "<70% hoặc vượt ngưỡng đỏ — dừng scale, họp khẩn trong 24h"),
    ]
    for i, (label, desc) in enumerate(legend, 16):
        ws_dash.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws_dash.cell(row=i, column=1).fill = zone_fills.get(label.split(" ")[0], fill_warn)
        ws_dash.cell(row=i, column=2, value=desc)
        ws_dash.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)

    # Dữ liệu biểu đồ cột (L=tên, M=# trục số, N/O=MT vs TT)
    ws_dash.cell(row=20, column=12, value="Chart_KPI").font = Font(bold=True)
    chart_kpi_labels = ["CPL", "MQL", "MQL→SQL", "ROAS", "ROMI", "Scorecard"]
    chart_kpi_target = [100, 100, 100, 100, 100, 100]
    chart_kpi_actual = [106, 87, 76, 87, 85, 92]
    ws_dash.cell(row=21, column=12, value="Chỉ số")
    ws_dash.cell(row=21, column=13, value="#")
    ws_dash.cell(row=21, column=14, value="Mục tiêu (100%)")
    ws_dash.cell(row=21, column=15, value="Thực tế (% đạt)")
    for i, lbl in enumerate(chart_kpi_labels):
        ws_dash.cell(row=22 + i, column=12, value=lbl)
        ws_dash.cell(row=22 + i, column=13, value=i + 1)
        ws_dash.cell(row=22 + i, column=14, value=chart_kpi_target[i])
        ws_dash.cell(row=22 + i, column=15, value=chart_kpi_actual[i])

    # --- Sheet: Chiến lược (mở rộng) ---
    ws = wb.create_sheet("Chien_luoc")
    ws.cell(row=1, column=1, value=f"QUẢN LÝ KPI CHIẾN LƯỢC MARKETING — {dash_title}").font = Font(bold=True, size=13)
    strategy_fields = [
        ("Tên chiến dịch / quý", campaign_label or "Lead gen Q2 2026 — Đa kênh B2B"),
        ("Mục tiêu SMART", smart_goal),
        ("Phân khúc ICP", icp_val),
        ("Thời gian", duration_val),
        ("Ngân sách marketing", budget_val),
        ("Ngân sách dự phòng (10%)", "35.000.000 VNĐ"),
        ("Marketing Lead", ""),
        ("Sales Lead", ""),
        ("Finance / ROMI owner", ""),
        ("Cadence review", "Daily ads · Weekly KPI (T6) · Monthly scorecard · Quarterly ROMI"),
        ("Công cụ theo dõi", "CRM · Ads Manager · GA4 · Sheet KPI này"),
        ("Ngưỡng dừng chi (kill switch)", "CPL >350k 5 ngày liên tiếp HOẶC tracking lệch >15%"),
    ]
    for i, (label, val) in enumerate(strategy_fields, 3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=val)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 52

    # RACI
    ws.cell(row=16, column=1, value="RACI — quản lý KPI").font = Font(bold=True, size=12)
    raci_headers = ["Hoạt động", "R", "A", "C", "I", "Tần suất"]
    write_header(ws, raci_headers, row=17, fill=sub_fill, font=sub_font)
    raci_rows = [
        ("Theo dõi spend/CPL daily", "Ads", "MKT Lead", "Finance", "Sales", "Daily"),
        ("Họp KPI tuần (ritual T6)", "MKT", "MKT Lead", "Sales, Ads", "Leadership", "Weekly"),
        ("Scorecard tháng + ROMI", "MKT", "MKT Lead", "Finance", "CEO", "Monthly"),
        ("Review rủi ro & vùng RAG", "MKT Lead", "CEO", "Sales, Dev", "Team", "Monthly"),
        ("Quyết định scale/pause kênh", "Ads", "MKT Lead", "Sales", "Finance", "Ad-hoc"),
    ]
    for r, row in enumerate(raci_rows, 18):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    style_data_block(ws, 18, 18 + len(raci_rows) - 1, len(raci_headers))

    # --- Sheet 3: KPI funnel (mở rộng vùng) ---
    ws_f = wb.create_sheet("KPI_funnel")
    funnel_headers = [
        "Giai đoạn",
        "Loại KPI",
        "Chỉ số",
        "Công thức",
        "Mục tiêu",
        "Thực tế",
        "Đơn vị",
        "Vùng XANH (An toàn)",
        "Vùng VÀNG (Cảnh báo)",
        "Vùng ĐỎ (Nguy hiểm)",
        "Vùng hiện tại",
        "Tần suất đo",
        "Nguồn dữ liệu",
        "Owner",
        "Ghi chú",
    ]
    write_header(ws_f, funnel_headers, fill=sub_fill, font=sub_font)
    funnel_rows = [
        ("KHTN", "Leading", "CTR", "Clicks/Impressions", "≥3%", "2.8%", "%", "≥3%", "2–2.9%", "<2%", "Cảnh báo", "Daily", "Ads Manager", "Ads", "Test headline mới"),
        ("KHTN", "Leading", "CPC", "Spend/Clicks", "≤8.000", "7.200", "VNĐ", "≤8k", "8k–12k", ">12k", "An toàn", "Daily", "Ads Manager", "Ads", ""),
        ("KHTN", "Lagging", "CPL", "Spend/Leads", "≤250k", "235k", "VNĐ", "≤250k", "250k–300k", ">300k", "An toàn", "Daily", "Ads+CRM", "Ads Lead", ""),
        ("KHTN", "Lagging", "MQL rate", "MQL/Leads", "≥50%", "48%", "%", "≥50%", "35–49%", "<35%", "Cảnh báo", "Weekly", "CRM", "MKT", ""),
        ("KHQT", "Leading", "Contact rate", "Connected/MQL", "≥60%", "55%", "%", "≥60%", "45–59%", "<45%", "Cảnh báo", "Weekly", "CRM", "Sales", "SLA ≤4h"),
        ("KHQT", "Leading", "Demo book rate", "Demo booked/Connected", "≥30%", "28%", "%", "≥30%", "20–29%", "<20%", "Cảnh báo", "Weekly", "CRM", "Sales", ""),
        ("KHQT", "Lagging", "MQL→SQL", "SQL/MQL", "≥25%", "19%", "%", "≥25%", "15–24%", "<15%", "Nguy hiểm", "Weekly", "CRM", "Sales Lead", "Họp khẩn"),
        ("KHQT", "Lagging", "SQL→Win", "Win/SQL", "≥20%", "18%", "%", "≥20%", "10–19%", "<10%", "An toàn", "Monthly", "CRM", "Sales Lead", ""),
        ("CSKH", "Lagging", "NPS", "Promoter−Detractor", "≥40", "42", "điểm", "≥40", "30–39", "<30", "An toàn", "Quarterly", "Survey", "CS", ""),
        ("CSKH", "Lagging", "Retention", "Active Q/Q", "≥85%", "82%", "%", "≥85%", "75–84%", "<75%", "Cảnh báo", "Quarterly", "CRM", "CS", ""),
        ("Tài chính", "Lagging", "ROAS", "Revenue/Ad spend", "≥3", "2.6", "x", "≥3", "2–2.9", "<2", "Cảnh báo", "Weekly", "Finance+Ads", "MKT Lead", ""),
        ("Tài chính", "Lagging", "ROMI", "(Rev−Cost)/Cost", "≥150%", "128%", "%", "≥150%", "100–149%", "<100%", "Cảnh báo", "Monthly", "Finance", "MKT Lead", ""),
        ("Tài chính", "Lagging", "LTV/CAC", "LTV/CAC", "≥3", "2.8", "x", "≥3", "2–2.9", "<2", "Cảnh báo", "Quarterly", "Finance", "MKT Lead", ""),
        ("Vận hành", "Leading", "Tracking accuracy", "CRM leads/Ads leads", "≥95%", "91%", "%", "≥95%", "90–94%", "<90%", "Cảnh báo", "Weekly", "Dev+CRM", "Dev", "QA pixel"),
        ("Vận hành", "Leading", "SLA phản hồi lead", "Time to first contact", "≤4h", "5.2h", "giờ", "≤4h", "4–8h", ">8h", "Cảnh báo", "Daily", "CRM", "Sales Lead", ""),
    ]
    for r, row in enumerate(funnel_rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws_f.cell(row=r, column=c, value=val)
            if c == 11:
                cell.fill = zone_fills.get(str(val), fill_warn)
                cell.font = Font(bold=True)
    style_data_block(ws_f, 2, 1 + len(funnel_rows), len(funnel_headers))
    for i, w in enumerate([10, 10, 14, 20, 10, 10, 8, 14, 14, 14, 12, 10, 14, 12, 18], 1):
        ws_f.column_dimensions[get_column_letter(i)].width = w

    # --- Sheet 4: KPI kênh ---
    ws_ch = wb.create_sheet("KPI_kenh")
    ch_headers = [
        "Kênh",
        "Mục tiêu chiến lược",
        "KPI chính",
        "Mục tiêu",
        "% ngân sách",
        "T1 thực tế",
        "T2 thực tế",
        "T3 thực tế",
        "Vùng hiện tại",
        "Trend",
        "Owner",
        "Hành động",
    ]
    write_header(ws_ch, ch_headers, fill=sub_fill, font=sub_font)
    channel_rows = [
        ("Google Search", "Conversion", "CPL", "≤250k", "35%", "228k", "", "", "An toàn", "↗", "Ads", "Scale brand KW"),
        ("Meta Ads", "Consideration", "CPL", "≤180k", "25%", "195k", "", "", "Cảnh báo", "→", "Social", "Refresh creative"),
        ("LinkedIn", "B2B Lead", "CPL", "≤350k", "15%", "310k", "", "", "An toàn", "↗", "Ads", "Giữ"),
        ("TikTok", "Awareness", "CPV", "≤800", "8%", "720", "", "", "An toàn", "↗", "Social", "Test hook mới"),
        ("Email/CRM", "Nurture", "Open rate", "≥25%", "5%", "22%", "", "", "Cảnh báo", "↘", "CRM", "A/B subject"),
        ("SEO/Content", "Organic", "Organic leads", "≥20/th", "7%", "18", "", "", "Cảnh báo", "→", "Content", "2 bài/tháng"),
        ("Landing/CRO", "Conversion", "CVR", "≥4%", "—", "3.6%", "", "", "Cảnh báo", "→", "CRO", "Test form ngắn"),
        ("Telesales", "SQL", "Contact rate", "≥60%", "—", "55%", "", "", "Cảnh báo", "↘", "Sales", "Script SPIN"),
    ]
    plan_channels = plan.get("channels") if isinstance(plan.get("channels"), list) else []
    if plan_channels:
        channel_rows = []
        for ch in plan_channels[:10]:
            if not isinstance(ch, dict):
                continue
            pct = ch.get("budget_pct", "")
            pct_str = f"{pct}%" if pct != "" and "%" not in str(pct) else str(pct)
            channel_rows.append(
                (
                    ch.get("name", ""),
                    ch.get("goal", ""),
                    "KPI",
                    ch.get("kpi", ""),
                    pct_str,
                    "",
                    "",
                    "",
                    "Planned",
                    "→",
                    "MKT",
                    "",
                )
            )
    for r, row in enumerate(channel_rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws_ch.cell(row=r, column=c, value=val)
            if c == 9:
                cell.fill = zone_fills.get(str(val), fill_warn)
    style_data_block(ws_ch, 2, 1 + len(channel_rows), len(ch_headers))
    for i, w in enumerate([14, 16, 12, 12, 10, 12, 12, 12, 12, 8, 12, 22], 1):
        ws_ch.column_dimensions[get_column_letter(i)].width = w

    # --- Sheet 5: Vùng RAG (an toàn / cảnh báo / nguy hiểm) ---
    ws_rag = wb.create_sheet("Vung_RAG")
    ws_rag.cell(row=1, column=1, value="BẢNG NGƯỠNG VÙNG QUẢN LÝ — AN TOÀN · CẢNH BÁO · NGUY HIỂM").font = Font(
        bold=True, size=12, color="1F4E79"
    )
    rag_headers = [
        "Chỉ số",
        "Loại (cao/t thấp tốt)",
        "Vùng XANH — An toàn",
        "Vùng VÀNG — Cảnh báo",
        "Vùng ĐỎ — Nguy hiểm",
        "Giá trị hiện tại",
        "Vùng hiện tại",
        "Hành động bắt buộc",
        "Owner",
    ]
    write_header(ws_rag, rag_headers, row=3, fill=sub_fill, font=sub_font)
    rag_rows = [
        ("CPL", "Thấp tốt", "≤100% target", "100–120% target", ">120% target", "94%", "An toàn", "Duy trì / scale nhẹ", "Ads Lead"),
        ("MQL volume", "Cao tốt", "≥90% target", "70–89% target", "<70% target", "87%", "Cảnh báo", "Tăng reach hoặc budget test", "MKT Lead"),
        ("MQL→SQL", "Cao tốt", "≥90% target", "60–89% target", "<60% target", "76%", "Nguy hiểm", "Họp Sales+MKT trong 24h", "Sales Lead"),
        ("ROAS", "Cao tốt", "≥ target", "80–99% target", "<80% target", "87%", "Cảnh báo", "Audit funnel + offer", "MKT Lead"),
        ("Spend pacing", "Trong kế hoạch", "≤95% budget MTD", "95–110%", ">110%", "82%", "An toàn", "Giữ nhịp chi", "Finance"),
        ("Tracking accuracy", "Cao tốt", "≥95%", "90–94%", "<90%", "91%", "Cảnh báo", "QA pixel/CAPI ngay", "Dev"),
        ("Creative fatigue", "Thấp tốt", "Freq ≤2.5", "2.5–3.5", ">3.5", "2.8", "Cảnh báo", "Refresh creative 2 tuần", "Social"),
        ("Brand safety", "Thấp tốt", "0 vi phạm", "1 cảnh báo", "≥2 hoặc ban", "0", "An toàn", "Duy trì pre-approval", "MKT Lead"),
        ("Phụ thuộc 1 kênh", "Thấp tốt", "<50% leads", "50–70%", ">70%", "58%", "Cảnh báo", "Đa dạng hóa media mix", "MKT Lead"),
        ("NPS", "Cao tốt", "≥ target", "target−10 đến target", "<target−10", "42", "An toàn", "Duy trì CS playbook", "CS"),
    ]
    for r, row in enumerate(rag_rows, 4):
        for c, val in enumerate(row, 1):
            cell = ws_rag.cell(row=r, column=c, value=val)
            if c == 3:
                cell.fill = fill_safe
            elif c == 4:
                cell.fill = fill_warn
            elif c == 5:
                cell.fill = fill_danger
            elif c == 7:
                cell.fill = zone_fills.get(str(val), fill_warn)
                cell.font = Font(bold=True)
    style_data_block(ws_rag, 4, 3 + len(rag_rows), len(rag_headers))
    ws_rag.cell(row=16, column=1, value="Quy tắc: Chỉ số ở vùng ĐỎ → ghi vào Canh_bao + Rui_ro, không scale budget cho đến khi về VÀNG.").font = Font(
        italic=True, color="C00000"
    )
    for i, w in enumerate([18, 14, 18, 18, 18, 12, 12, 28, 12], 1):
        ws_rag.column_dimensions[get_column_letter(i)].width = w

    # --- Sheet 6: Scorecard tháng ---
    ws_sc = wb.create_sheet("Scorecard_thang")
    ws_sc.cell(row=1, column=1, value="Marketing Scorecard — weighted + vùng RAG").font = Font(bold=True, size=12)
    sc_headers = [
        "Chỉ số",
        "Trọng số",
        "T1 MT",
        "T1 TT",
        "T1 % đạt",
        "T1 Vùng",
        "T2 MT",
        "T2 TT",
        "T2 %",
        "T2 Vùng",
        "T3 MT",
        "T3 TT",
        "T3 %",
        "T3 Vùng",
    ]
    write_header(ws_sc, sc_headers, row=3, fill=sub_fill, font=sub_font)
    score_rows = [
        ("CPL trung bình", "20%", "250k", "235k", "106%", "An toàn", "", "", "", "", "", "", "", ""),
        ("MQL", "15%", "180", "156", "87%", "Cảnh báo", "", "", "", "", "", "", "", ""),
        ("SQL", "15%", "45", "30", "67%", "Nguy hiểm", "", "", "", "", "", "", "", ""),
        ("Win rate", "15%", "20%", "18%", "90%", "An toàn", "", "", "", "", "", "", "", ""),
        ("ROAS", "15%", "3.0", "2.6", "87%", "Cảnh báo", "", "", "", "", "", "", "", ""),
        ("ROMI", "10%", "150%", "128%", "85%", "Cảnh báo", "", "", "", "", "", "", "", ""),
        ("Brand search lift", "5%", "15%", "12%", "80%", "Cảnh báo", "", "", "", "", "", "", "", ""),
        ("NPS", "5%", "40", "42", "105%", "An toàn", "", "", "", "", "", "", "", ""),
    ]
    zone_cols = {6, 10, 14}
    for r, row in enumerate(score_rows, 4):
        for c, val in enumerate(row, 1):
            cell = ws_sc.cell(row=r, column=c, value=val)
            if c in zone_cols and val in zone_fills:
                cell.fill = zone_fills[val]
                cell.font = Font(bold=True)
    ws_sc.cell(row=13, column=1, value="Tổng điểm scorecard (weighted)").font = Font(bold=True)
    for i, w in enumerate([18, 10, 10, 10, 10, 12, 10, 10, 8, 12, 10, 10, 8, 12], 1):
        ws_sc.column_dimensions[get_column_letter(i)].width = w

    # --- Sheet 7: KPI tuần (có số mẫu + biểu đồ line) ---
    ws_w = wb.create_sheet("KPI_tuan")
    w_headers = [
        "Tuần",
        "Spend (VNĐ)",
        "Leads",
        "MQL",
        "SQL",
        "Win",
        "CPL",
        "MQL→SQL %",
        "Vùng CPL",
        "Vùng funnel",
        "Ghi chú",
    ]
    write_header(ws_w, w_headers, fill=sub_fill, font=sub_font)
    weekly_data = [
        ("W1", 25_000_000, 120, 58, 14, 3, 208_333, "24%", "An toàn", "An toàn", "Launch Q2"),
        ("W2", 28_000_000, 115, 52, 12, 2, 243_478, "23%", "Cảnh báo", "Cảnh báo", "CPL tăng — audit LP"),
        ("W3", 26_500_000, 118, 55, 11, 2, 224_576, "20%", "An toàn", "Cảnh báo", "SQL giảm"),
        ("W4", 27_000_000, 122, 59, 13, 3, 221_311, "22%", "An toàn", "An toàn", ""),
        ("W5", 29_000_000, 110, 48, 9, 1, 263_636, "19%", "Cảnh báo", "Nguy hiểm", "Họp khẩn Sales"),
        ("W6", 28_500_000, 125, 62, 15, 4, 228_000, "24%", "An toàn", "An toàn", "Recovery"),
        ("W7", 30_000_000, 128, 64, 14, 3, 234_375, "22%", "An toàn", "Cảnh báo", ""),
        ("W8", 31_000_000, 130, 65, 13, 3, 238_462, "20%", "An toàn", "Cảnh báo", ""),
        ("W9", 32_000_000, 132, 68, 16, 4, 242_424, "24%", "Cảnh báo", "An toàn", ""),
        ("W10", 33_000_000, 135, 70, 15, 3, 244_444, "21%", "Cảnh báo", "Cảnh báo", ""),
        ("W11", 34_000_000, 138, 72, 17, 4, 246_377, "24%", "Cảnh báo", "An toàn", ""),
        ("W12", 35_000_000, 140, 74, 18, 5, 250_000, "24%", "An toàn", "An toàn", "Review quý"),
    ]
    for r, row in enumerate(weekly_data, 2):
        for c, val in enumerate(row, 1):
            if c in (7, 8, 9, 10):
                continue
            ws_w.cell(row=r, column=c, value=val)
        ws_w.cell(row=r, column=11, value=r - 1)
    style_data_block(ws_w, 2, 1 + len(weekly_data), len(w_headers))
    for i, w in enumerate([8, 14, 10, 10, 10, 8, 12, 12, 12, 12, 24], 1):
        ws_w.column_dimensions[get_column_letter(i)].width = w

    # --- Sheet: Rủi ro ---
    ws_risk = wb.create_sheet("Rui_ro")
    ws_risk.cell(row=1, column=1, value="SỔ ĐĂNG RỦI RO MARKETING — Ma trận xác suất × tác động").font = Font(
        bold=True, size=12, color="1F4E79"
    )
    risk_headers = [
        "ID",
        "Rủi ro",
        "Nhóm",
        "Xác suất (1–5)",
        "Tác động (1–5)",
        "Điểm rủi ro",
        "Vùng",
        "Biện pháp phòng ngừa",
        "Kế hoạch ứng phó",
        "Owner",
        "Trạng thái",
        "Ngày review",
    ]
    write_header(ws_risk, risk_headers, row=3, fill=sub_fill, font=sub_font)
    risk_rows = [
        ("R01", "Overspend ngân sách (>110% MTD)", "Tài chính", 3, 4, 12, "Nguy hiểm", "Daily cap, pacing rules", "Giảm bid 20%, pause ad set", "MKT Lead", "Mở", ""),
        ("R02", "Tracking/GA4/pixel lỗi", "Kỹ thuật", 3, 5, 15, "Nguy hiểm", "QA checklist weekly", "Rollback tag, CAPI backup", "Dev", "Mở", ""),
        ("R03", "MQL→SQL sụt <15%", "Funnel", 4, 5, 20, "Nguy hiểm", "SLA 4h, scoring rõ", "Họp Sales+MKT 24h", "Sales Lead", "Mở", ""),
        ("R04", "Creative fatigue", "Ads", 4, 3, 12, "Nguy hiểm", "Refresh 2–4 tuần", "Brief creative mới", "Social", "Theo dõi", ""),
        ("R05", "Phụ thuộc 1 kênh >70% lead", "Chiến lược", 3, 4, 12, "Nguy hiểm", "Media mix đa kênh", "Test 2 kênh mới", "MKT Lead", "Theo dõi", ""),
        ("R06", "Vi phạm policy ads", "Pháp lý", 2, 5, 10, "Cảnh báo", "Pre-approval copy", "Pause campaign, appeal", "MKT Lead", "Đóng", ""),
        ("R07", "Đối thủ tăng bid (CPL tăng)", "Thị trường", 3, 3, 9, "Cảnh báo", "Monitor auction insights", "Long-tail KW, LP CRO", "Ads Lead", "Theo dõi", ""),
        ("R08", "Thiếu nhân sự telesales", "Vận hành", 3, 4, 12, "Nguy hiểm", "Backup SDR, auto-assign", "Outsource SDR tạm", "Sales Lead", "Mở", ""),
        ("R09", "Brand reputation (review xấu)", "Thương hiệu", 2, 4, 8, "Cảnh báo", "Social listening", "PR response playbook", "MKT Lead", "Theo dõi", ""),
        ("R10", "Mất dữ liệu CRM", "Kỹ thuật", 1, 5, 5, "An toàn", "Backup daily", "Restore từ backup", "Dev", "Đóng", ""),
    ]
    for r, row in enumerate(risk_rows, 4):
        for c, val in enumerate(row, 1):
            cell = ws_risk.cell(row=r, column=c, value=val)
            if c == 7:
                cell.fill = zone_fills.get(str(val), fill_warn)
                cell.font = Font(bold=True)
            if c == 6 and isinstance(val, int) and val >= 15:
                cell.fill = fill_danger
            elif c == 6 and isinstance(val, int) and val >= 10:
                cell.fill = fill_warn
            elif c == 6 and isinstance(val, int):
                cell.fill = fill_safe
    style_data_block(ws_risk, 4, 3 + len(risk_rows), len(risk_headers))
    ws_risk.cell(row=16, column=1, value="Ma trận: Điểm = Xác suất × Tác động · ≥15 Đỏ · 10–14 Vàng · <10 Xanh").font = Font(italic=True)
    for i, w in enumerate([6, 28, 12, 10, 12, 10, 12, 28, 28, 12, 10, 12], 1):
        ws_risk.column_dimensions[get_column_letter(i)].width = w

    # --- Sheet 9: Cảnh báo vận hành ---
    ws_al = wb.create_sheet("Canh_bao")
    al_headers = [
        "Mã CB",
        "Chỉ số",
        "Điều kiện kích hoạt",
        "Vùng",
        "Mức ưu tiên",
        "Hành động ngay",
        "Thời hạn",
        "Owner",
        "Trạng thái",
        "Liên kết rủi ro",
    ]
    write_header(ws_al, al_headers, fill=sub_fill, font=sub_font)
    alert_rows = [
        ("CB-01", "CPL", "Vượt target >20% trong 7 ngày liên tiếp", "Nguy hiểm", "P1", "Pause variant + audit landing", "24h", "Ads Lead", "Active", "R01"),
        ("CB-02", "Spend pacing", ">110% budget giữa tuần", "Nguy hiểm", "P1", "Giảm daily cap 15%", "24h", "MKT Lead", "Active", "R01"),
        ("CB-03", "MQL→SQL", "<15% trong 2 tuần liên tiếp", "Nguy hiểm", "P1", "Review scoring + sales script", "24h", "Sales Lead", "Active", "R03"),
        ("CB-04", "Tracking", "Lead ads ≠ CRM >10%", "Nguy hiểm", "P1", "QA pixel/CAPI ngay", "12h", "Dev", "Active", "R02"),
        ("CB-05", "Creative fatigue", "Frequency >3 + CTR giảm >20%", "Cảnh báo", "P2", "Refresh creative + test hook", "48h", "Social", "Watch", "R04"),
        ("CB-06", "ROAS", "<80% target 2 tuần", "Cảnh báo", "P2", "Audit offer + LP CRO", "72h", "MKT Lead", "Watch", "R07"),
        ("CB-07", "SLA lead", "Phản hồi >8h", "Cảnh báo", "P2", "Escalate SDR lead", "24h", "Sales Lead", "Watch", "R08"),
        ("CB-08", "NPS", "Giảm >10 điểm so với quý trước", "Cảnh báo", "P3", "Khảo sát follow-up", "1 tuần", "CS", "Open", "R09"),
        ("CB-09", "Brand safety", "≥1 cảnh báo policy", "Nguy hiểm", "P1", "Pause + review copy", "12h", "MKT Lead", "Open", "R06"),
        ("CB-10", "Kênh đơn lẻ", ">70% lead từ 1 kênh", "Cảnh báo", "P2", "Kích hoạt test kênh mới", "1 tuần", "MKT Lead", "Watch", "R05"),
    ]
    for r, row in enumerate(alert_rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws_al.cell(row=r, column=c, value=val)
            if c == 4:
                cell.fill = zone_fills.get(str(val), fill_warn)
                cell.font = Font(bold=True)
            if c == 5 and val == "P1":
                cell.fill = fill_danger
                cell.font = Font(bold=True, color="FFFFFF")
    style_data_block(ws_al, 2, 1 + len(alert_rows), len(al_headers))
    for i, w in enumerate([8, 14, 32, 12, 10, 28, 10, 12, 10, 10], 1):
        ws_al.column_dimensions[get_column_letter(i)].width = w

    # --- Sheet 10: Quản trị ---
    ws_gov = wb.create_sheet("Quan_tri")
    ws_gov.cell(row=1, column=1, value="KHUNG QUẢN TRỊ KPI & RỦI RO").font = Font(bold=True, size=12)
    gov_sections = [
        ("Lịch họp cố định", [
            ("Daily standup Ads", "15 phút", "Spend, CPL, tracking", "Ads Lead", "Slack/Meet"),
            ("Weekly KPI (T6 14h)", "60 phút", "KPI_tuan + Canh_bao mở", "MKT Lead", "Phòng họp"),
            ("Monthly scorecard", "90 phút", "Scorecard_thang + ROMI + Rui_ro", "MKT Lead + CEO", "Phòng họp"),
            ("Quarterly strategy", "Half-day", "Review chiến lược + media mix", "Leadership", "Offsite"),
        ]),
        ("Escalation — leo thang", [
            ("Vùng VÀNG", "Owner xử lý trong 48h, báo cáo MKT Lead", "", "", ""),
            ("Vùng ĐỎ", "Họp khẩn 24h: MKT Lead + Sales + Finance", "", "", ""),
            ("P1 cảnh báo", "Pause scale, root cause analysis trước khi mở lại", "", "", ""),
            ("Rủi ro điểm ≥15", "Ghi vào Rui_ro, CEO được thông báo trong ngày", "", "", ""),
        ]),
    ]
    row_ptr = 3
    for title, items in gov_sections:
        ws_gov.cell(row=row_ptr, column=1, value=title).font = Font(bold=True, size=11, color="1F4E79")
        row_ptr += 1
        g_headers = ["Hạng mục", "Thời lượng / SLA", "Input / Output", "Chủ trì", "Ghi chú"]
        write_header(ws_gov, g_headers, row=row_ptr, fill=sub_fill, font=sub_font)
        row_ptr += 1
        for item in items:
            for c, val in enumerate(item, 1):
                ws_gov.cell(row=row_ptr, column=c, value=val)
            row_ptr += 1
        row_ptr += 1
    ws_gov.column_dimensions["A"].width = 22
    ws_gov.column_dimensions["B"].width = 18
    ws_gov.column_dimensions["C"].width = 36
    ws_gov.column_dimensions["D"].width = 16
    ws_gov.column_dimensions["E"].width = 16

    # --- Sheet 11: Hướng dẫn ---
    ws_info = wb.create_sheet("Huong_dan")
    ws_info.cell(row=1, column=1, value="Hướng dẫn sử dụng file KPI chiến lược (phiên bản mở rộng)").font = Font(
        bold=True, size=12
    )
    guide = [
        "0. Tham_so: CHỈNH SỐ TẠI ĐÂY — ngân sách, CPL/MQL target, ROAS/ROMI → mọi sheet tự cập nhật.",
        "1. Tong_quan: Dashboard + biểu đồ (cột L = tên chỉ số, cột M = # trên trục chart).",
        "2. Chien_luoc: Ngân sách = tham chiếu Tham_so (tự nhảy khi sửa ngân sách quý).",
        "3. KPI_funnel: KPI leading/lagging — nhập cột Thực tế.",
        "4. KPI_kenh: Sửa % ngân sách cột E → cột M (Ngân sách VNĐ) tự tính.",
        "5. Vung_RAG: Giá trị & vùng tham chiếu Dashboard.",
        "6. Scorecard_thang: Nhập MT/TT — % đạt & vùng & tổng điểm tự tính (SUMPRODUCT).",
        "7. KPI_tuan: Nhập Spend/Leads/MQL/SQL/Win — CPL, MQL→SQL, vùng RAG tự tính.",
        "8. Rui_ro: Nhập xác suất × tác động — điểm & vùng tự tính.",
        "9. Canh_bao: Playbook cảnh báo P1–P3.",
        "10. Quan_tri: Lịch họp, escalation.",
        "Workflow: Sửa Tham_so (ngân sách) → KPI_tuan (chi tuần) → xem Tong_quan + biểu đồ.",
    ]
    for i, line in enumerate(guide, 3):
        ws_info.cell(row=i, column=1, value=line)
    ws_info.column_dimensions["A"].width = 100

    budget_num = _parse_budget_vnd(budget_val)
    _create_tham_so_sheet(wb, budget=budget_num, plan=plan)
    _wire_kpi_workbook_formulas(wb)
    _add_kpi_workbook_charts(
        wb,
        n_weekly_rows=len(weekly_data),
        n_chart_kpis=len(chart_kpi_labels),
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def execution_rule_reply(question: str) -> str | None:
    step_id = extract_step_id(question)
    if step_id and step_id in STEP_TEMPLATES:
        return STEP_TEMPLATES[step_id]

    q = normalize_query(question)

    m = re.search(r"\bbuoc\s*(\d)\b", q)
    if m:
        mapped = STEP_NUMBER_MAP.get(m.group(1))
        if mapped and mapped in STEP_TEMPLATES:
            return STEP_TEMPLATES[mapped]

    rules: list[tuple[tuple[str, ...], str]] = [
        (
            ("facebook ads", "google ads", "mau noi dung ads", "mau ads", "primary text", "headline ads", "rsa", "buoc 1"),
            ADS_TEMPLATE,
        ),
        (
            ("tvc", "kol", "kich ban video", "video kol", "beat sheet", "shot list", "buoc 2"),
            TVC_KOL_TEMPLATE,
        ),
        (
            ("excel", "ke hoach tuan", "ke hoach marketing tuan", "file excel", "weekly plan", "xls", "buoc 3"),
            EXCEL_WEEKLY_TEMPLATE,
        ),
        (
            ("funnel", "lead den khach", "chuyen doi", "khtn khqt", "touchpoint funnel", "buoc 4"),
            FUNNEL_TEMPLATE,
        ),
        (
            ("telesales", "script sales", "tu van ban hang", "goi dien", "spin", "phan doi", "buoc 5"),
            TELESALES_TEMPLATE,
        ),
        (
            (
                "da kenh",
                "multichannel",
                "truyen thong da kenh",
                "ke hoach truyen thong",
                "omt",
                "paid owned",
                "channel mix",
                "lich phat sinh",
                "buoc 6",
            ),
            MULTICHANNEL_PLAN_TEMPLATE,
        ),
        (
            (
                "test kenh",
                "phuong phap test",
                "ab test",
                "a b test",
                "hypothesis test",
                "variant test",
                "kiem soat test",
                "channel test",
                "scale kill iterate",
                "buoc 7",
            ),
            CHANNEL_TEST_TEMPLATE,
        ),
    ]
    best_score = 0
    best: str | None = None
    for keywords, answer in rules:
        for kw in keywords:
            if kw in q:
                score = len(kw)
                if score > best_score:
                    best_score = score
                    best = answer
    return best
