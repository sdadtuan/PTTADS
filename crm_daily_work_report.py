"""Mẫu Excel báo cáo công việc hàng ngày của nhân viên — CRM PTT."""
from __future__ import annotations

from calendar import monthrange
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _thin_border() -> Border:
    thin = Side(style="thin", color="B4B4B4")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_row_header(ws, headers: list[str], row: int, *, fill: str = "2F7238") -> None:
    border = _thin_border()
    hfill = PatternFill("solid", fgColor=fill)
    hfont = Font(bold=True, color="FFFFFF")
    for col, label in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def _style_block(ws, min_row: int, max_row: int, max_col: int) -> None:
    border = _thin_border()
    for r in range(min_row, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_daily_work_report_workbook(
    *,
    brand: str = "PTT Advertising Solutions",
    staff: dict[str, Any] | None = None,
    year: int | None = None,
    month: int | None = None,
) -> Workbook:
    now = datetime.now()
    y = int(year or now.year)
    m = int(month or now.month)
    if m < 1 or m > 12:
        m = now.month
    if y < 2000 or y > 2100:
        y = now.year

    staff = staff or {}
    staff_name = str(staff.get("name") or "").strip()
    staff_code = str(staff.get("internal_code") or staff.get("code") or "").strip()
    staff_dept = str(staff.get("department") or "").strip()
    staff_title = str(staff.get("job_title") or "").strip()

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Bao_cao_ngay"

    title_font = Font(bold=True, size=14, color="2F7238")
    label_font = Font(bold=True, color="333333")
    ws.merge_cells("A1:H1")
    ws["A1"] = f"BÁO CÁO CÔNG VIỆC HÀNG NGÀY — {brand}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    info = [
        ("Họ và tên:", staff_name or "…………………………"),
        ("Mã NV:", staff_code or "…………"),
        ("Phòng ban:", staff_dept or "…………………………"),
        ("Chức danh:", staff_title or "…………………………"),
        ("Ngày báo cáo:", "… / … / …………"),
        ("Tuần (tháng):", f"Tháng {m:02d}/{y}"),
    ]
    row = 3
    for i, (label, val) in enumerate(info):
        r = row + (i // 2)
        c_label = 1 if i % 2 == 0 else 5
        c_val = c_label + 1
        ws.cell(row=r, column=c_label, value=label).font = label_font
        ws.merge_cells(start_row=r, start_column=c_val, end_row=r, end_column=c_val + 1)
        ws.cell(row=r, column=c_val, value=val)

    task_headers = [
        "STT",
        "Khung giờ",
        "Công việc / hạng mục",
        "Mô tả chi tiết",
        "Khách / Case / Dự án",
        "Tiến độ (%)",
        "Kết quả đạt được",
        "Ghi chú / vấn đề phát sinh",
    ]
    header_row = 7
    _write_row_header(ws, task_headers, header_row)
    data_start = header_row + 1
    task_rows = 12
    for i in range(task_rows):
        r = data_start + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="top")
    _style_block(ws, data_start, data_start + task_rows - 1, len(task_headers))

    summary_row = data_start + task_rows + 1
    ws.cell(row=summary_row, column=1, value="TỔNG KẾT CUỐI NGÀY").font = label_font
    ws.merge_cells(start_row=summary_row, start_column=2, end_row=summary_row, end_column=8)
    ws.cell(row=summary_row, column=2, value="")
    _style_block(ws, summary_row, summary_row, 8)

    plan_row = summary_row + 2
    ws.cell(row=plan_row, column=1, value="KẾ HOẠCH NGÀY MAI").font = label_font
    ws.merge_cells(start_row=plan_row, start_column=2, end_row=plan_row + 1, end_column=8)
    ws.cell(row=plan_row, column=2, value="")
    _style_block(ws, plan_row, plan_row + 1, 8)

    meta_row = plan_row + 3
    ws.cell(row=meta_row, column=1, value="Tổng giờ làm thực tế:").font = label_font
    ws.merge_cells(start_row=meta_row, start_column=2, end_row=meta_row, end_column=3)
    ws.cell(row=meta_row, column=4, value="Hỗ trợ cần từ quản lý:").font = label_font
    ws.merge_cells(start_row=meta_row, start_column=5, end_row=meta_row, end_column=8)

    sign_row = meta_row + 2
    ws.cell(row=sign_row, column=1, value="Nhân viên báo cáo").font = label_font
    ws.cell(row=sign_row, column=5, value="Quản lý xác nhận").font = label_font
    ws.cell(row=sign_row + 1, column=1, value="(Ký, ghi rõ họ tên)")
    ws.cell(row=sign_row + 1, column=5, value="(Ký, ghi rõ họ tên)")

    col_widths = [5, 12, 22, 28, 18, 10, 22, 24]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[header_row].height = 28

    # --- Sheet theo tháng ---
    ws_m = wb.create_sheet("Theo_thang")
    ws_m.merge_cells("A1:G1")
    ws_m["A1"] = f"THEO DÕI BÁO CÁO CÔNG VIỆC — {staff_name or 'Nhân viên'} — {m:02d}/{y}"
    ws_m["A1"].font = title_font
    ws_m["A1"].alignment = Alignment(horizontal="center")

    month_headers = [
        "Ngày",
        "Thứ",
        "Số công việc",
        "Hoàn thành",
        "Giờ làm",
        "Tóm tắt ngắn",
        "Đã nộp (Y/N)",
    ]
    _write_row_header(ws_m, month_headers, 3)
    weekdays_vi = ("T2", "T3", "T4", "T5", "T6", "T7", "CN")
    _, days_in_month = monthrange(y, m)
    for d in range(1, days_in_month + 1):
        r = 3 + d
        dt = datetime(y, m, d)
        ws_m.cell(row=r, column=1, value=f"{d:02d}/{m:02d}/{y}")
        ws_m.cell(row=r, column=2, value=weekdays_vi[dt.weekday()])
    _style_block(ws_m, 4, 3 + days_in_month, len(month_headers))
    for i, w in enumerate([12, 6, 12, 12, 10, 36, 12], 1):
        ws_m.column_dimensions[get_column_letter(i)].width = w

    # --- Hướng dẫn ---
    ws_h = wb.create_sheet("Huong_dan")
    guide = [
        "HƯỚNG DẪN SỬ DỤNG MẪU BÁO CÁO CÔNG VIỆC HÀNG NGÀY",
        "",
        "1. Sheet «Bao_cao_ngay» — điền mỗi ngày làm việc:",
        "   • Ghi rõ khung giờ, công việc, khách/case liên quan (nếu có).",
        "   • Tiến độ %: 0 = chưa làm, 100 = hoàn thành.",
        "   • Tổng kết cuối ngày: 3–5 dòng kết quả chính.",
        "   • Kế hoạch ngày mai: việc ưu tiên sáng hôm sau.",
        "",
        "2. Sheet «Theo_thang» — đánh dấu ngày đã nộp báo cáo (cột Y/N).",
        "",
        "3. Thời hạn nộp: trước 17h30 cùng ngày (hoặc theo quy định nội bộ).",
        "",
        "4. Lưu file: BC-cong-viec-<MãNV>-<YYYY-MM-DD>.xlsx",
        "",
        "5. Gửi quản lý trực tiếp qua email / nhóm nội bộ đã thống nhất.",
    ]
    for i, line in enumerate(guide, 1):
        cell = ws_h.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = title_font
    ws_h.column_dimensions["A"].width = 88

    return wb


def daily_work_report_xlsx_response(
    *,
    brand: str = "PTT Advertising Solutions",
    staff: dict[str, Any] | None = None,
    year: int | None = None,
    month: int | None = None,
    filename: str | None = None,
) -> tuple[BytesIO, str]:
    wb = build_daily_work_report_workbook(
        brand=brand, staff=staff, year=year, month=month
    )
    now = datetime.now()
    y = int(year or now.year)
    m = int(month or now.month)
    code = str((staff or {}).get("internal_code") or (staff or {}).get("code") or "NV").strip()
    code = code.replace(" ", "-")[:24] or "NV"
    fname = filename or f"mau-bao-cao-cong-viec-{code}-{y}-{m:02d}.xlsx"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, fname
